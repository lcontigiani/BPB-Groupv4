"""
Actualizar R019-03 desde un JSON y agregar una nueva fila.

Uso:
  python Codigos/R019-03/fill_r01903.py payload.json
  python Codigos/R019-03/fill_r01903.py payload.json --inplace

Por defecto genera una copia: R019-03-salida.xlsm en la carpeta superior.
"""

from __future__ import annotations

import os
import json
import re
import time
import warnings
from datetime import datetime
from pathlib import Path
import sys
from typing import Optional

LEGACY_ISO_DOCS_ROOT = Path(r"\\192.168.0.55\utn\REGISTROS\REG.DISEÑOS Y DESARROLLOS")


def _resolve_iso_root() -> Path:
    env_root = str(os.environ.get("BPB_ISO_ROOT", "") or "").strip()
    if env_root:
        return Path(env_root).expanduser()
    local_root = Path(__file__).resolve().parents[2]
    if local_root.exists():
        return local_root
    return LEGACY_ISO_DOCS_ROOT


ISO_DOCS_ROOT = _resolve_iso_root()
R01903_BASE_NAME = "R019-03 - Listado de Diseños y Desarrollos- Rev3"


def _load_workbook_quietly(*args, **kwargs):
    import openpyxl

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
            category=UserWarning,
        )
        return openpyxl.load_workbook(*args, **kwargs)


def parse_date(value: str | None) -> Optional[datetime]:
    if not value:
        return None
    value = value.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


class FileLockedError(Exception):
    pass


def _is_file_lock_error(err: Exception) -> bool:
    if isinstance(err, PermissionError):
        return True
    if isinstance(err, OSError):
        winerr = getattr(err, "winerror", None)
        if winerr in (32, 33):  # sharing violation / lock
            return True
    msg = str(err).lower()
    if "being used by another process" in msg or "permission denied" in msg:
        return True
    return False


def resolve_template_path(docs_root: Path) -> Path:
    search_dirs = [
        docs_root,
    ]
    candidates = []
    for d in search_dirs:
        candidates.extend([
            d / f"{R01903_BASE_NAME}.xlsm",
            d / f"{R01903_BASE_NAME}.xlsx",
            d / f"{R01903_BASE_NAME}.xls",
        ])
    for cand in candidates:
        if cand.exists():
            return cand
    matches = []
    for d in search_dirs:
        matches.extend(list(d.glob(f"{R01903_BASE_NAME}.*")))
    if matches:
        return matches[0]
    raise FileNotFoundError(f"No se encontró el archivo base R019-03 en {docs_root}")


def _extract_int(value) -> Optional[int]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    m = re.search(r"\d+", s)
    if not m:
        return None
    try:
        return int(m.group(0))
    except Exception:
        return None


def generate_r01903(payload: dict, inplace: bool = True, template_path: Optional[Path] = None) -> Path:
    import pythoncom
    from win32com.client import DispatchEx

    base = Path(__file__).resolve().parent
    docs_root = ISO_DOCS_ROOT
    src = template_path or resolve_template_path(docs_root)
    dest = src if inplace else docs_root / "R019-03-salida.xlsm"

    if not src.exists():
        raise FileNotFoundError(f"No se encontró el archivo base: {src}")
    if not inplace:
        dest.write_bytes(src.read_bytes())

    # Map fields (accept both R019-03 payloads and R019-01 payloads)
    dyd_val = _extract_int(payload.get("Numero_de_Registro") or payload.get("dyd") or payload.get("DyD"))
    descripcion = (
        payload.get("BP")
        or payload.get("codigo_producto")
        or payload.get("descripcion")
        or payload.get("Descripcion_Aplicacion_Producto")
        or ""
    )
    empresa = payload.get("empresa") or payload.get("Solicita") or ""
    fecha_inicio = parse_date(payload.get("Fecha_actual") or payload.get("fecha_inicio")) or datetime.now()
    etapa = payload.get("etapa") or payload.get("Etapa") or "Recolección de Datos"
    situacion = payload.get("situacion") or payload.get("Situacion_Declarada") or "-"
    fecha_fin = parse_date(payload.get("fecha_finalizacion") or payload.get("Fecha_Finalizacion") or payload.get("Fecha_finalizacion"))
    valoracion = payload.get("valoracion") or payload.get("Valoracion") or ""
    ubicacion = payload.get("ubicacion") or payload.get("Ubicacion") or ""

    def is_busy_error(err: Exception) -> bool:
        try:
            hresult = getattr(err, "hresult", None)
            if hresult is None and getattr(err, "args", None):
                hresult = err.args[0]
            return int(hresult) in (-2147418111, -2147417846)
        except Exception:
            return False

    def write_with_openpyxl():
        from copy import copy as shallow_copy
        import openpyxl

        keep_vba = dest.suffix.lower() == ".xlsm"
        wb = _load_workbook_quietly(dest, keep_vba=keep_vba)
        ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
        ws = wb[ws_name] if ws_name else wb.active

        def last_row_with_value():
            r = ws.max_row
            while r >= 7:
                val = ws.cell(row=r, column=1).value
                if val is not None and str(val).strip() != "":
                    return r
                r -= 1
            return 6

        last_row = last_row_with_value()
        new_row = last_row + 1
        template_row = last_row if last_row >= 7 else 7

        max_col = 9
        for col in range(1, max_col + 1):
            src = ws.cell(row=template_row, column=col)
            dst = ws.cell(row=new_row, column=col)
            if src.has_style:
                dst.font = shallow_copy(src.font)
                dst.border = shallow_copy(src.border)
                dst.fill = shallow_copy(src.fill)
                dst.number_format = src.number_format
                dst.protection = shallow_copy(src.protection)
                dst.alignment = shallow_copy(src.alignment)
        ws.row_dimensions[new_row].height = ws.row_dimensions[template_row].height

        nonlocal_dyd = dyd_val
        if nonlocal_dyd is None:
            last_dyd = ws.cell(row=last_row, column=1).value
            try:
                nonlocal_dyd = int(last_dyd) + 1
            except Exception:
                nonlocal_dyd = 1

        ws.cell(row=new_row, column=1).value = nonlocal_dyd
        ws.cell(row=new_row, column=2).value = descripcion
        ws.cell(row=new_row, column=3).value = empresa
        ws.cell(row=new_row, column=4).value = fecha_inicio
        ws.cell(row=new_row, column=5).value = etapa
        ws.cell(row=new_row, column=6).value = situacion
        ws.cell(row=new_row, column=7).value = fecha_fin if fecha_fin else ""
        ws.cell(row=new_row, column=8).value = valoracion
        ws.cell(row=new_row, column=9).value = ubicacion

        wb.save(dest)

    # Prefer openpyxl (no Excel/COM). If it fails, fall back to COM.
    try:
        write_with_openpyxl()
        return dest
    except Exception as e:
        if _is_file_lock_error(e):
            raise FileLockedError(str(e))
        last_err = e

    for attempt in range(1, 6):
        pythoncom.CoInitialize()
        excel = DispatchEx("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False
        try:
            wb = excel.Workbooks.Open(str(dest))
            try:
                ws = wb.Worksheets("Listado") if "Listado" in [s.Name for s in wb.Worksheets] else wb.Worksheets(1)

                # Buscar la última fila con DyD N° (columna A)
                xl_up = -4162
                last_row = ws.Cells(ws.Rows.Count, 1).End(xl_up).Row
                if last_row < 7:
                    last_row = 6
                new_row = last_row + 1

                # Copiar solo formato de la última fila con datos (o fila 7 como plantilla)
                template_row = last_row if last_row >= 7 else 7
                ws.Rows(template_row).Copy()
                ws.Rows(new_row).PasteSpecial(Paste=-4122)  # xlPasteFormats
                excel.CutCopyMode = False

                # DyD N° = último + 1 si no se proveyó
                if dyd_val is None:
                    last_dyd = ws.Cells(last_row, 1).Value
                    try:
                        dyd_val = int(last_dyd) + 1
                    except Exception:
                        dyd_val = 1

                ws.Cells(new_row, 1).Value = dyd_val
                ws.Cells(new_row, 2).Value = descripcion
                ws.Cells(new_row, 3).Value = empresa
                ws.Cells(new_row, 4).Value = fecha_inicio
                ws.Cells(new_row, 5).Value = etapa
                ws.Cells(new_row, 6).Value = situacion
                ws.Cells(new_row, 7).Value = fecha_fin if fecha_fin else ""
                ws.Cells(new_row, 8).Value = valoracion
                ws.Cells(new_row, 9).Value = ubicacion

                wb.Save()
                return dest
            finally:
                wb.Close()
        except Exception as e:
            last_err = e
            if is_busy_error(e) and attempt < 5:
                time.sleep(0.6 * attempt)
                continue
            raise
        finally:
            excel.Quit()
            pythoncom.CoUninitialize()

    if last_err:
        if is_busy_error(last_err):
            raise FileLockedError(str(last_err))
        raise last_err

    return dest


def update_r01903_status(
    bp: str,
    etapa: Optional[str] = None,
    situacion: Optional[str] = None,
    fecha_fin: Optional[str] = None,
    template_path: Optional[Path] = None
) -> bool:
    import openpyxl

    base = Path(__file__).resolve().parent
    docs_root = ISO_DOCS_ROOT
    src = template_path or resolve_template_path(docs_root)
    if not src.exists():
        raise FileNotFoundError(f"No se encontró el archivo base: {src}")

    keep_vba = src.suffix.lower() == ".xlsm"
    wb = _load_workbook_quietly(src, keep_vba=keep_vba)
    ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
    ws = wb[ws_name] if ws_name else wb.active

    target_row = None
    bp_norm = (bp or "").strip().lower()
    if bp_norm:
        for r in range(7, (ws.max_row or 0) + 1):
            val = ws.cell(row=r, column=2).value
            if val is None:
                continue
            s = str(val).strip().lower()
            if s == bp_norm or bp_norm in s:
                target_row = r
                break

    if not target_row:
        wb.close()
        return False

    if etapa is not None and str(etapa).strip() != "":
        ws.cell(row=target_row, column=5).value = etapa
    if situacion is not None and str(situacion).strip() != "":
        ws.cell(row=target_row, column=6).value = situacion
    if fecha_fin is not None and str(fecha_fin).strip() != "":
        ws.cell(row=target_row, column=7).value = parse_date(str(fecha_fin))

    wb.save(src)
    try:
        wb.close()
    except Exception:
        pass
    return True


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python Codigos/R019-03/fill_r01903.py payload.json [--inplace]")
        sys.exit(1)

    payload_path = Path(sys.argv[1]).resolve()
    inplace = "--inplace" in sys.argv

    payload = json.loads(payload_path.read_text(encoding="utf-8"))
    dest = generate_r01903(payload, inplace=inplace)
    print(f"Actualizado: {dest}")


if __name__ == "__main__":
    main()
