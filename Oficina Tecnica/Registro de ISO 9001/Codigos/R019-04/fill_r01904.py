"""
Actualizar R019-04 (Project) desde un CSV de eventos de R019-02 y un
JSON de duraciones modelo para comparar (baseline vs real).

Uso:
  python Codigos/R019-04/fill_r01904.py eventos.csv --modelo modelo.json
  python Codigos/R019-04/fill_r01904.py eventos.csv --modelo modelo.json --out R019-04-ejemplo.mpp
  python Codigos/R019-04/fill_r01904.py eventos.csv --modelo modelo.json --inplace
  python Codigos/R019-04/fill_r01904.py eventos.csv --modelo modelo.json --mpp
"""

from __future__ import annotations

import csv
import json
from datetime import datetime, time, timedelta
from pathlib import Path
import os
import sys
import tempfile
import time as pytime
from typing import Dict, List, Tuple
import pythoncom
import openpyxl

LEGACY_ISO_DOCS_ROOT = Path(r"\\192.168.0.55\utn\REGISTROS\REG.DISEÑOS Y DESARROLLOS")


def _resolve_iso_root() -> Path:
    env_root = str(os.environ.get("BPB_ISO_ROOT", "") or "").strip()
    if env_root:
        return Path(env_root).expanduser()
    local_root = Path(__file__).resolve().parents[2]
    if local_root.exists():
        return local_root
    return LEGACY_ISO_DOCS_ROOT


def _resolve_output_dir(root: Path, prefix: str, preferred_name: str) -> Path:
    preferred = root / preferred_name
    if preferred.exists():
        return preferred
    try:
        for entry in root.iterdir():
            if entry.is_dir() and entry.name.upper().startswith(prefix.upper()):
                return entry
    except Exception:
        pass
    return preferred


ISO_DOCS_ROOT = _resolve_iso_root()
R01904_OUTPUT_DIR = _resolve_output_dir(ISO_DOCS_ROOT, "R019-04", "R019-04 - Planificación de Diseño")

from win32com.client import Dispatch, DispatchEx, gencache


EXPECTED_COLS = [
    "fecha",
    "etapa",
    "area",
    "empresa",
    "descripcion",
    "resultado",
    "accion",
    "usuario",
]


def normalize_col(name: str) -> str:
    if name is None:
        return ""
    text = name.strip().lower()
    replacements = {
        "á": "a",
        "é": "e",
        "í": "i",
        "ó": "o",
        "ú": "u",
        "ñ": "n",
    }
    for src, dst in replacements.items():
        text = text.replace(src, dst)
    return text


def parse_csv(path: Path) -> List[Dict[str, str]]:
    raw = path.read_text(encoding="utf-8-sig", errors="replace")
    try:
        dialect = csv.Sniffer().sniff(raw.splitlines()[0])
        delimiter = dialect.delimiter
    except Exception:
        delimiter = "," if "," in raw.splitlines()[0] else ";"

    rows: List[Dict[str, str]] = []
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh, delimiter=delimiter)
        if not reader.fieldnames:
            raise ValueError("El CSV no tiene encabezados.")
        colmap = {normalize_col(n): n for n in reader.fieldnames}
        missing = [c for c in EXPECTED_COLS if c not in colmap]
        if missing:
            raise ValueError(f"Faltan columnas en CSV: {', '.join(missing)}")
        for row in reader:
            out = {c: (row.get(colmap[c]) or "").strip() for c in EXPECTED_COLS}
            rows.append(out)
    return rows


def parse_date(value: str):
    if not value:
        return None
    for fmt in (
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%d/%m/%Y",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d %H:%M:%S",
    ):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(value)
    except ValueError:
        return None


def load_workbook_quietly(*args, **kwargs):
    return openpyxl.load_workbook(*args, **kwargs)


def read_r01902_rows(path: Path) -> List[Dict[str, str]]:
    wb = load_workbook_quietly(str(path), read_only=True, data_only=True)
    try:
        ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
        ws = wb[ws_name] if ws_name else wb.active
        rows = []
        empty_streak = 0
        max_row = ws.max_row or 0
        for row_idx in range(4, max_row + 1):
            values = [ws.cell(row=row_idx, column=col).value for col in range(1, 9)]
            if all(v is None or str(v).strip() == "" for v in values):
                empty_streak += 1
                if empty_streak >= 5:
                    break
                continue
            empty_streak = 0
            rows.append({
                "fecha": str(values[0]).strip() if values[0] is not None else "",
                "etapa": str(values[1]).strip() if values[1] is not None else "",
                "area": str(values[2]).strip() if values[2] is not None else "",
                "empresa": str(values[3]).strip() if values[3] is not None else "",
                "descripcion": str(values[4]).strip() if values[4] is not None else "",
                "resultado": str(values[5]).strip() if values[5] is not None else "",
                "accion": str(values[6]).strip() if values[6] is not None else "",
                "usuario": str(values[7]).strip() if values[7] is not None else "",
            })
        return rows
    finally:
        wb.close()


def task_name(event: Dict[str, str]) -> str:
    desc = (event.get("descripcion") or "").strip()
    if desc:
        return desc
    etapa = (event.get("etapa") or "").strip()
    if etapa:
        return etapa
    parts = [event.get("etapa"), event.get("area"), event.get("empresa")]
    name = " | ".join([p for p in parts if p])
    return name if name else "Evento"


def build_notes(event: Dict[str, str]) -> str:
    lines = []
    if event.get("descripcion"):
        lines.append(f"Descripcion: {event['descripcion']}")
    if event.get("resultado"):
        lines.append(f"Resultado: {event['resultado']}")
    if event.get("accion"):
        lines.append(f"Accion: {event['accion']}")
    if event.get("usuario"):
        lines.append(f"Usuario: {event['usuario']}")
    return "\n".join(lines)


def compute_stage_stats_durations(root: Path) -> Dict[str, int]:
    r01902_dir = _resolve_output_dir(root, "R019-02", "R019-02 - Revisión, Verificación y Validación")
    if not r01902_dir.exists():
        return {}

    stage_durations: Dict[str, List[int]] = {}
    for path in r01902_dir.glob("R019-02 Rev03 - *.xls*"):
        try:
            rows = read_r01902_rows(path)
        except Exception:
            continue
        dates_by_stage: Dict[str, List[datetime]] = {}
        for row in rows:
            etapa = (row.get("etapa") or "").strip()
            dt = parse_date(row.get("fecha", ""))
            if not etapa or not dt:
                continue
            dates_by_stage.setdefault(etapa, []).append(dt)
        for etapa, dates in dates_by_stage.items():
            if not dates:
                continue
            min_dt = min(dates)
            max_dt = max(dates)
            duration = round((max_dt - min_dt).total_seconds() / 86400)
            stage_durations.setdefault(etapa, []).append(duration)

    avg_map: Dict[str, int] = {}
    for etapa, values in stage_durations.items():
        if not values:
            continue
        avg = sum(values) / len(values)
        avg_map[etapa] = max(1, int(round(avg)))
    return avg_map


def build_stage_blocks(eventos: List[Dict[str, str]]) -> List[Dict[str, object]]:
    dates: List[Tuple[Dict[str, str], datetime | None]] = [(e, parse_date(e.get("fecha", ""))) for e in eventos]
    blocks: List[Dict[str, object]] = []
    current = None
    for event, dt in dates:
        etapa = (event.get("etapa") or "").strip() or "Sin etapa"
        if current is None or current["etapa"] != etapa:
            current = {"etapa": etapa, "start": dt, "end": dt, "events": [event]}
            blocks.append(current)
        else:
            current["events"].append(event)
            if dt:
                current["end"] = dt
    return blocks


def load_model(path: Path | None) -> Tuple[Dict[str, int], int, datetime | None]:
    if not path:
        return {}, 1, None
    data = json.loads(path.read_text(encoding="utf-8"))
    durations = data.get("duraciones", data)
    default = int(data.get("duracion_default", 1))
    start_raw = data.get("fecha_inicio_proyecto")
    start_dt = parse_date(start_raw) if start_raw else None
    cleaned = {}
    for k, v in durations.items():
        try:
            cleaned[str(k)] = int(v)
        except (TypeError, ValueError):
            continue
    return cleaned, default, start_dt


def _call_with_retries(target, attr_name: str, *args, retries: int = 6, sleep_base: float = 0.6, **kwargs):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            attr = getattr(target, attr_name)
            return attr(*args, **kwargs) if callable(attr) else attr
        except Exception as e:
            last_err = e
            pytime.sleep(sleep_base * attempt)
            try:
                pythoncom.PumpWaitingMessages()
            except Exception:
                pass
    raise RuntimeError(f"No se pudo ejecutar '{attr_name}' en MS Project. Error: {last_err}")


def _create_msproject_app():
    last_err = None
    for factory in (
        lambda: gencache.EnsureDispatch("MSProject.Application"),
        lambda: DispatchEx("MSProject.Application"),
        lambda: Dispatch("MSProject.Application"),
    ):
        try:
            return factory()
        except Exception as e:
            last_err = e
    raise RuntimeError(f"No se pudo iniciar la automatizacion COM de Microsoft Project. Error: {last_err}")


def open_project_with_retries(app, project_path: Path, retries: int = 4):
    last_err = None
    for attempt in range(1, retries + 1):
        try:
            if hasattr(app, "FileOpenEx"):
                _call_with_retries(app, "FileOpenEx", str(project_path), False, retries=1)
            elif hasattr(app, "FileOpen"):
                _call_with_retries(app, "FileOpen", str(project_path), ReadOnly=False, retries=1)
            else:
                raise RuntimeError("La instalacion de Microsoft Project no expone FileOpen/FileOpenEx via COM.")
        except Exception as e:
            last_err = e

        for accessor in (
            lambda: getattr(app, "ActiveProject"),
            lambda: getattr(getattr(app, "Application"), "ActiveProject"),
            lambda: getattr(app, "Projects")(1),
            lambda: getattr(getattr(app, "Application"), "Projects")(1),
        ):
            try:
                proj = accessor()
                if proj is not None:
                    return proj
            except Exception as e:
                last_err = e

        pytime.sleep(0.4 * attempt)
        try:
            pythoncom.PumpWaitingMessages()
        except Exception:
            pass

    raise RuntimeError(
        f"No se pudo abrir el modelo R019-04 en MS Project ({project_path}). "
        f"Error: {last_err}"
    )


def main() -> None:
    if len(sys.argv) < 2:
        print("Uso: python Codigos/fill_r01904.py eventos.csv --modelo modelo.json [--out archivo.mpp|--inplace]")
        sys.exit(1)

    csv_path = Path(sys.argv[1]).resolve()
    inplace = "--inplace" in sys.argv
    out_path = None
    model_path = None
    include_notes = "--notes" in sys.argv
    if "--out" in sys.argv:
        idx = sys.argv.index("--out")
        if idx + 1 < len(sys.argv):
            out_path = Path(sys.argv[idx + 1]).resolve()
    if "--modelo" in sys.argv:
        idx = sys.argv.index("--modelo")
        if idx + 1 < len(sys.argv):
            model_path = Path(sys.argv[idx + 1]).resolve()

    docs_root = _resolve_iso_root()
    output_root = R01904_OUTPUT_DIR
    model = docs_root / "R019-04-Modelo.mpp"
    if not model.exists():
        alt = docs_root / "R019-04" / "R019-04-Modelo.mpp"
        if alt.exists():
            model = alt
        else:
            mpp_files = list(docs_root.glob("*.mpp"))
            if not mpp_files:
                raise FileNotFoundError("No se encontró ningún .mpp en la carpeta.")
            model = mpp_files[0]

    dest = model if inplace else (out_path or (output_root / "R019-04-salida.mpp"))
    if not inplace:
        output_root.mkdir(parents=True, exist_ok=True)

    eventos = parse_csv(csv_path)
    if not eventos:
        raise ValueError("El CSV no tiene filas de eventos.")

    # Copia local para evitar problemas con rutas UNC y acentos (archivo temporal único)
    fd, tmp_name = tempfile.mkstemp(prefix="r01904-", suffix=".mpp")
    os.close(fd)
    tmp = Path(tmp_name)
    tmp.write_bytes(model.read_bytes())

    output_path = dest
    try:
        pythoncom.CoInitialize()
        app = _create_msproject_app()
        try:
            try:
                app.Visible = False
            except Exception:
                pass
            try:
                app.DisplayAlerts = False
            except Exception:
                pass

            proj = open_project_with_retries(app, tmp)

            model_durations, default_duration, model_start = load_model(model_path)
            stats_durations = compute_stage_stats_durations(ISO_DOCS_ROOT)
            dates: List[Tuple[Dict[str, str], datetime | None]] = [
                (e, parse_date(e.get("fecha", ""))) for e in eventos
            ]
            stage_blocks = build_stage_blocks(eventos)
            project_start = model_start or parse_date(eventos[0].get("fecha", "")) or datetime.now()
            project_start = datetime.combine(project_start.date(), time(9, 0))

            for i in range(proj.Tasks.Count, 0, -1):
                t = proj.Tasks(i)
                if t is not None:
                    t.Delete()

            baseline_cursor = None
            if model_start:
                proj.ProjectStart = datetime.combine(model_start.date(), time(9, 0))
                baseline_cursor = proj.ProjectStart

            for idx, (event, dt) in enumerate(dates, start=1):
                task = proj.Tasks.Add(task_name(event))
                try:
                    task.Estimated = False
                except Exception:
                    pass
                if include_notes:
                    notes = build_notes(event)
                    if notes:
                        task.Notes = notes
                if event.get("etapa"):
                    task.Text1 = event["etapa"]
                if event.get("area"):
                    task.Text2 = event["area"]
                if event.get("empresa"):
                    task.Text3 = event["empresa"]
                if event.get("descripcion"):
                    task.Text4 = event["descripcion"]
                if event.get("resultado"):
                    task.Text5 = event["resultado"]

                if dt:
                    start_dt = datetime.combine(dt.date(), time(9, 0))
                    next_dt = dates[idx][1] if idx < len(dates) else None
                    if next_dt:
                        finish_dt = datetime.combine(next_dt.date(), time(18, 0))
                    else:
                        finish_dt = start_dt
                        task.Milestone = True
                    task.Start = start_dt
                    task.Finish = finish_dt

                etapa = event.get("etapa", "")
                dur_days = model_durations.get(etapa, default_duration)
                if baseline_cursor is None and dt:
                    baseline_cursor = start_dt
                if baseline_cursor:
                    base_start = baseline_cursor
                    base_finish = base_start + timedelta(days=dur_days)
                    try:
                        task.BaselineStart = base_start
                        task.BaselineFinish = base_finish
                        try:
                            task.BaselineDuration = f"{dur_days}d"
                        except Exception:
                            pass
                    except Exception:
                        pass
                    baseline_cursor = base_finish

            for block in stage_blocks:
                stage_name = str(block["etapa"])
                stage_start = block["start"] or project_start
                avg_days = stats_durations.get(stage_name, model_durations.get(stage_name, default_duration))
                milestone_date = datetime.combine(stage_start.date(), time(18, 0)) + timedelta(days=max(1, int(avg_days)))
                milestone = proj.Tasks.Add(f"Milestone objetivo - {stage_name}")
                try:
                    milestone.Estimated = False
                except Exception:
                    pass
                milestone.Start = milestone_date
                milestone.Finish = milestone_date
                milestone.Milestone = True
                try:
                    milestone.Notes = f"Hito objetivo calculado desde estadisticas para la etapa {stage_name}."
                except Exception:
                    pass

            if inplace:
                _call_with_retries(app, "FileSave")
            else:
                _call_with_retries(app, "FileSaveAs", str(dest))
        finally:
            try:
                _call_with_retries(app, "FileCloseAll", True, retries=1)
            except Exception:
                pass
            try:
                app.Quit()
            except Exception:
                pass
            try:
                pythoncom.CoUninitialize()
            except Exception:
                pass
    finally:
        try:
            tmp.unlink()
        except Exception:
            pass

    print(f"Actualizado: {output_path}")


if __name__ == "__main__":
    main()
