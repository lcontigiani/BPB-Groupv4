"""

Actualizar R019-02 desde un JSON (cabecera) y un CSV (eventos).



Uso:

  python Codigos/R019-02/fill_r01902.py cabecera.json eventos.csv

  python Codigos/R019-02/fill_r01902.py cabecera.json eventos.csv --inplace



Por defecto genera una copia: R019-02-salida.xls en la carpeta superior.

"""



from __future__ import annotations



import os
import csv

import json
import warnings

from datetime import datetime

import time

from pathlib import Path

import sys

from typing import Dict, List, Optional

LEGACY_ISO_DOCS_ROOT = Path(r"\\192.168.0.55\utn\REGISTROS\REG.DISEÑOS Y DESARROLLOS")


def _resolve_iso_root() -> Path:
    env_root = str(os.environ.get("BPB_ISO_ROOT", "") or "").strip()
    if env_root:
        return Path(env_root).expanduser()
    local_root = Path(__file__).resolve().parents[2]
    if local_root.exists():
        return local_root
    return LEGACY_ISO_DOCS_ROOT


def _resolve_output_dir(root: Path, prefix: str, preferred_name: str) -> Path:
    preferred = root / preferred_name
    if preferred.exists():
        return preferred
    try:
        for entry in root.iterdir():
            if entry.is_dir() and entry.name.upper().startswith(prefix.upper()):
                return entry
    except Exception:
        pass
    return preferred


ISO_DOCS_ROOT = _resolve_iso_root()
R01902_OUTPUT_DIR = _resolve_output_dir(ISO_DOCS_ROOT, "R019-02", "R019-02 - Revisi?n, Verificaci?n y Validaci?n")


def _load_workbook_quietly(*args, **kwargs):
    import openpyxl

    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
            category=UserWarning,
        )
        return openpyxl.load_workbook(*args, **kwargs)





EXPECTED_COLS = [

    "fecha",

    "etapa",

    "area",

    "empresa",

    "descripcion",

    "resultado",

    "accion",

    "usuario",

]





def normalize_col(name: str) -> str:

    if name is None:

        return ""

    text = name.strip().lower()

    # Normalizar acentos comunes para comparar

    replacements = {

        "á": "a",

        "é": "e",

        "í": "i",

        "ó": "o",

        "ú": "u",

        "ñ": "n",

    }

    for src, dst in replacements.items():

        text = text.replace(src, dst)

    return text





def parse_csv(path: Path) -> List[Dict[str, str]]:

    raw = path.read_text(encoding="utf-8-sig", errors="replace")

    try:

        dialect = csv.Sniffer().sniff(raw.splitlines()[0])

        delimiter = dialect.delimiter

    except Exception:

        delimiter = "," if "," in raw.splitlines()[0] else ";"



    rows: List[Dict[str, str]] = []

    with path.open(newline="", encoding="utf-8-sig") as fh:

        reader = csv.DictReader(fh, delimiter=delimiter)

        if not reader.fieldnames:

            raise ValueError("El CSV no tiene encabezados.")

        colmap = {normalize_col(n): n for n in reader.fieldnames}

        missing = [c for c in EXPECTED_COLS if c not in colmap]

        if missing:

            raise ValueError(f"Faltan columnas en CSV: {', '.join(missing)}")

        for row in reader:

            out = {c: (row.get(colmap[c]) or "").strip() for c in EXPECTED_COLS}

            rows.append(out)

    return rows





def parse_date(value: str):

    if not value:

        return ""

    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d/%m/%Y", "%Y-%m-%d %H:%M", "%Y-%m-%d %H:%M:%S"):

        try:

            return datetime.strptime(value, fmt)

        except ValueError:

            continue

    try:

        return datetime.fromisoformat(value)

    except ValueError:

        return value





def resolve_template_path(docs_root: Path) -> Path:

    candidates = [

        docs_root / "R019-02-Modelo.xlsm",

        docs_root / "R019-02-Modelo.xlsx",

        docs_root / "R019-02-Modelo.xls",

    ]

    for cand in candidates:

        if cand.exists():

            return cand

    matches = list(docs_root.glob("R019-02-Modelo.*"))

    if matches:

        return matches[0]

    raise FileNotFoundError(f"No se encontró el archivo base R019-02 en {docs_root}")





def generate_r01902(payload: dict, output_path: Path, template_path: Optional[Path] = None) -> Path:

    base = Path(__file__).resolve().parent

    docs_root = base.parent.parent / "R019-02"

    src = template_path or resolve_template_path(docs_root)



    if not src.exists():

        raise FileNotFoundError(f"No se encontr?? el archivo base: {src}")

    if output_path.exists():

        raise FileExistsError(f"El archivo ya existe: {output_path}")



    output_path.parent.mkdir(parents=True, exist_ok=True)
    last_copy_err = None
    for attempt in range(1, 4):
        try:
            output_path.write_bytes(src.read_bytes())
            last_copy_err = None
            break
        except PermissionError as e:
            last_copy_err = e
            # If the file already exists (locked/readonly), try to update it in-place below.
            if output_path.exists():
                break
            time.sleep(0.4 * attempt)
    if last_copy_err and not output_path.exists():
        raise last_copy_err


    producto = payload.get("producto") or payload.get("Descripcion_Aplicacion_Producto") or ""

    codigo = payload.get("BP") or payload.get("codigo_producto") or ""

    ultima_rev = payload.get("ultima_revision_plano") or ""

    ultima_texto = f"Ultima Revisión Plano: {ultima_rev}" if ultima_rev else "Ultima Revisión Plano:"



    def _write_with_openpyxl():
        import openpyxl

        last_err = None
        for attempt in range(1, 4):
            try:
                keep_vba = output_path.suffix.lower() == ".xlsm"
                wb = _load_workbook_quietly(output_path, keep_vba=keep_vba)
                ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active

                ws.cell(row=2, column=2).value = producto  # B2
                ws.cell(row=2, column=5).value = ultima_texto  # E2
                ws.cell(row=2, column=8).value = codigo  # H2

                wb.save(output_path)
                try:
                    wb.close()
                except Exception:
                    pass
                return
            except PermissionError as e:
                last_err = e
                time.sleep(0.4 * attempt)
            except Exception as e:
                last_err = e
                break
        if last_err:
            raise last_err


    def _is_busy_error(err: Exception) -> bool:
        try:
            hresult = getattr(err, "hresult", None)
            if hresult is None and getattr(err, "args", None):
                hresult = err.args[0]
            return int(hresult) in (-2147418111, -2147417846)
        except Exception:
            return False

    def _is_permission_error(err: Exception) -> bool:
        if isinstance(err, PermissionError):
            return True
        if isinstance(err, OSError):
            winerr = getattr(err, "winerror", None)
            if winerr in (32, 33):
                return True
            if getattr(err, "errno", None) == 13:
                return True
        msg = str(err).lower()
        if "permission denied" in msg or "being used by another process" in msg:
            return True
        return False


    def _write_with_com():
        import pythoncom
        from win32com.client import DispatchEx
        import gc

        last_err = None
        for attempt in range(1, 6):
            pythoncom.CoInitialize()
            excel = DispatchEx("Excel.Application")
            excel.Visible = False

            excel.DisplayAlerts = False

            wb = None

            try:

                wb = excel.Workbooks.Open(str(output_path))

                ws = wb.Worksheets(1)

                ws.Range("B2").Value = producto

                ws.Range("E2").Value = ultima_texto

                ws.Range("H2").Value = codigo

                wb.Save()

                wb.Close(False)

                excel.Quit()

                return

            except Exception as e:

                last_err = e

                try:
                    if wb is not None:
                        wb.Close(False)
                except Exception:
                    pass
                excel.Quit()
                try:
                    del wb
                    del ws
                    del excel
                except Exception:
                    pass
                gc.collect()
                if _is_busy_error(e):
                    time.sleep(0.4 * attempt)
                    continue
                break
            finally:
                try:
                    pythoncom.CoUninitialize()
                except Exception:
                    pass
        if last_err:
            raise last_err


    # Prefer openpyxl to avoid leaving Excel COM processes open.
    # It preserves formatting when only values are written.
    _write_with_openpyxl()

    return output_path




def append_r01902_event(output_path: Path, event: dict) -> int:
    import openpyxl
    from copy import copy as shallow_copy

    if not output_path.exists():
        raise FileNotFoundError(f"No se encontro el archivo: {output_path}")

    def _is_permission_error(err: Exception) -> bool:
        if isinstance(err, PermissionError):
            return True
        if isinstance(err, OSError):
            winerr = getattr(err, 'winerror', None)
            if winerr in (32, 33):
                return True
            if getattr(err, 'errno', None) == 13:
                return True
        msg = str(err).lower()
        return 'permission denied' in msg or 'being used by another process' in msg

    last_err = None
    for attempt in range(1, 4):
        wb = None
        try:
            keep_vba = output_path.suffix.lower() == '.xlsm'
            wb = _load_workbook_quietly(output_path, keep_vba=keep_vba)
            ws = wb['Listado'] if 'Listado' in wb.sheetnames else wb.active

            start_row = 4
            max_row = ws.max_row or start_row
            target_row = None

            def _row_empty(r):
                for c in range(1, 9):
                    val = ws.cell(row=r, column=c).value
                    if val is not None and str(val).strip() != '':
                        return False
                return True

            for r in range(start_row, max_row + 1):
                if _row_empty(r):
                    target_row = r
                    break

            if target_row is None:
                target_row = max_row + 1
                template_row = max(start_row, max_row)
                for c in range(1, 9):
                    src = ws.cell(row=template_row, column=c)
                    dst = ws.cell(row=target_row, column=c)
                    if src.has_style:
                        dst.font = shallow_copy(src.font)
                        dst.border = shallow_copy(src.border)
                        dst.fill = shallow_copy(src.fill)
                        dst.number_format = src.number_format
                        dst.protection = shallow_copy(src.protection)
                        dst.alignment = shallow_copy(src.alignment)
                ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height

            fecha_raw = (event or {}).get('fecha')
            fecha_val = parse_date(str(fecha_raw)) if fecha_raw else ''

            def _val(key):
                val = (event or {}).get(key)
                return str(val).strip() if val is not None else ''

            ws.cell(row=target_row, column=1).value = fecha_val
            ws.cell(row=target_row, column=2).value = _val('etapa')
            ws.cell(row=target_row, column=3).value = _val('area')
            ws.cell(row=target_row, column=4).value = _val('empresa')
            desc_cell = ws.cell(row=target_row, column=5)
            desc_cell.value = _val('descripcion')
            try:
                desc_cell.alignment = desc_cell.alignment.copy(wrap_text=True)
            except Exception:
                pass
            ws.cell(row=target_row, column=6).value = _val('resultados') or _val('resultado')
            ws.cell(row=target_row, column=7).value = _val('accion')
            ws.cell(row=target_row, column=8).value = _val('usuario')

            wb.save(output_path)
            return target_row
        except Exception as e:
            last_err = e
            if _is_permission_error(e):
                time.sleep(0.4 * attempt)
                continue
            break
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    if last_err:
        raise last_err
    raise RuntimeError('No se pudo guardar el evento en R019-02.')


def main() -> None:

    if len(sys.argv) < 3:

        print("Uso: python Codigos/fill_r01902.py cabecera.json eventos.csv [--inplace]")

        sys.exit(1)



    header_path = Path(sys.argv[1]).resolve()

    csv_path = Path(sys.argv[2]).resolve()

    inplace = "--inplace" in sys.argv



    header = json.loads(header_path.read_text(encoding="utf-8"))

    eventos = parse_csv(csv_path)



    base = Path(__file__).resolve().parent
    template_root = base.parent.parent / "R019-02"
    src = resolve_template_path(template_root)
    docs_root = R01902_OUTPUT_DIR
    dest = src if inplace else docs_root / "R019-02-salida.xlsm"



    if not inplace:
        docs_root.mkdir(parents=True, exist_ok=True)

        dest.write_bytes(src.read_bytes())



    # Use openpyxl for updates

    import openpyxl

    keep_vba = dest.suffix.lower() == ".xlsm"

    wb = _load_workbook_quietly(dest, keep_vba=keep_vba)

    ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active



    # Cabecera

    producto = header.get("producto", "")

    codigo = header.get("codigo_producto", "")

    ultima_rev = header.get("ultima_revision_plano", "")



    ws.cell(row=2, column=2).value = producto  # B2

    ws.cell(row=2, column=5).value = f"Ultima Revisión Plano: {ultima_rev}" if ultima_rev else "Ultima Revisión Plano:"

    ws.cell(row=2, column=8).value = codigo  # H2



    # Encabezados tabla (fila 3)

    ws.cell(row=3, column=1).value = "Fecha"

    ws.cell(row=3, column=2).value = "Etapa"

    ws.cell(row=3, column=3).value = "Area"

    ws.cell(row=3, column=4).value = "Empresa"

    ws.cell(row=3, column=5).value = "Descripción"

    ws.cell(row=3, column=6).value = "Resultados"

    ws.cell(row=3, column=7).value = "Acción"

    ws.cell(row=3, column=8).value = "Ord / Usuario"



    # Buscar última fila con fecha en columna A

    last_row = ws.max_row

    if last_row < 4:

        last_row = 3

    row = last_row + 1



    for e in eventos:

        ws.cell(row=row, column=1).value = parse_date(e["fecha"])

        ws.cell(row=row, column=2).value = e["etapa"]

        ws.cell(row=row, column=3).value = e["area"]

        ws.cell(row=row, column=4).value = e["empresa"]

        ws.cell(row=row, column=5).value = e["descripcion"]

        ws.cell(row=row, column=6).value = e["resultado"]

        ws.cell(row=row, column=7).value = e["accion"]

        ws.cell(row=row, column=8).value = e["usuario"]

        row += 1



    wb.save(dest)

    print(f"Actualizado: {dest}")





if __name__ == "__main__":

    main()

