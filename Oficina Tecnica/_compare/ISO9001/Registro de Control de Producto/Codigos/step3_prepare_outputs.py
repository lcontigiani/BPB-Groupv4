import argparse
import base64
import json
import logging
import re
import io
import difflib
import os
import shutil
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Iterable

import csv
import yaml
from datetime import datetime

try:
    import pdfplumber
except ImportError:
    pdfplumber = None

try:
    import pypdfium2 as pdfium
except ImportError:
    pdfium = None

try:
    import openpyxl
except ImportError:
    openpyxl = None
    
try:
    import xlrd
except ImportError:
    xlrd = None

import requests
from PIL import Image

import extract_to_csv
import extract_to_csv_bolas_fixed
import extract_to_csv_especial
import extract_to_csv_rodillo
import extract_to_csv_agricola
import extract_to_csv_engranaje
import extract_to_csv_sellos_jaulas
from export_auxiliar_csv import export_aux_sheet


def load_config(path: Path) -> Dict:
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    
    # Determine Root Directory
    # Priority 1: Environment Variable
    env_root = os.environ.get("BPB_BASE_DIR")
    if env_root:
        root = Path(env_root)
    else:
        # Priority 2: Relative to config file (assuming in Codigos/)
        root = path.parent.parent

    def resolve(p: Optional[str]) -> Optional[Path]:
        if not p:
            return None
        pth = Path(p)
        if not pth.is_absolute():
            pth = root / pth
        return pth

    return {
        "processed_dir": resolve(raw.get("processed_dir") or raw.get("in process_dir")),
        "output_root": resolve(raw.get("output_root")),
        "registros_root": resolve(raw.get("registros_root_dir") or raw.get("register_root_dir") or "\\\\192.168.0.55\\utn\\REGISTROS"),
        "aux_indices_dir": resolve(raw.get("aux_indices_dir") or "Auxiliares\\indices_auxiliar"),
        "register_password": raw.get("register_password", "bpb"),
        "fabricas_csv": resolve(raw.get("fabricas_csv") or "Auxiliares\\Fabricas\\Listado Maestro de Codificacion Fabricas.csv"),
        "log_file": resolve(raw.get("logging", {}).get("file", "step3.log")),
        "log_level": str(raw.get("logging", {}).get("level", "INFO")).upper(),
    }


def setup_logging(log_file: Path, level: str) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    logging.basicConfig(
        level=getattr(logging, level, logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=[logging.FileHandler(log_file, encoding="utf-8"), logging.StreamHandler()],
    )


def latest_registro_dir(processed_dir: Path) -> Path:
    dirs = [p for p in processed_dir.iterdir() if p.is_dir() and "Registro - R" in p.name]
    if not dirs:
        raise FileNotFoundError("No se encontraron carpetas 'Registro - Rxxxx - fecha' en in process.")

    def seq(p: Path) -> int:
        m = re.search(r"R(\d{4})", p.name)
        return int(m.group(1)) if m else -1

    dirs.sort(key=seq, reverse=True)
    return dirs[0]


def read_rows(csv_path: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    with open(csv_path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for row in reader:
            rows.append(
                {
                    "code": (row.get("Codigo") or "").strip(),
                    "po": (row.get("PO") or "").strip(),
                    "obs": (row.get("Observac_PO") or "").strip(),
                }
            )
    return rows


def key_from_row(row: Dict[str, str]) -> Optional[str]:
    obs = row.get("obs", "")
    po = row.get("po", "")
    if obs:
        m = re.findall(r"\d+", obs)
        if m:
            return m[0]
    if po and po.lower() != "sin oc":
        m = re.findall(r"\d+", po)
        if m:
            return m[0]
    return None


def group_rows_by_key(rows: List[Dict[str, str]]) -> Dict[str, List[Dict[str, str]]]:
    grouped: Dict[str, List[Dict[str, str]]] = {}
    for r in rows:
        k = key_from_row(r)
        if not k:
            continue
        grouped.setdefault(k, []).append(r)
    return grouped


def norm_str(s: str) -> str:
    return re.sub(r"[^A-Z0-9]", "", s.upper())


def infer_type_from_name(path: Path) -> str:
    name = path.name.lower()
    if "bol" in name:
        return "bolas"
    if "rod" in name:
        return "rodillos"
    if "esp" in name:
        return "especiales"
    return "desconocido"


def infer_type_from_aux_name(name: str) -> str:
    s = (name or "").lower()
    if "inspeccion de especiales" in s:
        return "Especial"
    if "inspeccion rod. autocentrantes" in s:
        return "Autocentrante"
    if "inspeccion rod. agricola" in s:
        return "Agricola"
    if "inspeccion rod. de bolas" in s:
        return "Bolas"
    if "inspeccion rod. de rodillo" in s:
        return "Rodillo"
    if "bol" in s:
        return "Bolas"
    if "rod" in s:
        return "Rodillo"
    if "esp" in s:
        return "Especial"
    return "desconocido"


def detect_product_type(target: str, registros_root: Path) -> str:
    if not registros_root or not registros_root.exists() or openpyxl is None:
        return "desconocido"
    best_score = -1.0
    best_path = None
    target_norm = norm_str(target)
    for wb_path in registros_root.glob("R016-01*.xls*"):
        # saltar rutas que contengan obsoleto
        if any("obsole" in part.lower() for part in wb_path.parts):
            continue
        # Verify extension explicitly
        if wb_path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            continue
            
        if wb_path.suffix.lower() == '.xls':
             if xlrd is None: continue
             try:
                 wb_xls = xlrd.open_workbook(str(wb_path), formatting_info=False)
                 # Find Auxiliar sheet
                 sh_xls = None
                 for sname in wb_xls.sheet_names():
                     if sname and sname.strip().lower() == "auxiliar":
                         sh_xls = wb_xls.sheet_by_name(sname)
                         break
                 if sh_xls:
                     # iterate column 2 (index 1)
                     for rx in range(sh_xls.nrows):
                         row_vals = sh_xls.row_values(rx)
                         if len(row_vals) > 1:
                             cell = row_vals[1]
                             if not cell: continue
                             cell_norm = norm_str(str(cell))
                             if not cell_norm: continue
                             score = difflib.SequenceMatcher(None, target_norm, cell_norm).ratio()
                             if score > best_score:
                                 best_score = score
                                 best_path = wb_path
             except Exception:
                 pass
             continue
             
        # Else use openpyxl for xlsx/xlsm
        try:
            wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        except Exception:
            continue
        if "Auxiliar" not in wb.sheetnames:
            continue
        sh = wb["Auxiliar"]
        try:
            for cell in sh.iter_cols(min_col=2, max_col=2, values_only=True)[0]:
                if not cell:
                    continue
                cell_norm = norm_str(str(cell))
                if not cell_norm:
                    continue
                score = difflib.SequenceMatcher(None, target_norm, cell_norm).ratio()
                if score > best_score:
                    best_score = score
                    best_path = wb_path
        except Exception:
            pass
        finally:
            try:
                wb.close()
            except Exception:
                pass
    if best_path is None:
        return "desconocido"
    return infer_type_from_name(best_path)


def load_aux_indices_cache(indices_dir: Path) -> List[Dict[str, str]]:
    """
    Carga en memoria los valores de la columna B de todos los CSV de indices_auxiliar.
    Devuelve una lista de dicts: file, value, norm, tipo.
    """
    cache: List[Dict[str, str]] = []
    if not indices_dir or not indices_dir.exists():
        return cache
    for csv_path in indices_dir.glob("*.csv"):
        tipo = infer_type_from_aux_name(csv_path.name)
        try:
            with open(csv_path, newline="", encoding="utf-8", errors="ignore") as f:
                reader = csv.reader(f, delimiter=";")
                header_skipped = False
                for row in reader:
                    if not header_skipped:
                        header_skipped = True
                        continue
                    if len(row) < 2:
                        continue
                    val = str(row[1]).strip()
                    v_norm = norm_str(val)
                    if not v_norm:
                        continue
                    cache.append(
                        {
                            "file": csv_path.name,
                            "value": val,
                            "norm": v_norm,
                            "tipo": tipo,
                        }
                    )
        except Exception:
            continue
    return cache


def find_best_aux_match_csv(target: str, indices_dir: Path, cache: Optional[List[Dict[str, str]]] = None) -> Tuple[str, str, float]:
    """
    Busca coincidencia en los CSV de indices_auxiliar (columna B).
    Retorna (archivo_csv, valor, score)
    """
    entries = cache if cache is not None else load_aux_indices_cache(indices_dir)
    if not entries:
        return "", "", 0.0
    t_norm = norm_str(target)
    best_score = 0.0
    best_val = ""
    best_file = ""
    for row in entries:
        v_norm = row["norm"]
        ratio = difflib.SequenceMatcher(None, t_norm, v_norm).ratio()
        contain = t_norm in v_norm or v_norm in t_norm
        bonus = 0.5 if contain else 0.0
        if "especial" in (row.get("file") or "").lower():
            bonus += 0.05
        score = ratio + bonus
        if score > best_score:
            best_score = score
            best_val = row["value"]
            best_file = row["file"]
    return best_file, best_val, best_score


def _preprocess_pil(img: Image.Image) -> Image.Image:
    img = img.convert("L")
    img = img.point(lambda x: 255 if x > 200 else 0, "1").convert("L")
    bbox = img.getbbox()
    if bbox:
        x0, y0, x1, y1 = bbox
        m = 10
        x0 = max(x0 - m, 0)
        y0 = max(y0 - m, 0)
        x1 = min(x1 + m, img.width)
        y1 = min(y1 + m, img.height)
        img = img.crop((x0, y0, x1, y1))
    return img


def render_first_page_png(pdf_path: Path) -> Optional[bytes]:
    """
    Intenta renderizar la primera página a PNG.
    1) Usa pypdfium2 (rápido).
    2) Si falla, intenta pdfplumber como respaldo.
    """
    # 0. Check y lectura memoria (Robustez Red)
    try:
        if not pdf_path.exists() or pdf_path.stat().st_size == 0: return None
    except: return None
    
    import time, io
    file_bytes = None
    for i in range(3):
        try:
            file_bytes = pdf_path.read_bytes()
            break
        except: time.sleep(1.0)
        
    if not file_bytes: return None

    # 1. Primer intento: pdfplumber (Safe/Python) - PRIORIDAD ALTA
    if pdfplumber is not None:
        try:
            with pdfplumber.open(io.BytesIO(file_bytes)) as pdf:
                if not pdf.pages:
                    return None
                page = pdf.pages[0]
                pil_img = page.to_image(resolution=300).original
                pil_img = _preprocess_pil(pil_img)
                buf = io.BytesIO()
                pil_img.save(buf, format="PNG")
                return buf.getvalue()
        except Exception as exc:
            logging.warning("Fallo pdfplumber con %s: %s", pdf_path.name, exc)

    # 2. Respaldo: pdfium (Risky) - PRIORIDAD BAJA
    if pdfium is not None:
        try:
            doc = pdfium.PdfDocument(file_bytes)
            if len(doc) > 0:
                page = doc[0]
                bitmap = page.render(scale=3.0)
                img = bitmap.to_pil()
                img = _preprocess_pil(img)
                buf = io.BytesIO()
                img.save(buf, format="PNG")
                doc.close()
                return buf.getvalue()
        except Exception as exc:
            logging.warning("Fallo pdfium con %s: %s", pdf_path.name, exc)

    return None

def extract_text_first_page(pdf_path: Path) -> str:
    if pdfplumber is None:
        return ""
    try:
        with pdfplumber.open(pdf_path) as pdf:
            if not pdf.pages:
                return ""
            return pdf.pages[0].extract_text() or ""
    except Exception:
        return ""


def get_gemini_key() -> Optional[str]:
    env = (Path(__file__).parent / "gemini.key").read_text(encoding="utf-8").strip() if (Path(__file__).parent / "gemini.key").exists() else None
    if env:
        return env
    k = Path(__file__).parent / "gemini.key.txt"
    if k.exists():
        try:
            return k.read_text(encoding="utf-8").strip()
        except Exception:
            pass
    return None


def extract_notes_with_gemini(pdf_path: Path) -> Optional[Dict]:
    api_key = get_gemini_key()
    if not api_key:
        logging.error("GEMINI_API_KEY no disponible")
        return None

    img_bytes = render_first_page_png(pdf_path)
    if not img_bytes:
        logging.error("No se pudo generar imagen para %s", pdf_path)
        return None

    text_first = extract_text_first_page(pdf_path)
    prompt = (
        "Eres un extractor técnico de planos de rodamientos. Devuelve SOLO JSON plano (sin markdown) con EXACTAMENTE estos campos y estructura (igual al ejemplo existente):\n"
        "{\n"
        "  \"purchase_order\": \"<numero OC si aparece>\",\n"
        "  \"code\": \"<codigo exacto del campo Código/Code del rotulo, incluyendo prefijo numérico si existe ej 100-SAB211-55DG>\",\n"
        "  \"revision\": \"<letra o 0>\",\n"
        "  \"date\": \"dd/mm/yyyy\",\n"
        "  \"dimensions\": {\n"
        "     \"thickness\": {\"value_mm\": number, \"tol_plus_mm\": number, \"tol_minus_mm\": number},\n"
        "     \"outer_thickness\": {\"value_mm\": number, \"tol_plus_mm\": number, \"tol_minus_mm\": number},\n"
        "     \"inner_thickness\": {\"value_mm\": number, \"tol_plus_mm\": number, \"tol_minus_mm\": number},\n"
        "     \"outer_dia\": {\"value_mm\": number, \"tol_plus_mm\": number, \"tol_minus_mm\": number},\n"
        "     \"inner_dia\": {\"value_mm\": number, \"tol_plus_mm\": number, \"tol_minus_mm\": number}\n"
        "  },\n"
        "  \"radial_clearance\": {\"class\": \"C3\" u otra, \"range_um\": {\"min\": number, \"max\": number}},\n"
        "  \"axial_clearance\": null o {\"range_um\": {\"min\": number, \"max\": number}},\n"
        "  \"hardness_hrc\": {\"inner_plus\": number, \"inner_minus\": number, \"outer_plus\": number, \"outer_minus\": number},\n"
        "  \"seal_torque\": {\"low_Nm\": number, \"high_Nm\": number} (si aparece rango tipo 0.451-0.676 Nm),\n"
        "  \"lubricant\": {\"type\": \"texto\", \"fill\": \"texto\"} (fill puede incluir % y/o gramos en la misma cadena),\n"
        "  \"retainer\": {\"material\": \"Steel/Brass/...\"},\n"
        "  \"revision_history\": [{\"rev\":\"A/B/C...\", \"date\":\"dd/mm/yyyy\", \"by\":\"\"}, ...]\n"
        "}\n"
        "Reglas clave:\n"
        "- Usa mm y tolerancias con signo (ej -0.15). Si falta dato: null.\n"
        "- No inventes. Copia exactamente lo que veas (incluye PO y código marcado en pieza/rótulo).\n"
        "- Usa la última fecha/revisión si hay tabla de revisiones; incluye la lista completa en revision_history.\n"
        "- Para retainer, si ves 'Two steel cage' u otro texto con steel/brass, devuelve solo el material capitalizado.\n"
        "- Para lubricant, type: material base (ej 'Polyrex'), fill: texto con % y/o gramos tal como se ve.\n"
        "- Sin texto extra ni markdown, solo el JSON solicitado."
    )

    parts = [{"text": prompt}]
    parts.append({"inline_data": {"mime_type": "image/png", "data": base64.b64encode(img_bytes).decode()}})
    if text_first:
        parts.append({"text": f"TEXTO_PDFPLUMBER:\n{text_first[:6000]}"})

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    try:
        resp = requests.post(url, json={"contents": [{"parts": parts}]}, timeout=150)
        if resp.status_code != 200:
            logging.warning("Gemini status %s: %s", resp.status_code, resp.text[:400])
            return None
        data = resp.json()
        text_resp = data["candidates"][0]["content"]["parts"][0]["text"]
        m = re.search(r"\{.*\}", text_resp, re.S)
        payload = json.loads(m.group(0)) if m else json.loads(text_resp)
        return payload if isinstance(payload, dict) else None
    except Exception as exc:
        logging.warning("Fallo Gemini: %s", exc)
        return None


def normalize_retainer(material: str) -> str:
    s = material.lower()
    if "steel" in s:
        return "Steel"
    if "brass" in s:
        return "Brass"
    return material.strip()


def normalize_dimensions(dim: Dict) -> Dict:
    dims = dim or {}
    # Si outer/inner thickness existen, y thickness falta, rellenar thickness con outer
    if "thickness" not in dims:
        if "outer_thickness" in dims:
            dims["thickness"] = dims.get("outer_thickness")
    if "inner_thickness" not in dims and "outer_thickness" in dims:
        dims["inner_thickness"] = dims.get("outer_thickness")
    return dims


def load_existing_notes(pdf_path: Path) -> Optional[Dict]:
    """Busca un JSON existente junto al PDF (notes_extracted_gemini.json o <pdf>_notes.json)."""
    candidates = [
        pdf_path.with_name("notes_extracted_gemini.json"),
        pdf_path.with_name(f"{pdf_path.stem}_notes.json"),
        pdf_path.with_name(f"{pdf_path.stem}_notes_extracted_gemini.json"),
        pdf_path.with_name(f"{pdf_path.stem}_notes.json"),
    ]
    for c in candidates:
        if c.exists():
            try:
                return json.loads(c.read_text(encoding="utf-8"))
            except Exception:
                continue
    return None


def next_rev_path(base_path: Path) -> Path:
    """
    Si base_path existe, genera base_path- Rev.A, Rev.B, etc.
    """
    if not base_path.exists():
        return base_path
    letters = [chr(i) for i in range(ord("A"), ord("Z") + 1)]
    for letter in letters:
        cand = base_path.with_name(f"{base_path.stem} - Rev.{letter}{base_path.suffix}")
        if not cand.exists():
            return cand
    # fallback numerico
    idx = 1
    while True:
        cand = base_path.with_name(f"{base_path.stem} - Rev.{idx}{base_path.suffix}")
        if not cand.exists():
            return cand
        idx += 1


def load_factory_map(path: Path) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    if not path or not path.exists():
        return mapping
    try:
        with open(path, newline='', encoding='utf-8') as f:
            reader = csv.reader(f, delimiter=';')
            for row in reader:
                if len(row) >= 2:
                    key = str(row[0]).strip()
                    val = str(row[1]).strip()
                    if key:
                        mapping[key] = val
    except Exception as exc:
        logging.warning('No se pudo leer fabricas_csv %s: %s', path, exc)
    return mapping


def build_match_name(product_code: str, data: Dict, factory_map: Dict[str, str]) -> str:
    prov = str(data.get('proveedor') or '').strip()
    fac = factory_map.get(prov)
    return f"{product_code} - {fac}" if fac else product_code



def product_from_filename(stem: str) -> str:
    # Extrae el codigo de producto desde el nombre del PDF
    name = stem
    m = re.search(r'^PO\d+\s*[-_ ]\s*(.+)$', name, re.I)
    if m:
        name = m.group(1)
    name = re.split(r'\s*[-_ ]?\s*rev\b', name, flags=re.I)[0]
    return name.strip(' -_')

def ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def extract_product_key_from_pdf_name(name: str) -> str:
    """
    Desde un nombre de PDF tipo 'PO1291-51310-Rev B' devuelve '51310'.
    Si no hay 'Rev', devuelve lo que sigue al primer '-'.
    """
    stem = Path(name).stem
    m = re.search(r"^PO\d+\s*[-_ ]\s*(.+?)\s*[-_ ]?\s*rev\b", stem, flags=re.I)
    if m:
        return m.group(1).strip(" -_")
    # fallback: todo luego del primer '-'
    parts = stem.split("-")
    return "-".join(parts[1:]).strip() if len(parts) > 1 else stem


def pick_latest_json_for_pdf(pdf: Path, json_dir: Path) -> Optional[Path]:
    """
    Busca el json de ese PDF en json_dir, eligiendo la revisión más alta.
    Formatos esperados:
      <pdf.stem>_notes_extracted_gemini.json
      <pdf.stem>_notes_extracted_gemini - Rev.A.json
      <pdf.stem>_notes_extracted_gemini - Rev.B.json
    """
    prefix = f"{pdf.stem}_notes_extracted_gemini"
    best_path = None
    best_rank = -1
    pattern = re.compile(rf"^{re.escape(prefix)}(?: - Rev\.([A-Za-z0-9]+))?\.json$", re.I)
    for cand in json_dir.glob(f"{prefix}*.json"):
        m = pattern.match(cand.name)
        if not m:
            continue
        rev = m.group(1)
        if rev is None:
            rank = 0
        elif len(rev) == 1 and rev.isalpha():
            rank = ord(rev.upper()) - ord("A") + 1
        else:
            try:
                rank = int(rev)
            except Exception:
                rank = 1
        if rank > best_rank:
            best_rank = rank
            best_path = cand
    return best_path


def load_aux_matches(aux_csv_dir: Path) -> List[Dict[str, str]]:
    rows: List[Dict[str, str]] = []
    path = aux_csv_dir / "aux_matches.csv"
    if not path.exists():
        return rows
    with open(path, newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter=";")
        for r in reader:
            rows.append(r)
    return rows


def best_aux_row(target: str, rows: List[Dict[str, str]]) -> Optional[Dict[str, str]]:
    """
    Busca la mejor fila del aux_matches comparando contra columna 'Codigo'.
    Preferimos que contenga el target; se usa ratio como desempate.
    """
    t_norm = norm_str(target)
    best = None
    best_score = -1.0
    for r in rows:
        code = r.get("Codigo") or ""
        c_norm = norm_str(code)
        if not c_norm:
            continue
        contain = t_norm in c_norm or c_norm in t_norm
        ratio = difflib.SequenceMatcher(None, t_norm, c_norm).ratio()
        score = ratio + (0.5 if contain else 0.0)
        if score > best_score:
            best_score = score
            best = r
    return best


def adjust_code_with_provider(base_code: str, provider_code: str, factory_map: Dict[str, str]) -> str:
    """
    Reemplaza el sufijo (último segmento tras '-') por el código de fábrica (letras) si no coincide.
    provider_code: código numérico del JSON (ej: '195').
    factory_map: mapa {numero: letras} desde Listado Maestro de Codificacion Fabricas.
    """
    if not base_code:
        return base_code
    if not provider_code:
        return base_code
    letters = factory_map.get(str(provider_code)) if factory_map else None
    if not letters:
        return base_code

    parts = base_code.split("-")
    if len(parts) < 2:
        return base_code
    
    suffix = parts[-1].strip()
    new_letters = str(letters).strip()
    
    # 1. Exact match check
    if suffix.lower() == new_letters.lower():
        return base_code
        
    # 2. Substring check: If the new provider (e.g. "HL") is just a part of the existing suffix (e.g. "HLBR"),
    # we assume the existing suffix is more specific/correct and keep it.
    if new_letters.lower() in suffix.lower():
        return base_code
        
    parts[-1] = new_letters
    return "-".join(parts)


def generate_csv_from_json(product_code: str, data: Dict, tipo: str, out_path: Path) -> None:
    """
    Despacha al extractor correcto según tipo.
    """
    t = (tipo or "").lower()
    if "bola" in t or "autocentr" in t:
        row = extract_to_csv_bolas_fixed.build_row(product_code, data)
        extract_to_csv_bolas_fixed.write_csv(out_path, row)
    elif "especial" in t:
        row = extract_to_csv_especial.build_row(product_code, data)
        extract_to_csv_especial.write_csv(out_path, row)
    elif "agric" in t:
        row = extract_to_csv_agricola.extract_row(product_code, data)
        extract_to_csv_agricola.write_csv(out_path, row)
    elif "engran" in t:
        row = extract_to_csv_engranaje.extract_row(product_code, data)
        extract_to_csv_engranaje.write_csv(out_path, row)
    elif "sello" in t or "jaula" in t:
        row = extract_to_csv_sellos_jaulas.extract_row(product_code, data)
        extract_to_csv_sellos_jaulas.write_csv(out_path, row)
    elif "rodillo" in t:
        row = extract_to_csv_rodillo.extract_row(product_code, data)
        extract_to_csv_rodillo.write_csv(out_path, row)
    else:  # default otros
        row = extract_to_csv.extract_row(product_code, data)
        extract_to_csv.write_csv(out_path, row)


def generate_csvs_for_po(po_dir: Path, factory_map: Dict[str, str]) -> None:
    """
    Para cada PDF en la PO:
      - Elegir el JSON con mayor revisión.
      - Mapear al aux_matches para saber tipo de producto.
      - Generar CSV en carpeta csv/.
    """
    json_dir = po_dir / "json"
    aux_csv_dir = po_dir / "csv_Auxiliar"
    out_csv_dir = po_dir / "csv"
    ensure_dir(out_csv_dir)

    aux_rows = load_aux_matches(aux_csv_dir)
    if not aux_rows:
        logging.warning("Sin aux_matches en %s; se omiten CSV finales.", po_dir)
        return

    for pdf in [p for p in po_dir.iterdir() if p.suffix.lower() == ".pdf"]:
        product_key = extract_product_key_from_pdf_name(pdf.name)
        json_path = pick_latest_json_for_pdf(pdf, json_dir)
        if not json_path or not json_path.exists():
            logging.warning("No se encontró JSON para %s", pdf.name)
            continue
        try:
            data = json.loads(json_path.read_text(encoding="utf-8"))
        except Exception:
            logging.warning("No se pudo leer JSON %s", json_path)
            continue

        # Separar proveedor en caso de estar embebido en code
        raw_code = data.get("code") or ""
        m_code = re.match(r"^(\d{3})[- ]?(.*)$", str(raw_code).strip())
        if m_code:
            data["proveedor"] = m_code.group(1)
            data["code"] = m_code.group(2)

        match_row = best_aux_row(product_key, aux_rows)
        if not match_row:
            logging.warning("Sin match auxiliar para %s", product_key)
            continue
        aux_name = match_row.get("Archivo_Auxiliar") or ""
        tipo = match_row.get("Tipo") or infer_type_from_aux_name(aux_name)
        base_code = match_row.get("Coincidencia") or product_key
        code_for_csv = adjust_code_with_provider(base_code, data.get("proveedor"), factory_map)

        out_path = out_csv_dir / f"{pdf.stem}.csv"
        if out_path.exists():
            logging.info("CSV ya existe para %s, se omite.", pdf.name)
            continue
        try:
            generate_csv_from_json(code_for_csv, data, tipo, out_path)
            logging.info("CSV generado para %s (%s) en %s", pdf.name, tipo, out_path)
        except Exception as exc:
            logging.warning("No se pudo generar CSV para %s: %s", pdf.name, exc)


def main():
    parser = argparse.ArgumentParser(description="Paso 3: generar json/csv por PDF desde el último resumen.")
    parser.add_argument("--config", default="config.yaml", help="Ruta al config.yaml")
    args = parser.parse_args()

    cfg = load_config(Path(args.config).expanduser())
    setup_logging(cfg["log_file"], cfg["log_level"])

    processed_dir = cfg["processed_dir"]
    output_root = cfg["output_root"]
    registros_root = cfg["registros_root"]
    if not processed_dir or not processed_dir.exists():
        raise FileNotFoundError("processed_dir no encontrado en config.")
    if not output_root or not output_root.exists():
        raise FileNotFoundError("output_root no encontrado en config.")

    # Exportar snapshot de Auxiliar en este paso
    aux_dir = cfg.get("aux_indices_dir")
    if registros_root and aux_dir:
        try:
            aux_dir.mkdir(parents=True, exist_ok=True)
            for wb_path in registros_root.glob("R016-01*.xls*"):
                if wb_path.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
                    continue
                export_aux_sheet(wb_path, aux_dir, password=cfg.get("register_password", "bpb"))
        except Exception as exc:
            logging.warning("No se pudo exportar auxiliares en paso 3: %s", exc)

    aux_cache = load_aux_indices_cache(aux_dir) if aux_dir else []
    factory_map = load_factory_map(cfg.get("fabricas_csv"))

    csv_paths = []
    reg_dirs = [p for p in processed_dir.iterdir() if p.is_dir() and "Registro - R" in p.name]
    for rd in reg_dirs:
        cp = rd / "resumen.csv"
        if cp.exists():
            csv_paths.append(cp)

    if not csv_paths:
        raise FileNotFoundError(f"No se encontraron archivos resumen.csv en {processed_dir}")

    # Build global grouped rows from ALL summaries
    grouped = {} # PO -> list of rows
    all_rows = []
    for cp in csv_paths:
        r = read_rows(cp)
        # Unite with existing grouped
        g = group_rows_by_key(r)
        for k, v in g.items():
            if k not in grouped:
                grouped[k] = []
            grouped[k].extend(v)
        all_rows.extend(r)
        
    logging.info("Claves encontradas (Global): %s", list(grouped.keys()))

    entrada_dir = output_root / "Entrantes"
    destino_final_dir = output_root / "En Progreso"
    destino_final_dir.mkdir(parents=True, exist_ok=True)

    def _process_pdf(pdf: Path, rowlist: List[Dict[str, str]], po_dir: Path) -> Optional[List[str]]:
        try:
            prod_guess = product_from_filename(pdf.stem)
            if not prod_guess and rowlist:
                prod_guess = rowlist[0].get("code") or pdf.stem
            product_code = prod_guess or pdf.stem
            data = load_existing_notes(pdf)
            regen = False
            if not data:
                data = extract_notes_with_gemini(pdf)
                regen = True
            if not data:
                logging.warning("Sin datos para %s", pdf)
                return None
            dims = normalize_dimensions(data.get("dimensions") or {})
            data["dimensions"] = dims
            if "retainer" in data and isinstance(data["retainer"], dict):
                mat = data["retainer"].get("material")
                if mat:
                    data["retainer"]["material"] = normalize_retainer(mat)
            raw_code = data.get("code") or ""
            m_code = re.match(r"^(\d{3})[- ]?(.*)$", str(raw_code).strip())
            if m_code:
                data["proveedor"] = m_code.group(1)
                data["code"] = m_code.group(2)
            match_target = build_match_name(product_code, data, factory_map)
            file_name, match_val, score = find_best_aux_match_csv(match_target, cfg.get("aux_indices_dir") or Path(""), cache=aux_cache)
            prod_type = infer_type_from_aux_name(file_name) if file_name else detect_product_type(match_target, registros_root)
            json_base = (po_dir / "json") / f"{pdf.stem}_notes_extracted_gemini.json"
            json_path = next_rev_path(json_base) if json_base.exists() else json_base
            json_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
            logging.info("Generado json para %s en %s (regen=%s, tipo=%s)", pdf.name, po_dir, regen, prod_type)
            return [product_code, match_target, file_name, match_val, f"{score:.3f}", prod_type]
        except Exception as exc:
            logging.warning("No se pudo procesar %s: %s", pdf.name, exc)
            return None

    # Procesar todo lo que haya en Entrantes
    po_dirs = [p for p in entrada_dir.iterdir() if p.is_dir() and p.name.startswith("PO")]
    po_names = [p.name for p in po_dirs]
    logging.info("POs a procesar (Entrantes): %s", po_names)

    for po_name in po_names:
        rowlist = grouped.get(po_name.replace("PO", "", 1), [])
        po_dir = entrada_dir / po_name
        if not po_dir.exists():
            logging.warning("Carpeta %s no existe; se omite.", po_dir)
            continue
        json_dir = po_dir / "json"
        aux_csv_dir = po_dir / "csv_Auxiliar"
        ensure_dir(json_dir)
        ensure_dir(aux_csv_dir)
        pdfs = [p for p in po_dir.iterdir() if p.suffix.lower() == ".pdf"]
        if not pdfs:
            logging.warning("No hay PDFs en %s; se omite esta PO", po_dir)
            continue

        existing_rows = load_aux_matches(aux_csv_dir)
        existing_codes = {norm_str(r.get("Codigo", "")) for r in existing_rows}
        new_rows: List[List[str]] = []

        def _should_skip(pdf: Path, product_key: str) -> bool:
            if norm_str(product_key) in existing_codes:
                return True
            # Si ya existe algún json para este PDF, no regenerar
            prefix = f"{pdf.stem}_notes_extracted_gemini"
            if any(json_dir.glob(f"{prefix}*.json")):
                return True
            return False

        # Estrategia Concurrencia: Aumentamos a 8 workers (pdfplumber es estable)
        failed_pdfs = []
        
        with ThreadPoolExecutor(max_workers=8) as ex:
            # Map future to PDF to track failures
            future_to_pdf = {}
            for pdf in pdfs:
                product_key = extract_product_key_from_pdf_name(pdf.name)
                if _should_skip(pdf, product_key):
                    logging.info("Se omite %s (ya procesado)", pdf.name)
                    continue
                
                fut = ex.submit(_process_pdf, pdf, rowlist, po_dir)
                future_to_pdf[fut] = pdf
            
            # Process results
            for fut in as_completed(future_to_pdf):
                pdf = future_to_pdf[fut]
                try:
                    res = fut.result()
                    if res:
                        new_rows.append(res)
                except Exception as exc:
                    logging.warning("Fallo en %s con 4 workers: %s. Se agrupa para reintento secuencial.", pdf.name, exc)
                    failed_pdfs.append(pdf)

        # Fallback: Retry failed PDFs sequentially (1 worker equivalent)
        if failed_pdfs:
            logging.info("Iniciando reintento secuencial para %d archivos...", len(failed_pdfs))
            for pdf in failed_pdfs:
                try:
                    # Llamada directa sincrónica
                    res = _process_pdf(pdf, rowlist, po_dir)
                    if res:
                        new_rows.append(res)
                        logging.info("Reintento exitoso para %s", pdf.name)
                except Exception as exc:
                    logging.error("Fallo definitivo en %s: %s", pdf.name, exc)

        if new_rows:
            # combinar filas existentes + nuevas (evitando duplicados por Codigo)
            combined = existing_rows.copy()
            existing_codes_lower = {norm_str(r.get("Codigo", "")) for r in existing_rows}
            for r in new_rows:
                if norm_str(r[0]) in existing_codes_lower:
                    continue
                combined.append({
                    "Codigo": r[0],
                    "Match_Target": r[1],
                    "Archivo_Auxiliar": r[2],
                    "Coincidencia": r[3],
                    "Score": r[4],
                    "Tipo": r[5],
                })
            out_matches = aux_csv_dir / "aux_matches.csv"
            with open(out_matches, "w", newline="", encoding="utf-8") as f:
                w = csv.writer(f, delimiter=";")
                w.writerow(["Codigo", "Match_Target", "Archivo_Auxiliar", "Coincidencia", "Score", "Tipo"])
                for r in combined:
                    if isinstance(r, dict):
                        w.writerow([r.get("Codigo"), r.get("Match_Target"), r.get("Archivo_Auxiliar"), r.get("Coincidencia"), r.get("Score"), r.get("Tipo")])
                    else:
                        w.writerow(r)

        # Generar CSV finales solo para PDFs nuevos (si el CSV ya existe, no se reescribe)
        try:
            generate_csvs_for_po(po_dir, factory_map)
        except Exception as exc:
            logging.warning("No se pudieron generar CSV finales en %s: %s", po_dir, exc)

        # Inicializar approved_items vacío si no existe
        approval_file = po_dir / "approval_info.json"
        if not approval_file.exists():
            try:
                base_meta = {
                    "status": "Pendiente",
                    "approved_items": [],
                    "created_at": datetime.now().strftime("%d/%m/%Y %H:%M")
                }
                approval_file.write_text(json.dumps(base_meta, indent=4), encoding="utf-8")
            except Exception as e:
                logging.warning("No se pudo crear approval_info.json: %s", e)

        # MOVER A EN PROGRESO (Destino Final)
        try:
            final_path = destino_final_dir / po_name
            if final_path.exists():
                logging.info("Destino %s existe, se reemplaza.", final_path)
                shutil.rmtree(final_path)
            shutil.move(str(po_dir), str(final_path))
            logging.info("PO MOVIDA A FINAL: %s -> %s", po_dir, final_path)
        except Exception as exc:
            logging.error("FALLO AL MOVER %s a %s: %s", po_dir, destino_final_dir, exc)

if __name__ == "__main__":
    main()
