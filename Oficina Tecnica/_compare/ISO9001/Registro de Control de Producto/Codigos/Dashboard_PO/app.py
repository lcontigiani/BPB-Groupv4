import eventlet
eventlet.monkey_patch()
from flask import Flask, render_template, jsonify, request, send_from_directory, session, redirect, url_for, send_file, make_response
from flask_socketio import SocketIO, emit, join_room, leave_room
import heapq

from decimal import Decimal

import os

import json

import csv

import time

import subprocess

import shutil
import threading
from datetime import datetime, timedelta, date
import tempfile

from pathlib import Path

from io import StringIO

from werkzeug.utils import secure_filename

from werkzeug.security import check_password_hash, generate_password_hash

from werkzeug.security import check_password_hash, generate_password_hash

import re

import openpyxl
import smtplib
import uuid
import secrets
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Logistics Solver Import
from logistics_solver import Packer, Bin, Item, RotationType

def _do_pack_internal(container_data, items_data, config):
    """Core Logistic Solver logic extracted for reuse in maximization."""
    print(f"DEBUG: _do_pack_internal called. Items: {len(items_data)}")
    try:
        pallet_type_key = config.get('pallet_type', 'none')
        container_type_key = config.get('container_type', '20ft')
        user_max_pallet_h = float(config.get('max_pallet_height', 0))
        
        PALLET_DIMS = {
            'europallet': {'w': 120, 'd': 80, 'h': 15, 'weight': 25},
            'american': {'w': 120, 'd': 100, 'h': 15, 'weight': 25}
        }
        
        COLLARS_TYPES = ('collars', 'collars_120x100')
        is_collars = pallet_type_key in COLLARS_TYPES
        allow_stacking = config.get('stack_load', True)

        # Forced orientation (face down) for items
        forced_rotations = None
        if config.get('force_orientation'):
            face = str(config.get('orientation_face', '') or '').strip().lower()
            if face in ('lxa', 'largo x ancho', 'largo_x_ancho'):
                forced_rotations = [RotationType.RT_WHD, RotationType.RT_DHW]
            elif face in ('lxh', 'largo x alto', 'largo_x_alto'):
                forced_rotations = [RotationType.RT_WDH, RotationType.RT_HDW]
            elif face in ('axh', 'ancho x alto', 'ancho_x_alto'):
                forced_rotations = [RotationType.RT_DWH, RotationType.RT_HWD]
        
        # 0. Safety Factors & Item Preparation
        sf_percent_dims = float(config.get('safety_factor_dims', 0))
        sf_percent_weight = float(config.get('safety_factor_weight', 0))
        
        sf_mult_dims = Decimal(1.0 + (sf_percent_dims / 100.0))
        sf_mult_weight = Decimal(1.0 + (sf_percent_weight / 100.0))
        
        # 1. Prepare Items
        # 1. Prepare Items
        # --- OPTIMIZATION: Heuristic Cap for Maximization ---
        max_items_cap = float('inf')
        if config.get('maximize') and items_data and not config.get('skip_max_cap'):
            try:
                # Estimate Volume Limit
                c_vol = 0
                if container_type_key != 'none':
                    c_vol = float(container_data.get('width', 0)) * float(container_data.get('height', 0)) * float(container_data.get('depth', 0))
                else:
                    # Single Pallet Mode: Use safe upper bound dims
                    cw = 120
                    cd = 80
                    ch = 250 # Safe upper bound height
                    if is_collars:
                        cd = 100 if pallet_type_key == 'collars_120x100' else 80
                    elif pallet_type_key in PALLET_DIMS:
                        cw = PALLET_DIMS[pallet_type_key]['w']
                        cd = PALLET_DIMS[pallet_type_key]['d']
                    c_vol = float(cw) * float(cd) * float(ch)

                min_vol = float('inf')
                for itm in items_data:
                    v = float(itm.get('w', 1)) * float(itm.get('h', 1)) * float(itm.get('d', 1))
                    if v > 0 and v < min_vol: min_vol = v
                
                if c_vol > 0 and min_vol > 0 and min_vol != float('inf'):
                    max_items_cap = int((c_vol / min_vol) * 1.5) + 200 # 50% buffer + 200 items base safety
            except Exception as e:
                print(f"DEBUG: Optimization Error: {e}")
                pass

        raw_items = []
        total_items_generated = 0
        
        for item in items_data:
            qty = int(item.get('qty', 1))
            
            # Apply Cap
            if config.get('maximize') and (total_items_generated + qty) > max_items_cap:
                qty = max(0, max_items_cap - total_items_generated)
            
            if qty <= 0 and config.get('maximize'): continue
            
            for _ in range(qty):
                 itm = Item(
                    name=item.get('id'),
                    width=Decimal(item.get('w', 0)) * sf_mult_dims,
                    height=Decimal(item.get('h', 0)) * sf_mult_dims,
                    depth=Decimal(item.get('d', 0)) * sf_mult_dims,
                    weight=Decimal(item.get('weight', 0)) * sf_mult_weight,
                    allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                )
                 if forced_rotations:
                      itm.force_orientation = True
                 raw_items.append(itm)
            total_items_generated += qty
            if config.get('maximize') and total_items_generated >= max_items_cap: break

        items_to_pack_into_container = []
        unfitted_from_palletizing = []
        
        p_conf = None 
        p_conf_outer = None 
        is_stackable = True 
        
        if pallet_type_key != 'none':
            custom_dims = config.get('pallet_dims')
            if is_collars:
                boards = int(config.get('boards_count', 4))
                board_h_cm = Decimal('19.5')
                deck_h_cm = Decimal('0.9')
                base_h_cm = Decimal('15.0') + deck_h_cm
                wall_thick_cm = Decimal('1.9') 
                total_h = base_h_cm + (Decimal(boards) * board_h_cm)
                
                # Default preset weight for collars
                p_weight = Decimal('50')
                # OVERRIDE: If user provided a specific weight in UI, use it.
                if custom_dims and custom_dims.get('weight', 0) > 0:
                     p_weight = Decimal(custom_dims['weight'])
                
                outer_w = Decimal('120')
                outer_d = Decimal('100') if pallet_type_key == 'collars_120x100' else Decimal('80')

                p_conf = {
                    'w': outer_w - (wall_thick_cm * 2),
                    'd': outer_d - (wall_thick_cm * 2),
                    'h': total_h,
                    'loading_h': (Decimal(boards) * board_h_cm),
                    'weight': p_weight
                }
                p_conf_outer = {'w': outer_w, 'd': outer_d, 'h': total_h, 'weight': p_weight}
                is_stackable = True
            elif custom_dims and custom_dims.get('w', 0) > 0 and custom_dims.get('d', 0) > 0:
                p_conf = {'w': Decimal(custom_dims['w']), 'd': Decimal(custom_dims['d']), 'h': Decimal(custom_dims.get('h', 15)), 'loading_h': Decimal(0), 'weight': Decimal(custom_dims.get('weight', 25))}
                p_conf_outer = p_conf.copy()
                is_stackable = False 
            elif pallet_type_key in PALLET_DIMS:
                preset = PALLET_DIMS[pallet_type_key]
                # Default preset weight
                p_weight = Decimal(preset['weight'])
                # OVERRIDE: If user provided a specific weight in UI, use it.
                if custom_dims and custom_dims.get('weight', 0) > 0:
                     p_weight = Decimal(custom_dims['weight'])
                
                p_conf = {'w': Decimal(preset['w']), 'd': Decimal(preset['d']), 'h': Decimal(preset['h']), 'loading_h': Decimal(0), 'weight': p_weight}
                p_conf_outer = p_conf.copy()
                is_stackable = False 

        if p_conf:
            # FORCE NON-STACKABLE if Single Pallet Maximization (No Container + Maximize)
            # This ensures only ONE pallet is filled logic-wise, preventing "Collars" from stacking
            if container_type_key == 'none' and config.get('maximize'):
                is_stackable = False

            cont_h = Decimal(container_data.get('height', 0)) if container_type_key != 'none' else Decimal(0)
            if p_conf.get('loading_h', 0) > 0:
                 pallet_loading_height = p_conf['loading_h']
            else:
                if cont_h > 0: physical_limit_load = cont_h - p_conf['h']
                else: physical_limit_load = Decimal(250)
                user_max_h_dec = Decimal(user_max_pallet_h)
                if user_max_h_dec > 0:
                    user_limit_load = user_max_h_dec - p_conf['h']
                    pallet_loading_height = min(user_limit_load, physical_limit_load)
                else: 
                     # Fix for Efficiency > 100%:
                     # If user doesn't specify max height, we default efficiency calc to 180cm.
                     # So we must also limit the pallet loading to 180cm - base_h.
                     default_std_limit = Decimal(180) - p_conf['h']
                     pallet_loading_height = min(default_std_limit, physical_limit_load)
            
            user_max_w = Decimal(config.get('max_pallet_weight', 0))
            pallet_max_weight = user_max_w if user_max_w > 0 else Decimal(2000)
            
            # The Packer checks 'items weight' vs 'max_weight'.
            # But 'pallet_max_weight' is GROSS WEIGHT (Load + Base).
            # So we must subtract the base weight to get the allowed LOAD weight.
            pallet_base_weight = p_conf['weight']
            allowed_load_weight = pallet_max_weight - pallet_base_weight
            if allowed_load_weight < 0: allowed_load_weight = Decimal(0)

            def pallet_factory():
                return Bin(name=f"Pallet-{pallet_type_key}", width=p_conf['w'], depth=p_conf['d'], height=pallet_loading_height, max_weight=allowed_load_weight, allow_stacking=allow_stacking)
            
            # --- PALLETIZING LOGIC (Mixed vs Pure) ---
            filled_pallets = []
            if config.get('mixed_pallets', True):
                # Standard: Mix everything
                temp_packer = Packer()
                filled_pallets = temp_packer.pack_to_many_bins(pallet_factory, raw_items, sort_items=config.get('sort_items', True))
                unfitted_from_palletizing = temp_packer.unfit_items
            else:
                # No Mixing: Group by Item Name/ID and pack separately
                groups = {}
                for it in raw_items:
                    groups.setdefault(it.name, []).append(it)
                
                unfitted_from_palletizing = []
                # Pack each group
                for g_name, g_items in groups.items():
                    grp_packer = Packer()
                    grp_pallets = grp_packer.pack_to_many_bins(pallet_factory, g_items, sort_items=config.get('sort_items', True))
                    filled_pallets.extend(grp_pallets)
                    unfitted_from_palletizing.extend(grp_packer.unfit_items)

            # Pre-calculate Max Height for Single Pallet Maximization Logic
            target_total_h_for_max = Decimal(0)
            if container_type_key == 'none' and config.get('maximize'):
                if is_collars:
                    target_total_h_for_max = p_conf['h']
                else:
                    user_max_h = float(config.get('max_pallet_height', 0))
                    target_total_h_for_max = Decimal(user_max_h) if user_max_h > 0 else Decimal(180)

            for i, p_bin in enumerate(filled_pallets):
                max_h_used = Decimal(0)
                for pit in p_bin.items:
                     dims = pit.get_dimension()
                     top = Decimal(pit.position[1]) + Decimal(dims[1])
                     if top > max_h_used: max_h_used = top
                
                if is_collars: real_total_h = p_conf_outer['h']
                else: real_total_h = Decimal(p_conf['h']) + max_h_used

                if not is_stackable and cont_h > 0: packing_h = cont_h
                elif not is_stackable and cont_h == 0: 
                    # If Maximizing (Single Pallet Mode), do NOT inflate height, we need fit in strict bin
                    # CRITICAL FIX: To prevent stacking, the item must consume the ENTIRE vertical space
                    # of the Single Pallet Bin. So set packing_h = target_total_h_for_max.
                    if config.get('maximize'): 
                         packing_h = target_total_h_for_max
                         if packing_h <= 0: packing_h = real_total_h # Safety fallback
                    else: packing_h = real_total_h + Decimal(200)
                else: packing_h = real_total_h

                p_item = Item(name=f"Pallet_{i+1}", width=p_conf_outer['w'], depth=p_conf_outer['d'], height=packing_h, weight=Decimal(p_conf['weight']) + Decimal(p_bin.get_total_weight()), allowed_rotations=[RotationType.RT_WHD, RotationType.RT_DHW])
                p_item.inner_items = p_bin.items 
                if is_collars: p_item.base_height = Decimal('15.9') 
                else: p_item.base_height = p_conf['h']       
                p_item.visual_height = float(real_total_h)
                p_item.pallet_type = pallet_type_key
                # Tag it with the "Type" of content (if unmixed)
                # If unmixed, all items have same name. 
                if p_bin.items:
                     p_item.content_type = p_bin.items[0].name 
                else: p_item.content_type = 'unknown'

                items_to_pack_into_container.append(p_item)
        else:
            items_to_pack_into_container = raw_items
            # Tag raw items with their own name as content type
            for it in items_to_pack_into_container:
                it.content_type = it.name

        final_packer = Packer()
        if container_type_key != 'none':
            main_bin = Bin(name=container_data.get('name', 'Container'), width=Decimal(container_data.get('width', 0)), height=Decimal(container_data.get('height', 0)), depth=Decimal(container_data.get('depth', 0)), max_weight=Decimal(container_data.get('max_weight', 999999)), allow_stacking=allow_stacking)
            final_packer.add_bin(main_bin)
            
            # --- CONTAINER LOADING LOGIC (Mixed vs Pure) ---
            items_for_container = []
            remaining_unfitted = []

            if config.get('mixed_containers', True):
                 # Mix everything
                 items_for_container = items_to_pack_into_container
            else:
                 # No Mixing: Can only pack ONE type into this container.
                 groups = {}
                 for it in items_to_pack_into_container:
                      groups.setdefault(getattr(it, 'content_type', 'unknown'), []).append(it)
                 if groups:
                      first_group_key = list(groups.keys())[0]
                      items_for_container = groups[first_group_key]
                      for k, v in groups.items():
                           if k != first_group_key:
                                remaining_unfitted.extend(v)
            
            # --- START TEMPLATE LOGIC ---
            # Recognize combinations of Standard Container + Standard Pallet
            applied_template = False
            
            # Container Dims
            c_l = Decimal(container_data.get('width', 0))
            c_d = Decimal(container_data.get('depth', 0))
            
            # 1. Component Detection
            is_20ft = (585 <= c_l < 600 and 230 <= c_d < 240)
            is_40ft = (1200 <= c_l < 1210 and 230 <= c_d < 240)
            is_40pw = (1200 <= c_l < 1210 and 244 <= c_d < 255)
            
            p_w, p_d = Decimal(0), Decimal(0)
            is_europallet = False
            is_american = False
            
            if p_conf_outer:
                p_w, p_d = p_conf_outer['w'], p_conf_outer['d']
                is_europallet = (p_w == 120 and p_d == 80)
                is_american = (p_w == 120 and p_d == 100)
            
            # 2. Template Selection & Capacity
            tpl_name = None
            max_cap = 0
            
            # FOOTPRINT DETECTION: Europallet and Collars share 120x80.
            is_euro_footprint = (is_europallet or (is_collars and p_d == 80))
            
            if is_20ft:
                if is_american:   tpl_name, max_cap = "20ft_american", 10
                elif is_euro_footprint: tpl_name, max_cap = "20ft_europallet", 11
            elif is_40ft:
                if is_american:   tpl_name, max_cap = "40ft_american", 21
                elif is_euro_footprint: tpl_name, max_cap = "40ft_europallet", 25
            elif is_40pw:
                if is_american:   tpl_name, max_cap = "40pw_american", 24
                elif is_euro_footprint: tpl_name, max_cap = "40pw_europallet", 30

            # 3. Decision Logic
            is_max = config.get('maximize', False)
            load_count = len(items_for_container)
            
            if tpl_name:
                # If maximizing, we ALWAYS use the template to its capacity
                if is_max:
                    applied_template = True
                    target_count = max_cap
                else:
                    # Always use template when available.
                    # If load exceeds capacity, fill template and mark the rest as unfitted.
                    applied_template = True
                    target_count = min(load_count, max_cap)
            
            if applied_template:
                template_coords = []
                
                if tpl_name == "20ft_american":
                    # 10 american (5 sets of alternating pairs)
                    for i in range(5):
                        cycle = i // 2
                        is_second_in_cycle = i % 2 == 1
                        if not is_second_in_cycle:
                            template_coords.append({'x': Decimal(cycle * 220), 'z': Decimal(0), 'rot': RotationType.RT_DHW}) # 100x120
                            template_coords.append({'x': Decimal(cycle * 220), 'z': Decimal(135), 'rot': RotationType.RT_WHD}) # 120x100
                        else:
                            template_coords.append({'x': Decimal(cycle * 220 + 100), 'z': Decimal(0), 'rot': RotationType.RT_WHD}) # 120x100
                            template_coords.append({'x': Decimal(cycle * 220 + 120), 'z': Decimal(115), 'rot': RotationType.RT_DHW}) # 100x120
                
                elif tpl_name == "20ft_europallet":
                    # 11 europallets (7 side + 4 face)
                    for i in range(7):
                        template_coords.append({'x': Decimal(i * 80), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                    for i in range(4):
                        template_coords.append({'x': Decimal(i * 120), 'z': Decimal(155), 'rot': RotationType.RT_WHD})

                elif tpl_name == "40ft_american":
                    # 21 american (10 pairs + 1 end)
                    for i in range(10):
                        cycle = i // 2
                        is_second_in_cycle = i % 2 == 1
                        if not is_second_in_cycle:
                            template_coords.append({'x': Decimal(cycle * 220), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                            template_coords.append({'x': Decimal(cycle * 220), 'z': Decimal(135), 'rot': RotationType.RT_WHD})
                        else:
                            template_coords.append({'x': Decimal(cycle * 220 + 100), 'z': Decimal(0), 'rot': RotationType.RT_WHD})
                            template_coords.append({'x': Decimal(cycle * 220 + 120), 'z': Decimal(115), 'rot': RotationType.RT_DHW})
                    template_coords.append({'x': Decimal(1100), 'z': Decimal(0), 'rot': RotationType.RT_DHW})

                elif tpl_name == "40ft_europallet":
                    # 25 europallets
                    for i in range(15):
                        template_coords.append({'x': Decimal(i * 80), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                    for i in range(10):
                        template_coords.append({'x': Decimal(i * 120), 'z': Decimal(155), 'rot': RotationType.RT_WHD})

                elif tpl_name == "40pw_american":
                    # 24 american (2x12) - Both rows oriented 120 wide (RT_WHD)
                    for i in range(12):
                        template_coords.append({'x': Decimal(i * 100), 'z': Decimal(0), 'rot': RotationType.RT_DHW})   # 100x120
                        template_coords.append({'x': Decimal(i * 100), 'z': Decimal(125), 'rot': RotationType.RT_DHW}) # 100x120
                
                elif tpl_name == "40pw_europallet":
                    # 30 europallets (2x15)
                    for i in range(15):
                        template_coords.append({'x': Decimal(i * 80), 'z': Decimal(0), 'rot': RotationType.RT_DHW})   # 80x120
                        template_coords.append({'x': Decimal(i * 80), 'z': Decimal(125), 'rot': RotationType.RT_DHW}) # 80x120

                # STACKING LOGIC FOR MAXIMIZATION
                if applied_template and is_max and is_stackable and config.get('maximize') and allow_stacking:
                    # Calculate how many layers fit vertically
                    total_h_pallet = p_conf_outer['h'] # This is total height of one unit
                    container_h = Decimal(container_data.get('height', 0))
                    
                    if container_h > 0 and total_h_pallet > 0:
                        layers = int(container_h // total_h_pallet)
                        if layers > 1:
                            print(f"DEBUG: Stacking Detected. Layers: {layers}")
                            # Expand template_coords for multiple layers
                            base_coords = list(template_coords)
                            template_coords = []
                            
                            for layer in range(layers):
                                y_pos = Decimal(layer) * total_h_pallet
                                for coord in base_coords:
                                    # Clone coord and add Y
                                    new_coord = coord.copy()
                                    new_coord['y'] = y_pos
                                    template_coords.append(new_coord)
                            
                            # Update target count to reflect full capacity
                            target_count = len(template_coords)
                            print(f"DEBUG: New Target Count with Stacking: {target_count}")

                print(f"DEBUG: Applying Template {tpl_name} for {target_count} slots.")
                result_bin = main_bin
                packed_top_level = []
                unfitted_final = remaining_unfitted
                
                # Take items to fill target_count
                items_to_use = items_for_container[:target_count]
                
                # Default limiting factor is None (Optimistic). 
                # If we have more items than slots, then Volume is the limit.
                limiting_factor = 'volume' if len(items_for_container) > target_count else None

                current_total_weight = Decimal(0)
                container_max_w = Decimal(container_data.get('max_weight', 999999))
                if container_max_w <= 0: container_max_w = Decimal(999999999)

                for idx, coord in enumerate(template_coords):
                    if idx >= target_count: break
                    if idx >= len(items_to_use): break
                    
                    it = items_to_use[idx]
                    
                    # Weight Check
                    if current_total_weight + it.weight > container_max_w:
                        print(f"DEBUG: Weight Limit Reached! Cur: {current_total_weight} + Item: {it.weight} > Max: {container_max_w}")
                        limiting_factor = 'weight'
                        # Add remaining items to unfitted since we stopped early
                        if len(items_to_use) > idx:
                             unfitted_final.extend(items_to_use[idx:])
                        break

                    current_total_weight += it.weight
                    it.position = [coord['x'], coord.get('y', Decimal(0)), coord['z']]
                    it.rotation_type = coord['rot']
                    dims = it.get_dimension()
                    it.lx, it.rx = it.position[0], it.position[0] + dims[0]
                    it.ly, it.ry = it.position[1], it.position[1] + dims[1]
                    it.lz, it.rz = it.position[2], it.position[2] + dims[2]
                    
                    result_bin.items.append(it)
                    packed_top_level.append(it)
                
                # Any leftover items in items_for_container that didn't fit in template are unfitted
                if len(items_for_container) > len(template_coords):
                    unfitted_final.extend(items_for_container[len(template_coords):])
                    
            else:
                # Generic fallback packing
                for it in items_for_container: final_packer.add_item(it)
                final_packer.pack(bigger_first=True)
                result_bin = final_packer.bins[0]
                unfitted_final = final_packer.unfit_items + remaining_unfitted
                packed_top_level = result_bin.items
        else:
            # NO CONTAINER (Virtual Floor) - No mixed constraints apply (infinite space)
            if config.get('maximize') and p_conf_outer:
                # Check if this is multi-item maximization (each getting their own pallet)
                is_multi_item_max = not config.get('mixed_pallets', True)
                
                if is_multi_item_max:
                    # Multiple Items, Each Maximized on Own Pallet
                    # Use LARGE virtual floor to accommodate all pallets side by side
                    result_bin = Bin("Virtual Floor", Decimal(2000), Decimal(500), Decimal(2000), Decimal(999999), allow_stacking=allow_stacking)
                else:
                    # Single Pallet Maximization: Restrict bin to the pallet dimensions
                    
                    # COLLARS Case: Use fixed total height
                    if is_collars:
                         limit_h = p_conf['h'] # Total Height (Base + Walls)
                         # For packing, we need the bin to be big enough to hold the 'Pallet Item'
                         # The 'Pallet Item' created by pallet factory has height = total_h
                         # So bin height must be at least total_h.
                         result_bin = Bin("Single Pallet Limit", p_conf_outer['w'], limit_h, p_conf_outer['d'], Decimal(999999), allow_stacking=allow_stacking)
                    
                    else:
                        # STANDARD Case: Use User Max Height (Default 180cm)
                        user_max_h = float(config.get('max_pallet_height', 0))
                        target_total_h = Decimal(user_max_h) if user_max_h > 0 else Decimal(180)
                        
                        # For packing, Bin Height must be the TOTAL HEIGHT to accommodate the pallet + cargo
                        limit_h = target_total_h
                        
                        result_bin = Bin("Single Pallet Limit", p_conf_outer['w'] + Decimal(1), limit_h + Decimal(1), p_conf_outer['d'] + Decimal(1), Decimal(999999), allow_stacking=allow_stacking)
            else:
                result_bin = Bin("Virtual Floor", Decimal(2000), Decimal(500), Decimal(2000), Decimal(999999), allow_stacking=allow_stacking)
            final_packer.add_bin(result_bin)
            for it in items_to_pack_into_container: final_packer.add_item(it)
            final_packer.pack(bigger_first=True)
            result_bin = final_packer.bins[0]
            unfitted_final = final_packer.unfit_items
            packed_top_level = result_bin.items

        flattened_packed_items = []
        rot_map_90 = {RotationType.RT_WHD: RotationType.RT_DHW, RotationType.RT_DHW: RotationType.RT_WHD, RotationType.RT_HWD: RotationType.RT_DWH, RotationType.RT_DWH: RotationType.RT_HWD, RotationType.RT_HDW: RotationType.RT_WDH, RotationType.RT_WDH: RotationType.RT_HDW}

        for pkg in packed_top_level:
            pkg_dims = pkg.get_dimension()
            pkg_x, pkg_y, pkg_z = float(pkg.position[0]), float(pkg.position[1]), float(pkg.position[2])
            if hasattr(pkg, 'inner_items'):
                total_h = float(pkg_dims[1])
                base_obj = {
                    'name': 'PALLET_BASE', 
                    'x': pkg_x, 'y': pkg_y, 'z': pkg_z, 
                    'w': float(pkg_dims[0]), 'h': float(pkg.base_height), 'd': float(pkg_dims[2]), 
                    'total_h': total_h,
                    'rt': pkg.rotation_type, 
                    'is_pallet': True, 
                    'pallet_type': getattr(pkg, 'pallet_type', 'unknown')
                }
                if getattr(pkg, 'pallet_type', '') in ('collars', 'collars_120x100'): base_obj['h_visual'] = getattr(pkg, 'visual_height', total_h)
                flattened_packed_items.append(base_obj)
                is_rotated = (pkg.rotation_type == RotationType.RT_DHW)
                for inner in pkg.inner_items:
                    i_dims = inner.get_dimension()
                    ix, iy, iz = float(inner.position[0]), float(inner.position[1]), float(inner.position[2])
                    if is_rotated:
                        offset = 1.9 if getattr(pkg, 'pallet_type', '') in ('collars', 'collars_120x100') else 0
                        final_x, final_z, final_y = pkg_x + iz + offset, pkg_z + ix + offset, pkg_y + float(pkg.base_height) + iy
                        final_rt = rot_map_90.get(inner.rotation_type, inner.rotation_type)
                        final_w, final_h, final_d = float(i_dims[2]), float(i_dims[1]), float(i_dims[0])
                    else:
                        final_x = pkg_x + ix + (1.9 if getattr(pkg, 'pallet_type', '') in ('collars', 'collars_120x100') else 0)
                        final_z = pkg_z + iz + (1.9 if getattr(pkg, 'pallet_type', '') in ('collars', 'collars_120x100') else 0)
                        final_y = pkg_y + float(pkg.base_height) + iy
                        final_rt = inner.rotation_type
                        final_w, final_h, final_d = float(i_dims[0]), float(i_dims[1]), float(i_dims[2])
                    flattened_packed_items.append({'name': inner.name, 'x': final_x, 'y': final_y, 'z': final_z, 'w': final_w, 'h': final_h, 'd': final_d, 'rt': final_rt})
            else:
                flattened_packed_items.append({'name': pkg.name, 'x': pkg_x, 'y': pkg_y, 'z': pkg_z, 'w': float(pkg_dims[0]), 'h': float(pkg_dims[1]), 'd': float(pkg_dims[2]), 'rt': pkg.rotation_type})
        
        all_unfitted = unfitted_from_palletizing + unfitted_final
        unfitted_serialized = [{'name': i.name, 'w': float(i.width), 'h': float(i.height), 'd': float(i.depth)} for i in all_unfitted]

        vol_bin = float(result_bin.width * result_bin.height * result_bin.depth)
        
        # EFFICIENCY CALCULATION REFINEMENT
        if container_type_key == 'none' and p_conf_outer:
             # Override vol_bin to reflect NET AVAILABLE VOLUME
             if is_collars:
                  # Net Volume = Inner Width * Inner Depth * Inner Loading Height
                  # p_conf for collars has loading_h correctly defined as "boards * board_h"
                  # and w/d as inner dimensions.
                  vol_bin = float(p_conf['w'] * p_conf['d'] * p_conf['loading_h'])
             else:
                  # Standard Pallet:
                  # Net Volume = Outer Width * Outer Depth * (Total Height - Base Height)
                  # Re-read configs (safe)
                  user_max_h = float(config.get('max_pallet_height', 0))
                  target_total_h = Decimal(user_max_h) if user_max_h > 0 else Decimal(180)
                  base_h = p_conf.get('h', Decimal(15))
                  net_h = target_total_h - base_h
                  if net_h < 0: net_h = Decimal(1)
                  vol_bin = float(p_conf_outer['w'] * p_conf_outer['d'] * net_h)
            
             vol_items = 0
             for it in packed_top_level:
                if hasattr(it, 'inner_items'):
                    vol_items += sum([float(i.width * i.height * i.depth) for i in it.inner_items])
                else:
                    vol_items += float(it.width * it.height * it.depth)
        else:
            vol_items = sum([float(it.width * it.height * it.depth) for it in packed_top_level])
        efficiency = (vol_items / vol_bin * 100) if vol_bin > 0 else 0

        res = {
            'status': 'success',
            'bin_name': result_bin.name,
            'bin_dims': {'w': float(result_bin.width), 'h': float(result_bin.height), 'd': float(result_bin.depth)},
            'packed_items': flattened_packed_items,
            'unfitted_items': unfitted_serialized,
            'unfitted_count': len(all_unfitted),
            'unfitted_from_palletizing_count': len(unfitted_from_palletizing),
            'unfitted_final_count': len(unfitted_final),
            'limiting_factor': locals().get('limiting_factor', None),
            'efficiency': round(efficiency, 2),
            'packed_top_level_count': len(packed_top_level),
            # NEW KPIs
            'kpis': {
                'container_vol': round(efficiency, 2), # Already calculated
                'pallet_vol_avg': 0, # To be calc below
                'container_weight': 0,
                'pallet_weight_avg': 0
            },
            'container_info': {
                'length': float(container_data.get('length', container_data.get('depth', 0))),
                'width': float(container_data.get('width', 0)),
                'height': float(container_data.get('height', 0))
            },
            'grouped_pallets': []
        }

        # --- KPI & GROUPING LOGIC ---
        print(f"DEBUG: Starting KPI Logic. Packed Top Level Count: {len(packed_top_level)}")
        
        # 1. Container Weight KPI
        total_weight_loaded = sum([float(i.weight) for i in packed_top_level])
        container_max_w = float(result_bin.max_weight)
        print(f"DEBUG: Cont Max W: {container_max_w}, Total Loaded W: {total_weight_loaded}")
        
        if container_max_w > 0:
            res['kpis']['container_weight'] = round((total_weight_loaded / container_max_w) * 100, 2)

        # 2. Pallet Metrics (Vol Efficiency & Weight) & Grouping
        # distinct_pallets hash map -> { signature: {count: N, data: p_item, items: {name: qty}} }
        pallet_groups = {}
        
        pallet_vol_efficiencies = []
        pallet_weight_ratios = []

        # Iterate TOP LEVEL packed items (which are Pallets or Loose Items)
        packed_loose_items = []

        for p_item in packed_top_level:
            is_pallet = getattr(p_item, 'is_pallet', False) or hasattr(p_item, 'inner_items')
            print(f"DEBUG: Item {p_item.name} IsPallet: {is_pallet}")
            
            if is_pallet:
                # Calculate Efficiency for this individual pallet
                # Net Volume of Pallet (Inner)
                # p_conf uses 'h' as Base Height, but 'loading_h' is what we want for volume calc if collars?
                # Actually, easy way: Sum volume of items inside / (W*D*LoadingH)
                
                inner_vol = sum([float(i.width * i.height * i.depth) for i in p_item.inner_items])
                
                # Pallet Usable Volume
                # config['pallet_dims'] might not be persistent if we have multiple types, but here we usually have 1 type per calc
                # We can deduce usable volume from the p_item dimensions but p_item height might be total.
                # Let's rely on p_conf if available (it was defined above in _do_pack_internal scope)
                
                pallet_usable_vol = 1
                if p_conf:
                    # p_conf['h'] is Base Height, 'loading_h' is Loading Height (for Collars). 
                    # For standard pallets, loading_h is 0 in definition, calculated dynamically?
                    # Let's use the actual height of the pallet item minus base height?
                    # Or better: if collars, use fixed loading_h. If standard, use (VisualHeight - BaseHeight) or (MaxHeight - BaseHeight).
                    
                    if is_collars:
                         pallet_usable_vol = float(p_conf['w'] * p_conf['d'] * p_conf['loading_h'])
                    else:
                         # Standard: Use the "Max User Height" or "Physical Limit" defined as limit
                         # We calculated `pallet_loading_height` variable earlier!
                         # Accessing it might be tricky if scope is lost, but we are in same func.
                         # `pallet_loading_height` variable is available in local scope.
                         pallet_usable_vol = float(p_conf['w'] * p_conf['d'] * pallet_loading_height)

                if pallet_usable_vol > 0:
                    pallet_vol_efficiencies.append((inner_vol / pallet_usable_vol) * 100)
                
                # Weight Ratio
                curr_w = float(p_item.weight) # Total weight including tare
                # Max weight is in p_conf['max_weight']? No, user_max_w variable.
                # `pallet_max_weight` variable available.
                if pallet_max_weight > 0:
                    pallet_weight_ratios.append((curr_w / float(pallet_max_weight)) * 100)

                # Grouping Signature
                # Signature based on: Content (Items + Qty)
                # We can sort inner items by name/id and build a string
                content_map = {}
                for sub in p_item.inner_items:
                    content_map[sub.name] = content_map.get(sub.name, 0) + 1
                
                # Create sorted signature tuple
                sig_items = sorted(content_map.items()) # [('001', 50), ('002', 10)]
                sig_str = str(sig_items) 
                
                # Also include pallet type/dims in signature just in case
                sig = f"{pallet_type_key}_{p_item.width}x{p_item.depth}_{sig_str}"

                if sig not in pallet_groups:
                    # Calc Load Weight (Sum of inner items)
                    load_weight = sum(sub.weight for sub in p_item.inner_items)
                    
                    pallet_groups[sig] = {
                        'count': 0,
                        'name_display': f"Pallet {len(pallet_groups) + 1} ({p_item.width}x{p_item.depth})", # Placeholder name
                        'dims': f"{p_item.width}x{p_item.depth}x{p_item.height}", # Outer dims
                        'weight': float(p_item.weight), # Gross Weight
                        'load_weight': float(load_weight), # Net Load Weight
                        'items_map': content_map,
                        'base_height': float(getattr(p_item, 'base_height', 15))
                    }
                pallet_groups[sig]['count'] += 1
            
            else:
                # Loose top level item
                packed_loose_items.append({
                     'name': p_item.name,
                     'dims': f"{p_item.width}x{p_item.height}x{p_item.depth}",
                     'weight': float(p_item.weight),
                     'qty': 1
                })

        # Calculate Averages
        if pallet_vol_efficiencies:
            res['kpis']['pallet_vol_avg'] = round(sum(pallet_vol_efficiencies) / len(pallet_vol_efficiencies), 2)
        if pallet_weight_ratios:
             res['kpis']['pallet_weight_avg'] = round(sum(pallet_weight_ratios) / len(pallet_weight_ratios), 2)

        # Finalize Groups List
        final_groups = []
        
        # Add Pallets First
        for sig, data in pallet_groups.items():
            # Format Items for this group
            # We want: Name, Dims(lookup?), Qty Per Pallet, Total Qty
            group_items = []
            for it_name, qty_per in data['items_map'].items():
                # Need dimensions of the item. Only have name 'it_name'.
                
                it_info = next((x for x in items_data if x['id'] == it_name), None)
                dims_str = "???"
                unit_weight = 0
                if it_info:
                    dims_str = f"{it_info.get('w')}x{it_info.get('d')}x{it_info.get('h')}"
                    unit_weight = float(it_info.get('weight', 0))
                
                group_items.append({
                    'name': it_name,
                    'dims': dims_str,
                    'weight': unit_weight, # ADDED WEIGHT
                    'qty_per_pallet': qty_per,
                    'total_qty': qty_per * data['count']
                })
            
            final_groups.append({
                'type': 'pallet',
                'name': data['name_display'],
                'dims': data['dims'],
                'weight_per_pallet': round(data['weight'], 2), # Gross
                'load_weight_per_pallet': round(data['load_weight'], 2), # Net
                'count': data['count'],
                'items': group_items
            })
            
        # Add Loose Items (Grouped by name)
        if packed_loose_items:
            # Group loose items by name
            loose_grouped = {}
            for li in packed_loose_items:
                k = li['name']
                if k not in loose_grouped:
                    loose_grouped[k] = {'name': k, 'dims': li['dims'], 'weight': li['weight'], 'count': 0}
                loose_grouped[k]['count'] += 1
            
            # Add as a special "Loose" group? Or just individual line items at root?
            # User wants: "Si no hay pallet... se mantiene tal cual".
            # If we mixed pallets and loose items?
            # Let's add them as 'loose_item' type entries in the root list
            
            for k, lg in loose_grouped.items():
                final_groups.append({
                    'type': 'loose_item',
                    'name': lg['name'],
                    'dims': lg['dims'],
                    'weight_unit': lg['weight'], # Weight per unit
                    'total_qty': lg['count']
                })

        res['grouped_pallets'] = final_groups
        print(f"DEBUG: Final Grouped Pallets Count: {len(final_groups)}")
        print(f"DEBUG: Grouped Data: {json.dumps(final_groups, default=str)}")

        # 3. LIMITING FACTOR DETECTION
        # Logic: We can have multiple limiting factors.
        # 1. Violations (>100%) - Critical
        # 2. Bottlenecks (Unfitted items) - If no violations, what prevented packing?
        
        limiting_ids = []
        
        avg_p_w = res['kpis']['pallet_weight_avg']
        cont_w = res['kpis']['container_weight']
        avg_p_v = res['kpis']['pallet_vol_avg']
        cont_v = res['kpis']['container_vol']

        # 1. Check Violations (Hard Limits Exceeded)
        if avg_p_w > 100.0: limiting_ids.append('kpi-pallet-weight')
        if cont_w > 100.0: limiting_ids.append('kpi-container-weight')
        if avg_p_v > 100.0: limiting_ids.append('kpi-pallet-vol')
        if cont_v > 100.0: limiting_ids.append('kpi-container-vol')
        
        # 2. If no Hard Violations, check Soft Limits if items couldn't fit
        if len(limiting_ids) == 0 and len(all_unfitted) > 0:
            # Did we fail at Palletizing?
            if len(unfitted_from_palletizing) > 0:
                 # Pallet Bottleneck
                if avg_p_w > 90: limiting_ids.append('kpi-pallet-weight')
                else: limiting_ids.append('kpi-pallet-vol')
            
            # Did we fail at Container Loading? (Even if palletizing failed, container might also be full)
            # Logic: If we have unfitted final items (that were palletized but didn't fit in container)
            if len(unfitted_final) > 0:
                # Container Bottleneck
                if cont_w > 95: limiting_ids.append('kpi-container-weight')
                else: limiting_ids.append('kpi-container-vol')
        
        if container_type_key == 'none' and config.get('maximize'):
            # FIX: In Single Pallet Maximization, there is no container.
            # The "Container Vol/Weight" KPIs are actually measuring the Single Pallet (Bin).
            # But the UI Card for Container is hidden or irrelevant.
            # We must redirect the highlighted KPI to the Pallet Card.
            
            new_limits = []
            for limit_id in limiting_ids:
                if limit_id == 'kpi-container-vol':
                    new_limits.append('kpi-pallet-vol')
                elif limit_id == 'kpi-container-weight':
                    new_limits.append('kpi-pallet-weight')
                else:
                    new_limits.append(limit_id)
            limiting_ids = list(set(new_limits)) # Deduplicate

        res['limiting_kpis'] = limiting_ids # Return LIST
        
        return res
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'status': 'error', 'message': str(e)}



# Determine absolute path to this script's directory

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SCRIPT_DIR_PATH = Path(SCRIPT_DIR)
LOGISTICS_RECORDS_FILE = os.path.join(SCRIPT_DIR, 'logistics_records.json')

# Initialize Flask with explicit folder paths

# --- DEBUG STARTUP ---
try:
    with open("global_debug_startup.log", "a") as f:
        f.write(f"{datetime.now()}: App starting from {os.getcwd()}\n")
except: pass
# ---------------------

app = Flask(__name__, 
            template_folder=os.path.join(SCRIPT_DIR, 'templates'),
            static_folder=os.path.join(SCRIPT_DIR, 'static'))
socketio = SocketIO(app, async_mode='eventlet', cors_allowed_origins="*")

app.config['JSON_SORT_KEYS'] = False
app.config['JSON_AS_ASCII'] = False
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

app.secret_key = 'super_secret_key_bpb_group_2026'

# Force UTF-8 encoding in all responses
@app.after_request
def add_utf8_header(response):
    if 'Content-Type' in response.headers:
        content_type = response.headers['Content-Type']
        if 'charset' not in content_type.lower():
            response.headers['Content-Type'] = content_type + '; charset=utf-8'
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# CONFIGURATION

# SEARCH LOGIC FOR BASE_DIR
# 1. Environment Variable (Explicit Override)
env_base = os.environ.get("BPB_BASE_DIR")

if env_base:
    BASE_DIR = Path(env_base)
    print(f"DEBUG: Using Configured BASE_DIR (Env Var): {BASE_DIR}")

else:
    # 2. Try Network Path
    network_path = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto")
    
    # Check if network path is actually accessible (cheap check)
    network_available = False
    try:
        if network_path.exists():
            network_available = True
    except: pass

    if network_available:
        BASE_DIR = network_path
        print(f"DEBUG: Using Network BASE_DIR: {BASE_DIR}")
    else:
        # 3. Try Local Relative Path (Assuming script is in Codigos/Dashboard_PO)
        # We go up 2 levels: Codigos/Dashboard_PO -> Codigos -> Root
        local_fallback = SCRIPT_DIR_PATH.parent.parent
        
        # Verify if this looks like the real repo (check for P2 or Auxiliares)
        if (local_fallback / "P2 - Purchase Order").exists():
            BASE_DIR = local_fallback
            print(f"DEBUG: Network unavailable. Auto-detected Local BASE_DIR: {BASE_DIR}")
            
            # Auto-set env var for child processes
            os.environ["BPB_BASE_DIR"] = str(BASE_DIR)
        else:
            # 4. Total Failure -> Mock/Default (Will likely fail later but we keep the structure)
            BASE_DIR = network_path
            print("DEBUG: Could not detect authentic local paths. Defaulting to Network Path (which may fail).")

# Update SCRIPT_DIR usage to be consistent? 
# SCRIPT_DIR was defined as os.path.dirname... at top.


# --- DEBUG FORCE ABSOLUTE ---
try:
    force_log = BASE_DIR / "Codigos/debug_force.txt"
    with open(force_log, "a") as f:
        f.write(f"{datetime.now()}: APP STARTED. CWD={os.getcwd()}\n")
    print(f"!!! DEBUG LOG WRITTEN TO {force_log} !!!")
except Exception as e:
    print(f"!!! FAILED TO WRITE LOG: {e} !!!")
# ----------------------------

PRODUCTION_PATH = BASE_DIR / "P2 - Purchase Order/En progreso"

PROCESSED_PATH = BASE_DIR / "P2 - Purchase Order/Procesado"

UPLOAD_FOLDER = BASE_DIR / "P1 - Registros Solicitados"

IN_PROCESS_DIR = UPLOAD_FOLDER / "in process"

USERS_FILE = BASE_DIR / "Codigos/Usuarios/users.json"
NOTIFICATIONS_FILE = BASE_DIR / "Datos/notifications.json"

SOLIDS_DIR = str(BASE_DIR / "Solidos")
PROJECTS_DB_PATH = str(BASE_DIR / "Registro de Proyectos" / "projects.json")

def load_db():
    if not os.path.exists(PROJECTS_DB_PATH):
        return {"projects": {}}
    try:
        with open(PROJECTS_DB_PATH, 'r', encoding='utf-8') as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading DB: {e}")
        return {"projects": {}}

def save_db(data):
    try:
        with open(PROJECTS_DB_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=4, ensure_ascii=False)
    except Exception as e:
        print(f"Error saving DB: {e}")

LOG_FILE = BASE_DIR / "Codigos/Usuarios/audit_log.json"

AUXILIAR_DIR = BASE_DIR / "Auxiliares/indices_auxiliar"

HISTORIAL_PO_DIR = BASE_DIR / "Auxiliares/Historial PO"

PROFILE_PICS_DIR = BASE_DIR / "Codigos/Usuarios/profile_pics"

MOCK_PATH = Path("mock_data")

RESET_TOKENS_FILE = BASE_DIR / "Codigos/Usuarios/reset_tokens.json"

SMTP_CONFIG = {
    "SMTP_SERVER": "smtp.gmail.com",
    "SMTP_PORT": 587,
    "SMTP_USER": "costos@bpbargentina.com",
    "SMTP_PASS": "ksoe dybt byya zcrz",
    "SMTP_FROM_NAME": "Oficina Tecnica"
}

def send_reset_email(to_email, token):
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{SMTP_CONFIG['SMTP_FROM_NAME']} <{SMTP_CONFIG['SMTP_USER']}>"
        msg['To'] = to_email
        msg['Subject'] = "Recuperar Contraseña - BPB Group"

        # Dynamically build link (assuming standard port/host, can be improved with url_for external)
        # Use request.host_url from context if available, otherwise guess.
        # Since this is called from a route, request context is active.
        reset_link = f"{request.host_url}validate-reset?token={token}"

        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Restablecimiento de Contraseña</h2>
            <p>Se ha solicitado un cambio de contraseña para su cuenta ({to_email}).</p>
            <p>Para confirmar el cambio y establecer su nueva contraseña, haga clic en el siguiente enlace:</p>
            <p>
                <a href="{reset_link}" style="background-color: #e74c3c; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    Confirmar Cambio de Contraseña
                </a>
            </p>
            <p style="font-size: 0.9em; color: #777;">Si usted no solicitó este cambio, ignore este correo.</p>
            <p>Este enlace es válido por 1 hora.</p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        server = smtplib.SMTP(SMTP_CONFIG['SMTP_SERVER'], SMTP_CONFIG['SMTP_PORT'])
        server.starttls()
        server.login(SMTP_CONFIG['SMTP_USER'], SMTP_CONFIG['SMTP_PASS'])
        server.sendmail(SMTP_CONFIG['SMTP_USER'], to_email, msg.as_string())
        server.quit()
        print(f"DEBUG: Reset email sent to {to_email}")
        return True, "Correo enviado"
    except Exception as e:
        print(f"ERROR: Failed to send email: {e}")
        # Log the link for debugging if SMTP fails
        print(f"DEBUG LINK (Fallback): {request.host_url}validate-reset?token={token}")
        return False, str(e)

@app.route('/api/request-reset', methods=['POST'])
def request_reset():
    try:
        # RAW DEBUG
        # print(f"DEBUG RAW PAYLOAD: {request.data}")
        # print(f"DEBUG JSON: {request.json}")

        data = request.json
        if data is None:
             return jsonify({'status': 'error', 'message': 'No JSON data received'}), 400

        if isinstance(data, str):
            try:
                data = json.loads(data)
            except:
                return jsonify({'status': 'error', 'message': 'Invalid JSON format (string parse failed)'}), 400
            
        if not isinstance(data, dict):
             return jsonify({'status': 'error', 'message': f'Invalid data format: Expected dict, got {type(data)}'}), 400

        email_addr = data.get('email', '').strip()
        new_password = data.get('new_password', '').strip()

        if not email_addr or not new_password:
            return jsonify({'status': 'error', 'message': 'Faltan datos.'}), 400

        if not os.path.exists(USERS_FILE):
            return jsonify({'status': 'error', 'message': 'Error interno de usuarios.'}), 500

        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            users = json.load(f)
        
        # Check if user exists
        # Users file is a dict: {username: user_data}
        user_exists = False
        for username, u_data in users.items():
            # Check if key (username) matches email OR if 'email' field inside matches
            if username.lower() == email_addr.lower():
                user_exists = True
                break
            if isinstance(u_data, dict) and u_data.get('email', '').lower() == email_addr.lower():
                user_exists = True
                break
        
        if not user_exists:
             return jsonify({'status': 'error', 'message': 'Email no registrado.'}), 404

        # Generate Token
        token = secrets.token_urlsafe(32)
        
        # Load existing tokens
        tokens = {}
        if os.path.exists(RESET_TOKENS_FILE):
            try:
                with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
                    tokens = json.load(f)
            except: tokens = {}

        # Clean expired tokens (older than 1h)
        now = time.time()
        clean_tokens = {}
        for k, v in tokens.items():
            if isinstance(v, dict) and (now - v.get('timestamp', 0) < 3600):
                clean_tokens[k] = v
        tokens = clean_tokens

        # Add new token
        tokens[token] = {
            'email': email_addr,
            'new_password_hash': generate_password_hash(new_password),
            'timestamp': now
        }

        # Save tokens
        try:
            os.makedirs(RESET_TOKENS_FILE.parent, exist_ok=True)
            with open(RESET_TOKENS_FILE, 'w', encoding='utf-8') as f:
                json.dump(tokens, f, indent=4)
        except Exception as file_err:
             import traceback
             print(traceback.format_exc())
             return jsonify({'status': 'error', 'message': f'Error guardando token: {file_err}'}), 500

        # Send Email
        success, msg = send_reset_email(email_addr, token)
        
        if success:
            return jsonify({'status': 'success', 'message': 'Revise su correo para confirmar.'})
        else:
            return jsonify({'status': 'error', 'message': f'Error enviando correo: {msg}'})

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"CRITICAL ERROR in request_reset: {tb}")
        return jsonify({'status': 'error', 'message': f'Error CRITICO: {str(e)} -> {tb}'}), 500

@app.route('/validate-reset', methods=['GET'])
def validate_reset():
    token = request.args.get('token')
    
    if not token or not os.path.exists(RESET_TOKENS_FILE):
        return "Enlace inválido o expirado.", 400

    try:
        with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
            tokens = json.load(f)

        token_data = tokens.get(token)
        if not token_data:
             return "Enlace inválido o expirado.", 400
        
        # Check expiration again just in case
        if time.time() - token_data.get('timestamp', 0) > 3600:
            return "Enlace expirado.", 400

        # Update User Password
        target_email = token_data['email']
        new_hash = token_data['new_password_hash']

        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            users = json.load(f)

        updated = False
        target_email = target_email.strip().lower()
        
        # Iterate safely over keys and values
        matched_username = None
        for username, u_data in users.items():
            # Check if key matches email
            if username.lower() == target_email:
                matched_username = username
                break
            # Check if inner email matches
            if isinstance(u_data, dict) and u_data.get('email', '').strip().lower() == target_email:
                matched_username = username
                break

        if matched_username and matched_username in users:
             users[matched_username]['password'] = new_hash
             updated = True
        
        if updated:
            with open(USERS_FILE, 'w', encoding='utf-8') as f:
                json.dump(users, f, indent=4, ensure_ascii=False)
            
            # Delete used token
            del tokens[token]
            with open(RESET_TOKENS_FILE, 'w', encoding='utf-8') as f:
                json.dump(tokens, f, indent=4)
                
            return """
            <html>
            <head>
                <meta http-equiv="refresh" content="3;url=/">
                <style>body{font-family:sans-serif;text-align:center;padding-top:50px;}</style>
            </head>
            <body>
                <h1 style="color:green;">¡Contraseña Actualizada!</h1>
                <p>Será redirigido al inicio de sesión en 3 segundos...</p>
            </body>
            </html>
            """
        else:
             return "Usuario no encontrado.", 404

    except Exception as e:
        print(f"Error validating reset: {e}")
        return "Error interno.", 500

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

MANUAL_INGRESOS_FILE = BASE_DIR / r"Auxiliares/Ingresos/resumen_all.csv"



# Ensure profile pics dir exists

if not PROFILE_PICS_DIR.exists():

    try:

        PROFILE_PICS_DIR.mkdir(parents=True, exist_ok=True)

    except Exception as e:

        print(f"Warning: Could not create profile pics folder: {e}")



# Determine which path to use

if os.environ.get('FLASK_TEST_MODE'):

    DATA_DIR = MOCK_PATH

    print("Running in TEST MODE: Using mock data")

else:

    if PRODUCTION_PATH.exists():

        DATA_DIR = PRODUCTION_PATH

        print(f"Production path found: {DATA_DIR}")

    else:

        DATA_DIR = MOCK_PATH

        print(f"Production path NOT found. Falling back to Mock Data: {DATA_DIR}")

    

    if not UPLOAD_FOLDER.exists():

        try:

            UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

        except Exception as e:

            print(f"Warning: Could not create upload folder {UPLOAD_FOLDER}: {e}")



print(f"Using Data Directory: {DATA_DIR.resolve()}")



def log_action(username, action, details):

    print(f"[LOG ACTION] User: {username}, Action: {action}")

    try:

        logs = []

        if LOG_FILE.exists():

            with open(LOG_FILE, 'r') as f:

                logs = json.load(f)

        

        entry = {

            "timestamp": time.strftime('%Y-%m-%d %H:%M:%S'),

            "user": username,

            "action": action,

            "details": details

        }

        logs.append(entry)

        

        if len(logs) > 1000:

            logs = logs[-1000:]

            

        with open(LOG_FILE, 'w') as f:

            json.dump(logs, f, indent=4)

        print("[LOG ACTION] Saved successfully.")

            

    except Exception as e:

        print(f"Logging Error: {e}")



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



def get_entry_dates_cache():

    """

    Scans IN_PROCESS_DIR for folders matching 'Registro - Rxxxx - ...'

    Sorts them by R number (ascending) so latest overwrites oldest.

    Reads resumen.csv and builds {po_number: date} map.

    """

    cache = {}

    if not IN_PROCESS_DIR.exists():

        return cache

        

    # Find all relevant folders

    r_folders = []

    for item in IN_PROCESS_DIR.iterdir():

        if item.is_dir():
            # Match R number. Pattern assumed: "Registro - R0060 - ..."
            match = re.search(r'Registro\s*-\s*R(\d+)', item.name, re.IGNORECASE)
            if match:
                r_num = int(match.group(1))
                r_folders.append((r_num, item))

    # Sort by R number Ascending (Oldest to Newest)
    r_folders.sort(key=lambda x: x[0])

    for _, folder in r_folders:
        resumen_file = folder / "resumen.csv"
        if resumen_file.exists():
            try:
                # Try reading with different encodings
                content = None
                for enc in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        with open(resumen_file, 'r', encoding=enc) as f:
                            content = f.read()
                        break
                    except UnicodeDecodeError:
                        continue
                
                if content:
                    delimiter = ';' if ';' in content.split('\n')[0] else ','
                    lines = content.split('\n')
                    lines = [l.strip() for l in lines if l.strip()]
                    
                    for line in lines:
                        parts = line.split(delimiter)
                        if len(parts) >= 4:
                            date_val = parts[3].strip()
                            po_val = parts[2].strip()
                            if not po_val: po_val = parts[1].strip()
                            
                            if po_val and date_val:
                                digits = "".join(filter(str.isdigit, po_val))
                                if digits:
                                    cache[digits] = date_val
            except Exception as e:
                print(f"Error reading resumen in {folder}: {e}")

    return cache

# --- SYNC GLOBAL STATE ---
SYNC_STATE = {
    "running": False,
    "stage": "idle",
    "progress": 0,
    "total": 0,
    "current": 0,
    "message": "",
    "source": "auto", 
    "data_version": 0
}

def count_files_recursive(directory, extension):
    count = 0
    try:
        for path in Path(directory).rglob(f'*{extension}'):
            if path.is_file():
                count += 1
    except: pass
    return count

def count_subdirs(directory):
    count = 0
    try:
        # Count direct subdirectories
        for path in Path(directory).iterdir():
            if path.is_dir():
                count += 1
    except: pass
    return count



def background_sync_task():

    global SYNC_STATE

    SYNC_STATE["running"] = True

    SYNC_STATE["progress"] = 0

    SYNC_STATE["data_version"] = 0 # Reset version

    # script_dir = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Codigos")
    script_dir = BASE_DIR / "Codigos"

    

    try:

        import sys

        # --- ETAPA 1: STEP 2 (Descarga de PDFs) ---

        SYNC_STATE["stage"] = "step2"

        SYNC_STATE["message"] = "Descargando archivos PDF..."

        

        # LOGGING FOR DEBUGGING

        print(f"Background Sync: Using interpreter {sys.executable}")

        print(f"Background Sync: CWD is {script_dir}")

        

        # Keep progress at 0 during fetch since we can't estimate easily yet

        SYNC_STATE["progress"] = 0 

        

        # Use sys.executable to ensure same environment

        step2_cmd = [sys.executable, 'step2_fetch_pdfs.py', '--config', 'config.yaml']

        

        # Run Step 2

        result2 = subprocess.run(step2_cmd, cwd=str(script_dir), capture_output=True, text=True)

        if result2.returncode != 0:

            print(f"Step 2 Failed: {result2.stderr}")

            raise Exception(f"Fallo descarga PDF: {result2.stderr}")

        

        # --- ETAPA 2: STEP 3 (Procesar Outputs en EN PROGRESO) ---

        SYNC_STATE["stage"] = "step3"

        SYNC_STATE["message"] = "Procesando registros..."

        

        # entrantes_dir = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\P2 - Purchase Order\Entrantes")
        entrantes_dir = BASE_DIR / "P2 - Purchase Order/Entrantes"
        
        # en_progreso_dir = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\P2 - Purchase Order\En Progreso")
        en_progreso_dir = BASE_DIR / "P2 - Purchase Order/En Progreso"

        

        total_pdfs = count_files_recursive(entrantes_dir, ".pdf")

        SYNC_STATE["total"] = total_pdfs if total_pdfs > 0 else 1 # Avoid div/0

        

        step3_cmd = [sys.executable, 'step3_prepare_outputs.py', '--config', 'config.yaml']


        if process3.returncode != 0:

            print(f"Step 3 Failed with return code {process3.returncode}")

            raise Exception(f"Fallo script procesamiento (Ver consola para detalles)")



        SYNC_STATE["progress"] = 100

        SYNC_STATE["message"] = "Sincronización completada"

        

    except Exception as e:

        print(f"Sync Error: {e}")

        SYNC_STATE["message"] = f"Error: {str(e)}"

        SYNC_STATE["stage"] = "error"

    finally:

        SYNC_STATE["running"] = False

        print("Background sync task finished.")



def start_sync_thread(source="auto"):

    if SYNC_STATE["running"]:

        return False

    SYNC_STATE["source"] = source

    thread = threading.Thread(target=background_sync_task)

    thread.daemon = True

    thread.start()

    return True


# --- ROUTES ---

@app.route('/solids/<path:filename>')
def serve_solids(filename):
    return send_from_directory(SOLIDS_DIR, filename)

@app.route('/api/projects', methods=['GET', 'POST'])
def api_projects():
    db = load_db()
    if request.method == 'GET':
        # Convert dict to list for frontend
        projects_list = list(db['projects'].values())
        return jsonify(projects_list)
    
    if request.method == 'POST':
        data = request.json
        p_id = str(data.get('id'))
        if not p_id:
            return jsonify({"status": "error", "message": "Missing ID"}), 400
        
        db['projects'][p_id] = data
        save_db(db)
        return jsonify({"status": "success", "message": "Project saved"})

@app.route('/api/add-solid', methods=['POST'])
def api_add_solid():
    try:
        project_id = request.form.get('projectId')
        solid_name = request.form.get('name')
        revision = request.form.get('revision')
        
        if 'file' not in request.files:
             return jsonify({'status': 'error', 'message': 'No file part'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No selected file'}), 400

        # Save File
        filename = secure_filename(file.filename)
        save_path = os.path.join(SOLIDS_DIR, filename)
        file.save(save_path)
        
        # Save Metadata
        db = load_db()
        if project_id not in db['projects']:
             return jsonify({'status': 'error', 'message': 'Project not found'}), 404
             
        if 'solids' not in db['projects'][project_id]:
            db['projects'][project_id]['solids'] = []
            
        new_solid = {
            "id": f"s{int(time.time()*1000)}",
            "name": solid_name,
            "revision": revision,
            "filename": filename,
            "hasFile": True,
            "date": datetime.now().strftime("%Y-%m-%d")
        }
        
        db['projects'][project_id]['solids'].append(new_solid)
        save_db(db)
        
        return jsonify({'status': 'success', 'solid': new_solid})
        
    except Exception as e:
        print(f"Error adding solid: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/solids/<project_id>', methods=['GET'])
def api_get_solids(project_id):
    db = load_db()
    if project_id in db['projects']:
        solids = db['projects'][project_id].get('solids', [])
        return jsonify(solids)
    return jsonify([])

@app.route('/projects-viewer')
def projects_viewer():
    # Redirect or serve a specific template if needed, for now use index
    return redirect(url_for('index'))

@app.route('/')
def index():
    """Serve the main page with no-cache headers to prevent stale HTML"""
    from datetime import datetime
    response = make_response(render_template('index.html', cache_buster=int(datetime.now().timestamp())))
    # Prevent browser from caching HTML to ensure users always get latest version
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/api/fetch-pdfs', methods=['POST'])

def run_fetch_pdfs():

    if not start_sync_thread(source="auto"):

        return jsonify({"status": "running", "message": "Ya hay una sincronización en curso"})

    


    return jsonify({"status": "started", "message": "Iniciando proceso en segundo plano"})


@app.route('/api/logistics/calculate', methods=['POST'])
def calculate_logistics():
    print("DEBUG: >>> ENTERING calculate_logistics <<<")
    try:
        data = request.json
        if not data: return jsonify({'status': 'error', 'message': 'No data received'}), 400
        container_data = data.get('container')
        items_data = data.get('items')
        config = data.get('config', {})
        if not container_data or not items_data: return jsonify({'status': 'error', 'message': 'Missing data'}), 400
        
        # Save historical trace for internal logs if needed
        # (Already happening in debug logs)

        container_type_key = config.get('container_type', '20ft')

        if config.get('maximize'):
            print(f"DEBUG: Maximization requested for {len(items_data)} items.")
            
            # SPECIAL CASE: Multi-Item Single Pallet Maximization
            # When: No Container + Maximize + Pallets NOT Mixed
            # Each item gets its own maximized pallet
            if container_type_key == 'none' and not config.get('mixed_pallets', True):
                print(f"DEBUG: Per-Item Single Pallet Maximization Mode (mixed_pallets=false)")
                print(f"DEBUG: Input items: {[it['id'] for it in items_data]}")
                
                # Strategy: Find max quantity for each item, then pack all together in ONE call
                new_quantities = {it['id']: 0 for it in items_data}
                maximized_items = []
                
                # For binary search, use a test config that ALLOWS mixing (single item anyway)
                test_config = {k: v for k, v in config.items()}
                test_config['mixed_pallets'] = True  # Override for testing
                
                # IMPORTANT: Process ALL items, even those with qty=0
                # In maximization mode, qty=0 means "find the maximum", not "skip this item"
                for item in items_data:
                    print(f"DEBUG: Finding max for item: {item['id']}")
                    # Use binary search to find max quantity for this item alone
                    lo, hi = 1, 10000
                    best = 0
                    
                    while lo <= hi:
                        mid = (lo + hi) // 2
                        test_item = dict(item, qty=mid)
                        res_test = _do_pack_internal(container_data, [test_item], test_config)
                        
                        if res_test.get('status') == 'success' and res_test.get('unfitted_count', 0) == 0:
                            best = mid
                            lo = mid + 1
                        else:
                            hi = mid - 1
                    
                    print(f"DEBUG: Max found for {item['id']}: {best}")
                    if best > 0:
                        new_quantities[item['id']] = best
                        maximized_items.append(dict(item, qty=best))
                
                print(f"DEBUG: Maximized items to pack: {[(it['id'], it['qty']) for it in maximized_items]}")
                
                # Now pack ALL maximized items together in ONE call
                # Use the ORIGINAL config which has mixed_pallets=False
                # This will separate each item type into its own pallet(s)
                if len(maximized_items) > 0:
                    res = _do_pack_internal(container_data, maximized_items, config)
                    
                    print(f"DEBUG: Final packing result status: {res.get('status')}")
                    print(f"DEBUG: Number of pallets: {len(res.get('grouped_pallets', []))}")
                    print(f"DEBUG: Number of packed items: {len(res.get('packed_items', []))}")
                    
                    if res.get('status') == 'success':
                        res['new_quantities'] = new_quantities
                        return jsonify(res)
                    else:
                        return jsonify({'status': 'error', 'message': 'No se pudo maximizar ningún ítem.'}), 500
                else:
                    return jsonify({'status': 'error', 'message': 'No se pudo maximizar ningún ítem.'}), 500
            
            # Refined Strategy: Adaptive Single-Pass Greedy
            # 1. If Ratios defined: Interleave & Disable Sort (Fairness)
            # 2. If No Ratios: Block Gen & Enable Sort (Volume Max)
            try:
                c_w = float(container_data.get('width', 0) or 120)
                c_h = float(container_data.get('height', 0) or 250)
                c_d = float(container_data.get('depth', 0) or 80)
                container_vol = float(c_w * c_h * c_d)
                if container_vol <= 0: container_vol = 2400000.0
            except:
                container_vol = 2400000.0

            # Check if any relative quantity is set
            has_ratios = False
            for it in items_data:
                try:
                    if float(it.get('rel_qty') or 0) > 0:
                        has_ratios = True
                        break
                except: pass

            def get_base_qty(it):
                try:
                    rel = float(it.get('rel_qty') or 0)
                    if rel > 0: return rel
                    cur = float(it.get('qty') or 0)
                    return cur if cur > 0 else 1.0
                except: return 1.0

            total_set_vol = 0
            for it in items_data:
                try:
                    i_vol = float(it.get('w', 0)) * float(it.get('h', 0)) * float(it.get('d', 0))
                    total_set_vol += get_base_qty(it) * i_vol
                except: pass

            multiplier = (container_vol / total_set_vol * 1.3) if total_set_vol > 0 else 100
            global_item_limit = 25000 if container_type_key != 'none' else 8000
            
            max_items = []
            
            if has_ratios:
                # RATIO-SET MAXIMIZATION (Enforce complete sets)
                ratio_items = []
                fixed_items = []
                for it in items_data:
                    rel = float(it.get('rel_qty') or 0)
                    if rel > 0:
                        ratio_items.append((it, rel))
                    else:
                        q = int(it.get('qty') or 0)
                        if q > 0:
                            fixed_items.append(it)

                def build_items_for_sets(k):
                    built = []
                    for it, rel in ratio_items:
                        q = int(rel * k)
                        if q > 0:
                            new_it = dict(it)
                            new_it['qty'] = q
                            built.append(new_it)
                    for it in fixed_items:
                        q = int(it.get('qty') or 0)
                        if q > 0:
                            new_it = dict(it)
                            new_it['qty'] = q
                            built.append(new_it)
                    return built

                total_rel = sum(rel for _, rel in ratio_items)
                max_sets_limit = int(40000 / max(total_rel, 1))
                if max_sets_limit < 1:
                    max_sets_limit = 1

                config_for_test = dict(config)
                config_for_test['maximize'] = True
                config_for_test['skip_max_cap'] = True

                best = 0
                hi = 1
                while hi <= max_sets_limit:
                    test_items = build_items_for_sets(hi)
                    res_test = _do_pack_internal(container_data, test_items, config_for_test)
                    if res_test.get('status') != 'success' or res_test.get('unfitted_count', 0) > 0:
                        break
                    best = hi
                    hi *= 2

                lo = best + 1
                hi = min(hi - 1, max_sets_limit)
                while lo <= hi:
                    mid = (lo + hi) // 2
                    test_items = build_items_for_sets(mid)
                    res_mid = _do_pack_internal(container_data, test_items, config_for_test)
                    if res_mid.get('status') == 'success' and res_mid.get('unfitted_count', 0) == 0:
                        best = mid
                        lo = mid + 1
                    else:
                        hi = mid - 1

                final_items = build_items_for_sets(best)
                res = _do_pack_internal(container_data, final_items, config_for_test)

                if res.get('status') == 'success':
                    new_quantities = {it['id']: 0 for it in items_data}
                    for it, rel in ratio_items:
                        new_quantities[it['id']] = int(rel * best)
                    for it in fixed_items:
                        new_quantities[it['id']] = int(it.get('qty') or 0)
                    res['new_quantities'] = new_quantities
                    res['set_count'] = best
                    res['unfitted_items'] = []
                    res['unfitted_count'] = 0
                    return jsonify(res)

                return jsonify({'status': 'error', 'message': 'Fallo en el cálculo de maximización por sets.'}), 500
                
            else:
                # BLOCK GENERATION (Volume Efficiency)
                # Standard generation, let the packer sort by volume
                for it in items_data:
                    q = int(get_base_qty(it) * multiplier)
                    q = min(q, global_item_limit)
                    if q < 1: q = 1
                    max_items.append(dict(it, qty=q))
                should_sort = True # Enable sorting in packer

            print(f"DEBUG: Packing Mode: {'Interleaved/NoSort' if has_ratios else 'Block/Sort'}. Items: {len(max_items)}")
            
            # Pass 'sort_items' through config or arguments?
            # _do_pack_internal takes config. We can inject it there.
            config['sort_items'] = should_sort
            res = _do_pack_internal(container_data, max_items, config)
            
            if res.get('status') == 'success':
                packed_counts = {it['id']: 0 for it in items_data}
                for pi in res.get('packed_items', []):
                    name = pi.get('name')
                    if name in packed_counts: packed_counts[name] += 1
                
                res['new_quantities'] = packed_counts
                res['multiplier'] = multiplier
                res['unfitted_items'] = []
                res['unfitted_count'] = 0
                return jsonify(res)
            
            return jsonify({'status': 'error', 'message': 'Fallo en el cálculo de maximización.'}), 500
        else:
            res = _do_pack_internal(container_data, items_data, config)
            return jsonify(res) if res.get('status') == 'success' else (jsonify(res), 500)
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/save', methods=['POST'])
def save_logistics_calculation():
    try:
        data = request.json
        if not data: return jsonify({'status': 'error', 'message': 'No data received'}), 400

        # Sanitize payload to prevent accidental huge saves (packed_items, kpis, etc.)
        def _clean_items(raw_items):
            clean = []
            if not isinstance(raw_items, list):
                return clean
            for it in raw_items:
                if not isinstance(it, dict):
                    continue
                clean.append({
                    'name': it.get('name', ''),
                    'l': it.get('l', 0),
                    'w': it.get('w', 0),
                    'h': it.get('h', 0),
                    'weight': it.get('weight', 0),
                    'qty': it.get('qty', 0),
                    'rel_qty': it.get('rel_qty', 0)
                })
            return clean

        record = {
            'id': data.get('id') or str(uuid.uuid4()),
            'save_name': data.get('save_name', ''),
            'save_description': data.get('save_description', ''),
            'timestamp': data.get('timestamp', datetime.utcnow().isoformat() + 'Z'),
            'author': data.get('author', 'Usuario'),
            'container': {
                'type': data.get('container', {}).get('type'),
                'l': data.get('container', {}).get('l', 0),
                'w': data.get('container', {}).get('w', 0),
                'h': data.get('container', {}).get('h', 0),
                'weight': data.get('container', {}).get('weight', 0)
            },
            'pallet': {
                'type': data.get('pallet', {}).get('type'),
                'boards': data.get('pallet', {}).get('boards', 0),
                'l': data.get('pallet', {}).get('l', 0),
                'w': data.get('pallet', {}).get('w', 0),
                'h': data.get('pallet', {}).get('h', 0),
                'weight': data.get('pallet', {}).get('weight', 0),
                'max_weight': data.get('pallet', {}).get('max_weight', 0),
                'limit_height': data.get('pallet', {}).get('limit_height', False),
                'max_height': data.get('pallet', {}).get('max_height', 0)
            },
            'optimization': {
                'maximize': data.get('optimization', {}).get('maximize', False),
                'mixed_pallets': data.get('optimization', {}).get('mixed_pallets', True),
                'stack_load': data.get('optimization', {}).get('stack_load', True),
                'force_orientation': data.get('optimization', {}).get('force_orientation', False),
                'orientation_face': data.get('optimization', {}).get('orientation_face', 'LxA'),
                'mixed_containers': data.get('optimization', {}).get('mixed_containers', True),
                'sort_items': data.get('optimization', {}).get('sort_items', True),
                'visual_only': data.get('optimization', {}).get('visual_only', False)
            },
            'safety_factors': {
                'dims': data.get('safety_factors', {}).get('dims',
                        data.get('safety_factor_dims', 0)),
                'weight': data.get('safety_factors', {}).get('weight',
                          data.get('safety_factor_weight', 0))
            },
            'items': _clean_items(data.get('items', []))
        }
        
        records = []
        if os.path.exists(LOGISTICS_RECORDS_FILE):
            try:
                with open(LOGISTICS_RECORDS_FILE, 'r', encoding='utf-8') as f:
                    records = json.load(f)
            except:
                records = []
        
        records.insert(0, record) # Newest first
        
        # Keep only last 1000 records to avoid huge files
        records = records[:1000]
        
        with open(LOGISTICS_RECORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(records, f, indent=4, ensure_ascii=False)
            
        return jsonify({'status': 'success', 'message': 'Calculation saved'})
    except Exception as e:
        print(f"ERROR saving logistics: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/logistics/records', methods=['GET'])
def get_logistics_records():
    try:
        if not os.path.exists(LOGISTICS_RECORDS_FILE):
            return jsonify([])
        with open(LOGISTICS_RECORDS_FILE, 'r', encoding='utf-8') as f:
            records = json.load(f)
        return jsonify(records)
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/logistics/delete', methods=['POST'])
def delete_logistics_record():
    try:
        data = request.json or {}
        rec_id = data.get('id')
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400
        if not os.path.exists(LOGISTICS_RECORDS_FILE):
            return jsonify({'status': 'error', 'message': 'No records file'}), 404

        with open(LOGISTICS_RECORDS_FILE, 'r', encoding='utf-8') as f:
            records = json.load(f)

        new_records = [r for r in records if r.get('id') != rec_id]
        if len(new_records) == len(records):
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        with open(LOGISTICS_RECORDS_FILE, 'w', encoding='utf-8') as f:
            json.dump(new_records, f, indent=4, ensure_ascii=False)

        return jsonify({'status': 'success', 'message': 'Record deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/check-ingresos', methods=['POST'])
def check_ingresos():
    """Check for new ingresos data by running procesar_ingresos.py"""
    import sys
    try:
        script_dir = BASE_DIR / "Codigos"
        script_name = "procesar_ingresos.py"
        
        if not script_dir.exists():
             return jsonify({"status": "error", "message": "Script directory not found"}), 404

        # Run checking script using the SAME python interpreter as this app
        result = subprocess.run([sys.executable, script_name], cwd=str(script_dir), capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"Check Script Error: {result.stderr}")
            return jsonify({"status": "error", "message": f"Error en script: {result.stderr}"}), 500
            
        # Analyze output
        if "NO se crea" in result.stdout:
            return jsonify({"status": "no_changes", "message": "Sin cambios", "debug_output": result.stdout})
        else:
            print(f"DEBUG: Check determined CHANGES DETECTED. Stdout chars: {len(result.stdout)}")
            return jsonify({"status": "changes_detected", "message": "Nuevos registros detectados", "debug_output": result.stdout})

    except Exception as e:
        print(f"Check Process Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/sync-progress', methods=['GET'])

def get_sync_progress():

    return jsonify(SYNC_STATE)



@app.route('/api/run-automation', methods=['POST'])

def run_automation():

    try:

        # script_dir = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Codigos")
        script_dir = BASE_DIR / "Codigos"

        script_name = "automation.py"

        

        # Args as requested: --config ".\config.yaml" --once

        cmd = ['python', script_name, '--config', r'.\config.yaml', '--once']



        if not (script_dir / script_name).exists():

            return jsonify({"status": "error", "message": "Automation script not found"}), 404



        # Run non-blocking or blocking? 

        # User implies it's triggered after upload. 

        # Usually automation takes time. Blocking might timeout.

        # But for simplicity let's do blocking first as per previous pattern.

        result = subprocess.run(cmd, cwd=str(script_dir), capture_output=True, text=True)

        

        if result.returncode == 0:

             return jsonify({"status": "success", "message": "Automación ejecutada", "output": result.stdout})

        else:

             print(f"Auto Error: {result.stderr}")

             return jsonify({"status": "error", "message": f"Error automación: {result.stderr}"}), 500



    except Exception as e:

        print(f"Auto Run Error: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/data')

def get_data():

    results = []

    if not DATA_DIR.exists():

        return jsonify({"error": "Data directory not found", "path": str(DATA_DIR)}), 404



    # Pre-fetch entry dates

    entry_dates_cache = get_entry_dates_cache()

        

    for folder in DATA_DIR.iterdir():

        if folder.is_dir():

            try:

                stats = folder.stat()

                # On Windows st_ctime is creation time, st_mtime is modification. 

                # "Fecha Ingreso" usually implies creation/arrival. Let's use mtime as it reflects the folder's last touch.

                sys_date = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y')

            except:

                sys_date = '-'

            

            # --- OVERRIDE DATE WITH REGISTRO INFO IF AVAILABLE ---

            # Extract number from folder name (e.g. PO1291 -> 1291)

            po_digits = "".join(filter(str.isdigit, folder.name))

            if po_digits and po_digits in entry_dates_cache:

                sys_date = entry_dates_cache[po_digits]

            # -----------------------------------------------------



            po_data = {

                "id": folder.name,

                "system_date": sys_date,

                "path": str(folder),

                "files": {"pdfs": [], "csvs": [], "json": None, "approved_items": []},

                "content": {"json": {}, "csvs": []}

            }



            # Read Approval Info

            meta_path = folder / "approval_info.json"

            if meta_path.exists():

                try:

                    with open(meta_path, 'r', encoding='utf-8') as f:

                        app_meta = json.load(f)

                        po_data["files"]["approved_items"] = app_meta.get("approved_items", [])

                except: pass

            

            # Scan files

            for file in folder.iterdir():

                if file.is_file():

                    if file.suffix.lower() == '.pdf': po_data["files"]["pdfs"].append(file.name)

                    elif file.suffix.lower() == '.json':

                        po_data["files"]["json"] = file.name

                        try:

                            with open(file, 'r', encoding='utf-8') as f: po_data["content"]["json"] = json.load(f)

                        except: pass

                    elif file.suffix.lower() == '.csv':

                        po_data["files"]["csvs"].append(file.name)

                        try:

                            with open(file, 'r', encoding='utf-8', errors='replace') as f:

                                reader = csv.DictReader(f, delimiter=';')

                                headers = [str(h) for h in (reader.fieldnames or [])]

                                rows = [{str(k): v for k, v in row.items()} for row in reader]

                                po_data["content"]["csvs"].append({"filename": file.name, "headers": headers, "rows": rows})

                        except: pass

            

            # Scan csv subdir

            csv_subdir = folder / 'csv'

            if csv_subdir.exists():

                for file in csv_subdir.iterdir():

                    if file.is_file() and file.suffix.lower() == '.csv':

                        po_data["files"]["csvs"].append(file.name)

                        try:

                            with open(file, 'r', encoding='utf-8', errors='replace') as f:

                                reader = csv.DictReader(f, delimiter=';')

                                headers = [str(h) for h in (reader.fieldnames or [])]

                                rows = [{str(k): v for k, v in row.items()} for row in reader]

                                po_data["content"]["csvs"].append({"filename": file.name, "headers": headers, "rows": rows})

                        except: pass

            results.append(po_data)

    return jsonify(results)



@app.route('/api/profile/history')

def get_my_history():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "Login required"}), 401

    return get_user_history(session.get('user'))



@app.route('/api/admin/history/<username>', methods=['GET'])

def get_user_history(username):

    current_user = session.get('user')

    current_role = session.get('role')

    

    # Permitir si es admin O si el usuario pide su propio historial

    if current_role != 'admin' and current_user != username:

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    try:

        if not LOG_FILE.exists():

            return jsonify([])

            

        with open(LOG_FILE, 'r') as f:

            logs = json.load(f)

            

        user_logs = [l for l in logs if l['user'] == username]

        return jsonify(user_logs[::-1])

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/update', methods=['POST'])

def update_data():

    data = request.json

    po_id = data.get('po_id')

    field = data.get('field')

    value = data.get('value')

    

    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        return jsonify({"status": "error", "message": "Folder not found"}), 404

        

    json_files = list(target_folder.glob("*.json"))

    if not json_files:

         return jsonify({"status": "error", "message": "JSON file not found"}), 404

    

    target_file = json_files[0]

    

    try:

        with open(target_file, 'r', encoding='utf-8') as f:

            content = json.load(f)

        

        content[field] = value

        

        with open(target_file, 'w', encoding='utf-8') as f:

            json.dump(content, f, indent=4)

            

        actor = session.get('user', 'system')

        log_action(actor, "Updated JSON Field", f"PO: {po_id}, Field: {field}, New Value: {value}")

            

        return jsonify({"status": "success", "new_value": value})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/update_csv', methods=['POST'])

def update_csv():

    data = request.json

    po_id = data.get('po_id')

    filename = data.get('filename')

    rows = data.get('rows')



    if not po_id or not filename:

        return jsonify({"status": "error", "message": "Missing PO ID or Filename"}), 400



    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        return jsonify({"status": "error", "message": "PO Folder not found"}), 404



    target_file = target_folder / filename

    if not target_file.exists():

        target_file = target_folder / 'csv' / filename

        

    if not target_file.exists():

         return jsonify({"status": "error", "message": "File not found"}), 404



    try:

        if not rows:

             with open(target_file, 'w', encoding='utf-8', newline='') as f:

                 pass

             actor = session.get('user', 'system')

             log_action(actor, "Cleared CSV", f"PO: {po_id}, File: {filename}")

             return jsonify({"status": "success", "message": "File cleared"})



        headers = list(rows[0].keys())



        with open(target_file, 'w', encoding='utf-8', newline='', errors='replace') as f:

            writer = csv.DictWriter(f, fieldnames=headers, delimiter=';')

            writer.writeheader()

            writer.writerows(rows)

            

        actor = session.get('user', 'system')

        log_action(actor, "Updated CSV Data", f"PO: {po_id}, File: {filename}, Rows: {len(rows)}")



        return jsonify({"status": "success"})



    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/upload', methods=['POST'])

def upload_file():

    if 'file' not in request.files:

        return jsonify({'status': 'error', 'message': 'No file part'}), 400

        

    file = request.files['file']

    

    if file.filename == '':

        return jsonify({'status': 'error', 'message': 'No selected file'}), 400

        

    if file and allowed_file(file.filename):

        try:

            original_filename = secure_filename(file.filename)

            timestamp = int(time.time())

            filename, ext = os.path.splitext(original_filename)

            new_filename = f"{filename}_{timestamp}{ext}"

            

            save_path = UPLOAD_FOLDER / new_filename

            

            if not UPLOAD_FOLDER.exists():

                UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

                

            file.save(str(save_path))

            

            actor = session.get('user', 'system')

            log_action(actor, "Uploaded File", f"Filename: {new_filename}")

            

            return jsonify({'status': 'success', 'filename': new_filename})

        except Exception as e:

            return jsonify({'status': 'error', 'message': str(e)}), 500

            

    return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400



@app.route('/api/login', methods=['POST'])

def login():

    data = request.json

    username = data.get('username')

    password = data.get('password')

    

    if not username or not password:

        return jsonify({"status": "error", "message": "Faltan credenciales"}), 400

        

    try:

        if not USERS_FILE.exists():

             return jsonify({"status": "error", "message": "Error del sistema: Falta archivo de usuarios"}), 500

             

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        user_data = users.get(username)

        if user_data and check_password_hash(user_data['password'], password):

            if user_data.get('status') != 'approved':

                return jsonify({"status": "error", "message": "Debe esperar aprobación."}), 403

                

            session['user'] = username

            session['role'] = user_data.get('role', 'user')

            return jsonify({"status": "success", "role": session['role']})

        else:

            return jsonify({"status": "error", "message": "Credenciales inválidas"}), 401

            

    except Exception as e:

        print(f"Auth Error: {e}")

        return jsonify({"status": "error", "message": "Error de autenticación"}), 500



@app.route('/api/logout', methods=['POST'])

def logout():

    session.pop('user', None)

    session.pop('role', None)

    return jsonify({"status": "success"})



@app.route('/api/register', methods=['POST'])

def register():

    data = request.json

    username = data.get('username')

    password = data.get('password')

    

    if not username or not password:

        return jsonify({"status": "error", "message": "Faltan credenciales"}), 400

        

    try:

        if not USERS_FILE.exists():

             users = {}

        else:

            with open(USERS_FILE, 'r') as f:

                users = json.load(f)

        

        if username in users:

            return jsonify({"status": "error", "message": "El usuario ya existe"}), 400

            

        users[username] = {

            "password": generate_password_hash(password),

            "role": "user",

            "status": "pending",

            "created_at": time.strftime('%Y-%m-%d %H:%M:%S')

        }

        

        with open(USERS_FILE, 'w') as f:

            json.dump(users, f, indent=4)

            

        return jsonify({"status": "success", "message": "Usuario creado. Debe esperar aprobacion del area para ingresar. Vuelva a intentarlo mas tarde."})

            

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/users', methods=['GET'])

def get_users():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    try:

        if not USERS_FILE.exists():

            return jsonify([])

            

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        user_list = []

        for uname, udata in users.items():

            user_list.append({

                "username": uname,

                "role": udata.get('role', 'user'),

                "status": udata.get('status', 'approved'),

                "created_at": udata.get('created_at', '-'),

                "profile_pic": udata.get('profile_pic')

            })

            

        return jsonify(user_list)

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/profile/upload_photo', methods=['POST'])

def upload_profile_photo():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "No autheticado"}), 401

        

    requester = session.get('user')

    requester_role = session.get('role')

    

    target_user = request.form.get('target_user')

    

    if target_user and requester_role == 'admin':

        username = target_user

    else:

        username = requester

    

    if 'file' not in request.files:

         return jsonify({"status": "error", "message": "No file part"}), 400

         

    file = request.files['file']

    if file.filename == '':

         return jsonify({"status": "error", "message": "No selected file"}), 400

         

    if file:

        try:

            import hashlib

            ext = os.path.splitext(file.filename)[1]

            safe_name = f"{hashlib.md5(username.encode()).hexdigest()}_{int(time.time())}{ext}"

            target_path = PROFILE_PICS_DIR / safe_name

            

            file.save(target_path)

            

            with open(USERS_FILE, 'r') as f:

                users = json.load(f)

            

            if username in users:

                users[username]['profile_pic'] = safe_name

                with open(USERS_FILE, 'w') as f:

                    json.dump(users, f, indent=4)

                    

            return jsonify({"status": "success", "filename": safe_name})

            

        except Exception as e:

             return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/profile/update', methods=['POST'])

def update_profile():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "No autheticado"}), 401

    

    username = session.get('user')

    data = request.json

    

    new_display = data.get('display_name')

    new_pass = data.get('password')

    

    try:

        if not USERS_FILE.exists():

            return jsonify({"status": "error", "message": "User db not found"}), 500

            

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if username not in users:

            return jsonify({"status": "error", "message": "User not found"}), 404

            

        changes = False

        if new_display:

            users[username]['display_name'] = new_display

            session['display_name'] = new_display

            changes = True

            

        if new_pass:

            users[username]['password'] = generate_password_hash(new_pass)

            changes = True

            

        if changes:

            with open(USERS_FILE, 'w') as f:

                json.dump(users, f, indent=4)

                

        return jsonify({"status": "success"})

        

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/approve', methods=['POST'])

def approve_user():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    

    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            users[target_user]['status'] = 'approved'

            with open(USERS_FILE, 'w') as f:

                json.dump(users, f, indent=4)

                

            actor = session.get('user', 'system')

            log_action(actor, "Approved User", f"Target: {target_user}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

            

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/update_role', methods=['POST'])

def update_role():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    new_role = data.get('role')

    

    if new_role not in ['admin', 'user', 'externos']:

         return jsonify({"status": "error", "message": "Rol inválido"}), 400



    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            users[target_user]['role'] = new_role

            with open(USERS_FILE, 'w') as f:

                json.dump(users, f, indent=4)

                

            actor = session.get('user', 'system')

            log_action(actor, "Updated Role", f"Target: {target_user}, New Role: {new_role}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/delete_user', methods=['POST'])

def delete_user():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    

    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            del users[target_user]

            with open(USERS_FILE, 'w') as f:

                json.dump(users, f, indent=4)

                

            actor = session.get('user', 'system')

            log_action(actor, "Deleted User", f"Target: {target_user}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/approve-po', methods=['POST'])
def approve_po():
    try:
        debug_path = BASE_DIR / "Codigos/global_debug.log"
        with open(debug_path, "a") as f:
            f.write(f"{datetime.now()}: Request received for /api/approve-po. Body: {request.get_data(as_text=True)}\n")
    except Exception as e:
        print(f"Log Error: {e}")

    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401
         
    data = request.json
    po_id = data.get('po_id')
    
    if not po_id:
        return jsonify({"status": "error", "message": "Falta PO ID"}), 400
        
    source_folder = DATA_DIR / po_id
    if not source_folder.exists():
        # Check if already processed
        if (PROCESSED_PATH / po_id).exists():
             return jsonify({"status": "success", "message": "Ya fue aprobado"}), 200
        return jsonify({"status": "error", "message": "Carpeta de PO no encontrada"}), 404

    # Resolve User Display Name
    actor = session.get('user', 'system')
    actor_name = actor
    if USERS_FILE.exists():
        try:
            with open(USERS_FILE, 'r') as f:
                users = json.load(f)
                if actor in users:
                    actor_name = users[actor].get('display_name', actor)
        except: pass

    # Start Background Thread
    print(f"[APPROVE] Launching background thread for {po_id}")
    thread = threading.Thread(target=approve_po_background, args=(po_id, data, actor, actor_name))
    thread.daemon = True
    thread.start()
    print(f"[APPROVE] Thread started for {po_id}")
    
    return jsonify({"status": "queued", "message": "Aprobación iniciada en segundo plano"}), 202

def approve_po_background(po_id, data, actor, actor_name):
    import sys
    # Helper for dual logging (File + Console)
    def log_trace(msg, folder=None):
        print(msg, flush=True)
        try: sys.stdout.flush()
        except: pass
        if folder and folder.exists():
            try:
                with open(folder / "_trace_approval.txt", "a", encoding="utf-8") as f:
                    f.write(f"{datetime.now().strftime('%H:%M:%S')} - {msg}\n")
            except: pass
    """
    Background worker:
    1. MOVE folder to Processed (User request: immediate move).
    2. Process Excel Updates.
    3. Mark as Complete.
    """
    try:
        # Determine initial folder to start logging
        log_folder = DATA_DIR / po_id
        if not log_folder.exists():
            log_folder = PROCESSED_PATH / po_id

        log_trace(f"[BACKGROUND] === START Approval for {po_id} by {actor} ===", log_folder)
        
        source_folder = DATA_DIR / po_id
        destination_folder = PROCESSED_PATH / po_id
        
        PROCESSED_PATH.mkdir(parents=True, exist_ok=True)
            
        log_trace(f"[BACKGROUND] Step 1: Requesting move to {destination_folder}", log_folder)
        
        current_work_folder = source_folder
        
        if source_folder.exists():
            if destination_folder.exists():
                log_trace(f"[BACKGROUND] Error: Destination {po_id} already exists!", log_folder)
                # Fallback: Work in source, but this shouldn't happen based on route checks
                current_work_folder = source_folder
            else:
                try:
                    log_trace(f"[BACKGROUND] Moving now...", log_folder)
                    shutil.move(str(source_folder), str(destination_folder))
                    current_work_folder = destination_folder
                    log_trace(f"[BACKGROUND] Folder moved successfully.", current_work_folder)
                    log_action(actor, "Approved PO", f"Moved {po_id} to Processed (Processing Start)")
                except Exception as e:
                    log_trace(f"[BACKGROUND] Failed to move folder: {e}", log_folder)
                    with open(source_folder / "approval_error.log", "w") as f:
                        f.write(f"Move Error: {e}")
                    return

        elif destination_folder.exists():
            # Already moved (maybe retry?)
            current_work_folder = destination_folder
            log_trace(f"[BACKGROUND] Folder was already in Processed. Resuming...", current_work_folder)

        # Create LOCK file to indicate processing is still active
        lock_file = current_work_folder / "processing.lock"
        with open(lock_file, "w") as f:
            f.write("Processing in progress...")
        
        try:
            approved_count = data.get('approved_count', 0)
            total_count = data.get('total_count', 0)
            approved_items = data.get('approved_items', [])

            # Normalize approved items list
            if not isinstance(approved_items, list): approved_items = []
            approved_items = [str(item) for item in approved_items if item]
            if approved_items: approved_count = len(approved_items)

            # Create Metadata
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')
            if total_count == 0:
                pdfs = list(current_work_folder.glob('*.pdf'))
                total_count = len(pdfs)
                approved_count = total_count

            metadata = {
                "approved_by": actor,
                "approved_by_name": actor_name,
                "approved_at": timestamp,
                "status": "Aprobado",
                "counts": {"approved": approved_count, "total": total_count},
                "approved_items": approved_items
            }
            
            with open(current_work_folder / "approval_info.json", "w", encoding='utf-8') as f:
                json.dump(metadata, f, indent=4)
            log_trace(f"[BACKGROUND] Step 2: Metadata created.", current_work_folder)

            # Update Auxiliary Excels (Heavy Operation)
            log_trace(f"[BACKGROUND] Step 3: Updating Excels (Batch Process)...", current_work_folder)
            process_excel_updates_v2(current_work_folder, approved_items, po_id)
            log_trace(f"[BACKGROUND] Step 3: Excel updates finished.", current_work_folder)
            
            # Remove Lock File => Signals Success
            if lock_file.exists():
                os.remove(lock_file)
            log_trace(f"[BACKGROUND] === FINISHED Approval for {po_id} ===", current_work_folder)

        except Exception as exc:
            log_trace(f"[BACKGROUND] Processing Error: {exc}", current_work_folder)
            print(f"[BACKGROUND] Processing Error: {exc}", flush=True)
            with open(current_work_folder / "approval_error.log", "w") as f:
                f.write(f"Processing Error: {exc}")
            # Remove lock so user sees the error
            if lock_file.exists(): os.remove(lock_file)
            return

    except Exception as e:
        print(f"[BACKGROUND] Critical Error for {po_id}: {e}", flush=True)
        try:
            # Try to write log wherever the folder is
            fail_path = PROCESSED_PATH / po_id if (PROCESSED_PATH / po_id).exists() else DATA_DIR / po_id
            if fail_path.exists():
                with open(fail_path / "approval_error.log", "w") as f:
                    f.write(f"System Error: {e}")
        except: pass

@app.route('/api/check-approval-status/<po_id>')
def check_approval_status(po_id):
    # 1. Check Processed Path
    dest = PROCESSED_PATH / po_id
    if dest.exists():
        # It's in Processed, but is it done?
        if (dest / "approval_error.log").exists():
            with open(dest / "approval_error.log", "r") as f: msg = f.read()
            return jsonify({"status": "error", "message": msg})
            
        if (dest / "processing.lock").exists():
            return jsonify({"status": "processing", "message": "Procesando Excels..."})
            
        return jsonify({"status": "success", "message": "Aprobado correctamente"})
        
    # 2. Check Source Path (if it hasn't moved yet or failed before move)
    source = DATA_DIR / po_id
    if source.exists():
        if (source / "approval_error.log").exists():
            with open(source / "approval_error.log", "r") as f: msg = f.read()
            return jsonify({"status": "error", "message": msg})
        return jsonify({"status": "processing", "message": "Iniciando..."})
        
    return jsonify({"status": "error", "message": "Carpeta no encontrada"})



@app.route('/api/save-po-progress', methods=['POST'])

def save_po_progress():

    """Guarda el estado parcial de los productos (aprobados/pendientes) sin mover la carpeta."""

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

         

    data = request.json

    po_id = data.get('po_id')

    approved_items = data.get('approved_items', [])

    

    print(f"[SAVE PROGRESS] Recibido para PO: '{po_id}' - Items: {len(approved_items)}")

    

    if not po_id: 

        print("[SAVE PROGRESS] Error: Falta PO ID")

        return jsonify({"status": "error", "message": "Falta PO ID"}), 400

    

    # Buscar carpeta (primero en En Progreso, fallback Procesado por si se edita historial)

    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        print(f"[SAVE PROGRESS] No en DATA_DIR ({DATA_DIR}), buscando en PROCESSED...")

        target_folder = PROCESSED_PATH / po_id



    if not target_folder.exists():

         print(f"[SAVE PROGRESS] Error: Carpeta no encontrada en ninguna ruta para '{po_id}'")

         return jsonify({"status": "error", "message": "Carpeta PO no encontrada"}), 404



    print(f"[SAVE PROGRESS] Carpeta destino: {target_folder}")



    try:

        # Load existing meta to preserve other fields if any

        meta_file = target_folder / "approval_info.json"

        metadata = {}

        if meta_file.exists():

            try:

                with open(meta_file, 'r', encoding='utf-8') as f:

                    metadata = json.load(f)

            except: pass

            

        # Calculate diff for logging

        old_items = set(metadata.get("approved_items", []))

        new_items_set = set(approved_items)

        added = new_items_set - old_items

        removed = old_items - new_items_set



        # Update fields

        metadata["approved_items"] = approved_items

        metadata["last_updated"] = datetime.now().strftime('%d/%m/%Y %H:%M')

        

        print(f"[SAVE PROGRESS] Escribiendo JSON en {meta_file}")

        with open(meta_file, "w", encoding='utf-8') as f:

            json.dump(metadata, f, indent=4)

            

        # Log Activity with specific details

        actor = session.get('user', 'system')

        if added:

            items_str = ", ".join(list(added))

            log_action(actor, "Aprobado", f"Item(s): {items_str} en {po_id}")

        elif removed:

            items_str = ", ".join(list(removed))

            log_action(actor, "Desaprobado", f"Item(s): {items_str} en {po_id}")

        else:

             log_action(actor, "Updated PO Progress", f"PO: {po_id}, Total: {len(approved_items)}")



        return jsonify({"status": "success"})

    except Exception as e:

        print(f"Error saving progress: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/reverse-approval', methods=['POST'])

def reverse_approval():

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

         

    data = request.json

    po_id = data.get('po_id')

    

    if not po_id:

        return jsonify({"status": "error", "message": "Falta PO ID"}), 400

        

    source_folder = PROCESSED_PATH / po_id

    if not source_folder.exists():

        return jsonify({"status": "error", "message": "Registro procesado no encontrado"}), 404

        

    destination_folder = DATA_DIR / po_id

    # Check if a folder with same name already exists in progress (unlikely but possible)

    if destination_folder.exists():

         return jsonify({"status": "error", "message": "Ya existe una carpeta activa con este ID"}), 400

         

    try:

        shutil.move(str(source_folder), str(destination_folder))

        

        # Update metadata to reflect modification status

        meta_file = destination_folder / "approval_info.json"

        if meta_file.exists():

             try:

                 with open(meta_file, 'r', encoding='utf-8') as f:

                     meta = json.load(f)

                 meta['status'] = 'En Modificación'

                 meta['modified_at'] = datetime.now().strftime('%d/%m/%Y %H:%M')

                 with open(meta_file, 'w', encoding='utf-8') as f:

                     json.dump(meta, f, indent=4)

             except: pass

        

        actor = session.get('user', 'system')

        log_action(actor, "Reversed Approval", f"Moved {po_id} back to In Progress")

        

        return jsonify({"status": "success", "message": "Registro movido a En Progreso"})

    except Exception as e:

        return jsonify({"status": "error", "message": f"Error al mover: {e}"}), 500



@app.route('/api/check_session')

def check_session():

    user = session.get('user')

    role = session.get('role')

    if user:

        # Get extended info

        display_name = user

        profile_pic = None

        email = None
        try:

            if USERS_FILE.exists():

                with open(USERS_FILE, 'r') as f:

                    users = json.load(f)

                    if user in users:

                        display_name = users[user].get('display_name', user)

                        profile_pic = users[user].get('profile_pic')

                        email = users[user].get('email') or users[user].get('correo')

        except: pass

        

        return jsonify({

            "status": "authenticated", 

            "user": user, 

            "role": role,

            "display_name": display_name,

            "profile_pic": profile_pic,

            "email": email

        })

    else:

        return jsonify({"status": "guest"})



@app.route('/files/<path:filename>')

def serve_file(filename):

    # print(f"Requested File: {filename}")

    try:

        # Try DATA_DIR first

        if (DATA_DIR / filename).exists():

            return send_from_directory(DATA_DIR, filename)

        # Try PROCESSED_PATH

        # filename might contain the PO ID as first folder

        # e.g. R016-01-xxxx/file.pdf

        parts = Path(filename).parts

        if len(parts) > 1:

            po_id = parts[0]

            if (PROCESSED_PATH / po_id).exists():

                return send_from_directory(PROCESSED_PATH, filename)

                

        # Last resort fallback (direct match in processed root? unlikely)

        if (PROCESSED_PATH / filename).exists():

             return send_from_directory(PROCESSED_PATH, filename)

             

        return jsonify({"error": "File not found"}), 404

        

    except Exception as e:

        print(f"Error serving file {filename}: {e}")

        return jsonify({"error": str(e)}), 404



# --- OPEN R016 FOLDER (Windows Explorer) ---

@app.route('/api/open-r016-folder')

def open_r016_folder():

    target_path = Path(r"\\192.168.0.55\utn\REGISTROS\R016-01")

    try:

        if not target_path.exists():

            return jsonify({"status": "error", "message": f"Ruta no encontrada: {target_path}"}), 404

        # Use cmd start to avoid explorer falling back to Documentos

        subprocess.Popen(['cmd', '/c', 'start', '', str(target_path)], shell=True)

        return jsonify({"status": "success"})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500


# --- OPEN ISO R019 FOLDERS (Windows Explorer) ---
@app.route('/api/iso-open-folder', methods=['POST'])
def open_iso_folder():
    try:
        data = request.get_json(silent=True) or {}
        key = str(data.get('key', '')).strip().upper()
        allowed = {'R019-01', 'R019-02', 'R019-03', 'R019-04'}
        if key not in allowed:
            return jsonify({"status": "error", "message": "Carpeta inválida."}), 400

        iso_root = BASE_DIR.parent / "Registro de ISO 9001"
        target_path = iso_root / key
        if not target_path.exists():
            return jsonify({"status": "error", "message": f"Ruta no encontrada: {target_path}"}), 404

        subprocess.Popen(['cmd', '/c', 'start', '', str(target_path)], shell=True)
        return jsonify({"status": "success"})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# --- ISO NEXT REGISTRY (R019-03 LISTADO) ---
@app.route('/api/iso-next-registry')
def iso_next_registry():
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    base_name = "R019-03 - Modelo"
    search_dirs = [
        iso_root,
        iso_root / "R019-03",
    ]
    candidates = []
    for d in search_dirs:
        candidates.extend([
            d / f"{base_name}.xlsm",
            d / f"{base_name}.xlsx",
            d / f"{base_name}.xls",
        ])
    target = next((p for p in candidates if p.exists()), None)
    if target is None:
        # Fallback: any file starting with the base name in known dirs
        matches = []
        for d in search_dirs:
            matches.extend(list(d.glob(f"{base_name}*")))
        target = matches[0] if matches else None

    if target is None or not target.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró el archivo R019-03 en {iso_root}"
        }), 404

    try:
        wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
        ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active

        last_num = None
        year_two = datetime.now().strftime("%y")
        last_seq = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            val_a = row[0] if len(row) > 0 else None
            val_b = row[1] if len(row) > 1 else None

            if val_a is not None and str(val_a).strip() != "":
                try:
                    last_num = int(float(str(val_a).strip()))
                except Exception:
                    pass

            if val_b:
                s = str(val_b).strip()
                m = re.match(r"BP-(\d{2})(\d{3})", s)
                if m and m.group(1) == year_two:
                    try:
                        seq = int(m.group(2))
                        if seq > last_seq:
                            last_seq = seq
                    except Exception:
                        pass

        wb.close()

        next_num = (last_num if last_num is not None else 174) + 1
        next_bp = f"BP-{year_two}{last_seq + 1:03d}"

        return jsonify({
            "status": "success",
            "next_num": next_num,
            "next_bp": next_bp,
            "source": str(target)
        })
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-r01903-records')
def iso_r01903_records():
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    base_name = "R019-03 - Modelo"
    search_dirs = [
        iso_root,
        iso_root / "R019-03",
    ]
    candidates = []
    for d in search_dirs:
        candidates.extend([
            d / f"{base_name}.xlsm",
            d / f"{base_name}.xlsx",
            d / f"{base_name}.xls",
        ])
    target = next((p for p in candidates if p.exists()), None)
    if target is None:
        matches = []
        for d in search_dirs:
            matches.extend(list(d.glob(f"{base_name}*")))
        target = matches[0] if matches else None

    if target is None or not target.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró el archivo R019-03 en {iso_root}"
        }), 404

    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        return str(val).strip() if val is not None else ""

    try:
        wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
        ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active

        records = []
        for row in ws.iter_rows(min_row=7, values_only=True):
            if not row or len(row) < 7:
                continue
            val_a = row[0]
            val_b = row[1] if len(row) > 1 else None
            if (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
                continue
            records.append({
                "numero": str(val_a).strip() if val_a is not None else "",
                "descripcion": str(val_b).strip() if val_b is not None else "",
                "solicitante": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                "fecha_inicio": fmt_date(row[3]) if len(row) > 3 else "",
                "etapa": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                "situacion": str(row[5]).strip() if len(row) > 5 and row[5] is not None else "",
                "fecha_fin": fmt_date(row[6]) if len(row) > 6 else "",
            })

        wb.close()
        return jsonify({"status": "success", "records": records, "source": str(target)})
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer R019-03: {e}"}), 500


@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    user = session.get('user')
    display_name = _resolve_display_name(user)
    email = _resolve_user_email(user)
    keys = {_normalize_user_key(user), _normalize_user_key(display_name)}
    if email:
        keys.add(_normalize_user_key(email))

    items = _load_notifications()
    unread = []
    for n in items:
        if n.get("read"):
            continue
        recip_key = _normalize_user_key(n.get("recipient"))
        recip_email = _normalize_user_key(n.get("recipient_email"))
        if recip_key in keys or (recip_email and recip_email in keys):
            unread.append(n)
    unread.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return jsonify({"status": "success", "notifications": unread, "unread": len(unread)})


@app.route('/api/notifications/mark-read', methods=['POST'])
def mark_notification_read():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    data = request.get_json(silent=True) or {}
    notif_id = data.get("id")
    if not notif_id:
        return jsonify({"status": "error", "message": "ID requerido"}), 400

    items = _load_notifications()
    updated = False
    for n in items:
        if n.get("id") == notif_id:
            n["read"] = True
            updated = True
            break
    if updated:
        _save_notifications(items)
    return jsonify({"status": "success", "updated": updated})


@app.route('/api/iso-r01902-events')
def iso_r01902_events():
    bp = (request.args.get("bp") or "").strip()
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400
    try:
        events, target = _read_r01902_events_for_bp(bp)
        return jsonify({"status": "success", "events": events, "file": str(target)})
    except FileNotFoundError as e:
        return jsonify({"status": "error", "message": str(e)}), 404
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer R019-02: {e}"}), 500


def _read_r01902_events_for_bp(bp: str):
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    r01902_dir = iso_root / "R019-02"
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp

    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)

    if target is None and r01902_dir.exists():
        for p in r01902_dir.glob("R019-02 Rev03 -*.*"):
            if safe_bp.lower() in p.stem.lower():
                target = p
                break

    if target is None or not target.exists():
        raise FileNotFoundError(f"No se encontró el archivo R019-02 para {bp}")

    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        return str(val).strip() if val is not None else ""

    wb = openpyxl.load_workbook(str(target), read_only=True, data_only=True)
    ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active

    events = []
    empty_streak = 0
    max_row = ws.max_row or 0

    for row in range(4, max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, 9)]
        if all(v is None or str(v).strip() == "" for v in values):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        events.append({
            "row": row,
            "fecha": fmt_date(values[0]),
            "etapa": str(values[1]).strip() if values[1] is not None else "",
            "area": str(values[2]).strip() if values[2] is not None else "",
            "empresa": str(values[3]).strip() if values[3] is not None else "",
            "descripcion": str(values[4]).strip() if values[4] is not None else "",
            "resultados": str(values[5]).strip() if values[5] is not None else "",
            "accion": str(values[6]).strip() if values[6] is not None else "",
            "usuario": str(values[7]).strip() if values[7] is not None else "",
        })

    wb.close()
    return events, target




def _find_r01904_model(iso_root: Path) -> Path:
    candidates = [
        iso_root / "R019-04" / "R019-04-Modelo.mpp",
        iso_root / "Codigos" / "R019-04" / "R019-04-Modelo.mpp",
    ]
    for cand in candidates:
        if cand.exists():
            return cand
    r01904_dir = iso_root / "R019-04"
    if r01904_dir.exists():
        matches = list(r01904_dir.glob("R019-04-Modelo*.mpp"))
        if matches:
            return matches[0]
        mpp_files = list(r01904_dir.glob("*.mpp"))
        if mpp_files:
            return mpp_files[0]
    raise FileNotFoundError("No se encontro el modelo R019-04.")


def _ensure_r01904_created(bp: str) -> Path:
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    r01904_dir = iso_root / "R019-04"
    r01904_dir.mkdir(parents=True, exist_ok=True)
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    dest = r01904_dir / f"R019-04 - {safe_bp}.mpp"
    if dest.exists():
        return dest
    model = _find_r01904_model(iso_root)
    last_err = None
    for attempt in range(1, 4):
        try:
            dest.write_bytes(model.read_bytes())
            return dest
        except PermissionError as e:
            last_err = e
            time.sleep(0.4 * attempt)
    if last_err:
        raise last_err
    return dest


def _generate_r01904(bp: str, allow_empty: bool = False) -> Path:
    import sys

    events, _ = _read_r01902_events_for_bp(bp)
    if not events:
        if allow_empty:
            return _ensure_r01904_created(bp)
        raise ValueError("R019-02 no tiene eventos para este BP.")

    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    r01904_dir = iso_root / "R019-04"
    r01904_dir.mkdir(parents=True, exist_ok=True)
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    dest = r01904_dir / f"R019-04 - {safe_bp}.mpp"

    csv_path = Path(tempfile.gettempdir()) / f"r01904_{safe_bp}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["fecha", "etapa", "area", "empresa", "descripcion", "resultado", "accion", "usuario"]
        )
        writer.writeheader()
        for ev in events:
            writer.writerow({
                "fecha": ev.get("fecha", ""),
                "etapa": ev.get("etapa", ""),
                "area": ev.get("area", ""),
                "empresa": ev.get("empresa", ""),
                "descripcion": ev.get("descripcion", ""),
                "resultado": ev.get("resultados", "") or ev.get("resultado", ""),
                "accion": ev.get("accion", ""),
                "usuario": ev.get("usuario", "")
            })

    script_path = iso_root / "Codigos" / "R019-04" / "fill_r01904.py"
    model_path = iso_root / "Codigos" / "R019-04" / "modelo_duraciones.json"
    if not script_path.exists():
        raise FileNotFoundError("No se encontro fill_r01904.py")

    cmd = [
        sys.executable,
        str(script_path),
        str(csv_path),
        "--modelo",
        str(model_path),
        "--out",
        str(dest)
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        err = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(err or "No se pudo generar R019-04.")
    return dest

@app.route('/api/iso-r01904-generate', methods=['POST'])
def iso_r01904_generate():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or "").strip()
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400

    try:
        dest = _generate_r01904(bp, allow_empty=False)
        return jsonify({"status": "success", "file": str(dest)})
    except FileNotFoundError as e:
        return jsonify({"status": "error", "message": str(e)}), 404
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo generar R019-04: {e}"}), 500


@app.route('/api/iso-r01902-approve', methods=['POST'])
def iso_r01902_approve():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or "").strip()
    row = data.get("row")
    status = (data.get("status") or "Aprobado").strip()
    accion = (data.get("accion") or "").strip()
    if not bp or not row:
        return jsonify({"status": "error", "message": "BP y fila requeridos"}), 400

    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    r01902_dir = iso_root / "R019-02"
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)
    if target is None:
        return jsonify({"status": "error", "message": f"No se encontr\u00f3 el archivo R019-02 para {bp}"}), 404

    def _is_permission_error(err: Exception) -> bool:
        if isinstance(err, PermissionError):
            return True
        if isinstance(err, OSError):
            winerr = getattr(err, "winerror", None)
            if winerr in (32, 33):
                return True
            if getattr(err, "errno", None) == 13:
                return True
        msg = str(err).lower()
        return "permission denied" in msg or "being used by another process" in msg

    last_err = None
    for attempt in range(1, 4):
        wb = None
        try:
            wb = openpyxl.load_workbook(str(target), keep_vba=target.suffix.lower() == ".xlsm")
            ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active
            ws.cell(row=int(row), column=6).value = status
            if accion:
                ws.cell(row=int(row), column=7).value = accion
            wb.save(target)
            return jsonify({"status": "success"})
        except Exception as e:
            last_err = e
            if _is_permission_error(e):
                time.sleep(0.4 * attempt)
                continue
            break
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    if last_err:
        if _is_permission_error(last_err):
            return jsonify({"status": "error", "message": "No se pudo aprobar el evento: el archivo R019-02 está en uso. Cierre Excel y reintente."}), 500
        return jsonify({"status": "error", "message": f"No se pudo aprobar el evento: {last_err}"}), 500
    return jsonify({"status": "error", "message": "No se pudo aprobar el evento."}), 500


@app.route('/api/iso-r01902-append', methods=['POST'])
def iso_r01902_append():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or data.get("BP") or "").strip()
    event = data.get("event") or {}
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400

    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    r01902_dir = iso_root / "R019-02"
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp

    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)

    if target is None and r01902_dir.exists():
        for p in r01902_dir.glob("R019-02 Rev03 -*.*"):
            if safe_bp.lower() in p.stem.lower():
                target = p
                break

    if target is None or not target.exists():
        return jsonify({"status": "error", "message": f"No se encontr\u00f3 el archivo R019-02 para {bp}"}), 404

    try:
        codes_r01902 = iso_root / "Codigos" / "R019-02"
        mod02 = _load_fill_r01902_module(codes_r01902)
        if hasattr(mod02, "append_r01902_event"):
            row = mod02.append_r01902_event(target, event)
        else:
            # Fallback: write with openpyxl if function not available
            wb = openpyxl.load_workbook(str(target), keep_vba=target.suffix.lower() == ".xlsm")
            try:
                ws = wb["Listado"] if "Listado" in wb.sheetnames else wb.active
                row = (ws.max_row or 3) + 1
                ws.cell(row=row, column=1).value = event.get("fecha") or ""
                ws.cell(row=row, column=2).value = event.get("etapa") or ""
                ws.cell(row=row, column=3).value = event.get("area") or ""
                ws.cell(row=row, column=4).value = event.get("empresa") or ""
                ws.cell(row=row, column=5).value = event.get("descripcion") or ""
                ws.cell(row=row, column=6).value = event.get("resultados") or ""
                ws.cell(row=row, column=7).value = event.get("accion") or ""
                ws.cell(row=row, column=8).value = event.get("usuario") or ""
                wb.save(target)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        # NotificaciÃ³n de aprobaciÃ³n si aplica
        if event.get("requiere_aprobacion"):
            try:
                for approver in ("Luciano Cochis", "Lorenzo Contigiani"):
                    _add_notification(
                        approver,
                        "approval",
                        f"Aprobacion pendiente para {bp}",
                        {"bp": bp, "row": row, "etapa": event.get("etapa", "")}
                    )
            except Exception:
                pass

        r01903_updated = False
        r01903_error = None
        try:
            codes_r01903 = iso_root / "Codigos" / "R019-03"
            mod03 = _load_fill_r01903_module(codes_r01903)
            if hasattr(mod03, "update_r01903_status"):
                etapa = (event.get("etapa") or "").strip()
                situacion = (event.get("situacion") or "").strip()
                fecha_fin = None
                if etapa.lower() == "cierre":
                    fecha_fin = (event.get("fecha") or "").strip()
                r01903_updated = mod03.update_r01903_status(bp, etapa, situacion, fecha_fin)
            else:
                r01903_error = "fill_r01903.py no expone update_r01903_status"
        except Exception as e:
            r01903_error = f"No se pudo actualizar R019-03: {e}"
        r01904_error = None
        try:
            _generate_r01904(str(bp), allow_empty=True)
        except Exception as e:
            r01904_error = f"No se pudo generar R019-04: {e}"

        resp = {"status": "success", "row": row}
        if r01903_error:
            resp["r01903_error"] = r01903_error
        if r01904_error:
            resp["r01904_error"] = r01904_error
        resp["r01903_updated"] = bool(r01903_updated)
        return jsonify(resp)
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo guardar evento en R019-02: {e}"}), 500


def _iso_r01901_paths():
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    docs_dir = iso_root / "R019-01"
    codes_dir = iso_root / "Codigos" / "R019-01"
    return iso_root, docs_dir, codes_dir


def _load_iso_payload_template(codes_dir: Path) -> dict:
    template_path = codes_dir / "payload_template.json"
    if not template_path.exists():
        return {}
    try:
        with template_path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}


def _load_fill_r01901_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01901.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontró fill_r01901.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01901", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el módulo fill_r01901.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _load_fill_r01903_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01903.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontró fill_r01903.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01903", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el módulo fill_r01903.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _load_fill_r01902_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01902.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontró fill_r01902.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01902", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el módulo fill_r01902.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _r01903_pending_path(iso_root: Path) -> Path:
    pending_dir = iso_root / "R019-03"
    return pending_dir / "R019-03-pendientes.json"


def _append_r01903_pending(iso_root: Path, payload: dict, reason=None) -> int:
    pending_path = _r01903_pending_path(iso_root)
    pending_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        if pending_path.exists():
            with open(pending_path, "r", encoding="utf-8") as f:
                items = json.load(f) or []
        else:
            items = []
    except Exception:
        items = []
    items.append({
        "created_at": datetime.now().isoformat(),
        "reason": reason or "",
        "payload": payload
    })
    with open(pending_path, "w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)
    return len(items)


def _notifications_path() -> Path:
    NOTIFICATIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    return NOTIFICATIONS_FILE


def _normalize_user_key(value: str) -> str:
    return str(value or "").strip().lower()


def _destinatarios_path() -> Path:
    return BASE_DIR.parent / "Registro de Actividad" / "Codigos" / "config" / "destinatarios.csv"


def _load_destinatarios() -> list:
    path = _destinatarios_path()
    if not path.exists():
        return []
    rows = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if not row or len(row) < 3:
                    continue
                nombre = str(row[0]).strip()
                apellido = str(row[1]).strip()
                email = str(row[2]).strip().lower()
                full = f"{nombre} {apellido}".strip()
                if full or email:
                    rows.append({
                        "nombre": nombre,
                        "apellido": apellido,
                        "email": email,
                        "full": full,
                        "key": _normalize_user_key(full)
                    })
    except Exception:
        return []
    return rows


def _resolve_email_from_name(full_name: str) -> str:
    target = _normalize_user_key(full_name)
    if not target:
        return ""
    for row in _load_destinatarios():
        if row.get("key") == target:
            return row.get("email", "")
    return ""


def _resolve_name_from_email(email: str) -> str:
    if not email:
        return ""
    email_key = email.strip().lower()
    for row in _load_destinatarios():
        if row.get("email", "").lower() == email_key:
            return row.get("full", "")
    return ""


def _resolve_user_email(user: str) -> str:
    if not user:
        return ""
    if "@" in user:
        return user.strip().lower()
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                users = json.load(f) or {}
            if user in users:
                email_val = users[user].get("email") or users[user].get("correo")
                if email_val:
                    return str(email_val).strip().lower()
    except Exception:
        return ""
    return ""


def _load_notifications() -> list:
    path = _notifications_path()
    if not path.exists():
        return []
    try:
        with path.open("r", encoding="utf-8") as f:
            data = json.load(f) or []
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_notifications(items: list) -> None:
    path = _notifications_path()
    with path.open("w", encoding="utf-8") as f:
        json.dump(items, f, ensure_ascii=False, indent=2)


def _resolve_display_name(user: str) -> str:
    if not user:
        return ""
    display_name = user
    email = ""
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                users = json.load(f) or {}
            if user in users:
                display_name = users[user].get("display_name", user)
                email = users[user].get("email") or users[user].get("correo") or ""
            else:
                for uname, data in users.items():
                    email = (data.get("email") or data.get("correo") or "").strip().lower()
                    if email and email == user.strip().lower():
                        display_name = data.get("display_name", uname)
                        break
    except Exception:
        pass
    if "@" in user:
        email = user.strip().lower()
    if email:
        name_from_email = _resolve_name_from_email(email)
        if name_from_email:
            return name_from_email
    return display_name or user


def _add_notification(recipient_name: str, ntype: str, message: str, payload=None) -> dict:
    items = _load_notifications()
    recipient_email = ""
    if recipient_name and "@" in recipient_name:
        recipient_email = recipient_name.strip().lower()
    else:
        recipient_email = _resolve_email_from_name(recipient_name)
    entry = {
        "id": str(uuid.uuid4()),
        "recipient": recipient_name,
        "recipient_key": _normalize_user_key(recipient_name),
        "recipient_email": recipient_email,
        "type": ntype,
        "message": message,
        "payload": payload or {},
        "created_at": datetime.now().isoformat(),
        "read": False,
    }
    items.append(entry)
    _save_notifications(items)
    return entry

# --- ISO RESPONSABLES (Destinatarios) ---
@app.route('/api/iso-responsables', methods=['GET'])
def iso_responsables():
    try:
        config_dir = BASE_DIR.parent / "Registro de Actividad" / "Codigos" / "config"
        destinatarios_file = config_dir / "destinatarios.csv"
        if not destinatarios_file.exists():
            return jsonify({
                "status": "error",
                "message": f"No se encontró destinatarios.csv en {config_dir}"
            }), 404

        names = []
        seen = set()
        with open(destinatarios_file, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if not row or len(row) < 2:
                    continue
                nombre = str(row[0]).strip()
                apellido = str(row[1]).strip()
                full = f"{nombre} {apellido}".strip()
                key = full.lower()
                if not full or key in seen:
                    continue
                seen.add(key)
                names.append(full)

        return jsonify({"status": "success", "data": names})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-destinatarios', methods=['GET'])
def iso_destinatarios():
    try:
        rows = _load_destinatarios()
        data = [{"name": r.get("full", ""), "email": r.get("email", "")} for r in rows if r.get("full") or r.get("email")]
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# --- ISO R019-03 PENDINGS SYNC ---
@app.route('/api/iso-r01903-pending-sync', methods=['POST'])
def iso_r01903_pending_sync():
    iso_root = BASE_DIR.parent / "Registro de ISO 9001"
    pending_path = _r01903_pending_path(iso_root)
    if not pending_path.exists():
        return jsonify({"status": "success", "processed": 0, "remaining": 0})

    try:
        with open(pending_path, "r", encoding="utf-8") as f:
            items = json.load(f) or []
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer pendientes: {e}"}), 500

    if not items:
        return jsonify({"status": "success", "processed": 0, "remaining": 0})

    codes_r01903 = iso_root / "Codigos" / "R019-03"
    mod03 = _load_fill_r01903_module(codes_r01903)
    locked_cls = getattr(mod03, "FileLockedError", None)

    processed = 0
    remaining = []
    for idx, item in enumerate(items):
        payload = (item or {}).get("payload") or {}
        try:
            mod03.generate_r01903(payload, inplace=True)
            processed += 1
        except Exception as e:
            if locked_cls and isinstance(e, locked_cls):
                remaining = items[idx:]
                break
            return jsonify({"status": "error", "message": f"Error al sincronizar pendientes: {e}"}), 500

    if remaining:
        with open(pending_path, "w", encoding="utf-8") as f:
            json.dump(remaining, f, ensure_ascii=False, indent=2)
    else:
        try:
            pending_path.unlink()
        except Exception:
            pass

    return jsonify({
        "status": "success",
        "processed": processed,
        "remaining": len(remaining)
    })


# --- ISO R019-01 GENERATION (WORD) ---
@app.route('/api/iso-generate-r01901', methods=['POST'])
def iso_generate_r01901():
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or data
    bp = data.get("bp") or payload.get("BP") or payload.get("bp") or "BP-XXXXX"

    iso_root, docs_dir, codes_dir = _iso_r01901_paths()
    if not docs_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró la carpeta R019-01 en {docs_dir}"
        }), 404
    if not codes_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró la carpeta de códigos R019-01 en {codes_dir}"
        }), 404

    template_payload = _load_iso_payload_template(codes_dir)
    merged = dict(template_payload)
    merged.update(payload or {})
    if bp:
        merged["BP"] = bp

    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", str(bp).strip()) or "BP-XXXXX"
    filename = f"R019-01-DATOS DE ENTRADA-{safe_bp}.docx"
    output_path = docs_dir / filename

    if output_path.exists():
        return jsonify({
            "status": "error",
            "message": f"El archivo ya existe: {output_path}",
            "path": str(output_path)
        }), 409

    try:
        mod = _load_fill_r01901_module(codes_dir)
        if not hasattr(mod, "generate_r01901"):
            return jsonify({
                "status": "error",
                "message": "fill_r01901.py no expone generate_r01901"
            }), 500
        mod.generate_r01901(merged, output_path)
        # Notificar firma pendiente si Responsable de DiseÃ±o es distinto al usuario actual
        try:
            responsable = (merged.get("Responsable_DiseÃ±o") or merged.get("Responsable_Diseño") or "").strip()
            if responsable:
                actor = session.get("user") if session.get("user") else ""
                actor_name = _resolve_display_name(actor)
                if _normalize_user_key(responsable) != _normalize_user_key(actor_name):
                    _add_notification(
                        responsable,
                        "signature",
                        f"Firma pendiente para {safe_bp}",
                        {"bp": safe_bp, "numero": merged.get("Numero_de_Registro") or ""}
                    )
        except Exception:
            pass
        # Actualizar R019-03 (in-place) o dejar pendiente si el archivo está en uso
        pending_info = None
        try:
            codes_r01903 = iso_root / "Codigos" / "R019-03"
            mod03 = _load_fill_r01903_module(codes_r01903)
            if hasattr(mod03, "generate_r01903"):
                mod03.generate_r01903(merged, inplace=True)
            else:
                raise RuntimeError("fill_r01903.py no expone generate_r01903")
        except Exception as e:
            locked_cls = getattr(mod03, "FileLockedError", None)
            if locked_cls and isinstance(e, locked_cls):
                pending_count = _append_r01903_pending(iso_root, merged, str(e))
                pending_info = {
                    "pending": True,
                    "pending_count": pending_count,
                    "pending_message": "R019-03 en cola (archivo en uso)."
                }
            else:
                return jsonify({"status": "error", "message": f"R019-01 generado, pero falló R019-03: {e}"}), 500

        # Crear R019-02 (in-place copia por BP)
        try:
            codes_r01902 = iso_root / "Codigos" / "R019-02"
            mod02 = _load_fill_r01902_module(codes_r01902)
            if hasattr(mod02, "generate_r01902"):
                r01902_dir = iso_root / "R019-02"
                r01902_dir.mkdir(parents=True, exist_ok=True)
                r01902_name = f"R019-02 Rev03 - {safe_bp}.xlsm"
                r01902_path = r01902_dir / r01902_name
                if not r01902_path.exists():
                    mod02.generate_r01902(merged, r01902_path)
            else:
                raise RuntimeError("fill_r01902.py no expone generate_r01902")
        except Exception as e:
            return jsonify({"status": "error", "message": f"R019-01 generado, pero falló R019-02: {e}"}), 500
        r01904_error = None
        try:
            _generate_r01904(str(safe_bp), allow_empty=True)
        except Exception as e:
            r01904_error = f"No se pudo generar R019-04: {e}"

        resp = {
            "status": "success",
            "file": filename,
            "path": str(output_path),
            "root": str(iso_root)
        }
        if pending_info:
            resp.update(pending_info)
        if r01904_error:
            resp["r01904_error"] = r01904_error
        return jsonify(resp)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-r01901-update', methods=['POST'])
def iso_update_r01901():
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or {}
    bp = data.get("bp") or payload.get("BP") or payload.get("bp") or "BP-XXXXX"

    iso_root, docs_dir, codes_dir = _iso_r01901_paths()
    if not docs_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró la carpeta R019-01 en {docs_dir}"
        }), 404
    if not codes_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró la carpeta de códigos R019-01 en {codes_dir}"
        }), 404

    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", str(bp).strip()) or "BP-XXXXX"
    filename = f"R019-01-DATOS DE ENTRADA-{safe_bp}.docx"
    output_path = docs_dir / filename
    if not output_path.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontró el archivo: {output_path}"
        }), 404

    try:
        mod = _load_fill_r01901_module(codes_dir)
        if not hasattr(mod, "update_r01901"):
            return jsonify({
                "status": "error",
                "message": "fill_r01901.py no expone update_r01901"
            }), 500
        mod.update_r01901(payload, output_path)
        return jsonify({"status": "success", "file": filename, "path": str(output_path)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500




# --- MANUAL UPLOAD: LIST POs FROM resumen_all ---

@app.route('/api/manual-register-po', methods=['POST'])

def manual_register_po():

    data = request.json

    po_id = data.get('po_id')

    if not po_id:

        return jsonify({"status": "error", "message": "Missing PO ID"}), 400



    if not MANUAL_INGRESOS_FILE.exists():

        return jsonify({"status": "error", "message": "File resumen_all.csv not found"}), 404



    # 1. Read & Filter CSV

    encodings = ['utf-8-sig', 'cp1252', 'latin-1']

    content = None

    for enc in encodings:

        try:

            content = MANUAL_INGRESOS_FILE.read_text(encoding=enc)

            break

        except Exception:

            continue

    

    if content is None:

        return jsonify({"status": "error", "message": "Could not read resumen_all.csv"}), 500



    lines = [ln for ln in content.splitlines() if ln.strip()]

    if not lines:

         return jsonify({"status": "error", "message": "resumen_all is empty"}), 400



    # Validate delimiter

    first_line = lines[0]

    delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','

    

    reader = csv.reader(lines, delimiter=delimiter)

    all_rows = list(reader)

    if not all_rows:

        return jsonify({"status": "error", "message": "No rows in CSV"}), 400



    header = all_rows[0]

    data_rows = all_rows[1:]

    

    target_po_clean = str(po_id).strip().lower()

    

    filtered_rows = []

    

    for i, row in enumerate(data_rows):

        match = False

        val1 = ""

        val2 = ""

        

        # Check col 2 (index 1)

        if len(row) > 1:

            val1 = row[1].strip().lower()

            if target_po_clean in val1: match = True

        

        # Check col 3 (index 2)

        if len(row) > 2:

            val2 = row[2].strip().lower()

            if target_po_clean in val2: match = True

            

        if match:

            filtered_rows.append(row)



    if not filtered_rows:

        return jsonify({"status": "error", "message": "No matching rows found for this PO"}), 404



    # PADRON DE DUPLICADOS (DEBUG MODE)

    print(f"--- Verificando Duplicados para PO {po_id} ---")

    roots_to_scan = [

        BASE_DIR / "P2 - Purchase Order/En Progreso",

        BASE_DIR / "P2 - Purchase Order/Procesado"

    ]

    po_target_str = str(po_id).strip()

    

    dupe_found = None

    for root in roots_to_scan:

        if not root.exists(): 

            print(f"Ruta no existe: {root}")

            continue

            

        print(f"Escaneando raiz: {root.name}")

        

        # Mapeo de nombres amigables

        friendly_name = "Registros Pendientes" if "En Progreso" in root.name else "Historial de PO"

        

        for folder in root.iterdir():

            if not folder.is_dir(): continue

            

            # 1. Nombre Carpeta

            if re.search(rf"\bPO\s*[-_]?\s*{re.escape(po_target_str)}\b", folder.name, re.IGNORECASE):

                dupe_found = friendly_name

                print(f"!!! DUPLICADO ENCONTRADO POR NOMBRE: {dupe_found} ({folder.name})")

                break



            # 2. Contenido CSV

            res_csv = folder / "resumen.csv"

            if res_csv.exists():

                try:

                    c_txt = res_csv.read_text(encoding='latin-1', errors='ignore') 

                    if re.search(rf"\b{re.escape(po_target_str)}\b", c_txt):

                         dupe_found = friendly_name

                         print(f"!!! DUPLICADO ENCONTRADO POR CSV: {dupe_found} ({folder.name})")

                         break

                except Exception as e:

                    print(f"Error leyendo {folder.name}: {e}")

        if dupe_found: break

        

    if dupe_found:

         print(f"RECHAZADO: Duplicado en {dupe_found}")

         return jsonify({"status": "error", "message": f"La PO {po_id} ya existe en {dupe_found}"}), 400

    

    print("--- PO Valida, procediendo ---")



    # 2. Determine Next R Number

    if not IN_PROCESS_DIR.exists():

        IN_PROCESS_DIR.mkdir(parents=True, exist_ok=True)



    max_r = 0

    for item in IN_PROCESS_DIR.iterdir():

        if item.is_dir():

             # Match R number. Pattern: "Registro - R0060 - ..."

            match = re.search(r'Registro\s*-\s*R(\d+)', item.name, re.IGNORECASE)

            if match:

                try:

                    r_num = int(match.group(1))

                    if r_num > max_r: max_r = r_num

                except: pass

    

    next_r = max_r + 1

    # PREVIOUS: folder_name = f"Registro - R{next_r:04d} - Manual PO {po_id.upper()}"

    # REQUESTED FORMAT: "Registro - R0070 - 2026-01-14"

    today_str = datetime.now().strftime("%Y-%m-%d")

    folder_name = f"Registro - R{next_r:04d} - {today_str}"

    new_folder_path = IN_PROCESS_DIR / folder_name

    

    try:

        new_folder_path.mkdir(parents=True, exist_ok=True)

        

        # 3. Write resumen.csv

        out_csv = new_folder_path / "resumen.csv"

        

        # We write simply with utf-8-sig and delimiter ';'

        with open(out_csv, 'w', encoding='utf-8-sig', newline='') as f:

            writer = csv.writer(f, delimiter=';')

            writer.writerow(header)

            writer.writerows(filtered_rows)

            

        # Log it

        actor = session.get('user', 'system')

        log_action(actor, "Manual Register", f"Created {folder_name} with {len(filtered_rows)} rows")



        return jsonify({"status": "success", "message": f"Registro creado: {folder_name}"})

        

    except Exception as e:

        return jsonify({"status": "error", "message": f"Error creating folder/file: {e}"}), 500



@app.route('/api/run-step2-manual', methods=['POST'])

def run_step2_manual():

    try:

        script_dir = BASE_DIR / "Codigos"

        import sys

        

        # Reutilizamos la lógica de sincronización en segundo plano: Step 2 (Entrantes) + Step 3 (En Progreso)

        start_sync_thread(source="manual")

        return jsonify({"status": "success", "message": "Sincronización manual iniciada en segundo plano"})





        

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/manual-po-list')

def manual_po_list():

    import sys

    # Actualizar CSVs (Run Step 1 dry-run)

    try:

        script_dir = BASE_DIR / "Codigos"

        updated = subprocess.run(

            [sys.executable, 'procesar_ingresos.py', '--just-update-csv'],

            cwd=str(script_dir),

            capture_output=True,

            text=True

        )

        if updated.returncode != 0:

            print(f"Error updating CSVs: {updated.stderr}")

    except Exception as e:

         print(f"Failed to run procesar_ingresos update: {e}")



    if not MANUAL_INGRESOS_FILE.exists():

        return jsonify([])



    encodings = ['utf-8-sig', 'cp1252', 'latin-1']

    content = None

    for enc in encodings:

        try:

            content = MANUAL_INGRESOS_FILE.read_text(encoding=enc)

            break

        except Exception:

            continue



    if content is None:

        return jsonify({"status": "error", "message": "No se pudo leer resumen_all.csv"}), 500



    # Detect delimiter using first non-empty line

    lines = [ln for ln in content.splitlines() if ln.strip()]

    if not lines:

         return jsonify([])



    first_line = lines[0]

    delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','



    reader = csv.DictReader(StringIO(content), delimiter=delimiter)



    # Aggregation Dict: PO -> { date: '...', count: 0 }

    po_data = {}



    for row in reader:

        if not isinstance(row, dict):

            continue

            

        # 1. Extract PO Number

        obs = (row.get('Observac_PO') or '').strip()

        po_val = row.get('PO') or ''

        

        po_str = None

        # Try to find digits in Observac_PO first, then PO column

        m = re.search(r'(\d+)', obs)

        if m:

            po_str = m.group(1)

        else:

            m2 = re.search(r'(\d+)', po_val)

            if m2:

                po_str = m2.group(1)

                

        if not po_str:

            continue



        # Clean leading zeros

        cleaned_po = po_str.lstrip('0') or po_str

        

        # 2. Extract Date (Try common keys or index based fallback if needed, but DictReader relies on keys)

        # Based on file check: Header is 'Columna_D' for the 4th column

        date_val = (row.get('Columna_D') or '').strip()

        

        # Fallback if 'Columna_D' key missing but 'Fecha' present or generic

        if not date_val:

            date_val = (row.get('Fecha') or '').strip()



        # 3. Aggregate

        if cleaned_po not in po_data:

            po_data[cleaned_po] = {

                "po": cleaned_po,

                "fecha_ingreso": date_val,

                "count": 0

            }

        

        po_data[cleaned_po]["count"] += 1

        # Update date if missing (or overwrite? usually dates are same for a PO block)

        if not po_data[cleaned_po]["fecha_ingreso"] and date_val:

            po_data[cleaned_po]["fecha_ingreso"] = date_val



    # Convert to list and sort

    final_list = list(po_data.values())



    def sort_key(item):

        try:

            return int(item['po'])

        except:

            return item['po']



    final_list.sort(key=sort_key, reverse=True) # Newest POs first usually



    return jsonify(final_list)



# --- API FOR HISTORIAL DETAILS ---

@app.route('/api/historial-po-details/<po_id>')

def get_historial_po_details(po_id):

    if not PROCESSED_PATH.exists():

        return jsonify([])

        

    target_folder = PROCESSED_PATH / po_id

    if not target_folder.exists():

        return jsonify({"status": "error", "message": "PO not found"}), 404

        

    products = []

    

    # Read approval metadata for global info

    app_meta = {}

    meta_path = target_folder / "approval_info.json"

    if meta_path.exists():

        try:

             with open(meta_path, 'r', encoding='utf-8') as f:

                 app_meta = json.load(f)

        except: pass



    # Scan for PDFs

    approved_items = []

    if isinstance(app_meta, dict):

        tmp = app_meta.get('approved_items', [])

        if isinstance(tmp, list):

            approved_items = [str(x) for x in tmp]

            

    for pdf in target_folder.glob('*.pdf'):

        stats = pdf.stat()

        mtime = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y %H:%M')

        

        # Derive status per product

        status_val = app_meta.get('status', 'Pendiente')

        if approved_items:

            status_val = "Aprobado" if pdf.name in approved_items else "Pendiente"

        

        products.append({

            "name": pdf.name,

            "status": status_val,

            "fecha_aprobado": app_meta.get('approved_at', '-'),

            "aprobado_por": app_meta.get('approved_by_name', app_meta.get('approved_by', '-')),

            "path": f"{po_id}/{pdf.name}",

            "size": stats.st_size

        })

        

    return jsonify(products)



@app.route('/api/historial-product-csv-match', methods=['POST'])

def get_historial_product_csv_match():

    data = request.json

    po_id = data.get('po_id')

    product_code = data.get('product_code')

    

    if not po_id or not product_code:

        return jsonify({"status": "error", "message": "Missing params"}), 400

        

    target_folder = PROCESSED_PATH / po_id

    if not target_folder.exists():

         return jsonify({"status": "error", "message": "PO not found"}), 404

         

    # Scan all CSVs in folder recursively

    match_data = None

    all_csvs = list(target_folder.rglob('*.csv'))

    

    # 1. Try Filename Match (Prioritize this to match Client-Side behavior)

    p_name_norm = re.sub(r'\.[a-z0-9]+$', '', product_code, flags=re.IGNORECASE)

    p_name_norm = re.sub(r'[^a-z0-9]', '', p_name_norm).lower()

    

    for csv_file in all_csvs:

        if "aux_matches" in csv_file.name.lower(): continue

        if "resumen" in csv_file.name.lower(): continue

        

        c_name_norm = re.sub(r'\.[a-z0-9]+$', '', csv_file.name, flags=re.IGNORECASE)

        c_name_norm = re.sub(r'[^a-z0-9]', '', c_name_norm).lower()

        

        # Loose match logic

        if p_name_norm and (p_name_norm == c_name_norm or p_name_norm in c_name_norm or c_name_norm in p_name_norm):

            try:

                content = None

                for enc in ['utf-8-sig', 'cp1252', 'latin-1']:

                    try:

                        with open(csv_file, 'r', encoding=enc) as f:

                            content = f.read()

                            break

                    except: continue

                

                if content:

                    from io import StringIO

                    f = StringIO(content)

                    delimiter = ';' if content.count(';') > content.count(',') else ','

                    reader = csv.reader(f, delimiter=delimiter)

                    rows = list(reader)

                    if len(rows) > 1:

                         headers = rows[0]

                         found_row = None

                         

                         # Try to find row matching product_code

                         p_code_search = product_code.strip().lower() if product_code else ""

                         

                         for r in rows[1:]:

                             if not r or len(r) == 0: continue

                             cell_val = r[0].strip().lower()

                             if p_code_search and p_code_search in cell_val:

                                 found_row = r

                                 break

                         

                         # Fallback to first row if not found

                         final_row = found_row if found_row else rows[1]

                         

                         match_data = {

                            "filename": csv_file.name,

                            "headers": headers,

                            "row": final_row

                        }

                         break

            except: continue



    if not match_data:

        # 2. Fallback: Content Scan

        for csv_file in all_csvs:

            if "aux_matches" in csv_file.name.lower(): continue

            if "resumen" in csv_file.name.lower(): continue

            

            try:

                content = None

                for enc in ['utf-8-sig', 'cp1252', 'latin-1']:

                    try:

                        with open(csv_file, 'r', encoding=enc) as f:

                            content = f.read()

                            break

                    except: continue

                    

                if not content: continue

                

                from io import StringIO

                f = StringIO(content)

                delimiter = ';' if content.count(';') > content.count(',') else ','

                reader = csv.reader(f, delimiter=delimiter)

                rows = list(reader)

                

                if len(rows) < 2: continue

                

                headers = rows[0]

                for row in rows[1:]:

                    if not row: continue

                    cell_val = row[0].strip().lower()

                    p_code = product_code.strip().lower()

                    if p_code in cell_val:

                        match_data = {

                            "filename": csv_file.name,

                            "headers": headers,

                            "row": row

                        }

                        break

                if match_data: break

            except: continue

        

    if match_data:

        return jsonify({"status": "success", "data": match_data})

    else:

        return jsonify({"status": "not_found", "message": "No matched CSV data"}), 200





# --- AUXILIAR ROUTES ---



@app.route('/api/auxiliar-indices')

def get_auxiliar_indices():

    files_data = []

    if not AUXILIAR_DIR.exists():

        return jsonify([])

        

    for file in AUXILIAR_DIR.glob('*.csv'):

        if file.is_file():

            stats = file.stat()

            mtime = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y %H:%M')

            author = "Oficina Técnica" 

            files_data.append({

                "name": file.name,

                "date": mtime,

                "author": author

            })

    return jsonify(files_data)



@app.route('/api/historial-po')
def get_historial_po():
    try:
        print("[HISTORY DEBUG] Starting get_historial_po...")
        po_list = []
        
        if not PROCESSED_PATH.exists():
            print(f"[HISTORY DEBUG] PROCESSED_PATH does not exist: {PROCESSED_PATH}")
            return jsonify([])

        # Load users for fallback name resolution
        users_map = {}
        if USERS_FILE.exists():
            try:
                with open(USERS_FILE, 'r') as f:
                    users_map = json.load(f)
            except Exception as e:
                print(f"[HISTORY DEBUG] Failed to load users map: {e}")

        # Scan PROCESSED_PATH for directories (POs)
        print(f"[HISTORY DEBUG] Scanning {PROCESSED_PATH}")
        
        for folder in PROCESSED_PATH.iterdir():
            if folder.is_dir():
                try:
                    # Basic Info
                    po_id = folder.name
                    # print(f"[HISTORY DEBUG] Found PO folder: {po_id}")
                    
                    po_data = {
                        "id": po_id,
                        "proveedor": "-", # Will try to fetch
                        "fecha_po": "-",
                        "fecha_ingreso": "-",
                        "fecha_aprobado": "-",
                        "aprobado_por": "-", # Display Name
                        "estado": "Procesado",
                        "counts": {"approved": 0, "total": 0}
                    }
                    
                    # 1. Folder Date (Ingreso approximation)
                    try:
                        stats = folder.stat()
                        # Use st_mtime as it's more reliable on some Windows configs for 'last write'
                        po_data["fecha_ingreso"] = datetime.fromtimestamp(stats.st_ctime).strftime('%d/%m/%Y')
                    except: pass
                    
                    # 2. Read Approval Metadata
                    app_meta = folder / "approval_info.json"
                    if app_meta.exists():
                        try:
                            with open(app_meta, 'r', encoding='utf-8') as f:
                                meta = json.load(f)
                                po_data["fecha_aprobado"] = meta.get("approved_at", "-")
                                po_data["estado"] = meta.get("status", "Procesado")
                                
                                # Author Name Logic
                                saved_name = meta.get("approved_by_name")
                                if saved_name:
                                    po_data["aprobado_por"] = saved_name
                                else:
                                    # Fallback to ID and lookup
                                    uid = meta.get("approved_by", "-")
                                    if uid in users_map:
                                        po_data["aprobado_por"] = users_map[uid].get('display_name', uid)
                                    else:
                                        po_data["aprobado_por"] = uid
                                        
                                # Counts
                                if "counts" in meta:
                                    po_data["counts"] = meta["counts"]
                        except Exception as e_meta:
                            print(f"[HISTORY DEBUG] Error reading metadata for {po_id}: {e_meta}")
                    
                    # 3. Read Provider/Date from resumen.csv (if available)
                    # (Placeholder logic preserved)

                    # Fallback Counts if not in metadata
                    if po_data["counts"]["total"] == 0:
                        pdfs = list(folder.glob('*.pdf'))
                        count_pdfs = len(pdfs)
                        if count_pdfs > 0:
                            po_data["counts"]["total"] = count_pdfs
                            po_data["counts"]["approved"] = count_pdfs # Assume 100% if no metadata
                    
                    po_list.append(po_data)
                except Exception as e_inner:
                    print(f"[HISTORY DEBUG] Error processing folder {folder.name}: {e_inner}")
                    continue

        print(f"[HISTORY DEBUG] Returning {len(po_list)} records.")
        return jsonify(po_list)

    except Exception as e:
        print(f"[HISTORY CRITICAL ERROR]: {str(e)}")
        # Return empty list or error object to prevent frontend crash if possible, or 500
        return jsonify({"error": str(e), "status": "error"}), 500



@app.route('/api/auxiliar-csv/<filename>')

def get_auxiliar_csv(filename):

    filename = os.path.basename(filename)

    target = AUXILIAR_DIR / filename

    

    if not target.exists():

        return jsonify({"status": "error", "message": f"File not found: {filename}"}), 404

        

    try:

        rows = []

        headers = []

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        

        content = None

        used_encoding = 'utf-8'

        

        for enc in encodings:

            try:

                with open(target, 'r', encoding=enc) as f:

                    content = f.read()

                    used_encoding = enc

                    break

            except UnicodeDecodeError:

                continue

                

        if content is None:

             return jsonify({"status": "error", "message": "Could not determine file encoding"}), 500

             

        from io import StringIO

        f = StringIO(content)

        

        try:

            sample = content[:1024]

            sniffer = csv.Sniffer()

            dialect = sniffer.sniff(sample)

        except csv.Error:

            dialect = 'excel' 

            

        if ';' in content.split('\n')[0] or ';' in content.split('\n')[1] if len(content.split('\n')) > 1 else False:

             if content.count(';') > content.count(','):

                 reader = csv.reader(f, delimiter=';')

             else:

                 reader = csv.reader(f, dialect)

        else:

             reader = csv.reader(f, dialect)



        rows_list = list(reader)

        

        if rows_list:

            if len(rows_list) > 1:

                raw_headers = rows_list[1]

                data_rows = rows_list[2:]

                

                valid_indices = [i for i, h in enumerate(raw_headers) if h and h.strip()]

                

            if valid_indices:

                headers = [raw_headers[i].strip() for i in valid_indices]

                rows = []

                for r in data_rows:

                    if not any(cell.strip() for cell in r):

                        continue

                    filtered_row = [r[i] if i < len(r) else "" for i in valid_indices]

                    rows.append(filtered_row)

            else:

                headers = rows_list[0]

                rows = rows_list[1:]

        else:

            headers = rows_list[0]

            rows = [row for row in rows if any(cell.strip() for cell in row)]



        return jsonify({"headers": headers, "rows": rows})

    except Exception as e:

        print(f"Error parsing CSV: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



# --- PLANILLA ROUTES ---



PLA_ROOT = BASE_DIR / "Auxiliares"



def get_planilla_folder(context_filename):

    stem = Path(context_filename).stem

    code = stem.split('(')[0].strip() if '(' in stem else stem.split(' ')[0]

    

    family_folder = None

    for item in PLA_ROOT.iterdir():

        if item.is_dir() and code in item.name and "Registro Planillas" in item.name:

            family_folder = item

            break

            

    if not family_folder:

        for item in PLA_ROOT.iterdir():

            if item.is_dir() and code in item.name:

                family_folder = item

                break

    

    if not family_folder:

        return PLA_ROOT / stem



    target_specific = family_folder / stem

    if not target_specific.exists():

        try:

            target_specific.mkdir(parents=True, exist_ok=True)

        except: pass

        

    return target_specific



@app.route('/api/planilla/list/<filename>')

def list_planillas(filename):

    filename = secure_filename(filename) 

    filename = os.path.basename(filename)

    

    folder = get_planilla_folder(filename)

    

    files = []

    if folder.exists():

        for f in folder.iterdir():

            if f.is_file():

                files.append(f.name)

                

    return jsonify({"folder": folder.name, "files": files})



@app.route('/api/planilla/upload', methods=['POST'])

def upload_planilla():

    if 'file' not in request.files or 'context' not in request.form:

        return jsonify({"status": "error", "message": "Missing file or context"}), 400

        

    file = request.files['file']

    context = request.form['context'] 

    

    if file.filename == '':

        return jsonify({"status": "error", "message": "No selected file"}), 400

        

    folder = get_planilla_folder(os.path.basename(context))

    if not folder.exists():

        try:

            folder.mkdir(parents=True, exist_ok=True)

        except Exception as e:

             return jsonify({"status": "error", "message": f"Could not create directory: {str(e)}"}), 500

             

    filename = secure_filename(file.filename)

    try:

        file.save(folder / filename)

        return jsonify({"status": "success", "message": "Planilla uploaded successfully"})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/planilla/search-product', methods=['POST'])

def search_product_planilla():

    data = request.json

    context = data.get('context')

    query = data.get('query', '').lower()

    

    if not context or not query:

        return jsonify([])

        

    target = AUXILIAR_DIR / os.path.basename(context)

    if not target.exists():

        return jsonify([])

        

    results = []

    try:

        content = None

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        for enc in encodings:

            try:

                with open(target, 'r', encoding=enc) as f:

                    content = f.read()

                    break

            except: continue

            

        if content:

            from io import StringIO

            f = StringIO(content)

            delimiter = ';' if content.count(';') > content.count(',') else ','

            reader = csv.reader(f, delimiter=delimiter)

            rows = list(reader)

            

            if len(rows) > 1:

                headers = rows[1]

                data_rows = rows[2:]

                

                prod_idx = -1

                for i, h in enumerate(headers):

                    if 'PRODUCTO' in h.upper() or 'CODIGO' in h.upper() or 'ITEM' in h.upper():

                        prod_idx = i

                        break

                

                if prod_idx == -1: prod_idx = 1

                

                for row in data_rows:

                    if len(row) > prod_idx:

                        val = row[prod_idx].strip()

                        if query in val.lower():

                            if val not in results:

                                results.append(val)

                                if len(results) > 10: break

    except: pass

    

    return jsonify(results)



@app.route('/api/planilla/generate-pdf', methods=['POST'])

def generate_planilla_pdf():

    data = request.json

    product = data.get('product')

    # Validated, return URL to download

    # STUB RECOVERY

    return jsonify({"status": "error", "message": "Funcionalidad de PDF en mantenimiento"})



# --- ADVANCED AUXILIAR LOOKUP (PO LINKED) ---



@app.route('/api/product-aux-details', methods=['POST'])

def get_product_aux_details():

    try:

        data = request.json

        po_id = data.get('po_id')

        product_code_a = data.get('product_code_a')

        

        try:

            with open(os.path.join(SCRIPT_DIR, 'debug_backend.txt'), 'a') as df:

                df.write(f"\\n[{datetime.now()}] REQ: po={po_id}, code={product_code_a}\\n")

        except: pass



        print(f"DEBUG_AUX: START get_product_aux_details. po_id={po_id}, product={product_code_a}")



        if not po_id or not product_code_a:

            return jsonify({"status": "error", "message": "Missing params"}), 400



        po_path = DATA_DIR / po_id

        if not po_path.exists():

             # Check PROCESSED

             po_path = PROCESSED_PATH / po_id

             if not po_path.exists():

                return jsonify({"status": "error", "message": "PO not found"}), 404



        # 1. Find aux_matches.csv

        # 1. Find aux_matches.csv (More robust search)

        aux_matches_path = None

        for f in po_path.rglob('aux_matches.csv'):

            aux_matches_path = f

            break

            

        if not aux_matches_path or not aux_matches_path.exists():

             return jsonify({"status": "error", "message": "aux_matches.csv not found in PO"}), 404



        target_c = None

        target_d = None

        

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        content = None

        for enc in encodings:

            try:

                with open(aux_matches_path, 'r', encoding=enc) as f:

                    content = f.read()

                    break

            except: continue

            

        if not content: return jsonify({"status": "error", "message": "Could not read aux_matches"}), 500

        

        from io import StringIO

        f = StringIO(content)

        delimiter = ';' if content.count(';') > content.count(',') else ','

        reader = csv.reader(f, delimiter=delimiter)

        

        product_code_clean = product_code_a.strip()

        

        for i, row in enumerate(reader):

            if len(row) < 4: continue

            if row[0].strip() == product_code_clean:

                target_c = row[2].strip() # Filename base

                target_d = row[3].strip() # Search Key

                break

                

        if not target_c:

             # Fallback: Brute force search in AUXILIAR_DIR

             print(f"Fallback Search for {product_code_clean} in {AUXILIAR_DIR}")

             try:

                with open(os.path.join(SCRIPT_DIR, 'debug_backend.txt'), 'a') as df:

                    df.write(f"Fallback Search Start for {product_code_clean} in {AUXILIAR_DIR}\\n")

             except: pass



             from difflib import SequenceMatcher

             

             candidates = [] # List of dicts: {score, headers, row, filename}

             seen_rows = set() # To avoid duplicates if multiple matches in same file/row



             if AUXILIAR_DIR.exists():

                 for item in AUXILIAR_DIR.iterdir():

                     if item.is_file() and item.suffix.lower() == '.csv':

                         try:

                             content_bf = None

                             for enc in ['utf-8-sig', 'cp1252', 'latin-1']:

                                 try:

                                     with open(item, 'r', encoding=enc) as f:

                                         content_bf = f.read()

                                         break

                                 except: continue

                             if not content_bf: continue



                             f_bf = StringIO(content_bf)

                             delim_bf = ';' if content_bf.count(';') > content_bf.count(',') else ','

                             reader_bf = csv.reader(f_bf, delimiter=delim_bf)

                             rows_bf = list(reader_bf)

                             

                             if len(rows_bf) > 1:

                                 # Iterate all rows.

                                 for r_idx, row in enumerate(rows_bf):

                                     if len(row) > 1:

                                         cell_val = row[1].strip()

                                         if not cell_val: continue

                                         

                                         # Calculate Score

                                         score = 0

                                         if cell_val == product_code_clean:

                                             score = 100

                                         else:

                                             # General Similarity Check

                                             # Optimization: Check basic length or prefix to speed up?

                                             # Let's just calculate ratio for all, but maybe skip if length diff is huge

                                             if abs(len(cell_val) - len(product_code_clean)) < 20: 

                                                 sm_ratio = SequenceMatcher(None, product_code_clean, cell_val).ratio()

                                                 score = sm_ratio * 100

                                                 

                                                 # CRITICAL PENALTY 1: Product Number Mismatch

                                                 # We look for the longest digit sequence (e.g. 51310) to avoid comparing partials like '040'

                                                 import re

                                                 def get_main_number(s):

                                                     nums = re.findall(r'\d+', s)

                                                     if not nums: return None

                                                     return max(nums, key=len)



                                                 n1 = get_main_number(product_code_clean)

                                                 n2 = get_main_number(cell_val)

                                                 

                                                 if n1 and n2 and n1 != n2:

                                                     score -= 50 # Critical Error (User request)

                                                 

                                                 # PENALTY 2: Substring Check (User: "Si no esta exactamente... baja el puntaje")

                                                 # We check if one is contained in the other (normalized)

                                                 s_clean = product_code_clean.strip().lower()

                                                 c_clean = cell_val.strip().lower()

                                                 if c_clean not in s_clean and s_clean not in c_clean:

                                                     score -= 15

                                                 

                                                 if score < 0: score = 0

                                         

                                         if score >= 50:

                                             # Unique Key: Filename + Row Index

                                             unique_key = f"{item.name}_{r_idx}"

                                             if unique_key in seen_rows: continue

                                             seen_rows.add(unique_key)



                                             raw_h = rows_bf[1] if len(rows_bf) > 1 else []

                                             final_h = raw_h[1:] if len(raw_h) > 1 else raw_h

                                             row_d = row[1:]

                                             

                                             candidates.append({

                                                 "score": score,

                                                 "headers": final_h,

                                                 "row": row_d,

                                                 "filename": item.name,

                                                 "match_val": cell_val

                                             })

                                                 

                         except Exception as e:

                             print(f"Error scanning {item.name}: {e}")

                             continue

             

             # Sort by score desc

             candidates.sort(key=lambda x: x['score'], reverse=True)

             

             # Take top 3

             top_matches = candidates[:3]

             

             if top_matches:

                 print(f"Returning top {len(top_matches)} matches")

                 return jsonify({

                     "status": "success",

                     "matches": top_matches # List of match objects

                 })



             return jsonify({"status": "not_found", "message": "No match found"}), 200



        # 2. Find Aux File (Smart Match)

        import re

        target_c_base = os.path.splitext(target_c)[0]

        

        found_path = None

        

        # A. Try exact construction first

        try_name_1 = f"{target_c_base}_Auxiliar.csv"

        path_1 = AUXILIAR_DIR / try_name_1

        if path_1.exists():

            found_path = path_1

            aux_filename = try_name_1

            

        # B. Smart Match: Use text inside parentheses

        if not found_path and AUXILIAR_DIR.exists():

            paren_match = re.search(r'\((.*?)\)', target_c_base)

            if paren_match:

                key_text = paren_match.group(1).lower().strip() # e.g. "inspeccion de especiales"

                if len(key_text) > 3: # Safety check

                    for item in AUXILIAR_DIR.iterdir():

                        if item.is_file() and item.suffix.lower() == '.csv':

                            if key_text in item.name.lower():

                                found_path = item

                                aux_filename = item.name

                                break

        

        # C. Fallback: Loose partial match of the base name

        if not found_path and AUXILIAR_DIR.exists():

             search_base = target_c_base.lower().replace('_auxiliar','').strip()

             for item in AUXILIAR_DIR.iterdir():

                 if item.is_file() and item.suffix.lower() == '.csv':

                     if search_base in item.name.lower():

                         found_path = item

                         aux_filename = item.name

                         break

                         

        if found_path:

             aux_file_path = found_path

        else:

             return jsonify({"status": "error", "message": f"Aux file match not found for: {target_c}"}), 404

             

        # 3. Read Aux File and find Target D in Col B

        row_data = None

        final_headers = None

        

        content_aux = None

        for enc in encodings:

            try:

                with open(aux_file_path, 'r', encoding=enc) as f:

                    content_aux = f.read()

                    break

            except: continue

            

        if not content_aux: return jsonify({"status": "error", "message": "Could not read aux file"}), 500

        

        f_aux = StringIO(content_aux)

        delimiter_aux = ';' if content_aux.count(';') > content_aux.count(',') else ','

        reader_aux = csv.reader(f_aux, delimiter=delimiter_aux)

        rows_aux = list(reader_aux)

        

        if len(rows_aux) > 1:

            # User Request: Use Row 2 (index 1) as header

            # User Request: Skip Col 1 (index 0)

            

            raw_headers = rows_aux[1]

            # Slicing: Skip first column

            final_headers = raw_headers[1:]

            

            for row in rows_aux:

                if len(row) < 2: continue

                # Match Target D in Col B (index 1) - Search remains in original structure

                if row[1].strip() == target_d:

                    # Return data skipping first column

                    row_data = row[1:]

                    break

                    

        if row_data:

            return jsonify({

                "status": "success",

                "headers": final_headers,

                "row": row_data,

                "filename": aux_filename

            })

        else:

            return jsonify({"status": "not_found", "message": "No match in aux file"}), 200



    except Exception as e:

        print(f"Aux Lookup Error: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



# --- API FOR SAVING AUX TARGET ---

@app.route('/api/save-aux-target', methods=['POST'])

def save_aux_target():

    try:

        data = request.json

        po_id = data.get('po_id')

        product_code = data.get('product_code')

        target = data.get('target')



        if not po_id or not product_code:

            return jsonify({'status': 'error', 'message': 'Missing fields'}), 400



        # Create/Update Metadata in PO Folder

        po_folder = DATA_DIR / po_id

        if not po_folder.exists():

            # Try PROCESSED? No, usually in process

            po_folder = PROCESSED_PATH / po_id

            if not po_folder.exists():

                return jsonify({'status': 'error', 'message': 'PO folder not found'}), 404



        target_file = po_folder / "aux_targets.json"

        

        current_data = {}

        if target_file.exists():

            try:

                with open(target_file, 'r', encoding='utf-8') as f:

                    current_data = json.load(f)

            except: pass



        current_data[product_code] = target

        

        with open(target_file, 'w', encoding='utf-8') as f:

            json.dump(current_data, f, indent=4)

            

        return jsonify({'status': 'success'})

    except Exception as e:

        print(f"Save Target Error: {e}")

        return jsonify({'status': 'error', 'message': str(e)}), 500





# --- EXCEL UPDATE LOGIC ---

def process_excel_updates(po_folder, approved_items, po_id):

    """

    Updates the auxiliary Excel files (COPIES) based on approved items and saved targets.

    """

    from openpyxl.styles import Font, Alignment



    targets_file = po_folder / "aux_targets.json"

    matches_file = po_folder / "csv_Auxiliar/aux_matches.csv"

    

    if not targets_file.exists(): 

        print(f"[EXCEL] No aux_targets.json for {po_id}, skipping excel updates.")

        return



    try:

        with open(targets_file, 'r', encoding='utf-8') as f:

            targets = json.load(f)

    except Exception as e:

        print(f"[EXCEL] Error reading targets: {e}")

        return



    # Load Matches (to resolve Type/Filename for New Lines)

    matches_map = {} # Code -> {Type, File...}

    if matches_file.exists():

        try:

            with open(matches_file, 'r', encoding='utf-8') as f:

                reader = csv.DictReader(f, delimiter=';')

                for row in reader:

                    matches_map[row['Codigo']] = row

        except Exception as e:

             print(f"[EXCEL] Error reading matches: {e}")



    # AUX_COPIES_DIR = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\Registro Planillas R016-01")
    AUX_COPIES_DIR = BASE_DIR / "Auxiliares/Registro Planillas R016-01"

    

    for prod_code in approved_items:

        # Resolve CSV Path & Real Code first

        clean_name = prod_code

        if clean_name.lower().endswith('.pdf'):

            clean_name = clean_name[:-4]

            

        csv_path = po_folder / f"csv/{clean_name}.csv"

        if not csv_path.exists():

             csv_path = po_folder / f"csv/{prod_code}.csv"

             

        if not csv_path.exists():

            print(f"[EXCEL] CSV not found for {prod_code}")

            continue



        row_data = None

        real_code = clean_name

        try:

            with open(csv_path, 'r', encoding='utf-8') as f:

                reader = csv.reader(f, delimiter=';')

                rows = list(reader)

                if len(rows) > 1:

                    row_data = rows[1]

                    if row_data: real_code = row_data[0]

        except Exception as e:

            print(f"[EXCEL] Error reading CSV {csv_path}: {e}")

            continue

            

        if not row_data: continue



        # Resolve Clean Code for Mapping (Fuzzy Match against Matches Map)

        # CSV might have "B&P-040-51310..." but Map has "51310"

        if real_code not in matches_map:

            sorted_keys = sorted(matches_map.keys(), key=len, reverse=True)

            for k in sorted_keys:

                if k in real_code:

                    print(f"[EXCEL] Fuzzy matched CSV code '{real_code}' to Map key '{k}'")

                    real_code = k

                    break



        # Resolve Target Action

        action_val = targets.get(prod_code)

        

        # --- FALLBACK ---

        if not action_val:

            match_info = matches_map.get(real_code) # Key by Code

            score = 0

            try: 

                score = float(match_info.get('Score', 0))

            except: pass

            

            if match_info and score >= 99.0:

                fname = match_info.get('Archivo_Auxiliar', '')

                if fname:

                    action_val = f"{fname}_0"

                    print(f"[EXCEL] Auto-default to Existing for {real_code} (Score {score})")

            

            if not action_val:

                action_val = "new"

                print(f"[EXCEL] Auto-default to New Line for {real_code}")

        

        # Determine Target File

        target_filename = None

        

        match_info = matches_map.get(real_code)

        if hasattr(match_info, 'get'):

             target_filename = match_info.get('Archivo_Auxiliar')



        if action_val == "new":

            # Direct use of identified file

            pass 

        else:

            # Existing: Extract filename from action ID if possible, mostly same as match

            try:

                parts = action_val.rsplit('_', 1)

                filename_from_val = parts[0]

                if filename_from_val:

                    target_filename = filename_from_val

            except: pass



        if not target_filename:

            print(f"[EXCEL] No target file (Archivo_Auxiliar) found for '{real_code}'. Skipping.")

            continue



        # Clean filename (sometimes it might have paths? verify step3 output)

        if "\\" in target_filename or "/" in target_filename:

             target_filename = Path(target_filename).name



        # Resolve Real Excel File from CSV name (e.g. remove _Auxiliar.csv)

        if "_Auxiliar.csv" in target_filename:

            target_filename = target_filename.replace("_Auxiliar.csv", "")

        elif target_filename.lower().endswith(".csv"):

             target_filename = target_filename[:-4]

             

        # Search for .xlsx or .xlsm match

        found_excel = None

        for ext in ["", ".xlsx", ".xlsm", ".xls"]:

            p = AUX_COPIES_DIR / (target_filename + ext)

            if p.exists():

                found_excel = p

                break

        

        if not found_excel:

             print(f"[EXCEL] Excel file not found for stem '{target_filename}'. Skipping.")

             continue

             

        excel_path = found_excel

            

        try:

            is_xlsm = excel_path.suffix.lower() == '.xlsm'

            

            # Load via binary stream to handle UNC/Network paths better

            wb = None

            try:

                with open(excel_path, 'rb') as f_stream:

                    wb = openpyxl.load_workbook(f_stream, keep_vba=is_xlsm)

            except Exception as e_load:

                # If failed (likely password protected), try with 'bpb'

                print(f"[EXCEL] Standard load failed ({e_load}). Trying with password 'bpb'...")

                try:

                    import msoffcrypto

                    import io

                    

                    decrypted = io.BytesIO()

                    with open(excel_path, 'rb') as f_enc:

                        office_file = msoffcrypto.OfficeFile(f_enc)

                        office_file.load_key(password='bpb')

                        office_file.decrypt(decrypted)

                    

                    wb = openpyxl.load_workbook(decrypted, keep_vba=is_xlsm)

                    print(f"[EXCEL] Decryption successful.")

                except ImportError:

                    print("[EXCEL] CRITICAL: File is password protected. Please run in cmd: pip install msoffcrypto-tool")

                    raise e_load

                except Exception as e_dec:

                    print(f"[EXCEL] Decryption failed: {e_dec}")

                    raise e_load

            

            # Select "Auxiliar" Sheet

            ws = None

            if "Auxiliar" in wb.sheetnames:

                ws = wb["Auxiliar"]

            elif "Aux" in wb.sheetnames:

                ws = wb["Aux"]

            else:

                # Fallback to active but warn

                print(f"[EXCEL] Sheet 'Auxiliar' not found in {target_filename}, using active.")

                ws = wb.active



            START_ROW = 3 # User requested to skip rows 1 & 2

            

            found_row = None

            max_num = 0

            

            # Scan for Existing & Max Number

            # Iterate from START_ROW

            for r in range(START_ROW, ws.max_row + 1):

                # Col 1: Numero

                num_cell = ws.cell(row=r, column=1).value

                if isinstance(num_cell, (int, float)):

                    if num_cell > max_num: max_num = int(num_cell)

                

                if action_val != "new":

                    # Check Code in Col 2 (B)

                    code_cell = ws.cell(row=r, column=2).value

                    if code_cell and str(code_cell).strip() == prod_code:

                        found_row = r

                        # Don't break immediately if we need max_num? 

                        # Only need max_num for NEW lines. For existing, we replace.

                        break

            

            # Logic Application

            if action_val == "new":

                # Determine Next Num

                next_num = max_num + 1

                

                # Append Row

                # [Number] + [Code, Data...]

                final_row_values = [next_num] + row_data

                ws.append(final_row_values)

                

                # Apply Styles to the appended row (last row) (and fix data type if needed)

                last_row = ws.max_row

                for c_idx, val in enumerate(final_row_values, start=1):

                    cell = ws.cell(row=last_row, column=c_idx)

                    # Col 1 is Num, Col 2 is Code

                    if c_idx == 2:

                        cell.font = Font(bold=True)

                    elif c_idx > 2:

                        cell.alignment = Alignment(horizontal='center')

                        

                print(f"[EXCEL] Appended {real_code} to {target_filename} (No. {next_num})")

                

            elif found_row:

                # Update Existing Row

                # Skip Col 1 (Numero), Update from Col 2 match row_data

                for c_idx, val in enumerate(row_data, start=2): # Start at Col 2

                     cell = ws.cell(row=found_row, column=c_idx)

                     cell.value = val

                     # Apply Styles

                     if c_idx == 2:

                         cell.font = Font(bold=True)

                     else:

                         cell.alignment = Alignment(horizontal='center')

                         

                print(f"[EXCEL] Updated {real_code} in {target_filename} at row {found_row}")

                

            else:

                 # Existing requested but not found -> Append?

                 print(f"[EXCEL] Code {real_code} not found for update, appending new.")

                 next_num = max_num + 1

                 final_row_values = [next_num] + row_data

                 ws.append(final_row_values)

                 # Apply Styles fallback

                 last_row = ws.max_row

                 for c_idx, val in enumerate(final_row_values, start=1):

                    cell = ws.cell(row=last_row, column=c_idx)

                    if c_idx == 2:

                        cell.font = Font(bold=True)

                    elif c_idx > 2:

                        cell.alignment = Alignment(horizontal='center')

                 

                 print(f"[EXCEL] Appended (fallback) {real_code} to {target_filename}")



            save_path_str = str(excel_path)

            print(f"[EXCEL] Saving workbook to: {save_path_str}")

            wb.save(save_path_str)

            print(f"[EXCEL] Save operation completed.")

            

        except Exception as e:

            print(f"[EXCEL] Failed to process {target_filename}: {e}")





# --- EXCEL UPDATE LOGIC V2 (SEARCH FIX + STYLES + CRYPTO) ---

def process_excel_updates_v2(po_folder, approved_items, po_id):
    """
    Updates the auxiliary Excel files (COPIES) based on approved items and saved targets.
    V2 Optimized: Batches updates per file to avoid opening/saving multiple times.
    """
    from openpyxl.styles import Font, Alignment
    import json
    import csv
    import openpyxl
    from pathlib import Path
    import msoffcrypto
    import io

    targets_file = po_folder / "aux_targets.json"
    matches_file = po_folder / "csv_Auxiliar/aux_matches.csv"
    
    if not targets_file.exists(): 
        print(f"[EXCEL] No aux_targets.json for {po_id}, skipping excel updates.", flush=True)
        return

    try:
        with open(targets_file, 'r', encoding='utf-8') as f:
            targets = json.load(f)
    except Exception as e:
        print(f"[EXCEL] Error reading targets: {e}")
        return

    # Load Matches
    matches_map = {} 
    if matches_file.exists():
        try:
            with open(matches_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    matches_map[row['Codigo']] = row
        except Exception as e:
             print(f"[EXCEL] Error reading matches: {e}")

    # AUX_COPIES_DIR = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\Registro Planillas R016-01")
    AUX_COPIES_DIR = BASE_DIR / "Auxiliares/Registro Planillas R016-01"
    
    # --- Phase 1: Group Updates by Target File ---
    updates_by_file = {} # { "filename": [ {item_data}, ... ] }

    for prod_code in approved_items:
        # Resolve CSV Path & Real Code
        clean_name = prod_code
        if clean_name.lower().endswith('.pdf'):
            clean_name = clean_name[:-4]
            
        csv_path = po_folder / f"csv/{clean_name}.csv"
        if not csv_path.exists():
             csv_path = po_folder / f"csv/{prod_code}.csv"
             
        if not csv_path.exists():
            print(f"[EXCEL] CSV not found for {prod_code}")
            continue

        row_data = None
        real_code = clean_name
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=';')
                rows = list(reader)
                if len(rows) > 1:
                    row_data = rows[1]
                    if row_data: real_code = row_data[0]
        except Exception as e:
            print(f"[EXCEL] Error reading CSV {csv_path}: {e}")
            continue
            
        if not row_data: continue

        # Capture raw code
        csv_raw_code = real_code

        # Fuzzy Match Resolution
        if real_code not in matches_map:
            sorted_keys = sorted(matches_map.keys(), key=len, reverse=True)
            for k in sorted_keys:
                if k in real_code:
                    # print(f"[EXCEL] Fuzzy matched CSV code '{real_code}' to Map key '{k}'")
                    real_code = k
                    break

        # Resolve Target Action
        action_val = targets.get(prod_code)
        if not action_val and csv_raw_code:
             action_val = targets.get(csv_raw_code)
        if not action_val:
             action_val = targets.get(clean_name)
        
        # Fallback Logic
        if not action_val:
            match_info = matches_map.get(real_code)
            score = 0
            try: score = float(match_info.get('Score', 0) if match_info else 0)
            except: pass
            
            if match_info and score >= 99.0:
                fname = match_info.get('Archivo_Auxiliar', '')
                if fname:
                    action_val = f"{fname}_0"
            
            if not action_val:
                action_val = "new"

        # Determine Target Filename & Forced Row
        target_filename = None
        forced_row = None
        
        match_info = matches_map.get(real_code)
        if hasattr(match_info, 'get'):
             target_filename = match_info.get('Archivo_Auxiliar')

        if action_val != "new":
            try:
                parts = action_val.rsplit('_', 1)
                filename_from_val = parts[0]
                if filename_from_val:
                    target_filename = filename_from_val
                
                if len(parts) > 1 and parts[1].isdigit():
                    parsed_row = int(parts[1])
                    if parsed_row > 0:
                        forced_row = parsed_row
            except: pass

        if not target_filename:
            print(f"[EXCEL] No target file found for '{real_code}'. Skipping.")
            continue

        # Clean filename
        if "\\" in target_filename or "/" in target_filename:
             target_filename = Path(target_filename).name
        if "_Auxiliar.csv" in target_filename:
            target_filename = target_filename.replace("_Auxiliar.csv", "")
        elif target_filename.lower().endswith(".csv"):
             target_filename = target_filename[:-4]

        # Add to Group
        if target_filename not in updates_by_file:
            updates_by_file[target_filename] = []
        
        updates_by_file[target_filename].append({
            "real_code": real_code,
            "row_data": row_data,
            "action_val": action_val,
            "forced_row": forced_row,
            "csv_raw_code": csv_raw_code
        })

    # --- Phase 2: Execute Updates per File ---
    print(f"[EXCEL] Starting Batch Update for {len(updates_by_file)} Unique Files.", flush=True)
    
    for target_filename, items in updates_by_file.items():
        # Search for .xlsx or .xlsm
        found_excel = None
        for ext in ["", ".xlsx", ".xlsm", ".xls"]:
            p = AUX_COPIES_DIR / (target_filename + ext)
            if p.exists():
                found_excel = p
                break
        
        if not found_excel:
             print(f"[EXCEL] File {target_filename} not found. Skipping {len(items)} items.")
             continue
             
        excel_path = found_excel
        print(f"[EXCEL] Processing {target_filename} ({len(items)} updates)...")

        try:
            is_xlsm = excel_path.suffix.lower() == '.xlsm'
            wb = None
            
            # Load Workbook (Once)
            try:
                with open(excel_path, 'rb') as f_stream:
                    wb = openpyxl.load_workbook(f_stream, keep_vba=is_xlsm)
            except Exception as e_load:
                # Decryption Retry
                print(f"[EXCEL] Load failed ({e_load}). Trying decryption...")
                try:
                    decrypted = io.BytesIO()
                    with open(excel_path, 'rb') as f_enc:
                        office_file = msoffcrypto.OfficeFile(f_enc)
                        office_file.load_key(password='bpb')
                        office_file.decrypt(decrypted)
                    wb = openpyxl.load_workbook(decrypted, keep_vba=is_xlsm)
                except Exception as e_dec:
                    print(f"[EXCEL] Decryption failed: {e_dec}")
                    raise e_load

            # Identify Sheet
            ws = None
            if "Auxiliar" in wb.sheetnames: ws = wb["Auxiliar"]
            elif "Aux" in wb.sheetnames: ws = wb["Aux"]
            else: ws = wb.active

            START_ROW = 3
            
            # Iterate Items for THIS file
            for item in items:
                real_code = item["real_code"]
                row_data = item["row_data"]
                action = item["action_val"]
                forced_row = item["forced_row"]
                csv_raw_code = item["csv_raw_code"]

                # Determine Last Row (Dynamic, changes as we append)
                last_real_row = START_ROW
                for r in range(ws.max_row, START_ROW - 1, -1):
                    val = ws.cell(row=r, column=2).value
                    if val and str(val).strip():
                        last_real_row = r
                        break
                
                # Determine Next Num
                last_num = 0
                try:
                    val_num = ws.cell(row=last_real_row, column=1).value
                    if isinstance(val_num, (int, float)):
                        last_num = int(val_num)
                except: pass
                next_num = last_num + 1

                # Find Row for Update
                found_row_idx = forced_row
                if not found_row_idx and action != "new":
                    for r in range(START_ROW, last_real_row + 1):
                        cell_val = str(ws.cell(row=r, column=2).value or "").strip()
                        search_term = str(csv_raw_code or real_code).strip()
                        if cell_val == search_term:
                            found_row_idx = r
                            break
                
                # Apply Write
                target_row = None
                if action == "new":
                    target_row = last_real_row + 1
                    ws.cell(row=target_row, column=1, value=next_num).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Appending {real_code} at Row {target_row}")
                elif found_row_idx:
                    target_row = found_row_idx
                    ws.cell(row=target_row, column=1).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Updating {real_code} at Row {target_row}")
                else:
                    # Fallback Append
                    target_row = last_real_row + 1
                    ws.cell(row=target_row, column=1, value=next_num).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Appending (Fallback) {real_code} at Row {target_row}")

                # Write Data Cols
                for c_idx, val in enumerate(row_data, start=2):
                    cell = ws.cell(row=target_row, column=c_idx)
                    cell.value = val
                    if c_idx == 2: cell.font = Font(bold=True)
                    else: cell.alignment = Alignment(horizontal='center')

            # Save Workbook (Once)
            wb.save(str(excel_path))
            print(f"[EXCEL] Saved {target_filename} successfully.")

            # Update Index CSV (Once)
            try:
                # INDICES_DIR = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\indices_auxiliar")
                INDICES_DIR = BASE_DIR / "Auxiliares/indices_auxiliar"
                if not INDICES_DIR.exists(): INDICES_DIR.mkdir(parents=True, exist_ok=True)
                
                csv_out = INDICES_DIR / f"{excel_path.stem}_Auxiliar.csv"
                with open(csv_out, 'w', newline='', encoding='utf-8') as f_csv:
                    writer = csv.writer(f_csv, delimiter=';')
                    for row in ws.iter_rows(values_only=True):
                        clean_row = [str(c) if c is not None else "" for c in row]
                        writer.writerow(clean_row)
            except Exception as e_csv:
                print(f"[EXCEL] Warning: CSV Index update failed: {e_csv}")

        except Exception as e:
            print(f"[EXCEL] Failed to process {target_filename}: {e}")
            if "Permission denied" in str(e):
                 raise Exception(f"El Registro {target_filename} se encuentra abierto.")
            raise Exception(f"Error en archivo '{target_filename}': {e}")



@app.route('/api/submit-activity', methods=['POST'])

def submit_activity():

    try:

        data = request.json

        date_str = data.get('date') # Format DD/MM/YYYY

        activities = data.get('activities')

        

        if not date_str or not activities:

            return jsonify({'status': 'error', 'content': 'Datos incompletos'})

            

        # Parse date to safe filename

        safe_date = date_str.replace('/', '-')

        

        # Base Path for Activities

        base_dir = os.path.dirname(os.path.abspath(__file__))

        activities_dir = os.path.join(base_dir, 'Datos', 'Actividades')

        if not os.path.exists(activities_dir):

            os.makedirs(activities_dir)

            

        file_path = os.path.join(activities_dir, f'registro_{safe_date}.json')

        

        record = {

            'date': date_str,

            'timestamp': datetime.now().isoformat(),

            'activities': activities,

            'user': session.get('user'),

            'status': 'completed'

        }

        

        with open(file_path, 'w', encoding='utf-8') as f:

            json.dump(record, f, indent=4, ensure_ascii=False)

            

        return jsonify({'status': 'success'})

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/approve-activity', methods=['POST'])
def approve_activity():
    try:
        data = request.json
        date_str = data.get('date')
        if not date_str:
            return jsonify({'status': 'error', 'content': 'Fecha requerida'})

        safe_date = date_str.replace('/', '-')
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        if not os.path.exists(file_path):
            return jsonify({'status': 'error', 'content': 'Registro no encontrado'})

        with open(file_path, 'r', encoding='utf-8') as f:
            record = json.load(f)

        if record.get('status') == 'approved':
            return jsonify({'status': 'error', 'content': 'El registro ya fue aprobado.'})

        # --- Guardado en CSV + estado mailer ---
        try:
            mailer_state_path = Path(r"\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Actividad\Codigos\data\activity_mailer_state.json")
            if not mailer_state_path.exists():
                return jsonify({'status': 'error', 'content': 'Estado de correos no encontrado.'})

            with open(mailer_state_path, 'r', encoding='utf-8') as f:
                mailer_data = json.load(f)

            try:
                dt_obj = datetime.strptime(date_str, '%d/%m/%Y')
                iso_date = dt_obj.strftime('%Y-%m-%d')
            except ValueError:
                return jsonify({'status': 'error', 'content': 'Formato de fecha invalido.'})

            token = None
            target_campaign = None
            for camp in mailer_data.get('campaigns', []):
                if camp.get('date') == iso_date:
                    token = camp.get('token')
                    target_campaign = camp
                    break

            if not token or not target_campaign:
                return jsonify({'status': 'error', 'content': 'No se encontro la solicitud (token) para esa fecha.'})

            user_email = record.get('user') or session.get('user')

            if user_email:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    users_db = json.load(f)

                user_data = users_db.get(user_email)
                if user_data:
                    full_name = user_data.get('display_name', user_email)

                    def clean_csv_val(val):
                        return str(val).replace(';', ',').replace('\n', ' ').replace('\r', ' ').strip()

                    legacy_root = Path(r"\\BPBSRV03\lcontigiani\Oficina Tecnica\Registro de Actividad\Codigos\data")
                    global_csv = legacy_root / "base_datos_respuestas.csv"
                    individual_csv = legacy_root / "respuestas_csv" / f"{full_name}.csv"

                    now = datetime.now()
                    time_str = now.strftime("%H:%M:%S")

                    activities = record.get('activities', [])

                    body_blocks = []
                    for act in activities:
                        proj = clean_csv_val(act.get('project', ''))
                        time_val = str(act.get('time', '0')).replace('.', ',')
                        desc = clean_csv_val(act.get('description', ''))
                        body_blocks.append(f"Proyecto: {proj}\\nTiempo destinado: {time_val} horas\\nDescripcion: {desc}")
                    body_text = "\\n\\n".join(body_blocks) if body_blocks else ""

                    responses = target_campaign.setdefault('responses', {})
                    responses[user_email] = {
                        "from": user_email,
                        "from_name": full_name,
                        "subject": f"Re: {target_campaign.get('subject')}",
                        "date": now.strftime("%a, %d %b %Y %H:%M:%S"),
                        "received_at": now.isoformat(),
                        "snippet": body_text,
                        "body": body_text,
                        "uid": 0,
                        "saved_path": "WEB_SUBMISSION"
                    }

                    with open(global_csv, 'a', newline='', encoding='utf-8') as lg:
                        writer = csv.writer(lg, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                        for act in activities:
                            proj = clean_csv_val(act.get('project', ''))
                            time_val = str(act.get('time', '0')).replace('.', ',')
                            desc = clean_csv_val(act.get('description', ''))
                            writer.writerow([date_str, time_str, token, full_name, user_email, proj, proj, time_val, desc])

                    individual_csv.parent.mkdir(parents=True, exist_ok=True)
                    with open(individual_csv, 'a', newline='', encoding='utf-8') as li:
                        writer = csv.writer(li, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                        for act in activities:
                            proj = clean_csv_val(act.get('project', ''))
                            time_val = str(act.get('time', '0')).replace('.', ',')
                            desc = clean_csv_val(act.get('description', ''))
                            writer.writerow([date_str, time_str, token, proj, time_val, proj, desc])

                    with open(mailer_state_path, 'w', encoding='utf-8') as f:
                        json.dump(mailer_data, f, ensure_ascii=False, indent=2)

        except Exception as e:
            print(f"Error saving legacy CSVs: {e}")
            raise e

        record['status'] = 'approved'
        record['approved_at'] = datetime.now().isoformat()

        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(record, f, indent=4, ensure_ascii=False)

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'status': 'error', 'content': str(e)})

@app.route('/api/get-activity-status', methods=['GET'])

def get_activity_status():

    try:

        date_str = request.args.get('date')

        if not date_str:

            return jsonify({'status': 'error'})

            

        safe_date = date_str.replace('/', '-')

        base_dir = os.path.dirname(os.path.abspath(__file__))

        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        

        if os.path.exists(file_path):

             with open(file_path, 'r', encoding='utf-8') as f:

                 data = json.load(f)

             is_approved = data.get('status') == 'approved'

             return jsonify({'status': 'success', 'completed': True, 'approved': is_approved})

        else:

             return jsonify({'status': 'success', 'completed': False})

             

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/get-activity', methods=['GET'])

def get_activity():

    try:

        date_str = request.args.get('date')

        if not date_str:

            return jsonify({'status': 'error', 'content': 'Fecha requerida'})

            

        safe_date = date_str.replace('/', '-')

        base_dir = os.path.dirname(os.path.abspath(__file__))

        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        

        if os.path.exists(file_path):

            with open(file_path, 'r', encoding='utf-8') as f:

                data = json.load(f)

                return jsonify({'status': 'success', 'activities': data.get('activities', [])})

        else:

             return jsonify({'status': 'not_found'})

             

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/glossary', methods=['GET'])

def get_glossary():

    # Dynamic path resolution for portability
    # BASE_DIR points to '.../Registro de Control de Producto'
    glossary_path = BASE_DIR.parent / "Registro de Actividad" / "Codigos" / "config" / "glosario_proyectos.log"

    try:

        if os.path.exists(glossary_path):

            with open(glossary_path, 'r', encoding='utf-8', errors='ignore') as f:

                content = f.read()

            return jsonify({'status': 'success', 'content': content})

        else:

            return jsonify({'status': 'error', 'content': 'El archivo de glosario (glosario_proyectos.log) no fue encontrado.'})

    except Exception as e:

        return jsonify({'status': 'error', 'content': f'Error al leer el glosario: {str(e)}'})



@app.route('/api/glossary/add', methods=['POST'])

def add_to_glossary():

    try:

        data = request.get_json()

        project = data.get('project', '').strip()

        

        if not project:

            return jsonify({'status': 'error', 'message': 'El nombre del proyecto no puede estar vacío'})

        

        # Dynamic path resolution for portability

        glossary_path = BASE_DIR.parent / "Registro de Actividad" / "Codigos" / "config" / "glosario_proyectos.log"

        

        # Check if file exists, create if not

        if not os.path.exists(glossary_path):

            with open(glossary_path, 'w', encoding='utf-8') as f:

                f.write('')

        

        # Append new project

        with open(glossary_path, 'a', encoding='utf-8') as f:

            f.write(f'\n{project}')

        

        return jsonify({'status': 'success', 'message': 'Proyecto agregado correctamente'})

    

    except Exception as e:

        return jsonify({'status': 'error', 'message': f'Error al agregar proyecto: {str(e)}'})




@app.route('/api/activity-history', methods=['GET'])

def get_activity_history():

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

    

    # Paths (Dynamic based on BASE_DIR)
    # BASE_DIR points to '.../Registro de Control de Producto'
    # We need to go up to 'Oficina Tecnica' then down to 'Registro de Actividad'
    ACTIVITY_BASE = BASE_DIR.parent / "Registro de Actividad" / "Codigos"
    
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"



    user_email = session.get('user')

    print(f"[HISTORY] User from session: {user_email}")



    target_filename = None

    

    try:

        # 1. Lookup User

        if not DESTINATARIOS_FILE.exists():

             print(f"[HISTORY] Destinatarios file not found: {DESTINATARIOS_FILE}")

             return jsonify({'status': 'error', 'message': 'Archivo de destinatarios no encontrado.'}), 500

             

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:

            reader = csv.DictReader(f) # Headers: NOMBRE,APELLIDO,EMAIL

            for row in reader:

                # print(f"Checking row: {row}")

                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():

                    # Found

                    target_filename = f"{row['NOMBRE'].strip()} {row['APELLIDO'].strip()}.csv"

                    print(f"[HISTORY] Found target filename: {target_filename}")

                    break

        

        if not target_filename:

             print(f"[HISTORY] User {user_email} not found in destinatarios.")

             return jsonify({'status': 'error', 'message': 'Usuario no encontrado en lista de destinatarios.'}), 404



        # 2. Read CSV

        csv_path = DATA_DIR / target_filename

        print(f"[HISTORY] Looking for CSV at: {csv_path}")

        

        if not csv_path.exists():

             print(f"[HISTORY] CSV file does not exist.")

             return jsonify({'status': 'success', 'data': [], 'columns': []})

        

        data_rows = []

        columns = []

        

        def read_csv_data(enc):

            local_rows = []

            with open(csv_path, 'r', encoding=enc) as f:

                reader = csv.reader(f, delimiter=';')

                local_rows = list(reader)

            return local_rows



        try:

            try:

                rows = read_csv_data('utf-8')

            except UnicodeDecodeError:

                print(f"[HISTORY] UTF-8 failed, trying latin-1")

                rows = read_csv_data('latin-1')

            

            if len(rows) > 0:

                columns = rows[0]

                data_rows = rows[1:] # Skip header

        except Exception as e:

            print(f"[HISTORY] Error reading CSV: {e}")

            return jsonify({'status': 'error', 'message': f'Error leyendo archivo de historial: {str(e)}'}), 500

        

        print(f"[HISTORY] Found {len(data_rows)} rows.")

    except Exception as e:
        print(f"[HISTORY] Error general: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', 'data': data_rows, 'columns': columns})


@app.route('/api/activity-stats', methods=['GET'])
def get_activity_stats():
    """
    Devuelve horas acumuladas por proyecto para el usuario logueado.
    Filtros opcionales ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD.
    """
    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401

    ACTIVITY_BASE = BASE_DIR.parent / "Registro de Actividad" / "Codigos"
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"

    user_email = session.get('user')
    target_filename = None

    # Rango de fechas
    start_param = request.args.get('start_date')
    end_param = request.args.get('end_date')
    start_dt = None
    end_dt = None
    try:
        if start_param:
            start_dt = datetime.strptime(start_param, '%Y-%m-%d').date()
        if end_param:
            end_dt = datetime.strptime(end_param, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'status': 'error', 'message': 'Fechas inválidas'}), 400

    # Buscar archivo del usuario
    try:
        if not DESTINATARIOS_FILE.exists():
             return jsonify({'status': 'error', 'message': 'Archivo de destinatarios no encontrado.'}), 500

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():
                    target_filename = f"{row['NOMBRE'].strip()} {row['APELLIDO'].strip()}.csv"
                    break

        if not target_filename:
             return jsonify({'status': 'error', 'message': 'Usuario no encontrado en lista de destinatarios.'}), 404

        csv_path = DATA_DIR / target_filename
        if not csv_path.exists():
             return jsonify({'status': 'success', 'data': [], 'total': 0})

        def read_csv(enc):
            with open(csv_path, 'r', encoding=enc) as f:
                return list(csv.reader(f, delimiter=';'))

        try:
            rows = read_csv('utf-8')
        except Exception:
            rows = read_csv('latin-1')

        totals = {}
        records = [] # Initialize records list
        for idx, row in enumerate(rows):
            if not row or len(row) < 5:
                continue
            # Skip header row explicitly (first line) even if text is not exactly "Fecha"
            if idx == 0 or str(row[0]).lower().strip() == 'fecha':
                continue

            # Fecha en CSV es dd/mm/yyyy
            try:
                d = datetime.strptime(row[0], '%d/%m/%Y').date()
            except Exception:
                d = None

            if start_dt and d and d < start_dt:
                continue
            if end_dt and d and d > end_dt:
                continue

            proj = str(row[3]).strip() if len(row) > 3 else ''
            # horas viene en la columna 7 (index 7) en base_datos_respuestas.csv; fallback a col 4 si no existe
            time_str = str(row[7]).strip() if len(row) > 7 else (str(row[4]).strip() if len(row) > 4 else '0')
            # Extraer número (soporta "1", "1,5", "1.5", "1 Horas")
            num_match = re.search(r'[-+]?[0-9]*[\\.,]?[0-9]+', time_str)
            time_clean = num_match.group(0) if num_match else '0'
            try:
                hours = float(time_clean.replace(',', '.'))
            except Exception:
                hours = 0.0

            if not proj:
                proj = 'Sin Proyecto Asignado'
            totals[proj] = totals.get(proj, 0.0) + hours
            
            # Add detailed record for Area Chart
            # Columnas CSV Personal: 0 Fecha, 1 Hora, 2 Token, 3 Proyecto, 4 Tiempo, 5 Descripcion
            records.append({
                "project": proj,
                "hours": round(hours, 2),
                "user": "Actual", # Personal view always shows own data
                "date": row[0] if len(row) > 0 else '',
                "time": str(row[1]).strip() if len(row) > 1 else '',       
                "token": str(row[2]).strip() if len(row) > 2 else '',  
                "description": str(row[5]).strip() if len(row) > 5 else '' 
            })

        data_rows = [{'project': k, 'hours': round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[1], reverse=True)]
        total_hours = round(sum(totals.values()), 2)

        return jsonify({'status': 'success', 'data': data_rows, 'total': total_hours, 'records': records})

    except Exception as e:
        print(f"[STATS] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



@app.route('/api/activity-stats-global', methods=['GET'])
def get_activity_stats_global():
    """
    Devuelve horas acumuladas por proyecto considerando todos los CSV de respuestas.
    Filtros opcionales ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD.
    """
    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401

    ACTIVITY_BASE = BASE_DIR.parent / "Registro de Actividad" / "Codigos"
    MASTER_CSV = ACTIVITY_BASE / "data/base_datos_respuestas.csv"

    start_param = request.args.get('start_date')
    end_param = request.args.get('end_date')
    start_dt = None
    end_dt = None
    try:
        if start_param:
            start_dt = datetime.strptime(start_param, '%Y-%m-%d').date()
        if end_param:
            end_dt = datetime.strptime(end_param, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'status': 'error', 'message': 'Fechas inválidas'}), 400

    def read_rows(csv_path):
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                return list(csv.reader(f, delimiter=';'))
        except Exception:
            with open(csv_path, 'r', encoding='latin-1') as f:
                return list(csv.reader(f, delimiter=';'))

    totals = {}
    records = []
    try:
        if not MASTER_CSV.exists():
             return jsonify({'status': 'success', 'data': [], 'total': 0, 'records': []})

        try:
            rows = read_rows(MASTER_CSV)
        except Exception as e:
            print(f"[STATS][GLOBAL] Error leyendo maestro: {e}")
            return jsonify({'status': 'error', 'message': 'No se pudo leer base_datos_respuestas.csv'}), 500

        for idx_row, row in enumerate(rows):
            if not row or len(row) < 5:
                continue
            if idx_row == 0 or str(row[0]).lower().strip() == 'fecha':
                continue

            try:
                d = datetime.strptime(row[0], '%d/%m/%Y').date()
            except Exception:
                d = None

            if start_dt and d and d < start_dt:
                continue
            if end_dt and d and d > end_dt:
                continue

            # Columnas: 0 Fecha, 1 Hora, 2 Token, 3 Nombre, 4 Email, 5 ProyectoInicial, 6 ProyectoFinal, 7 Tiempo, 8 Descripcion
            user_val = str(row[3]).strip() if len(row) > 3 else 'Desconocido'
            proj = str(row[6]).strip() if len(row) > 6 else (str(row[5]).strip() if len(row) > 5 else '')
            time_str = str(row[7]).strip() if len(row) > 7 else '0'
            num_match = re.search(r'[-+]?[0-9]*[\\.,]?[0-9]+', time_str)
            time_clean = num_match.group(0) if num_match else '0'
            try:
                hours = float(time_clean.replace(',', '.'))
            except Exception:
                hours = 0.0

            if not proj:
                proj = 'Sin Proyecto Asignado'
            totals[proj] = totals.get(proj, 0.0) + hours
            records.append({
                "project": proj,
                "hours": round(hours, 2),
                "user": user_val,
                "date": row[0] if len(row) > 0 else '',
                "time": str(row[1]).strip() if len(row) > 1 else '',       # NEW: Hora
                "token": str(row[2]).strip() if len(row) > 2 else '',      # NEW: Token
                "description": str(row[8]).strip() if len(row) > 8 else '' # NEW: Descripcion
            })

        data_rows = [{'project': k, 'hours': round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[1], reverse=True)]
        total_hours = round(sum(totals.values()), 2)

        return jsonify({'status': 'success', 'data': data_rows, 'total': total_hours, 'records': records})
    except Exception as e:
        print(f"[STATS][GLOBAL] Error general: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500







# -------------------------------------------------------------------------

# ACTIVITY TRACKING EXTENSION

# -------------------------------------------------------------------------



ACTIVITY_BASE_PATH = BASE_DIR.parent / "Registro de Actividad" / "Codigos"

ACTIVITY_STATE_FILE = ACTIVITY_BASE_PATH / "data/activity_mailer_state.json"

ACTIVITY_CSV_BASE = ACTIVITY_BASE_PATH / "data/base_datos_respuestas.csv"

ACTIVITY_CSV_USER_DIR = ACTIVITY_BASE_PATH / "data/respuestas_csv"

DESTINATARIOS_FILE = ACTIVITY_BASE_PATH / "config/destinatarios.csv"



@app.route('/api/activity-pending', methods=['GET'])
def get_pending_activity():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    # Resolver email real del usuario (en users.json el username puede no ser el email completo)
    raw_user = session.get('user')
    user_email = raw_user.strip().lower() if raw_user else ''
    try:
        if USERS_FILE.exists() and raw_user and '@' not in raw_user:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                users_db = json.load(f)
            if raw_user in users_db:
                email_val = users_db[raw_user].get('email') or users_db[raw_user].get('correo')
                if email_val:
                    user_email = email_val.strip().lower()
    except Exception as e:
        print(f"[ACTIVITY] Warning resolving email: {e}")

    pending = []

    try:
        if not ACTIVITY_STATE_FILE.exists():
            return jsonify({'status': 'success', 'data': [], 'debug_user': user_email})

        with open(ACTIVITY_STATE_FILE, 'r', encoding='utf-8') as f:
            state = json.load(f)

        campaigns = state.get('campaigns', [])
        today = datetime.now().date()

        # Tomar todas las campañas con fecha <= hoy donde este usuario no tenga respuesta (ignorando recipients)
        for camp in campaigns:
            camp_date = camp.get('date')
            show_camp = False
            try:
                if camp_date:
                    d_obj = datetime.strptime(camp_date, '%Y-%m-%d').date()
                    
                    # 3 Working Days Filter
                    wdays = 0
                    curr = d_obj
                    while curr < today:
                         curr += timedelta(days=1)
                         if curr.weekday() < 5:
                             wdays += 1
                    
                    if wdays < 3:
                        show_camp = True
            except Exception:
                pass
            
            if not show_camp:
                continue

            responses = camp.get('responses', {})
            
            # Check if user is in recipients list
            recipients = [r.strip().lower() for r in camp.get('recipients', [])]
            if user_email not in recipients:
                continue

            if any(email_key.strip().lower() == user_email for email_key in responses):
                continue

            pending.append({
                'token': camp.get('token'),
                'date': camp.get('date'),
                'subject': camp.get('subject'),
                'type': 'activity_record',
                'debug_recipients': camp.get('recipients', [])
            })

    except Exception as e:
        print(f"[ACTIVITY] Error checking pending: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', 'data': pending, 'debug_user': user_email, 'debug_pending_count': len(pending)})

@app.route('/api/activity-submit', methods=['POST'])
def submit_activity_record():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    data = request.json
    token = data.get('token') if data else None
    project = data.get('project') if data else None
    time_val = data.get('time') if data else None
    description = data.get('description') if data else None

    # Resolver email y nombre
    raw_user = session.get('user')
    user_email = raw_user.strip().lower() if raw_user else ''
    
    # 1. First, ensure we have the correct email from users.json if possible
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, 'r') as f:
                users_db = json.load(f)
            key = raw_user if raw_user in users_db else user_email
            if key in users_db:
                email_val = users_db[key].get('email') or users_db[key].get('correo')
                if email_val:
                    user_email = email_val.strip().lower()
    except Exception as e:
        print(f"[ACTIVITY] Warning resolving user email: {e}")

    user_display_name = None

    # 2. Try resolving name from DESTINATARIOS_FILE (Highest Priority)
    if DESTINATARIOS_FILE.exists():
         try:
             with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
                 reader = csv.reader(f)
                 # Expecting header: NOMBRE,APELLIDO,EMAIL
                 header = next(reader, None) 
                 for row in reader:
                     if len(row) >= 3:
                         csv_email = row[2].strip().lower()
                         if csv_email == user_email:
                             user_display_name = f"{row[0].strip()} {row[1].strip()}"
                             break
         except Exception as e:
             print(f"[ACTIVITY] Warning reading destinatarios.csv: {e}")

    # 3. Fallback to users.json (if present)
    if not user_display_name and USERS_FILE.exists():
        try:
             with open(USERS_FILE, 'r') as f:
                users_db = json.load(f)
             key = raw_user if raw_user in users_db else user_email
             if key in users_db:
                 user_display_name = users_db[key].get('display_name')
        except: pass

    # 4. Final Fallback: Format from email
    if not user_display_name:
        user_display_name = user_email.split('@')[0].replace('.', ' ').title() if '@' in user_email else user_email

    if not token or not project:
        return jsonify({'status': 'error', 'message': 'Datos incompletos'}), 400

    try:
        if not ACTIVITY_STATE_FILE.exists():
            return jsonify({'status': 'error', 'message': 'Archivo de estado no encontrado'}), 500

        with open(ACTIVITY_STATE_FILE, 'r', encoding='utf-8') as f:
            state = json.load(f)

        campaign_idx = -1
        target_camp = None
        for idx, camp in enumerate(state.get('campaigns', [])):
            if camp.get('token') == token:
                campaign_idx = idx
                target_camp = camp
                break

        if not target_camp:
            return jsonify({'status': 'error', 'message': 'Token invalido o expirado'}), 404

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        body_mock = f'Proyecto: {project}\nTiempo destinado: {time_val}\nDescripci?n: {description}'

        response_obj = {
            "from": user_email,
            "from_name": user_display_name,
            "subject": f"Re: {target_camp.get('subject')}",
            "date": now_str,
            "received_at": datetime.now().isoformat(),
            "snippet": body_mock,
            "body": body_mock,
            "uid": 0,
            "saved_path": "WEB_SUBMISSION"
        }

        state['campaigns'][campaign_idx].setdefault('responses', {})
        
        # Append logic if user already responded (to support multi-project submissions)
        if user_email in state['campaigns'][campaign_idx]['responses']:
            existing = state['campaigns'][campaign_idx]['responses'][user_email]
            existing['body'] += "\n----------------\n" + body_mock
            existing['snippet'] = existing['body']
            existing['received_at'] = response_obj['received_at']
        else:
            state['campaigns'][campaign_idx]['responses'][user_email] = response_obj

        with open(ACTIVITY_STATE_FILE, 'w', encoding='utf-8') as f:
            json.dump(state, f, indent=2, ensure_ascii=False)

        def clean_csv(val):
            return str(val).replace(';', ',').replace('\\n', ' ').strip()

        date_str = target_camp.get('date')
        try:
            dobj = datetime.strptime(date_str, '%Y-%m-%d')
            date_str_csv = dobj.strftime('%d/%m/%Y')
        except:
            date_str_csv = date_str

        time_now_str = datetime.now().strftime('%H:%M:%S')

        # Separar multiples proyectos (coma, punto y coma o salto de linea)
        projects_list = [p.strip() for p in re.split(r"[;,\n]+", str(project or '')) if p.strip()]
        if not projects_list:
            projects_list = [str(project or '')]

        # Master CSV (una fila por proyecto)
        try:
            with open(ACTIVITY_CSV_BASE, 'a', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                for proj in projects_list:
                    master_row = [
                        date_str_csv,
                        time_now_str,
                        token,
                        user_display_name,
                        user_email,
                        clean_csv(proj),
                        clean_csv(proj),
                        clean_csv(time_val),
                        clean_csv(description)
                    ]
                    writer.writerow(master_row)
        except Exception as e:
            print(f"[ACTIVITY] Error writing Master CSV: {e}")

        # User CSV (una fila por proyecto)
        safe_name = re.sub(r'[<>:"/\|?*]', '_', user_display_name).strip()
        user_csv_path = ACTIVITY_CSV_USER_DIR / f"{safe_name}.csv"
        try:
            write_header = not user_csv_path.exists()
            with open(user_csv_path, 'a', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                if write_header:
                    writer.writerow(["Fecha", "Hora", "Token", "ProyectoFinal", "Tiempo", "Registro", "Observaciones"])
                for proj in projects_list:
                    user_row = [
                        date_str_csv,
                        time_now_str,
                        token,
                        clean_csv(proj),
                        clean_csv(time_val),
                        clean_csv(description),
                        ""
                    ]
                    writer.writerow(user_row)
        except Exception as e:
            print(f"[ACTIVITY] Error writing User CSV: {e}")

        return jsonify({'status': 'success', 'message': 'Registro guardado correctamente'})

    except Exception as e:
        print(f"[ACTIVITY] Error submitting: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/activity-update', methods=['POST'])
def update_activity_entry():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    
    data = request.json
    token = data.get('token')
    original_project = data.get('original_project')  # NEW: original project to identify row
    new_project = data.get('project')  # NEW: allow updating project
    new_desc = data.get('description', '')  # Optional, default to empty string
    new_time = data.get('time')

    # Validation: Only token and time are required
    if not token or not new_time:
         return jsonify({"status": "error", "message": "Faltan datos requeridos"}), 400
    
    # If original_project not provided, we cannot safely identify the row
    if not original_project:
         return jsonify({"status": "error", "message": "Proyecto original requerido para identificar registro"}), 400

    ACTIVITY_BASE = BASE_DIR.parent / "Registro de Actividad" / "Codigos"
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"

    user_email = session.get('user')
    target_filename = None

    try:
        # 1. Lookup User Filename
        if not DESTINATARIOS_FILE.exists():
             return jsonify({'status': 'error', 'message': 'Configuración no encontrada.'}), 500

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():
                    target_filename = f"{row['NOMBRE'].strip()} {row['APELLIDO'].strip()}.csv"
                    break
        
        if not target_filename:
             return jsonify({'status': 'error', 'message': 'Usuario no autorizado.'}), 403

        csv_path = DATA_DIR / target_filename
        if not csv_path.exists():
             return jsonify({'status': 'error', 'message': 'Archivo de registros no encontrado.'}), 404

        # 2. Read All Rows
        updated_rows = []
        found = False
        
        # We need to detect encoding/delimiter or assume standard. 
        # Since get_activity_history implies successful read, we assume UTF-8 and comma (or whatever matches).
        # We'll stick to 'utf-8'.
        
        # Read and stored in memory
        with open(csv_path, 'r', encoding='utf-8', newline='') as f:
            reader = csv.reader(f, delimiter=';')
            header = next(reader, None)
            if header:
                updated_rows.append(header)
            
            for row in reader:
                # Row structure: Date, Time, Token, Project, Effort, Desc, Obs
                # Match using Token (index 2) AND Project (index 3)
                # This ensures we only update the specific activity, not all with same token
                if len(row) > 3 and row[2] == token and row[3] == original_project:
                    # Found the exact row to update
                    if len(row) > 5:
                        # Update Project (3), Time (4), and Description (5)
                        if new_project:  # Allow project updates
                            row[3] = new_project
                        row[4] = new_time
                        row[5] = new_desc
                        
                        # Log modification in Observation (Index 6)
                        timestamp = datetime.now().strftime("%d/%m %H:%M")
                        mod_note = f"[Modificado el {timestamp}]"
                        if len(row) > 6:
                             row[6] = mod_note
                        else:
                             row.append(mod_note)
                        
                        found = True
                updated_rows.append(row)

        if not found:
            return jsonify({'status': 'error', 'message': 'Registro no encontrado.'}), 404

        # 3. Write Back
        with open(csv_path, 'w', encoding='utf-8', newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerows(updated_rows)

        return jsonify({'status': 'success', 'message': 'Registro actualizado correctamente.'})

    except Exception as e:
        print(f"Update Error: {e}")
        return jsonify({'status': 'error', 'message': f'Error interno: {str(e)}'}), 500

# --- MULTIPLAYER GAME STATE ---
connected_users = {} # {sid: {'username': user, 'room': None}}
game_rooms = {} # {room_id: {'p1': sid, 'p2': sid, 'state': ...}}

@socketio.on('connect')
def handle_connect():
    print(f"Client Connected: {request.sid}")

@socketio.on('identify')
def handle_identify(data):
    username = data.get('username')
    connected_users[request.sid] = {'username': username, 'room': None}
    # Broadcast updated user list
    clean_list = [{'username': u['username'], 'sid': s} for s, u in connected_users.items()]
    emit('user_list_update', clean_list, broadcast=True)

@socketio.on('disconnect')
def handle_disconnect():
    if request.sid in connected_users:
        del connected_users[request.sid]
    clean_list = [{'username': u['username'], 'sid': s} for s, u in connected_users.items()]
    emit('user_list_update', clean_list, broadcast=True)
    print(f"Client Disconnected: {request.sid}")

@socketio.on('send_invite')
def handle_invite(data):
    target_sid = data.get('target_sid')
    sender_user = data.get('sender_user')
    if target_sid:
        emit('receive_invite', {'sender': sender_user, 'sender_sid': request.sid}, room=target_sid)

@socketio.on('accept_invite')
def handle_accept(data):
    host_sid = data.get('sender_sid')
    client_sid = request.sid
    
    room_id = str(uuid.uuid4())
    join_room(room_id, sid=host_sid)
    join_room(room_id, sid=client_sid)
    
    game_rooms[room_id] = {'p1': host_sid, 'p2': client_sid}
    
    # Notify Start
    emit('start_game', {'room_id': room_id, 'opponent': connected_users[client_sid]['username'], 'role': 'host'}, room=host_sid)
    emit('start_game', {'room_id': room_id, 'opponent': connected_users[host_sid]['username'], 'role': 'client'}, room=client_sid)

@socketio.on('game_update')
def handle_game_update(data):
    room = data.get('room_id')
    if room:
        emit('game_state_sync', data, room=room, include_self=False)

if __name__ == '__main__':
    import webbrowser
    from threading import Timer
    import os
    try:
        from waitress import serve
    except ImportError:
        serve = None

    # Only open browser in the main process
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        pass

    def open_browser():
        try:
            # Try to find Chrome
            chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe %s"
            if not os.path.exists("C:/Program Files/Google/Chrome/Application/chrome.exe"):
                chrome_path = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"
            
            webbrowser.get(chrome_path).open("http://127.0.0.1:5000")
        except Exception:
            webbrowser.open("http://127.0.0.1:5000")

    Timer(1.5, open_browser).start()

    print("----------------------------------------------------------------")
    print(" INICIANDO SERVIDOR CON SOCKET.IO (Eventlet)")
    print(" URL: http://0.0.0.0:5000")
    print("----------------------------------------------------------------")
    try:
        socketio.run(app, host='0.0.0.0', port=5000, debug=False)
    except Exception as e:
        print(f"Error iniciando SocketIO: {e}")
