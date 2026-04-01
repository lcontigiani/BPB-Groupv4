import argparse
import datetime as dt
"""
Zonas del modulo:
- Datos centrales (Config, carga de config, logging, dependencias).
- Utilidades de PDF (OCR, patrones) y OpenAI opcional.
- Utilidades de registros y formulas (mapa parte->producto).
- Analisis de partes (mapear partes B&P-00x/BPB-00x a producto base via formula).
- Analisis de producto completo (scoring y clustering de planos).
- Modo reporte y modo procesar (copiar/actualizar).
"""

import io
import json
import logging
import os
import re
import shutil
import time
import base64
import difflib
import requests
import requests
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import csv
import yaml


try:
    import pdfplumber
except ImportError:  # pragma: no cover - dependency check
    pdfplumber = None

try:
    import pypdfium2 as pdfium
except ImportError:  # pragma: no cover - dependency check
    pdfium = None

try:
    import openpyxl
except ImportError:  # pragma: no cover - dependency check
    openpyxl = None

try:
    import pytesseract
except ImportError:  # pragma: no cover - dependency check
    pytesseract = None

try:
    import msoffcrypto
except ImportError:  # pragma: no cover - dependency check
    msoffcrypto = None

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency check
    xlrd = None

SUPPORTED_EXTS = {".pdf", ".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}

try:
    import xlrd
except ImportError:  # pragma: no cover - dependency check
    xlrd = None
@dataclass
class Config:
    incoming_dir: Path
    processed_dir: Path
    failed_dir: Path
    output_root: Path
    batch_prefix: str
    register_excel: Optional[Path]
    register_root_dir: Optional[Path]
    register_file_prefix: str
    register_password: str
    formula_csv: Optional[Path]
    aux_excel: Optional[Path]
    product_pdf_dir: Optional[Path]
    watch_enabled: bool
    watch_scan_seconds: int
    pdf_max_pages: int
    pdf_max_products: int
    pdf_ocr_enabled: bool
    pdf_tesseract_cmd: str
    product_code_patterns: List[str]
    order_patterns: List[str]
    dimension_patterns: List[str]
    register_sheet_name: str
    register_code_column: int
    register_headers: Dict[str, str]
    openai_enabled: bool
    openai_model: str
    openai_temperature: float
    openai_api_key_env: str
    openai_max_text_chars: int
    log_file: Path
    log_level: str
    registros_root: Optional[Path]
    aux_indices_dir: Optional[Path]


def load_config(path: Path) -> Config:
    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}

    # Determine Root Directory
    # Priority 1: Environment Variable
    env_root = os.environ.get("BPB_BASE_DIR")
    if env_root:
        root = Path(env_root)
    else:
        # Priority 2: Relative to config file (assuming in Codigos/)
        root = path.parent.parent

    def resolve(p: Optional[str]) -> Optional[Path]:
        if not p:
            return None
        p = str(p)
        pth = Path(p)
        if not pth.is_absolute():
            pth = root / pth
        return pth

    logging_cfg = raw.get("logging", {})
    pdf_cfg = raw.get("pdf_parsing", {})
    register_cfg = raw.get("register_columns", {})
    watch_cfg = raw.get("watch", {})
    openai_cfg = raw.get("openai", {})

    return Config(
        incoming_dir=resolve(raw.get("incoming_dir", "P1 - Registros Solicitados")),
        processed_dir=resolve(raw.get("processed_dir", "P1 - Registros Solicitados\\in process")),
        failed_dir=resolve(raw.get("failed_dir", "P1 - Registros Solicitados\\failed")),
        output_root=resolve(raw.get("output_root", "P2 - Purchase Order")),
        batch_prefix=str(raw.get("batch_prefix", "Registro")),
        register_excel=resolve(raw.get("register_excel")),
        register_root_dir=resolve(raw.get("register_root_dir")),
        register_file_prefix=str(raw.get("register_file_prefix", "")),
        register_password=str(raw.get("register_password", "")),
        formula_csv=resolve(raw.get("formula_csv")),
        aux_excel=resolve(raw.get("aux_excel")),
        product_pdf_dir=resolve(raw.get("product_pdf_dir")),
        watch_enabled=bool(watch_cfg.get("enabled", False)),
        watch_scan_seconds=int(watch_cfg.get("scan_seconds", 10)),
        pdf_max_pages=int(pdf_cfg.get("max_pages", 3)),
        pdf_max_products=int(pdf_cfg.get("max_products", 0)),
        pdf_ocr_enabled=bool(pdf_cfg.get("ocr_enabled", False)),
        pdf_tesseract_cmd=str(pdf_cfg.get("tesseract_cmd", "")),
        product_code_patterns=list(pdf_cfg.get("product_code_patterns", [])),
        order_patterns=list(pdf_cfg.get("order_patterns", [])),
        dimension_patterns=list(pdf_cfg.get("dimension_patterns", [])),
        register_sheet_name=str(register_cfg.get("sheet_name", "")),
        register_code_column=int(register_cfg.get("code_column", 2)),
        register_headers=dict(
            code=register_cfg.get("headers", {}).get("code", "Codigo"),
            revision=register_cfg.get("headers", {}).get("revision", "Revision"),
            purchase_order=register_cfg.get("headers", {}).get("purchase_order", "OC"),
            notes=register_cfg.get("headers", {}).get("notes", "Notas"),
        ),
        openai_enabled=bool(openai_cfg.get("enabled", False)),
        openai_model=str(openai_cfg.get("model", "gpt-4o-mini")),
        openai_temperature=float(openai_cfg.get("temperature", 0.0)),
        openai_api_key_env=str(openai_cfg.get("api_key_env", "OPENAI_API_KEY")),
        openai_max_text_chars=int(openai_cfg.get("max_text_chars", 6000)),
        log_file=resolve(logging_cfg.get("file", "Codigos\\automation.log")),
        log_level=str(logging_cfg.get("level", "INFO")).upper(),
        registros_root=resolve(raw.get("register_root_dir") or raw.get("registros_root_dir") or r"\\192.168.0.55\utn\REGISTROS"),
        aux_indices_dir=resolve(raw.get("aux_indices_dir") or "Auxiliares\\indices_auxiliar"),
    )


def setup_logging(log_file: Path, level: str) -> None:
    log_file.parent.mkdir(parents=True, exist_ok=True)
    handlers = [
        logging.FileHandler(log_file, encoding="utf-8"),
        logging.StreamHandler(),
    ]
    logging.basicConfig(
        level=getattr(logging, level, logging.INFO),
        format="%(asctime)s [%(levelname)s] %(message)s",
        handlers=handlers,
    )


def ensure_dependencies() -> None:
    missing = []
    if pdfplumber is None:
        missing.append("pdfplumber")
    if openpyxl is None:
        missing.append("openpyxl")
    if pdfium is None:
        missing.append("pypdfium2")
    if pytesseract is None:
        missing.append("pytesseract")
    if msoffcrypto is None:
        missing.append("msoffcrypto-tool")
    if missing:
        raise RuntimeError(
            f"Faltan dependencias: {', '.join(missing)}. Instala con pip install -r requirements.txt"
        )


def normalize_header(value: str) -> str:
    return re.sub(r"\s+", "", value).lower()


def ensure_dirs(cfg: Config) -> None:
    for p in [cfg.incoming_dir, cfg.processed_dir, cfg.failed_dir, cfg.output_root]:
        p.mkdir(parents=True, exist_ok=True)


def load_image_bytes(path: Path) -> Optional[Tuple[bytes, str]]:
    """Carga una imagen y devuelve (bytes PNG, mime) para inline_data."""
    try:
        from PIL import Image
    except ImportError:
        logging.warning("Pillow no instalado; no se pueden leer imagenes %s", path)
        return None
    try:
        img = Image.open(path)
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        return buf.getvalue(), "image/png"
    except Exception as exc:
        logging.warning("No se pudo convertir imagen %s: %s", path, exc)
        return None


def next_sequential_id(seq_path: Path) -> str:
    """
    Lee/incrementa un contador en seq_path y devuelve el id con formato R0001.
    """
    seq_path.parent.mkdir(parents=True, exist_ok=True)
    current = 0
    if seq_path.exists():
        try:
            data = seq_path.read_text(encoding="utf-8").strip()
            m = re.search(r"(\d+)", data)
            if m:
                current = int(m.group(1))
        except Exception:
            current = 0
    new_val = current + 1
    seq_id = f"R{new_val:04d}"
    seq_path.write_text(seq_id, encoding="utf-8")
    return seq_id


def find_last_resumen(processed_dir: Path) -> Optional[Path]:
    """Devuelve la ruta del ultimo resumen.csv en processed_dir (ordenado por secuencia)."""
    if not processed_dir.exists():
        return None
    candidatos = []
    for entry in processed_dir.iterdir():
        if not entry.is_dir():
            continue
        name = entry.name
        m = re.match(r"Registro\s*-\s*(R\d{4})", name, re.IGNORECASE)
        if m:
            candidatos.append((m.group(1), entry))
    if not candidatos:
        return None
    # ordenar por secuencia numerica
    candidatos.sort(key=lambda t: int(re.sub(r"\D", "", t[0])), reverse=True)
    for _, folder in candidatos:
        resumen = folder / "resumen.csv"
        if resumen.exists():
            return resumen
    return None


def load_resumen_rows(path: Path) -> List[Tuple[str, str, str, str]]:
    """Carga filas canonicas (Codigo, PO, Obs, Fecha) de un resumen.csv."""
    rows: List[Tuple[str, str, str, str]] = []
    try:
        with open(path, newline="", encoding="utf-8") as f:
            reader = csv.reader(f, delimiter=";")
            header = next(reader, None)
            for row in reader:
                if not row:
                    continue
                # Soportar formatos antiguos (3 cols) o nuevos (4 cols)
                if len(row) >= 4:
                    code = row[-4].strip()
                    po = row[-3].strip()
                    obs = row[-2].strip()
                    fecha = row[-1].strip()
                    rows.append((code, po, obs, fecha))
                elif len(row) >= 3:
                    code = row[-3].strip()
                    po = row[-2].strip()
                    obs = row[-1].strip()
                    rows.append((code, po, obs, ""))
    except Exception:
        pass
    return rows


def extract_text(pdf_path: Path, max_pages: int) -> str:
    if pdfplumber is None:
        return ""
    texts = []
    with pdfplumber.open(pdf_path) as pdf:
        for idx, page in enumerate(pdf.pages):
            if idx >= max_pages:
                break
            page_text = page.extract_text() or ""
            texts.append(page_text)
    return "\n".join(texts)


def needs_ocr(text: str) -> bool:
    cleaned = re.sub(r"[\s\r\n]+", "", text or "")
    if len(cleaned) < 20:
        return True
    if not re.search(r"[A-Za-z0-9]{3,}", cleaned):
        return True
    return False


def extract_text_ocr(pdf_path: Path, cfg: Config) -> str:
    if not cfg.pdf_ocr_enabled or pdfium is None or pytesseract is None:
        return ""
    if cfg.pdf_tesseract_cmd:
        pytesseract.pytesseract.tesseract_cmd = cfg.pdf_tesseract_cmd
    texts: List[str] = []
    doc = pdfium.PdfDocument(pdf_path)
    try:
        pages = min(len(doc), cfg.pdf_max_pages)
        for idx in range(pages):
            page = doc[idx]
            bitmap = page.render(scale=3.0)
            pil_img = bitmap.to_pil()
            ocr_text = pytesseract.image_to_string(
                pil_img,
                lang="eng",
                config="--psm 6 -c tessedit_char_whitelist=ABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789-./&",
            )
            texts.append(ocr_text)
    finally:
        try:
            doc.close()
        except Exception:
            pass
    return "\n".join(texts)


def extract_with_patterns(text: str, patterns: List[str]) -> List[str]:
    found = []
    for pattern in patterns:
        try:
            matches = re.findall(pattern, text)
        except re.error as err:
            logging.warning("Patron invalido %s: %s", pattern, err)
            continue
        if not matches:
            continue
        if isinstance(matches[0], tuple):
            matches = [m[0] for m in matches]
        found.extend(matches)
    # preserve order, remove duplicates
    seen = set()
    ordered = []
    for item in found:
        item = item.strip()
        if not item:
            continue
        if item.lower() in seen:
            continue
        seen.add(item.lower())
        ordered.append(item)
    return ordered


def call_openai_for_pdf(text: str, cfg: Config) -> Dict:
    api_key = os.getenv(cfg.openai_api_key_env)
    if not api_key:
        logging.info("OPENAI_API_KEY no definido; se omite capa OpenAI")
        return {}
    try:
        from openai import OpenAI
    except ImportError:
        logging.warning("Paquete openai no instalado; omitiendo enriquecimiento")
        return {}

    client = OpenAI(api_key=api_key)
    prompt = (
        "Observa SOLO las IMÁGENES adjuntas (no inventes). Es una tabla con columnas en este orden: "
        "1.Emp, 2.Comp, 3.Numero, 4.Fecha, 5.Producto, 6.Cantidad, 7.Depo, 8.Ubicacion, "
        "9.Observac, 10.Proveed, 11.Usuario, 12.OC, 13.CodOC.\n"
        "- Extrae SOLO: columna 5 (Producto), columna 12 (OC), columna 9 (Observac) y columna 4 (Fecha exacta de la fila). No uses CodOC ni otras columnas.\n"
        "- La OC y la Observac deben ser SOLO números (1 a 8 dígitos). Si Observac viene como 'PO#1053 ESUN', devuelve 1053. Si no hay número, deja vacío.\n"
        "- Devuelve SOLO JSON plano: {\\\"items\\\": [{\\\"code\\\": \\\"<Producto>\\\", \\\"po\\\": \\\"<OC>\\\", \\\"obs\\\": \\\"<ObservacNumero>\\\", \\\"date\\\": \\\"<Fecha>\\\"}, ...]} sin texto extra ni markdown.\n"
        "- Copia exactamente los caracteres que veas (guiones, letras, números). No corrijas S/5 ni O/0; si dudas deja el campo vacío.\n"
        "- Recorre TODAS las filas visibles en todas las imágenes; mantiene el orden."
    )
    parts = [{"text": prompt}]
    imgs = render_pdf_pages_png(pdf_path, max_pages=cfg.pdf_max_pages)
    for img in imgs:
        parts.append(
            {
                "inline_data": {
                    "mime_type": "image/png",
                    "data": base64.b64encode(img).decode(),
                }
            }
        )

    payload = {"contents": [{"parts": parts}]}
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    try:
        import requests

        resp = requests.post(url, json=payload, timeout=150)
        if resp.status_code != 200:
            logging.warning("Gemini status %s: %s", resp.status_code, resp.text[:400])
            return []
        data = resp.json()
        text_resp = data["candidates"][0]["content"]["parts"][0]["text"]
        m = re.search(r"\{.*\}", text_resp, re.S)
        payload_json = json.loads(m.group(0)) if m else json.loads(text_resp)
        items = payload_json.get("items") if isinstance(payload_json, dict) else None
        if not isinstance(items, list):
            return []
        out: List[Dict[str, str]] = []
        for it in items:
            code = str(it.get("code") or "").strip()
            po = str(it.get("po") or "").strip()
            obs = str(it.get("obs") or "").strip()
            date_val = str(it.get("date") or "").strip()
            if obs:
                mnum = re.search(r"\d+", obs)
                obs = mnum.group(0) if mnum else ""
            if not po or po == "0":
                po = "Sin OC"
            if code or po or obs:
                out.append({"code": code, "po": po, "obs": obs, "date": date_val})
        return out
    except Exception as exc:
        logging.warning("Fallo Gemini: %s", exc)
        return []

def parse_codes_po_local(pdf_path: Path, cfg: Config) -> List[Dict[str, str]]:
    """Fallback local: usa texto extraido (pdfplumber + tesseract) para encontrar columnas Producto y OC."""
    texts: List[str] = []
    # OCR por imagen (m�s confiable para tablas escaneadas)
    img_ocr = ocr_images_text(pdf_path, cfg)
    if img_ocr:
        texts.append(img_ocr)
    # Texto con pdfplumber (por si hay capa de texto)
    try:
        if pdfplumber:
            with pdfplumber.open(pdf_path) as pdf:
                for idx, page in enumerate(pdf.pages):
                    if idx >= cfg.pdf_max_pages:
                        break
                    t = page.extract_text() or ""
                    if t:
                        texts.append(t)
    except Exception:
        pass
    # OCR texto (no imagen) como extra
    if cfg.pdf_ocr_enabled and pytesseract and pdfium:
        try:
            texts.append(extract_text_ocr(pdf_path, cfg))
        except Exception:
            pass
    text = "\n".join(texts)
    items: List[Dict[str, str]] = []
    if not text:
        return items
    lines = [ln for ln in text.splitlines() if ln.strip()]
    header_idx = None
    code_col = None
    po_col = None
    for idx, ln in enumerate(lines):
        parts = re.split(r"\s{2,}|\t+", ln.strip())
        parts_low = [p.lower() for p in parts]
        if any("producto" in p or p == "product" for p in parts_low) and any(p == "po" or "purchase" in p for p in parts_low):
            header_idx = idx
            for i, p in enumerate(parts_low):
                if code_col is None and ("producto" in p or p == "product"):
                    code_col = i
                if po_col is None and (p == "po" or "purchase" in p):
                    po_col = i
            break
    if header_idx is not None and code_col is not None and po_col is not None:
        for ln in lines[header_idx + 1 :]:
            parts = re.split(r"\s{2,}|\t+", ln.strip())
            if len(parts) > max(code_col, po_col):
                code = parts[code_col].strip()
                po = ""
                matches = re.findall(r"\d{2,6}", parts[po_col])
                if matches:
                    po = matches[-1]
                if code:
                    items.append({"code": code, "po": po})
            else:
                # fallback por l�nea: tomar c�digo(s) y la ULTIMA secuencia num�rica de 2-6 d�gitos como PO
                codes = re.findall(r"\b(?:B&P|BPB)[A-Z0-9\-]{3,}\b", ln, flags=re.IGNORECASE)
                nums = re.findall(r"\d{2,6}", ln)
                po_val = nums[-1] if nums else ""
                for c in codes:
                    items.append({"code": c.strip(), "po": po_val})
    # si no se encontr� encabezado, fallback regex general
    if not items:
        code_pattern = re.compile(r"\b(?:B&P|BPB)[A-Z0-9\-]{3,}\b", re.IGNORECASE)
        po_pattern = re.compile(r"\b(\d{2,6})\b")
        for line in lines:
            codes = code_pattern.findall(line)
            pos = po_pattern.findall(line)
            po_val = pos[0] if pos else ""
            for c in codes:
                items.append({"code": c.strip(), "po": po_val})
    return items




def copy_if_set(src: Optional[Path], dest_dir: Path) -> Optional[Path]:
    if not src:
        return None
    if not src.exists():
        logging.warning("No existe archivo origen: %s", src)
        return None
    dest_dir.mkdir(parents=True, exist_ok=True)
    dest = dest_dir / src.name
    shutil.copy2(src, dest)
    return dest


def parse_revision_letter(name: str) -> str:
    match = re.search(r"(?i)rev\s*\.?(\w)", name)
    if match:
        return match.group(1).upper()
    return ""


def normalize_string(s: str) -> str:
    return re.sub(r"[^a-z0-9]", "", s.lower())


def tokenize_code(code: str) -> List[str]:
    parts = re.split(r"[^A-Za-z0-9]+", code)
    return [p for p in parts if len(p) >= 3]


def normalize_o0(value: str) -> str:
    return value.replace("O", "0").replace("o", "0")


def extract_series_token(code: str) -> Optional[str]:
    tokens = tokenize_code(code)
    for tok in tokens:
        tok_norm = normalize_o0(tok)
        m = re.match(r"(?i)(ms|ls)\d{2}", tok_norm)
        if m:
            return tok_norm.upper()
    return None


def gather_tokens_from_path(path: Path) -> List[str]:
    tokens: List[str] = []
    parts = list(path.parts)
    for part in parts:
        for tok in tokenize_code(part):
            tokens.append(tok)
            norm = normalize_o0(tok)
            if norm != tok:
                tokens.append(norm)
    return tokens


def score_candidate(purchase_order: Optional[str], code: str, path: Path) -> float:
    total, _ = score_candidate_detail(purchase_order, code, path)
    return total


def score_candidate_detail(purchase_order: Optional[str], code: str, path: Path) -> Tuple[float, Dict[str, float]]:
    name_no_ext = path.stem.lower()
    score = 0.0
    detail: Dict[str, float] = {}
    series = extract_series_token(code)
    if purchase_order:
        po = purchase_order.lower()
        if po and po in name_no_ext:
            score += 8.0
            detail["po_in_name"] = 8.0
        if f"po{po}" in name_no_ext:
            score += 4.0
            detail["po_prefix"] = detail.get("po_prefix", 0) + 4.0
    code_tokens = tokenize_code(code)
    path_tokens = gather_tokens_from_path(path)
    token_score = 0.0
    for tok in code_tokens:
        tl = tok.lower()
        tl_norm = normalize_o0(tl)
        if tl_norm and any(tl_norm in pt.lower() or pt.lower() in tl_norm for pt in path_tokens):
            token_score += 3.0
    if token_score:
        score += token_score
        detail["token_overlap"] = token_score
    if series:
        series_lower = series.lower()
        if any(series_lower in pt.lower() for pt in path_tokens):
            score += 12.0
            detail["series_match"] = 12.0
        else:
            score -= 10.0
            detail["series_penalty"] = -10.0
    # similitud global
    sim = 0.0
    if code:
        try:
            import difflib

            sim = difflib.SequenceMatcher(None, normalize_string(code), normalize_string(name_no_ext)).ratio()
        except Exception:
            sim = 0.0
    sim_score = sim * 2.0
    score += sim_score  # peso suave
    if sim_score:
        detail["similarity"] = sim_score
    return score, detail


def find_product_pdf(purchase_order: Optional[str], code: str, source_dir: Optional[Path]) -> Optional[Path]:
    if not source_dir or not source_dir.exists():
        return None
    candidates = list(source_dir.rglob("*.pdf"))
    return find_product_pdf_cached(purchase_order, code, candidates)


def find_product_pdf_cached(purchase_order: Optional[str], code: str, candidates: List[Path]) -> Optional[Path]:
    best_path, best_score, _ = find_product_pdf_with_score(purchase_order, code, candidates)
    return best_path


def find_product_pdf_with_score(purchase_order: Optional[str], code: str, candidates: List[Path]) -> Tuple[Optional[Path], float, Dict[str, float]]:
    if not candidates:
        return None, -1.0, {}

    best = None
    best_score = -1.0
    best_rev = ""
    best_detail: Dict[str, float] = {}
    for path in candidates:
        sc, detail = score_candidate_detail(purchase_order, code, path)
        if sc <= 0:
            continue
        if purchase_order and f"po{purchase_order}".lower() not in path.name.lower():
            # si la PO no esta en el nombre, pero tokens son muy fuertes, darle chance reducida
            if sc < 8.0:
                continue
        rev = parse_revision_letter(path.name)
        if sc > best_score or (abs(sc - best_score) < 1e-6 and rev > best_rev):
            best = path
            best_score = sc
            best_rev = rev
            best_detail = detail
    if best_score < 3.0:
        return None, best_score, best_detail
    return best, best_score, best_detail


def ensure_headers(ws, headers: Dict[str, str], code_col_index: int) -> Dict[str, int]:
    header_row = 1
    positions: Dict[str, int] = {"code": code_col_index}
    existing = {normalize_header(str(cell.value)): cell.column for cell in ws[header_row] if cell.value}
    next_col = ws.max_column + 1 if ws.max_column else 1
    for key, header_name in headers.items():
        if key == "code":
            continue
        norm = normalize_header(header_name)
        if norm not in existing:
            ws.cell(row=header_row, column=next_col, value=header_name)
            existing[norm] = next_col
            next_col += 1
        positions[key] = existing[norm]
    return positions


def find_row_by_code(ws, code_col: int, code: str) -> Optional[int]:
    for row in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row, column=code_col).value
        if cell_val and str(cell_val).strip().lower() == code.lower():
            return row
    return None


def find_register_workbook_for_code(code: str, cfg: Config) -> Optional[Path]:
    if cfg.register_excel and cfg.register_excel.exists():
        return cfg.register_excel
    root = cfg.register_root_dir
    if not root or not root.exists():
        logging.warning("No se encontro register_root_dir configurado.")
        return None
    if openpyxl is None:
        logging.warning("openpyxl no disponible; no se puede leer registros.")
        return None
    prefix = cfg.register_file_prefix or ""
    patterns = [
        f"{prefix}*.xlsx",
        f"{prefix}*.xlsm",
        f"{prefix}*.xlsb",
    ]
    candidates: List[Path] = []
    for pat in patterns:
        candidates.extend(root.rglob(pat))
    # ordenar por fecha reciente para probar primero los ultimos
    candidates = sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)
    for path in candidates:
        wb = None
        try:
            wb = open_workbook(path, cfg.register_password, read_only=True, data_only=True)
            if wb is None:
                continue
            ws = wb[cfg.register_sheet_name] if cfg.register_sheet_name and cfg.register_sheet_name in wb.sheetnames else wb.active
            code_col = cfg.register_code_column
            for row in ws.iter_rows(min_row=2, max_col=code_col, min_col=code_col):
                cell = row[0]
                if cell.value and str(cell.value).strip().lower() == code.lower():
                    wb.close()
                    logging.info("Registro encontrado en: %s", path)
                    return path
        except Exception as exc:
            logging.debug("No se pudo abrir %s: %s", path, exc)
        finally:
            try:
                if wb:
                    wb.close()
            except Exception:
                pass
    return None


def update_register_excel(path: Path, items: List[Dict], cfg: Config) -> None:
    if openpyxl is None:
        logging.warning("openpyxl no disponible; no se actualiza Excel")
        return
    wb = open_workbook(path, cfg.register_password, read_only=False, data_only=False)
    if wb is None:
        logging.warning("No se pudo abrir el Excel de registro para escritura: %s", path)
        return
    ws = wb[cfg.register_sheet_name] if cfg.register_sheet_name and cfg.register_sheet_name in wb.sheetnames else wb.active
    headers_pos = ensure_headers(ws, cfg.register_headers, cfg.register_code_column)
    code_col = cfg.register_code_column
    for item in items:
        row = find_row_by_code(ws, code_col, item["code"])
        if row is None:
            row = ws.max_row + 1
        ws.cell(row=row, column=code_col, value=item["code"])
        if item.get("revision") is not None:
            ws.cell(row=row, column=headers_pos["revision"], value=item["revision"])
        if item.get("purchase_order") is not None:
            ws.cell(row=row, column=headers_pos["purchase_order"], value=item["purchase_order"])
        if item.get("notes"):
            ws.cell(row=row, column=headers_pos["notes"], value=item["notes"])
    wb.save(path)
    wb.close()


def write_notes(dest_dir: Path, notes: List[str]) -> None:
    notes_path = dest_dir / "notes.txt"
    timestamp = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    content = [f"[{timestamp}]"] + notes
    notes_path.write_text("\n".join(content), encoding="utf-8")


def open_workbook(path: Path, password: Optional[str], read_only: bool, data_only: bool):
    if openpyxl is None:
        return None
    try:
        return openpyxl.load_workbook(path, read_only=read_only, data_only=data_only)
    except Exception:
        if not password or msoffcrypto is None:
            raise
    # intentar desencriptar con msoffcrypto
    try:
        with open(path, "rb") as f:
            office = msoffcrypto.OfficeFile(f)
            office.load_key(password=password)
            output = io.BytesIO()
            office.decrypt(output)
            output.seek(0)
        return openpyxl.load_workbook(output, read_only=read_only, data_only=data_only)
    except Exception as exc:
        logging.debug("Fallo abriendo con password %s: %s", path, exc)
        return None


_formula_cache: Dict[Path, Dict[str, str]] = {}


def load_formula_map(path: Optional[Path]) -> Dict[str, str]:
    if not path or not path.exists():
        return {}
    if path in _formula_cache:
        return _formula_cache[path]
    mapping: Dict[str, str] = {}
    for enc in ("utf-8", "latin-1"):
        try:
            with open(path, newline="", encoding=enc, errors="ignore") as f:
                reader = csv.reader(f, delimiter=";")
                header = next(reader, None)
                for row in reader:
                    if len(row) < 8:
                        continue
                    product = row[2].strip()
                    part = row[7].strip()
                    if not part or not product:
                        continue
                    prev = mapping.get(part)
                    if prev:
                        try:
                            import difflib
                            if difflib.SequenceMatcher(None, part, product).ratio() <= difflib.SequenceMatcher(None, part, prev).ratio():
                                continue
                        except Exception:
                            continue
                    mapping[part] = product
            break
        except Exception as exc:
            logging.warning("No se pudo cargar formula_csv %s con %s: %s", path, enc, exc)
    _formula_cache[path] = mapping
    return mapping


def generate_report(cfg: Config) -> None:
    # Seleccionar PDF de entrada (Registros o in process como fallback)
    pdfs = sorted(cfg.incoming_dir.glob("*.pdf"))
    pdf_path = pdfs[0] if pdfs else (cfg.processed_dir / "Registro.pdf")
    if not pdf_path.exists():
        logging.error("No hay PDF de entrada para reporte.")
        return
    analysis = analyze_pdf(pdf_path, cfg)
    codes = analysis["product_codes"]
    formula_map = load_formula_map(cfg.formula_csv)
    groups: Dict[str, List[str]] = {}
    for code in codes:
        base = formula_map.get(code, code)
        groups.setdefault(base, []).append(code)
    candidates = list(cfg.product_pdf_dir.rglob("*.pdf")) if cfg.product_pdf_dir and cfg.product_pdf_dir.exists() else []
    report_lines = []
    report_lines.append(f"PDF: {pdf_path}")
    report_lines.append(f"OC detectada: {analysis.get('purchase_order')}")
    report_lines.append(f"Total codigos: {len(codes)}; Productos base agrupados: {len(groups)}")
    anchor_base = None
    anchor_score = 0.0
    po = analysis.get("purchase_order")
    anchor_by_series: Dict[str, Tuple[str, Path, float]] = {}
    anchor_global: Optional[Tuple[str, Path, float]] = None
    for base, parts in groups.items():
        path, score, detail = find_product_pdf_with_score(po, base, candidates)
        base_series = extract_series_token(base) or ""
        anchor_entry = anchor_by_series.get(base_series)
        if base.startswith("B&P-00"):
            # ancla por serie si existe
            if anchor_entry:
                anchor_base, anchor_path, anchor_score = anchor_entry
                if score < max(anchor_score - 4.0, 14.0):
                    report_lines.append(
                        f"{base} parts[{', '.join(parts)}] -> ANCLADO A {anchor_base} ({anchor_path}) orig_score={score:.2f} detail={detail} (formula map usado)"
                    )
                    continue
            # si no hay serie o no hay ancla de serie, usar ancla global
            if not base_series and anchor_global:
                g_base, g_path, g_score = anchor_global
                if score < max(g_score - 4.0, 14.0):
                    report_lines.append(
                        f"{base} parts[{', '.join(parts)}] -> ANCLADO A {g_base} ({g_path}) orig_score={score:.2f} detail={detail} (formula map usado)"
                    )
                    continue
        line = f"{base} parts[{', '.join(parts)}] -> {path if path else 'NO_ENCONTRADO'} score={score:.2f} detail={detail}"
        report_lines.append(line)
        if score >= 18.0 and path:
            anchor_by_series[base_series] = (base, path, score)
            if not anchor_global:
                anchor_global = (base, path, score)
    out_path = cfg.output_root / "mapping_report.txt"
    out_path.write_text("\n".join(report_lines), encoding="utf-8")
    logging.info("Reporte generado: %s", out_path)


def process_input_file(input_path: Path, cfg: Config) -> None:
    logging.info("Procesando archivo: %s", input_path.name)
    # Sin Gemini: usar solo parser local
    items = []
    if input_path.suffix.lower() == ".pdf":
        items = parse_codes_po_local(input_path, cfg)
    if not items:
        raise ValueError("No se detectaron codigos/PO.")

    def normalize_code(code: str) -> str:
        return (code or "").strip().upper()

    def normalize_po(po: str) -> str:
        if not po:
            return ""
        digits = re.findall(r"\d+", po)
        if digits:
            return digits[0]
        return po.strip()

    items_norm = []
    for it in items:
        items_norm.append(
            {
                "code": normalize_code(str(it.get("code") or "")),
                "po": normalize_po(str(it.get("po") or "")),
                "obs": str(it.get("obs") or "").strip(),
            }
        )
    items = items_norm

    # Comparar con el último resumen para evitar duplicados
    def canonical_rows(elems):
        rows = []
        for it in elems:
            obs_val = it.get("obs", "") or ""
            m = re.search(r"\d+", obs_val)
            obs_clean = m.group(0) if m else obs_val.strip()
            fecha = str(it.get("date", "") or "").strip()
            rows.append((it.get("code", ""), it.get("po", ""), obs_clean, fecha))
        return rows

    new_rows = canonical_rows(items)
    last_resumen = find_last_resumen(cfg.processed_dir)
    if last_resumen:
        prev_rows = load_resumen_rows(last_resumen)
        if prev_rows == new_rows:
            logging.info("Resumen idéntico al anterior (%s); no se crea nuevo registro.", last_resumen)
            return

    timestamp = dt.datetime.now()
    seq_id = next_sequential_id((cfg.incoming_dir / "Auxiliar") / ".registro_seq.txt")
    registro_dir = cfg.processed_dir / f"Registro - {seq_id} - {timestamp:%Y-%m-%d}"
    registro_dir.mkdir(parents=True, exist_ok=True)
    registro_file = registro_dir / input_path.name
    shutil.move(str(input_path), registro_file)
    csv_path = registro_dir / "resumen.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Codigo", "PO", "Observac_PO", "Fecha"])
        for it in items:
            obs_val = it.get("obs", "") or ""
            m = re.search(r"\d+", obs_val)
            obs_clean = m.group(0) if m else obs_val.strip()
            fecha = str(it.get("date", "") or "").strip()
            writer.writerow([it.get("code", ""), it.get("po", ""), obs_clean, fecha])
        # Coincidencias Auxiliar se generan ahora en el paso 3 (JSON)


def handle_file(input_path: Path, cfg: Config) -> None:
    try:
        process_input_file(input_path, cfg)
    except Exception as exc:
        logging.error("Error procesando %s: %s", input_path.name, exc)
        cfg.failed_dir.mkdir(parents=True, exist_ok=True)
        timestamp = dt.datetime.now()
        seq_id = next_sequential_id((cfg.incoming_dir / "Auxiliar") / ".registro_seq.txt")
        registro_dir = cfg.failed_dir / f"Registro - {seq_id} - {timestamp:%Y-%m-%d}"
        registro_dir.mkdir(parents=True, exist_ok=True)
        failed_dest = registro_dir / input_path.name
        try:
            shutil.move(str(input_path), failed_dest)
            # CSV con error
            csv_path = registro_dir / "resumen.csv"
            with open(csv_path, "w", newline="", encoding="utf-8") as f:
                writer = csv.writer(f, delimiter=";")
                writer.writerow(["Codigo", "PO", "Error"])
                writer.writerow(["", "", str(exc)])
            logging.info("Archivo movido a failed: %s y CSV resumen: %s", failed_dest, csv_path)
        except Exception:
            logging.error("No se pudo mover a failed: %s", failed_dest)




def extract_codes_po_with_gemini_multi(paths, cfg) -> list:
    """Envia multiples imagenes/PDF a Gemini y devuelve items combinados."""
    api_key = os.getenv("GEMINI_API_KEY") or (Path("gemini.key").read_text().strip() if Path("gemini.key").exists() else None)
    if not api_key:
        logging.info("GEMINI_API_KEY no definido; se omite extraccion Gemini.")
        return []
    parts = [{"text": "Extrae columnas Producto (codigo), OC (Purchase Order), Observac (columna 9) y Fecha (columna 4). La OC y Observac deben ser SOLO numeros (si Observac dice PO#1053 ESUN devuelve 1053). La Fecha debe venir tal cual aparece en la tabla. Devuelve JSON array: [{\"code\":\"...\",\"po\":\"...\",\"obs\":\"solo numero observac\",\"date\":\"fecha\"}]"}]
    for path in paths:
        suffix = path.suffix.lower()
        mime = None
        data = None
        if suffix in {".png", ".jpg", ".jpeg", ".tif", ".tiff", ".bmp"}:
            res = load_image_bytes(path)
            if res:
                data, mime = res
        elif suffix == ".pdf":
            imgs = render_pdf_pages_png(path, max_pages=cfg.pdf_max_pages)
            if imgs:
                data, mime = imgs[0], "image/png"
        if data and mime:
            parts.append({"inline_data": {"mime_type": mime, "data": base64.b64encode(data).decode()}})
    if len(parts) == 1:
        logging.warning("Sin imagenes para Gemini.")
        return []
    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.5-flash:generateContent?key={api_key}"
    try:
        resp = requests.post(url, json={"contents": [{"parts": parts}]}, timeout=120)
        if resp.status_code != 200:
            logging.warning("Gemini status %s: %s", resp.status_code, resp.text[:400])
            return []
        data = resp.json()
        text_resp = data["candidates"][0]["content"]["parts"][0]["text"]
        m = re.search(r"\[.*\]", text_resp, re.S)
        payload = json.loads(m.group(0)) if m else json.loads(text_resp)
        return payload if isinstance(payload, list) else []
    except Exception as exc:
        logging.warning("Fallo Gemini: %s", exc)
        return []


def extract_codes_po_with_gemini(path: Path, cfg: Config) -> list:
    """Wrapper simple para un solo archivo usando la lógica multi."""
    return extract_codes_po_with_gemini_multi([path], cfg)


def process_group(files, cfg: Config) -> None:
    logging.info("Procesando grupo de %s archivos", len(files))
    merged = []
    for path in files:
        if path.suffix.lower() == ".pdf":
            merged.extend(parse_codes_po_local(path, cfg))
        else:
            logging.warning("Sin Gemini: se omite archivo no-PDF %s", path.name)
    items = merged
    if not items:
        raise ValueError("No se detectaron codigos/PO en grupo.")
    def normalize_code(code: str) -> str:
        return (code or "").strip().upper()
    def normalize_po(po: str) -> str:
        if not po:
            return ""
        digits = re.findall(r"\d+", po)
        if digits:
            return digits[0]
        return po.strip()
    items_norm = []
    for it in items:
        items_norm.append({"code": normalize_code(str(it.get("code") or "")),
                           "po": normalize_po(str(it.get("po") or "")),
                           "obs": str(it.get("obs") or "").strip()})
    # Comparar con ultimo resumen
    def canonical_rows(elems):
        rows = []
        for it in elems:
            obs_val = it.get("obs", "") or ""
            m = re.search(r"\d+", obs_val)
            obs_clean = m.group(0) if m else obs_val.strip()
            fecha = str(it.get("date", "") or "").strip()
            rows.append((it.get("code",""), it.get("po",""), obs_clean, fecha))
        return rows

    new_rows = canonical_rows(items_norm)
    last_resumen = find_last_resumen(cfg.processed_dir)
    if last_resumen:
        prev_rows = load_resumen_rows(last_resumen)
        if prev_rows == new_rows:
            logging.info("Resumen idéntico al anterior (%s); no se crea nuevo registro.", last_resumen)
            return

    timestamp = dt.datetime.now()
    seq_id = next_sequential_id((cfg.incoming_dir / "Auxiliar") / ".registro_seq.txt")
    registro_dir = cfg.processed_dir / f"Registro - {seq_id} - {timestamp:%Y-%m-%d}"
    registro_dir.mkdir(parents=True, exist_ok=True)
    csv_path = registro_dir / "resumen.csv"
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerow(["Codigo", "PO", "Observac_PO", "Fecha"])
        for it in items_norm:
            obs_val = it.get("obs", "") or ""
            m = re.search(r"\d+", obs_val)
            obs_clean = m.group(0) if m else obs_val.strip()
            fecha = str(it.get("date", "") or "").strip()
            writer.writerow([it.get("code",""), it.get("po",""), obs_clean, fecha])
    for path in files:
        dest = registro_dir / path.name
        dest.parent.mkdir(parents=True, exist_ok=True)
        path.rename(dest)
    logging.info("Archivos movidos a in process: %s y CSV resumen: %s", registro_dir, csv_path)

def scan_and_process(cfg: Config) -> None:
    files = sorted([p for p in cfg.incoming_dir.iterdir() if p.suffix.lower() in SUPPORTED_EXTS])
    if not files:
        return
    if len(files)>1 and all(p.suffix.lower() in {".png",".jpg",".jpeg",".tif",".tiff",".bmp"} for p in files):
        process_group(files, cfg)
    else:
        for p in files:
            handle_file(p, cfg)


def main() -> None:
    parser = argparse.ArgumentParser(description="Procesa PDF y actualiza registros.")
    parser.add_argument("--config", default="config.yaml", help="Ruta al config.yaml")
    parser.add_argument("--once", action="store_true", help="Procesa PDFs pendientes y termina")
    parser.add_argument("--watch", action="store_true", help="Vigila la carpeta de entrada")
    parser.add_argument("--report", action="store_true", help="Solo genera reporte de mapeo sin copiar/mover archivos")
    args = parser.parse_args()

    cfg_path = Path(args.config).expanduser()
    cfg = load_config(cfg_path)
    setup_logging(cfg.log_file, cfg.log_level)
    ensure_dependencies()
    ensure_dirs(cfg)

    if args.report:
        generate_report(cfg)
        return

    if args.once or not args.watch:
        scan_and_process(cfg)
        return

    logging.info("Vigilando %s cada %ss", cfg.incoming_dir, cfg.watch_scan_seconds)
    while True:
        scan_and_process(cfg)
        time.sleep(cfg.watch_scan_seconds)


if __name__ == "__main__":
    main()











