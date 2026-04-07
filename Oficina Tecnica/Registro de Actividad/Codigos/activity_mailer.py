import argparse
import csv
import difflib
import email
import html
import imaplib
import json
import math
import os
import re
import sys
import time
import unicodedata
import uuid
from copy import copy
from datetime import date, datetime
from email import policy
from email.header import decode_header
from typing import Dict, List, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter

# Ruta del proyecto existente para reutilizar configuraciones de correo y feriados.
EXTERNAL_CODE_DIR = r"\\192.168.0.13\lcontigiani\Proyecto Costos\Codigo\Codigos"
if EXTERNAL_CODE_DIR not in sys.path:
    sys.path.insert(0, EXTERNAL_CODE_DIR)

# Permitir feriados adicionales locales (si el archivo no existe se ignora).
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DEFAULT_EXTRA_HOLIDAYS = os.path.join(BASE_DIR, "feriados_adicionales.csv")
if "HOLIDAYS_PATH" not in os.environ:
    os.environ["HOLIDAYS_PATH"] = DEFAULT_EXTRA_HOLIDAYS

from config import SMTP_DEFAULTS  # type: ignore
import utils  # type: ignore

CONFIG_DIR = os.path.join(BASE_DIR, "config")
DATA_DIR = os.path.join(BASE_DIR, "data")
LOG_DIR = os.path.join(BASE_DIR, "logs")
STATE_PATH = os.path.join(DATA_DIR, "activity_mailer_state.json")
RECIPIENTS_FILE = os.path.join(CONFIG_DIR, "destinatarios.csv")
PROFILES_PATH = os.path.join(CONFIG_DIR, "activity_mailer_profiles.json")
SUBJECT_TEMPLATE_FILE = os.path.join(CONFIG_DIR, "mail_subject.txt")
BODY_TEMPLATE_FILE = os.path.join(CONFIG_DIR, "mail_body.txt")
GLOSSARY_PATH = os.path.join(CONFIG_DIR, "glosario_proyectos.log")
RESPONSES_DIR = os.path.join(DATA_DIR, "responses")
EXCEL_PATH = os.path.abspath(os.path.join(BASE_DIR, os.pardir, "Oficina Tecnica Registros.xlsx"))
BASE_DATA_CSV = os.path.join(DATA_DIR, "base_datos_respuestas.csv")
USER_CSV_DIR = os.path.join(DATA_DIR, "respuestas_csv")
LOG_PATH = os.path.join(LOG_DIR, "activity_mailer.log")
_BODY_TEMPLATE_CACHE: Optional[str] = None
utils.LOG_PATH = LOG_PATH
from utils import is_business_day, load_holidays, log, parse_recipients  # type: ignore
import email_utils  # type: ignore

# Asegurar estructura de carpetas por orden
for _path in (CONFIG_DIR, DATA_DIR, LOG_DIR, RESPONSES_DIR, USER_CSV_DIR):
    os.makedirs(_path, exist_ok=True)
if not os.path.isfile(GLOSSARY_PATH):
    try:
        with open(GLOSSARY_PATH, "w", encoding="utf-8") as f:
            f.write("# Glosario de proyectos (una linea por proyecto)\n")
    except Exception:
        pass

# Horarios de envio: lunes-viernes 10:40 (lunes=0).
SCHEDULE_HOUR = {
    0: (17, 0),
    1: (17, 0),
    2: (17, 0),
    3: (17, 0),
    4: (16, 0),
}

TOKEN_PREFIX = "REGISTRO"
TOKEN_LABEL = f"{TOKEN_PREFIX}:"
IMAP_CHECK_INTERVAL_SECONDS = 90
MAX_CAMPAIGNS_IN_STATE = 30
LOCK_PATH = os.path.join(DATA_DIR, "activity_mailer.lock")
# Fechas bloqueadas manualmente: no se deben volver a persistir en el estado.
BLOCKED_CAMPAIGN_DATES = {"2026-02-16", "2026-02-17"}
_extra_blocked_dates = os.environ.get("ACTIVITY_BLOCKED_DATES", "").strip()
if _extra_blocked_dates:
    for _raw_day in _extra_blocked_dates.split(","):
        _day = _raw_day.strip()
        if _day:
            BLOCKED_CAMPAIGN_DATES.add(_day)
FORCE_PROFILE = None  # Forzar perfil fijo. Poner None para respetar el perfil marcado.
MANROPE_FONT = Font(name="Manrope", size=11)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)
ALIGN_WRAP_CENTER = Alignment(horizontal="left", vertical="center", wrap_text=True)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
TABLE_HEADERS = [
    "Fecha",
    "Hora",
    "Token",
    "Registro",
    "Proyectos",
    "Tiempo",
    "Observaciones",
]

# Columnas fijas en Excel (1-based)
COL_FECHA = 3      # C
COL_HORA = 4       # D
COL_TOKEN = 5      # E
COL_REGISTRO = 6   # F
COL_PROJ1 = 7      # G
COL_TIME1 = 8      # H
COL_PROJ2 = None
COL_TIME2 = None
COL_PROJ3 = None
COL_TIME3 = None
COL_OTROS = None


def _is_blocked_campaign_date(date_str: Optional[str]) -> bool:
    return bool(date_str) and date_str in BLOCKED_CAMPAIGN_DATES


def _filter_blocked_campaigns(campaigns: List[Dict]) -> List[Dict]:
    filtered: List[Dict] = []
    for camp in campaigns or []:
        if not isinstance(camp, dict):
            continue
        if _is_blocked_campaign_date(camp.get("date")):
            continue
        filtered.append(camp)
    return filtered


def ensure_state_struct(state: Optional[Dict]) -> Dict:
    if not isinstance(state, dict):
        state = {}
    state["campaigns"] = _filter_blocked_campaigns(state.get("campaigns", []))
    state.setdefault("imap", {})
    state["imap"].setdefault("last_uid", 0)
    state["imap"].setdefault("seen_uids", [])
    return state


def acquire_lock(max_age_seconds: int = 7200) -> bool:
    """
    Evita ejecuciones simultaneas que pisan el JSON.
    Si el lock tiene mucha antiguedad se limpia (protege de locks huerfanos).
    """
    try:
        if os.path.exists(LOCK_PATH):
            age = time.time() - os.path.getmtime(LOCK_PATH)
            if age < max_age_seconds:
                log(f"Otro proceso del mailer ya esta en ejecucion (lock en {LOCK_PATH}). Saliendo.", "WARN")
                return False
            else:
                os.remove(LOCK_PATH)
        fd = os.open(LOCK_PATH, os.O_CREAT | os.O_EXCL | os.O_RDWR)
        with os.fdopen(fd, "w") as f:
            f.write(str(os.getpid()))
        return True
    except FileExistsError:
        log(f"Lock detectado en {LOCK_PATH}. Saliendo para evitar pisar estado.", "WARN")
        return False
    except Exception as exc:
        log(f"No se pudo adquirir lock {LOCK_PATH}: {exc}", "WARN")
        return False


def release_lock():
    try:
        if os.path.exists(LOCK_PATH):
            os.remove(LOCK_PATH)
    except Exception:
        pass


def _default_profiles() -> Dict:
    return {
        "active_profile": "test",
        "profiles": {
            "test": {
                "description": "Perfil de pruebas",
                "recipients": [
                    {"name": "Lorenzo Contigiani", "email": "lcontigiani@bpbargentina.com"},
                ],
            },
            "real": {
                "description": "Perfil real (usa destinatarios.csv)",
                "use_csv": True,
            },
        },
    }


def load_profiles() -> Dict:
    if not os.path.isfile(PROFILES_PATH):
        try:
            os.makedirs(os.path.dirname(PROFILES_PATH), exist_ok=True)
            with open(PROFILES_PATH, "w", encoding="utf-8") as f:
                json.dump(_default_profiles(), f, indent=2, ensure_ascii=False)
        except Exception as exc:
            log(f"No se pudo crear archivo de perfiles {PROFILES_PATH}: {exc}", "WARN")
            return _default_profiles()
    try:
        with open(PROFILES_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return _default_profiles()
        if "active_profile" not in data or "profiles" not in data:
            return _default_profiles()
        return data
    except Exception as exc:
        log(f"No se pudo leer perfiles desde {PROFILES_PATH}: {exc}", "WARN")
        return _default_profiles()


def _recipients_from_profile(profile: Dict) -> List[Dict[str, Optional[str]]]:
    result: List[Dict[str, Optional[str]]] = []
    recs = profile.get("recipients")
    if isinstance(recs, list):
        for item in recs:
            if isinstance(item, dict):
                email_addr = (item.get("email") or "").strip()
                name = (item.get("name") or "").strip() or None
                if email_addr:
                    result.append({"email": email_addr, "name": name})
            elif isinstance(item, str):
                email_addr = item.strip()
                if email_addr:
                    result.append({"email": email_addr, "name": None})
    return result


def load_state() -> Dict:
    if not os.path.isfile(STATE_PATH):
        return ensure_state_struct({})
    try:
        with open(STATE_PATH, "r", encoding="utf-8-sig") as f:
            data = json.load(f)
        return ensure_state_struct(data)
    except Exception as exc:
        log(f"No se pudo cargar estado {STATE_PATH}: {exc}", "WARN")
        return ensure_state_struct({})


def save_state(state: Dict):
    try:
        # Merge defensively with latest on disk to avoid clobbering concurrent writers (web, etc.)
        def merge_states(old: Dict, new: Dict) -> Dict:
            merged = ensure_state_struct(old or {})
            new = ensure_state_struct(new or {})
            merged.setdefault("campaigns", [])
            idx = {c.get("token"): c for c in merged.get("campaigns", []) if c.get("token")}
            for camp in _filter_blocked_campaigns(new.get("campaigns", [])):
                token = camp.get("token")
                if token in idx:
                    dest = idx[token]
                    for k in ["date", "sent_at", "subject"]:
                        if camp.get(k):
                            dest[k] = camp[k]
                    if camp.get("recipients"):
                        dest["recipients"] = camp["recipients"]
                    dest.setdefault("responses", {})
                    for email_addr, resp in camp.get("responses", {}).items():
                        dest["responses"][email_addr] = resp
                else:
                    merged["campaigns"].append(camp)

            def _campaign_sort_key(c: Dict):
                # Prefer sent_at if present, fallback to date, else minimal
                ts = c.get("sent_at") or ""
                dt_str = ts or c.get("date") or ""
                try:
                    return datetime.fromisoformat(dt_str.replace("Z", ""))
                except Exception:
                    try:
                        return datetime.strptime(dt_str, "%Y-%m-%d")
                    except Exception:
                        return datetime.min

            # Dedup by token (keep latest info) and order by most recent
            deduped = {}
            for camp in _filter_blocked_campaigns(merged.get("campaigns", [])):
                token = camp.get("token")
                if not token:
                    continue
                deduped[token] = camp
            ordered = sorted(deduped.values(), key=_campaign_sort_key, reverse=True)
            if MAX_CAMPAIGNS_IN_STATE is not None:
                ordered = ordered[:MAX_CAMPAIGNS_IN_STATE]
            merged["campaigns"] = _filter_blocked_campaigns(ordered)
            merged.setdefault("imap", {})
            merged["imap"].update(new.get("imap", {}) or {})
            return merged

        try:
            with open(STATE_PATH, "r", encoding="utf-8-sig") as f:
                current_disk = json.load(f)
        except Exception:
            current_disk = {}

        merged_state = merge_states(current_disk, state)

        os.makedirs(os.path.dirname(STATE_PATH), exist_ok=True)
        with open(STATE_PATH, "w", encoding="utf-8") as f:
            json.dump(merged_state, f, indent=2, ensure_ascii=False)
    except Exception as exc:
        log(f"No se pudo guardar estado {STATE_PATH}: {exc}", "WARN")


def _decode_header(value: Optional[str]) -> str:
    if not value:
        return ""
    try:
        parts = decode_header(value)
        decoded = ""
        for text, enc in parts:
            if isinstance(text, bytes):
                decoded += text.decode(enc or "utf-8", errors="ignore")
            else:
                decoded += text
        return decoded
    except Exception:
        return str(value)


def _extract_plain_text(msg: email.message.Message) -> str:
    def html_to_text(ht: str) -> str:
        no_scripts = re.sub(r"(?is)<(script|style).*?>.*?(</\\1>)", "", ht)
        no_tags = re.sub(r"(?s)<[^>]+>", " ", no_scripts)
        unescaped = html.unescape(no_tags)
        return re.sub(r"\\s+", " ", unescaped).strip()

    html_candidate = ""

    if msg.is_multipart():
        for part in msg.walk():
            ctype = part.get_content_type()
            if ctype == "text/plain" and not part.get_filename():
                try:
                    payload = part.get_payload(decode=True)
                    charset = part.get_content_charset() or "utf-8"
                    return (payload or b"").decode(charset, errors="ignore")
                except Exception:
                    continue
            if ctype == "text/html" and not part.get_filename() and not html_candidate:
                try:
                    payload = part.get_payload(decode=True)
                    charset = part.get_content_charset() or "utf-8"
                    html_candidate = (payload or b"").decode(charset, errors="ignore")
                except Exception:
                    continue
    else:
        try:
            payload = msg.get_payload(decode=True)
            charset = msg.get_content_charset() or "utf-8"
            if msg.get_content_type() == "text/plain":
                return (payload or b"").decode(charset, errors="ignore")
            if msg.get_content_type() == "text/html":
                html_candidate = (payload or b"").decode(charset, errors="ignore")
        except Exception:
            return ""

    if html_candidate:
        return html_to_text(html_candidate)
    return ""


def _safe_filename(text: str) -> str:
    cleaned = re.sub(r"[^A-Za-z0-9._-]", "_", text)
    return cleaned.strip("_") or "respuesta"


def _extract_reply_text(body: str) -> str:
    if not body:
        return ""
    # Quitar tokens completos del prompt si vinieran en el cuerpo
    body = re.sub(rf"Token:\s*{TOKEN_PREFIX}:[A-Za-z0-9\-]+", "", body, flags=re.IGNORECASE)
    lines = body.splitlines()
    result_lines = []
    # Patrones de encabezados de respuesta (ingles/espanol)
    header_patterns = [
        re.compile(r"^on\s+", re.IGNORECASE),
        re.compile(r"^el\s+(lun|mar|mie|miÃ©|jue|vie|sab|sÃ¡b|dom)", re.IGNORECASE),
        re.compile(r"^de:", re.IGNORECASE),
        re.compile(r"^from:", re.IGNORECASE),
        re.compile(r"^wrote:", re.IGNORECASE),
        re.compile(r"escribi[oÃ³]", re.IGNORECASE),
    ]
    for line in lines:
        stripped = line.strip()
        lower = stripped.lower()
        # Cortar en citados y encabezados de respuesta
        if stripped.startswith(">"):
            break
        if any(p.search(stripped) for p in header_patterns):
            break
        if "original message" in lower or "mensaje original" in lower:
            break
        result_lines.append(line)
    cleaned = "\n".join(result_lines)
    # Cortar firma si aparece "--" en una linea sola
    signature_pos = cleaned.find("\n--")
    if signature_pos != -1:
        cleaned = cleaned[:signature_pos]
    return cleaned.strip()


def _strip_signature(text: str) -> str:
    """
    Remueve firmas comunes (celular, líneas rotativas, enviado desde, dominios, etc.)
    para no tomar la firma como parte de la respuesta.
    """
    markers = [
        r"^--+$",
        r"^__+$",
        r"^Enviado desde",
        r"^Enviado de",
        r"^From:",
        r"^De:",
        r"^On .+wrote:",
        r"^El .+ a la",
        r"Celular:",
        r"L[ií]neas rotativas",
        r"bpbargentina\.com",
        r"Oficina T[eé]cnica",
    ]
    regex = re.compile("|".join(markers), re.IGNORECASE)
    kept = []
    for line in text.splitlines():
        if regex.search(line):
            break
        kept.append(line)
    return "\n".join(kept).strip()


def _normalize_project_token(text: str) -> str:
    if not text:
        return ""
    txt = unicodedata.normalize("NFD", text)
    txt = "".join(ch for ch in txt if unicodedata.category(ch) != "Mn")
    txt = re.sub(r"[^a-z0-9]+", " ", txt.lower())
    return " ".join(txt.split())


def _load_project_glossary() -> List[Dict[str, List[str]]]:
    entries: List[Dict[str, List[str]]] = []
    if os.path.isfile(GLOSSARY_PATH):
        try:
            with open(GLOSSARY_PATH, "r", encoding="utf-8") as f:
                for line in f:
                    ln = line.strip()
                    if not ln or ln.startswith("#"):
                        continue
                    if "-" in ln:
                        main, rest = ln.split("-", 1)
                    else:
                        main, rest = ln, ""
                    main = main.strip()
                    aliases: List[str] = []
                    if rest.strip():
                        for part in re.split(r"[;,|]", rest):
                            val = part.strip()
                            if val:
                                aliases.append(val)
                    entries.append({"main": main, "aliases": aliases})
        except Exception:
            pass
    return entries


def _match_project_name(name: str, glossary: Optional[List[Dict[str, List[str]]]] = None, threshold: float = 0.7) -> str:
    raw = (name or "").strip()
    if not raw:
        return ""
    glossary = glossary or _load_project_glossary()
    if not glossary:
        return raw
    norm_name = _normalize_project_token(raw)
    best = raw
    best_score = 0.0
    for entry in glossary:
        main = entry.get("main") or ""
        candidates = [main] + entry.get("aliases", [])
        for cand in candidates:
            norm_item = _normalize_project_token(cand)
            if not norm_item:
                continue
            score = difflib.SequenceMatcher(None, norm_name, norm_item).ratio()
            if score > best_score:
                best_score = score
                best = main
    return best if best_score >= threshold else raw

def _match_project_with_flag(name: str, glossary: Optional[List[Dict[str, List[str]]]] = None, threshold: float = 0.7):
    """Devuelve (nombre_match, True) si matchea con el glosario; caso contrario (nombre_original, False)."""
    raw = (name or "").strip()
    if not raw:
        return "", False
    glossary = glossary or _load_project_glossary()
    if not glossary:
        return raw, False
    norm_name = _normalize_project_token(raw)
    best = raw
    best_score = 0.0
    for entry in glossary:
        main = entry.get("main") or ""
        candidates = [main] + entry.get("aliases", [])
        for cand in candidates:
            norm_item = _normalize_project_token(cand)
            if not norm_item:
                continue
            score = difflib.SequenceMatcher(None, norm_name, norm_item).ratio()
            if score > best_score:
                best_score = score
                best = main
    matched = best_score >= threshold
    return (best if matched else raw), matched


def _parse_time_to_hours(raw: str) -> Optional[float]:
    if not raw:
        return None
    txt = raw.strip().lower().replace(",", ".")
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(hora|horas|hr|hrs|h)\b", txt, flags=re.IGNORECASE)
    if m:
        try:
            return float(m.group(1))
        except Exception:
            return None
    m = re.search(r"([0-9]+(?:\.[0-9]+)?)\s*(min|mins|minuto|minutos|min|m)\b", txt, flags=re.IGNORECASE)
    if m:
        try:
            return float(m.group(1)) / 60.0
        except Exception:
            return None
    try:
        return float(txt)
    except Exception:
        return None


def _parse_projects_v2(body: str):
    lines = body.splitlines()
    segments = []
    free_lines = []
    current = None
    current_field = None
    desc_parts: List[str] = []

    def flush():
        nonlocal current
        if current and any(v.strip() for v in current.values()):
            tiempo_raw = current.get("tiempo", "").strip()
            tiempo_hours = _parse_time_to_hours(tiempo_raw)
            segments.append(
                {
                    "proyecto": current.get("proyecto", "").strip(),
                    "tiempo": tiempo_hours,
                    "tiempo_raw": tiempo_raw,
                    "descripcion": current.get("descripcion", "").strip(),
                }
            )
            if current.get("descripcion", "").strip():
                desc_parts.append(current["descripcion"].strip())
        current = None

    for line in lines:
        stripped = line.strip()
        lower = stripped.lower()
        if lower.startswith("proyecto:"):
            flush()
            current = {"proyecto": stripped[len("proyecto:") :].strip(), "tiempo": "", "descripcion": ""}
            current_field = None
            continue
        if lower.startswith("tiempo") and "destinado" in lower:
            if current is None:
                current = {"proyecto": "", "tiempo": "", "descripcion": ""}
            current["tiempo"] = stripped.split(":", 1)[-1].strip()
            current_field = None
            continue
        if re.match(r"descripci\S*:", lower):
            if current is None:
                current = {"proyecto": "", "tiempo": "", "descripcion": ""}
            current_field = "descripcion"
            tail = stripped.split(":", 1)[-1].strip()
            if tail:
                current["descripcion"] = (current.get("descripcion", "") + ("\n" if current.get("descripcion") else "") + tail).strip()
            continue
        if stripped == "-" and current:
            flush()
            current_field = None
            continue
        if current_field == "descripcion" and current is not None:
            if stripped:
                current["descripcion"] = (current.get("descripcion", "") + ("\n" if current.get("descripcion") else "") + stripped).strip()
            continue
        # Linea libre
        free_lines.append(line)

    flush()
    free_text_lines = [ln for ln in free_lines if ln.strip()]
    desc_text = "\n\n".join([p for p in desc_parts if p.strip()] + free_text_lines)
    return desc_text, segments


def _ensure_excel_sheet():
    if not os.path.isfile(EXCEL_PATH):
        wb = Workbook()
        # Crear hoja Base de datos con encabezados
        ws_bd = wb.active
        ws_bd.title = "Base de datos"
        ws_bd.append(["Fecha", "Hora", "Proyecto", "Tiempo", "Descripcion", "Proyecto2", "Tiempo2", "Descripcion2"])
        wb.save(EXCEL_PATH)
        return
    try:
        wb = None
        for attempt in range(5):
            try:
                wb = load_workbook(EXCEL_PATH)
                break
            except PermissionError:
                if attempt == 4:
                    log(f"Excel en uso, no se pudo abrir {EXCEL_PATH} tras 5 intentos.", "WARN")
                    return
                time.sleep(30)
            except Exception as exc:
                if attempt == 4:
                    log(f"No se pudo abrir Excel {EXCEL_PATH}: {exc}", "WARN")
                    return
                time.sleep(2)
        if wb is None:
            return
        # Asegurar hoja Base de datos
        if "Base de datos" not in wb.sheetnames:
            ws_bd = wb.create_sheet("Base de datos")
            ws_bd.append(["Fecha", "Hora", "Proyecto", "Tiempo", "Descripcion", "Proyecto2", "Tiempo2", "Descripcion2"])
            wb.save(EXCEL_PATH)
        wb.close()
    except Exception as exc:
        log(f"No se pudo preparar Excel {EXCEL_PATH}: {exc}", "WARN")


def _find_header(ws):
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 30)):
        values = [str(c.value).strip().lower() if c.value is not None else "" for c in row]
        if "fecha" in values and "hora" in values and "token" in values:
            headers = {}
            proj_count = 0
            time_count = 0
            for idx, val in enumerate(values, start=1):
                val_lower = val.lower()
                val_norm = re.sub(r"[^a-z0-9]", "", val_lower)
                if val in {"fecha", "hora", "token", "registro", "observaciones"}:
                    headers[val] = idx
                if val_norm.startswith("proyectos") or val_norm.startswith("proyecto"):
                    proj_count += 1
                    headers[f"proyectos{proj_count}"] = idx
                    continue
                if (
                    "tiempo" in val_norm
                    or "[h" in val_lower
                    or " h" in val_lower
                    or "hora" in val_norm
                    or "horas" in val_norm
                ):
                    time_count += 1
                    key = "tiempo" if time_count == 1 else f"tiempo{time_count}"
                    headers[key] = idx
                if "otros" in val_norm:
                    headers["otros"] = idx
            return row[0].row, headers
    return None, {}


def _first_empty_row(ws, start_row: int, header_cols: Dict[str, int]) -> int:
    max_row = ws.max_row or start_row
    for idx in range(start_row, max_row + 20):
        row_vals = []
        for key in ("fecha", "hora", "token", "registro"):
            col = header_cols.get(key)
            if col:
                row_vals.append(ws.cell(row=idx, column=col).value)
        if row_vals and all(v in (None, "") for v in row_vals):
            return idx
    return max_row + 1


def _copy_row_style(ws, source_row: int, target_row: int):
    try:
        src_h = ws.row_dimensions[source_row].height
        ws.row_dimensions[target_row].height = src_h if src_h is not None else 30
    except Exception:
        pass
    for col in range(1, ws.max_column + 1):
        src = ws.cell(row=source_row, column=col)
        tgt = ws.cell(row=target_row, column=col)
        tgt._style = copy(src._style)
        tgt.number_format = src.number_format
        # Forzar fuente Manrope y alineaciones en columnas claves
        if src.value is not None:
            try:
                tgt.font = Font(name="Manrope", size=src.font.size or 11, bold=src.font.bold, italic=src.font.italic)
            except Exception:
                tgt.font = src.font


def _apply_row_format(ws, row_idx: int, header_cols: Dict[str, int]):
    def _set_alignment(cell, **kwargs):
        current = cell.alignment or Alignment()
        cell.alignment = Alignment(
            horizontal=kwargs.get("horizontal", current.horizontal),
            vertical=kwargs.get("vertical", current.vertical),
            wrap_text=kwargs.get("wrap_text", current.wrap_text),
            text_rotation=current.text_rotation,
            shrink_to_fit=current.shrink_to_fit,
            indent=current.indent,
        )

    cols_center = []
    for key in ("fecha", "hora", "token", "tiempo"):
        c = header_cols.get(key)
        if c:
            cols_center.append(c)
    for col in cols_center:
        cell = ws.cell(row=row_idx, column=col)
        _set_alignment(cell, horizontal="center", vertical="center", wrap_text=True)
        cell.font = MANROPE_FONT
    reg_col = header_cols.get("registro")
    if reg_col:
        cell = ws.cell(row=row_idx, column=reg_col)
        _set_alignment(cell, horizontal="left", vertical="center", wrap_text=True)
        cell.font = MANROPE_FONT
    proj_col = header_cols.get("proyectos1")
    if proj_col:
        cell = ws.cell(row=row_idx, column=proj_col)
        _set_alignment(cell, horizontal="center", vertical="center", wrap_text=True)
        cell.font = MANROPE_FONT
    # Bordes finos en toda la fila (solo columnas que usamos)
    used_cols = sorted(set(header_cols.values()))
    for col in used_cols:
        ws.cell(row=row_idx, column=col).border = THIN_BORDER


def _adjust_row_height(ws, row_idx: int, header_cols: Dict[str, int]):
    """Ajusta la altura de fila segÃºn texto en Registro y Proyectos."""
    cols_to_check = []
    for key in ("registro", "proyectos1", "tiempo"):
        col = header_cols.get(key)
        if col:
            cols_to_check.append(col)
    if not cols_to_check:
        return

    max_lines_needed = 1
    for col in cols_to_check:
        val = ws.cell(row=row_idx, column=col).value
        if not val:
            continue
        text = str(val)
        lines = text.split("\n")
        letter = get_column_letter(col)
        col_width = ws.column_dimensions[letter].width or 20
        max_chars_per_line = max(1, int(col_width * 0.9))
        est_lines = 0
        for ln in lines:
            est_lines += max(1, math.ceil(len(ln) / max_chars_per_line))
        max_lines_needed = max(max_lines_needed, len(lines), est_lines)

    target_height = min(300, max(30, max_lines_needed * 18))
    try:
        ws.row_dimensions[row_idx].height = target_height
    except Exception:
        pass


def _safe_sheet_name(name: str) -> str:
    cleaned = re.sub(r"[\\/:*?\\[\\]]", "_", (name or "").strip())
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned or "Respuestas"


def _append_response_to_excel(campaign: Dict, payload: Dict, recipient_name: Optional[str]):
    try:
        os.makedirs(USER_CSV_DIR, exist_ok=True)

        def _csv_number(val: float) -> str:
            try:
                txt = format(val, "g")
            except Exception:
                txt = str(val)
            return txt.replace(".", ",")

        def _csv_clean(val):
            """Evita saltos de linea en el CSV para no romper filas al importar. Formatea decimales con coma."""
            if val is None:
                return ""
            if isinstance(val, (int, float)):
                return _csv_number(val)
            return str(val).replace("\r\n", " ").replace("\n", " ").replace("\r", " ").strip()

        def _append_to_base_csv(fecha_val: str, hora_val: str, token_val: str, sender_email: str, sender_name: Optional[str], segs):
            try:
                import csv

                need_header = not os.path.isfile(BASE_DATA_CSV)

                # utf-8-sig para que Excel detecte bien acentos
                with open(BASE_DATA_CSV, "a", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f, delimiter=";")
                    if need_header:
                        writer.writerow(["Fecha", "Hora", "Token", "Nombre", "Email", "ProyectoInicial", "ProyectoFinal", "Tiempo", "Descripcion"])
                    if not segs:
                        writer.writerow(
                            [
                                _csv_clean(fecha_val),
                                _csv_clean(hora_val),
                                _csv_clean(token_val),
                                _csv_clean(sender_name or sender_email),
                                _csv_clean(sender_email),
                                "",
                                "",
                                "",
                                "",
                            ]
                        )
                    else:
                        for seg in segs:
                            t_raw = seg.get("tiempo_raw") or ""
                            t_hours = _parse_time_to_hours(t_raw)
                            if t_hours is not None:
                                t_val = _csv_number(t_hours)
                            else:
                                t_val = _csv_clean(t_raw)
                            proj_initial = seg.get("proyecto_raw") or seg.get("proyecto") or ""
                            proj_final = seg.get("proyecto") or proj_initial
                            writer.writerow(
                                [
                                    _csv_clean(fecha_val),
                                    _csv_clean(hora_val),
                                    _csv_clean(token_val),
                                    _csv_clean(sender_name or sender_email),
                                    _csv_clean(sender_email),
                                    _csv_clean(proj_initial),
                                    _csv_clean(proj_final),
                                    t_val,
                                    _csv_clean(seg.get("descripcion") or ""),
                                ]
                            )
            except Exception as exc_inner:
                log(f"No se pudo guardar en CSV base datos: {exc_inner}", "WARN")

        def _append_to_user_csv(fecha_val: str, hora_val: str, token_val: str, sender_email: str, sender_name: Optional[str], segs):
            try:
                import csv

                resolved_name = _resolve_display_name(sender_email) or sender_name
                name_for_file = (resolved_name or sender_email or "sin_nombre").strip() or "sin_nombre"
                safe_filename = re.sub(r'[<>:"/\\|?*]', "_", name_for_file)
                user_csv_path = os.path.join(USER_CSV_DIR, f"{safe_filename}.csv")
                need_header = not os.path.isfile(user_csv_path)
                with open(user_csv_path, "a", encoding="utf-8-sig", newline="") as f:
                    writer = csv.writer(f, delimiter=";")
                    if need_header:
                        writer.writerow(["Fecha", "Hora", "Token", "ProyectoFinal", "Tiempo", "Registro", "Observaciones"])
                    if not segs:
                        writer.writerow([_csv_clean(fecha_val), _csv_clean(hora_val), _csv_clean(token_val), "", "", "", ""])
                    else:
                        for seg in segs:
                            t_raw_user = seg.get("tiempo_raw") or ""
                            t_hours = _parse_time_to_hours(t_raw_user)
                            # Si no hay unidad y es numero, agregar "horas" para el CSV del usuario
                            if t_hours is not None and not re.search(r"[a-zA-Z]", t_raw_user or ""):
                                t_disp = f"{_csv_number(t_hours)} horas"
                            else:
                                t_disp = _csv_clean(t_raw_user)
                            writer.writerow(
                                [
                                    _csv_clean(fecha_val),
                                    _csv_clean(hora_val),
                                    _csv_clean(token_val),
                                    _csv_clean(seg.get("proyecto") or ""),
                                    t_disp,
                                    _csv_clean(seg.get("descripcion") or ""),
                                    "",
                                ]
                            )
            except Exception as exc_inner:
                log(f"No se pudo guardar CSV por usuario: {exc_inner}", "WARN")

        body_text = payload.get("body") or payload.get("snippet") or ""
        body_text = _strip_signature(body_text)
        free_text, segments = _parse_projects_v2(body_text)
        # Si no se pudo parsear pero hay texto libre, registrar todo como un unico segmento.
        if not segments and free_text.strip():
            segments = [
                {
                    "proyecto": free_text.strip(),
                    "proyecto_raw": free_text.strip(),
                    "tiempo": None,
                    "tiempo_raw": "",
                    "descripcion": free_text.strip(),
                }
            ]
        # Si sigue sin haber segmentos, crear uno de fallback para no perder datos.
        if not segments:
            fallback_desc = (body_text or "").strip() or "Sin contenido"
            segments = [
                {
                    "proyecto": "Sin Proyecto Asignado",
                    "proyecto_raw": fallback_desc,
                    "tiempo": None,
                    "tiempo_raw": "",
                    "descripcion": fallback_desc,
                }
            ]

        glossary = _load_project_glossary()
        if glossary:
            for seg in segments:
                raw_orig = seg.get("proyecto") or ""
                seg["proyecto_raw"] = seg.get("proyecto_raw") or raw_orig
                matched_name, matched = _match_project_with_flag(raw_orig, glossary)
                seg["proyecto"] = matched_name if matched else "Sin Proyecto Asignado"
        else:
            for seg in segments:
                raw_orig = seg.get("proyecto") or ""
                seg["proyecto_raw"] = seg.get("proyecto_raw") or raw_orig
                seg["proyecto"] = raw_orig or "Sin Proyecto Asignado"

        # Normalizar campos para no dejar vacios.
        for seg in segments:
            seg["proyecto_raw"] = (seg.get("proyecto_raw") or seg.get("proyecto") or "Sin Proyecto Asignado") or "Sin Proyecto Asignado"
            seg["proyecto"] = (seg.get("proyecto") or "").strip() or "Sin Proyecto Asignado"
            if not seg.get("descripcion"):
                # Si no hay descripcion explicita para el proyecto, dejar en blanco (no rellenar con otros bloques).
                seg["descripcion"] = ""
            if seg.get("tiempo") is None and (seg.get("tiempo_raw") or "").strip() == "":
                seg["tiempo_raw"] = ""

        received_at = payload.get("received_at") or ""
        try:
            dt = datetime.fromisoformat(received_at)
            fecha = dt.strftime("%d/%m/%Y")
            hora = dt.strftime("%H:%M:%S")
        except Exception:
            fecha = campaign.get("date") or ""
            hora = ""

        resolved_name = _resolve_display_name(payload.get("from") or "") or recipient_name
        try:
            _append_to_base_csv(fecha, hora, campaign.get("token") or "", payload.get("from") or "", resolved_name, segments)
        except Exception as exc:
            log(f"No se pudo guardar en base_datos_respuestas.csv: {exc}", "WARN")
        try:
            _append_to_user_csv(fecha, hora, campaign.get("token") or "", payload.get("from") or "", resolved_name, segments)
        except Exception as exc:
            log(f"No se pudo guardar en CSV por usuario: {exc}", "WARN")
    except Exception as exc:
        log(f"No se pudo guardar respuesta en CSV: {exc}", "WARN")


def _recipients_from_csv() -> List[Dict[str, Optional[str]]]:
    recipients: List[Dict[str, Optional[str]]] = []
    if not os.path.isfile(RECIPIENTS_FILE):
        return recipients
    try:
        with open(RECIPIENTS_FILE, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.DictReader(f)
            if reader.fieldnames and "EMAIL" in [h.upper() for h in reader.fieldnames]:
                for row in reader:
                    email_cell = (row.get("EMAIL") or "").strip()
                    if not email_cell:
                        continue
                    name_parts = []
                    for key in ("NOMBRE", "APELLIDO"):
                        val = row.get(key) or row.get(key.lower()) or ""
                        if val.strip():
                            name_parts.append(val.strip())
                    name = " ".join(name_parts).strip() or None
                    recipients.append({"email": email_cell, "name": name})
            else:
                f.seek(0)
                for line in f:
                    line = line.strip()
                    if not line or line.startswith("#"):
                        continue
                    cells = [c.strip() for c in re.split(r"[;,]", line) if c.strip()]
                    email_cell = next((c for c in cells if "@" in c and c.lower() != "email"), None)
                    if email_cell:
                        name_cell = next((c for c in cells if "@" not in c and c.lower() != "email"), None)
                        recipients.append({"email": email_cell, "name": name_cell})
    except Exception as exc:
        log(f"No se pudo leer {RECIPIENTS_FILE}: {exc}", "WARN")
    return recipients


def load_recipients() -> List[Dict[str, Optional[str]]]:
    recipients: List[Dict[str, Optional[str]]] = []

    # Perfil forzado (ignora variables de entorno y override externo).
    if FORCE_PROFILE:
        profiles = load_profiles()
        profile_data = profiles.get("profiles", {}).get(FORCE_PROFILE, {})
        recipients.extend(_recipients_from_profile(profile_data))
        if not recipients and profile_data.get("use_csv", False):
            recipients.extend(_recipients_from_csv())
        if not recipients:
            recipients.extend(_recipients_from_csv())
    else:
        # 1) Variable de entorno (override total).
        raw = os.environ.get("ACTIVITY_RECIPIENTS")
        if raw:
            for addr in parse_recipients(raw):
                recipients.append({"email": addr, "name": None})
            if recipients:
                return recipients

        # 2) Perfil activo desde archivo o env ACTIVITY_PROFILE.
        profiles = load_profiles()
        active_profile = os.environ.get("ACTIVITY_PROFILE") or profiles.get("active_profile") or "test"
        profile_data = profiles.get("profiles", {}).get(active_profile, {})

        recipients.extend(_recipients_from_profile(profile_data))
        if not recipients and profile_data.get("use_csv", False):
            recipients.extend(_recipients_from_csv())

        # 3) Fallback a CSV si el perfil no aporto nada.
        if not recipients:
            recipients.extend(_recipients_from_csv())

    cleaned: List[Dict[str, Optional[str]]] = []
    seen = set()
    for rec in recipients:
        email_addr = (rec.get("email") or "").strip()
        if not email_addr:
            continue
        key = email_addr.lower()
        if "@" not in key:
            continue
        if key not in seen:
            cleaned.append({"email": email_addr, "name": rec.get("name")})
            seen.add(key)

    if not cleaned:
        log("No hay destinatarios definidos (perfil activo, CSV o env ACTIVITY_RECIPIENTS).", "WARN")
    return cleaned


def render_subject(fecha_corta: str, fecha_larga: str, token: str, recipient_name: Optional[str] = None) -> str:
    subject = ""
    if os.path.isfile(SUBJECT_TEMPLATE_FILE):
        try:
            template = open(SUBJECT_TEMPLATE_FILE, "r", encoding="utf-8").read().strip()
            if template:
                subject = template.format(
                    fecha=fecha_corta,
                    Fecha=fecha_corta,
                    fecha_corta=fecha_corta,
                    Fecha_corta=fecha_corta,
                    fecha_larga=fecha_larga,
                    Fecha_larga=fecha_larga,
                    token=token,
                    Token=token,
                    nombre=recipient_name or "",
                    Nombre=recipient_name or "",
                )
        except Exception as exc:
            log(f"No se pudo leer {SUBJECT_TEMPLATE_FILE}: {exc}", "WARN")

    if not subject:
        subject = f"Registro de actividad {fecha_corta}"
    return subject


def render_body(fecha_corta: str, fecha_larga: str, token: str, recipient_name: Optional[str] = None) -> str:
    token_full = f"{TOKEN_PREFIX}:{token}"
    body = ""
    if os.path.isfile(BODY_TEMPLATE_FILE):
        try:
            template = open(BODY_TEMPLATE_FILE, "r", encoding="utf-8").read()
            if template:
                body = template.format(
                    fecha=fecha_corta,
                    Fecha=fecha_corta,
                    fecha_corta=fecha_corta,
                    Fecha_corta=fecha_corta,
                    fecha_larga=fecha_larga,
                    Fecha_larga=fecha_larga,
                    token=token,
                    Token=token_full,
                    Token_full=token_full,
                    nombre=recipient_name or "",
                    Nombre=recipient_name or "",
                )
                global _BODY_TEMPLATE_CACHE
                _BODY_TEMPLATE_CACHE = body
        except Exception as exc:
            log(f"No se pudo leer {BODY_TEMPLATE_FILE}: {exc}", "WARN")

    if not body:
        body = (
            "Hola,\n\n"
            "Por favor responde a este correo con tu registro diario. Incluimos un token unico para poder\n"
            "vincular tu respuesta automaticamente.\n\n"
            f"Fecha: {fecha_corta}\n"
            f"Token: {token_full}\n\n"
            "Gracias."
        )
    if TOKEN_PREFIX not in body and token not in body:
        body = body.rstrip() + f"\n\nToken: {token_full}"
    _BODY_TEMPLATE_CACHE = body
    return body


def scheduled_time_for(day: date, override: Optional[datetime] = None) -> Optional[datetime]:
    if override:
        return override
    target = SCHEDULE_HOUR.get(day.weekday())
    if not target:
        return None
    hour, minute = target
    return datetime(day.year, day.month, day.day, hour, minute, 0)


def already_sent(state: Dict, day: date) -> bool:
    for camp in state.get("campaigns", []):
        if camp.get("date") == day.isoformat():
            return True
    return False


def _build_token() -> str:
    return uuid.uuid4().hex[:8]


def maybe_send_daily(state: Dict, now: datetime, holidays, *, bypass_schedule: bool = False, force: bool = False, allow_duplicate_today: bool = False) -> Dict:
    today = now.date()
    if _is_blocked_campaign_date(today.isoformat()):
        return state
    target_dt = scheduled_time_for(today, override=now if bypass_schedule else None)
    if not target_dt:
        return state
    if not force and not is_business_day(today, holidays):
        return state
    if now < target_dt:
        return state
    if already_sent(state, today) and not allow_duplicate_today:
        return state

    recipients_info = load_recipients()
    if not recipients_info:
        log("Envio diario omitido: no hay destinatarios configurados.", "WARN")
        return state

    fecha_corta = f"{today:%d/%m/%y}"
    fecha_larga = f"{today:%d/%m/%Y}"
    token = _build_token()
    sent_count = 0
    successful_recipients: List[str] = []
    for rec in recipients_info:
        email_addr = rec.get("email")
        if not email_addr:
            continue
        name = rec.get("name")
        subject = render_subject(fecha_corta, fecha_larga, token, name)
        body = render_body(fecha_corta, fecha_larga, token, name)
        if email_utils.send_email(subject, body, None, recipients_override=[email_addr]):
            sent_count += 1
            successful_recipients.append(email_addr)
    if sent_count == 0:
        log("Fallo el envio del correo diario (ningun destinatario recibio).", "ERROR")
        return state

    campaign = {
        "date": today.isoformat(),
        "sent_at": now.isoformat(timespec="seconds"),
        "subject": render_subject(fecha_corta, fecha_larga, token, None),
        "token": token,
        "recipients": successful_recipients,
        "responses": {},
    }
    state.setdefault("campaigns", []).append(campaign)
    state["campaigns"] = state["campaigns"][-MAX_CAMPAIGNS_IN_STATE:]
    log(f"Correo diario enviado a {sent_count}/{len(recipients_info)} destinatarios. Token {token}.", "INFO")
    return state


def _imap_connect() -> Optional[imaplib.IMAP4_SSL]:
    host = os.environ.get("IMAP_HOST", "imap.gmail.com")
    port = int(os.environ.get("IMAP_PORT", "993"))
    user = os.environ.get("IMAP_USER") or os.environ.get("SMTP_USER") or SMTP_DEFAULTS.get("user")
    password = os.environ.get("IMAP_PASS") or os.environ.get("SMTP_PASS") or SMTP_DEFAULTS.get("pass")
    if not user or not password:
        log("IMAP no configurado (IMAP_USER/IMAP_PASS).", "WARN")
        return None
    try:
        client = imaplib.IMAP4_SSL(host, port)
        client.login(user, password)
        client.select("INBOX")
        return client
    except Exception as exc:
        log(f"No se pudo conectar a IMAP {host}:{port} como {user}: {exc}", "ERROR")
        return None


def _own_addresses() -> set:
    addrs = set()
    for key in ("IMAP_USER", "SMTP_USER"):
        val = os.environ.get(key)
        if val and "@" in val:
            addrs.add(val.lower().strip())
    default_user = SMTP_DEFAULTS.get("user")
    if default_user and "@" in default_user:
        addrs.add(default_user.lower().strip())
    return addrs


def _extract_token_from_text(text: str) -> Optional[str]:
    if not text:
        return None
    match = re.search(rf"{TOKEN_PREFIX}:([A-Za-z0-9\-]+)", str(text))
    if match:
        return match.group(1)
    return None


def _campaign_by_token(state: Dict, token: str) -> Optional[Dict]:
    for camp in state.get("campaigns", []):
        if camp.get("token") == token:
            return camp
    return None


def _recipient_name_from_profiles(sender_email: str) -> Optional[str]:
    for rec in load_recipients():
        if (rec.get("email") or "").lower() == sender_email.lower():
            return rec.get("name")
    return None


def _recipient_name_from_any_source(sender_email: str) -> Optional[str]:
    sender_email_l = (sender_email or "").lower()
    # 1) Perfiles activos/inactivos
    try:
        profiles = load_profiles().get("profiles", {})
        for p in profiles.values():
            recs = p.get("recipients") or []
            for rec in recs:
                if isinstance(rec, dict):
                    em = (rec.get("email") or "").lower()
                    if em == sender_email_l:
                        nm = (rec.get("name") or "").strip()
                        if nm:
                            return nm
    except Exception:
        pass
    # 2) CSV destinatarios
    try:
        for rec in _recipients_from_csv():
            em = (rec.get("email") or "").lower()
            if em == sender_email_l:
                nm = (rec.get("name") or "").strip()
                if nm:
                    return nm
    except Exception:
        pass
    # 3) Activo
    return _recipient_name_from_profiles(sender_email)


def _resolve_display_name(sender_email: str) -> Optional[str]:
    """
    Devuelve el nombre legible buscando en perfiles y destinatarios.csv.
    Evita crear CSVs por email si ya tenemos nombre conocido.
    """
    nm = _recipient_name_from_any_source(sender_email)
    if nm:
        return nm.strip() or None
    return None


def _record_response(campaign: Dict, sender_email: str, payload: Dict):
    responses = campaign.setdefault("responses", {})
    key = sender_email.lower()
    existing = responses.get(key)
    if existing and isinstance(existing, dict):
        try:
            existing_uid = int(existing.get("uid") or 0)
            new_uid = int(payload.get("uid") or 0)
            if new_uid <= existing_uid:
                return
        except Exception:
            return
    responses[key] = payload
    try:
        os.makedirs(RESPONSES_DIR, exist_ok=True)
        token = campaign.get("token") or "token"
        filename = f"{_safe_filename(token)}_{_safe_filename(sender_email)}_{payload.get('uid','')}.txt"
        path = os.path.join(RESPONSES_DIR, filename)
        body = payload.get("body") or payload.get("snippet") or ""
        with open(path, "w", encoding="utf-8") as f:
            f.write(body)
        payload["saved_path"] = path
    except Exception as exc:
        log(f"No se pudo guardar respuesta en archivo: {exc}", "WARN")
    try:
        recipient_name = _recipient_name_from_any_source(sender_email)
        _append_response_to_excel(campaign, payload, recipient_name)
    except Exception as exc:
        log(f"No se pudo registrar respuesta en Excel: {exc}", "WARN")


def _response_file_exists(token: str, sender_email: str, uid) -> bool:
    filename = f"{_safe_filename(token)}_{_safe_filename(sender_email)}_{uid}.txt"
    path = os.path.join(RESPONSES_DIR, filename)
    return os.path.isfile(path)


def check_replies(state: Dict) -> Dict:
    client = _imap_connect()
    if client is None:
        return state

    last_uid = int(state.get("imap", {}).get("last_uid") or 0)
    seen_uids = set()
    try:
        seen_list = state.get("imap", {}).get("seen_uids") or []
        for item in seen_list:
            try:
                seen_uids.add(int(item))
            except Exception:
                continue
    except Exception:
        seen_uids = set()
    max_seen_uid = last_uid
    try:
        # Buscar siempre desde (last_uid - 500) para tolerar resets/cambios de cuenta IMAP.
        start_uid = max(1, last_uid - 500)
        typ, data = client.uid("search", None, f"UID {start_uid}:*")
        if typ != "OK":
            return state
        uid_list = [int(x) for x in data[0].split() if x]
        uid_list.sort()
        if not uid_list:
            return state
        for uid in uid_list:
            if uid in seen_uids:
                continue
            if uid > max_seen_uid:
                max_seen_uid = uid
            typ, fetched = client.uid("fetch", str(uid), "(RFC822)")
            if typ != "OK" or not fetched or fetched[0] is None:
                continue
            raw = fetched[0][1]
            msg = email.message_from_bytes(raw, policy=policy.default)
            subj = _decode_header(msg.get("Subject"))
            sender_name, sender_addr = email.utils.parseaddr(msg.get("From") or "")
            sender = sender_addr
            own_sender = sender and sender.lower().strip() in _own_addresses()
            text = _extract_plain_text(msg)
            token = _extract_token_from_text(subj) or _extract_token_from_text(text) or _extract_token_from_text(raw)
            if not token:
                continue
            # Si es el correo saliente original (sin hilo) desde nuestra cuenta, saltarlo
            if own_sender and not msg.get("In-Reply-To"):
                continue
            campaign = _campaign_by_token(state, token)
            if not campaign:
                campaign = {
                    "date": datetime.now().date().isoformat(),
                    "sent_at": None,
                    "subject": subj or "",
                    "token": token,
                    "recipients": [],
                    "responses": {},
                }
                state.setdefault("campaigns", []).append(campaign)
                state["campaigns"] = state["campaigns"][-MAX_CAMPAIGNS_IN_STATE:]
            reply_body = _extract_reply_text(text)
            if not reply_body:
                continue
            if _BODY_TEMPLATE_CACHE and _BODY_TEMPLATE_CACHE.strip() and reply_body.strip() == _BODY_TEMPLATE_CACHE.strip():
                continue
            payload = {
                "from": sender,
                "from_name": sender_name,
                "subject": subj,
                "date": msg.get("Date"),
                "received_at": datetime.now().isoformat(timespec="seconds"),
                "snippet": reply_body.strip()[:500],
                "body": reply_body.strip(),
                "uid": uid,
            }
            # Evitar duplicados si ya tenemos un UID igual o mayor para ese remitente
            existing = campaign.get("responses", {}).get(sender.lower())
            if existing:
                try:
                    existing_uid = int(existing.get("uid") or 0)
                    if uid <= existing_uid:
                        continue
                except Exception:
                    pass
            # Evitar reproceso si ya existe el archivo de respuesta (proteccion contra multiples procesos)
            if _response_file_exists(campaign.get("token") or token, sender, uid):
                continue
            _record_response(campaign, sender, payload)
            seen_uids.add(uid)
            state.setdefault("imap", {})["seen_uids"] = sorted(seen_uids)[-500:]
            state["imap"]["last_uid"] = max(state["imap"].get("last_uid", 0), uid)
            save_state(state)
            log(f"Respuesta registrada de {sender} para token {token} (UID {uid}).", "INFO")
    except Exception as exc:
        log(f"Error al revisar respuestas IMAP: {exc}", "WARN")
    finally:
        try:
            client.logout()
        except Exception:
            pass
    if max_seen_uid and max_seen_uid != last_uid:
        state.setdefault("imap", {})["last_uid"] = max_seen_uid
    # Persistir UIDs vistos para evitar reprocesarlos
    state.setdefault("imap", {})["seen_uids"] = sorted(seen_uids)[-500:]
    return state


def main_loop():
    holidays = load_holidays()
    state = load_state()
    last_imap_check = 0.0
    last_holiday_refresh = date.today()
    log("Activity mailer iniciado en modo loop.", "INFO")

    while True:
        now = datetime.now()
        try:
            if now.date() != last_holiday_refresh:
                holidays = load_holidays()
                last_holiday_refresh = now.date()

            state = ensure_state_struct(state)
            state = maybe_send_daily(state, now, holidays)

            if (time.time() - last_imap_check) >= IMAP_CHECK_INTERVAL_SECONDS:
                state = check_replies(state)
                last_imap_check = time.time()

            save_state(state)
        except Exception as exc:
            log(f"Error en loop principal: {exc}", "ERROR")
        time.sleep(30)


def send_once(force: bool = False, bypass_schedule: bool = False, allow_duplicate_today: bool = False):
    holidays = load_holidays()
    now = datetime.now()
    state = load_state()
    state = maybe_send_daily(state, now, holidays, bypass_schedule=bypass_schedule, force=force, allow_duplicate_today=allow_duplicate_today)
    save_state(state)


def check_replies_once():
    state = load_state()
    state = check_replies(state)
    save_state(state)


def parse_args():
    parser = argparse.ArgumentParser(description="Envio y captura de registros diarios por email.")
    parser.add_argument("--loop", action="store_true", help="Ejecuta el ciclo continuo (default).")
    parser.add_argument("--send-now", action="store_true", help="Forzar un envio inmediato si corresponde al dia.")
    parser.add_argument("--force", action="store_true", help="Permite enviar aunque no sea dia habil (solo con --send-now).")
    parser.add_argument("--bypass-schedule", action="store_true", help="Ignora el horario programado y envia ya (solo con --send-now).")
    parser.add_argument("--allow-duplicate", action="store_true", help="Permite otro envio en el mismo dia (solo con --send-now).")
    parser.add_argument("--check-replies", action="store_true", help="Solo revisa respuestas y sale.")
    return parser.parse_args()


if __name__ == "__main__":
    if not acquire_lock():
        sys.exit(0)
    args = parse_args()
    try:
        if args.check_replies:
            check_replies_once()
        elif args.send_now:
            send_once(force=args.force, bypass_schedule=args.bypass_schedule, allow_duplicate_today=args.allow_duplicate)
        else:
            main_loop()
    finally:
        release_lock()


