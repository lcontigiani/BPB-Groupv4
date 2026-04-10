from flask import Flask, render_template, jsonify, request, send_from_directory, session, redirect, url_for, send_file, make_response
import heapq

from decimal import Decimal

import os

import json
import sys

import urllib.request
import urllib.error
from urllib.parse import quote, unquote, urlparse

import csv

import time

import subprocess

import shutil
import threading
from datetime import datetime, timedelta, date, timezone
import tempfile
import warnings

from pathlib import Path
from typing import Optional, Tuple

from io import StringIO
import hashlib
import hmac

from werkzeug.utils import secure_filename

from werkzeug.security import check_password_hash, generate_password_hash

from werkzeug.security import check_password_hash, generate_password_hash

import re
import unicodedata
import math

import openpyxl
import smtplib
import uuid
import secrets
from copy import deepcopy
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

# Logistics Solver Import
from logistics_solver import Packer, Bin, Item, RotationType
from chatbot_bridge import build_chatbot_response
from path_config import (
    resolve_activity_codigos_dir,
    resolve_control_base_dir,
    resolve_iso_code_root,
    resolve_iso_docs_root,
    resolve_oficina_root,
    resolve_project_costos_analysis_root,
    resolve_quality_root,
    resolve_r016_dir,
    resolve_workspace_root,
)

PBKDF2_METHOD = 'pbkdf2:sha256:600000'


class LogisticsCalculationCancelled(Exception):
    pass


def _has_native_scrypt():
    return hasattr(hashlib, 'scrypt')


def _utc_now_iso_z():
    return datetime.now(timezone.utc).isoformat().replace('+00:00', 'Z')


def _load_workbook_quietly(*args, **kwargs):
    with warnings.catch_warnings():
        warnings.filterwarnings(
            "ignore",
            message="Cannot parse header or footer so it will be ignored",
            category=UserWarning,
        )
        return openpyxl.load_workbook(*args, **kwargs)


def _generate_password_hash_compatible(password):
    method = 'scrypt' if _has_native_scrypt() else PBKDF2_METHOD
    return generate_password_hash(password, method=method)


def _check_password_hash_compatible(stored_hash, password):
    if not stored_hash:
        return False

    try:
        return check_password_hash(stored_hash, password)
    except AttributeError:
        if not str(stored_hash).startswith('scrypt:'):
            raise

        try:
            import pyscrypt

            method, salt, digest = str(stored_hash).split('$', 2)
            _, n_str, r_str, p_str = method.split(':', 3)
            derived = pyscrypt.hash(
                password.encode('utf-8'),
                salt.encode('utf-8'),
                N=int(n_str),
                r=int(r_str),
                p=int(p_str),
                dkLen=len(digest) // 2,
            )
            return hmac.compare_digest(derived.hex(), digest)
        except Exception:
            raise

JSON_IO_LOCKS = {}
JSON_IO_LOCKS_GUARD = threading.Lock()


def _json_path(path_value) -> Path:
    return path_value if isinstance(path_value, Path) else Path(path_value)


def _get_json_lock(path_value) -> threading.RLock:
    path = _json_path(path_value)
    try:
        key = str(path.resolve())
    except Exception:
        key = str(path)
    with JSON_IO_LOCKS_GUARD:
        lock = JSON_IO_LOCKS.get(key)
        if lock is None:
            lock = threading.RLock()
            JSON_IO_LOCKS[key] = lock
        return lock


def _json_backup_path(path_value) -> Path:
    path = _json_path(path_value)
    suffix = f"{path.suffix}.bak" if path.suffix else ".bak"
    return path.with_suffix(suffix)


def _load_json_file(path_value, default_value, expected_type=None):
    path = _json_path(path_value)
    lock = _get_json_lock(path)
    with lock:
        candidates = [path, _json_backup_path(path)]
        for candidate in candidates:
            if not candidate.exists():
                continue
            try:
                with candidate.open('r', encoding='utf-8') as f:
                    data = json.load(f)
                if expected_type is not None and not isinstance(data, expected_type):
                    continue
                return data
            except Exception as exc:
                print(f"JSON LOAD WARNING [{candidate}]: {exc}")
        return deepcopy(default_value)


def _write_json_file_atomic(path_value, data, *, indent=4, ensure_ascii=False):
    path = _json_path(path_value)
    lock = _get_json_lock(path)
    with lock:
        path.parent.mkdir(parents=True, exist_ok=True)
        backup_path = _json_backup_path(path)
        temp_fd = None
        temp_path = None
        try:
            payload = json.dumps(data, indent=indent, ensure_ascii=ensure_ascii)
            temp_fd, temp_path = tempfile.mkstemp(
                prefix=f".{path.name}.",
                suffix=".tmp",
                dir=str(path.parent),
                text=True
            )
            with os.fdopen(temp_fd, 'w', encoding='utf-8', newline='\n') as f:
                temp_fd = None
                f.write(payload)
                f.flush()
                os.fsync(f.fileno())

            if path.exists():
                try:
                    shutil.copy2(path, backup_path)
                except Exception as backup_exc:
                    print(f"JSON BACKUP WARNING [{path}]: {backup_exc}")

            os.replace(temp_path, path)
            temp_path = None
        finally:
            if temp_fd is not None:
                try:
                    os.close(temp_fd)
                except Exception:
                    pass
            if temp_path:
                try:
                    os.unlink(temp_path)
                except Exception:
                    pass

def _do_pack_internal(container_data, items_data, config):
    """Core Logistic Solver logic extracted for reuse in maximization."""
    try:
        job_id = config.get('job_id')
        progress_detail = config.get('progress_detail', False)
        debug_verbose = config.get('debug_verbose', True)

        def debug_log(message):
            if debug_verbose:
                print(message, flush=True)

        debug_log(f"DEBUG: _do_pack_internal called. Items: {len(items_data)}")

        def progress(progress_value, message, stage=None):
            _raise_if_logistics_calc_cancelled(job_id)
            if progress_detail and job_id:
                if stage:
                    _update_logistics_calc_progress(job_id, stage=stage, progress=progress_value, message=message, status='running')
                else:
                    _set_logistics_calc_state(job_id, progress=progress_value, message=message, status='running')

        load_type_key = config.get('load_type', 'loose')
        pallet_type_key = config.get('pallet_type', 'none')
        container_type_key = config.get('container_type', '20ft')
        user_max_pallet_h = float(config.get('max_pallet_height', 0))
        
        PALLET_DIMS = {
            'europallet': {'w': 1200, 'd': 800, 'h': 150, 'weight': 25},
            'american': {'w': 1200, 'd': 1000, 'h': 150, 'weight': 25}
        }
        
        COLLARS_TYPES = ('collars', 'collars_120x100')
        is_collars = pallet_type_key in COLLARS_TYPES
        is_tray_load = load_type_key in ('tray', 'tray_euro', 'tray_american', 'tray_custom')
        allow_stacking = config.get('stack_load', True)
        stack_trays = config.get('stack_trays', allow_stacking)
        stack_collars = config.get('stack_collars', allow_stacking)
        mixed_trays = config.get('mixed_trays', True)
        tray_kanban_enabled = bool(config.get('tray_kanban')) and is_tray_load
        if tray_kanban_enabled:
            mixed_trays = False

        # Forced orientation (face down) for items
        forced_rotations = None
        if config.get('force_orientation'):
            face = str(config.get('orientation_face', '') or '').strip().lower()
            if face in ('lxa', 'largo x ancho', 'largo_x_ancho'):
                forced_rotations = [RotationType.RT_WHD, RotationType.RT_DHW]
            elif face in ('lxh', 'largo x alto', 'largo_x_alto'):
                forced_rotations = [RotationType.RT_WDH, RotationType.RT_HDW]
            elif face in ('axh', 'ancho x alto', 'ancho_x_alto'):
                forced_rotations = [RotationType.RT_DWH, RotationType.RT_HWD]
        
        # 0. Safety Factors & Item Preparation
        sf_percent_dims = float(config.get('safety_factor_dims', 0))
        sf_percent_weight = float(config.get('safety_factor_weight', 0))
        
        sf_mult_dims = Decimal(1.0 + (sf_percent_dims / 100.0))
        sf_mult_weight = Decimal(1.0 + (sf_percent_weight / 100.0))
        
        # 1. Prepare Items
        # 1. Prepare Items
        # --- OPTIMIZATION: Heuristic Cap for Maximization ---
        max_items_cap = float('inf')
        if config.get('maximize') and items_data and not config.get('skip_max_cap'):
            try:
                # Estimate Volume Limit
                c_vol = 0
                if container_type_key != 'none':
                    c_vol = float(container_data.get('width', 0)) * float(container_data.get('height', 0)) * float(container_data.get('depth', 0))
                else:
                    # Single Pallet Mode: Use safe upper bound dims
                    cw = 1200
                    cd = 800
                    ch = 2500 # Safe upper bound height
                    if is_collars:
                        cd = 1000 if pallet_type_key == 'collars_120x100' else 800
                    elif pallet_type_key in PALLET_DIMS:
                        cw = PALLET_DIMS[pallet_type_key]['w']
                        cd = PALLET_DIMS[pallet_type_key]['d']
                    c_vol = float(cw) * float(cd) * float(ch)

                min_vol = float('inf')
                for itm in items_data:
                    v = float(itm.get('w', 1)) * float(itm.get('h', 1)) * float(itm.get('d', 1))
                    if v > 0 and v < min_vol: min_vol = v
                
                if c_vol > 0 and min_vol > 0 and min_vol != float('inf'):
                    max_items_cap = int((c_vol / min_vol) * 1.5) + 200 # 50% buffer + 200 items base safety
            except Exception as e:
                debug_log(f"DEBUG: Optimization Error: {e}")
                pass

        raw_items = []
        total_items_generated = 0
        if debug_verbose:
            qty_summary = [(str(item.get('id', '?')), int(item.get('qty', 1) or 1)) for item in items_data]
            debug_log(f"DEBUG: Preparing raw items. Requested: {qty_summary}, MaxCap: {max_items_cap}")

        def iter_leaf_items(unit):
            if hasattr(unit, 'content_map') and hasattr(unit, 'leaf_item_template'):
                content_map = getattr(unit, 'content_map', {}) or {}
                template = getattr(unit, 'leaf_item_template', None)
                if template:
                    for item_name, qty in content_map.items():
                        try:
                            qty_int = int(qty or 0)
                        except Exception:
                            qty_int = 0
                        for _ in range(max(qty_int, 0)):
                            virtual_leaf = Item(
                                name=item_name,
                                width=template.width,
                                height=template.height,
                                depth=template.depth,
                                weight=template.weight,
                                allowed_rotations=template.allowed_rotations
                            )
                            yield virtual_leaf
                return
            if hasattr(unit, 'inner_items'):
                for child in unit.inner_items:
                    yield from iter_leaf_items(child)
            else:
                yield unit

        def sum_leaf_volume(unit):
            if hasattr(unit, 'leaf_total_volume'):
                return float(getattr(unit, 'leaf_total_volume', 0) or 0)
            if hasattr(unit, 'inner_items'):
                return sum(sum_leaf_volume(child) for child in unit.inner_items)
            return sum([float(i.width * i.height * i.depth) for i in iter_leaf_items(unit)])

        def sum_leaf_weight(unit):
            if hasattr(unit, 'leaf_total_weight'):
                return float(getattr(unit, 'leaf_total_weight', 0) or 0)
            if hasattr(unit, 'inner_items'):
                return sum(sum_leaf_weight(child) for child in unit.inner_items)
            return sum(float(i.weight) for i in iter_leaf_items(unit))

        def collect_content_map(unit):
            if hasattr(unit, 'content_map'):
                return dict(getattr(unit, 'content_map', {}) or {})
            if hasattr(unit, 'inner_items'):
                content_map = {}
                for child in unit.inner_items:
                    child_map = collect_content_map(child)
                    for item_name, qty in child_map.items():
                        try:
                            qty_int = int(qty or 0)
                        except Exception:
                            qty_int = 0
                        content_map[item_name] = content_map.get(item_name, 0) + qty_int
                return content_map
            content_map = {}
            for sub in iter_leaf_items(unit):
                content_map[sub.name] = content_map.get(sub.name, 0) + 1
            return content_map

        def count_leaf_items(unit):
            if hasattr(unit, 'content_map'):
                try:
                    return sum(int(v or 0) for v in getattr(unit, 'content_map', {}).values())
                except Exception:
                    return 0
            if hasattr(unit, 'leaf_total_count'):
                try:
                    return int(getattr(unit, 'leaf_total_count', 0) or 0)
                except Exception:
                    return 0
            if hasattr(unit, 'inner_items'):
                return sum(count_leaf_items(child) for child in unit.inner_items)
            return sum(1 for _ in iter_leaf_items(unit))

        def estimate_item_upper_bound(item_payload, bin_width, bin_depth, bin_height, max_load_weight):
            try:
                i_w = float(item_payload.get('w', 0) or 0)
                i_d = float(item_payload.get('d', 0) or 0)
                i_h = float(item_payload.get('h', 0) or 0)
                i_weight = float(item_payload.get('weight', 0) or 0)
                bounds = []
                if i_w > 0 and i_d > 0 and i_h > 0 and bin_width > 0 and bin_depth > 0 and bin_height > 0:
                    item_vol = i_w * i_d * i_h
                    bin_vol = bin_width * bin_depth * bin_height
                    if item_vol > 0 and bin_vol > 0:
                        bounds.append(max(int(bin_vol // item_vol), 1))
                if max_load_weight > 0 and i_weight > 0:
                    bounds.append(max(int(max_load_weight // i_weight), 1))
                if not bounds:
                    return 512
                return max(1, min(min(bounds) + 1, 50000))
            except Exception:
                return 512

        def max_qty_per_single_bin(item_payload, bin_factory, bin_width, bin_depth, bin_height, max_load_weight):
            upper_bound = estimate_item_upper_bound(item_payload, bin_width, bin_depth, bin_height, max_load_weight)
            probe_cache = {}

            def fits(qty_probe):
                if qty_probe in probe_cache:
                    return probe_cache[qty_probe]
                if qty_probe <= 0:
                    probe_cache[qty_probe] = False
                    return False
                packer = Packer()
                try:
                    probe_bin = bin_factory()
                    packer.add_bin(probe_bin)
                    for _ in range(int(qty_probe)):
                        probe_item = Item(
                            name=item_payload.get('id'),
                            width=Decimal(item_payload.get('w', 0)) * sf_mult_dims,
                            height=Decimal(item_payload.get('h', 0)) * sf_mult_dims,
                            depth=Decimal(item_payload.get('d', 0)) * sf_mult_dims,
                            weight=Decimal(item_payload.get('weight', 0)) * sf_mult_weight,
                            allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                        )
                        if forced_rotations:
                            probe_item.force_orientation = True
                        packer.add_item(probe_item)
                    packer.pack(bigger_first=True)
                    fits_result = len(packer.unfit_items) == 0
                except Exception:
                    fits_result = False
                probe_cache[qty_probe] = fits_result
                return fits_result

            lo_probe = 1
            hi_probe = 1
            best_probe = 0
            while hi_probe <= upper_bound and fits(hi_probe):
                best_probe = hi_probe
                if hi_probe == upper_bound:
                    return best_probe
                hi_probe = min(hi_probe * 2, upper_bound)
            lo_probe = max(1, best_probe + 1)
            hi_probe = max(min(hi_probe - 1, upper_bound), lo_probe)
            while lo_probe <= hi_probe:
                mid_probe = (lo_probe + hi_probe) // 2
                if fits(mid_probe):
                    best_probe = mid_probe
                    lo_probe = mid_probe + 1
                else:
                    hi_probe = mid_probe - 1
            return best_probe

        def build_tray_groups(tray_units):
            tray_groups_map = {}
            item_lookup = {str(it.get('id')): it for it in items_data if isinstance(it, dict)}

            for tray_unit in tray_units:
                if not hasattr(tray_unit, 'inner_items') and not hasattr(tray_unit, 'content_map'):
                    continue

                content_map = collect_content_map(tray_unit)

                dims_str = f"{float(tray_unit.width):.0f}x{float(getattr(tray_unit, 'visual_height', tray_unit.height)):.0f}x{float(tray_unit.depth):.0f}"
                sig = f"{dims_str}_{sorted(content_map.items())}"

                if sig not in tray_groups_map:
                    load_weight = sum_leaf_weight(tray_unit)
                    tray_groups_map[sig] = {
                        'type': 'tray',
                        'name': 'Bandejas',
                        'dims': dims_str,
                        'weight_per_tray': float(tray_unit.weight),
                        'load_weight_per_tray': load_weight,
                        'count': 0,
                        'items_map': content_map
                    }
                tray_groups_map[sig]['count'] += 1

            tray_groups = []
            for _, grp in tray_groups_map.items():
                grp_items = []
                for it_name, qty_per in sorted(grp['items_map'].items()):
                    it_info = item_lookup.get(str(it_name), {})
                    dims_str = "???"
                    unit_weight = 0
                    if it_info:
                        dims_str = f"{it_info.get('w')}x{it_info.get('d')}x{it_info.get('h')}"
                        unit_weight = float(it_info.get('weight', 0) or 0)

                    grp_items.append({
                        'name': it_name,
                        'dims': dims_str,
                        'weight': unit_weight,
                        'qty_per_tray': qty_per,
                        'qty_per_pallet': qty_per,
                        'total_qty': qty_per * grp['count']
                    })

                tray_groups.append({
                    'type': 'tray',
                    'name': grp['name'],
                    'dims': grp['dims'],
                    'weight_per_tray': round(grp['weight_per_tray'], 2),
                    'load_weight_per_tray': round(grp['load_weight_per_tray'], 2),
                    'count': grp['count'],
                    'items': grp_items
                })

            return tray_groups

        kanban_layout_registry = {}

        def build_kanban_tray_layout(source_item, qty_in_tray):
            cache_key = f"{source_item.get('id')}::{int(qty_in_tray)}"
            if cache_key in kanban_layout_registry:
                return cache_key

            tray_layout_packer = Packer()
            tray_layout_packer.add_bin(tray_factory())
            for _ in range(int(qty_in_tray)):
                layout_item = Item(
                    name=source_item.get('id'),
                    width=Decimal(source_item.get('w', 0)) * sf_mult_dims,
                    height=Decimal(source_item.get('h', 0)) * sf_mult_dims,
                    depth=Decimal(source_item.get('d', 0)) * sf_mult_dims,
                    weight=Decimal(source_item.get('weight', 0)) * sf_mult_weight,
                    allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                )
                if forced_rotations:
                    layout_item.force_orientation = True
                tray_layout_packer.add_item(layout_item)

            tray_layout_packer.pack(bigger_first=True)
            tray_layout_bin = tray_layout_packer.bins[0] if tray_layout_packer.bins else None
            if tray_layout_bin is None or tray_layout_packer.unfit_items:
                raise Exception(f'No se pudo generar layout Kanban para {source_item.get("id")} x {qty_in_tray}.')

            kanban_layout_registry[cache_key] = [{
                'name': packed_piece.name,
                'x': float(packed_piece.position[0]),
                'y': float(packed_piece.position[1]),
                'z': float(packed_piece.position[2]),
                'w': float(packed_piece.get_dimension()[0]),
                'h': float(packed_piece.get_dimension()[1]),
                'd': float(packed_piece.get_dimension()[2]),
                'rt': packed_piece.rotation_type,
                'weight': float(getattr(packed_piece, 'weight', 0) or 0)
            } for packed_piece in tray_layout_bin.items]
            return cache_key

        def build_kanban_tray_unit(source_item, qty_in_tray):
            piece_w = Decimal(source_item.get('w', 0)) * sf_mult_dims
            piece_h = Decimal(source_item.get('h', 0)) * sf_mult_dims
            piece_d = Decimal(source_item.get('d', 0)) * sf_mult_dims
            piece_weight = Decimal(source_item.get('weight', 0)) * sf_mult_weight
            layout_id = build_kanban_tray_layout(source_item, qty_in_tray)
            piece_template = Item(
                name=source_item.get('id'),
                width=piece_w,
                height=piece_h,
                depth=piece_d,
                weight=piece_weight,
                allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
            )
            if forced_rotations:
                piece_template.force_orientation = True

            t_item = Item(
                name=f"Tray_{len(tray_items) + 1}",
                width=tray_conf_outer['w'],
                depth=tray_conf_outer['d'],
                height=tray_conf_outer['h'],
                weight=Decimal(tray_conf['weight']) + (piece_weight * Decimal(qty_in_tray)),
                allowed_rotations=[RotationType.RT_WHD, RotationType.RT_DHW]
            )
            t_item.inner_items = []
            t_item.base_height = tray_conf.get('base_h', Decimal('0'))
            t_item.visual_height = float(tray_conf_outer['h'])
            t_item.pallet_type = 'tray'
            t_item.wall_thickness = float(tray_conf.get('wall_thickness', Decimal('0')))
            t_item.content_type = source_item.get('id')
            t_item.content_map = {str(source_item.get('id')): int(qty_in_tray)}
            t_item.kanban_layout_id = layout_id
            t_item.leaf_total_count = int(qty_in_tray)
            t_item.leaf_total_volume = float(piece_w * piece_h * piece_d) * float(qty_in_tray)
            t_item.leaf_total_weight = float(piece_weight) * float(qty_in_tray)
            t_item.leaf_item_template = piece_template
            t_item.kanban_qty_per_tray = int(qty_in_tray)
            return t_item
        
        deferred_tray_kanban_inputs = []
        for item in items_data:
            _raise_if_logistics_calc_cancelled(job_id)
            qty = int(item.get('qty', 1))

            # Apply Cap
            if config.get('maximize') and (total_items_generated + qty) > max_items_cap:
                qty = max(0, max_items_cap - total_items_generated)

            if qty <= 0 and config.get('maximize'):
                continue
            if debug_verbose:
                debug_log(f"DEBUG: Generating units for {item.get('id', '?')}: qty={qty}")

            total_items_generated += qty
            if tray_kanban_enabled:
                deferred_tray_kanban_inputs.append((item, qty))
                if config.get('maximize') and total_items_generated >= max_items_cap:
                    break
                continue

            for _ in range(qty):
                itm = Item(
                    name=item.get('id'),
                    width=Decimal(item.get('w', 0)) * sf_mult_dims,
                    height=Decimal(item.get('h', 0)) * sf_mult_dims,
                    depth=Decimal(item.get('d', 0)) * sf_mult_dims,
                    weight=Decimal(item.get('weight', 0)) * sf_mult_weight,
                    allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                )
                if forced_rotations:
                    itm.force_orientation = True
                raw_items.append(itm)
            if config.get('maximize') and total_items_generated >= max_items_cap:
                break

        if tray_kanban_enabled:
            debug_log(
                f"DEBUG: Raw item materialization skipped by Tray Kanban. "
                f"Requested units: {sum(qty for _, qty in deferred_tray_kanban_inputs)}"
            )
        else:
            debug_log(f"DEBUG: Raw items prepared: {len(raw_items)}")

        progress(18, 'Items preparados...', 'validation')

        items_to_pack_into_container = []
        unfitted_from_palletizing = []
        tray_items = []
        tray_conf = None
        tray_conf_outer = None

        if is_tray_load:
            if tray_kanban_enabled:
                debug_log(
                    f"DEBUG: Tray mode enabled (Kanban). Groups entering tray stage: "
                    f"{[(str(it.get('id', '?')), int(qty)) for it, qty in deferred_tray_kanban_inputs]}"
                )
            else:
                debug_log(f"DEBUG: Tray mode enabled. Raw items entering tray stage: {len(raw_items)}")
            tray_dims = config.get('tray_dims') or {}
            inner_w = Decimal(str(tray_dims.get('tray_inner_w', 0) or 0))
            inner_d = Decimal(str(tray_dims.get('tray_inner_d', 0) or 0))
            inner_h = Decimal(str(tray_dims.get('tray_inner_h', 0) or 0))
            outer_w = Decimal(str(tray_dims.get('tray_outer_w', 0) or 0))
            outer_d = Decimal(str(tray_dims.get('tray_outer_d', 0) or 0))
            outer_h = Decimal(str(tray_dims.get('tray_outer_h', 0) or 0))
            tray_weight = Decimal(str(tray_dims.get('weight', 0) or 0))

            if inner_w > 0 and inner_d > 0 and inner_h > 0 and outer_w > 0 and outer_d > 0 and outer_h > 0:
                wall_thick_w = max((outer_w - inner_w) / Decimal(2), Decimal('0'))
                wall_thick_d = max((outer_d - inner_d) / Decimal(2), Decimal('0'))
                base_h = max(outer_h - inner_h, Decimal('0'))
                wall_thick = max(wall_thick_w, wall_thick_d)

                tray_conf = {
                    'w': inner_w,
                    'd': inner_d,
                    'h': outer_h,
                    'base_h': base_h,
                    'loading_h': inner_h,
                    'weight': tray_weight,
                    'wall_thickness': wall_thick,
                    'max_load_weight': Decimal(str(tray_dims.get('max_weight', 25) or 25))
                }
                tray_conf_outer = {
                    'w': outer_w,
                    'd': outer_d,
                    'h': outer_h,
                    'weight': tray_weight
                }

                def tray_factory():
                    return Bin(name="Tray", width=tray_conf['w'], depth=tray_conf['d'], height=tray_conf['loading_h'], max_weight=tray_conf['max_load_weight'], allow_stacking=True)

                filled_trays = []
                if tray_kanban_enabled:
                    debug_log(f"DEBUG: Tray Kanban enabled. Input groups: {[(str(it.get('id', '?')), int(qty)) for it, qty in deferred_tray_kanban_inputs]}")
                    progress(24, 'Calculando bandejas Kanban...', 'tray')
                    for source_item, requested_qty in deferred_tray_kanban_inputs:
                        _raise_if_logistics_calc_cancelled(job_id)
                        if requested_qty <= 0:
                            continue
                        tray_capacity = max_qty_per_single_bin(
                            source_item,
                            tray_factory,
                            float(tray_conf['w']),
                            float(tray_conf['d']),
                            float(tray_conf['loading_h']),
                            float(tray_conf['max_load_weight'])
                        )
                        debug_log(
                            f"DEBUG: Tray Kanban capacity for {source_item.get('id', '?')}: "
                            f"{tray_capacity} units/tray. Requested={requested_qty}"
                        )
                        if tray_capacity <= 0:
                            for _ in range(requested_qty):
                                unfitted_item = Item(
                                    name=source_item.get('id'),
                                    width=Decimal(source_item.get('w', 0)) * sf_mult_dims,
                                    height=Decimal(source_item.get('h', 0)) * sf_mult_dims,
                                    depth=Decimal(source_item.get('d', 0)) * sf_mult_dims,
                                    weight=Decimal(source_item.get('weight', 0)) * sf_mult_weight,
                                    allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                                )
                                if forced_rotations:
                                    unfitted_item.force_orientation = True
                                unfitted_from_palletizing.append(unfitted_item)
                            continue

                        if config.get('maximize'):
                            tray_quantities = [tray_capacity] * (requested_qty // tray_capacity)
                            remainder_qty = requested_qty % tray_capacity
                            if remainder_qty > 0:
                                debug_log(
                                    f"DEBUG: Tray Kanban remainder for {source_item.get('id', '?')}: "
                                    f"{remainder_qty} units left unfitted because maximization requires trays completas."
                                )
                                piece_w = Decimal(source_item.get('w', 0)) * sf_mult_dims
                                piece_h = Decimal(source_item.get('h', 0)) * sf_mult_dims
                                piece_d = Decimal(source_item.get('d', 0)) * sf_mult_dims
                                piece_weight = Decimal(source_item.get('weight', 0)) * sf_mult_weight
                                for _ in range(remainder_qty):
                                    unfitted_item = Item(
                                        name=source_item.get('id'),
                                        width=piece_w,
                                        height=piece_h,
                                        depth=piece_d,
                                        weight=piece_weight,
                                        allowed_rotations=forced_rotations if forced_rotations else RotationType.ALL
                                    )
                                    if forced_rotations:
                                        unfitted_item.force_orientation = True
                                    unfitted_from_palletizing.append(unfitted_item)
                        else:
                            tray_count = max(int(math.ceil(requested_qty / tray_capacity)), 1)
                            base_qty = requested_qty // tray_count
                            extra_qty = requested_qty % tray_count
                            tray_quantities = [
                                base_qty + (1 if idx < extra_qty else 0)
                                for idx in range(tray_count)
                                if (base_qty + (1 if idx < extra_qty else 0)) > 0
                            ]
                            debug_log(
                                f"DEBUG: Tray Kanban balanced distribution for {source_item.get('id', '?')}: "
                                f"{tray_quantities}"
                            )

                        for qty_in_tray in tray_quantities:
                            tray_items.append(build_kanban_tray_unit(source_item, qty_in_tray))
                    debug_log(
                        f"DEBUG: Tray Kanban stage done. Trays built: {len(tray_items)}, "
                        f"Unfitted after trays: {len(unfitted_from_palletizing)}"
                    )
                elif mixed_trays:
                    tray_packer = Packer()
                    filled_trays = tray_packer.pack_to_many_bins(tray_factory, raw_items, sort_items=config.get('sort_items', True))
                    unfitted_from_palletizing.extend(tray_packer.unfit_items)
                else:
                    tray_groups = {}
                    for raw_item in raw_items:
                        tray_groups.setdefault(raw_item.name, []).append(raw_item)
                    for _, group_items in tray_groups.items():
                        tray_packer = Packer()
                        grp_trays = tray_packer.pack_to_many_bins(tray_factory, group_items, sort_items=config.get('sort_items', True))
                        filled_trays.extend(grp_trays)
                        unfitted_from_palletizing.extend(tray_packer.unfit_items)

                if not tray_kanban_enabled:
                    debug_log(f"DEBUG: Tray stage done. Trays built: {len(filled_trays)}, Unfitted after trays: {len(unfitted_from_palletizing)}")

                for i, t_bin in enumerate(filled_trays):
                    t_item = Item(
                        name=f"Tray_{i+1}",
                        width=tray_conf_outer['w'],
                        depth=tray_conf_outer['d'],
                        height=tray_conf_outer['h'],
                        weight=Decimal(tray_conf['weight']) + Decimal(t_bin.get_total_weight()),
                        allowed_rotations=[RotationType.RT_WHD, RotationType.RT_DHW]
                    )
                    t_item.inner_items = t_bin.items
                    t_item.base_height = tray_conf.get('base_h', Decimal('0'))
                    t_item.visual_height = float(tray_conf_outer['h'])
                    t_item.pallet_type = 'tray'
                    t_item.wall_thickness = float(tray_conf.get('wall_thickness', Decimal('0')))
                    if t_bin.items:
                        t_item.content_type = t_bin.items[0].name
                    else:
                        t_item.content_type = 'unknown'
                    tray_items.append(t_item)
            else:
                tray_items = raw_items

        progress(36, 'Bandejas calculadas...', 'tray')
        
        p_conf = None 
        p_conf_outer = None 
        is_stackable = True 
        
        if pallet_type_key != 'none':
            custom_dims = config.get('pallet_dims') or {}
            if is_collars:
                boards = int(config.get('boards_count', 4))
                board_h_cm = Decimal('195')
                deck_h_cm = Decimal('9')
                base_h_cm = Decimal('150') + deck_h_cm
                wall_thick_cm = Decimal('19') 
                total_h = base_h_cm + (Decimal(boards) * board_h_cm)
                
                # Default preset weight for collars
                p_weight = Decimal('50')
                # OVERRIDE: If user provided a specific weight in UI, use it.
                if custom_dims and custom_dims.get('weight', 0) > 0:
                     p_weight = Decimal(custom_dims['weight'])
                
                outer_w = Decimal('1200')
                outer_d = Decimal('1000') if pallet_type_key == 'collars_120x100' else Decimal('800')

                p_conf = {
                    'w': outer_w - (wall_thick_cm * 2),
                    'd': outer_d - (wall_thick_cm * 2),
                    'h': total_h,
                    'loading_h': (Decimal(boards) * board_h_cm),
                    'weight': p_weight
                }
                p_conf_outer = {'w': outer_w, 'd': outer_d, 'h': total_h, 'weight': p_weight}
                is_stackable = True
            elif custom_dims and custom_dims.get('w', 0) > 0 and custom_dims.get('d', 0) > 0:
                p_conf = {'w': Decimal(custom_dims['w']), 'd': Decimal(custom_dims['d']), 'h': Decimal(custom_dims.get('h', 150)), 'loading_h': Decimal(0), 'weight': Decimal(custom_dims.get('weight', 25))}
                p_conf_outer = p_conf.copy()
                is_stackable = False 
            elif pallet_type_key in PALLET_DIMS:
                preset = PALLET_DIMS[pallet_type_key]
                # Default preset weight
                p_weight = Decimal(preset['weight'])
                # OVERRIDE: If user provided a specific weight in UI, use it.
                if custom_dims and custom_dims.get('weight', 0) > 0:
                     p_weight = Decimal(custom_dims['weight'])
                
                p_conf = {'w': Decimal(preset['w']), 'd': Decimal(preset['d']), 'h': Decimal(preset['h']), 'loading_h': Decimal(0), 'weight': p_weight}
                p_conf_outer = p_conf.copy()
                is_stackable = False 

        if p_conf:
            items_for_palletizing = tray_items if is_tray_load and tray_items else raw_items
            debug_log(f"DEBUG: Pallet stage starting. Units entering palletizing: {len(items_for_palletizing)}")

            if is_tray_load and items_for_palletizing:
                for tray_unit in items_for_palletizing:
                    if not hasattr(tray_unit, 'width') or not hasattr(tray_unit, 'depth'):
                        continue
                    default_fit = int(p_conf['w'] // tray_unit.width) * int(p_conf['d'] // tray_unit.depth)
                    rotated_fit = int(p_conf['w'] // tray_unit.depth) * int(p_conf['d'] // tray_unit.width)
                    if rotated_fit > default_fit:
                        tray_unit.allowed_rotations = [RotationType.RT_DHW]
                        tray_unit.force_orientation = True
                    elif default_fit > 0:
                        tray_unit.allowed_rotations = [RotationType.RT_WHD]
                        tray_unit.force_orientation = True

            # FORCE NON-STACKABLE if Single Pallet Maximization (No Container + Maximize)
            # This ensures only ONE pallet is filled logic-wise, preventing "Collars" from stacking
            if container_type_key == 'none' and config.get('maximize'):
                is_stackable = False

            cont_h = Decimal(container_data.get('height', 0)) if container_type_key != 'none' else Decimal(0)
            if p_conf.get('loading_h', 0) > 0:
                 pallet_loading_height = p_conf['loading_h']
            else:
                if cont_h > 0: physical_limit_load = cont_h - p_conf['h']
                else: physical_limit_load = Decimal(2500)
                user_max_h_dec = Decimal(user_max_pallet_h)
                if user_max_h_dec > 0:
                    user_limit_load = user_max_h_dec - p_conf['h']
                    pallet_loading_height = min(user_limit_load, physical_limit_load)
                else: 
                     # Fix for Efficiency > 100%:
                     # If user doesn't specify max height, we default efficiency calc to 180cm.
                     # So we must also limit the pallet loading to 180cm - base_h.
                     default_std_limit = Decimal(1800) - p_conf['h']
                     pallet_loading_height = min(default_std_limit, physical_limit_load)
            
            user_max_w = Decimal(config.get('max_pallet_weight', 0))
            pallet_max_weight = user_max_w if user_max_w > 0 else Decimal(2000)
            
            # The Packer checks 'items weight' vs 'max_weight'.
            # But 'pallet_max_weight' is GROSS WEIGHT (Load + Base).
            # So we must subtract the base weight to get the allowed LOAD weight.
            pallet_base_weight = p_conf['weight']
            allowed_load_weight = pallet_max_weight - pallet_base_weight
            if allowed_load_weight < 0: allowed_load_weight = Decimal(0)

            pallet_bin_allow_stacking = stack_trays if is_tray_load else allow_stacking

            def pallet_factory():
                return Bin(name=f"Pallet-{pallet_type_key}", width=p_conf['w'], depth=p_conf['d'], height=pallet_loading_height, max_weight=allowed_load_weight, allow_stacking=pallet_bin_allow_stacking)
            
            # --- PALLETIZING LOGIC (Mixed vs Pure) ---
            filled_pallets = []
            if config.get('mixed_pallets', True):
                # Standard: Mix everything
                temp_packer = Packer()
                filled_pallets = temp_packer.pack_to_many_bins(pallet_factory, items_for_palletizing, sort_items=config.get('sort_items', True))
                unfitted_from_palletizing.extend(temp_packer.unfit_items)
            else:
                # No Mixing: Group by Item Name/ID and pack separately
                groups = {}
                for it in items_for_palletizing:
                    group_key = getattr(it, 'content_type', None) or it.name
                    groups.setdefault(group_key, []).append(it)
                
                # Pack each group
                for g_name, g_items in groups.items():
                    _raise_if_logistics_calc_cancelled(job_id)
                    grp_packer = Packer()
                    grp_pallets = grp_packer.pack_to_many_bins(pallet_factory, g_items, sort_items=config.get('sort_items', True))
                    filled_pallets.extend(grp_pallets)
                    unfitted_from_palletizing.extend(grp_packer.unfit_items)

            debug_log(f"DEBUG: Pallet stage done. Pallets built: {len(filled_pallets)}, Unfitted after pallets: {len(unfitted_from_palletizing)}")

            progress(58, 'Pallets calculados...', 'pallet')

            # Pre-calculate Max Height for Single Pallet Maximization Logic
            target_total_h_for_max = Decimal(0)
            if container_type_key == 'none' and config.get('maximize'):
                if is_collars:
                    target_total_h_for_max = p_conf['h']
                else:
                    user_max_h = float(config.get('max_pallet_height', 0))
                    target_total_h_for_max = Decimal(user_max_h) if user_max_h > 0 else Decimal(1800)

            for i, p_bin in enumerate(filled_pallets):
                max_h_used = Decimal(0)
                for pit in p_bin.items:
                     dims = pit.get_dimension()
                     top = Decimal(pit.position[1]) + Decimal(dims[1])
                     if top > max_h_used: max_h_used = top
                
                if is_collars: real_total_h = p_conf_outer['h']
                else: real_total_h = Decimal(p_conf['h']) + max_h_used

                if not is_stackable and cont_h > 0: packing_h = cont_h
                elif not is_stackable and cont_h == 0: 
                    # If Maximizing (Single Pallet Mode), do NOT inflate height, we need fit in strict bin
                    # CRITICAL FIX: To prevent stacking, the item must consume the ENTIRE vertical space
                    # of the Single Pallet Bin. So set packing_h = target_total_h_for_max.
                    if config.get('maximize'): 
                         packing_h = target_total_h_for_max
                         if packing_h <= 0: packing_h = real_total_h # Safety fallback
                    else: packing_h = real_total_h + Decimal(2000)
                else: packing_h = real_total_h

                p_item = Item(name=f"Pallet_{i+1}", width=p_conf_outer['w'], depth=p_conf_outer['d'], height=packing_h, weight=Decimal(p_conf['weight']) + Decimal(p_bin.get_total_weight()), allowed_rotations=[RotationType.RT_WHD, RotationType.RT_DHW])
                p_item.inner_items = p_bin.items 
                if is_collars:
                    p_item.base_height = Decimal('159')
                else:
                    p_item.base_height = p_conf['h']
                p_item.visual_height = float(real_total_h)
                p_item.pallet_type = pallet_type_key
                # Tag it with the "Type" of content (if unmixed)
                # If unmixed, all items have same name. 
                if p_bin.items:
                     p_item.content_type = p_bin.items[0].name 
                else: p_item.content_type = 'unknown'

                items_to_pack_into_container.append(p_item)
        else:
            items_to_pack_into_container = tray_items if is_tray_load and tray_items else raw_items
            # Tag raw items with their own name as content type
            for it in items_to_pack_into_container:
                if not hasattr(it, 'content_type'):
                    it.content_type = it.name

        final_packer = Packer()
        if container_type_key != 'none':
            debug_log(f"DEBUG: Final packing into container. Top-level units: {len(items_to_pack_into_container)}")
            final_allow_stacking = allow_stacking
            if is_collars:
                final_allow_stacking = stack_collars
            elif is_tray_load:
                final_allow_stacking = stack_trays
            main_bin = Bin(name=container_data.get('name', 'Container'), width=Decimal(container_data.get('width', 0)), height=Decimal(container_data.get('height', 0)), depth=Decimal(container_data.get('depth', 0)), max_weight=Decimal(container_data.get('max_weight', 999999)), allow_stacking=final_allow_stacking)
            final_packer.add_bin(main_bin)
            
            # --- CONTAINER LOADING LOGIC (Mixed vs Pure) ---
            items_for_container = []
            remaining_unfitted = []

            if config.get('mixed_containers', True):
                 # Mix everything
                 items_for_container = items_to_pack_into_container
            else:
                 # No Mixing: Can only pack ONE type into this container.
                 groups = {}
                 for it in items_to_pack_into_container:
                      groups.setdefault(getattr(it, 'content_type', 'unknown'), []).append(it)
                 if groups:
                      first_group_key = list(groups.keys())[0]
                      items_for_container = groups[first_group_key]
                      for k, v in groups.items():
                           if k != first_group_key:
                                remaining_unfitted.extend(v)
            
            # --- START TEMPLATE LOGIC ---
            # Recognize combinations of Standard Container + Standard Pallet
            applied_template = False
            
            # Container Dims
            c_l = Decimal(container_data.get('width', 0))
            c_d = Decimal(container_data.get('depth', 0))
            
            # 1. Component Detection
            is_20ft = (5850 <= c_l < 6000 and 2300 <= c_d < 2400)
            is_40ft = (12000 <= c_l < 12100 and 2300 <= c_d < 2400)
            is_40pw = (12000 <= c_l < 12100 and 2440 <= c_d < 2550)
            
            p_w, p_d = Decimal(0), Decimal(0)
            is_europallet = False
            is_american = False
            
            if p_conf_outer:
                p_w, p_d = p_conf_outer['w'], p_conf_outer['d']
                is_europallet = (p_w == 1200 and p_d == 800)
                is_american = (p_w == 1200 and p_d == 1000)
            
            # 2. Template Selection & Capacity
            tpl_name = None
            max_cap = 0
            
            # FOOTPRINT DETECTION: Europallet and Collars share 120x80.
            is_euro_footprint = (is_europallet or (is_collars and p_d == 800))
            
            if is_20ft:
                if is_american:   tpl_name, max_cap = "20ft_american", 10
                elif is_euro_footprint: tpl_name, max_cap = "20ft_europallet", 11
            elif is_40ft:
                if is_american:   tpl_name, max_cap = "40ft_american", 21
                elif is_euro_footprint: tpl_name, max_cap = "40ft_europallet", 25
            elif is_40pw:
                if is_american:   tpl_name, max_cap = "40pw_american", 24
                elif is_euro_footprint: tpl_name, max_cap = "40pw_europallet", 30

            # 3. Decision Logic
            is_max = config.get('maximize', False)
            load_count = len(items_for_container)
            
            if tpl_name:
                # If maximizing, we ALWAYS use the template to its capacity
                if is_max:
                    applied_template = True
                    target_count = max_cap
                else:
                    # Always use template when available.
                    # If load exceeds capacity, fill template and mark the rest as unfitted.
                    applied_template = True
                    target_count = min(load_count, max_cap)
            
            if applied_template:
                template_coords = []
                
                if tpl_name == "20ft_american":
                    # 10 american (5 sets of alternating pairs)
                    for i in range(5):
                        cycle = i // 2
                        is_second_in_cycle = i % 2 == 1
                        if not is_second_in_cycle:
                            template_coords.append({'x': Decimal(cycle * 2200), 'z': Decimal(0), 'rot': RotationType.RT_DHW}) # 1000x1200
                            template_coords.append({'x': Decimal(cycle * 2200), 'z': Decimal(1350), 'rot': RotationType.RT_WHD}) # 1200x1000
                        else:
                            template_coords.append({'x': Decimal(cycle * 2200 + 1000), 'z': Decimal(0), 'rot': RotationType.RT_WHD}) # 1200x1000
                            template_coords.append({'x': Decimal(cycle * 2200 + 1200), 'z': Decimal(1150), 'rot': RotationType.RT_DHW}) # 1000x1200
                
                elif tpl_name == "20ft_europallet":
                    # 11 europallets (7 side + 4 face)
                    for i in range(7):
                        template_coords.append({'x': Decimal(i * 800), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                    for i in range(4):
                        template_coords.append({'x': Decimal(i * 1200), 'z': Decimal(1550), 'rot': RotationType.RT_WHD})

                elif tpl_name == "40ft_american":
                    # 21 american (10 pairs + 1 end)
                    for i in range(10):
                        cycle = i // 2
                        is_second_in_cycle = i % 2 == 1
                        if not is_second_in_cycle:
                            template_coords.append({'x': Decimal(cycle * 2200), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                            template_coords.append({'x': Decimal(cycle * 2200), 'z': Decimal(1350), 'rot': RotationType.RT_WHD})
                        else:
                            template_coords.append({'x': Decimal(cycle * 2200 + 1000), 'z': Decimal(0), 'rot': RotationType.RT_WHD})
                            template_coords.append({'x': Decimal(cycle * 2200 + 1200), 'z': Decimal(1150), 'rot': RotationType.RT_DHW})
                    template_coords.append({'x': Decimal(11000), 'z': Decimal(0), 'rot': RotationType.RT_DHW})

                elif tpl_name == "40ft_europallet":
                    # 25 europallets
                    for i in range(15):
                        template_coords.append({'x': Decimal(i * 800), 'z': Decimal(0), 'rot': RotationType.RT_DHW})
                    for i in range(10):
                        template_coords.append({'x': Decimal(i * 1200), 'z': Decimal(1550), 'rot': RotationType.RT_WHD})

                elif tpl_name == "40pw_american":
                    # 24 american (2x12) - Both rows oriented 120 wide (RT_WHD)
                    for i in range(12):
                        template_coords.append({'x': Decimal(i * 1000), 'z': Decimal(0), 'rot': RotationType.RT_DHW})   # 1000x1200
                        template_coords.append({'x': Decimal(i * 1000), 'z': Decimal(1250), 'rot': RotationType.RT_DHW}) # 1000x1200
                
                elif tpl_name == "40pw_europallet":
                    # 30 europallets (2x15)
                    for i in range(15):
                        template_coords.append({'x': Decimal(i * 800), 'z': Decimal(0), 'rot': RotationType.RT_DHW})   # 800x1200
                        template_coords.append({'x': Decimal(i * 800), 'z': Decimal(1250), 'rot': RotationType.RT_DHW}) # 800x1200

                # STACKING LOGIC FOR MAXIMIZATION
                if applied_template and is_max and is_stackable and config.get('maximize') and allow_stacking:
                    # Calculate how many layers fit vertically
                    total_h_pallet = p_conf_outer['h'] # This is total height of one unit
                    container_h = Decimal(container_data.get('height', 0))
                    
                    if container_h > 0 and total_h_pallet > 0:
                        layers = int(container_h // total_h_pallet)
                        if layers > 1:
                            debug_log(f"DEBUG: Stacking Detected. Layers: {layers}")
                            # Expand template_coords for multiple layers
                            base_coords = list(template_coords)
                            template_coords = []
                            
                            for layer in range(layers):
                                y_pos = Decimal(layer) * total_h_pallet
                                for coord in base_coords:
                                    # Clone coord and add Y
                                    new_coord = coord.copy()
                                    new_coord['y'] = y_pos
                                    template_coords.append(new_coord)
                            
                            # Update target count to reflect full capacity
                            target_count = len(template_coords)
                            debug_log(f"DEBUG: New Target Count with Stacking: {target_count}")

                debug_log(f"DEBUG: Applying Template {tpl_name} for {target_count} slots.")
                result_bin = main_bin
                packed_top_level = []
                unfitted_final = remaining_unfitted
                
                # Take items to fill target_count
                items_to_use = items_for_container[:target_count]
                
                # Default limiting factor is None (Optimistic). 
                # If we have more items than slots, then Volume is the limit.
                limiting_factor = 'volume' if len(items_for_container) > target_count else None

                current_total_weight = Decimal(0)
                container_max_w = Decimal(container_data.get('max_weight', 999999))
                if container_max_w <= 0: container_max_w = Decimal(999999999)

                for idx, coord in enumerate(template_coords):
                    _raise_if_logistics_calc_cancelled(job_id)
                    if idx >= target_count: break
                    if idx >= len(items_to_use): break
                    
                    it = items_to_use[idx]
                    
                    # Weight Check
                    if current_total_weight + it.weight > container_max_w:
                        debug_log(f"DEBUG: Weight Limit Reached! Cur: {current_total_weight} + Item: {it.weight} > Max: {container_max_w}")
                        limiting_factor = 'weight'
                        # Add remaining items to unfitted since we stopped early
                        if len(items_to_use) > idx:
                             unfitted_final.extend(items_to_use[idx:])
                        break

                    current_total_weight += it.weight
                    it.position = [coord['x'], coord.get('y', Decimal(0)), coord['z']]
                    it.rotation_type = coord['rot']
                    dims = it.get_dimension()
                    it.lx, it.rx = it.position[0], it.position[0] + dims[0]
                    it.ly, it.ry = it.position[1], it.position[1] + dims[1]
                    it.lz, it.rz = it.position[2], it.position[2] + dims[2]
                    
                    result_bin.items.append(it)
                    packed_top_level.append(it)
                
                # Any leftover items in items_for_container that didn't fit in template are unfitted
                if len(items_for_container) > len(template_coords):
                    unfitted_final.extend(items_for_container[len(template_coords):])
                    
            else:
                # Generic fallback packing
                for it in items_for_container: final_packer.add_item(it)
                debug_log(f"DEBUG: Running generic container solver with {len(items_for_container)} top-level units")
                final_packer.pack(bigger_first=True)
                result_bin = final_packer.bins[0]
                unfitted_final = final_packer.unfit_items + remaining_unfitted
                packed_top_level = result_bin.items
        else:
            # NO CONTAINER (Virtual Floor) - No mixed constraints apply (infinite space)
            debug_log(f"DEBUG: Final packing on virtual floor. Top-level units: {len(items_to_pack_into_container)}")
            virtual_allow_stacking = allow_stacking
            if is_collars:
                virtual_allow_stacking = stack_collars
            elif is_tray_load:
                virtual_allow_stacking = stack_trays
            if config.get('maximize') and p_conf_outer:
                # Check if this is multi-item maximization (each getting their own pallet)
                is_multi_item_max = not config.get('mixed_pallets', True)
                
                if is_multi_item_max:
                    # Multiple Items, Each Maximized on Own Pallet
                    # Use LARGE virtual floor to accommodate all pallets side by side
                    result_bin = Bin("Virtual Floor", Decimal(20000), Decimal(5000), Decimal(20000), Decimal(999999), allow_stacking=allow_stacking)
                else:
                    # Single Pallet Maximization: Restrict bin to the pallet dimensions
                    
                    # COLLARS Case: Use fixed total height
                    if is_collars:
                         limit_h = p_conf['h'] # Total Height (Base + Walls)
                         # For packing, we need the bin to be big enough to hold the 'Pallet Item'
                         # The 'Pallet Item' created by pallet factory has height = total_h
                         # So bin height must be at least total_h.
                         result_bin = Bin("Single Pallet Limit", p_conf_outer['w'], limit_h, p_conf_outer['d'], Decimal(999999), allow_stacking=virtual_allow_stacking)
                    else:
                        # STANDARD Case: Use User Max Height (Default 180cm)
                        user_max_h = float(config.get('max_pallet_height', 0))
                        target_total_h = Decimal(user_max_h) if user_max_h > 0 else Decimal(1800)
                        
                        # For packing, Bin Height must be the TOTAL HEIGHT to accommodate the pallet + cargo
                        limit_h = target_total_h
                        
                        result_bin = Bin("Single Pallet Limit", p_conf_outer['w'] + Decimal(1), limit_h + Decimal(1), p_conf_outer['d'] + Decimal(1), Decimal(999999), allow_stacking=virtual_allow_stacking)
            elif config.get('maximize') and is_tray_load and tray_conf_outer:
                result_bin = Bin("Single Tray Limit", tray_conf_outer['w'] + Decimal(1), tray_conf_outer['h'] + Decimal(1), tray_conf_outer['d'] + Decimal(1), Decimal(999999), allow_stacking=False)
            else:
                result_bin = Bin("Virtual Floor", Decimal(20000), Decimal(5000), Decimal(20000), Decimal(999999), allow_stacking=virtual_allow_stacking)
            final_packer.add_bin(result_bin)
            for it in items_to_pack_into_container: final_packer.add_item(it)
            debug_log(f"DEBUG: Running virtual-floor solver with {len(items_to_pack_into_container)} top-level units")
            final_packer.pack(bigger_first=True)
            result_bin = final_packer.bins[0]
            unfitted_final = final_packer.unfit_items
            packed_top_level = result_bin.items

        progress(78, 'Empaquetado principal resuelto...', 'solver')

        flattened_packed_items = []
        kanban_visual_threshold = 250000
        rot_map_90 = {RotationType.RT_WHD: RotationType.RT_DHW, RotationType.RT_DHW: RotationType.RT_WHD, RotationType.RT_HWD: RotationType.RT_DWH, RotationType.RT_DWH: RotationType.RT_HWD, RotationType.RT_HDW: RotationType.RT_WDH, RotationType.RT_WDH: RotationType.RT_HDW}

        total_kanban_piece_count = 0
        for pkg in packed_top_level:
            total_kanban_piece_count += count_leaf_items(pkg)

        simplify_kanban_visual = total_kanban_piece_count > kanban_visual_threshold

        def dims_for_rotation(unit, rotation_type):
            base_w = float(getattr(unit, 'width', 0) or 0)
            base_h = float(getattr(unit, 'height', 0) or 0)
            base_d = float(getattr(unit, 'depth', 0) or 0)
            rotation_dims = {
                RotationType.RT_WHD: (base_w, base_h, base_d),
                RotationType.RT_HWD: (base_h, base_w, base_d),
                RotationType.RT_HDW: (base_h, base_d, base_w),
                RotationType.RT_DHW: (base_d, base_h, base_w),
                RotationType.RT_DWH: (base_d, base_w, base_h),
                RotationType.RT_WDH: (base_w, base_d, base_h),
            }
            return rotation_dims.get(rotation_type, (base_w, base_h, base_d))

        def transform_nested_position(parent_unit, local_x, local_z, child_local_w, child_local_d, abs_x, abs_z, shell_offset, is_rotated):
            if not is_rotated:
                return abs_x + local_x + shell_offset, abs_z + local_z + shell_offset

            parent_inner_w = max(float(getattr(parent_unit, 'width', 0) or 0) - (shell_offset * 2), 0)
            rotated_x = abs_x + local_z + shell_offset
            rotated_z = abs_z + max(parent_inner_w - local_x - child_local_w, 0) + shell_offset
            return rotated_x, rotated_z

        def emit_virtual_kanban_items(unit, abs_x, abs_y, abs_z, is_rotated, shell_offset):
            layout_id = getattr(unit, 'kanban_layout_id', None)
            if layout_id and layout_id in kanban_layout_registry:
                emitted = 0
                for entry in kanban_layout_registry.get(layout_id, []):
                    local_x = float(entry.get('x', 0) or 0)
                    local_y = float(entry.get('y', 0) or 0)
                    local_z = float(entry.get('z', 0) or 0)
                    entry_w = float(entry.get('w', 0) or 0)
                    entry_h = float(entry.get('h', 0) or 0)
                    entry_d = float(entry.get('d', 0) or 0)
                    entry_rt = entry.get('rt', RotationType.RT_WHD)

                    if is_rotated:
                        child_x, child_z = transform_nested_position(
                            unit,
                            local_x,
                            local_z,
                            entry_w,
                            entry_d,
                            abs_x,
                            abs_z,
                            shell_offset,
                            True
                        )
                        final_w, final_h, final_d = entry_d, entry_h, entry_w
                        final_rt = rot_map_90.get(entry_rt, entry_rt)
                    else:
                        child_x = abs_x + local_x + shell_offset
                        child_z = abs_z + local_z + shell_offset
                        final_w, final_h, final_d = entry_w, entry_h, entry_d
                        final_rt = entry_rt
                    child_y = abs_y + float(getattr(unit, 'base_height', 0)) + local_y

                    flattened_packed_items.append({
                        'name': entry.get('name', getattr(unit, 'leaf_item_name', 'KANBAN_ITEM')),
                        'x': child_x,
                        'y': child_y,
                        'z': child_z,
                        'w': final_w,
                        'h': final_h,
                        'd': final_d,
                        'rt': final_rt,
                        'weight': float(entry.get('weight', 0) or 0)
                    })
                    emitted += 1
                return emitted

            content_map = getattr(unit, 'content_map', {}) or {}
            template = getattr(unit, 'leaf_item_template', None)
            if not content_map or template is None:
                return 0

            piece_w = float(template.width)
            piece_h = float(template.height)
            piece_d = float(template.depth)
            piece_weight = float(getattr(template, 'weight', 0) or 0)
            if piece_w <= 0 or piece_h <= 0 or piece_d <= 0:
                return 0

            inner_w = max(float(unit.width) - (shell_offset * 2), piece_w)
            inner_d = max(float(unit.depth) - (shell_offset * 2), piece_d)
            usable_h = max(float(getattr(unit, 'visual_height', float(unit.height))) - float(getattr(unit, 'base_height', 0)), piece_h)

            slots_x = max(int(inner_w // piece_w), 1)
            slots_z = max(int(inner_d // piece_d), 1)
            slots_per_layer = max(slots_x * slots_z, 1)
            max_layers = max(int(usable_h // piece_h), 1)

            emitted = 0
            for item_name, qty in content_map.items():
                try:
                    qty_int = int(qty or 0)
                except Exception:
                    qty_int = 0
                for idx in range(max(qty_int, 0)):
                    layer_idx = min(idx // slots_per_layer, max_layers - 1)
                    slot_idx = idx % slots_per_layer
                    row_idx = slot_idx // slots_x
                    col_idx = slot_idx % slots_x

                    local_x = col_idx * piece_w
                    local_z = row_idx * piece_d
                    local_y = layer_idx * piece_h

                    if is_rotated:
                        child_x, child_z = transform_nested_position(
                            unit,
                            local_x,
                            local_z,
                            piece_w,
                            piece_d,
                            abs_x,
                            abs_z,
                            shell_offset,
                            True
                        )
                        rt = RotationType.RT_DHW
                    else:
                        child_x = abs_x + local_x + shell_offset
                        child_z = abs_z + local_z + shell_offset
                        rt = RotationType.RT_WHD
                    child_y = abs_y + float(getattr(unit, 'base_height', 0)) + local_y

                    flattened_packed_items.append({
                        'name': item_name,
                        'x': child_x,
                        'y': child_y,
                        'z': child_z,
                        'w': piece_w,
                        'h': piece_h,
                        'd': piece_d,
                        'rt': rt,
                        'weight': piece_weight
                    })
                    emitted += 1
            return emitted

        def serialize_unit(unit, abs_x, abs_y, abs_z):
            unit_rt = getattr(unit, 'effective_rotation_type', unit.rotation_type)
            unit_dims = dims_for_rotation(unit, unit_rt)

            if hasattr(unit, 'inner_items'):
                total_h = float(unit_dims[1])
                base_obj = {
                    'name': 'PALLET_BASE',
                    'x': abs_x, 'y': abs_y, 'z': abs_z,
                    'w': float(unit_dims[0]), 'h': float(unit.base_height), 'd': float(unit_dims[2]),
                    'total_h': total_h,
                    'rt': unit_rt,
                    'is_pallet': True,
                    'pallet_type': getattr(unit, 'pallet_type', 'unknown')
                }
                if hasattr(unit, 'wall_thickness'):
                    base_obj['wall_thickness'] = getattr(unit, 'wall_thickness', 0)
                if getattr(unit, 'pallet_type', '') in ('collars', 'collars_120x100', 'tray'):
                    base_obj['h_visual'] = getattr(unit, 'visual_height', total_h)
                flattened_packed_items.append(base_obj)

                is_rotated = (unit_rt == RotationType.RT_DHW)
                shell_offset = float(getattr(unit, 'wall_thickness', 19)) if getattr(unit, 'pallet_type', '') in ('collars', 'collars_120x100', 'tray') else 0

                for child in unit.inner_items:
                    c_dims = child.get_dimension()
                    ix, iy, iz = float(child.position[0]), float(child.position[1]), float(child.position[2])
                    if is_rotated:
                        child_x, child_z = transform_nested_position(
                            unit,
                            ix,
                            iz,
                            float(c_dims[0]),
                            float(c_dims[2]),
                            abs_x,
                            abs_z,
                            shell_offset,
                            True
                        )
                        child_y = abs_y + float(unit.base_height) + iy
                        child.effective_rotation_type = rot_map_90.get(child.rotation_type, child.rotation_type)
                    else:
                        child_x = abs_x + ix + shell_offset
                        child_z = abs_z + iz + shell_offset
                        child_y = abs_y + float(unit.base_height) + iy
                        child.effective_rotation_type = child.rotation_type

                    if hasattr(child, 'inner_items'):
                        serialize_unit(child, child_x, child_y, child_z)
                    else:
                        final_w, final_h, final_d = dims_for_rotation(child, child.effective_rotation_type)
                        flattened_packed_items.append({'name': child.name, 'x': child_x, 'y': child_y, 'z': child_z, 'w': final_w, 'h': final_h, 'd': final_d, 'rt': child.effective_rotation_type, 'weight': float(getattr(child, 'weight', 0) or 0)})
                if hasattr(unit, 'content_map'):
                    base_obj['content_map'] = dict(getattr(unit, 'content_map', {}) or {})
                    base_obj['leaf_total_count'] = int(getattr(unit, 'leaf_total_count', count_leaf_items(unit)) or 0)
                    base_obj['leaf_total_weight'] = float(getattr(unit, 'leaf_total_weight', sum_leaf_weight(unit)) or 0)
                    base_obj['kanban_qty_per_tray'] = int(getattr(unit, 'kanban_qty_per_tray', 0) or 0)
                    if hasattr(unit, 'kanban_layout_id'):
                        base_obj['kanban_layout_id'] = getattr(unit, 'kanban_layout_id', '')
                    template = getattr(unit, 'leaf_item_template', None)
                    if template is not None:
                        base_obj['leaf_item_name'] = getattr(template, 'name', '')
                        base_obj['leaf_item_dims'] = {
                            'w': float(getattr(template, 'width', 0) or 0),
                            'h': float(getattr(template, 'height', 0) or 0),
                            'd': float(getattr(template, 'depth', 0) or 0)
                        }
                        base_obj['leaf_item_weight'] = float(getattr(template, 'weight', 0) or 0)
                    if len(getattr(unit, 'inner_items', []) or []) == 0:
                        if simplify_kanban_visual:
                            base_obj['kanban_simplified'] = True
                        else:
                            emitted = emit_virtual_kanban_items(unit, abs_x, abs_y, abs_z, is_rotated, shell_offset)
                            debug_log(f"DEBUG: Virtual Kanban items emitted for {unit.name}: {emitted}")
            else:
                flattened_packed_items.append({'name': unit.name, 'x': abs_x, 'y': abs_y, 'z': abs_z, 'w': float(unit_dims[0]), 'h': float(unit_dims[1]), 'd': float(unit_dims[2]), 'rt': unit_rt, 'weight': float(getattr(unit, 'weight', 0) or 0)})

        for pkg in packed_top_level:
            serialize_unit(pkg, float(pkg.position[0]), float(pkg.position[1]), float(pkg.position[2]))
        
        all_unfitted = unfitted_from_palletizing + unfitted_final
        unfitted_serialized = [{'name': i.name, 'w': float(i.width), 'h': float(i.height), 'd': float(i.depth)} for i in all_unfitted]

        def _build_freight_basis(top_level_units):
            packages = []
            total_weight_kg = 0.0
            consolidated_volume_m3 = 0.0
            max_x = 0.0
            max_y = 0.0
            max_z = 0.0

            for unit in top_level_units:
                unit_rt = getattr(unit, 'effective_rotation_type', unit.rotation_type)
                unit_dims = dims_for_rotation(unit, unit_rt)
                unit_w = float(unit_dims[0] or 0)
                unit_h = float(unit_dims[1] or 0)
                unit_d = float(unit_dims[2] or 0)
                unit_weight = float(getattr(unit, 'weight', 0) or 0)
                unit_volume_m3 = max((unit_w * unit_h * unit_d) / 1000000000.0, 0.0)
                unit_x = float(getattr(unit, 'position', [0, 0, 0])[0] or 0)
                unit_y = float(getattr(unit, 'position', [0, 0, 0])[1] or 0)
                unit_z = float(getattr(unit, 'position', [0, 0, 0])[2] or 0)

                max_x = max(max_x, unit_x + unit_w)
                max_y = max(max_y, unit_y + unit_h)
                max_z = max(max_z, unit_z + unit_d)
                total_weight_kg += unit_weight
                consolidated_volume_m3 += unit_volume_m3

                unit_kind = 'loose'
                pallet_type = str(getattr(unit, 'pallet_type', '') or '').strip().lower()
                if hasattr(unit, 'inner_items'):
                    if pallet_type == 'tray':
                        unit_kind = 'tray'
                    elif pallet_type in ('collars', 'collars_120x100'):
                        unit_kind = 'collars'
                    else:
                        unit_kind = 'pallet'

                packages.append({
                    'kind': unit_kind,
                    'name': str(getattr(unit, 'name', unit_kind.upper()) or unit_kind.upper()),
                    'dims_mm': {
                        'w': round(unit_w, 3),
                        'h': round(unit_h, 3),
                        'd': round(unit_d, 3),
                    },
                    'gross_weight_kg': round(unit_weight, 3),
                    'outer_volume_m3': round(unit_volume_m3, 6),
                })

            basis_metrics = _logistics_load_basis(total_weight_kg, consolidated_volume_m3)
            chargeable_weight_kg = total_weight_kg
            load_fraction = basis_metrics['load_fraction']
            bbox_volume_m3 = max((max_x * max_y * max_z) / 1000000000.0, 0.0)

            return {
                'source': 'post_packing_top_level_units',
                'package_count': len(packages),
                'packages': packages,
                'actual_weight_kg': round(total_weight_kg, 3),
                'consolidated_volume_m3': round(consolidated_volume_m3, 6),
                'chargeable_weight_kg': round(chargeable_weight_kg, 3),
                'load_fraction': round(load_fraction, 6),
                'weight_fraction': basis_metrics['weight_fraction'],
                'volume_fraction': basis_metrics['volume_fraction'],
                'dominant_basis': basis_metrics['dominant_basis'],
                'standard_payload_kg': round(float(LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG), 3),
                'standard_volume_m3': round(float(LOGISTICS_FREIGHT_STANDARD_VOLUME_M3), 6),
                'shipment_bbox_dims_mm': {
                    'w': round(max_x, 3),
                    'h': round(max_y, 3),
                    'd': round(max_z, 3),
                },
                'shipment_bbox_volume_m3': round(bbox_volume_m3, 6),
            }

        vol_bin = float(result_bin.width * result_bin.height * result_bin.depth)
        
        # EFFICIENCY CALCULATION REFINEMENT
        if container_type_key == 'none' and p_conf_outer:
             # Override vol_bin to reflect NET AVAILABLE VOLUME
             if is_collars:
                  # Net Volume = Inner Width * Inner Depth * Inner Loading Height
                  vol_bin = float(p_conf['w'] * p_conf['d'] * p_conf['loading_h'])
             else:
                  # Standard Pallet:
                  # Net Volume = Outer Width * Outer Depth * (Total Height - Base Height)
                  # Re-read configs (safe)
                  user_max_h = float(config.get('max_pallet_height', 0))
                  target_total_h = Decimal(user_max_h) if user_max_h > 0 else Decimal(1800)
                  base_h = p_conf.get('h', Decimal(150))
                  net_h = target_total_h - base_h
                  if net_h < 0: net_h = Decimal(1)
                  vol_bin = float(p_conf_outer['w'] * p_conf_outer['d'] * net_h)
            
             vol_items = 0
             for it in packed_top_level:
                if hasattr(it, 'inner_items'):
                    vol_items += sum_leaf_volume(it)
                else:
                    vol_items += float(it.width * it.height * it.depth)
        else:
            vol_items = sum([float(it.width * it.height * it.depth) for it in packed_top_level])
        efficiency = (vol_items / vol_bin * 100) if vol_bin > 0 else 0

        res = {
            'status': 'success',
            'bin_name': result_bin.name,
            'bin_dims': {'w': float(result_bin.width), 'h': float(result_bin.height), 'd': float(result_bin.depth)},
            'packed_items': flattened_packed_items,
            'kanban_layouts': kanban_layout_registry if simplify_kanban_visual else {},
            'unfitted_items': unfitted_serialized,
            'unfitted_count': len(all_unfitted),
            'unfitted_from_palletizing_count': len(unfitted_from_palletizing),
            'unfitted_final_count': len(unfitted_final),
            'limiting_factor': locals().get('limiting_factor', None),
            'efficiency': round(efficiency, 2),
            'packed_top_level_count': len(packed_top_level),
            # NEW KPIs
            'kpis': {
                'container_vol': round(efficiency, 2), # Already calculated
                'pallet_vol_avg': None, # To be calc below
                'container_weight': 0,
                'pallet_weight_avg': None,
                'tray_vol_avg': None,
                'tray_weight_avg': None
            },
            'container_info': {
                'length': float(container_data.get('length', container_data.get('depth', 0))),
                'width': float(container_data.get('width', 0)),
                'height': float(container_data.get('height', 0))
            },
            'grouped_pallets': [],
            'freight_basis': _build_freight_basis(packed_top_level)
        }

        # --- KPI & GROUPING LOGIC ---
        debug_log(f"DEBUG: Starting KPI Logic. Packed Top Level Count: {len(packed_top_level)}")
        
        # 1. Container Weight KPI
        total_weight_loaded = sum([float(i.weight) for i in packed_top_level])
        container_max_w = float(result_bin.max_weight)
        debug_log(f"DEBUG: Cont Max W: {container_max_w}, Total Loaded W: {total_weight_loaded}")
        
        if container_max_w > 0:
            res['kpis']['container_weight'] = round((total_weight_loaded / container_max_w) * 100, 2)

        # 2. Pallet Metrics (Vol Efficiency & Weight) & Grouping
        # distinct_pallets hash map -> { signature: {count: N, data: p_item, items: {name: qty}} }
        pallet_groups = {}
        
        pallet_vol_efficiencies = []
        pallet_weight_ratios = []
        tray_vol_efficiencies = []
        tray_weight_ratios = []

        # Iterate TOP LEVEL packed items (which are Pallets or Loose Items)
        packed_loose_items = []
        packed_top_level_trays = []

        for p_item in packed_top_level:
            is_pallet = (pallet_type_key != 'none') and (getattr(p_item, 'is_pallet', False) or hasattr(p_item, 'inner_items'))
            is_top_level_tray = (getattr(p_item, 'pallet_type', '') == 'tray') and hasattr(p_item, 'inner_items')
            debug_log(f"DEBUG: Item {p_item.name} IsPallet: {is_pallet}")
            
            if is_pallet:
                tray_children = [sub for sub in p_item.inner_items if hasattr(sub, 'inner_items') and getattr(sub, 'pallet_type', '') == 'tray']
                if tray_children and tray_conf:
                    tray_usable_vol = float(tray_conf['w'] * tray_conf['d'] * tray_conf['loading_h'])
                    tray_max_load_weight = float(tray_conf.get('max_load_weight', Decimal('25')))
                    for tray_unit in tray_children:
                        tray_inner_vol = sum_leaf_volume(tray_unit)
                        if tray_usable_vol > 0:
                            tray_vol_efficiencies.append((tray_inner_vol / tray_usable_vol) * 100)
                        if tray_max_load_weight > 0:
                            tray_net_weight = max(float(tray_unit.weight) - float(tray_conf.get('weight', 0)), 0.0)
                            tray_weight_ratios.append((tray_net_weight / tray_max_load_weight) * 100)

                # Calculate Efficiency for this individual pallet
                # Net Volume of Pallet (Inner)
                # p_conf uses 'h' as Base Height, but 'loading_h' is what we want for volume calc if collars?
                # Actually, easy way: Sum volume of items inside / (W*D*LoadingH)
                
                inner_vol = sum_leaf_volume(p_item)
                
                # Pallet Usable Volume
                # config['pallet_dims'] might not be persistent if we have multiple types, but here we usually have 1 type per calc
                # We can deduce usable volume from the p_item dimensions but p_item height might be total.
                # Let's rely on p_conf if available (it was defined above in _do_pack_internal scope)
                
                pallet_usable_vol = 1
                if p_conf:
                    # p_conf['h'] is Base Height, 'loading_h' is Loading Height (for Collars). 
                    # For standard pallets, loading_h is 0 in definition, calculated dynamically?
                    # Let's use the actual height of the pallet item minus base height?
                    # Or better: if collars, use fixed loading_h. If standard, use (VisualHeight - BaseHeight) or (MaxHeight - BaseHeight).
                    
                    if is_collars:
                         pallet_usable_vol = float(p_conf['w'] * p_conf['d'] * p_conf['loading_h'])
                    else:
                         # Standard: Use the "Max User Height" or "Physical Limit" defined as limit
                         # We calculated `pallet_loading_height` variable earlier!
                         # Accessing it might be tricky if scope is lost, but we are in same func.
                         # `pallet_loading_height` variable is available in local scope.
                         pallet_usable_vol = float(p_conf['w'] * p_conf['d'] * pallet_loading_height)

                if pallet_usable_vol > 0:
                    pallet_vol_efficiencies.append((inner_vol / pallet_usable_vol) * 100)
                
                # Weight Ratio
                curr_w = float(p_item.weight) # Total weight including tare
                # Max weight is in p_conf['max_weight']? No, user_max_w variable.
                # `pallet_max_weight` variable available.
                if pallet_max_weight > 0:
                    pallet_weight_ratios.append((curr_w / float(pallet_max_weight)) * 100)

                # Grouping Signature
                # Signature based on: Content (Items + Qty)
                # We can sort inner items by name/id and build a string
                content_map = collect_content_map(p_item)
                
                # Create sorted signature tuple
                sig_items = sorted(content_map.items()) # [('001', 50), ('002', 10)]
                sig_str = str(sig_items) 
                tray_sig = str(sorted([(float(t.width), float(getattr(t, 'visual_height', t.height)), float(t.depth), count_leaf_items(t)) for t in tray_children]))
                
                # Also include pallet type/dims in signature just in case
                sig = f"{pallet_type_key}_{p_item.width}x{p_item.depth}_{tray_sig}_{sig_str}"

                if sig not in pallet_groups:
                    # Calc Load Weight (Sum of inner items)
                    load_weight = sum(float(sub.weight) for sub in p_item.inner_items)
                    
                    pallet_groups[sig] = {
                        'count': 0,
                        'name_display': f"Pallet {len(pallet_groups) + 1} ({p_item.width}x{p_item.depth})", # Placeholder name
                        'dims': f"{p_item.width}x{p_item.depth}x{p_item.height}", # Outer dims
                        'weight': float(p_item.weight), # Gross Weight
                        'load_weight': float(load_weight), # Net Load Weight
                        'items_map': content_map,
                        'base_height': float(getattr(p_item, 'base_height', 15)),
                        'trays': build_tray_groups(tray_children)
                    }
                pallet_groups[sig]['count'] += 1
            
            elif is_top_level_tray:
                packed_top_level_trays.append(p_item)
                if tray_conf:
                    tray_usable_vol = float(tray_conf['w'] * tray_conf['d'] * tray_conf['loading_h'])
                    tray_max_load_weight = float(tray_conf.get('max_load_weight', Decimal('25')))
                    tray_inner_vol = sum_leaf_volume(p_item)
                    if tray_usable_vol > 0:
                        tray_vol_efficiencies.append((tray_inner_vol / tray_usable_vol) * 100)
                    if tray_max_load_weight > 0:
                        tray_net_weight = max(float(p_item.weight) - float(tray_conf.get('weight', 0)), 0.0)
                        tray_weight_ratios.append((tray_net_weight / tray_max_load_weight) * 100)
            
            else:
                # Loose top level item
                packed_loose_items.append({
                     'name': p_item.name,
                     'dims': f"{p_item.width}x{p_item.height}x{p_item.depth}",
                     'weight': float(p_item.weight),
                     'qty': 1
                })

        # Calculate Averages
        if pallet_vol_efficiencies:
            res['kpis']['pallet_vol_avg'] = round(sum(pallet_vol_efficiencies) / len(pallet_vol_efficiencies), 2)
        if pallet_weight_ratios:
             res['kpis']['pallet_weight_avg'] = round(sum(pallet_weight_ratios) / len(pallet_weight_ratios), 2)
        if tray_vol_efficiencies:
            res['kpis']['tray_vol_avg'] = round(sum(tray_vol_efficiencies) / len(tray_vol_efficiencies), 2)
        if tray_weight_ratios:
            res['kpis']['tray_weight_avg'] = round(sum(tray_weight_ratios) / len(tray_weight_ratios), 2)

        # Finalize Groups List
        final_groups = []
        
        # Add Pallets First
        for sig, data in pallet_groups.items():
            # Format Items for this group
            # We want: Name, Dims(lookup?), Qty Per Pallet, Total Qty
            group_items = []
            for it_name, qty_per in data['items_map'].items():
                # Need dimensions of the item. Only have name 'it_name'.
                
                it_info = next((x for x in items_data if x['id'] == it_name), None)
                dims_str = "???"
                unit_weight = 0
                if it_info:
                    dims_str = f"{it_info.get('w')}x{it_info.get('d')}x{it_info.get('h')}"
                    unit_weight = float(it_info.get('weight', 0))
                
                group_items.append({
                    'name': it_name,
                    'dims': dims_str,
                    'weight': unit_weight, # ADDED WEIGHT
                    'qty_per_pallet': qty_per,
                    'total_qty': qty_per * data['count']
                })
            
            final_groups.append({
                'type': 'pallet',
                'name': data['name_display'],
                'dims': data['dims'],
                'weight_per_pallet': round(data['weight'], 2), # Gross
                'load_weight_per_pallet': round(data['load_weight'], 2), # Net
                'count': data['count'],
                'items': group_items,
                'trays': data.get('trays', [])
            })

        if packed_top_level_trays:
            final_groups.extend(build_tray_groups(packed_top_level_trays))
            
        # Add Loose Items (Grouped by name)
        if packed_loose_items:
            # Group loose items by name
            loose_grouped = {}
            for li in packed_loose_items:
                k = li['name']
                if k not in loose_grouped:
                    loose_grouped[k] = {'name': k, 'dims': li['dims'], 'weight': li['weight'], 'count': 0}
                loose_grouped[k]['count'] += 1
            
            # Add as a special "Loose" group? Or just individual line items at root?
            # User wants: "Si no hay pallet... se mantiene tal cual".
            # If we mixed pallets and loose items?
            # Let's add them as 'loose_item' type entries in the root list
            
            for k, lg in loose_grouped.items():
                final_groups.append({
                    'type': 'loose_item',
                    'name': lg['name'],
                    'dims': lg['dims'],
                    'weight_unit': lg['weight'], # Weight per unit
                    'total_qty': lg['count']
                })

        res['grouped_pallets'] = final_groups
        debug_log(f"DEBUG: Final Grouped Pallets Count: {len(final_groups)}")
        debug_log(f"DEBUG: Grouped Data: {json.dumps(final_groups, default=str)}")
        progress(92, 'Resumiendo resultados...', 'summary')

        # 3. LIMITING FACTOR DETECTION
        # Logic: We can have multiple limiting factors.
        # 1. Violations (>100%) - Critical
        # 2. Bottlenecks (Unfitted items) - If no violations, what prevented packing?
        
        limiting_ids = []
        
        avg_p_w = res['kpis']['pallet_weight_avg'] or 0
        cont_w = res['kpis']['container_weight']
        avg_p_v = res['kpis']['pallet_vol_avg'] or 0
        cont_v = res['kpis']['container_vol']
        avg_t_w = res['kpis']['tray_weight_avg'] or 0
        avg_t_v = res['kpis']['tray_vol_avg'] or 0

        # 1. Check Violations (Hard Limits Exceeded)
        if avg_t_w >= 100.0: limiting_ids.append('kpi-tray-weight')
        if avg_t_v >= 100.0: limiting_ids.append('kpi-tray-vol')
        if avg_p_w >= 100.0: limiting_ids.append('kpi-pallet-weight')
        if cont_w >= 100.0: limiting_ids.append('kpi-container-weight')
        if avg_p_v >= 100.0: limiting_ids.append('kpi-pallet-vol')
        if cont_v >= 100.0: limiting_ids.append('kpi-container-vol')
        
        # 2. If no Hard Violations, check Soft Limits if items couldn't fit
        if len(limiting_ids) == 0 and len(all_unfitted) > 0:
            # Did we fail at Palletizing?
            if len(unfitted_from_palletizing) > 0:
                if is_tray_load:
                    if avg_t_w > 90: limiting_ids.append('kpi-tray-weight')
                    else: limiting_ids.append('kpi-tray-vol')
                elif avg_p_w > 90:
                    limiting_ids.append('kpi-pallet-weight')
                else:
                    limiting_ids.append('kpi-pallet-vol')
            
            # Did we fail at Container Loading? (Even if palletizing failed, container might also be full)
            # Logic: If we have unfitted final items (that were palletized but didn't fit in container)
            if len(unfitted_final) > 0:
                # Container Bottleneck
                if cont_w > 95: limiting_ids.append('kpi-container-weight')
                else: limiting_ids.append('kpi-container-vol')
        
        if container_type_key == 'none' and config.get('maximize'):
            # FIX: In Single Pallet Maximization, there is no container.
            # The "Container Vol/Weight" KPIs are actually measuring the Single Pallet (Bin).
            # But the UI Card for Container is hidden or irrelevant.
            # We must redirect the highlighted KPI to the active level.
            
            new_limits = []
            for limit_id in limiting_ids:
                if limit_id == 'kpi-container-vol':
                    if p_conf_outer:
                        new_limits.append('kpi-pallet-vol')
                    elif is_tray_load and tray_conf_outer:
                        if avg_t_w >= avg_t_v:
                            new_limits.append('kpi-tray-weight')
                        else:
                            new_limits.append('kpi-tray-vol')
                elif limit_id == 'kpi-container-weight':
                    if p_conf_outer:
                        new_limits.append('kpi-pallet-weight')
                    elif is_tray_load and tray_conf_outer:
                        new_limits.append('kpi-tray-weight')
                else:
                    new_limits.append(limit_id)
            limiting_ids = list(set(new_limits)) # Deduplicate

        if container_type_key == 'none' and config.get('maximize') and len(all_unfitted) > 0 and p_conf_outer:
            limiting_ids = ['kpi-pallet-weight'] if avg_p_w >= avg_p_v else ['kpi-pallet-vol']

        if config.get('maximize'):
            if is_tray_load and res['kpis']['tray_vol_avg'] is not None:
                limiting_ids.append('kpi-tray-weight' if avg_t_w >= avg_t_v else 'kpi-tray-vol')
            if p_conf_outer and res['kpis']['pallet_vol_avg'] is not None:
                limiting_ids.append('kpi-pallet-weight' if avg_p_w >= avg_p_v else 'kpi-pallet-vol')
            if container_type_key != 'none':
                limiting_ids.append('kpi-container-weight' if cont_w >= cont_v else 'kpi-container-vol')
            limiting_ids = list(set(limiting_ids))

        limiting_summary = []
        if is_tray_load and res['kpis']['tray_vol_avg'] is not None:
            tray_reasons = []
            if 'kpi-tray-weight' in limiting_ids:
                tray_reasons.append('peso')
            if 'kpi-tray-vol' in limiting_ids:
                tray_reasons.append('volumen')
            if not tray_reasons:
                if avg_t_w >= avg_t_v:
                    tray_reasons.append('peso cercano')
                else:
                    tray_reasons.append('volumen cercano')
            limiting_summary.append({
                'level': 'Bandeja',
                'volume_pct': round(avg_t_v, 2),
                'weight_pct': round(avg_t_w, 2),
                'reasons': tray_reasons,
                'detail': f"Bandeja promedio al {avg_t_v:.2f}% de volumen y {avg_t_w:.2f}% de peso."
            })

        if p_conf_outer and res['kpis']['pallet_vol_avg'] is not None:
            pallet_reasons = []
            if 'kpi-pallet-weight' in limiting_ids:
                pallet_reasons.append('peso')
            if 'kpi-pallet-vol' in limiting_ids:
                pallet_reasons.append('volumen')
            if not pallet_reasons:
                if avg_p_w >= avg_p_v:
                    pallet_reasons.append('peso cercano')
                else:
                    pallet_reasons.append('volumen cercano')
            limiting_summary.append({
                'level': 'Pallet',
                'volume_pct': round(avg_p_v, 2),
                'weight_pct': round(avg_p_w, 2),
                'reasons': pallet_reasons,
                'detail': f"Pallet promedio al {avg_p_v:.2f}% de volumen y {avg_p_w:.2f}% de peso."
            })

        if container_type_key != 'none':
            container_reasons = []
            if 'kpi-container-weight' in limiting_ids:
                container_reasons.append('peso')
            if 'kpi-container-vol' in limiting_ids:
                container_reasons.append('volumen')
            if not container_reasons:
                if cont_w >= cont_v:
                    container_reasons.append('peso cercano')
                else:
                    container_reasons.append('volumen cercano')
            limiting_summary.append({
                'level': 'Contenedor',
                'volume_pct': round(cont_v, 2),
                'weight_pct': round(cont_w, 2),
                'reasons': container_reasons,
                'detail': f"Contenedor al {cont_v:.2f}% de volumen y {cont_w:.2f}% de peso."
            })

        res['limiting_kpis'] = limiting_ids # Return LIST
        res['limiting_summary'] = limiting_summary
        progress(98, 'Resultado listo...')
        
        return res
    except Exception as e:
        import traceback
        traceback.print_exc()
        return {'status': 'error', 'message': str(e)}



# Determine absolute path to this script's directory

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
SCRIPT_DIR_PATH = Path(SCRIPT_DIR)
LOGISTICS_RECORDS_FILE = os.path.join(SCRIPT_DIR, 'logistics_records.json')
COTIZACION_RECORDS_FILE = os.path.join(SCRIPT_DIR, 'cotizacion_records.json')
LOGISTICS_DEFAULT_FOLDER = 'Sin Carpeta'
COTIZACION_DEFAULT_FOLDER = 'Sin Carpeta'

# Initialize Flask with explicit folder paths

if os.environ.get("BPB_DEBUG_STARTUP_LOG", "").strip() == "1":
    try:
        with open(SCRIPT_DIR_PATH / "global_debug_startup.log", "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()}: App starting from {os.getcwd()}\n")
    except Exception:
        pass

app = Flask(__name__, 
            template_folder=os.path.join(SCRIPT_DIR, 'templates'),
            static_folder=os.path.join(SCRIPT_DIR, 'static'))

app.config['JSON_SORT_KEYS'] = False
app.config['JSON_AS_ASCII'] = False
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0

app.secret_key = 'super_secret_key_bpb_group_2026'

# Force UTF-8 encoding in all responses
@app.after_request
def add_utf8_header(response):
    if 'Content-Type' in response.headers:
        content_type = response.headers['Content-Type']
        if 'charset' not in content_type.lower():
            response.headers['Content-Type'] = content_type + '; charset=utf-8'
    if request.path.startswith('/static/'):
        response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response


# CONFIGURATION

BASE_DIR = resolve_control_base_dir(SCRIPT_DIR_PATH)
os.environ.setdefault("BPB_BASE_DIR", str(BASE_DIR))
OFFICE_ROOT = resolve_oficina_root(BASE_DIR)
WORKSPACE_ROOT = resolve_workspace_root(BASE_DIR)
PROJECT_COSTOS_ANALYSIS_ROOT = resolve_project_costos_analysis_root(BASE_DIR)
ACTIVITY_CODE_DIR = resolve_activity_codigos_dir(BASE_DIR)
QUALITY_ROOT = resolve_quality_root(BASE_DIR)
R016_REGISTROS_DIR = resolve_r016_dir(BASE_DIR)

print(f"DEBUG: Using resolved BASE_DIR: {BASE_DIR}")


def _resolve_cotizacion_indices_files():
    relative_paths = {
        'produccion': Path('Analisis de Costos Produccion/Analisis de Costos Produccion.xlsm'),
        'ensamble': Path('Analisis de Costos Ensamble/Analisis de Costos Ensamble.xlsm'),
        'embalaje': Path('Analisis de Costos Embalaje/Analisis de Costos Embalaje.xlsm'),
        'deposito_logistica': Path('Analisis de Costos Deposito y Logistica/Analisis de Costos Deposito y Logistica.xlsm')
    }

    resolved = {}
    for key, relative_path in relative_paths.items():
        resolved[key] = PROJECT_COSTOS_ANALYSIS_ROOT / relative_path

    return resolved


COTIZACION_INDICES_FILES = _resolve_cotizacion_indices_files()


if os.environ.get("BPB_DEBUG_STARTUP_LOG", "").strip() == "1":
    try:
        force_log = BASE_DIR / "Codigos/debug_force.txt"
        with open(force_log, "a", encoding="utf-8") as f:
            f.write(f"{datetime.now()}: APP STARTED. CWD={os.getcwd()}\n")
        print(f"DEBUG: startup log written to {force_log}")
    except Exception as exc:
        print(f"DEBUG: failed to write startup log: {exc}")

def _resolve_iso_root() -> Path:
    return resolve_iso_docs_root(BASE_DIR)


def _resolve_iso_subdir(root: Path, prefix: str, preferred_name: str) -> Path:
    preferred = root / preferred_name
    try:
        if preferred.exists():
            return preferred
    except Exception:
        pass
    try:
        for d in root.iterdir():
            if d.is_dir() and d.name.upper().startswith(prefix.upper()):
                return d
    except Exception:
        pass
    return preferred


ISO_DOCS_ROOT = _resolve_iso_root()
ISO_CODE_ROOT = resolve_iso_code_root(BASE_DIR)
ISO_R01901_DOCS_DIR = _resolve_iso_subdir(ISO_DOCS_ROOT, "R019-01", "R019-01- Datos de entrada")
ISO_R01902_DOCS_DIR = _resolve_iso_subdir(ISO_DOCS_ROOT, "R019-02", "R019-02")
ISO_R01904_DOCS_DIR = _resolve_iso_subdir(ISO_DOCS_ROOT, "R019-04", "R019-04")
ISO_R01903_BASE_NAME = "R019-03"


def _find_iso_r01903_target() -> Optional[Path]:
    iso_root = ISO_DOCS_ROOT
    base_name = ISO_R01903_BASE_NAME
    candidates = []
    for ext in ("xlsm", "xlsx", "xls"):
        candidates.append(iso_root / f"{base_name}.{ext}")
    target = next((p for p in candidates if p.exists()), None)
    if target is not None:
        return target
    matches = list(iso_root.glob(f"{base_name}*"))
    return matches[0] if matches else None


def _parse_iso_registry_int(value) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    match = re.search(r"(\d+)", text)
    if not match:
        return None
    try:
        return int(match.group(1))
    except Exception:
        return None


def _extract_bp_year_seq(value) -> Optional[Tuple[str, int]]:
    text = str(value or "").strip().upper()
    if not text:
        return None
    match = re.search(r"BP-?(\d{2})(\d{3})", text)
    if not match:
        return None
    try:
        return match.group(1), int(match.group(2))
    except Exception:
        return None


def _compute_iso_next_registry() -> dict:
    year_two = datetime.now().strftime("%y")
    last_num = 174
    last_seq = 0
    source = ""
    source_details = []

    target = _find_iso_r01903_target()
    if target is not None and target.exists():
        source = str(target)
        source_details.append(f"R019-03:{target}")
        wb = _load_workbook_quietly(str(target), read_only=True, data_only=True)
        try:
            ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
            ws = wb[ws_name] if ws_name else wb.active
            for row in ws.iter_rows(min_row=5, values_only=True):
                if not row:
                    continue
                val_a = row[0] if len(row) > 0 else None
                val_b = row[1] if len(row) > 1 else None

                parsed_num = _parse_iso_registry_int(val_a)
                if parsed_num is not None:
                    last_num = max(last_num, parsed_num)

                bp_parts = _extract_bp_year_seq(val_b)
                if bp_parts and bp_parts[0] == year_two:
                    last_seq = max(last_seq, bp_parts[1])
        finally:
            wb.close()

    try:
        if ISO_R01901_DOCS_DIR.exists():
            source_details.append(f"R019-01:{ISO_R01901_DOCS_DIR}")
            for doc in ISO_R01901_DOCS_DIR.iterdir():
                if not doc.is_file():
                    continue
                bp_parts = _extract_bp_year_seq(doc.name)
                if bp_parts and bp_parts[0] == year_two:
                    last_seq = max(last_seq, bp_parts[1])
    except Exception:
        pass

    try:
        payloads = _load_iso_payloads()
        if payloads:
            source_details.append(f"payloads:{ISO_PAYLOADS_FILE}")
        for payload in payloads.values():
            if not isinstance(payload, dict):
                continue
            parsed_num = _parse_iso_registry_int(
                payload.get("Numero_de_Registro") or payload.get("numero")
            )
            if parsed_num is not None:
                last_num = max(last_num, parsed_num)

            bp_parts = _extract_bp_year_seq(payload.get("BP") or payload.get("bp"))
            if bp_parts and bp_parts[0] == year_two:
                last_seq = max(last_seq, bp_parts[1])
    except Exception:
        pass

    return {
        "next_num": last_num + 1,
        "next_bp": f"BP-{year_two}{last_seq + 1:03d}",
        "source": source,
        "source_details": source_details,
    }

PRODUCTION_PATH = BASE_DIR / "P2 - Purchase Order/En Progreso"

PROCESSED_PATH = BASE_DIR / "P2 - Purchase Order/Procesado"

UPLOAD_FOLDER = BASE_DIR / "P1 - Registros Solicitados"

IN_PROCESS_DIR = UPLOAD_FOLDER / "in process"

USERS_FILE = BASE_DIR / "Codigos/Usuarios/users.json"
NOTIFICATIONS_FILE = BASE_DIR / "Datos/notifications.json"
ISO_PAYLOADS_FILE = BASE_DIR / "Datos/iso_payloads.json"

for _required_dir in (PRODUCTION_PATH, PROCESSED_PATH, UPLOAD_FOLDER, IN_PROCESS_DIR):
    try:
        _required_dir.mkdir(parents=True, exist_ok=True)
    except Exception as _dir_exc:
        print(f"Warning: Could not ensure directory {_required_dir}: {_dir_exc}")


SOLIDS_DIR = str(BASE_DIR / "Solidos")
PROJECTS_DB_PATH = str(BASE_DIR / "Registro de Proyectos" / "projects.json")
PLM_SHORTCUTS_DIR = str(BASE_DIR / "Registro de Proyectos" / "plm_shortcuts")


def _path_to_file_uri(path_value):
    raw = os.path.abspath(str(path_value))
    if raw.startswith('\\'):
        unc = raw.lstrip('\\')
        parts = unc.split('\\')
        if len(parts) < 2:
            return ''
        host = parts[0]
        encoded_parts = [quote(seg) for seg in parts[1:] if seg]
        if not encoded_parts:
            return ''
        return f"file://{host}/" + '/'.join(encoded_parts)
    return Path(raw).as_uri()


def _extract_plm_shortcut_filename(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return ''

    parsed = urlparse(value)
    path_value = str(parsed.path or value)
    marker = '/plm-shortcuts/'
    idx = path_value.lower().find(marker)
    if idx == -1:
        return ''

    candidate = path_value[idx + len(marker):].strip('/\\')
    return secure_filename(unquote(os.path.basename(candidate)))


def _plm_shortcut_to_local_path(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return ''

    filename = _extract_plm_shortcut_filename(value)
    if filename:
        return os.path.join(PLM_SHORTCUTS_DIR, filename)

    if value.lower().startswith('file://'):
        parsed = urlparse(value)
        netloc = unquote(str(parsed.netloc or '')).strip()
        path_value = unquote(str(parsed.path or ''))

        if netloc:
            return ('\\\\' + netloc + path_value.replace('/', '\\')).rstrip('\\')

        if path_value.startswith('/') and len(path_value) >= 3 and path_value[2] == ':':
            path_value = path_value[1:]

        return path_value.replace('/', '\\').strip()

    return ''


def _pick_file_with_windows_dialog(kind):
    kind_value = str(kind or '').strip().lower()

    filters = {
        'cad': "Archivos CAD|*.sldprt;*.sldasm;*.slddrw;*.step;*.stp;*.iges;*.igs;*.x_t;*.x_b;*.dxf;*.dwg|Todos los archivos|*.*",
        'drawing': "Archivos PDF|*.pdf|Todos los archivos|*.*"
    }

    titles = {
        'cad': 'Seleccionar archivo CAD',
        'drawing': 'Seleccionar archivo Plano'
    }

    file_filter = filters.get(kind_value, filters['cad']).replace("'", "''")
    title = titles.get(kind_value, 'Seleccionar archivo').replace("'", "''")

    ps_script = (
        "Add-Type -AssemblyName System.Windows.Forms; "
        "$dialog = New-Object System.Windows.Forms.OpenFileDialog; "
        f"$dialog.Title = '{title}'; "
        f"$dialog.Filter = '{file_filter}'; "
        "$dialog.Multiselect = $false; "
        "$dialog.CheckFileExists = $true; "
        "$dialog.CheckPathExists = $true; "
        "$result = $dialog.ShowDialog(); "
        "if ($result -eq [System.Windows.Forms.DialogResult]::OK) { "
        "[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; "
        "Write-Output $dialog.FileName }"
    )

    proc = subprocess.run(
        ['powershell.exe', '-NoProfile', '-STA', '-Command', ps_script],
        capture_output=True,
        text=True,
        timeout=300
    )

    if proc.returncode != 0:
        raise RuntimeError((proc.stderr or '').strip() or 'No se pudo abrir el explorador de archivos.')

    selected = str(proc.stdout or '').strip()
    if not selected:
        return ''

    lines = [line.strip() for line in selected.splitlines() if line.strip()]
    return lines[-1] if lines else ''


def _open_path_in_file_manager(target_path: Path) -> None:
    target = str(target_path)
    if os.name == 'nt':
        subprocess.Popen(['cmd', '/c', 'start', '', target], shell=True)
    elif sys.platform == 'darwin':
        subprocess.Popen(['open', target], shell=False)
    else:
        subprocess.Popen(['xdg-open', target], shell=False)

def load_db():
    default_db = {"projects": {}}
    if not os.path.exists(PROJECTS_DB_PATH):
        return default_db
    try:
        data = _load_json_file(PROJECTS_DB_PATH, default_db, dict)
        if not isinstance(data, dict):
            return default_db
        if not isinstance(data.get('projects'), dict):
            data['projects'] = {}
        return data
    except Exception as e:
        print(f"Error loading DB: {e}")
        return default_db

def save_db(data):
    try:
        _write_json_file_atomic(PROJECTS_DB_PATH, data, indent=4, ensure_ascii=False)
        return True
    except Exception as e:
        print(f"Error saving DB: {e}")
        return False

LOG_FILE = BASE_DIR / "Codigos/Usuarios/audit_log.json"

AUXILIAR_DIR = BASE_DIR / "Auxiliares/indices_auxiliar"
COTIZACION_CONSULTAS_DIR = BASE_DIR / "Auxiliares/Consultas"
COTIZACION_COMPLEMENTARIOS_FILE = COTIZACION_CONSULTAS_DIR / "Facturas compras - Ultimo precio.xlsx"
COTIZACION_COMPLEMENTARIOS_CSV_FILE = COTIZACION_CONSULTAS_DIR / "Facturas compras - Ultimo precio.csv"
COTIZACION_COMPLEMENTARIOS_CACHE_LOCK = threading.Lock()
COTIZACION_COMPLEMENTARIOS_CACHE = {
    "source": "",
    "mtime": 0.0,
    "records": []
}

LOGISTICS_CALC_LOCK = threading.Lock()
LOGISTICS_CALC_JOBS = {}
LOGISTICS_PROGRESS_STAGES = [
    ('queued', 'En cola'),
    ('validation', 'Preparación'),
    ('maximize', 'Maximización'),
    ('tray', 'Bandejas'),
    ('pallet', 'Pallets'),
    ('solver', 'Empaquetado'),
    ('summary', 'Resumen'),
    ('done', 'Completado')
]
LOGISTICS_PROGRESS_STAGE_INDEX = {
    key: idx for idx, (key, _) in enumerate(LOGISTICS_PROGRESS_STAGES)
}
LOGISTICS_PROGRESS_STAGE_LABELS = {
    key: label for key, label in LOGISTICS_PROGRESS_STAGES
}

def _set_logistics_calc_state(job_id, **updates):
    if not job_id:
        return
    with LOGISTICS_CALC_LOCK:
        state = LOGISTICS_CALC_JOBS.setdefault(job_id, {
            'status': 'queued',
            'progress': 0,
            'message': 'En cola...',
            'result': None,
            'cancel_requested': False,
            'stage': 'queued',
            'stage_label': LOGISTICS_PROGRESS_STAGE_LABELS['queued'],
            'stage_index': 0,
            'stage_total': len(LOGISTICS_PROGRESS_STAGES),
            'detail': 'En cola...'
        })
        state.update(updates)

def _update_logistics_calc_progress(job_id, *, stage=None, progress=None, message=None, status=None, result=None, detail=None):
    updates = {}
    if stage:
        updates['stage'] = stage
        updates['stage_label'] = LOGISTICS_PROGRESS_STAGE_LABELS.get(stage, stage)
        updates['stage_index'] = LOGISTICS_PROGRESS_STAGE_INDEX.get(stage, 0)
        updates['stage_total'] = len(LOGISTICS_PROGRESS_STAGES)
    if progress is not None:
        updates['progress'] = progress
    if message is not None:
        updates['message'] = message
    if detail is not None:
        updates['detail'] = detail
    elif message is not None:
        updates['detail'] = message
    if status is not None:
        updates['status'] = status
    if result is not None:
        updates['result'] = result
    _set_logistics_calc_state(job_id, **updates)

def _get_logistics_calc_state(job_id):
    with LOGISTICS_CALC_LOCK:
        return deepcopy(LOGISTICS_CALC_JOBS.get(job_id, {}))


def _is_logistics_calc_cancel_requested(job_id):
    if not job_id:
        return False
    with LOGISTICS_CALC_LOCK:
        return bool((LOGISTICS_CALC_JOBS.get(job_id) or {}).get('cancel_requested'))


def _raise_if_logistics_calc_cancelled(job_id):
    if _is_logistics_calc_cancel_requested(job_id):
        raise LogisticsCalculationCancelled('Cálculo cancelado por el usuario.')

def _extract_flask_json_response(resp):
    status_code = 200
    payload = None
    if isinstance(resp, tuple):
        response_obj = resp[0]
        if len(resp) > 1:
            status_code = resp[1]
    else:
        response_obj = resp
    try:
        payload = response_obj.get_json()
    except Exception:
        payload = None
    return payload, status_code


def _logistics_now_iso():
    return _utc_now_iso_z()


def _logistics_safe_timestamp(raw_value):
    value = str(raw_value or '').strip()
    if value:
        return value
    return _logistics_now_iso()


def _normalize_logistics_folder_name(raw_folder):
    folder = str(raw_folder or '').strip()
    return folder or LOGISTICS_DEFAULT_FOLDER


def _clean_logistics_items(raw_items):
    clean = []
    if not isinstance(raw_items, list):
        return clean
    for it in raw_items:
        if not isinstance(it, dict):
            continue
        clean.append({
            'name': it.get('name', ''),
            'l': it.get('l', 0),
            'w': it.get('w', 0),
            'h': it.get('h', 0),
            'weight': it.get('weight', 0),
            'qty': it.get('qty', 0),
            'rel_qty': it.get('rel_qty', 0)
        })
    return clean


def _format_logistics_timestamp_display(raw_timestamp, fallback=''):
    text = str(raw_timestamp or '').strip()
    if not text:
        return str(fallback or '').strip()
    try:
        parsed = datetime.fromisoformat(text.replace('Z', '+00:00'))
        return parsed.strftime('%d/%m/%Y, %H:%M:%S')
    except Exception:
        return text


def _get_logistics_largest_unit_label(source):
    data = source if isinstance(source, dict) else {}
    container = data.get('container', {}) if isinstance(data.get('container', {}), dict) else {}
    pallet = data.get('pallet', {}) if isinstance(data.get('pallet', {}), dict) else {}
    load = data.get('load', {}) if isinstance(data.get('load', {}), dict) else {}

    container_type = str(container.get('type') or '').strip().lower()
    pallet_type = str(pallet.get('type') or '').strip().lower()
    load_type = str(load.get('type') or '').strip().lower()

    if container_type and container_type != 'none':
        return 'Contenedor'
    if pallet_type and pallet_type != 'none':
        return 'Pallet'
    if load_type and load_type != 'loose':
        return 'Bandeja'
    return 'Piezas Sueltas'


def _build_logistics_history_label(version):
    data = version if isinstance(version, dict) else {}
    version_number = int(data.get('version_number') or 1)
    timestamp_display = str(data.get('timestamp_display') or '').strip()
    if not timestamp_display:
        timestamp_display = _format_logistics_timestamp_display(data.get('timestamp'))
    author = str(data.get('author') or 'Usuario').strip() or 'Usuario'
    is_kanban = bool(((data.get('optimization') or {}).get('tray_kanban')))
    largest_label = str(data.get('largest_unit_label') or _get_logistics_largest_unit_label(data)).strip() or 'Piezas Sueltas'
    kanban_label = ' - KANBAN' if is_kanban else ''
    return f"v{version_number} - {timestamp_display} - {author}{kanban_label} - {largest_label}"


def _build_logistics_version_snapshot(source, version_number=1):
    data = source if isinstance(source, dict) else {}
    timestamp = _logistics_safe_timestamp(data.get('timestamp'))
    timestamp_display = str(data.get('timestamp_display') or '').strip() or _format_logistics_timestamp_display(timestamp)
    snapshot = {
        'version_id': str(data.get('version_id') or uuid.uuid4()),
        'version_number': int(data.get('version_number') or version_number or 1),
        'timestamp': timestamp,
        'timestamp_display': timestamp_display,
        'save_name': str(data.get('save_name') or '').strip(),
        'save_description': str(data.get('save_description') or '').strip(),
        'author': str(data.get('author') or 'Usuario').strip() or 'Usuario',
        'container': {
            'type': data.get('container', {}).get('type'),
            'l': data.get('container', {}).get('l', 0),
            'w': data.get('container', {}).get('w', 0),
            'h': data.get('container', {}).get('h', 0),
            'weight': data.get('container', {}).get('weight', 0)
        },
        'load': {
            'type': data.get('load', {}).get('type', 'loose'),
            'tray_inner_l': data.get('load', {}).get('tray_inner_l', 0),
            'tray_inner_w': data.get('load', {}).get('tray_inner_w', 0),
            'tray_inner_h': data.get('load', {}).get('tray_inner_h', 0),
            'tray_outer_l': data.get('load', {}).get('tray_outer_l', 0),
            'tray_outer_w': data.get('load', {}).get('tray_outer_w', 0),
            'tray_outer_h': data.get('load', {}).get('tray_outer_h', 0),
            'tray_weight': data.get('load', {}).get('tray_weight', 0),
            'tray_max_weight': data.get('load', {}).get('tray_max_weight', 25)
        },
        'pallet': {
            'type': data.get('pallet', {}).get('type'),
            'boards': data.get('pallet', {}).get('boards', 0),
            'l': data.get('pallet', {}).get('l', 0),
            'w': data.get('pallet', {}).get('w', 0),
            'h': data.get('pallet', {}).get('h', 0),
            'weight': data.get('pallet', {}).get('weight', 0),
            'max_weight': data.get('pallet', {}).get('max_weight', 0),
            'limit_height': data.get('pallet', {}).get('limit_height', False),
            'max_height': data.get('pallet', {}).get('max_height', 0)
        },
        'optimization': {
            'maximize': data.get('optimization', {}).get('maximize', False),
            'mixed_trays': data.get('optimization', {}).get('mixed_trays', True),
            'tray_kanban': data.get('optimization', {}).get('tray_kanban', False),
            'mixed_pallets': data.get('optimization', {}).get('mixed_pallets', True),
            'stack_load': data.get('optimization', {}).get('stack_load', True),
            'stack_trays': data.get('optimization', {}).get('stack_trays', True),
            'stack_collars': data.get('optimization', {}).get('stack_collars', True),
            'force_orientation': data.get('optimization', {}).get('force_orientation', False),
            'orientation_face': data.get('optimization', {}).get('orientation_face', 'LxA'),
            'mixed_containers': data.get('optimization', {}).get('mixed_containers', True),
            'sort_items': data.get('optimization', {}).get('sort_items', True),
            'visual_only': data.get('optimization', {}).get('visual_only', False)
        },
        'safety_factors': {
            'dims': data.get('safety_factors', {}).get('dims', data.get('safety_factor_dims', 0)),
            'weight': data.get('safety_factors', {}).get('weight', data.get('safety_factor_weight', 0))
        },
        'freight': {
            'origin': data.get('freight', {}).get('origin'),
            'destination': data.get('freight', {}).get('destination'),
            'adjustment_usd': data.get('freight', {}).get('adjustment_usd', 0),
            'total_cost_usd': data.get('freight', {}).get('total_cost_usd', 0),
            'last_estimate': data.get('freight', {}).get('last_estimate')
        },
        'items': _clean_logistics_items(data.get('items', []))
    }
    snapshot['largest_unit_label'] = _get_logistics_largest_unit_label(snapshot)
    snapshot['history_label'] = _build_logistics_history_label(snapshot)
    return snapshot


def _normalize_logistics_group(raw_group):
    if not isinstance(raw_group, dict):
        return None

    group_id = str(raw_group.get('id') or uuid.uuid4())
    raw_versions = raw_group.get('versions')

    if isinstance(raw_versions, list):
        versions = []
        fallback_counter = 1
        for raw_version in raw_versions:
            if not isinstance(raw_version, dict):
                continue
            versions.append(_build_logistics_version_snapshot(
                raw_version,
                version_number=raw_version.get('version_number') or fallback_counter
            ))
            fallback_counter += 1
        if not versions:
            versions = [_build_logistics_version_snapshot(raw_group, version_number=1)]
    else:
        versions = [_build_logistics_version_snapshot(raw_group, version_number=1)]

    versions.sort(key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', '')))
    latest = versions[-1]

    return {
        'id': group_id,
        'folder': _normalize_logistics_folder_name(raw_group.get('folder')),
        'created_at': _logistics_safe_timestamp(raw_group.get('created_at') or versions[0].get('timestamp')),
        'updated_at': _logistics_safe_timestamp(raw_group.get('updated_at') or latest.get('timestamp')),
        'latest_version': int(latest.get('version_number') or len(versions) or 1),
        'save_name': latest.get('save_name', ''),
        'save_description': latest.get('save_description', ''),
        'author': latest.get('author', 'Usuario'),
        'versions': versions
    }


def _merge_logistics_groups(groups):
    merged = {}
    for group in groups if isinstance(groups, list) else []:
        normalized = _normalize_logistics_group(group)
        if not normalized:
            continue
        gid = str(normalized.get('id') or '')
        if not gid:
            continue
        existing = merged.get(gid)
        if not existing:
            merged[gid] = normalized
            continue

        combined_versions = []
        seen = set()
        for source in [existing, normalized]:
            for version in source.get('versions', []):
                version_id = str(version.get('version_id') or '')
                key = version_id or f"{version.get('version_number')}|{version.get('timestamp')}"
                if key in seen:
                    continue
                seen.add(key)
                combined_versions.append(version)

        existing['versions'] = combined_versions
        existing['folder'] = _normalize_logistics_folder_name(normalized.get('folder') or existing.get('folder'))
        merged[gid] = _normalize_logistics_group(existing)

    merged_list = [group for group in merged.values() if isinstance(group, dict)]
    merged_list.sort(key=lambda group: group.get('updated_at', ''), reverse=True)
    return merged_list


def _read_logistics_records_store():
    empty_store = {'folders': [LOGISTICS_DEFAULT_FOLDER], 'groups': []}
    if not os.path.exists(LOGISTICS_RECORDS_FILE):
        return empty_store

    try:
        with open(LOGISTICS_RECORDS_FILE, 'r', encoding='utf-8') as f:
            raw_records = json.load(f)
    except Exception:
        return empty_store

    if isinstance(raw_records, list):
        raw_groups = raw_records
        raw_folders = []
    elif isinstance(raw_records, dict):
        raw_groups = raw_records.get('groups', [])
        raw_folders = raw_records.get('folders', [])
    else:
        return empty_store

    groups = _merge_logistics_groups(raw_groups if isinstance(raw_groups, list) else [])
    folder_set = {LOGISTICS_DEFAULT_FOLDER}
    if isinstance(raw_folders, list):
        for folder in raw_folders:
            folder_set.add(_normalize_logistics_folder_name(folder))
    for group in groups:
        folder_set.add(_normalize_logistics_folder_name(group.get('folder')))

    folders = sorted(folder_set, key=lambda name: (name != LOGISTICS_DEFAULT_FOLDER, name.lower()))
    return {'folders': folders, 'groups': groups}


def _read_logistics_record_groups():
    return _read_logistics_records_store().get('groups', [])


def _write_logistics_record_groups(groups):
    safe_groups = _merge_logistics_groups(groups if isinstance(groups, list) else [])
    folder_set = {LOGISTICS_DEFAULT_FOLDER}
    existing_store = _read_logistics_records_store()
    for folder in existing_store.get('folders', []):
        folder_set.add(_normalize_logistics_folder_name(folder))
    for group in safe_groups:
        folder_set.add(_normalize_logistics_folder_name(group.get('folder')))
    payload = {
        'folders': sorted(folder_set, key=lambda name: (name != LOGISTICS_DEFAULT_FOLDER, name.lower())),
        'groups': safe_groups
    }
    _write_json_file_atomic(LOGISTICS_RECORDS_FILE, payload, indent=4, ensure_ascii=False)


def _logistics_get_latest_version(group):
    versions = group.get('versions', []) if isinstance(group, dict) else []
    if not versions:
        return None
    return sorted(versions, key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', '')))[-1]


def _logistics_group_to_summary(group):
    latest = _logistics_get_latest_version(group) or {}
    versions = group.get('versions', []) if isinstance(group, dict) else []
    created_at = group.get('created_at') or latest.get('timestamp') or _logistics_now_iso()
    modified_at = group.get('updated_at') or latest.get('timestamp') or created_at
    return {
        'id': group.get('id', ''),
        'folder': _normalize_logistics_folder_name(group.get('folder')),
        'save_name': latest.get('save_name', '') or group.get('save_name', ''),
        'save_description': latest.get('save_description', '') or group.get('save_description', ''),
        'author': latest.get('author', 'Usuario') or group.get('author', 'Usuario'),
        'created_at': created_at,
        'modified_at': modified_at,
        'latest_version': int(group.get('latest_version') or latest.get('version_number') or len(versions) or 1),
        'version_count': len(versions),
        'latest_history_label': latest.get('history_label', '')
    }


def _logistics_version_to_editor_record(group, version):
    data = deepcopy(version if isinstance(version, dict) else {})
    data['id'] = group.get('id', '')
    data['folder'] = _normalize_logistics_folder_name(group.get('folder'))
    data['created_at'] = group.get('created_at', '')
    data['modified_at'] = group.get('updated_at', '')
    data['latest_version'] = int(group.get('latest_version') or version.get('version_number') or 1)
    data['version_count'] = len(group.get('versions', [])) if isinstance(group.get('versions', []), list) else 1
    return data

HISTORIAL_PO_DIR = BASE_DIR / "Auxiliares/Historial PO"

PROFILE_PICS_DIR = BASE_DIR / "Codigos/Usuarios/profile_pics"

MOCK_PATH = SCRIPT_DIR_PATH / "mock_data"

RESET_TOKENS_FILE = BASE_DIR / "Codigos/Usuarios/reset_tokens.json"

SMTP_CONFIG = {
    "SMTP_SERVER": "smtp.gmail.com",
    "SMTP_PORT": 465,
    "SMTP_USER": "no-reply@bpbargentina.com",
    "SMTP_PASS": "mtky inyj bntn oxii",
    "SMTP_FROM_EMAIL": "no-reply@bpbargentina.com",
    "SMTP_FROM_NAME": "Oficina Tecnica",
    "SMTP_SSL": True,
    "SMTP_STARTTLS": False,
}

def send_reset_email(to_email, token):
    try:
        msg = MIMEMultipart()
        msg['From'] = f"{SMTP_CONFIG['SMTP_FROM_NAME']} <{SMTP_CONFIG['SMTP_FROM_EMAIL']}>"
        msg['To'] = to_email
        msg['Subject'] = "Recuperar ContraseÃ±a - BPB Group"

        # Dynamically build link (assuming standard port/host, can be improved with url_for external)
        # Use request.host_url from context if available, otherwise guess.
        # Since this is called from a route, request context is active.
        reset_link = f"{request.host_url}validate-reset?token={token}"

        body = f"""
        <html>
        <body style="font-family: Arial, sans-serif;">
            <h2>Restablecimiento de ContraseÃ±a</h2>
            <p>Se ha solicitado un cambio de contraseÃ±a para su cuenta ({to_email}).</p>
            <p>Para confirmar el cambio y establecer su nueva contraseÃ±a, haga clic en el siguiente enlace:</p>
            <p>
                <a href="{reset_link}" style="background-color: #e74c3c; color: white; padding: 10px 20px; text-decoration: none; border-radius: 5px; font-weight: bold;">
                    Confirmar Cambio de ContraseÃ±a
                </a>
            </p>
            <p style="font-size: 0.9em; color: #777;">Si usted no solicitÃ³ este cambio, ignore este correo.</p>
            <p>Este enlace es vÃ¡lido por 1 hora.</p>
        </body>
        </html>
        """
        msg.attach(MIMEText(body, 'html'))

        if SMTP_CONFIG.get('SMTP_SSL'):
            server = smtplib.SMTP_SSL(SMTP_CONFIG['SMTP_SERVER'], SMTP_CONFIG['SMTP_PORT'])
        else:
            server = smtplib.SMTP(SMTP_CONFIG['SMTP_SERVER'], SMTP_CONFIG['SMTP_PORT'])
            if SMTP_CONFIG.get('SMTP_STARTTLS'):
                server.starttls()
        server.login(SMTP_CONFIG['SMTP_USER'], SMTP_CONFIG['SMTP_PASS'])
        server.sendmail(SMTP_CONFIG['SMTP_FROM_EMAIL'], to_email, msg.as_string())
        server.quit()
        print(f"DEBUG: Reset email sent to {to_email}")
        return True, "Correo enviado"
    except Exception as e:
        print(f"ERROR: Failed to send email: {e}")
        # Log the link for debugging if SMTP fails
        print(f"DEBUG LINK (Fallback): {request.host_url}validate-reset?token={token}")
        return False, str(e)

@app.route('/api/request-reset', methods=['POST'])
def request_reset():
    try:
        # RAW DEBUG
        # print(f"DEBUG RAW PAYLOAD: {request.data}")
        # print(f"DEBUG JSON: {request.json}")

        data = request.json
        if data is None:
             return jsonify({'status': 'error', 'message': 'No JSON data received'}), 400

        if isinstance(data, str):
            try:
                data = json.loads(data)
            except:
                return jsonify({'status': 'error', 'message': 'Invalid JSON format (string parse failed)'}), 400
            
        if not isinstance(data, dict):
             return jsonify({'status': 'error', 'message': f'Invalid data format: Expected dict, got {type(data)}'}), 400

        email_addr = data.get('email', '').strip()
        new_password = data.get('new_password', '').strip()

        if not email_addr or not new_password:
            return jsonify({'status': 'error', 'message': 'Faltan datos.'}), 400

        if not os.path.exists(USERS_FILE):
            return jsonify({'status': 'error', 'message': 'Error interno de usuarios.'}), 500

        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            users = json.load(f)
        
        # Check if user exists
        # Users file is a dict: {username: user_data}
        user_exists = False
        for username, u_data in users.items():
            # Check if key (username) matches email OR if 'email' field inside matches
            if username.lower() == email_addr.lower():
                user_exists = True
                break
            if isinstance(u_data, dict) and u_data.get('email', '').lower() == email_addr.lower():
                user_exists = True
                break
        
        if not user_exists:
             return jsonify({'status': 'error', 'message': 'Email no registrado.'}), 404

        # Generate Token
        token = secrets.token_urlsafe(32)
        
        # Load existing tokens
        tokens = {}
        if os.path.exists(RESET_TOKENS_FILE):
            try:
                with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
                    tokens = json.load(f)
            except: tokens = {}

        # Clean expired tokens (older than 1h)
        now = time.time()
        clean_tokens = {}
        for k, v in tokens.items():
            if isinstance(v, dict) and (now - v.get('timestamp', 0) < 3600):
                clean_tokens[k] = v
        tokens = clean_tokens

        # Add new token
        tokens[token] = {
            'email': email_addr,
            'new_password_hash': _generate_password_hash_compatible(new_password),
            'timestamp': now
        }

        # Save tokens
        try:
            _write_json_file_atomic(RESET_TOKENS_FILE, tokens, indent=4, ensure_ascii=False)
        except Exception as file_err:
             import traceback
             print(traceback.format_exc())
             return jsonify({'status': 'error', 'message': f'Error guardando token: {file_err}'}), 500

        # Send Email
        success, msg = send_reset_email(email_addr, token)
        
        if success:
            return jsonify({'status': 'success', 'message': 'Revise su correo para confirmar.'})
        else:
            return jsonify({'status': 'error', 'message': f'Error enviando correo: {msg}'})

    except Exception as e:
        import traceback
        tb = traceback.format_exc()
        print(f"CRITICAL ERROR in request_reset: {tb}")
        return jsonify({'status': 'error', 'message': f'Error CRITICO: {str(e)} -> {tb}'}), 500

@app.route('/validate-reset', methods=['GET'])
def validate_reset():
    token = request.args.get('token')
    
    if not token or not os.path.exists(RESET_TOKENS_FILE):
        return "Enlace invÃ¡lido o expirado.", 400

    try:
        with open(RESET_TOKENS_FILE, 'r', encoding='utf-8') as f:
            tokens = json.load(f)

        token_data = tokens.get(token)
        if not token_data:
             return "Enlace invÃ¡lido o expirado.", 400
        
        # Check expiration again just in case
        if time.time() - token_data.get('timestamp', 0) > 3600:
            return "Enlace expirado.", 400

        # Update User Password
        target_email = token_data['email']
        new_hash = token_data['new_password_hash']

        with open(USERS_FILE, 'r', encoding='utf-8') as f:
            users = json.load(f)

        updated = False
        target_email = target_email.strip().lower()
        
        # Iterate safely over keys and values
        matched_username = None
        for username, u_data in users.items():
            # Check if key matches email
            if username.lower() == target_email:
                matched_username = username
                break
            # Check if inner email matches
            if isinstance(u_data, dict) and u_data.get('email', '').strip().lower() == target_email:
                matched_username = username
                break

        if matched_username and matched_username in users:
             users[matched_username]['password'] = new_hash
             updated = True
        
        if updated:
            _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)
            
            # Delete used token
            del tokens[token]
            _write_json_file_atomic(RESET_TOKENS_FILE, tokens, indent=4, ensure_ascii=False)
                
            return """
            <html>
            <head>
                <meta http-equiv="refresh" content="3;url=/">
                <style>body{font-family:sans-serif;text-align:center;padding-top:50px;}</style>
            </head>
            <body>
                <h1 style="color:green;">Â¡ContraseÃ±a Actualizada!</h1>
                <p>SerÃ¡ redirigido al inicio de sesiÃ³n en 3 segundos...</p>
            </body>
            </html>
            """
        else:
             return "Usuario no encontrado.", 404

    except Exception as e:
        print(f"Error validating reset: {e}")
        return "Error interno.", 500

ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

MANUAL_INGRESOS_FILE = BASE_DIR / r"Auxiliares/Ingresos/resumen_all.csv"



# Ensure profile pics dir exists

if not PROFILE_PICS_DIR.exists():

    try:

        PROFILE_PICS_DIR.mkdir(parents=True, exist_ok=True)

    except Exception as e:

        print(f"Warning: Could not create profile pics folder: {e}")



# Determine which path to use

if os.environ.get('FLASK_TEST_MODE'):

    DATA_DIR = MOCK_PATH

    print("Running in TEST MODE: Using mock data")

else:

    if PRODUCTION_PATH.exists():

        DATA_DIR = PRODUCTION_PATH

        print(f"Production path found: {DATA_DIR}")

    else:

        DATA_DIR = MOCK_PATH

        print(f"Production path NOT found. Falling back to Mock Data: {DATA_DIR}")

    

    if not UPLOAD_FOLDER.exists():

        try:

            UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

        except Exception as e:

            print(f"Warning: Could not create upload folder {UPLOAD_FOLDER}: {e}")



print(f"Using Data Directory: {DATA_DIR.resolve()}")



def log_action(username, action, details):

    print(f"[LOG ACTION] User: {username}, Action: {action}")

    try:

        logs = []

        if LOG_FILE.exists():

            with open(LOG_FILE, 'r') as f:

                logs = json.load(f)

        

        entry = {

            "timestamp": time.strftime('%Y-%m-%d %H:%M:%S'),

            "user": username,

            "action": action,

            "details": details

        }

        logs.append(entry)

        

        if len(logs) > 1000:

            logs = logs[-1000:]

            

        _write_json_file_atomic(LOG_FILE, logs, indent=4, ensure_ascii=False)

        print("[LOG ACTION] Saved successfully.")

            

    except Exception as e:

        print(f"Logging Error: {e}")



def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS



def get_entry_dates_cache():

    """

    Scans IN_PROCESS_DIR for folders matching 'Registro - Rxxxx - ...'

    Sorts them by R number (ascending) so latest overwrites oldest.

    Reads resumen.csv and builds {po_number: date} map.

    """

    cache = {}

    if not IN_PROCESS_DIR.exists():

        return cache

        

    # Find all relevant folders

    r_folders = []

    for item in IN_PROCESS_DIR.iterdir():

        if item.is_dir():
            # Match R number. Pattern assumed: "Registro - R0060 - ..."
            match = re.search(r'Registro\s*-\s*R(\d+)', item.name, re.IGNORECASE)
            if match:
                r_num = int(match.group(1))
                r_folders.append((r_num, item))

    # Sort by R number Ascending (Oldest to Newest)
    r_folders.sort(key=lambda x: x[0])

    for _, folder in r_folders:
        resumen_file = folder / "resumen.csv"
        if resumen_file.exists():
            try:
                # Try reading with different encodings
                content = None
                for enc in ['utf-8', 'latin-1', 'cp1252']:
                    try:
                        with open(resumen_file, 'r', encoding=enc) as f:
                            content = f.read()
                        break
                    except UnicodeDecodeError:
                        continue
                
                if content:
                    delimiter = ';' if ';' in content.split('\n')[0] else ','
                    lines = content.split('\n')
                    lines = [l.strip() for l in lines if l.strip()]
                    
                    for line in lines:
                        parts = line.split(delimiter)
                        if len(parts) >= 4:
                            date_val = parts[3].strip()
                            po_val = parts[2].strip()
                            if not po_val: po_val = parts[1].strip()
                            
                            if po_val and date_val:
                                digits = "".join(filter(str.isdigit, po_val))
                                if digits:
                                    cache[digits] = date_val
            except Exception as e:
                print(f"Error reading resumen in {folder}: {e}")

    return cache

# --- SYNC GLOBAL STATE ---
SYNC_STATE = {
    "running": False,
    "stage": "idle",
    "progress": 0,
    "total": 0,
    "current": 0,
    "message": "",
    "source": "auto", 
    "data_version": 0
}

def count_files_recursive(directory, extension):
    count = 0
    try:
        for path in Path(directory).rglob(f'*{extension}'):
            if path.is_file():
                count += 1
    except: pass
    return count

def count_subdirs(directory):
    count = 0
    try:
        # Count direct subdirectories
        for path in Path(directory).iterdir():
            if path.is_dir():
                count += 1
    except: pass
    return count



def background_sync_task():

    global SYNC_STATE

    SYNC_STATE["running"] = True

    SYNC_STATE["progress"] = 0

    SYNC_STATE["data_version"] = 0 # Reset version

    target_registro = SYNC_STATE.get("target_registro")

    # script_dir = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Codigos")
    script_dir = BASE_DIR / "Codigos"

    

    try:

        import sys

        # --- ETAPA 1: STEP 2 (Descarga de PDFs) ---

        SYNC_STATE["stage"] = "step2"

        SYNC_STATE["message"] = "Descargando archivos PDF..."

        

        # LOGGING FOR DEBUGGING

        print(f"Background Sync: Using interpreter {sys.executable}")

        print(f"Background Sync: CWD is {script_dir}")

        

        # Keep progress at 0 during fetch since we can't estimate easily yet

        SYNC_STATE["progress"] = 0 

        

        # Use sys.executable to ensure same environment

        step2_cmd = [sys.executable, 'step2_fetch_pdfs.py', '--config', 'config.yaml']

        if target_registro:
            step2_cmd += ['--registro', str(target_registro)]

        

        # Run Step 2

        result2 = subprocess.run(step2_cmd, cwd=str(script_dir), capture_output=True, text=True)

        if result2.returncode != 0:

            print(f"Step 2 Failed: {result2.stderr}")
            if result2.stdout:
                print(f"Step 2 stdout: {result2.stdout}")

            err_msg = (result2.stderr or result2.stdout or '').strip()
            if err_msg:
                raise Exception(f"Fallo descarga PDF: {err_msg}")
            raise Exception(f"Fallo descarga PDF (Ver consola para detalles)")

        

        # --- ETAPA 2: STEP 3 (Procesar Outputs en EN PROGRESO) ---

        SYNC_STATE["stage"] = "step3"

        SYNC_STATE["message"] = "Procesando registros..."

        

        # entrantes_dir = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\P2 - Purchase Order\Entrantes")
        entrantes_dir = BASE_DIR / "P2 - Purchase Order/Entrantes"
        
        # en_progreso_dir = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\P2 - Purchase Order\En Progreso")
        en_progreso_dir = BASE_DIR / "P2 - Purchase Order/En Progreso"

        

        total_pdfs = count_files_recursive(entrantes_dir, ".pdf")

        SYNC_STATE["total"] = total_pdfs if total_pdfs > 0 else 1 # Avoid div/0

        

        step3_cmd = [sys.executable, '-X', 'faulthandler', 'step3_prepare_outputs.py', '--config', 'config.yaml']

        def _run_step3_once():
            process3 = subprocess.Popen(
                step3_cmd,
                cwd=str(script_dir),
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                bufsize=1
            )

            step3_lines = []
            if process3.stdout:
                for raw_line in process3.stdout:
                    line = (raw_line or '').strip()
                    if not line:
                        continue
                    if line.startswith('PROGRESS '):
                        try:
                            token = line.split()[1]
                            done_s, total_s = token.split('/')
                            done_i = int(done_s)
                            total_i = int(total_s) if int(total_s) > 0 else 1
                            SYNC_STATE["total"] = total_i
                            SYNC_STATE["progress"] = min(100, int((done_i / total_i) * 100))
                            SYNC_STATE["message"] = f"Generando JSON {done_i}/{total_i}"
                        except Exception:
                            step3_lines.append(line)
                    else:
                        step3_lines.append(line)

            process3.wait()
            return process3.returncode, "\n".join(step3_lines).strip()

        step3_attempts = 2
        step3_returncode = None
        step3_stdout = ""

        for attempt in range(1, step3_attempts + 1):
            if attempt > 1:
                print(f"Step 3 retry attempt {attempt}/{step3_attempts}...")
                SYNC_STATE["message"] = f"Reintentando procesamiento ({attempt}/{step3_attempts})..."
                SYNC_STATE["progress"] = 0
                time.sleep(1.0)

            step3_returncode, step3_stdout = _run_step3_once()
            if step3_returncode == 0:
                break

        if step3_returncode != 0:

            print(f"Step 3 Failed with return code {step3_returncode}")
            if step3_stdout:
                print(f"Step 3 output: {step3_stdout}")

            err_msg = (step3_stdout or '').strip()
            if err_msg:
                raise Exception(f"Fallo script procesamiento (rc={step3_returncode}): {err_msg}")
            raise Exception(f"Fallo script procesamiento (rc={step3_returncode}) (Ver consola para detalles)")

        SYNC_STATE["progress"] = 100

        SYNC_STATE["message"] = "SincronizaciÃ³n completada"

        

    except Exception as e:

        print(f"Sync Error: {e}")

        SYNC_STATE["message"] = f"Error: {str(e)}"

        SYNC_STATE["stage"] = "error"

    finally:

        SYNC_STATE["target_registro"] = None
        SYNC_STATE["running"] = False

        print("Background sync task finished.")



def start_sync_thread(source="auto", target_registro=None):

    if SYNC_STATE["running"]:

        return False

    SYNC_STATE["source"] = source
    SYNC_STATE["target_registro"] = target_registro

    thread = threading.Thread(target=background_sync_task)

    thread.daemon = True

    thread.start()

    return True


# --- ROUTES ---

@app.route('/solids/<path:filename>')
def serve_solids(filename):
    return send_from_directory(SOLIDS_DIR, filename)

@app.route('/plm-shortcuts/<path:filename>')
def serve_plm_shortcut(filename):
    safe_name = secure_filename(filename)
    if not safe_name:
        return jsonify({'status': 'error', 'message': 'Archivo invalido'}), 400
    return send_from_directory(PLM_SHORTCUTS_DIR, safe_name)

@app.route('/api/projects', methods=['GET', 'POST'])
def api_projects():
    db = load_db()
    projects_map = db.get('projects', {}) if isinstance(db, dict) else {}

    if request.method == 'GET':
        projects_list = []
        for p_id, payload in projects_map.items():
            if not isinstance(payload, dict):
                continue
            row = dict(payload)
            row['id'] = str(row.get('id') or p_id)
            projects_list.append(row)
        return jsonify(projects_list)

    if request.method == 'POST':
        data = request.json or {}
        p_id = str(data.get('id') or '').strip()
        if not p_id:
            return jsonify({"status": "error", "message": "Missing ID"}), 400

        row = dict(data)
        row['id'] = p_id
        if 'description' not in row:
            row['description'] = ''

        if 'projects' not in db or not isinstance(db.get('projects'), dict):
            db['projects'] = {}

        db['projects'][p_id] = row
        if not save_db(db):
            return jsonify({"status": "error", "message": "No se pudo guardar el proyecto"}), 500
        return jsonify({"status": "success", "message": "Project saved", "project": row})

@app.route('/api/add-solid', methods=['POST'])
def api_add_solid():
    try:
        project_id = request.form.get('projectId')
        solid_name = request.form.get('name')
        revision = request.form.get('revision')
        
        if 'file' not in request.files:
             return jsonify({'status': 'error', 'message': 'No file part'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'status': 'error', 'message': 'No selected file'}), 400

        # Save File
        filename = secure_filename(file.filename)
        save_path = os.path.join(SOLIDS_DIR, filename)
        file.save(save_path)
        
        # Save Metadata
        db = load_db()
        if project_id not in db['projects']:
             return jsonify({'status': 'error', 'message': 'Project not found'}), 404
             
        if 'solids' not in db['projects'][project_id]:
            db['projects'][project_id]['solids'] = []
            
        new_solid = {
            "id": f"s{int(time.time()*1000)}",
            "name": solid_name,
            "revision": revision,
            "filename": filename,
            "hasFile": True,
            "date": datetime.now().strftime("%Y-%m-%d")
        }
        
        db['projects'][project_id]['solids'].append(new_solid)
        save_db(db)
        
        return jsonify({'status': 'success', 'solid': new_solid})
        
    except Exception as e:
        print(f"Error adding solid: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/plm-pick-file', methods=['POST'])
def api_plm_pick_file():
    try:
        data = request.get_json(silent=True) or {}
        kind = str(data.get('kind') or 'cad').strip().lower()
        if kind not in ('cad', 'drawing'):
            kind = 'cad'

        selected_path = _pick_file_with_windows_dialog(kind)
        if not selected_path:
            return jsonify({'status': 'cancelled'})

        if not os.path.exists(selected_path):
            return jsonify({'status': 'error', 'message': f'Archivo no encontrado: {selected_path}'}), 404

        shortcut = _path_to_file_uri(selected_path)
        if not shortcut:
            return jsonify({'status': 'error', 'message': 'No se pudo generar el acceso directo del archivo.'}), 500

        return jsonify({
            'status': 'success',
            'shortcut': shortcut,
            'path': selected_path,
            'filename': os.path.basename(selected_path)
        })
    except Exception as exc:
        print(f"Error selecting PLM file: {exc}")
        return jsonify({'status': 'error', 'message': str(exc)}), 500

@app.route('/api/plm-shortcut', methods=['POST'])
def api_create_plm_shortcut():
    try:
        if 'file' not in request.files:
            return jsonify({'status': 'error', 'message': 'No file part'}), 400

        file = request.files['file']
        if not file or file.filename == '':
            return jsonify({'status': 'error', 'message': 'No selected file'}), 400

        kind = str(request.form.get('kind') or 'file').strip().lower()
        if kind not in ('drawing', 'cad'):
            kind = 'file'

        original_name = secure_filename(file.filename)
        if not original_name:
            return jsonify({'status': 'error', 'message': 'Nombre de archivo invalido'}), 400

        _, ext = os.path.splitext(original_name)
        unique_name = f"{kind}-{uuid.uuid4().hex}{ext.lower()}"

        os.makedirs(PLM_SHORTCUTS_DIR, exist_ok=True)
        file.save(os.path.join(PLM_SHORTCUTS_DIR, unique_name))

        web_shortcut = f"/plm-shortcuts/{unique_name}"
        file_shortcut = ''
        try:
            file_shortcut = _path_to_file_uri(os.path.join(PLM_SHORTCUTS_DIR, unique_name))
        except Exception:
            file_shortcut = ''

        shortcut = web_shortcut
        open_mode = 'web'
        if kind == 'cad':
            if not file_shortcut:
                return jsonify({'status': 'error', 'message': 'No se pudo generar acceso directo de archivo para CAD.'}), 500
            shortcut = file_shortcut
            open_mode = 'file'

        return jsonify({
            'status': 'success',
            'shortcut': shortcut,
            'web_shortcut': web_shortcut,
            'file_shortcut': file_shortcut,
            'open_mode': open_mode,
            'filename': original_name
        })
    except Exception as e:
        print(f"Error creating PLM shortcut: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/plm-open-cad', methods=['POST'])
def api_open_plm_cad():
    try:
        import sys
        data = request.get_json(silent=True) or {}
        shortcut = data.get('shortcut')
        cad_path = _plm_shortcut_to_local_path(shortcut)

        if not cad_path:
            return jsonify({'status': 'error', 'message': 'Acceso CAD invalido.'}), 400

        if not os.path.exists(cad_path):
            return jsonify({'status': 'error', 'message': f'Archivo CAD no encontrado: {cad_path}'}), 404

        if hasattr(os, 'startfile'):
            os.startfile(cad_path)
        elif sys.platform == 'darwin':
            subprocess.Popen(['open', cad_path], shell=False)
        else:
            subprocess.Popen(['xdg-open', cad_path], shell=False)

        return jsonify({'status': 'success'})
    except Exception as exc:
        print(f"Error opening CAD: {exc}")
        return jsonify({'status': 'error', 'message': str(exc)}), 500

@app.route('/api/solids/<project_id>', methods=['GET'])
def api_get_solids(project_id):
    db = load_db()
    if project_id in db['projects']:
        solids = db['projects'][project_id].get('solids', [])
        return jsonify(solids)
    return jsonify([])

@app.route('/projects-viewer')
def projects_viewer():
    # Redirect or serve a specific template if needed, for now use index
    return redirect(url_for('index'))

@app.route('/')
def index():
    """Serve the main page with no-cache headers to prevent stale HTML"""
    from datetime import datetime
    response = make_response(render_template('index.html', cache_buster=int(datetime.now().timestamp())))
    # Prevent browser from caching HTML to ensure users always get latest version
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '0'
    return response


@app.route('/api/chatbot/chat', methods=['POST'])
def chatbot_chat():
    if not session.get('user'):
        return jsonify({
            'status': 'error',
            'response': 'Primero necesitas iniciar sesion para usar el asistente.',
            'actions': []
        }), 401

    try:
        payload = request.get_json(silent=True) or {}
        result = build_chatbot_response(
            payload,
            base_dir=BASE_DIR,
            cotizacion_records_file=COTIZACION_RECORDS_FILE,
            projects_file=PROJECTS_DB_PATH,
            user_email=str(session.get('user') or ''),
            user_role=str(session.get('role') or ''),
        )
        return jsonify(result)
    except Exception as e:
        print(f"[CHATBOT] Error: {e}")
        return jsonify({
            'status': 'error',
            'response': f'No pude procesar la consulta del asistente: {e}',
            'actions': []
        }), 500


@app.route('/api/fetch-pdfs', methods=['POST'])

def run_fetch_pdfs():

    if not start_sync_thread(source="auto"):

        return jsonify({"status": "running", "message": "Ya hay una sincronizaciÃ³n en curso"})

    


    return jsonify({"status": "started", "message": "Iniciando proceso en segundo plano"})

def _run_logistics_calc_job(job_id, data):
    try:
        _update_logistics_calc_progress(job_id, status='running', stage='validation', progress=2, message='Iniciando cálculo...')
        payload = deepcopy(data or {})
        config = payload.setdefault('config', {})
        config['job_id'] = job_id
        config['progress_detail'] = True
        with app.test_request_context('/api/logistics/calculate', method='POST', json=payload):
            resp = calculate_logistics()
        result, status_code = _extract_flask_json_response(resp)
        if status_code == 200 and isinstance(result, dict) and result.get('status') == 'success':
            _update_logistics_calc_progress(job_id, status='done', stage='done', progress=100, message='Cálculo completado.', result=result)
        elif status_code == 499 or (isinstance(result, dict) and result.get('status') == 'cancelled'):
            _update_logistics_calc_progress(job_id, status='cancelled', stage='done', progress=100, message='Cálculo cancelado.', result=result)
        else:
            error_message = (result or {}).get('message') if isinstance(result, dict) else 'Error en cálculo.'
            _update_logistics_calc_progress(job_id, status='error', stage='done', progress=100, message=error_message, result=result)
    except LogisticsCalculationCancelled as exc:
        _update_logistics_calc_progress(job_id, status='cancelled', stage='done', progress=100, message=str(exc), result={'status': 'cancelled', 'message': str(exc)})
    except Exception as exc:
        _update_logistics_calc_progress(job_id, status='error', stage='done', progress=100, message=str(exc), result={'status': 'error', 'message': str(exc)})

@app.route('/api/logistics/calculate-start', methods=['POST'])
def start_logistics_calculation():
    data = request.json
    if not data:
        return jsonify({'status': 'error', 'message': 'No data received'}), 400
    job_id = uuid.uuid4().hex
    _update_logistics_calc_progress(job_id, status='queued', stage='queued', progress=0, message='En cola...', result=None)
    thread = threading.Thread(target=_run_logistics_calc_job, args=(job_id, data))
    thread.daemon = True
    thread.start()
    return jsonify({'status': 'started', 'job_id': job_id})

@app.route('/api/logistics/calculate-progress/<job_id>', methods=['GET'])
def get_logistics_calculation_progress(job_id):
    state = _get_logistics_calc_state(job_id)
    if not state:
        return jsonify({'status': 'error', 'message': 'Job no encontrado'}), 404
    return jsonify(state)


@app.route('/api/logistics/calculate-cancel/<job_id>', methods=['POST'])
def cancel_logistics_calculation(job_id):
    state = _get_logistics_calc_state(job_id)
    if not state:
        return jsonify({'status': 'error', 'message': 'Job no encontrado'}), 404
    if state.get('status') in ('done', 'error', 'cancelled'):
        return jsonify({'status': 'ok', 'message': 'El cálculo ya finalizó.'})
    _set_logistics_calc_state(job_id, cancel_requested=True, status='cancelling', message='Cancelando cálculo...', detail='Cancelando cálculo...')
    return jsonify({'status': 'ok', 'message': 'Solicitud de cancelación enviada.'})


def _validate_logistics_weight_limits(container_data, items_data, config):
    tray_load_types = ('tray', 'tray_euro', 'tray_american', 'tray_custom')
    load_type = config.get('load_type', 'loose')
    pallet_type = config.get('pallet_type', 'none')
    container_type = config.get('container_type', 'none')

    item_weights = []
    for item in items_data or []:
        try:
            weight = float(item.get('weight', 0) or 0)
        except Exception:
            weight = 0
        if weight > 0:
            item_weights.append({
                'name': item.get('id') or item.get('name') or 'Item',
                'weight': weight
            })

    if not item_weights:
        return None

    lightest_item = min(item_weights, key=lambda item: item['weight'])
    is_tray_type = load_type in tray_load_types
    has_pallet = pallet_type != 'none'
    has_container = container_type != 'none'

    tray_dims = config.get('tray_dims') or {}
    tray_load_limit = float(tray_dims.get('max_weight', 25) or 25)
    tray_tare = float(tray_dims.get('weight', 0) or 0)

    pallet_dims = config.get('pallet_dims') or {}
    pallet_base_weight = float(pallet_dims.get('weight', 0) or 0)
    try:
        pallet_gross_limit = float(config.get('max_pallet_weight', 0) or 0)
    except Exception:
        pallet_gross_limit = 0
    if has_pallet and pallet_gross_limit <= 0:
        pallet_gross_limit = 2000

    try:
        container_gross_limit = float(container_data.get('max_weight', 0) or 0)
    except Exception:
        container_gross_limit = 0

    if is_tray_type:
        tray_blocked_item = next((item for item in item_weights if item['weight'] > tray_load_limit), None)
        if tray_blocked_item:
            return f'El item "{tray_blocked_item["name"]}" pesa {tray_blocked_item["weight"]:.2f} kg y supera el límite de carga de la bandeja ({tray_load_limit:.2f} kg).'

    if has_pallet:
        if is_tray_type:
            min_tray_gross = tray_tare + lightest_item['weight']
            if (pallet_base_weight + min_tray_gross) > pallet_gross_limit:
                return (
                    f'El pallet no admite ni una bandeja cargada mínima: '
                    f'base {pallet_base_weight:.2f} kg + bandeja {tray_tare:.2f} kg + '
                    f'item {lightest_item["weight"]:.2f} kg > límite de pallet {pallet_gross_limit:.2f} kg.'
                )
        else:
            pallet_blocked_item = next((item for item in item_weights if (pallet_base_weight + item['weight']) > pallet_gross_limit), None)
            if pallet_blocked_item:
                return (
                    f'El item "{pallet_blocked_item["name"]}" supera el límite del pallet: '
                    f'base {pallet_base_weight:.2f} kg + item {pallet_blocked_item["weight"]:.2f} kg > {pallet_gross_limit:.2f} kg.'
                )

    if has_container and container_gross_limit > 0:
        if has_pallet:
            min_top_level_gross = pallet_base_weight + lightest_item['weight']
            if is_tray_type:
                min_top_level_gross += tray_tare
            if min_top_level_gross > container_gross_limit:
                return (
                    f'El contenedor no admite la unidad mínima de carga: '
                    f'{min_top_level_gross:.2f} kg > límite de contenedor {container_gross_limit:.2f} kg.'
                )
        elif is_tray_type:
            min_tray_gross = tray_tare + lightest_item['weight']
            if min_tray_gross > container_gross_limit:
                return (
                    f'El contenedor no admite una bandeja cargada mínima: '
                    f'{min_tray_gross:.2f} kg > límite de contenedor {container_gross_limit:.2f} kg.'
                )
        else:
            container_blocked_item = next((item for item in item_weights if item['weight'] > container_gross_limit), None)
            if container_blocked_item:
                return (
                    f'El item "{container_blocked_item["name"]}" pesa {container_blocked_item["weight"]:.2f} kg '
                    f'y supera el límite del contenedor ({container_gross_limit:.2f} kg).'
                )

    return None


@app.route('/api/logistics/calculate', methods=['POST'])
def calculate_logistics():
    print("DEBUG: >>> ENTERING calculate_logistics <<<", flush=True)
    try:
        data = request.json
        if not data: return jsonify({'status': 'error', 'message': 'No data received'}), 400
        container_data = data.get('container')
        items_data = data.get('items')
        config = data.get('config', {})
        job_id = config.get('job_id')
        _raise_if_logistics_calc_cancelled(job_id)
        if job_id:
            _update_logistics_calc_progress(job_id, status='running', stage='validation', progress=4, message='Validando datos...')
        if not container_data or not items_data: return jsonify({'status': 'error', 'message': 'Missing data'}), 400
        weight_limit_error = _validate_logistics_weight_limits(container_data, items_data, config)
        if weight_limit_error:
            return jsonify({'status': 'error', 'message': weight_limit_error}), 400
        
        # Save historical trace for internal logs if needed
        # (Already happening in debug logs)

        container_type_key = config.get('container_type', '20ft')

        if config.get('maximize'):
            print(f"DEBUG: Maximization requested for {len(items_data)} items.", flush=True)
            if job_id:
                _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=8, message='Calculando maximización...')

            def _result_fits_fully(result):
                return result.get('status') == 'success' and result.get('unfitted_count', 0) == 0

            def _max_rectangles_on_floor(floor_w, floor_d, rect_w, rect_d):
                try:
                    floor_w = int(float(floor_w or 0))
                    floor_d = int(float(floor_d or 0))
                    rect_w = int(float(rect_w or 0))
                    rect_d = int(float(rect_d or 0))
                except Exception:
                    return 0

                if floor_w <= 0 or floor_d <= 0 or rect_w <= 0 or rect_d <= 0:
                    return 0

                best = 0
                step_candidates = [rect_w, rect_d]
                step = max(1, math.gcd(*step_candidates))

                best = max(best, (floor_w // rect_w) * (floor_d // rect_d))
                best = max(best, (floor_w // rect_d) * (floor_d // rect_w))

                for split_w in range(0, floor_w + 1, step):
                    left = (split_w // rect_w) * (floor_d // rect_d)
                    right = ((floor_w - split_w) // rect_d) * (floor_d // rect_w)
                    best = max(best, left + right)

                    left_rot = (split_w // rect_d) * (floor_d // rect_w)
                    right_rot = ((floor_w - split_w) // rect_w) * (floor_d // rect_d)
                    best = max(best, left_rot + right_rot)

                for split_d in range(0, floor_d + 1, step):
                    top = (floor_w // rect_w) * (split_d // rect_d)
                    bottom = (floor_w // rect_d) * ((floor_d - split_d) // rect_w)
                    best = max(best, top + bottom)

                    top_rot = (floor_w // rect_d) * (split_d // rect_w)
                    bottom_rot = (floor_w // rect_w) * ((floor_d - split_d) // rect_d)
                    best = max(best, top_rot + bottom_rot)

                return int(best)

            def _estimate_single_qty_upper_bound(test_container, test_item_payload, test_conf):
                try:
                    item_w = float(test_item_payload.get('w', 0) or 0)
                    item_d = float(test_item_payload.get('d', 0) or 0)
                    item_h = float(test_item_payload.get('h', 0) or 0)
                    item_weight = float(test_item_payload.get('weight', 0) or 0)
                    if item_w <= 0 or item_d <= 0 or item_h <= 0:
                        return 4096

                    bin_w = float(test_container.get('width', 0) or 0)
                    bin_d = float(test_container.get('depth', 0) or 0)
                    bin_h = float(test_container.get('height', 0) or 0)
                    max_weight = float(test_container.get('max_weight', 0) or 0)

                    load_type = test_conf.get('load_type', 'loose')
                    pallet_type = test_conf.get('pallet_type', 'none')
                    container_type = test_conf.get('container_type', 'none')

                    if load_type in ('tray', 'tray_euro', 'tray_american', 'tray_custom'):
                        tray_dims = test_conf.get('tray_dims') or {}
                        bin_w = float(tray_dims.get('tray_inner_w', 0) or 0)
                        bin_d = float(tray_dims.get('tray_inner_d', 0) or 0)
                        bin_h = float(tray_dims.get('tray_inner_h', 0) or 0)
                        tray_tare = float(tray_dims.get('weight', 0) or 0)
                        tray_max_weight = float(tray_dims.get('max_weight', 25) or 25)
                        max_weight = max(tray_max_weight - tray_tare, 0)
                    elif pallet_type != 'none' and container_type == 'none':
                        pallet_dims = test_conf.get('pallet_dims') or {}
                        if float(pallet_dims.get('w', 0) or 0) > 0 and float(pallet_dims.get('d', 0) or 0) > 0:
                            bin_w = float(pallet_dims.get('w', 0) or 0)
                            bin_d = float(pallet_dims.get('d', 0) or 0)
                            pallet_base_h = float(pallet_dims.get('h', 150) or 150)
                            user_max_h = float(test_conf.get('max_pallet_height', 0) or 1800)
                            bin_h = max(user_max_h - pallet_base_h, 0)
                            pallet_base_weight = float(pallet_dims.get('weight', 25) or 25)
                            pallet_max_weight = float(test_conf.get('max_pallet_weight', 0) or 2000)
                            max_weight = max(pallet_max_weight - pallet_base_weight, 0)

                    bounds = []
                    if bin_w > 0 and bin_d > 0 and bin_h > 0:
                        item_vol = item_w * item_d * item_h
                        bin_vol = bin_w * bin_d * bin_h
                        if item_vol > 0 and bin_vol > 0:
                            bounds.append(max(int(bin_vol // item_vol), 1))
                    if max_weight > 0 and item_weight > 0:
                        bounds.append(max(int(max_weight // item_weight), 1))

                    if not bounds:
                        return 4096

                    upper_bound = min(bounds)
                    return max(1, min(upper_bound + 1, 50000))
                except Exception:
                    return 4096

            has_ratios = False
            for it in items_data:
                try:
                    if float(it.get('rel_qty') or 0) > 0:
                        has_ratios = True
                        break
                except:
                    pass

            if len(items_data) == 1 and not has_ratios:
                item = items_data[0]
                test_config = {k: v for k, v in config.items()}
                test_config['skip_max_cap'] = True
                test_config['progress_detail'] = False
                test_config['debug_verbose'] = False

                lo = 1
                hi = 1
                best = 0
                fit_cache = {}

                def _fits_single_qty(qty):
                    if qty in fit_cache:
                        return fit_cache[qty]
                    test_item = dict(item, qty=qty)
                    res_test = _do_pack_internal(container_data, [test_item], test_config)
                    fits = _result_fits_fully(res_test)
                    fit_cache[qty] = fits
                    return fits

                def _max_single_qty_for_config(test_container, test_item_payload, test_conf):
                    local_conf = {k: v for k, v in test_conf.items()}
                    local_conf['skip_max_cap'] = True
                    local_conf['progress_detail'] = False
                    local_conf['debug_verbose'] = False
                    upper_bound_local = _estimate_single_qty_upper_bound(test_container, test_item_payload, local_conf)
                    lo_local = 1
                    hi_local = 1
                    best_local = 0
                    local_fit_cache = {}

                    def _fits_local(qty_local):
                        if qty_local in local_fit_cache:
                            return local_fit_cache[qty_local]
                        payload = dict(test_item_payload, qty=qty_local)
                        res_local = _do_pack_internal(test_container, [payload], local_conf)
                        fits_local = _result_fits_fully(res_local)
                        local_fit_cache[qty_local] = fits_local
                        return fits_local

                    while hi_local <= upper_bound_local and _fits_local(hi_local):
                        _raise_if_logistics_calc_cancelled(job_id)
                        best_local = hi_local
                        if hi_local == upper_bound_local:
                            return best_local
                        hi_local = min(hi_local * 2, upper_bound_local)

                    lo_local = max(1, best_local + 1)
                    hi_local = max(min(hi_local - 1, upper_bound_local), lo_local)
                    while lo_local <= hi_local:
                        _raise_if_logistics_calc_cancelled(job_id)
                        mid_local = (lo_local + hi_local) // 2
                        if _fits_local(mid_local):
                            best_local = mid_local
                            lo_local = mid_local + 1
                        else:
                            hi_local = mid_local - 1
                    return best_local

                def _resolve_forced_rotations_local(local_conf):
                    if not local_conf.get('force_orientation'):
                        return None
                    face = str(local_conf.get('orientation_face', 'LxA') or 'LxA').strip().upper()
                    if face == 'LXA':
                        return [RotationType.RT_WHD, RotationType.RT_DHW]
                    if face == 'LXH':
                        return [RotationType.RT_WDH, RotationType.RT_HDW]
                    if face == 'AXH':
                        return [RotationType.RT_HWD, RotationType.RT_DWH]
                    return None

                def _max_single_qty_in_tray_direct(test_item_payload, tray_conf_local):
                    tray_dims_cfg = tray_conf_local.get('tray_dims') or {}
                    inner_w = float(tray_dims_cfg.get('tray_inner_w', 0) or 0)
                    inner_d = float(tray_dims_cfg.get('tray_inner_d', 0) or 0)
                    inner_h = float(tray_dims_cfg.get('tray_inner_h', 0) or 0)
                    tray_tare = float(tray_dims_cfg.get('weight', 0) or 0)
                    tray_max_weight = float(tray_dims_cfg.get('max_weight', 25) or 25)
                    max_load_weight = max(tray_max_weight - tray_tare, 0)
                    if inner_w <= 0 or inner_d <= 0 or inner_h <= 0:
                        return 0

                    forced_rot_local = _resolve_forced_rotations_local(tray_conf_local)
                    sf_mult_dims_local = Decimal(1) + (Decimal(str(float(tray_conf_local.get('safety_factor_dims', 0) or 0))) / Decimal(100))
                    sf_mult_weight_local = Decimal(1) + (Decimal(str(float(tray_conf_local.get('safety_factor_weight', 0) or 0))) / Decimal(100))

                    direct_upper = _estimate_single_qty_upper_bound(
                        {'width': inner_w, 'depth': inner_d, 'height': inner_h, 'max_weight': max_load_weight},
                        test_item_payload,
                        {'load_type': 'loose', 'pallet_type': 'none', 'container_type': 'none'}
                    )
                    direct_cache = {}

                    def _fits_direct(qty_direct):
                        if qty_direct in direct_cache:
                            return direct_cache[qty_direct]
                        if qty_direct <= 0:
                            direct_cache[qty_direct] = False
                            return False
                        try:
                            tray_bin = Bin(
                                name='TrayDirect',
                                width=Decimal(str(inner_w)),
                                depth=Decimal(str(inner_d)),
                                height=Decimal(str(inner_h)),
                                max_weight=Decimal(str(max_load_weight)),
                                allow_stacking=True
                            )
                            tray_packer = Packer()
                            tray_packer.add_bin(tray_bin)
                            for _ in range(int(qty_direct)):
                                direct_item = Item(
                                    name=test_item_payload.get('id'),
                                    width=Decimal(test_item_payload.get('w', 0)) * sf_mult_dims_local,
                                    height=Decimal(test_item_payload.get('h', 0)) * sf_mult_dims_local,
                                    depth=Decimal(test_item_payload.get('d', 0)) * sf_mult_dims_local,
                                    weight=Decimal(test_item_payload.get('weight', 0)) * sf_mult_weight_local,
                                    allowed_rotations=forced_rot_local if forced_rot_local else RotationType.ALL
                                )
                                if forced_rot_local:
                                    direct_item.force_orientation = True
                                tray_packer.add_item(direct_item)
                            tray_packer.pack(bigger_first=True)
                            fits_direct = len(tray_packer.unfit_items) == 0
                        except Exception:
                            fits_direct = False
                        direct_cache[qty_direct] = fits_direct
                        return fits_direct

                    lo_direct = 1
                    hi_direct = 1
                    best_direct = 0
                    while hi_direct <= direct_upper and _fits_direct(hi_direct):
                        best_direct = hi_direct
                        if hi_direct == direct_upper:
                            return best_direct
                        hi_direct = min(hi_direct * 2, direct_upper)

                    lo_direct = max(1, best_direct + 1)
                    hi_direct = max(min(hi_direct - 1, direct_upper), lo_direct)
                    while lo_direct <= hi_direct:
                        mid_direct = (lo_direct + hi_direct) // 2
                        if _fits_direct(mid_direct):
                            best_direct = mid_direct
                            lo_direct = mid_direct + 1
                        else:
                            hi_direct = mid_direct - 1
                    return best_direct

                def _maximize_tray_kanban_chain(single_item_payload):
                    def _extract_tray_groups_for_item(result_payload, item_name):
                        tray_qty_counter = {}
                        for grouped in (result_payload or {}).get('grouped_pallets', []) or []:
                            trays_to_scan = []
                            if grouped.get('type') in ('pallet', 'pallet_group'):
                                trays_to_scan = grouped.get('trays', []) or []
                            elif grouped.get('type') == 'tray':
                                trays_to_scan = [grouped]

                            for tray_group in trays_to_scan:
                                tray_count = int(tray_group.get('count', 0) or 0)
                                if tray_count <= 0:
                                    continue
                                for tray_item in tray_group.get('items', []) or []:
                                    if str(tray_item.get('name', '')) != str(item_name):
                                        continue
                                    qty_per_tray = int(tray_item.get('qty_per_tray', 0) or 0)
                                    if qty_per_tray > 0:
                                        tray_qty_counter[qty_per_tray] = tray_qty_counter.get(qty_per_tray, 0) + tray_count
                        return tray_qty_counter

                    def _adjust_result_to_full_kanban_trays(result_payload):
                        tray_qty_counter = _extract_tray_groups_for_item(result_payload, single_item_payload.get('id'))
                        if not tray_qty_counter:
                            return None, 0
                        full_tray_qty = max(tray_qty_counter.keys())
                        full_tray_count = int(tray_qty_counter.get(full_tray_qty, 0) or 0)
                        adjusted_qty = int(full_tray_qty * full_tray_count)
                        print(
                            f"DEBUG: [KANBAN] Tray groups for {single_item_payload.get('id')}: "
                            f"{tray_qty_counter}, chosen={full_tray_count}x{full_tray_qty} => {adjusted_qty}",
                            flush=True
                        )
                        if adjusted_qty <= 0:
                            return None, 0
                        return adjusted_qty, int(full_tray_qty)

                    standard_conf = {k: v for k, v in config.items()}
                    standard_conf['tray_kanban'] = False
                    standard_conf['progress_detail'] = False
                    standard_conf['debug_verbose'] = False

                    existing_qty = int(single_item_payload.get('qty', 0) or 0)
                    if existing_qty > 0:
                        existing_res = _do_pack_internal(
                            container_data,
                            [dict(single_item_payload, qty=existing_qty)],
                            standard_conf
                        )
                        if existing_res.get('status') == 'success':
                            adjusted_from_existing, existing_tray_qty = _adjust_result_to_full_kanban_trays(existing_res)
                            if adjusted_from_existing and existing_tray_qty > 0:
                                print(
                                    f"DEBUG: [KANBAN] Using existing quantity for {single_item_payload.get('id')}: "
                                    f"existing={existing_qty} adjusted={adjusted_from_existing}",
                                    flush=True
                                )
                                return adjusted_from_existing, None, existing_tray_qty

                    standard_best = _max_single_qty_for_config(container_data, single_item_payload, standard_conf)
                    if standard_best > 0:
                        standard_res = _do_pack_internal(
                            container_data,
                            [dict(single_item_payload, qty=int(standard_best))],
                            standard_conf
                        )
                        if standard_res.get('status') == 'success':
                            adjusted_qty, adjusted_tray_qty = _adjust_result_to_full_kanban_trays(standard_res)
                            if adjusted_qty and adjusted_tray_qty > 0:
                                print(
                                    f"DEBUG: [KANBAN] Standard maximize for {single_item_payload.get('id')}: "
                                    f"qty={standard_best} adjusted={adjusted_qty}",
                                    flush=True
                                )
                                return adjusted_qty, None, adjusted_tray_qty

                    tray_dims_cfg = config.get('tray_dims') or {}
                    tray_only_conf = {k: v for k, v in config.items()}
                    tray_only_conf['pallet_type'] = 'none'
                    tray_only_conf['container_type'] = 'none'
                    tray_only_conf['load_type'] = config.get('load_type')
                    tray_only_conf['tray_kanban'] = False

                    tray_qty = _max_single_qty_in_tray_direct(single_item_payload, tray_only_conf)
                    if tray_qty <= 0:
                        tray_qty = _max_single_qty_for_config(
                            {'type': 'none', 'name': 'Sin Contenedor', 'width': 0, 'height': 0, 'depth': 0, 'max_weight': 0},
                            single_item_payload,
                            tray_only_conf
                        )
                    print(f"DEBUG: [KANBAN] Tray capacity for {single_item_payload.get('id')}: {tray_qty}", flush=True)
                    if tray_qty <= 0:
                        return None, f'No entra una bandeja Kanban completa para {single_item_payload.get("id")}.', 0

                    tray_outer_w = float(tray_dims_cfg.get('tray_outer_w', 0) or 0)
                    tray_outer_d = float(tray_dims_cfg.get('tray_outer_d', 0) or 0)
                    tray_outer_h = float(tray_dims_cfg.get('tray_outer_h', 0) or 0)
                    tray_weight_unit = float(tray_dims_cfg.get('weight', 0) or 0)
                    item_weight_unit = float(single_item_payload.get('weight', 0) or 0)
                    tray_total_weight = tray_weight_unit + (item_weight_unit * tray_qty)
                    print(
                        f"DEBUG: [KANBAN] Tray gross weight for {single_item_payload.get('id')}: "
                        f"tare={tray_weight_unit}, item_unit={item_weight_unit}, qty_per_tray={tray_qty}, gross={tray_total_weight}",
                        flush=True
                    )
                    tray_payload = {
                        'id': 'TRAY_UNIT',
                        'w': tray_outer_w,
                        'd': tray_outer_d,
                        'h': tray_outer_h,
                        'weight': tray_total_weight,
                        'qty': 1
                    }

                    final_qty = tray_qty

                    if config.get('pallet_type', 'none') != 'none' and tray_outer_w > 0 and tray_outer_d > 0 and tray_outer_h > 0:
                        if job_id:
                            _update_logistics_calc_progress(job_id, status='running', stage='tray', progress=28, message='Calculando bandejas por pallet...')
                        pallet_type_key = config.get('pallet_type', 'none')
                        pallet_dims_cfg = config.get('pallet_dims') or {}
                        pallet_outer_w = float(pallet_dims_cfg.get('w', 0) or 0)
                        pallet_outer_d = float(pallet_dims_cfg.get('d', 0) or 0)
                        pallet_base_h = float(pallet_dims_cfg.get('h', 150) or 150)
                        pallet_base_weight = float(pallet_dims_cfg.get('weight', 25) or 25)

                        if pallet_type_key == 'europallet':
                            pallet_outer_w = 1200
                            pallet_outer_d = 800
                            pallet_base_h = 150
                            if float(pallet_dims_cfg.get('weight', 0) or 0) <= 0:
                                pallet_base_weight = 25
                        elif pallet_type_key == 'american':
                            pallet_outer_w = 1200
                            pallet_outer_d = 1000
                            pallet_base_h = 150
                            if float(pallet_dims_cfg.get('weight', 0) or 0) <= 0:
                                pallet_base_weight = 25

                        user_max_h = float(config.get('max_pallet_height', 0) or 1800)
                        # In Kanban mode we treat the configured pallet max height as usable cargo height,
                        # matching the business rule used by OT for tray stacking calculations.
                        usable_load_h = max(user_max_h, 0)
                        layers_by_height = max(int(usable_load_h // tray_outer_h), 1 if tray_outer_h > 0 else 0)
                        trays_per_layer = _max_rectangles_on_floor(pallet_outer_w, pallet_outer_d, tray_outer_w, tray_outer_d)
                        tray_weight_for_limit = tray_total_weight
                        max_pallet_weight = float(config.get('max_pallet_weight', 0) or 2000)
                        allowed_load_weight = max(max_pallet_weight - pallet_base_weight, 0)
                        trays_by_weight = int(allowed_load_weight // tray_weight_for_limit) if tray_weight_for_limit > 0 else 0
                        trays_by_volume = trays_per_layer * layers_by_height
                        if trays_by_weight > 0:
                            pallet_per_qty = min(trays_by_volume, trays_by_weight)
                        else:
                            pallet_per_qty = trays_by_volume
                        print(
                            f"DEBUG: [KANBAN] Pallet analytic capacity for {single_item_payload.get('id')}: "
                            f"per_layer={trays_per_layer}, layers={layers_by_height}, by_volume={trays_by_volume}, "
                            f"by_weight={trays_by_weight}, chosen={pallet_per_qty}",
                            flush=True
                        )
                        print(f"DEBUG: [KANBAN] Trays per pallet for {single_item_payload.get('id')}: {pallet_per_qty}", flush=True)
                        if pallet_per_qty <= 0:
                            return None, f'No entra una bandeja Kanban completa en el pallet para {single_item_payload.get("id")}.', tray_qty

                        final_qty *= pallet_per_qty

                        if container_type_key != 'none':
                            if job_id:
                                _update_logistics_calc_progress(job_id, status='running', stage='pallet', progress=42, message='Calculando pallets por contenedor...')
                            pallet_w = float(pallet_outer_w or 0)
                            pallet_d = float(pallet_outer_d or 0)
                            pallet_h = float(config.get('max_pallet_height', 0) or 1800)
                            pallet_weight_unit = float(pallet_base_weight or 0) + (tray_total_weight * pallet_per_qty)

                            c_l = float(container_data.get('width', 0) or 0)
                            c_d = float(container_data.get('depth', 0) or 0)
                            template_cap = 0
                            is_20ft = (5850 <= c_l < 6000 and 2300 <= c_d < 2400)
                            is_40ft = (12000 <= c_l < 12100 and 2300 <= c_d < 2400)
                            is_40pw = (12000 <= c_l < 12100 and 2440 <= c_d < 2550)
                            is_euro = (pallet_w == 1200 and pallet_d == 800)
                            is_american = (pallet_w == 1200 and pallet_d == 1000)
                            if is_20ft:
                                if is_american:
                                    template_cap = 10
                                elif is_euro:
                                    template_cap = 11
                            elif is_40ft:
                                if is_american:
                                    template_cap = 21
                                elif is_euro:
                                    template_cap = 25
                            elif is_40pw:
                                if is_american:
                                    template_cap = 24
                                elif is_euro:
                                    template_cap = 30

                            if template_cap > 0:
                                print(f"DEBUG: [KANBAN] Template pallets per container for {single_item_payload.get('id')}: {template_cap}", flush=True)
                                final_qty *= template_cap
                            else:
                                pallet_payload = {
                                    'id': 'PALLET_UNIT',
                                    'w': pallet_w,
                                    'd': pallet_d,
                                    'h': pallet_h,
                                    'weight': pallet_weight_unit,
                                    'qty': 1
                                }
                                container_conf = {k: v for k, v in config.items()}
                                container_conf['load_type'] = 'loose'
                                container_conf['pallet_type'] = 'none'
                                container_conf['tray_kanban'] = False
                                container_conf['force_orientation'] = True
                                container_conf['orientation_face'] = 'lxa'
                                pallet_container_qty = _max_single_qty_for_config(container_data, pallet_payload, container_conf)
                                print(f"DEBUG: [KANBAN] Pallets per container for {single_item_payload.get('id')}: {pallet_container_qty}", flush=True)
                                if pallet_container_qty <= 0:
                                    return None, f'No entra un pallet con bandejas Kanban en el contenedor para {single_item_payload.get("id")}.', tray_qty
                                final_qty *= pallet_container_qty
                    elif container_type_key != 'none':
                        if job_id:
                            _update_logistics_calc_progress(job_id, status='running', stage='pallet', progress=42, message='Calculando bandejas por contenedor...')
                        container_conf = {k: v for k, v in config.items()}
                        container_conf['load_type'] = 'loose'
                        container_conf['pallet_type'] = 'none'
                        container_conf['tray_kanban'] = False
                        container_conf['force_orientation'] = True
                        container_conf['orientation_face'] = 'lxa'
                        trays_per_container = _max_single_qty_for_config(container_data, tray_payload, container_conf)
                        print(f"DEBUG: [KANBAN] Trays per container for {single_item_payload.get('id')}: {trays_per_container}", flush=True)
                        if trays_per_container <= 0:
                            return None, f'No entra una bandeja Kanban completa en el contenedor para {single_item_payload.get("id")}.', tray_qty
                        final_qty *= trays_per_container

                    print(f"DEBUG: [KANBAN] Final maximized quantity for {single_item_payload.get('id')}: {final_qty}", flush=True)

                    return int(final_qty), None, int(tray_qty)

                tray_load_types = ('tray', 'tray_euro', 'tray_american', 'tray_custom')
                if config.get('load_type') in tray_load_types and config.get('tray_kanban'):
                    final_qty, kanban_error, kanban_tray_qty = _maximize_tray_kanban_chain(item)
                    if final_qty and final_qty > 0:
                        total_trays_upper = max(int(final_qty // max(kanban_tray_qty, 1)), 1)
                        verify_cache = {}

                        def _fits_kanban_total_trays(tray_count_probe):
                            if tray_count_probe in verify_cache:
                                return verify_cache[tray_count_probe]
                            if tray_count_probe <= 0:
                                verify_cache[tray_count_probe] = False
                                return False
                            probe_qty = int(kanban_tray_qty * tray_count_probe)
                            probe_conf = {k: v for k, v in config.items()}
                            probe_conf['progress_detail'] = False
                            probe_conf['debug_verbose'] = False
                            probe_res = _do_pack_internal(container_data, [dict(item, qty=probe_qty)], probe_conf)
                            fits_probe = (
                                probe_res.get('status') == 'success' and
                                int(probe_res.get('unfitted_count', 0) or 0) == 0 and
                                int(probe_res.get('unfitted_from_palletizing_count', 0) or 0) == 0 and
                                int(probe_res.get('unfitted_final_count', 0) or 0) == 0
                            )
                            verify_cache[tray_count_probe] = fits_probe
                            print(
                                f"DEBUG: [KANBAN] Verify trays={tray_count_probe} qty={probe_qty} fits={fits_probe}",
                                flush=True
                            )
                            return fits_probe

                        lo_trays = 1
                        hi_trays = total_trays_upper
                        best_trays = 0
                        while lo_trays <= hi_trays:
                            _raise_if_logistics_calc_cancelled(job_id)
                            mid_trays = (lo_trays + hi_trays) // 2
                            if job_id:
                                _update_logistics_calc_progress(
                                    job_id,
                                    status='running',
                                    stage='maximize',
                                    progress=48,
                                    message=f'Validando {mid_trays} bandejas Kanban...'
                                )
                            if _fits_kanban_total_trays(mid_trays):
                                best_trays = mid_trays
                                lo_trays = mid_trays + 1
                            else:
                                hi_trays = mid_trays - 1

                        final_qty = int(best_trays * kanban_tray_qty)
                        print(
                            f"DEBUG: [KANBAN] Verified maximum trays for {item.get('id')}: {best_trays} "
                            f"(qty={final_qty})",
                            flush=True
                        )
                        if final_qty <= 0:
                            return jsonify({'status': 'error', 'message': f'No entra una bandeja Kanban completa para {item.get("id")}.'}), 400

                        if job_id:
                            _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=58, message='Empaquetando resultado final...')
                        final_res = _do_pack_internal(container_data, [dict(item, qty=int(final_qty))], config)
                        if final_res.get('status') == 'success':
                            final_res['new_quantities'] = {item['id']: int(final_qty)}
                            return jsonify(final_res)
                        print(
                            f"DEBUG: [KANBAN] Final packing returned non-success for {item.get('id')}: "
                            f"{final_res.get('message', final_res.get('status'))}",
                            flush=True
                        )
                        return jsonify({'status': 'error', 'message': final_res.get('message', 'No se pudo maximizar el ítem con Bandeja Kanban.')}), 500
                    if kanban_error:
                        return jsonify({'status': 'error', 'message': kanban_error}), 400
                elif config.get('load_type') in tray_load_types:
                    try:
                        tray_dims_cfg = config.get('tray_dims') or {}
                        tray_only_conf = {k: v for k, v in config.items()}
                        tray_only_conf['pallet_type'] = 'none'
                        tray_only_conf['container_type'] = 'none'
                        tray_only_conf['load_type'] = config.get('load_type')
                        tray_only_conf['tray_kanban'] = False
                        tray_qty = _max_single_qty_for_config(
                            {'type': 'none', 'name': 'Sin Contenedor', 'width': 0, 'height': 0, 'depth': 0, 'max_weight': 0},
                            item,
                            tray_only_conf
                        )

                        if tray_qty > 0:
                            final_qty = tray_qty
                            tray_outer_w = float(tray_dims_cfg.get('tray_outer_w', 0) or 0)
                            tray_outer_d = float(tray_dims_cfg.get('tray_outer_d', 0) or 0)
                            tray_outer_h = float(tray_dims_cfg.get('tray_outer_h', 0) or 0)
                            tray_weight_unit = float(tray_dims_cfg.get('weight', 0) or 0)
                            item_weight_unit = float(item.get('weight', 0) or 0)
                            tray_total_weight = tray_weight_unit + (item_weight_unit * tray_qty)

                            if config.get('pallet_type', 'none') != 'none' and tray_outer_w > 0 and tray_outer_d > 0 and tray_outer_h > 0:
                                if job_id:
                                    _update_logistics_calc_progress(job_id, status='running', stage='tray', progress=28, message='Calculando bandejas por pallet...')
                                tray_payload = {
                                    'id': 'TRAY_UNIT',
                                    'w': tray_outer_w,
                                    'd': tray_outer_d,
                                    'h': tray_outer_h,
                                    'weight': tray_total_weight,
                                    'qty': 1
                                }
                                pallet_conf = {k: v for k, v in config.items()}
                                pallet_conf['load_type'] = 'loose'
                                pallet_conf['container_type'] = 'none'
                                pallet_per_qty = _max_single_qty_for_config(
                                    {'type': 'none', 'name': 'Sin Contenedor', 'width': 0, 'height': 0, 'depth': 0, 'max_weight': 0},
                                    tray_payload,
                                    pallet_conf
                                )
                                if pallet_per_qty > 0:
                                    final_qty *= pallet_per_qty

                                    if container_type_key != 'none':
                                        if job_id:
                                            _update_logistics_calc_progress(job_id, status='running', stage='pallet', progress=42, message='Calculando pallets por contenedor...')
                                        pallet_dims_cfg = config.get('pallet_dims') or {}
                                        pallet_w = float(pallet_dims_cfg.get('w', 0) or 0)
                                        pallet_d = float(pallet_dims_cfg.get('d', 0) or 0)
                                        pallet_h = float(config.get('max_pallet_height', 0) or 1800)
                                        pallet_weight_unit = float(pallet_dims_cfg.get('weight', 0) or 0) + (tray_total_weight * pallet_per_qty)

                                        c_l = float(container_data.get('width', 0) or 0)
                                        c_d = float(container_data.get('depth', 0) or 0)
                                        template_cap = 0
                                        is_20ft = (5850 <= c_l < 6000 and 2300 <= c_d < 2400)
                                        is_40ft = (12000 <= c_l < 12100 and 2300 <= c_d < 2400)
                                        is_40pw = (12000 <= c_l < 12100 and 2440 <= c_d < 2550)
                                        is_euro = (pallet_w == 1200 and pallet_d == 800)
                                        is_american = (pallet_w == 1200 and pallet_d == 1000)
                                        if is_20ft:
                                            if is_american: template_cap = 10
                                            elif is_euro: template_cap = 11
                                        elif is_40ft:
                                            if is_american: template_cap = 21
                                            elif is_euro: template_cap = 25
                                        elif is_40pw:
                                            if is_american: template_cap = 24
                                            elif is_euro: template_cap = 30

                                        if template_cap > 0:
                                            final_qty *= template_cap
                                        else:
                                            pallet_payload = {
                                                'id': 'PALLET_UNIT',
                                                'w': pallet_w,
                                                'd': pallet_d,
                                                'h': pallet_h,
                                                'weight': pallet_weight_unit,
                                                'qty': 1
                                            }
                                            container_conf = {k: v for k, v in config.items()}
                                            container_conf['load_type'] = 'loose'
                                            container_conf['pallet_type'] = 'none'
                                            pallet_container_qty = _max_single_qty_for_config(container_data, pallet_payload, container_conf)
                                            if pallet_container_qty > 0:
                                                final_qty *= pallet_container_qty

                                    if job_id:
                                        _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=58, message='Empaquetando resultado final...')
                                    final_res = _do_pack_internal(container_data, [dict(item, qty=int(final_qty))], config)
                                    if final_res.get('status') == 'success':
                                        final_res['new_quantities'] = {item['id']: int(final_qty)}
                                        return jsonify(final_res)
                                    print(
                                        f"DEBUG: Single-item tray fast path produced non-success result for {item.get('id')}: "
                                        f"{final_res.get('message', final_res.get('status'))}",
                                        flush=True
                                    )
                        elif config.get('tray_kanban'):
                            return jsonify({
                                'status': 'error',
                                'message': f'No entra una bandeja Kanban completa para {item.get("id")}.'
                            }), 400
                    except Exception as fast_path_exc:
                        print(f"DEBUG: Single-item tray fast path fallback: {fast_path_exc}")

                if job_id:
                    _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=12, message=f"Buscando máximo para {item.get('id')}...")

                upper_bound = _estimate_single_qty_upper_bound(container_data, item, test_config)
                print(f"DEBUG: [MAX] Upper bound estimated for {item.get('id')}: {upper_bound}", flush=True)
                hi = 1
                while hi <= upper_bound and _fits_single_qty(hi):
                    _raise_if_logistics_calc_cancelled(job_id)
                    best = hi
                    if hi == upper_bound:
                        break
                    hi = min(hi * 2, upper_bound)
                    print(f"DEBUG: [MAX] Quantity {best} fits for {item.get('id')}. Next probe: {hi}", flush=True)
                    if job_id:
                        capped_progress = min(40, 12 + int(math.log2(max(hi, 1))) * 4)
                        _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=capped_progress, message=f"Probando {hi} unidades...")

                lo = max(1, best + 1)
                hi = max(min(hi - 1, upper_bound), lo)
                while lo <= hi:
                    _raise_if_logistics_calc_cancelled(job_id)
                    mid = (lo + hi) // 2
                    print(f"DEBUG: [MAX] Binary search for {item.get('id')}. Range=({lo},{hi}) Mid={mid}", flush=True)
                    if job_id:
                        _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=46, message=f"Afinando capacidad: {mid} unidades...")
                    if _fits_single_qty(mid):
                        best = mid
                        print(f"DEBUG: [MAX] Mid {mid} fits for {item.get('id')}", flush=True)
                        lo = mid + 1
                    else:
                        print(f"DEBUG: [MAX] Mid {mid} does NOT fit for {item.get('id')}", flush=True)
                        hi = mid - 1

                if best <= 0:
                    return jsonify({'status': 'error', 'message': 'No se pudo maximizar el ítem.'}), 500

                print(f"DEBUG: [MAX] Best quantity found for {item.get('id')}: {best}", flush=True)
                if job_id:
                    _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=58, message='Empaquetando resultado final...')
                final_res = _do_pack_internal(container_data, [dict(item, qty=best)], config)
                if final_res.get('status') == 'success':
                    final_res['new_quantities'] = {item['id']: best}
                    return jsonify(final_res)
                return jsonify({'status': 'error', 'message': 'No se pudo maximizar el ítem.'}), 500
            
            # SPECIAL CASE: Multi-Item Single Pallet Maximization
            # When: No Container + Maximize + Pallets NOT Mixed
            # Each item gets its own maximized pallet
            if container_type_key == 'none' and config.get('pallet_type', 'none') != 'none' and not config.get('mixed_pallets', True):
                print(f"DEBUG: Per-Item Single Pallet Maximization Mode (mixed_pallets=false)")
                print(f"DEBUG: Input items: {[it['id'] for it in items_data]}")
                
                # Strategy: Find max quantity for each item, then pack all together in ONE call
                new_quantities = {it['id']: 0 for it in items_data}
                maximized_items = []
                
                # For binary search, use a test config that ALLOWS mixing (single item anyway)
                test_config = {k: v for k, v in config.items()}
                test_config['mixed_pallets'] = True  # Override for testing
                test_config['progress_detail'] = False
                test_config['debug_verbose'] = False
                
                # IMPORTANT: Process ALL items, even those with qty=0
                # In maximization mode, qty=0 means "find the maximum", not "skip this item"
                for idx, item in enumerate(items_data):
                    _raise_if_logistics_calc_cancelled(job_id)
                    if job_id and items_data:
                        _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=10 + int((idx / max(len(items_data), 1)) * 35), message=f"Maximizando {item['id']}...")
                    print(f"DEBUG: Finding max for item: {item['id']}")
                    if config.get('load_type') in ('tray', 'tray_euro', 'tray_american', 'tray_custom') and config.get('tray_kanban'):
                        best, kanban_error, kanban_tray_qty = _maximize_tray_kanban_chain(item)
                        if kanban_error:
                            print(f"DEBUG: [KANBAN] Skipping {item['id']} in per-item maximize: {kanban_error}", flush=True)
                            best = 0
                        elif best and kanban_tray_qty > 0:
                            tray_upper = max(int(best // kanban_tray_qty), 1)
                            lo_trays = 1
                            hi_trays = tray_upper
                            best_trays = 0
                            verify_cache = {}

                            def _fits_item_trays(tray_count_probe):
                                if tray_count_probe in verify_cache:
                                    return verify_cache[tray_count_probe]
                                probe_conf = {k: v for k, v in config.items()}
                                probe_conf['progress_detail'] = False
                                probe_conf['debug_verbose'] = False
                                probe_qty = int(tray_count_probe * kanban_tray_qty)
                                probe_res = _do_pack_internal(container_data, [dict(item, qty=probe_qty)], probe_conf)
                                fits_probe = (
                                    probe_res.get('status') == 'success' and
                                    int(probe_res.get('unfitted_count', 0) or 0) == 0 and
                                    int(probe_res.get('unfitted_from_palletizing_count', 0) or 0) == 0 and
                                    int(probe_res.get('unfitted_final_count', 0) or 0) == 0
                                )
                                verify_cache[tray_count_probe] = fits_probe
                                return fits_probe

                            while lo_trays <= hi_trays:
                                _raise_if_logistics_calc_cancelled(job_id)
                                mid_trays = (lo_trays + hi_trays) // 2
                                if _fits_item_trays(mid_trays):
                                    best_trays = mid_trays
                                    lo_trays = mid_trays + 1
                                else:
                                    hi_trays = mid_trays - 1
                            best = int(best_trays * kanban_tray_qty)
                    else:
                        best = _max_single_qty_for_config(container_data, item, test_config)
                    
                    print(f"DEBUG: Max found for {item['id']}: {best}")
                    if best > 0:
                        new_quantities[item['id']] = best
                        maximized_items.append(dict(item, qty=best))
                
                print(f"DEBUG: Maximized items to pack: {[(it['id'], it['qty']) for it in maximized_items]}")
                
                # Now pack ALL maximized items together in ONE call
                # Use the ORIGINAL config which has mixed_pallets=False
                # This will separate each item type into its own pallet(s)
                if len(maximized_items) > 0:
                    if job_id:
                        _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=52, message='Empaquetando resultado final...')
                    res = _do_pack_internal(container_data, maximized_items, config)
                    
                    print(f"DEBUG: Final packing result status: {res.get('status')}")
                    print(f"DEBUG: Number of pallets: {len(res.get('grouped_pallets', []))}")
                    print(f"DEBUG: Number of packed items: {len(res.get('packed_items', []))}")
                    
                    if res.get('status') == 'success':
                        res['new_quantities'] = new_quantities
                        return jsonify(res)
                    else:
                        return jsonify({'status': 'error', 'message': 'No se pudo maximizar ningÃºn Ã­tem.'}), 500
                else:
                    return jsonify({'status': 'error', 'message': 'No se pudo maximizar ningÃºn Ã­tem.'}), 500
            
            # Refined Strategy: Adaptive Single-Pass Greedy
            # 1. If Ratios defined: Interleave & Disable Sort (Fairness)
            # 2. If No Ratios: Block Gen & Enable Sort (Volume Max)
            try:
                c_w = float(container_data.get('width', 0) or 120)
                c_h = float(container_data.get('height', 0) or 250)
                c_d = float(container_data.get('depth', 0) or 80)
                container_vol = float(c_w * c_h * c_d)
                if container_vol <= 0: container_vol = 2400000.0
            except:
                container_vol = 2400000.0

            def get_base_qty(it):
                try:
                    rel = float(it.get('rel_qty') or 0)
                    if rel > 0: return rel
                    cur = float(it.get('qty') or 0)
                    return cur if cur > 0 else 1.0
                except: return 1.0

            total_set_vol = 0
            for it in items_data:
                try:
                    i_vol = float(it.get('w', 0)) * float(it.get('h', 0)) * float(it.get('d', 0))
                    total_set_vol += get_base_qty(it) * i_vol
                except: pass

            effective_capacity_vol = container_vol
            try:
                tray_dims = config.get('tray_dims') or {}
                if config.get('load_type') in ('tray', 'tray_euro', 'tray_american', 'tray_custom'):
                    tray_inner_vol = float((tray_dims.get('tray_inner_w', 0) or 0) * (tray_dims.get('tray_inner_d', 0) or 0) * (tray_dims.get('tray_inner_h', 0) or 0))
                    tray_outer_vol = float((tray_dims.get('tray_outer_w', 0) or 0) * (tray_dims.get('tray_outer_d', 0) or 0) * (tray_dims.get('tray_outer_h', 0) or 0))
                    if tray_inner_vol > 0 and tray_outer_vol > 0:
                        effective_capacity_vol *= (tray_inner_vol / tray_outer_vol)

                pallet_type_for_cap = config.get('pallet_type', 'none')
                pallet_dims = config.get('pallet_dims') or {}
                if pallet_type_for_cap != 'none':
                    pallet_w = float(pallet_dims.get('w', 0) or 0)
                    pallet_d = float(pallet_dims.get('d', 0) or 0)
                    pallet_h = float(pallet_dims.get('h', 150) or 150)
                    max_h = float(config.get('max_pallet_height', 0) or 1800)
                    if pallet_w > 0 and pallet_d > 0 and max_h > pallet_h:
                        effective_capacity_vol *= max((max_h - pallet_h) / max_h, 0.05)
            except Exception:
                effective_capacity_vol = container_vol

            multiplier = (effective_capacity_vol / total_set_vol * 1.3) if total_set_vol > 0 else 100
            global_item_limit = float('inf')
            
            max_items = []
            
            if has_ratios:
                # RATIO-SET MAXIMIZATION (Enforce complete sets)
                ratio_items = []
                fixed_items = []
                for it in items_data:
                    rel = float(it.get('rel_qty') or 0)
                    if rel > 0:
                        ratio_items.append((it, rel))
                    else:
                        q = int(it.get('qty') or 0)
                        if q > 0:
                            fixed_items.append(it)

                def build_items_for_sets(k):
                    built = []
                    for it, rel in ratio_items:
                        q = int(rel * k)
                        if q > 0:
                            new_it = dict(it)
                            new_it['qty'] = q
                            built.append(new_it)
                    for it in fixed_items:
                        q = int(it.get('qty') or 0)
                        if q > 0:
                            new_it = dict(it)
                            new_it['qty'] = q
                            built.append(new_it)
                    return built

                total_rel = sum(rel for _, rel in ratio_items)
                max_sets_limit = int(40000 / max(total_rel, 1))
                if max_sets_limit < 1:
                    max_sets_limit = 1

                config_for_test = dict(config)
                config_for_test['maximize'] = True
                config_for_test['skip_max_cap'] = True
                config_for_test['progress_detail'] = False

                best = 0
                hi = 1
                while hi <= max_sets_limit:
                    _raise_if_logistics_calc_cancelled(job_id)
                    if job_id:
                        _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=min(40, 12 + hi), message='Buscando capacidad máxima...')
                    test_items = build_items_for_sets(hi)
                    res_test = _do_pack_internal(container_data, test_items, config_for_test)
                    if res_test.get('status') != 'success' or res_test.get('unfitted_count', 0) > 0:
                        break
                    best = hi
                    hi *= 2

                lo = best + 1
                hi = min(hi - 1, max_sets_limit)
                while lo <= hi:
                    _raise_if_logistics_calc_cancelled(job_id)
                    if job_id:
                        _update_logistics_calc_progress(job_id, status='running', stage='maximize', progress=45, message='Afinando capacidad máxima...')
                    mid = (lo + hi) // 2
                    test_items = build_items_for_sets(mid)
                    res_mid = _do_pack_internal(container_data, test_items, config_for_test)
                    if res_mid.get('status') == 'success' and res_mid.get('unfitted_count', 0) == 0:
                        best = mid
                        lo = mid + 1
                    else:
                        hi = mid - 1

                final_items = build_items_for_sets(best)
                if job_id:
                    _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=55, message='Empaquetando resultado final...')
                res = _do_pack_internal(container_data, final_items, config_for_test)

                if res.get('status') == 'success':
                    new_quantities = {it['id']: 0 for it in items_data}
                    for it, rel in ratio_items:
                        new_quantities[it['id']] = int(rel * best)
                    for it in fixed_items:
                        new_quantities[it['id']] = int(it.get('qty') or 0)
                    res['new_quantities'] = new_quantities
                    res['set_count'] = best
                    res['unfitted_items'] = []
                    res['unfitted_count'] = 0
                    return jsonify(res)

                return jsonify({'status': 'error', 'message': 'Fallo en el cÃ¡lculo de maximizaciÃ³n por sets.'}), 500
                
            else:
                # BLOCK GENERATION (Volume Efficiency)
                # Standard generation, let the packer sort by volume
                for it in items_data:
                    q = int(get_base_qty(it) * multiplier)
                    q = min(q, global_item_limit)
                    if q < 1: q = 1
                    max_items.append(dict(it, qty=q))
                should_sort = True # Enable sorting in packer

            print(f"DEBUG: Packing Mode: {'Interleaved/NoSort' if has_ratios else 'Block/Sort'}. Items: {len(max_items)}")
            
            # Pass 'sort_items' through config or arguments?
            # _do_pack_internal takes config. We can inject it there.
            config['sort_items'] = should_sort
            if job_id:
                _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=20, message='Ejecutando solver...')
            res = _do_pack_internal(container_data, max_items, config)
            
            if res.get('status') == 'success':
                packed_counts = {it['id']: 0 for it in items_data}
                for pi in res.get('packed_items', []):
                    name = pi.get('name')
                    if name in packed_counts: packed_counts[name] += 1
                
                res['new_quantities'] = packed_counts
                res['multiplier'] = multiplier
                res['unfitted_items'] = []
                res['unfitted_count'] = 0
                return jsonify(res)
            
            return jsonify({'status': 'error', 'message': 'Fallo en el cÃ¡lculo de maximizaciÃ³n.'}), 500
        else:
            if job_id:
                _update_logistics_calc_progress(job_id, status='running', stage='solver', progress=12, message='Calculando distribución...')
            res = _do_pack_internal(container_data, items_data, config)
            return jsonify(res) if res.get('status') == 'success' else (jsonify(res), 500)
    except LogisticsCalculationCancelled as e:
        return jsonify({'status': 'cancelled', 'message': str(e)}), 499
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


LOGISTICS_GEOREF_LOOKUP_URL = 'https://apis.datos.gob.ar/georef/api/localidades'
LOGISTICS_ROUTE_DISTANCE_URL = 'https://router.project-osrm.org/route/v1/driving/{lon1},{lat1};{lon2},{lat2}?overview=false'
LOGISTICS_FREIGHT_REFERENCE_DATE = '2025-01-01'
LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG = 30000.0
LOGISTICS_FREIGHT_STANDARD_VOLUME_M3 = 90.0
LOGISTICS_FREIGHT_COST_BASE_DATE = '2024-03'
LOGISTICS_FREIGHT_COST_UPDATE_TARGET = '2025-01'
LOGISTICS_FREIGHT_COST_COMPONENTS = [
    {
        'key': 'mano_obra',
        'label': 'Mano de obra',
        'ars_per_km_mar_2024': 190.70,
        'share_pct_mar_2024': 15.9,
        'annual_adjustment_pct_2024': 187.7,
    },
    {
        'key': 'combustibles',
        'label': 'Combustibles',
        'ars_per_km_mar_2024': 343.20,
        'share_pct_mar_2024': 28.7,
        'annual_adjustment_pct_2024': 67.0,
    },
    {
        'key': 'neumaticos',
        'label': 'Neumáticos',
        'ars_per_km_mar_2024': 140.50,
        'share_pct_mar_2024': 11.7,
        'annual_adjustment_pct_2024': 11.7,
    },
    {
        'key': 'mantenimiento',
        'label': 'Mantenimiento',
        'ars_per_km_mar_2024': 74.10,
        'share_pct_mar_2024': 6.2,
        'annual_adjustment_pct_2024': 32.3,
    },
    {
        'key': 'material_rodante',
        'label': 'Material rodante',
        'ars_per_km_mar_2024': 100.70,
        'share_pct_mar_2024': 8.4,
        'annual_adjustment_pct_2024': 85.6,
    },
    {
        'key': 'patentes_registros',
        'label': 'Patentes y registros',
        'ars_per_km_mar_2024': 17.20,
        'share_pct_mar_2024': 1.4,
        'annual_adjustment_pct_2024': 270.0,
    },
    {
        'key': 'seguros',
        'label': 'Seguros',
        'ars_per_km_mar_2024': 29.90,
        'share_pct_mar_2024': 2.5,
        'annual_adjustment_pct_2024': 239.6,
    },
    {
        'key': 'gastos_generales',
        'label': 'Gastos generales',
        'ars_per_km_mar_2024': 37.20,
        'share_pct_mar_2024': 3.1,
        'annual_adjustment_pct_2024': 193.3,
    },
    {
        'key': 'costos_financieros',
        'label': 'Costos financieros',
        'ars_per_km_mar_2024': 262.90,
        'share_pct_mar_2024': 22.0,
        'annual_adjustment_pct_2024': -13.4,
    },
]
LOGISTICS_FREIGHT_EXCLUDED_COMPONENTS = [
    {
        'label': 'Peajes y otros',
        'reason': 'La lámina base de marzo 2024 indica que los peajes no están incluidos en el costo por km utilizado.',
        'annual_adjustment_pct_2024': 332.5,
    }
]
LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS = {
    'openrouteservice': 'OpenRouteService',
    'graphhopper': 'GraphHopper',
    'mapbox': 'Mapbox Directions',
    'osrm': 'OSRM',
    'haversine': 'Haversine ajustada',
}
LOGISTICS_FREIGHT_MODEL_CACHE = {}


def _logistics_margin_for_fraction(load_fraction):
    fraction = max(0.0, min(float(load_fraction or 0), 1.0))
    if fraction < 0.05:
        return 1.00
    if fraction <= 0.20:
        return 0.60
    if fraction <= 0.60:
        return 0.40
    if fraction <= 0.90:
        return 0.30
    return 0.20


def _logistics_load_basis(actual_weight_kg, consolidated_volume_m3):
    weight_fraction = 0.0
    volume_fraction = 0.0
    if LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG > 0:
        weight_fraction = max(float(actual_weight_kg or 0) / float(LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG), 0.0)
    if LOGISTICS_FREIGHT_STANDARD_VOLUME_M3 > 0:
        volume_fraction = max(float(consolidated_volume_m3 or 0) / float(LOGISTICS_FREIGHT_STANDARD_VOLUME_M3), 0.0)

    dominant_basis = 'weight' if weight_fraction >= volume_fraction else 'volume'
    load_fraction = min(max(weight_fraction, volume_fraction), 1.0)

    return {
        'weight_fraction': round(weight_fraction, 6),
        'volume_fraction': round(volume_fraction, 6),
        'load_fraction': round(load_fraction, 6),
        'dominant_basis': dominant_basis,
    }


def _normalize_logistics_city_payload(city):
    if isinstance(city, str):
        city = {'label': city, 'nombre': city}
    if not isinstance(city, dict):
        return None

    centroide = city.get('centroide', {}) if isinstance(city.get('centroide', {}), dict) else {}
    provincia = city.get('provincia', {}) if isinstance(city.get('provincia', {}), dict) else {}

    try:
        lat = float(centroide.get('lat'))
        lon = float(centroide.get('lon'))
    except Exception:
        return None

    nombre = str(city.get('nombre') or city.get('city') or city.get('label') or '').strip()
    provincia_nombre = str(provincia.get('nombre') or city.get('province') or '').strip()
    label = str(city.get('label') or '').strip() or ', '.join(part for part in [nombre, provincia_nombre] if part)
    if not label:
        return None

    return {
        'id': str(city.get('id') or '').strip(),
        'name': nombre,
        'province': provincia_nombre,
        'label': label,
        'lat': lat,
        'lon': lon
    }


def _normalize_logistics_compare_text(value):
    text = unicodedata.normalize('NFKD', str(value or '')).encode('ascii', 'ignore').decode('ascii')
    text = re.sub(r'\s+', ' ', text).strip().lower()
    return text


def _get_official_sale_dollar_rate(requested_date=''):
    safe_date = str(requested_date or '').strip()
    if safe_date:
        normalized_date = datetime.strptime(safe_date, '%Y-%m-%d').strftime('%Y/%m/%d')
        payload = _fetch_external_json(
            f'https://api.argentinadatos.com/v1/cotizaciones/dolares/oficial/{normalized_date}',
            timeout=6
        )
        venta = payload.get('venta') if isinstance(payload, dict) else None
        fecha = str(payload.get('fecha') or safe_date).strip() if isinstance(payload, dict) else safe_date
        numeric = float(venta)
        if not math.isfinite(numeric) or numeric <= 0:
            raise ValueError('valor oficial venta invalido')
        return {
            'value': round(numeric, 2),
            'date': fecha,
            'source': 'argentinadatos-historico-oficial',
            'rate_type': 'oficial_venta'
        }

    providers = [
        ('dolarapi-oficial', 'https://dolarapi.com/v1/dolares/oficial'),
        ('bluelytics-oficial', 'https://api.bluelytics.com.ar/v2/latest'),
    ]
    errors = []
    for source_name, url in providers:
        try:
            payload = _fetch_external_json(url, timeout=6)
            venta = None
            fecha = ''
            if source_name == 'dolarapi-oficial':
                venta = payload.get('venta')
                fecha = str(payload.get('fecha') or '').strip()
            else:
                oficial = payload.get('oficial', {}) if isinstance(payload, dict) else {}
                venta = oficial.get('value_sell')
                fecha = str(payload.get('last_update') or '').strip()
            numeric = float(venta)
            if not math.isfinite(numeric) or numeric <= 0:
                raise ValueError('valor oficial venta invalido')
            if fecha:
                fecha = fecha.split('T')[0]
            return {
                'value': round(numeric, 2),
                'date': fecha,
                'source': source_name,
                'rate_type': 'oficial_venta'
            }
        except Exception as exc:
            errors.append(f'{source_name}: {exc}')
    raise RuntimeError('; '.join(errors) or 'sin proveedores disponibles')


def _resolve_logistics_city(city):
    normalized = _normalize_logistics_city_payload(city)
    if normalized:
        return normalized

    query = ''
    if isinstance(city, str):
        query = city
    elif isinstance(city, dict):
        query = city.get('label') or city.get('nombre') or city.get('name') or ''

    safe_query = str(query or '').strip()
    if len(safe_query) < 2:
        return None

    query_variants = [safe_query]
    comma_parts = [part.strip() for part in safe_query.split(',') if str(part).strip()]
    if comma_parts:
        query_variants.append(comma_parts[0])
    if len(comma_parts) >= 2:
        query_variants.append(f'{comma_parts[0]} {comma_parts[1]}')

    compact_query = unicodedata.normalize('NFKD', safe_query).encode('ascii', 'ignore').decode('ascii')
    compact_query = re.sub(r'\s+', ' ', compact_query).strip()
    if compact_query and compact_query not in query_variants:
        query_variants.append(compact_query)
        compact_parts = [part.strip() for part in compact_query.split(',') if str(part).strip()]
        if compact_parts:
            query_variants.append(compact_parts[0])

    if _normalize_logistics_compare_text(safe_query) in ('caba', 'capital federal', 'capital'):
        query_variants.append('Ciudad de Buenos Aires')

    seen = set()
    for candidate_query in query_variants:
        normalized_query = str(candidate_query or '').strip()
        if len(normalized_query) < 2:
            continue
        dedupe_key = _normalize_logistics_compare_text(normalized_query)
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        results = _lookup_logistics_cities(normalized_query, max_items=5)
        if not results:
            continue

        query_lower = _normalize_logistics_compare_text(normalized_query)
        exact = next((item for item in results if _normalize_logistics_compare_text(item.get('label', '')) == query_lower), None)
        if exact:
            return exact

        if comma_parts and len(comma_parts) >= 2:
            wanted_name = _normalize_logistics_compare_text(comma_parts[0])
            wanted_province = _normalize_logistics_compare_text(comma_parts[1])
            name_and_province = next((
                item for item in results
                if _normalize_logistics_compare_text(item.get('name', '')) == wanted_name and wanted_province in _normalize_logistics_compare_text(item.get('province', ''))
            ), None)
            if name_and_province:
                return name_and_province

        starts = next((item for item in results if _normalize_logistics_compare_text(item.get('label', '')).startswith(query_lower)), None)
        if starts:
            return starts

        name_exact = next((item for item in results if _normalize_logistics_compare_text(item.get('name', '')) == query_lower), None)
        if name_exact:
            return name_exact

        return results[0]

    return None


def _lookup_logistics_cities(query, max_items=8):
    safe_query = str(query or '').strip()
    if len(safe_query) < 2:
        return []

    params = f'?campos=id,nombre,provincia,centroide&max={max(1, min(int(max_items or 8), 12))}&nombre={quote(safe_query)}'
    payload = _fetch_external_json(f'{LOGISTICS_GEOREF_LOOKUP_URL}{params}', timeout=8)
    localities = payload.get('localidades', []) if isinstance(payload, dict) else []

    results = []
    seen = set()
    for item in localities:
        normalized = _normalize_logistics_city_payload(item)
        if not normalized:
            continue
        dedupe_key = f"{normalized['label'].lower()}|{normalized['lat']}|{normalized['lon']}"
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)
        results.append(normalized)
    return results


def _logistics_haversine_km(lat1, lon1, lat2, lon2):
    radius_km = 6371.0088
    lat1_rad = math.radians(lat1)
    lat2_rad = math.radians(lat2)
    delta_lat = math.radians(lat2 - lat1)
    delta_lon = math.radians(lon2 - lon1)
    a = (math.sin(delta_lat / 2) ** 2) + math.cos(lat1_rad) * math.cos(lat2_rad) * (math.sin(delta_lon / 2) ** 2)
    c = 2 * math.atan2(math.sqrt(a), math.sqrt(1 - a))
    return radius_km * c


def _fetch_external_json_request(url, timeout=4, headers=None, method='GET', data=None):
    req_headers = {'User-Agent': 'BPB-Dashboard/1.0'}
    if isinstance(headers, dict):
        req_headers.update(headers)
    req = urllib.request.Request(url, headers=req_headers, data=data, method=method)
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        charset = resp.headers.get_content_charset() or 'utf-8'
        raw = resp.read().decode(charset, errors='replace')
    return json.loads(raw)


def _build_logistics_freight_cost_model():
    cache_key = LOGISTICS_FREIGHT_REFERENCE_DATE
    cached = LOGISTICS_FREIGHT_MODEL_CACHE.get(cache_key)
    if cached:
        return deepcopy(cached)

    exchange_rate = _get_official_sale_dollar_rate(LOGISTICS_FREIGHT_REFERENCE_DATE)
    ars_per_usd = float(exchange_rate.get('value') or 0)
    if not math.isfinite(ars_per_usd) or ars_per_usd <= 0:
        raise ValueError('dólar oficial de referencia inválido para el modelo de flete')

    components = []
    total_ars_per_km = 0.0
    for component in LOGISTICS_FREIGHT_COST_COMPONENTS:
        base_ars = float(component.get('ars_per_km_mar_2024') or 0)
        annual_adjustment_pct = float(component.get('annual_adjustment_pct_2024') or 0)
        updated_ars = base_ars * (1.0 + (annual_adjustment_pct / 100.0))
        total_ars_per_km += updated_ars
        components.append({
            'key': component.get('key'),
            'label': component.get('label'),
            'ars_per_km_mar_2024': round(base_ars, 4),
            'share_pct_mar_2024': round(float(component.get('share_pct_mar_2024') or 0), 2),
            'annual_adjustment_pct_2024': round(annual_adjustment_pct, 2),
            'ars_per_km_jan_2025': round(updated_ars, 4),
            'usd_per_km_jan_2025': round(updated_ars / ars_per_usd, 6),
        })

    model = {
        'reference_date': LOGISTICS_FREIGHT_REFERENCE_DATE,
        'exchange_rate_ars': round(ars_per_usd, 2),
        'exchange_date': str(exchange_rate.get('date') or LOGISTICS_FREIGHT_REFERENCE_DATE),
        'exchange_source': str(exchange_rate.get('source') or ''),
        'base_cost_date': LOGISTICS_FREIGHT_COST_BASE_DATE,
        'updated_cost_date': LOGISTICS_FREIGHT_COST_UPDATE_TARGET,
        'standard_payload_kg': round(LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG, 3),
        'standard_volume_m3': round(LOGISTICS_FREIGHT_STANDARD_VOLUME_M3, 6),
        'components': components,
        'excluded_components': deepcopy(LOGISTICS_FREIGHT_EXCLUDED_COMPONENTS),
        'total_ars_per_km_jan_2025': round(total_ars_per_km, 4),
        'total_usd_per_km_jan_2025': round(total_ars_per_km / ars_per_usd, 6),
        'methodology_note': (
            'Costo por km base de marzo 2024 actualizado rubro por rubro con la variación acumulada 2024 '
            'y convertido a USD con dólar oficial venta del 2025-01-01.'
        ),
    }
    LOGISTICS_FREIGHT_MODEL_CACHE[cache_key] = deepcopy(model)
    return model


def _logistics_route_distance_openrouteservice(origin, destination):
    api_key = str(os.environ.get('OPENROUTESERVICE_API_KEY') or '').strip()
    if not api_key:
        return None

    payload = {
        'coordinates': [
            [float(origin['lon']), float(origin['lat'])],
            [float(destination['lon']), float(destination['lat'])],
        ]
    }
    data = _fetch_external_json_request(
        'https://api.openrouteservice.org/v2/directions/driving-car/json',
        timeout=10,
        method='POST',
        headers={'Authorization': api_key, 'Content-Type': 'application/json'},
        data=json.dumps(payload).encode('utf-8')
    )
    routes = data.get('routes', []) if isinstance(data, dict) else []
    if not routes:
        return None
    summary = routes[0].get('summary', {}) if isinstance(routes[0], dict) else {}
    distance_m = float(summary.get('distance') or 0)
    if distance_m <= 0:
        return None
    return distance_m / 1000.0, LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS['openrouteservice']


def _logistics_route_distance_graphhopper(origin, destination):
    api_key = str(os.environ.get('GRAPHHOPPER_API_KEY') or '').strip()
    if not api_key:
        return None

    origin_point = f"{float(origin['lat'])},{float(origin['lon'])}"
    destination_point = f"{float(destination['lat'])},{float(destination['lon'])}"
    url = (
        'https://graphhopper.com/api/1/route'
        f'?point={quote(origin_point)}'
        f'&point={quote(destination_point)}'
        '&profile=car&locale=es&calc_points=false'
        f'&key={quote(api_key)}'
    )
    data = _fetch_external_json_request(url, timeout=10)
    paths = data.get('paths', []) if isinstance(data, dict) else []
    if not paths:
        return None
    distance_m = float(paths[0].get('distance') or 0)
    if distance_m <= 0:
        return None
    return distance_m / 1000.0, LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS['graphhopper']


def _logistics_route_distance_mapbox(origin, destination):
    api_key = str(os.environ.get('MAPBOX_API_TOKEN') or '').strip()
    if not api_key:
        return None

    url = (
        'https://api.mapbox.com/directions/v5/mapbox/driving/'
        f'{float(origin["lon"])},{float(origin["lat"])};{float(destination["lon"])},{float(destination["lat"])}'
        f'?overview=false&access_token={quote(api_key)}'
    )
    data = _fetch_external_json_request(url, timeout=10)
    routes = data.get('routes', []) if isinstance(data, dict) else []
    if not routes:
        return None
    distance_m = float(routes[0].get('distance') or 0)
    if distance_m <= 0:
        return None
    return distance_m / 1000.0, LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS['mapbox']


def _logistics_route_distance_km(origin, destination):
    lat1 = float(origin['lat'])
    lon1 = float(origin['lon'])
    lat2 = float(destination['lat'])
    lon2 = float(destination['lon'])

    providers = [
        _logistics_route_distance_openrouteservice,
        _logistics_route_distance_graphhopper,
        _logistics_route_distance_mapbox,
    ]
    for provider in providers:
        try:
            result = provider(origin, destination)
            if result:
                return result
        except Exception:
            continue

    try:
        route_url = LOGISTICS_ROUTE_DISTANCE_URL.format(lon1=lon1, lat1=lat1, lon2=lon2, lat2=lat2)
        payload = _fetch_external_json(route_url, timeout=10)
        routes = payload.get('routes', []) if isinstance(payload, dict) else []
        if routes:
            distance_m = float(routes[0].get('distance') or 0)
            if distance_m > 0:
                return distance_m / 1000.0, LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS['osrm']
    except Exception:
        pass

    direct_km = _logistics_haversine_km(lat1, lon1, lat2, lon2)
    road_estimate_km = max(1.0, direct_km * 1.22)
    return road_estimate_km, LOGISTICS_FREIGHT_DISTANCE_PROVIDER_LABELS['haversine']


@app.route('/api/logistics/cities-lookup', methods=['GET'])
def logistics_cities_lookup():
    try:
        query = str(request.args.get('q', '')).strip()
        if len(query) < 2:
            return jsonify({'status': 'success', 'cities': []})

        return jsonify({
            'status': 'success',
            'cities': _lookup_logistics_cities(query, max_items=request.args.get('max', 8))
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/freight-estimate', methods=['POST'])
def logistics_freight_estimate():
    try:
        data = request.json or {}
        origin = _resolve_logistics_city(data.get('origin'))
        destination = _resolve_logistics_city(data.get('destination'))
        if not origin or not destination:
            return jsonify({'status': 'error', 'message': 'Origen o destino inválido.'}), 400

        freight_basis = data.get('freight_basis', {}) if isinstance(data.get('freight_basis', {}), dict) else {}
        actual_weight_kg = max(0.0, float(
            freight_basis.get('actual_weight_kg', data.get('weight_kg') or 0)
        ))
        consolidated_volume_m3 = max(0.0, float(
            freight_basis.get('consolidated_volume_m3', data.get('volume_m3') or 0)
        ))
        basis_metrics = _logistics_load_basis(actual_weight_kg, consolidated_volume_m3)
        chargeable_weight_kg = actual_weight_kg
        load_fraction = basis_metrics['load_fraction']
        profit_margin = _logistics_margin_for_fraction(load_fraction)

        distance_km, distance_source = _logistics_route_distance_km(origin, destination)
        cost_model = _build_logistics_freight_cost_model()
        cost_per_km_ars = float(cost_model.get('total_ars_per_km_jan_2025') or 0)
        cost_per_km_usd = float(cost_model.get('total_usd_per_km_jan_2025') or 0)
        if not math.isfinite(cost_per_km_ars) or cost_per_km_ars <= 0:
            raise ValueError('costo por km inválido en modelo de flete')
        if not math.isfinite(cost_per_km_usd) or cost_per_km_usd <= 0:
            raise ValueError('costo por km USD inválido en modelo de flete')
        sell_cost_per_km_ars = cost_per_km_ars * (1.0 + profit_margin)
        sell_cost_per_km_usd = cost_per_km_usd * (1.0 + profit_margin)

        estimated_cost_ars = distance_km * sell_cost_per_km_ars * load_fraction
        estimated_cost_usd = distance_km * sell_cost_per_km_usd * load_fraction

        return jsonify({
            'status': 'success',
            'estimate': {
                'origin': origin,
                'destination': destination,
                'distance_km': round(distance_km, 1),
                'distance_source': distance_source,
                'actual_weight_kg': round(actual_weight_kg, 3),
                'consolidated_volume_m3': round(consolidated_volume_m3, 6),
                'chargeable_weight_kg': round(chargeable_weight_kg, 3),
                'load_fraction': round(load_fraction, 6),
                'weight_fraction': basis_metrics['weight_fraction'],
                'volume_fraction': basis_metrics['volume_fraction'],
                'dominant_basis': basis_metrics['dominant_basis'],
                'reference_payload_kg': round(float(LOGISTICS_FREIGHT_STANDARD_PAYLOAD_KG), 3),
                'reference_volume_m3': round(float(LOGISTICS_FREIGHT_STANDARD_VOLUME_M3), 6),
                'reference_cost_per_km_ars': round(cost_per_km_ars, 4),
                'reference_cost_per_km_usd': round(cost_per_km_usd, 6),
                'sell_cost_per_km_ars': round(sell_cost_per_km_ars, 4),
                'sell_cost_per_km_usd': round(sell_cost_per_km_usd, 6),
                'profit_margin_pct': round(profit_margin * 100.0, 2),
                'estimated_cost_ars': round(estimated_cost_ars, 2),
                'estimated_cost_usd': round(estimated_cost_usd, 4),
                'exchange_rate_ars': round(float(cost_model.get('exchange_rate_ars') or 0), 2),
                'exchange_date': str(cost_model.get('exchange_date') or LOGISTICS_FREIGHT_REFERENCE_DATE),
                'source_label': '',
                'note': (
                    f"Base: {'peso' if basis_metrics['dominant_basis'] == 'weight' else 'volumen'} | "
                    f"Fracción: {round(load_fraction * 100.0, 1)}% | "
                    f"Utilidad: {round(profit_margin * 100.0, 1)}%"
                ),
                'freight_basis': {
                    'source': freight_basis.get('source') or 'request_payload',
                    'package_count': int(freight_basis.get('package_count') or 0),
                    'shipment_bbox_dims_mm': freight_basis.get('shipment_bbox_dims_mm') if isinstance(freight_basis.get('shipment_bbox_dims_mm'), dict) else None,
                    'shipment_bbox_volume_m3': round(float(freight_basis.get('shipment_bbox_volume_m3') or 0), 6) if freight_basis else 0.0,
                },
                'cost_model': cost_model,
            }
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/save', methods=['POST'])
def save_logistics_calculation():
    try:
        data = request.json or {}
        if not data:
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        rec_id = str(data.get('id') or uuid.uuid4())
        groups = _read_logistics_record_groups()
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        folder = _normalize_logistics_folder_name(data.get('folder') or (group.get('folder') if group else ''))

        if group:
            new_version_number = int(group.get('latest_version') or len(group.get('versions', [])) or 0) + 1
            snapshot = _build_logistics_version_snapshot(data, version_number=new_version_number)
            group.setdefault('versions', []).append(snapshot)
            group['folder'] = folder
            group['latest_version'] = new_version_number
            group['updated_at'] = snapshot.get('timestamp', _logistics_now_iso())
            group['save_name'] = snapshot.get('save_name', '')
            group['save_description'] = snapshot.get('save_description', '')
            group['author'] = snapshot.get('author', 'Usuario')
        else:
            snapshot = _build_logistics_version_snapshot(data, version_number=1)
            created_at = _logistics_safe_timestamp(data.get('created_at') or snapshot.get('timestamp'))
            group = {
                'id': rec_id,
                'folder': folder,
                'created_at': created_at,
                'updated_at': snapshot.get('timestamp', created_at),
                'latest_version': 1,
                'save_name': snapshot.get('save_name', ''),
                'save_description': snapshot.get('save_description', ''),
                'author': snapshot.get('author', 'Usuario'),
                'versions': [snapshot]
            }
            groups.append(group)

        groups = _merge_logistics_groups(groups)[:1000]
        _write_logistics_record_groups(groups)

        return jsonify({
            'status': 'success',
            'message': 'Calculation saved',
            'id': rec_id,
            'version_id': snapshot.get('version_id', ''),
            'version_number': int(snapshot.get('version_number') or 1),
            'history_label': snapshot.get('history_label', ''),
            'folder': folder
        })
    except Exception as e:
        print(f"ERROR saving logistics: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/logistics/records', methods=['GET'])
def get_logistics_records():
    try:
        store = _read_logistics_records_store()
        payload = [_logistics_group_to_summary(group) for group in store.get('groups', [])]
        return jsonify({
            'status': 'success',
            'folders': store.get('folders', [LOGISTICS_DEFAULT_FOLDER]),
            'records': payload
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/folders', methods=['POST'])
def create_logistics_folder():
    try:
        data = request.json or {}
        folder_name = _normalize_logistics_folder_name(data.get('name'))
        store = _read_logistics_records_store()
        folder_set = {_normalize_logistics_folder_name(name) for name in store.get('folders', [])}
        folder_set.add(LOGISTICS_DEFAULT_FOLDER)
        folder_set.add(folder_name)
        _write_json_file_atomic(LOGISTICS_RECORDS_FILE, {
            'folders': sorted(folder_set, key=lambda name: (name != LOGISTICS_DEFAULT_FOLDER, name.lower())),
            'groups': _merge_logistics_groups(store.get('groups', []))
        }, indent=4, ensure_ascii=False)
        return jsonify({'status': 'success', 'folder': folder_name})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/move-folder', methods=['POST'])
def move_logistics_record_folder():
    try:
        data = request.json or {}
        rec_id = str(data.get('id') or '').strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        target_folder = _normalize_logistics_folder_name(data.get('folder'))
        store = _read_logistics_records_store()
        groups = store.get('groups', [])
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        group['folder'] = target_folder
        folder_set = {_normalize_logistics_folder_name(name) for name in store.get('folders', [])}
        folder_set.add(LOGISTICS_DEFAULT_FOLDER)
        folder_set.add(target_folder)

        _write_json_file_atomic(LOGISTICS_RECORDS_FILE, {
            'folders': sorted(folder_set, key=lambda name: (name != LOGISTICS_DEFAULT_FOLDER, name.lower())),
            'groups': _merge_logistics_groups(groups)
        }, indent=4, ensure_ascii=False)

        return jsonify({'status': 'success', 'folder': target_folder})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/history', methods=['GET'])
def get_logistics_history():
    try:
        rec_id = str(request.args.get('id', '')).strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        groups = _read_logistics_record_groups()
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        versions = sorted(
            group.get('versions', []),
            key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', '')),
            reverse=True
        )
        latest = versions[0] if versions else {}

        history = []
        for version in versions:
            history.append({
                'version_id': version.get('version_id', ''),
                'version_number': int(version.get('version_number') or 1),
                'timestamp': version.get('timestamp', ''),
                'timestamp_display': version.get('timestamp_display', ''),
                'author': version.get('author', 'Usuario'),
                'largest_unit_label': version.get('largest_unit_label', ''),
                'history_label': version.get('history_label', ''),
                'save_name': version.get('save_name', ''),
                'save_description': version.get('save_description', '')
            })

        return jsonify({
            'status': 'success',
            'id': rec_id,
            'save_name': latest.get('save_name', '') or group.get('save_name', ''),
            'created_at': group.get('created_at', ''),
            'modified_at': group.get('updated_at', ''),
            'versions': history
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/version', methods=['GET'])
def get_logistics_version():
    try:
        rec_id = str(request.args.get('id', '')).strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        version_number_raw = str(request.args.get('version', '')).strip()
        version_id = str(request.args.get('version_id', '')).strip()
        latest_flag = str(request.args.get('latest', '')).strip().lower() in ('1', 'true', 'yes', 'si')

        groups = _read_logistics_record_groups()
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        versions = group.get('versions', [])
        selected = None

        if version_id:
            selected = next((v for v in versions if str(v.get('version_id')) == version_id), None)

        if not selected and version_number_raw and not latest_flag:
            try:
                version_number = int(version_number_raw)
            except Exception:
                version_number = None
            if version_number is not None:
                selected = next((v for v in versions if int(v.get('version_number') or 0) == version_number), None)

        if not selected:
            selected = _logistics_get_latest_version(group)

        if not selected:
            return jsonify({'status': 'error', 'message': 'Version not found'}), 404

        return jsonify({
            'status': 'success',
            'record': _logistics_version_to_editor_record(group, selected)
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/logistics/delete', methods=['POST'])
def delete_logistics_record():
    try:
        data = request.json or {}
        rec_id = data.get('id')
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400
        groups = _read_logistics_record_groups()
        new_groups = [g for g in groups if str(g.get('id')) != str(rec_id)]
        if len(new_groups) == len(groups):
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        _write_logistics_record_groups(new_groups)

        return jsonify({'status': 'success', 'message': 'Record deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500



# --- COTIZACION ROUTES ---
def _fetch_external_json(url, timeout=4):
    return _fetch_external_json_request(url, timeout=timeout)


def _to_positive_float(value):
    try:
        if value is None:
            return None
        if isinstance(value, str):
            value = value.strip().replace(',', '.')
        numeric = float(value)
        return numeric if numeric > 0 else None
    except Exception:
        return None


def _extract_dolar_blue_sell_value(payload):
    if isinstance(payload, dict):
        blue_node = payload.get('blue')
        if isinstance(blue_node, dict):
            for key in ('value_sell', 'venta', 'sell'):
                numeric = _to_positive_float(blue_node.get(key))
                if numeric is not None:
                    return numeric

        for key in ('value_sell', 'venta', 'sell'):
            numeric = _to_positive_float(payload.get(key))
            if numeric is not None:
                return numeric

    if isinstance(payload, list):
        for item in payload:
            if not isinstance(item, dict):
                continue
            name = str(item.get('casa') or item.get('nombre') or item.get('name') or '').strip().lower()
            if 'blue' not in name:
                continue
            for key in ('value_sell', 'venta', 'sell'):
                numeric = _to_positive_float(item.get(key))
                if numeric is not None:
                    return numeric

    return None


def _extract_dolar_official_sell_value(payload):
    if isinstance(payload, dict):
        official_node = payload.get('oficial') or payload.get('official')
        if isinstance(official_node, dict):
            for key in ('value_sell', 'venta', 'sell'):
                numeric = _to_positive_float(official_node.get(key))
                if numeric is not None:
                    return numeric

        for key in ('value_sell', 'venta', 'sell'):
            numeric = _to_positive_float(payload.get(key))
            if numeric is not None:
                return numeric

    if isinstance(payload, list):
        for item in payload:
            if not isinstance(item, dict):
                continue
            name = str(item.get('casa') or item.get('nombre') or item.get('name') or '').strip().lower()
            if 'oficial' not in name and 'official' not in name:
                continue
            for key in ('value_sell', 'venta', 'sell'):
                numeric = _to_positive_float(item.get(key))
                if numeric is not None:
                    return numeric

    return None




def _to_non_negative_float(value):
    try:
        if value is None:
            return None
        if isinstance(value, str):
            clean = value.strip()
            if not clean:
                return None
            clean = clean.replace(',', '.')
            value = clean
        numeric = float(value)
        return numeric if numeric >= 0 else None
    except Exception:
        return None


def _normalize_cotizacion_lookup_text(raw_value):
    text = str(raw_value or '').strip()
    if not text:
        return ''
    normalized = unicodedata.normalize('NFKD', text)
    ascii_only = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return ' '.join(ascii_only.lower().split())


def _normalize_cotizacion_complementario_unit(raw_value):
    text = str(raw_value or '').strip()
    if not text:
        return ''

    normalized = unicodedata.normalize('NFKD', text)
    ascii_only = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    compact = ascii_only.upper().replace(' ', '').replace('.', '')

    aliases = {
        'C/UNO': 'Unidades',
        'C/1': 'Unidades',
        'CUNO': 'Unidades',
        'UNI': 'Unidades',
        'UND': 'Unidades',
        'UN': 'Unidades',
        'UNIDAD': 'Unidades',
        'UNIDADES': 'Unidades',
        'MTS': 'Metros',
        'MT': 'Metros',
        'METRO': 'Metros',
        'METROS': 'Metros',
        'KG': 'Kilogramos',
        'KGS': 'Kilogramos',
        'KILOGRAMO': 'Kilogramos',
        'KILOGRAMOS': 'Kilogramos',
        'FT': 'Pies',
        'PIE': 'Pies',
        'PIES': 'Pies',
        'LTS': 'Litros',
        'LT': 'Litros',
        'LITRO': 'Litros',
        'LITROS': 'Litros',
        'TN': 'Toneladas',
        'TON': 'Toneladas',
        'TONELADA': 'Toneladas',
        'TONELADAS': 'Toneladas',
        'MM': 'Milimetros',
        'MILIMETRO': 'Milimetros',
        'MILIMETROS': 'Milimetros',
        'MT3': 'Metros Cubicos',
        'MTS3': 'Metros Cubicos',
        'M3': 'Metros Cubicos',
        'METROCUBICO': 'Metros Cubicos',
        'METROSCUBICOS': 'Metros Cubicos',
        'MTS2': 'Metros Cuadrados',
        'MT2': 'Metros Cuadrados',
        'M2': 'Metros Cuadrados',
        'METROCUADRADO': 'Metros Cuadrados',
        'METROSCUADRADOS': 'Metros Cuadrados',
        'HORA': 'Horas',
        'HORAS': 'Horas',
        'MINUTO': 'Minutos',
        'MINUTOS': 'Minutos',
        'SEGUNDO': 'Segundos',
        'SEGUNDOS': 'Segundos',
        'CENTIMETRO': 'Centimetros',
        'CENTIMETROS': 'Centimetros',
        'CM': 'Centimetros',
        'GRAMO': 'Gramos',
        'GRAMOS': 'Gramos',
        'GR': 'Gramos',
        'G': 'Gramos'
    }

    return aliases.get(compact, text)


def _read_cotizacion_complementarios_catalog(workbook_path):
    if not isinstance(workbook_path, Path):
        workbook_path = Path(workbook_path)

    if not workbook_path.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {workbook_path}')

    wb = _load_workbook_quietly(
        str(workbook_path),
        read_only=True,
        data_only=True,
        keep_vba=workbook_path.suffix.lower() == '.xlsm'
    )

    try:
        if not wb.sheetnames:
            return []

        ws = wb[wb.sheetnames[0]]
        records = []
        dedupe = set()

        for product_raw, description_raw, price_raw, _, unit_raw in ws.iter_rows(
            min_row=2,
            min_col=2,
            max_col=6,
            values_only=True
        ):
            product = str(product_raw or '').strip()
            description = str(description_raw or '').strip()
            unidad = _normalize_cotizacion_complementario_unit(unit_raw)
            if not product and not description:
                continue

            price = _to_non_negative_float(price_raw)
            key = (_normalize_cotizacion_lookup_text(product), _normalize_cotizacion_lookup_text(description))
            if key in dedupe:
                continue
            dedupe.add(key)

            records.append({
                'pieza': product,
                'descripcion': description,
                'precio': round(price, 6) if price is not None else None,
                'unidad': unidad,
                '_pieza_key': key[0],
                '_descripcion_key': key[1]
            })

        records.sort(key=lambda rec: (rec.get('_pieza_key', ''), rec.get('_descripcion_key', '')))
        return records
    finally:
        wb.close()


def _write_cotizacion_complementarios_csv(records, csv_path):
    if not isinstance(csv_path, Path):
        csv_path = Path(csv_path)

    csv_path.parent.mkdir(parents=True, exist_ok=True)

    with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['pieza', 'descripcion', 'precio', 'unidad'], delimiter=';')
        writer.writeheader()
        for item in records:
            price = item.get('precio')
            if price is None:
                price_text = ''
            else:
                try:
                    price_text = f"{float(price):.6f}".rstrip('0').rstrip('.')
                except Exception:
                    price_text = str(price)
            writer.writerow({
                'pieza': str(item.get('pieza') or ''),
                'descripcion': str(item.get('descripcion') or ''),
                'precio': price_text,
                'unidad': str(item.get('unidad') or '')
            })


def _refresh_cotizacion_complementarios_csv_from_excel(excel_path=None, csv_path=None):
    source = Path(excel_path or COTIZACION_COMPLEMENTARIOS_FILE)
    target = Path(csv_path or COTIZACION_COMPLEMENTARIOS_CSV_FILE)

    records = _read_cotizacion_complementarios_catalog(source)
    _write_cotizacion_complementarios_csv(records, target)

    target_stat = target.stat()
    return {
        'excel_path': str(source),
        'csv_path': str(target),
        'rows': len(records),
        'csv_mtime': float(target_stat.st_mtime)
    }


def _read_cotizacion_complementarios_catalog_from_csv(csv_path):
    if not isinstance(csv_path, Path):
        csv_path = Path(csv_path)

    if not csv_path.exists():
        raise FileNotFoundError(f'CSV no encontrado: {csv_path}')

    records = []
    dedupe = set()

    with open(csv_path, 'r', encoding='utf-8-sig', newline='') as f:
        sample = f.read(4096)
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample, delimiters=';,|\t,')
        except Exception:
            dialect = csv.excel
            dialect.delimiter = ';'

        reader = csv.DictReader(f, dialect=dialect)
        for row in reader:
            if not isinstance(row, dict):
                continue

            product = str(
                row.get('pieza')
                or row.get('produc')
                or row.get('producto')
                or row.get('product')
                or ''
            ).strip()
            description = str(
                row.get('descripcion')
                or row.get('descrip_larga')
                or row.get('descripcion_larga')
                or row.get('description')
                or ''
            ).strip()
            price = _to_non_negative_float(
                row.get('precio')
                or row.get('precio_uni')
                or row.get('price')
            )
            unidad = _normalize_cotizacion_complementario_unit(
                row.get('unidad')
                or row.get('unit')
                or row.get('indice')
                or row.get('índice')
                or ''
            )

            if not product and not description:
                continue

            key = (_normalize_cotizacion_lookup_text(product), _normalize_cotizacion_lookup_text(description))
            if key in dedupe:
                continue
            dedupe.add(key)

            records.append({
                'pieza': product,
                'descripcion': description,
                'precio': round(price, 6) if price is not None else None,
                'unidad': unidad,
                '_pieza_key': key[0],
                '_descripcion_key': key[1]
            })

    records.sort(key=lambda rec: (rec.get('_pieza_key', ''), rec.get('_descripcion_key', '')))
    return records


def _load_cotizacion_complementarios_catalog(force_reload=False):
    source = Path(COTIZACION_COMPLEMENTARIOS_CSV_FILE)

    if force_reload or (not source.exists()):
        _refresh_cotizacion_complementarios_csv_from_excel(
            excel_path=COTIZACION_COMPLEMENTARIOS_FILE,
            csv_path=source
        )

    stat = source.stat()
    mtime = float(stat.st_mtime)

    with COTIZACION_COMPLEMENTARIOS_CACHE_LOCK:
        same_source = COTIZACION_COMPLEMENTARIOS_CACHE.get('source') == str(source)
        same_mtime = float(COTIZACION_COMPLEMENTARIOS_CACHE.get('mtime') or 0) == mtime
        cached = COTIZACION_COMPLEMENTARIOS_CACHE.get('records')
        if (not force_reload) and same_source and same_mtime and isinstance(cached, list):
            return cached

        try:
            records = _read_cotizacion_complementarios_catalog_from_csv(source)
        except Exception:
            # Fallback: regenerate CSV from Excel and retry once
            _refresh_cotizacion_complementarios_csv_from_excel(
                excel_path=COTIZACION_COMPLEMENTARIOS_FILE,
                csv_path=source
            )
            records = _read_cotizacion_complementarios_catalog_from_csv(source)
            stat = source.stat()
            mtime = float(stat.st_mtime)

        COTIZACION_COMPLEMENTARIOS_CACHE['source'] = str(source)
        COTIZACION_COMPLEMENTARIOS_CACHE['mtime'] = mtime
        COTIZACION_COMPLEMENTARIOS_CACHE['records'] = records
        return records


def _score_cotizacion_complementarios_match(piece_key, description_key, query_key, field):
    if not query_key:
        return 99

    piece_starts = piece_key.startswith(query_key)
    piece_contains = query_key in piece_key
    description_starts = description_key.startswith(query_key)
    description_contains = query_key in description_key

    if field == 'descripcion':
        if description_starts:
            return 0
        if description_contains:
            return 1
        if piece_starts:
            return 2
        if piece_contains:
            return 3
        return None

    if piece_starts:
        return 0
    if piece_contains:
        return 1
    if description_starts:
        return 2
    if description_contains:
        return 3
    return None


def _read_cotizacion_indices_file(workbook_path):
    if not isinstance(workbook_path, Path):
        workbook_path = Path(workbook_path)

    if not workbook_path.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {workbook_path}')

    wb = _load_workbook_quietly(
        str(workbook_path),
        read_only=True,
        data_only=True,
        keep_vba=workbook_path.suffix.lower() == '.xlsm'
    )

    try:
        if 'INDICE' not in wb.sheetnames:
            raise ValueError(f'Hoja INDICE no encontrada en {workbook_path.name}')

        ws = wb['INDICE']
        records = []
        seen = set()

        for machine_raw, rate_raw in ws.iter_rows(min_row=1, min_col=3, max_col=4, values_only=True):
            machine_name = str(machine_raw or '').strip()
            if not machine_name:
                continue

            lower_name = machine_name.lower()
            if lower_name.startswith('info'):
                continue
            if lower_name in ('maquina / proceso', 'maquina/proceso'):
                continue

            rate_value = _to_non_negative_float(rate_raw)
            if rate_value is None:
                continue

            dedupe_key = lower_name
            if dedupe_key in seen:
                continue
            seen.add(dedupe_key)

            records.append({
                'subconcepto': machine_name,
                'rate': round(rate_value, 6)
            })

        return records
    finally:
        wb.close()


@app.route('/api/cotizacion/indices', methods=['GET'])
def get_cotizacion_indices():
    data = {}
    errors = []

    for category_key, workbook_path in COTIZACION_INDICES_FILES.items():
        try:
            data[category_key] = _read_cotizacion_indices_file(workbook_path)
        except Exception as e:
            data[category_key] = []
            errors.append(f"{category_key}: {str(e)}")

    status = 'success' if not errors else ('partial' if any(data.values()) else 'error')
    response = {
        'status': status,
        'data': data,
        'errors': errors
    }

    if status == 'error':
        return jsonify(response), 500

    return jsonify(response)


@app.route('/api/cotizacion/complementarios-refresh-csv', methods=['GET', 'POST'])
def refresh_cotizacion_complementarios_csv():
    try:
        info = _refresh_cotizacion_complementarios_csv_from_excel(
            excel_path=COTIZACION_COMPLEMENTARIOS_FILE,
            csv_path=COTIZACION_COMPLEMENTARIOS_CSV_FILE
        )

        with COTIZACION_COMPLEMENTARIOS_CACHE_LOCK:
            COTIZACION_COMPLEMENTARIOS_CACHE['source'] = ''
            COTIZACION_COMPLEMENTARIOS_CACHE['mtime'] = 0.0
            COTIZACION_COMPLEMENTARIOS_CACHE['records'] = []

        return jsonify({
            'status': 'success',
            'message': 'CSV de complementarios actualizado.',
            'excel_path': info.get('excel_path'),
            'csv_path': info.get('csv_path'),
            'rows': info.get('rows', 0),
            'csv_mtime': info.get('csv_mtime')
        })
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'No se pudo actualizar el CSV de complementarios: {str(e)}'
        }), 500


@app.route('/api/cotizacion/complementarios-lookup', methods=['GET'])
def get_cotizacion_complementarios_lookup():
    query_raw = str(request.args.get('q', '') or '').strip()
    query_key = _normalize_cotizacion_lookup_text(query_raw)
    field = str(request.args.get('field', 'pieza') or 'pieza').strip().lower()
    if field not in ('pieza', 'descripcion'):
        field = 'pieza'

    try:
        limit = int(str(request.args.get('limit', '10') or '10').strip())
    except Exception:
        limit = 10
    limit = max(1, min(50, limit))

    try:
        catalog = _load_cotizacion_complementarios_catalog(force_reload=False)
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': f'No se pudo cargar la base de complementarios: {str(e)}',
            'data': []
        }), 500

    matches = []
    for item in catalog:
        piece_key = item.get('_pieza_key', '')
        description_key = item.get('_descripcion_key', '')
        score = _score_cotizacion_complementarios_match(piece_key, description_key, query_key, field)
        if score is None:
            continue
        matches.append((score, item))

    matches.sort(key=lambda row: (row[0], row[1].get('_pieza_key', ''), row[1].get('_descripcion_key', '')))
    data = []
    for _, item in matches[:limit]:
        data.append({
            'pieza': item.get('pieza', ''),
            'descripcion': item.get('descripcion', ''),
            'precio': item.get('precio', None),
            'unidad': item.get('unidad', '')
        })

    return jsonify({
        'status': 'success',
        'query': query_raw,
        'field': field,
        'count': len(data),
        'data': data
    })


@app.route('/api/cotizacion/dolar-rate', methods=['GET'])
def get_cotizacion_dolar_rate():
    requested_date = str(request.args.get('date', '') or '').strip()
    try:
        if requested_date:
            datetime.strptime(requested_date, '%Y-%m-%d')
        rate_data = _get_official_sale_dollar_rate(requested_date)
        return jsonify({'status': 'success', **rate_data})
    except ValueError:
        return jsonify({
            'status': 'error',
            'message': 'Formato de fecha invalido. Use YYYY-MM-DD.'
        }), 400
    except Exception as e:
        return jsonify({
            'status': 'error',
            'message': 'No se pudo obtener el valor del dolar Oficial venta para la fecha seleccionada.' if requested_date else 'No se pudo obtener el valor del dolar Oficial venta desde internet',
            'errors': [str(e)]
        }), 502


@app.route('/api/cotizacion/providers-lookup', methods=['GET'])
def get_cotizacion_provider_lookup():
    query_raw = str(request.args.get('q', '') or '').strip()
    query_key = query_raw.lower()

    try:
        limit = int(str(request.args.get('limit', '12') or '12').strip())
    except Exception:
        limit = 12
    limit = max(1, min(50, limit))

    providers = _collect_cotizacion_provider_suggestions()
    if not query_key:
        data = providers[:limit]
    else:
        starts = [item for item in providers if item.lower().startswith(query_key)]
        contains = [item for item in providers if query_key in item.lower() and item not in starts]
        data = (starts + contains)[:limit]

    return jsonify({
        'status': 'success',
        'query': query_raw,
        'count': len(data),
        'data': data
    })



def _clean_cotizacion_items(raw_items):
    clean = []
    if not isinstance(raw_items, list):
        return clean

    for item in raw_items:
        if not isinstance(item, dict):
            continue
        clean.append({
            'category': item.get('category', ''),
            'categoria': _cotizacion_fix_text(item.get('categoria', '')),
            'subCategoria': _cotizacion_fix_text(item.get('subCategoria', '')),
            'subconcepto': _cotizacion_fix_text(item.get('subconcepto', '')),
            'pieza': _cotizacion_fix_text(item.get('pieza', item.get('piece', ''))),
            'descripcion': _cotizacion_fix_text(item.get('descripcion', '')),
            'proveedor': _cotizacion_fix_text(item.get('proveedor', item.get('provider', ''))),
            'provider_mix': item.get('provider_mix', []),
            'source_summary': _cotizacion_normalize_source_summary(item.get('source_summary')),
            'source_record_id': str(item.get('source_record_id', '')),
            'source_version_id': str(item.get('source_version_id', '')),
            'is_muda': bool(item.get('is_muda')),
            'rate_formula': str(item.get('rate_formula', '')),
            'cantidad_formula': str(item.get('cantidad_formula', '')),
            'costo_formula': str(item.get('costo_formula', '')),
            'vida_util': item.get('vida_util', 0),
            'rate': item.get('rate', 0),
            'cantidad': item.get('cantidad', 0),
            'indice': _cotizacion_fix_text(item.get('indice', '')),
            'observaciones': _cotizacion_fix_text(item.get('observaciones', '')),
            'costo_unitario': item.get('costo_unitario', 0)
        })
    return clean


def _collect_cotizacion_provider_suggestions():
    groups = _read_cotizacion_record_groups()
    seen = set()
    providers = []

    for group in groups:
        versions = group.get('versions', []) if isinstance(group, dict) else []
        for version in versions:
            items = version.get('items', []) if isinstance(version, dict) else []
            for item in items:
                if not isinstance(item, dict):
                    continue
                provider = str(item.get('proveedor') or item.get('provider') or '').strip()
                if not provider:
                    continue
                key = provider.lower()
                if key in seen:
                    continue
                seen.add(key)
                providers.append(provider)

    providers.sort(key=lambda value: value.lower())
    return providers


def _clean_cotizacion_attachments(raw_attachments):
    clean = []
    if not isinstance(raw_attachments, list):
        return clean

    max_files = 40
    max_data_url_len = 16 * 1024 * 1024  # chars

    for attachment in raw_attachments[:max_files]:
        if not isinstance(attachment, dict):
            continue

        name = str(attachment.get('name', '')).strip()
        file_type = str(attachment.get('type', '')).strip().lower()
        size = attachment.get('size', 0)
        data_url = str(attachment.get('data_url', attachment.get('dataURL', ''))).strip()

        if not name or not data_url:
            continue
        if len(data_url) > max_data_url_len:
            continue
        if not (data_url.startswith('data:image/') or data_url.startswith('data:application/pdf')):
            continue

        # Keep only the MIME types requested by the user (image/PDF).
        if data_url.startswith('data:application/pdf'):
            file_type = 'application/pdf'
        elif not file_type.startswith('image/'):
            file_type = 'image/*'

        try:
            size_value = int(size)
        except Exception:
            size_value = 0

        clean.append({
            'id': str(attachment.get('id') or uuid.uuid4()),
            'name': name,
            'type': file_type,
            'size': max(size_value, 0),
            'data_url': data_url
        })

    return clean


def _cotizacion_now_iso():
    return _utc_now_iso_z()


def _cotizacion_safe_timestamp(raw_value):
    value = str(raw_value or '').strip()
    if value:
        return value
    return _cotizacion_now_iso()


def _cotizacion_fix_text(value):
    text = str(value or '')
    if not text:
        return ''
    if any(token in text for token in ('Ã', 'Â', 'â', '�')):
        try:
            repaired = text.encode('latin-1', errors='ignore').decode('utf-8', errors='ignore')
            if repaired:
                return repaired
        except Exception:
            pass
    return text


def _cotizacion_normalize_delivery_unit(value):
    unit = _cotizacion_fix_text(value).strip().lower()
    aliases = {
        'hora': 'Horas',
        'horas': 'Horas',
        'dia': 'Dias',
        'dias': 'Dias',
        'días': 'Dias',
        'semana': 'Semanas',
        'semanas': 'Semanas',
        'mes': 'Meses',
        'meses': 'Meses'
    }
    return aliases.get(unit, 'Dias')


def _cotizacion_normalize_source_summary(raw_summary):
    if not isinstance(raw_summary, dict):
        return None

    keys = ['materia_prima', 'complementarios', 'transformacion', 'desperdicios', 'externo', 'importacion', 'extra', 'comadm', 'produccion', 'venta', 'total']
    normalized = {}
    has_value = False

    for key in keys:
        try:
            value = float(raw_summary.get(key, 0) or 0)
        except Exception:
            value = 0.0
        if not math.isfinite(value):
            value = 0.0
        normalized[key] = value
        if abs(value) > 0:
            has_value = True

    return normalized if has_value else None


def _cotizacion_apply_source_summary_to_bucket(bucket, source_summary, multiplier=1.0):
    if not isinstance(bucket, dict):
        return False

    summary = _cotizacion_normalize_source_summary(source_summary)
    if not summary:
        return False

    try:
        factor = float(multiplier or 1)
    except Exception:
        factor = 1.0
    if not math.isfinite(factor) or factor <= 0:
        factor = 1.0

    for key in ('materia_prima', 'complementarios', 'transformacion', 'desperdicios', 'externo', 'importacion', 'extra', 'comadm'):
        bucket[key] = float(bucket.get(key, 0) or 0) + (float(summary.get(key, 0) or 0) * factor)
    return True


def _cotizacion_safe_header(raw_header):
    header = raw_header if isinstance(raw_header, dict) else {}
    code_value = _cotizacion_fix_text(header.get('piece_code', header.get('code', '')))
    requester_value = _cotizacion_fix_text(header.get('requested_by', header.get('requester', '')))
    return {
        'code': code_value,
        'piece_code': code_value,
        'piece': _cotizacion_fix_text(header.get('piece', '')),
        'analysis_date': header.get('analysis_date', ''),
        'usd_value': header.get('usd_value', 0),
        'responsible_name': _cotizacion_fix_text(header.get('responsible_name', '')),
        'responsible_position': _cotizacion_fix_text(header.get('responsible_position', '')),
        'responsible_department': _cotizacion_fix_text(header.get('responsible_department', '')),
        'responsible_owner': _cotizacion_fix_text(header.get('responsible_owner', '')),
        'requester': requester_value,
        'requested_by': requester_value,
        'delivery_time': _cotizacion_fix_text(header.get('delivery_time', '')),
        'delivery_unit': _cotizacion_normalize_delivery_unit(header.get('delivery_unit', 'Dias'))
    }


def _cotizacion_safe_settings(raw_settings):
    settings = raw_settings if isinstance(raw_settings, dict) else {}
    safe_settings = dict(settings)
    safe_settings['piece_quantity'] = safe_settings.get('piece_quantity', 1)
    safe_settings['mayorista_descuento'] = safe_settings.get('mayorista_descuento', 0)
    return safe_settings


def _cotizacion_is_truthy(value):
    if isinstance(value, bool):
        return value
    return str(value or '').strip().lower() in ('1', 'true', 'yes', 'si', 'on')


def _cotizacion_normalize_revision_label(raw_label):
    cleaned = re.sub(r'[^A-Za-z]+', '', str(raw_label or '').upper())
    return cleaned


def _cotizacion_revision_index_to_label(index):
    try:
        value = int(index or 0)
    except Exception:
        value = 0
    if value < 1:
        return ''

    chars = []
    while value > 0:
        value -= 1
        chars.append(chr(ord('A') + (value % 26)))
        value //= 26
    return ''.join(reversed(chars))


def _cotizacion_apply_revision_metadata(versions):
    safe_versions = versions if isinstance(versions, list) else []
    revision_count = 0
    latest_revision_label = '0'

    for version in safe_versions:
        if not isinstance(version, dict):
            continue

        is_revision = _cotizacion_is_truthy(version.get('is_revision'))
        version['is_revision'] = is_revision
        if not is_revision:
            version['revision_label'] = ''
            continue

        revision_count += 1
        revision_label = (
            _cotizacion_normalize_revision_label(version.get('revision_label'))
            or _cotizacion_revision_index_to_label(revision_count)
        )
        version['revision_label'] = revision_label
        latest_revision_label = revision_label or latest_revision_label

    return {
        'revision_count': revision_count,
        'latest_revision_label': latest_revision_label if revision_count > 0 else '0'
    }


def _cotizacion_get_group_revision_info(group):
    if not isinstance(group, dict):
        return {'revision_count': 0, 'latest_revision_label': '0'}

    versions = group.get('versions', [])
    if isinstance(versions, list) and versions:
        sorted_versions = sorted(
            versions,
            key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', ''))
        )
        return _cotizacion_apply_revision_metadata(sorted_versions)

    revision_count = int(group.get('revision_count') or 0)
    latest_revision_label = _cotizacion_normalize_revision_label(group.get('latest_revision_label'))
    return {
        'revision_count': revision_count,
        'latest_revision_label': latest_revision_label if revision_count > 0 and latest_revision_label else '0'
    }


def _build_cotizacion_version_snapshot(source, version_number=1):
    data = source if isinstance(source, dict) else {}
    is_revision = _cotizacion_is_truthy(data.get('is_revision'))
    return {
        'version_id': str(data.get('version_id') or uuid.uuid4()),
        'version_number': int(data.get('version_number') or version_number or 1),
        'timestamp': _cotizacion_safe_timestamp(data.get('timestamp')),
        'is_revision': is_revision,
        'revision_label': _cotizacion_normalize_revision_label(data.get('revision_label')) if is_revision else '',
        'save_name': _cotizacion_fix_text(data.get('save_name', '')),
        'save_description': _cotizacion_fix_text(data.get('save_description', '')),
        'save_category': _normalize_cotizacion_record_category(data.get('save_category', '')),
        'author': _cotizacion_fix_text(data.get('author', 'Usuario')),
        'header': _cotizacion_safe_header(data.get('header', {})),
        'approved_by': _cotizacion_fix_text(data.get('approved_by', '')),
        'notes': _cotizacion_fix_text(data.get('notes', '')),
        'settings': _cotizacion_safe_settings(data.get('settings', {})),
        'summary': data.get('summary', {}),
        'attachments': _clean_cotizacion_attachments(data.get('attachments', [])),
        'items': _clean_cotizacion_items(data.get('items', []))
    }


def _normalize_cotizacion_folder_name(raw_folder):
    folder = str(raw_folder or '').strip()
    return folder or COTIZACION_DEFAULT_FOLDER


def _normalize_cotizacion_record_category(raw_category):
    value = str(raw_category or '').strip().lower()
    if value == 'conjunto':
        return 'Conjunto'
    if value in ('sub-conjunto', 'sub conjunto', 'subconjunto'):
        return 'Sub-Conjunto'
    if value == 'pieza':
        return 'Pieza'
    return ''


def _cotizacion_build_save_name(code_value, piece_value):
    code = _cotizacion_fix_text(code_value).strip()
    piece = _cotizacion_fix_text(piece_value).strip()
    if code and piece:
        return f"{code} - {piece}"
    return piece or code


def _normalize_cotizacion_group(raw_group):
    if not isinstance(raw_group, dict):
        return None

    group_id = str(raw_group.get('id') or uuid.uuid4())
    raw_versions = raw_group.get('versions')

    if isinstance(raw_versions, list):
        versions = []
        fallback_counter = 1
        for raw_version in raw_versions:
            if not isinstance(raw_version, dict):
                continue
            snapshot = _build_cotizacion_version_snapshot(
                raw_version,
                version_number=raw_version.get('version_number') or fallback_counter
            )
            versions.append(snapshot)
            fallback_counter += 1

        if not versions:
            versions = [_build_cotizacion_version_snapshot(raw_group, version_number=1)]
    else:
        # Legacy flat record (single snapshot)
        versions = [_build_cotizacion_version_snapshot(raw_group, version_number=1)]

    versions.sort(key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', '')))
    revision_info = _cotizacion_apply_revision_metadata(versions)
    latest = versions[-1]
    latest_version = int(latest.get('version_number') or len(versions) or 1)

    created_at = _cotizacion_safe_timestamp(raw_group.get('created_at') or versions[0].get('timestamp'))
    updated_at = _cotizacion_safe_timestamp(raw_group.get('updated_at') or latest.get('timestamp'))

    return {
        'id': group_id,
        'folder': _normalize_cotizacion_folder_name(raw_group.get('folder')),
        'created_at': created_at,
        'updated_at': updated_at,
        'latest_version': latest_version,
        'latest_revision_label': revision_info.get('latest_revision_label', '0'),
        'revision_count': int(revision_info.get('revision_count') or 0),
        'save_name': latest.get('save_name', ''),
        'save_description': latest.get('save_description', ''),
        'save_category': latest.get('save_category', ''),
        'record_category': latest.get('save_category', ''),
        'author': latest.get('author', 'Usuario'),
        'versions': versions
    }


def _merge_cotizacion_groups(groups):
    merged = {}
    for group in groups:
        if not isinstance(group, dict):
            continue
        gid = str(group.get('id') or '')
        if not gid:
            continue

        normalized = _normalize_cotizacion_group(group)
        if not normalized:
            continue

        existing = merged.get(gid)
        if not existing:
            merged[gid] = normalized
            continue

        combined_versions = []
        seen = set()
        for source in [existing, normalized]:
            for version in source.get('versions', []):
                v_id = str(version.get('version_id') or '')
                v_key = v_id or f"{version.get('version_number')}|{version.get('timestamp')}"
                if v_key in seen:
                    continue
                seen.add(v_key)
                combined_versions.append(version)

        existing['versions'] = combined_versions
        existing['folder'] = _normalize_cotizacion_folder_name(normalized.get('folder') or existing.get('folder'))
        merged[gid] = _normalize_cotizacion_group(existing)

    merged_list = [g for g in merged.values() if isinstance(g, dict)]
    merged_list.sort(key=lambda g: g.get('updated_at', ''), reverse=True)
    return merged_list


def _read_cotizacion_records_store():
    empty_store = {'folders': [COTIZACION_DEFAULT_FOLDER], 'groups': []}
    if not os.path.exists(COTIZACION_RECORDS_FILE):
        return empty_store

    try:
        with open(COTIZACION_RECORDS_FILE, 'r', encoding='utf-8') as f:
            raw_records = json.load(f)
    except Exception:
        return empty_store

    if isinstance(raw_records, list):
        raw_groups = raw_records
        raw_folders = []
    elif isinstance(raw_records, dict):
        raw_groups = raw_records.get('groups', [])
        raw_folders = raw_records.get('folders', [])
    else:
        return empty_store

    normalized = []
    for raw in raw_groups if isinstance(raw_groups, list) else []:
        group = _normalize_cotizacion_group(raw)
        if group:
            normalized.append(group)

    groups = _merge_cotizacion_groups(normalized)
    folder_set = {COTIZACION_DEFAULT_FOLDER}
    if isinstance(raw_folders, list):
        for folder in raw_folders:
            folder_set.add(_normalize_cotizacion_folder_name(folder))
    for group in groups:
        folder_set.add(_normalize_cotizacion_folder_name(group.get('folder')))

    folders = sorted(folder_set, key=lambda name: (name != COTIZACION_DEFAULT_FOLDER, name.lower()))
    return {'folders': folders, 'groups': groups}


def _read_cotizacion_record_groups():
    return _read_cotizacion_records_store().get('groups', [])


def _read_cotizacion_folder_names():
    return _read_cotizacion_records_store().get('folders', [COTIZACION_DEFAULT_FOLDER])


def _write_cotizacion_record_groups(groups):
    safe_groups = _merge_cotizacion_groups(groups if isinstance(groups, list) else [])
    folder_set = {COTIZACION_DEFAULT_FOLDER}
    existing_store = _read_cotizacion_records_store()
    for folder in existing_store.get('folders', []):
        folder_set.add(_normalize_cotizacion_folder_name(folder))
    for group in safe_groups:
        folder_set.add(_normalize_cotizacion_folder_name(group.get('folder')))
    safe_payload = {
        'folders': sorted(folder_set, key=lambda name: (name != COTIZACION_DEFAULT_FOLDER, name.lower())),
        'groups': safe_groups
    }
    _write_json_file_atomic(COTIZACION_RECORDS_FILE, safe_payload, indent=4, ensure_ascii=False)


def _cotizacion_get_latest_version(group):
    versions = group.get('versions', []) if isinstance(group, dict) else []
    if not versions:
        return None
    return sorted(
        versions,
        key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', ''))
    )[-1]


def _cotizacion_group_to_summary(group):
    latest = _cotizacion_get_latest_version(group) or {}
    versions = group.get('versions', []) if isinstance(group, dict) else []
    created_at = group.get('created_at') or latest.get('timestamp') or _cotizacion_now_iso()
    modified_at = group.get('updated_at') or latest.get('timestamp') or created_at
    return {
        'id': group.get('id', ''),
        'folder': _normalize_cotizacion_folder_name(group.get('folder')),
        'save_name': latest.get('save_name', '') or group.get('save_name', ''),
        'save_description': latest.get('save_description', '') or group.get('save_description', ''),
        'save_category': latest.get('save_category', '') or group.get('save_category', ''),
        'record_category': latest.get('save_category', '') or group.get('record_category', '') or group.get('save_category', ''),
        'author': latest.get('author', 'Usuario') or group.get('author', 'Usuario'),
        'header': latest.get('header', {}) if isinstance(latest.get('header', {}), dict) else {},
        'created_at': created_at,
        'modified_at': modified_at,
        'latest_version': int(group.get('latest_version') or latest.get('version_number') or len(versions) or 1),
        'version_count': len(versions),
        'latest_revision_label': group.get('latest_revision_label', '0') or '0',
        'revision_count': int(group.get('revision_count') or 0)
    }


def _cotizacion_version_to_editor_record(group, version):
    return {
        'id': group.get('id', ''),
        'folder': _normalize_cotizacion_folder_name(group.get('folder')),
        'version_id': version.get('version_id', ''),
        'version_number': int(version.get('version_number') or 1),
        'is_revision': _cotizacion_is_truthy(version.get('is_revision')),
        'revision_label': version.get('revision_label', ''),
        'latest_revision_label': group.get('latest_revision_label', '0') or '0',
        'created_at': group.get('created_at', ''),
        'modified_at': group.get('updated_at', ''),
        'save_name': version.get('save_name', ''),
        'save_description': version.get('save_description', ''),
        'save_category': version.get('save_category', ''),
        'record_category': version.get('save_category', ''),
        'timestamp': version.get('timestamp', ''),
        'author': version.get('author', 'Usuario'),
        'header': version.get('header', {}),
        'approved_by': version.get('approved_by', ''),
        'notes': version.get('notes', ''),
        'settings': version.get('settings', {}),
        'summary': version.get('summary', {}),
        'attachments': version.get('attachments', []),
        'items': version.get('items', [])
    }


def _cotizacion_normalize_category(raw_category):
    value = str(raw_category or '').strip().lower()
    normalized = unicodedata.normalize('NFD', value)
    normalized = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    normalized = normalized.replace('&', 'y')
    normalized = re.sub(r'[^a-z0-9]+', '_', normalized).strip('_')
    aliases = {
        'materia_prima': 'materia_prima',
        'conjunto': 'conjunto',
        'complementarios': 'complementarios',
        'complementario': 'complementarios',
        'produccion': 'produccion',
        'ensamble': 'ensamble',
        'embalaje': 'embalaje',
        'deposito': 'deposito_logistica',
        'deposito_y_logistica': 'deposito_logistica',
        'deposito_logistica': 'deposito_logistica',
        'externos': 'externos',
        'externos_a_bpb': 'externos',
        'externo_a_bpb': 'externos',
        'gastos_importacion': 'importacion',
        'importacion': 'importacion',
        'extras': 'extras',
        'extra': 'extras',
        'costo_bpb': 'costo_byp',
        'costo_byp': 'costo_byp',
        'com_y_adm': 'costo_byp',
        'com_adm': 'costo_byp',
    }
    return aliases.get(normalized, normalized)


def _cotizacion_empty_summary_bucket():
    return {
        'materia_prima': 0.0,
        'complementarios': 0.0,
        'transformacion': 0.0,
        'desperdicios': 0.0,
        'produccion': 0.0,
        'externo': 0.0,
        'importacion': 0.0,
        'extra': 0.0,
        'venta': 0.0,
        'comadm': 0.0,
        'total': 0.0
    }


def _cotizacion_add_cost_to_bucket(bucket, category_key, cost):
    try:
        numeric_cost = float(cost or 0)
    except Exception:
        numeric_cost = 0.0
    if not math.isfinite(numeric_cost):
        return

    category = _cotizacion_normalize_category(category_key)
    if category in ('materia_prima', 'conjunto'):
        bucket['materia_prima'] += numeric_cost
    elif category == 'complementarios':
        bucket['complementarios'] += numeric_cost
    elif category in ('produccion', 'ensamble', 'embalaje', 'deposito_logistica'):
        bucket['transformacion'] += numeric_cost
    elif category == 'externos':
        bucket['externo'] += numeric_cost
    elif category == 'importacion':
        bucket['importacion'] += numeric_cost
    elif category == 'extras':
        bucket['extra'] += numeric_cost
    elif category == 'costo_byp':
        bucket['comadm'] += numeric_cost


def _cotizacion_finalize_summary_bucket(raw_bucket, piece_qty=1):
    raw = _cotizacion_empty_summary_bucket()
    if isinstance(raw_bucket, dict):
        for key in raw.keys():
            try:
                raw[key] = float(raw_bucket.get(key, 0) or 0)
            except Exception:
                raw[key] = 0.0

    raw['produccion'] = raw['materia_prima'] + raw['complementarios'] + raw['transformacion']
    raw['venta'] = raw['produccion'] + raw['desperdicios'] + raw['externo'] + raw['importacion'] + raw['extra']
    raw['total'] = raw['venta'] + raw['comadm']

    try:
        qty = int(piece_qty or 1)
    except Exception:
        qty = 1
    qty = max(1, qty)

    piezas = {
        key: float(value) * qty
        for key, value in raw.items()
    }
    return {
        'unitario': raw,
        'piezas': piezas
    }


def _cotizacion_normalize_provider_mix(raw_mix):
    clean = []
    total_ratio = 0.0
    if not isinstance(raw_mix, list):
        return clean

    for item in raw_mix:
        if not isinstance(item, dict):
            continue
        provider = _cotizacion_fix_text(item.get('provider', '')).strip() or 'Sin proveedor'
        try:
            ratio = float(item.get('ratio', 0) or 0)
        except Exception:
            ratio = 0.0
        if not math.isfinite(ratio) or ratio <= 0:
            continue
        total_ratio += ratio
        clean.append({
            'provider': provider,
            'ratio': ratio
        })

    if total_ratio > 0:
        for item in clean:
            item['ratio'] = item['ratio'] / total_ratio
    return clean


def _cotizacion_build_provider_breakdown_from_items(items):
    totals = {}
    grand_total = 0.0

    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        try:
            cost = float(item.get('costo_unitario', 0) or 0)
        except Exception:
            cost = 0.0
        if not math.isfinite(cost) or cost <= 0:
            continue

        provider_mix = _cotizacion_normalize_provider_mix(item.get('provider_mix', []))
        if provider_mix:
            for provider_item in provider_mix:
                provider_cost = cost * provider_item['ratio']
                provider_name = provider_item['provider']
                totals[provider_name] = float(totals.get(provider_name, 0)) + provider_cost
                grand_total += provider_cost
            continue

        provider_name = _cotizacion_fix_text(item.get('proveedor', '')).strip() or 'Sin proveedor'
        totals[provider_name] = float(totals.get(provider_name, 0)) + cost
        grand_total += cost

    breakdown = []
    for provider_name, value in totals.items():
        breakdown.append({
            'provider': provider_name,
            'value': value,
            'ratio': (value / grand_total) if grand_total > 0 else 0
        })

    breakdown.sort(key=lambda item: float(item.get('value', 0) or 0), reverse=True)
    return breakdown


def _cotizacion_get_provider_breakdown_from_version(version):
    summary = version.get('summary', {}) if isinstance(version, dict) else {}
    raw_breakdown = summary.get('provider_breakdown', []) if isinstance(summary, dict) else []
    valid_breakdown = []
    if isinstance(raw_breakdown, list):
        value_total = 0.0
        ratio_total = 0.0
        for item in raw_breakdown:
            if not isinstance(item, dict):
                continue
            provider = _cotizacion_fix_text(item.get('provider', '')).strip() or 'Sin proveedor'
            try:
                value = float(item.get('value', 0) or 0)
            except Exception:
                value = 0.0
            try:
                ratio = float(item.get('ratio', 0) or 0)
            except Exception:
                ratio = 0.0
            if not math.isfinite(value):
                value = 0.0
            if not math.isfinite(ratio):
                ratio = 0.0
            if value <= 0 and ratio <= 0:
                continue
            value_total += max(value, 0.0)
            ratio_total += max(ratio, 0.0)
            valid_breakdown.append({
                'provider': provider,
                'value': max(value, 0.0),
                'ratio': max(ratio, 0.0)
            })

        if valid_breakdown:
            if value_total > 0:
                for item in valid_breakdown:
                    item['ratio'] = item['value'] / value_total
            elif ratio_total > 0:
                for item in valid_breakdown:
                    item['ratio'] = item['ratio'] / ratio_total
                    item['value'] = 0.0
            valid_breakdown.sort(key=lambda item: float(item.get('value', 0) or 0), reverse=True)
            return valid_breakdown

    return _cotizacion_build_provider_breakdown_from_items(version.get('items', []))


def _cotizacion_build_conjunto_item_from_group(group, base_item=None):
    latest = _cotizacion_get_latest_version(group) or {}
    summary = latest.get('summary', {}) if isinstance(latest, dict) else {}
    unitario = summary.get('unitario', {}) if isinstance(summary, dict) else {}
    try:
        sale_cost = float(unitario.get('venta', 0) or 0)
    except Exception:
        sale_cost = 0.0
    if not math.isfinite(sale_cost):
        sale_cost = 0.0

    provider_breakdown = _cotizacion_get_provider_breakdown_from_version(latest)
    provider_mix = [{
        'provider': item.get('provider', 'Sin proveedor'),
        'ratio': float(item.get('ratio', 0) or 0)
    } for item in provider_breakdown if float(item.get('ratio', 0) or 0) > 0]
    provider_label = ''
    if len(provider_mix) == 1:
        provider_label = provider_mix[0]['provider']
    elif len(provider_mix) > 1:
        provider_label = 'Mixto'

    source_code = _cotizacion_fix_text(latest.get('header', {}).get('code', '')).strip()
    source_piece = _cotizacion_fix_text(latest.get('header', {}).get('piece', '')).strip()
    source_name = _cotizacion_fix_text(latest.get('save_name', '')).strip()
    piece_label = source_code or 'Conjunto'
    description = source_piece or source_name

    base = base_item if isinstance(base_item, dict) else {}
    try:
        quantity = float(base.get('cantidad', 1) or 1)
    except Exception:
        quantity = 1.0
    if not math.isfinite(quantity) or quantity <= 0:
        quantity = 1.0

    indice = _cotizacion_fix_text(base.get('indice', '')).strip() or 'Unidad'
    observaciones = _cotizacion_fix_text(base.get('observaciones', '')).strip()
    rate_formula = str(base.get('rate_formula', '') or '').strip()
    cantidad_formula = str(base.get('cantidad_formula', '') or '').strip()
    costo_formula = str(base.get('costo_formula', '') or '').strip()

    return {
        'category': 'conjunto',
        'categoria': 'Conjunto',
        'subCategoria': '-',
        'subconcepto': '-',
        'pieza': piece_label,
        'descripcion': description,
        'proveedor': provider_label,
        'provider_mix': provider_mix,
        'source_summary': _cotizacion_normalize_source_summary(unitario),
        'source_record_id': str(group.get('id') or ''),
        'source_version_id': str(latest.get('version_id') or ''),
        'is_muda': False,
        'rate_formula': rate_formula,
        'cantidad_formula': cantidad_formula,
        'costo_formula': costo_formula,
        'vida_util': 0,
        'rate': sale_cost,
        'cantidad': quantity,
        'indice': indice,
        'observaciones': observaciones,
        'costo_unitario': sale_cost * quantity
    }


def _cotizacion_build_combined_header_piece(items, fallback=''):
    source_names = []
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        label = _cotizacion_fix_text(item.get('pieza', '')).strip()
        if label:
            source_names.append(label)
    if source_names:
        return f"Conjunto - {' + '.join(source_names)}"
    return _cotizacion_fix_text(fallback or 'Conjunto').strip() or 'Conjunto'


def _cotizacion_resolve_source_summary_from_records(source_record_id, records):
    source_id = str(source_record_id or '').strip()
    if not source_id:
        return None
    for group in records if isinstance(records, list) else []:
        if not isinstance(group, dict):
            continue
        if str(group.get('id') or '').strip() != source_id:
            continue
        latest = _cotizacion_get_latest_version(group) or {}
        summary = latest.get('summary', {}) if isinstance(latest, dict) else {}
        unitario = summary.get('unitario', {}) if isinstance(summary, dict) else {}
        return _cotizacion_normalize_source_summary(unitario)
    return None


def _cotizacion_fill_missing_source_summaries(items, records):
    if not isinstance(items, list):
        return items
    for item in items:
        if not isinstance(item, dict):
            continue
        category = _cotizacion_normalize_category(item.get('category') or item.get('categoria'))
        if category != 'conjunto':
            continue
        if bool(item.get('is_muda')):
            continue
        existing = _cotizacion_normalize_source_summary(item.get('source_summary'))
        if existing:
            item['source_summary'] = existing
            continue
        resolved = _cotizacion_resolve_source_summary_from_records(item.get('source_record_id'), records)
        if resolved:
            item['source_summary'] = resolved
    return items


def _cotizacion_build_combined_summary_from_items(items, piece_qty=1):
    raw_bucket = _cotizacion_empty_summary_bucket()
    piece_list = []
    piece_summaries = {}
    desperdicios_list = []
    desperdicios_summaries = {}
    seen_piece_keys = set()
    seen_desperdicio_keys = set()

    piece_counter = 0
    desperdicio_counter = 0
    for item in items if isinstance(items, list) else []:
        if not isinstance(item, dict):
            continue
        category = _cotizacion_normalize_category(item.get('category') or item.get('categoria'))
        is_muda = bool(item.get('is_muda'))
        try:
            cost = float(item.get('costo_unitario', 0) or 0)
        except Exception:
            cost = 0.0
        if not math.isfinite(cost):
            cost = 0.0
        source_summary = item.get('source_summary')
        quantity_value = item.get('cantidad', 1)
        try:
            quantity = float(quantity_value or 1)
        except Exception:
            quantity = 1.0
        if not math.isfinite(quantity) or quantity <= 0:
            quantity = 1.0

        used_source_summary = False
        if is_muda:
            summary = _cotizacion_normalize_source_summary(source_summary)
            if summary:
                try:
                    summary_waste = float(summary.get('desperdicios', 0) or 0)
                except Exception:
                    summary_waste = 0.0
                if math.isfinite(summary_waste) and summary_waste != 0:
                    raw_bucket['desperdicios'] = float(raw_bucket.get('desperdicios', 0) or 0) + (summary_waste * quantity)
                    used_source_summary = True
            if not used_source_summary:
                raw_bucket['desperdicios'] = float(raw_bucket.get('desperdicios', 0) or 0) + cost

            desperdicio_label = _cotizacion_fix_text(item.get('pieza', '')).strip() or f"Desperdicio {desperdicio_counter + 1}"
            desperdicio_key = f"waste:{desperdicio_label.lower()}"
            if desperdicio_key not in seen_desperdicio_keys:
                seen_desperdicio_keys.add(desperdicio_key)
                desperdicios_list.append({'key': desperdicio_key, 'label': desperdicio_label})
                desperdicios_summaries[desperdicio_key] = {'label': desperdicio_label, 'raw': _cotizacion_empty_summary_bucket()}
            desperdicio_value = cost
            if summary:
                try:
                    summary_waste = float(summary.get('desperdicios', 0) or 0)
                except Exception:
                    summary_waste = 0.0
                if math.isfinite(summary_waste) and summary_waste != 0:
                    desperdicio_value = summary_waste * quantity
            desperdicios_summaries[desperdicio_key]['raw']['desperdicios'] = float(desperdicios_summaries[desperdicio_key]['raw'].get('desperdicios', 0) or 0) + desperdicio_value
            desperdicio_counter += 1
        else:
            if category == 'conjunto':
                used_source_summary = _cotizacion_apply_source_summary_to_bucket(raw_bucket, source_summary, quantity)
            if not used_source_summary:
                _cotizacion_add_cost_to_bucket(raw_bucket, category, cost)

        if is_muda:
            continue

        if category not in ('materia_prima', 'conjunto'):
            continue

        label = _cotizacion_fix_text(item.get('pieza', '')).strip()
        if not label:
            piece_counter += 1
            label = f"Conjunto {piece_counter}"
        piece_key = f"piece:{label.lower()}"
        if piece_key not in seen_piece_keys:
            seen_piece_keys.add(piece_key)
            piece_list.append({'key': piece_key, 'label': label})
            piece_summaries[piece_key] = {'label': label, 'raw': _cotizacion_empty_summary_bucket()}
        if category == 'conjunto' and _cotizacion_apply_source_summary_to_bucket(piece_summaries[piece_key]['raw'], source_summary, quantity):
            continue
        _cotizacion_add_cost_to_bucket(piece_summaries[piece_key]['raw'], category, cost)

    finalized_global = _cotizacion_finalize_summary_bucket(raw_bucket, piece_qty)
    finalized_piece_summaries = {}
    for piece_key, piece_data in piece_summaries.items():
        finalized_piece = _cotizacion_finalize_summary_bucket(piece_data.get('raw', {}), piece_qty)
        finalized_piece_summaries[piece_key] = {
            'label': piece_data.get('label', ''),
            'unitario': finalized_piece['unitario'],
            'piezas': finalized_piece['piezas']
        }
    finalized_desperdicios_summaries = {}
    for waste_key, waste_data in desperdicios_summaries.items():
        finalized_waste = _cotizacion_finalize_summary_bucket(waste_data.get('raw', {}), piece_qty)
        finalized_desperdicios_summaries[waste_key] = {
            'label': waste_data.get('label', ''),
            'unitario': finalized_waste['unitario'],
            'piezas': finalized_waste['piezas']
        }

    return {
        'pieceQty': max(1, int(piece_qty or 1)),
        'unitario': finalized_global['unitario'],
        'piezas': finalized_global['piezas'],
        'pieceList': piece_list,
        'pieceSummaries': finalized_piece_summaries,
        'complementariosList': [],
        'complementariosSummaries': {},
        'desperdiciosList': desperdicios_list,
        'desperdiciosSummaries': finalized_desperdicios_summaries,
        'provider_breakdown': _cotizacion_build_provider_breakdown_from_items(items)
    }


def _cotizacion_item_links_to_source(item, source_group):
    if not isinstance(item, dict) or not isinstance(source_group, dict):
        return False
    if _cotizacion_normalize_category(item.get('category') or item.get('categoria')) != 'conjunto':
        return False

    source_group_id = str(source_group.get('id') or '').strip()
    item_source_id = str(item.get('source_record_id') or '').strip()
    if source_group_id and item_source_id:
        return item_source_id == source_group_id

    latest = _cotizacion_get_latest_version(source_group) or {}
    candidate_labels = {
        _cotizacion_fix_text(latest.get('save_name', '')).strip().lower(),
        _cotizacion_fix_text(latest.get('header', {}).get('piece', '')).strip().lower()
    }
    candidate_labels = {label for label in candidate_labels if label}
    if not candidate_labels:
        return False

    item_label = _cotizacion_fix_text(item.get('pieza', '')).strip().lower()
    return item_label in candidate_labels


def _cotizacion_sync_linked_combined_groups(groups, source_group_id):
    safe_groups = _merge_cotizacion_groups(groups if isinstance(groups, list) else [])
    groups_by_id = {
        str(group.get('id') or ''): group
        for group in safe_groups
        if isinstance(group, dict) and str(group.get('id') or '').strip()
    }

    queue = [str(source_group_id or '').strip()]
    processed = set()

    while queue:
        current_source_id = queue.pop(0)
        if not current_source_id or current_source_id in processed:
            continue
        processed.add(current_source_id)

        source_group = groups_by_id.get(current_source_id)
        if not source_group:
            continue

        downstream_changes = []
        for target_group in safe_groups:
            if not isinstance(target_group, dict):
                continue
            target_id = str(target_group.get('id') or '').strip()
            if not target_id or target_id == current_source_id:
                continue

            latest = _cotizacion_get_latest_version(target_group)
            if not isinstance(latest, dict):
                continue

            original_items = latest.get('items', [])
            if not isinstance(original_items, list) or not original_items:
                continue

            touched = False
            updated_items = []
            for item in original_items:
                if not isinstance(item, dict):
                    updated_items.append(item)
                    continue

                if not _cotizacion_item_links_to_source(item, source_group):
                    updated_items.append(deepcopy(item))
                    continue

                rebuilt_item = _cotizacion_build_conjunto_item_from_group(source_group, item)
                updated_items.append(rebuilt_item)
                if json.dumps(item, ensure_ascii=False, sort_keys=True) != json.dumps(rebuilt_item, ensure_ascii=False, sort_keys=True):
                    touched = True

            if not touched:
                continue

            settings = _cotizacion_safe_settings(latest.get('settings', {}))
            piece_qty = settings.get('piece_quantity', 1)
            updated_summary = _cotizacion_build_combined_summary_from_items(updated_items, piece_qty)
            updated_header = _cotizacion_safe_header(latest.get('header', {}))

            new_version_number = int(target_group.get('latest_version') or len(target_group.get('versions', [])) or 0) + 1
            new_snapshot = _build_cotizacion_version_snapshot({
                'timestamp': _cotizacion_now_iso(),
                'save_name': latest.get('save_name', '') or target_group.get('save_name', ''),
                'save_description': latest.get('save_description', '') or target_group.get('save_description', ''),
                'save_category': latest.get('save_category', '') or target_group.get('save_category', ''),
                'author': latest.get('author', 'Sistema') or 'Sistema',
                'header': updated_header,
                'approved_by': latest.get('approved_by', ''),
                'notes': latest.get('notes', ''),
                'settings': settings,
                'summary': updated_summary,
                'attachments': latest.get('attachments', []),
                'items': updated_items
            }, version_number=new_version_number)

            target_group.setdefault('versions', []).append(new_snapshot)
            target_group['latest_version'] = new_version_number
            target_group['updated_at'] = new_snapshot.get('timestamp', _cotizacion_now_iso())
            target_group['save_name'] = new_snapshot.get('save_name', '')
            target_group['save_description'] = new_snapshot.get('save_description', '')
            target_group['save_category'] = new_snapshot.get('save_category', '')
            target_group['author'] = new_snapshot.get('author', 'Sistema')
            downstream_changes.append(target_id)
            groups_by_id[target_id] = target_group

        queue.extend(downstream_changes)

    return _merge_cotizacion_groups(safe_groups)


@app.route('/api/cotizacion/save', methods=['POST'])
def save_cotizacion_record():
    try:
        data = request.json or {}
        if not data:
            return jsonify({'status': 'error', 'message': 'No data received'}), 400

        header_payload = data.get('header', {}) if isinstance(data.get('header', {}), dict) else {}
        normalized_code = _cotizacion_fix_text(header_payload.get('piece_code', header_payload.get('code', ''))).strip()
        normalized_piece = _cotizacion_fix_text(header_payload.get('piece', '')).strip()
        normalized_category = _normalize_cotizacion_record_category(data.get('save_category', data.get('record_category', '')))
        normalized_requester = _cotizacion_fix_text(header_payload.get('requested_by', header_payload.get('requester', ''))).strip()

        header_payload['code'] = normalized_code
        header_payload['piece_code'] = normalized_code
        header_payload['piece'] = normalized_piece
        header_payload['requester'] = normalized_requester
        header_payload['requested_by'] = normalized_requester
        data['header'] = header_payload
        data['save_category'] = normalized_category
        data['record_category'] = normalized_category
        data['save_name'] = _cotizacion_build_save_name(normalized_code, normalized_piece)

        rec_id = str(data.get('id') or uuid.uuid4())
        records = _read_cotizacion_record_groups()
        data['items'] = _cotizacion_fill_missing_source_summaries(data.get('items', []), records)
        generate_revision = _cotizacion_is_truthy(data.get('generate_revision'))

        group = None
        for item in records:
            if str(item.get('id')) == rec_id:
                group = item
                break

        if group:
            current_revision_info = _cotizacion_get_group_revision_info(group)
            next_revision_label = _cotizacion_revision_index_to_label(int(current_revision_info.get('revision_count') or 0) + 1)
        else:
            current_revision_info = {'revision_count': 0, 'latest_revision_label': '0'}
            next_revision_label = _cotizacion_revision_index_to_label(1)

        data['is_revision'] = generate_revision
        data['revision_label'] = next_revision_label if generate_revision else ''

        if group:
            new_version_number = int(group.get('latest_version') or len(group.get('versions', [])) or 0) + 1
            snapshot = _build_cotizacion_version_snapshot(data, version_number=new_version_number)
            group.setdefault('versions', []).append(snapshot)
            group['folder'] = _normalize_cotizacion_folder_name(data.get('folder') or group.get('folder'))
            group['latest_version'] = new_version_number
            group['updated_at'] = snapshot.get('timestamp', _cotizacion_now_iso())
            group['save_name'] = snapshot.get('save_name', '')
            group['save_description'] = snapshot.get('save_description', '')
            group['save_category'] = snapshot.get('save_category', '')
            group['record_category'] = snapshot.get('save_category', '')
            group['author'] = snapshot.get('author', 'Usuario')
            revision_info = _cotizacion_get_group_revision_info(group)
            group['latest_revision_label'] = revision_info.get('latest_revision_label', '0')
            group['revision_count'] = int(revision_info.get('revision_count') or 0)
        else:
            snapshot = _build_cotizacion_version_snapshot(data, version_number=1)
            created_at = _cotizacion_safe_timestamp(data.get('created_at') or snapshot.get('timestamp'))
            initial_revision_label = snapshot.get('revision_label', '') if _cotizacion_is_truthy(snapshot.get('is_revision')) else ''
            group = {
                'id': rec_id,
                'folder': _normalize_cotizacion_folder_name(data.get('folder')),
                'created_at': created_at,
                'updated_at': snapshot.get('timestamp', created_at),
                'latest_version': 1,
                'latest_revision_label': initial_revision_label or '0',
                'revision_count': 1 if initial_revision_label else 0,
                'save_name': snapshot.get('save_name', ''),
                'save_description': snapshot.get('save_description', ''),
                'save_category': snapshot.get('save_category', ''),
                'record_category': snapshot.get('save_category', ''),
                'author': snapshot.get('author', 'Usuario'),
                'versions': [snapshot]
            }
            records.append(group)

        if str(data.get('update_linked', '')).lower() in ('1', 'true', 'yes', 'si') or bool(data.get('update_linked')):
            records = _cotizacion_sync_linked_combined_groups(records, rec_id)

        records = _merge_cotizacion_groups(records)[:1000]
        _write_cotizacion_record_groups(records)

        return jsonify({
            'status': 'success',
            'message': 'Cotizacion saved',
            'id': rec_id,
            'version_id': snapshot.get('version_id', ''),
            'version_number': int(snapshot.get('version_number') or 1),
            'is_revision': _cotizacion_is_truthy(snapshot.get('is_revision')),
            'revision_label': snapshot.get('revision_label', ''),
            'latest_revision_label': group.get('latest_revision_label', '0') or '0'
        })
    except Exception as e:
        print(f"ERROR saving cotizacion: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/records', methods=['GET'])
def get_cotizacion_records():
    try:
        store = _read_cotizacion_records_store()
        groups = store.get('groups', [])
        payload = [_cotizacion_group_to_summary(group) for group in groups]
        return jsonify({
            'status': 'success',
            'folders': store.get('folders', [COTIZACION_DEFAULT_FOLDER]),
            'records': payload
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/folders', methods=['POST'])
def create_cotizacion_folder():
    try:
        data = request.json or {}
        folder_name = _normalize_cotizacion_folder_name(data.get('name'))
        store = _read_cotizacion_records_store()
        folder_set = {_normalize_cotizacion_folder_name(name) for name in store.get('folders', [])}
        folder_set.add(COTIZACION_DEFAULT_FOLDER)
        folder_set.add(folder_name)
        _write_json_file_atomic(COTIZACION_RECORDS_FILE, {
            'folders': sorted(folder_set, key=lambda name: (name != COTIZACION_DEFAULT_FOLDER, name.lower())),
            'groups': _merge_cotizacion_groups(store.get('groups', []))
        }, indent=4, ensure_ascii=False)
        return jsonify({'status': 'success', 'folder': folder_name})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/move-folder', methods=['POST'])
def move_cotizacion_record_folder():
    try:
        data = request.json or {}
        rec_id = str(data.get('id') or '').strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        target_folder = _normalize_cotizacion_folder_name(data.get('folder'))
        store = _read_cotizacion_records_store()
        groups = store.get('groups', [])
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        group['folder'] = target_folder
        folder_set = {_normalize_cotizacion_folder_name(name) for name in store.get('folders', [])}
        folder_set.add(COTIZACION_DEFAULT_FOLDER)
        folder_set.add(target_folder)

        _write_json_file_atomic(COTIZACION_RECORDS_FILE, {
            'folders': sorted(folder_set, key=lambda name: (name != COTIZACION_DEFAULT_FOLDER, name.lower())),
            'groups': _merge_cotizacion_groups(groups)
        }, indent=4, ensure_ascii=False)

        return jsonify({'status': 'success', 'folder': target_folder})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/history', methods=['GET'])
def get_cotizacion_history():
    try:
        rec_id = str(request.args.get('id', '')).strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        groups = _read_cotizacion_record_groups()
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        versions = sorted(
            group.get('versions', []),
            key=lambda v: (int(v.get('version_number') or 0), v.get('timestamp', '')),
            reverse=True
        )
        latest = versions[0] if versions else {}

        history = []
        for version in versions:
            history.append({
                'version_id': version.get('version_id', ''),
                'version_number': int(version.get('version_number') or 1),
                'is_revision': _cotizacion_is_truthy(version.get('is_revision')),
                'revision_label': version.get('revision_label', ''),
                'timestamp': version.get('timestamp', ''),
                'author': version.get('author', 'Usuario'),
                'save_name': version.get('save_name', ''),
                'save_description': version.get('save_description', '')
            })

        return jsonify({
            'status': 'success',
            'id': rec_id,
            'save_name': latest.get('save_name', '') or group.get('save_name', ''),
            'created_at': group.get('created_at', ''),
            'modified_at': group.get('updated_at', ''),
            'latest_revision_label': group.get('latest_revision_label', '0') or '0',
            'versions': history
        })
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/version', methods=['GET'])
def get_cotizacion_version():
    try:
        rec_id = str(request.args.get('id', '')).strip()
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        version_number_raw = str(request.args.get('version', '')).strip()
        version_id = str(request.args.get('version_id', '')).strip()
        latest_flag = str(request.args.get('latest', '')).strip().lower() in ('1', 'true', 'yes', 'si')

        groups = _read_cotizacion_record_groups()
        group = next((item for item in groups if str(item.get('id')) == rec_id), None)
        if not group:
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        versions = group.get('versions', [])
        selected = None

        if version_id:
            selected = next((v for v in versions if str(v.get('version_id')) == version_id), None)

        if not selected and version_number_raw and not latest_flag:
            try:
                version_number = int(version_number_raw)
            except Exception:
                version_number = None
            if version_number is not None:
                candidates = [v for v in versions if int(v.get('version_number') or 0) == version_number]
                if candidates:
                    candidates.sort(key=lambda v: v.get('timestamp', ''), reverse=True)
                    selected = candidates[0]

        if not selected:
            selected = _cotizacion_get_latest_version(group)

        if not selected:
            return jsonify({'status': 'error', 'message': 'No versions available'}), 404

        record = _cotizacion_version_to_editor_record(group, selected)
        return jsonify({'status': 'success', 'record': record})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/cotizacion/delete', methods=['POST'])
def delete_cotizacion_record():
    try:
        data = request.json or {}
        rec_id = data.get('id')
        if not rec_id:
            return jsonify({'status': 'error', 'message': 'Missing id'}), 400

        records = _read_cotizacion_record_groups()
        if not records:
            return jsonify({'status': 'error', 'message': 'No records file'}), 404

        new_records = [r for r in records if r.get('id') != rec_id]
        if len(new_records) == len(records):
            return jsonify({'status': 'error', 'message': 'Record not found'}), 404

        _write_cotizacion_record_groups(new_records)

        return jsonify({'status': 'success', 'message': 'Record deleted'})
    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)}), 500

@app.route('/api/check-ingresos', methods=['POST'])
def check_ingresos():
    """Check for new ingresos data by running procesar_ingresos.py"""
    import sys
    try:
        script_dir = BASE_DIR / "Codigos"
        script_name = "procesar_ingresos.py"
        
        if not script_dir.exists():
             return jsonify({"status": "error", "message": "Script directory not found"}), 404

        # Run checking script using the SAME python interpreter as this app
        result = subprocess.run([sys.executable, script_name, '--refresh-excel'], cwd=str(script_dir), capture_output=True, text=True)
        
        if result.returncode != 0:
            print(f"Check Script Error: {result.stderr}")
            return jsonify({"status": "error", "message": f"Error en script: {result.stderr}"}), 500
            
        # Analyze output
        if "NO se crea" in result.stdout:
            return jsonify({"status": "no_changes", "message": "Sin cambios", "debug_output": result.stdout})
        else:
            print(f"DEBUG: Check determined CHANGES DETECTED. Stdout chars: {len(result.stdout)}")
            return jsonify({"status": "changes_detected", "message": "Nuevos registros detectados", "debug_output": result.stdout})

    except Exception as e:
        print(f"Check Process Error: {e}")
        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/sync-progress', methods=['GET'])

def get_sync_progress():

    return jsonify(SYNC_STATE)



@app.route('/api/run-automation', methods=['POST'])

def run_automation():

    try:

        # script_dir = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Codigos")
        script_dir = BASE_DIR / "Codigos"

        script_name = "automation.py"

        

        # Args as requested: --config ".\config.yaml" --once

        cmd = ['python', script_name, '--config', r'.\config.yaml', '--once']



        if not (script_dir / script_name).exists():

            return jsonify({"status": "error", "message": "Automation script not found"}), 404



        # Run non-blocking or blocking? 

        # User implies it's triggered after upload. 

        # Usually automation takes time. Blocking might timeout.

        # But for simplicity let's do blocking first as per previous pattern.

        result = subprocess.run(cmd, cwd=str(script_dir), capture_output=True, text=True)

        

        if result.returncode == 0:

             return jsonify({"status": "success", "message": "AutomaciÃ³n ejecutada", "output": result.stdout})

        else:

             print(f"Auto Error: {result.stderr}")

             return jsonify({"status": "error", "message": f"Error automaciÃ³n: {result.stderr}"}), 500



    except Exception as e:

        print(f"Auto Run Error: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/data')

def get_data():

    results = []

    if not DATA_DIR.exists():

        return jsonify({"error": "Data directory not found", "path": str(DATA_DIR)}), 404



    # Pre-fetch entry dates

    entry_dates_cache = get_entry_dates_cache()

        

    for folder in DATA_DIR.iterdir():

        if folder.is_dir():

            try:

                stats = folder.stat()

                # On Windows st_ctime is creation time, st_mtime is modification. 

                # "Fecha Ingreso" usually implies creation/arrival. Let's use mtime as it reflects the folder's last touch.

                sys_date = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y')

            except:

                sys_date = '-'

            

            # --- OVERRIDE DATE WITH REGISTRO INFO IF AVAILABLE ---

            # Extract number from folder name (e.g. PO1291 -> 1291)

            po_digits = "".join(filter(str.isdigit, folder.name))

            if po_digits and po_digits in entry_dates_cache:

                sys_date = entry_dates_cache[po_digits]

            # -----------------------------------------------------



            po_data = {

                "id": folder.name,

                "system_date": sys_date,

                "path": str(folder),

                "files": {"pdfs": [], "csvs": [], "json": None, "approved_items": []},

                "content": {"json": {}, "csvs": []}

            }



            # Read Approval Info

            meta_path = folder / "approval_info.json"

            if meta_path.exists():

                try:

                    with open(meta_path, 'r', encoding='utf-8') as f:

                        app_meta = json.load(f)

                        po_data["files"]["approved_items"] = app_meta.get("approved_items", [])

                except: pass

            

            # Scan files

            for file in folder.iterdir():

                if file.is_file():

                    if file.suffix.lower() == '.pdf': po_data["files"]["pdfs"].append(file.name)

                    elif file.suffix.lower() == '.json':

                        po_data["files"]["json"] = file.name

                        try:

                            with open(file, 'r', encoding='utf-8') as f: po_data["content"]["json"] = json.load(f)

                        except: pass

                    elif file.suffix.lower() == '.csv':

                        po_data["files"]["csvs"].append(file.name)

                        try:

                            with open(file, 'r', encoding='utf-8', errors='replace') as f:

                                reader = csv.DictReader(f, delimiter=';')

                                headers = [str(h) for h in (reader.fieldnames or [])]

                                rows = [{str(k): v for k, v in row.items()} for row in reader]

                                po_data["content"]["csvs"].append({"filename": file.name, "headers": headers, "rows": rows})

                        except: pass

            

            # Scan csv subdir

            csv_subdir = folder / 'csv'

            if csv_subdir.exists():

                for file in csv_subdir.iterdir():

                    if file.is_file() and file.suffix.lower() == '.csv':

                        po_data["files"]["csvs"].append(file.name)

                        try:

                            with open(file, 'r', encoding='utf-8', errors='replace') as f:

                                reader = csv.DictReader(f, delimiter=';')

                                headers = [str(h) for h in (reader.fieldnames or [])]

                                rows = [{str(k): v for k, v in row.items()} for row in reader]

                                po_data["content"]["csvs"].append({"filename": file.name, "headers": headers, "rows": rows})

                        except: pass

            results.append(po_data)

    return jsonify(results)



@app.route('/api/profile/history')

def get_my_history():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "Login required"}), 401

    return get_user_history(session.get('user'))



@app.route('/api/admin/history/<username>', methods=['GET'])

def get_user_history(username):

    current_user = session.get('user')

    current_role = session.get('role')

    

    # Permitir si es admin O si el usuario pide su propio historial

    if current_role != 'admin' and current_user != username:

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    try:

        if not LOG_FILE.exists():

            return jsonify([])

            

        with open(LOG_FILE, 'r') as f:

            logs = json.load(f)

            

        user_logs = [l for l in logs if l['user'] == username]

        return jsonify(user_logs[::-1])

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/update', methods=['POST'])

def update_data():

    data = request.json

    po_id = data.get('po_id')

    field = data.get('field')

    value = data.get('value')

    

    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        return jsonify({"status": "error", "message": "Folder not found"}), 404

        

    json_files = list(target_folder.glob("*.json"))

    if not json_files:

         return jsonify({"status": "error", "message": "JSON file not found"}), 404

    

    target_file = json_files[0]

    

    try:

        with open(target_file, 'r', encoding='utf-8') as f:

            content = json.load(f)

        

        content[field] = value

        

        _write_json_file_atomic(target_file, content, indent=4, ensure_ascii=False)

            

        actor = session.get('user', 'system')

        log_action(actor, "Updated JSON Field", f"PO: {po_id}, Field: {field}, New Value: {value}")

            

        return jsonify({"status": "success", "new_value": value})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/update_csv', methods=['POST'])

def update_csv():

    data = request.json

    po_id = data.get('po_id')

    filename = data.get('filename')

    rows = data.get('rows')



    if not po_id or not filename:

        return jsonify({"status": "error", "message": "Missing PO ID or Filename"}), 400



    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        alt_folder = PROCESSED_PATH / po_id

        if alt_folder.exists():

            target_folder = alt_folder

        else:

            return jsonify({"status": "error", "message": "PO Folder not found"}), 404



    target_file = target_folder / filename

    if not target_file.exists():

        target_file = target_folder / 'csv' / filename

        

    if not target_file.exists():

         return jsonify({"status": "error", "message": "File not found"}), 404



    try:

        if not rows:

             with open(target_file, 'w', encoding='utf-8', newline='') as f:

                 pass

             actor = session.get('user', 'system')

             log_action(actor, "Cleared CSV", f"PO: {po_id}, File: {filename}")

             return jsonify({"status": "success", "message": "File cleared"})



        headers = list(rows[0].keys())



        with open(target_file, 'w', encoding='utf-8', newline='', errors='replace') as f:

            writer = csv.DictWriter(f, fieldnames=headers, delimiter=';')

            writer.writeheader()

            writer.writerows(rows)

            

        actor = session.get('user', 'system')

        log_action(actor, "Updated CSV Data", f"PO: {po_id}, File: {filename}, Rows: {len(rows)}")



        return jsonify({"status": "success"})



    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/upload', methods=['POST'])

def upload_file():

    if 'file' not in request.files:

        return jsonify({'status': 'error', 'message': 'No file part'}), 400

        

    file = request.files['file']

    

    if file.filename == '':

        return jsonify({'status': 'error', 'message': 'No selected file'}), 400

        

    if file and allowed_file(file.filename):

        try:

            original_filename = secure_filename(file.filename)

            timestamp = int(time.time())

            filename, ext = os.path.splitext(original_filename)

            new_filename = f"{filename}_{timestamp}{ext}"

            

            save_path = UPLOAD_FOLDER / new_filename

            

            if not UPLOAD_FOLDER.exists():

                UPLOAD_FOLDER.mkdir(parents=True, exist_ok=True)

                

            file.save(str(save_path))

            

            actor = session.get('user', 'system')

            log_action(actor, "Uploaded File", f"Filename: {new_filename}")

            

            return jsonify({'status': 'success', 'filename': new_filename})

        except Exception as e:

            return jsonify({'status': 'error', 'message': str(e)}), 500

            

    return jsonify({'status': 'error', 'message': 'Invalid file type'}), 400



@app.route('/api/login', methods=['POST'])

def login():

    data = request.json

    username = data.get('username')

    password = data.get('password')

    

    if not username or not password:

        return jsonify({"status": "error", "message": "Faltan credenciales"}), 400

        

    try:

        if not USERS_FILE.exists():

             return jsonify({"status": "error", "message": "Error del sistema: Falta archivo de usuarios"}), 500

             

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        normalized_username = str(username).strip().lower()
        user_data = users.get(normalized_username)

        if user_data and _check_password_hash_compatible(user_data['password'], password):

            if user_data.get('status') != 'approved':

                return jsonify({"status": "error", "message": "Debe esperar aprobaciÃ³n."}), 403

                

            session['user'] = normalized_username

            session['role'] = user_data.get('role', 'user')

            return jsonify({"status": "success", "role": session['role']})

        else:

            return jsonify({"status": "error", "message": "Credenciales invÃ¡lidas"}), 401

            

    except Exception as e:

        print(f"Auth Error: {e}")

        return jsonify({"status": "error", "message": "Error de autenticaciÃ³n"}), 500



@app.route('/api/logout', methods=['POST'])

def logout():

    session.pop('user', None)

    session.pop('role', None)

    return jsonify({"status": "success"})



@app.route('/api/register', methods=['POST'])

def register():

    data = request.json

    username = data.get('username')

    password = data.get('password')

    

    if not username or not password:

        return jsonify({"status": "error", "message": "Faltan credenciales"}), 400

        

    try:

        if not USERS_FILE.exists():

             users = {}

        else:

            with open(USERS_FILE, 'r') as f:

                users = json.load(f)

        

        normalized_username = str(username).strip().lower()

        if normalized_username in users:

            return jsonify({"status": "error", "message": "El usuario ya existe"}), 400

            

        users[normalized_username] = {

            "password": _generate_password_hash_compatible(password),

            "display_name": username,

            "position": "",

            "department": "",

            "signature": "",

            "role": "user",

            "status": "pending",

            "created_at": time.strftime('%Y-%m-%d %H:%M:%S')

        }

        

        _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

            

        return jsonify({"status": "success", "message": "Usuario creado. Debe esperar aprobacion del area para ingresar. Vuelva a intentarlo mas tarde."})

            

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/users', methods=['GET'])

def get_users():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    try:

        if not USERS_FILE.exists():

            return jsonify([])

            

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        user_list = []

        for uname, udata in users.items():

            user_list.append({

                "username": uname,

                "role": udata.get('role', 'user'),

                "status": udata.get('status', 'approved'),

                "created_at": udata.get('created_at', '-'),

                "profile_pic": udata.get('profile_pic')

            })

            

        return jsonify(user_list)

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/profile/upload_photo', methods=['POST'])

def upload_profile_photo():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "No autheticado"}), 401

        

    requester = session.get('user')

    requester_role = session.get('role')

    

    target_user = request.form.get('target_user')

    

    if target_user and requester_role == 'admin':

        username = target_user

    else:

        username = requester

    

    if 'file' not in request.files:

         return jsonify({"status": "error", "message": "No file part"}), 400

         

    file = request.files['file']

    if file.filename == '':

         return jsonify({"status": "error", "message": "No selected file"}), 400

         

    if file:

        try:

            import hashlib

            ext = os.path.splitext(file.filename)[1]

            safe_name = f"{hashlib.md5(username.encode()).hexdigest()}_{int(time.time())}{ext}"

            target_path = PROFILE_PICS_DIR / safe_name

            

            file.save(target_path)

            

            with open(USERS_FILE, 'r') as f:

                users = json.load(f)

            

            if username in users:

                users[username]['profile_pic'] = safe_name

                _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

                    

            return jsonify({"status": "success", "filename": safe_name})

            

        except Exception as e:

             return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/profile/update', methods=['POST'])

def update_profile():

    if not session.get('user'):

        return jsonify({"status": "error", "message": "No autheticado"}), 401

    

    username = session.get('user')

    data = request.json

    

    new_display = data.get('display_name')

    new_position = data.get('position')

    new_department = data.get('department')

    new_signature = data.get('signature')

    new_pass = data.get('password')

    

    try:

        if not USERS_FILE.exists():

            return jsonify({"status": "error", "message": "User db not found"}), 500

            

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if username not in users:

            return jsonify({"status": "error", "message": "User not found"}), 404

            

        changes = False

        if new_display is not None:

            new_display = str(new_display).strip()

            if new_display:

                users[username]['display_name'] = new_display

                session['display_name'] = new_display

                changes = True

        if new_position is not None:

            users[username]['position'] = str(new_position).strip()

            changes = True

        if new_department is not None:

            users[username]['department'] = str(new_department).strip()

            changes = True

        if new_signature is not None:

            users[username]['signature'] = str(new_signature).strip()

            changes = True

        if new_pass:

            users[username]['password'] = _generate_password_hash_compatible(new_pass)

            changes = True

            

        if changes:

            _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

                

        return jsonify({"status": "success"})

        

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/approve', methods=['POST'])

def approve_user():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    

    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            users[target_user]['status'] = 'approved'

            _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

                

            actor = session.get('user', 'system')

            log_action(actor, "Approved User", f"Target: {target_user}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

            

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/update_role', methods=['POST'])

def update_role():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    new_role = data.get('role')

    

    if new_role not in ['admin', 'user', 'externos', 'calidad']:

         return jsonify({"status": "error", "message": "Rol invÃ¡lido"}), 400



    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            users[target_user]['role'] = new_role

            _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

                

            actor = session.get('user', 'system')

            log_action(actor, "Updated Role", f"Target: {target_user}, New Role: {new_role}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/admin/delete_user', methods=['POST'])

def delete_user():

    if session.get('role') != 'admin':

        return jsonify({"status": "error", "message": "No autorizado"}), 403

        

    data = request.json

    target_user = data.get('username')

    

    try:

        with open(USERS_FILE, 'r') as f:

            users = json.load(f)

            

        if target_user in users:

            del users[target_user]

            _write_json_file_atomic(USERS_FILE, users, indent=4, ensure_ascii=False)

                

            actor = session.get('user', 'system')

            log_action(actor, "Deleted User", f"Target: {target_user}")

            

            return jsonify({"status": "success"})

        else:

            return jsonify({"status": "error", "message": "Usuario no encontrado"}), 404

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/approve-po', methods=['POST'])
def approve_po():
    try:
        debug_path = BASE_DIR / "Codigos/global_debug.log"
        with open(debug_path, "a") as f:
            f.write(f"{datetime.now()}: Request received for /api/approve-po. Body: {request.get_data(as_text=True)}\n")
    except Exception as e:
        print(f"Log Error: {e}")

    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401
         
    data = request.json
    po_id = data.get('po_id')
    
    if not po_id:
        return jsonify({"status": "error", "message": "Falta PO ID"}), 400
        
    source_folder = DATA_DIR / po_id
    if not source_folder.exists():
        # Check if already processed
        if (PROCESSED_PATH / po_id).exists():
             return jsonify({"status": "success", "message": "Ya fue aprobado"}), 200
        return jsonify({"status": "error", "message": "Carpeta de PO no encontrada"}), 404

    # Resolve User Display Name
    actor = session.get('user', 'system')
    actor_name = actor
    if USERS_FILE.exists():
        try:
            with open(USERS_FILE, 'r') as f:
                users = json.load(f)
                if actor in users:
                    actor_name = users[actor].get('display_name', actor)
        except: pass

    # Start Background Thread
    print(f"[APPROVE] Launching background thread for {po_id}")
    thread = threading.Thread(target=approve_po_background, args=(po_id, data, actor, actor_name))
    thread.daemon = True
    thread.start()
    print(f"[APPROVE] Thread started for {po_id}")
    
    return jsonify({"status": "queued", "message": "AprobaciÃ³n iniciada en segundo plano"}), 202

def approve_po_background(po_id, data, actor, actor_name):
    import sys
    # Helper for dual logging (File + Console)
    def log_trace(msg, folder=None):
        print(msg, flush=True)
        try: sys.stdout.flush()
        except: pass
        if folder and folder.exists():
            try:
                with open(folder / "_trace_approval.txt", "a", encoding="utf-8") as f:
                    f.write(f"{datetime.now().strftime('%H:%M:%S')} - {msg}\n")
            except: pass
    """
    Background worker:
    1. MOVE folder to Processed (User request: immediate move).
    2. Process Excel Updates.
    3. Mark as Complete.
    """
    try:
        # Determine initial folder to start logging
        log_folder = DATA_DIR / po_id
        if not log_folder.exists():
            log_folder = PROCESSED_PATH / po_id

        log_trace(f"[BACKGROUND] === START Approval for {po_id} by {actor} ===", log_folder)
        
        source_folder = DATA_DIR / po_id
        destination_folder = PROCESSED_PATH / po_id
        
        PROCESSED_PATH.mkdir(parents=True, exist_ok=True)
            
        log_trace(f"[BACKGROUND] Step 1: Requesting move to {destination_folder}", log_folder)
        
        current_work_folder = source_folder
        
        if source_folder.exists():
            if destination_folder.exists():
                log_trace(f"[BACKGROUND] Error: Destination {po_id} already exists!", log_folder)
                # Fallback: Work in source, but this shouldn't happen based on route checks
                current_work_folder = source_folder
            else:
                try:
                    log_trace(f"[BACKGROUND] Moving now...", log_folder)
                    shutil.move(str(source_folder), str(destination_folder))
                    current_work_folder = destination_folder
                    log_trace(f"[BACKGROUND] Folder moved successfully.", current_work_folder)
                    log_action(actor, "Approved PO", f"Moved {po_id} to Processed (Processing Start)")
                except Exception as e:
                    log_trace(f"[BACKGROUND] Failed to move folder: {e}", log_folder)
                    with open(source_folder / "approval_error.log", "w") as f:
                        f.write(f"Move Error: {e}")
                    return

        elif destination_folder.exists():
            # Already moved (maybe retry?)
            current_work_folder = destination_folder
            log_trace(f"[BACKGROUND] Folder was already in Processed. Resuming...", current_work_folder)

        # Create LOCK file to indicate processing is still active
        lock_file = current_work_folder / "processing.lock"
        with open(lock_file, "w") as f:
            f.write("Processing in progress...")
        
        try:
            approved_count = data.get('approved_count', 0)
            total_count = data.get('total_count', 0)
            approved_items = data.get('approved_items', [])

            # Normalize approved items list
            if not isinstance(approved_items, list):
                approved_items = []
            approved_items = [str(item) for item in approved_items if item]

            try:
                approved_count = int(approved_count or 0)
            except Exception:
                approved_count = 0
            try:
                total_count = int(total_count or 0)
            except Exception:
                total_count = 0

            # Always resolve total from real PDFs when frontend does not send it.
            if total_count <= 0:
                pdfs = list(current_work_folder.glob('*.pdf'))
                total_count = len(pdfs)

            # If a concrete list is provided, it has priority over counters.
            if approved_items:
                approved_count = len(approved_items)
            elif approved_count <= 0 and total_count > 0:
                # Backward compatibility: old clients that approved all without sending list.
                approved_count = total_count

            if total_count > 0:
                approved_count = max(0, min(approved_count, total_count))

            # Create Metadata
            timestamp = datetime.now().strftime('%d/%m/%Y %H:%M')

            metadata = {
                "approved_by": actor,
                "approved_by_name": actor_name,
                "approved_at": timestamp,
                "status": "Aprobado",
                "counts": {"approved": approved_count, "total": total_count},
                "approved_items": approved_items
            }
            
            _write_json_file_atomic(current_work_folder / "approval_info.json", metadata, indent=4, ensure_ascii=False)
            log_trace(f"[BACKGROUND] Step 2: Metadata created.", current_work_folder)

            # Update Auxiliary Excels (Heavy Operation)
            log_trace(f"[BACKGROUND] Step 3: Updating Excels (Batch Process)...", current_work_folder)
            process_excel_updates_v2(current_work_folder, approved_items, po_id)
            log_trace(f"[BACKGROUND] Step 3: Excel updates finished.", current_work_folder)
            
            # Remove Lock File => Signals Success
            if lock_file.exists():
                os.remove(lock_file)
            log_trace(f"[BACKGROUND] === FINISHED Approval for {po_id} ===", current_work_folder)

        except Exception as exc:
            log_trace(f"[BACKGROUND] Processing Error: {exc}", current_work_folder)
            print(f"[BACKGROUND] Processing Error: {exc}", flush=True)
            with open(current_work_folder / "approval_error.log", "w") as f:
                f.write(f"Processing Error: {exc}")
            # Remove lock so user sees the error
            if lock_file.exists(): os.remove(lock_file)
            return

    except Exception as e:
        print(f"[BACKGROUND] Critical Error for {po_id}: {e}", flush=True)
        try:
            # Try to write log wherever the folder is
            fail_path = PROCESSED_PATH / po_id if (PROCESSED_PATH / po_id).exists() else DATA_DIR / po_id
            if fail_path.exists():
                with open(fail_path / "approval_error.log", "w") as f:
                    f.write(f"System Error: {e}")
        except: pass

@app.route('/api/check-approval-status/<po_id>')
def check_approval_status(po_id):
    # 1. Check Processed Path
    dest = PROCESSED_PATH / po_id
    if dest.exists():
        # It's in Processed, but is it done?
        if (dest / "approval_error.log").exists():
            with open(dest / "approval_error.log", "r") as f: msg = f.read()
            return jsonify({"status": "error", "message": msg})
            
        if (dest / "processing.lock").exists():
            return jsonify({"status": "processing", "message": "Procesando Excels..."})
            
        return jsonify({"status": "success", "message": "Aprobado correctamente"})
        
    # 2. Check Source Path (if it hasn't moved yet or failed before move)
    source = DATA_DIR / po_id
    if source.exists():
        if (source / "approval_error.log").exists():
            with open(source / "approval_error.log", "r") as f: msg = f.read()
            return jsonify({"status": "error", "message": msg})
        return jsonify({"status": "processing", "message": "Iniciando..."})
        
    return jsonify({"status": "error", "message": "Carpeta no encontrada"})



@app.route('/api/save-po-progress', methods=['POST'])

def save_po_progress():

    """Guarda el estado parcial de los productos (aprobados/pendientes) sin mover la carpeta."""

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

         

    data = request.json

    po_id = data.get('po_id')

    approved_items = data.get('approved_items', [])

    

    print(f"[SAVE PROGRESS] Recibido para PO: '{po_id}' - Items: {len(approved_items)}")

    

    if not po_id: 

        print("[SAVE PROGRESS] Error: Falta PO ID")

        return jsonify({"status": "error", "message": "Falta PO ID"}), 400

    

    # Buscar carpeta (primero en En Progreso, fallback Procesado por si se edita historial)

    target_folder = DATA_DIR / po_id

    if not target_folder.exists():

        print(f"[SAVE PROGRESS] No en DATA_DIR ({DATA_DIR}), buscando en PROCESSED...")

        target_folder = PROCESSED_PATH / po_id



    if not target_folder.exists():

         print(f"[SAVE PROGRESS] Error: Carpeta no encontrada en ninguna ruta para '{po_id}'")

         return jsonify({"status": "error", "message": "Carpeta PO no encontrada"}), 404



    print(f"[SAVE PROGRESS] Carpeta destino: {target_folder}")



    try:

        # Load existing meta to preserve other fields if any

        meta_file = target_folder / "approval_info.json"

        metadata = {}

        if meta_file.exists():

            try:

                with open(meta_file, 'r', encoding='utf-8') as f:

                    metadata = json.load(f)

            except: pass

            

        # Calculate diff for logging

        old_items = set(metadata.get("approved_items", []))

        new_items_set = set(approved_items)

        added = new_items_set - old_items

        removed = old_items - new_items_set



        # Update fields

        metadata["approved_items"] = approved_items

        metadata["last_updated"] = datetime.now().strftime('%d/%m/%Y %H:%M')

        

        print(f"[SAVE PROGRESS] Escribiendo JSON en {meta_file}")

        _write_json_file_atomic(meta_file, metadata, indent=4, ensure_ascii=False)

            

        # Log Activity with specific details

        actor = session.get('user', 'system')

        if added:

            items_str = ", ".join(list(added))

            log_action(actor, "Aprobado", f"Item(s): {items_str} en {po_id}")

        elif removed:

            items_str = ", ".join(list(removed))

            log_action(actor, "Desaprobado", f"Item(s): {items_str} en {po_id}")

        else:

             log_action(actor, "Updated PO Progress", f"PO: {po_id}, Total: {len(approved_items)}")



        return jsonify({"status": "success"})

    except Exception as e:

        print(f"Error saving progress: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/reverse-approval', methods=['POST'])

def reverse_approval():

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

         

    data = request.json

    po_id = data.get('po_id')

    

    if not po_id:

        return jsonify({"status": "error", "message": "Falta PO ID"}), 400

        

    source_folder = PROCESSED_PATH / po_id

    if not source_folder.exists():

        return jsonify({"status": "error", "message": "Registro procesado no encontrado"}), 404

        

    destination_folder = DATA_DIR / po_id

    # Check if a folder with same name already exists in progress (unlikely but possible)

    if destination_folder.exists():

         return jsonify({"status": "error", "message": "Ya existe una carpeta activa con este ID"}), 400

         

    try:

        shutil.move(str(source_folder), str(destination_folder))

        

        # Update metadata to reflect modification status

        meta_file = destination_folder / "approval_info.json"

        if meta_file.exists():

             try:

                 with open(meta_file, 'r', encoding='utf-8') as f:

                     meta = json.load(f)

                 meta['status'] = 'En ModificaciÃ³n'

                 meta['modified_at'] = datetime.now().strftime('%d/%m/%Y %H:%M')

                 _write_json_file_atomic(meta_file, meta, indent=4, ensure_ascii=False)

             except: pass

        

        actor = session.get('user', 'system')

        log_action(actor, "Reversed Approval", f"Moved {po_id} back to In Progress")

        

        return jsonify({"status": "success", "message": "Registro movido a En Progreso"})

    except Exception as e:

        return jsonify({"status": "error", "message": f"Error al mover: {e}"}), 500



@app.route('/api/check_session')

def check_session():

    user = session.get('user')

    role = session.get('role')

    if user:

        # Get extended info

        display_name = user

        profile_pic = None

        email = None

        position = ''

        department = ''

        signature = ''
        try:

            if USERS_FILE.exists():

                with open(USERS_FILE, 'r') as f:

                    users = json.load(f)

                    if user in users:

                        display_name = users[user].get('display_name', user)

                        profile_pic = users[user].get('profile_pic')

                        email = users[user].get('email') or users[user].get('correo')

                        position = users[user].get('position') or ''

                        department = users[user].get('department') or ''

                        signature = users[user].get('signature') or ''

        except: pass

        

        return jsonify({

            "status": "authenticated", 

            "user": user, 

            "role": role,

            "display_name": display_name,

            "profile_pic": profile_pic,

            "email": email,

            "position": position,

            "department": department,

            "signature": signature

        })

    else:

        return jsonify({"status": "guest"})



@app.route('/files/<path:filename>')

def serve_file(filename):

    # print(f"Requested File: {filename}")

    try:

        # Try DATA_DIR first

        if (DATA_DIR / filename).exists():

            return send_from_directory(DATA_DIR, filename)

        # Try PROCESSED_PATH

        # filename might contain the PO ID as first folder

        # e.g. R016-01-xxxx/file.pdf

        parts = Path(filename).parts

        if len(parts) > 1:

            po_id = parts[0]

            if (PROCESSED_PATH / po_id).exists():

                return send_from_directory(PROCESSED_PATH, filename)

                

        # Last resort fallback (direct match in processed root? unlikely)

        if (PROCESSED_PATH / filename).exists():

             return send_from_directory(PROCESSED_PATH, filename)

             

        return jsonify({"error": "File not found"}), 404

        

    except Exception as e:

        print(f"Error serving file {filename}: {e}")

        return jsonify({"error": str(e)}), 404



# --- OPEN R016 FOLDER (Windows Explorer) ---

@app.route('/api/open-r016-folder')

def open_r016_folder():

    target_path = R016_REGISTROS_DIR

    try:

        if not target_path.exists():

            return jsonify({"status": "error", "message": f"Ruta no encontrada: {target_path}"}), 404

        _open_path_in_file_manager(target_path)

        return jsonify({"status": "success"})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500


# --- ISO R019 FOLDERS (Return Path) ---
@app.route('/api/iso-open-folder', methods=['POST'])
def open_iso_folder():
    try:
        data = request.get_json(silent=True) or {}
        key = str(data.get('key', '')).strip().upper()
        allowed = {'R019-01', 'R019-02', 'R019-03', 'R019-04'}
        if key not in allowed:
            return jsonify({"status": "error", "message": "Carpeta invÃ¡lida."}), 400

        folder_map = {

            'R019-01': ISO_R01901_DOCS_DIR,

            'R019-02': ISO_R01902_DOCS_DIR,

            'R019-03': ISO_DOCS_ROOT,

            'R019-04': ISO_R01904_DOCS_DIR,

        }

        target_path = folder_map[key]
        if not target_path.exists():
            return jsonify({"status": "error", "message": f"Ruta no encontrada: {target_path}"}), 404

        return jsonify({"status": "success", "path": str(target_path)})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# --- ISO NEXT REGISTRY (R019-03 LISTADO) ---
@app.route('/api/iso-next-registry')
def iso_next_registry():
    target = _find_iso_r01903_target()
    if target is None or not target.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontro el archivo R019-03 en {ISO_DOCS_ROOT}"
        }), 404

    try:
        result = _compute_iso_next_registry()
        result["status"] = "success"
        return jsonify(result)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-r01903-records')
def iso_r01903_records():
    target = _find_iso_r01903_target()

    if target is None or not target.exists():
        return jsonify({
            "status": "success",
            "records": [],
            "source": "",
            "warning": f"No se encontro el archivo R019-03 en {ISO_DOCS_ROOT}"
        })

    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        return str(val).strip() if val is not None else ""

    try:
        wb = _load_workbook_quietly(str(target), read_only=True, data_only=True)
        ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
        ws = wb[ws_name] if ws_name else wb.active

        records = []
        for row in ws.iter_rows(min_row=5, values_only=True):
            if not row or len(row) < 7:
                continue
            val_a = row[0]
            val_b = row[1] if len(row) > 1 else None
            if (val_a is None or str(val_a).strip() == "") and (val_b is None or str(val_b).strip() == ""):
                continue
            records.append({
                "numero": str(val_a).strip() if val_a is not None else "",
                "descripcion": str(val_b).strip() if val_b is not None else "",
                "solicitante": str(row[2]).strip() if len(row) > 2 and row[2] is not None else "",
                "fecha_inicio": fmt_date(row[3]) if len(row) > 3 else "",
                "etapa": str(row[4]).strip() if len(row) > 4 and row[4] is not None else "",
                "situacion": str(row[5]).strip() if len(row) > 5 and row[5] is not None else "",
                "fecha_fin": fmt_date(row[6]) if len(row) > 6 else "",
            })

        wb.close()
        return jsonify({"status": "success", "records": records, "source": str(target)})
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer R019-03: {e}"}), 500


@app.route('/api/notifications', methods=['GET'])
def get_notifications():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    user = session.get('user')
    display_name = _resolve_display_name(user)
    email = _resolve_user_email(user)
    keys = {_normalize_user_key(user), _normalize_user_key(display_name)}
    if email:
        keys.add(_normalize_user_key(email))

    items = _load_notifications()
    unread = []
    for n in items:
        if n.get("read"):
            continue
        recip_key = _normalize_user_key(n.get("recipient"))
        recip_email = _normalize_user_key(n.get("recipient_email"))
        if recip_key in keys or (recip_email and recip_email in keys):
            unread.append(n)
    unread.sort(key=lambda x: x.get("created_at") or "", reverse=True)
    return jsonify({"status": "success", "notifications": unread, "unread": len(unread)})


@app.route('/api/notifications/mark-read', methods=['POST'])
def mark_notification_read():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    data = request.get_json(silent=True) or {}
    notif_id = data.get("id")
    if not notif_id:
        return jsonify({"status": "error", "message": "ID requerido"}), 400

    items = _load_notifications()
    updated = False
    for n in items:
        if n.get("id") == notif_id:
            n["read"] = True
            updated = True
            break
    if updated:
        _save_notifications(items)
    return jsonify({"status": "success", "updated": updated})


@app.route('/api/iso-r01902-events')
def iso_r01902_events():
    bp = (request.args.get("bp") or "").strip()
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400
    try:
        events, target = _read_r01902_events_for_bp(bp)
        return jsonify({"status": "success", "events": events, "file": str(target)})
    except FileNotFoundError as e:
        return jsonify({"status": "success", "events": [], "file": "", "warning": str(e)})
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer R019-02: {e}"}), 500


def _read_r01902_events_for_bp(bp: str):
    r01902_dir = ISO_R01902_DOCS_DIR
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp

    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)

    if target is None and r01902_dir.exists():
        for p in r01902_dir.glob("R019-02 Rev03 -*.*"):
            if safe_bp.lower() in p.stem.lower():
                target = p
                break

    if target is None or not target.exists():
        raise FileNotFoundError(f"No se encontrÃ³ el archivo R019-02 para {bp}")

    def fmt_date(val):
        if isinstance(val, datetime):
            return val.strftime("%d/%m/%Y")
        if isinstance(val, date):
            return val.strftime("%d/%m/%Y")
        return str(val).strip() if val is not None else ""

    wb = _load_workbook_quietly(str(target), read_only=True, data_only=True)
    ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
    ws = wb[ws_name] if ws_name else wb.active

    events = []
    empty_streak = 0
    max_row = ws.max_row or 0

    for row in range(4, max_row + 1):
        values = [ws.cell(row=row, column=col).value for col in range(1, 9)]
        if all(v is None or str(v).strip() == "" for v in values):
            empty_streak += 1
            if empty_streak >= 5:
                break
            continue
        empty_streak = 0

        events.append({
            "row": row,
            "fecha": fmt_date(values[0]),
            "etapa": str(values[1]).strip() if values[1] is not None else "",
            "area": str(values[2]).strip() if values[2] is not None else "",
            "empresa": str(values[3]).strip() if values[3] is not None else "",
            "descripcion": str(values[4]).strip() if values[4] is not None else "",
            "resultados": str(values[5]).strip() if values[5] is not None else "",
            "accion": str(values[6]).strip() if values[6] is not None else "",
            "usuario": str(values[7]).strip() if values[7] is not None else "",
        })

    wb.close()
    return events, target




def _find_r01904_model(iso_root: Path) -> Path:
    candidates = [
        iso_root / "R019-04" / "R019-04-Modelo.mpp",
        iso_root / "Codigos" / "R019-04" / "R019-04-Modelo.mpp",
    ]
    for cand in candidates:
        if cand.exists():
            return cand
    r01904_dir = iso_root / "R019-04"
    if r01904_dir.exists():
        matches = list(r01904_dir.glob("R019-04-Modelo*.mpp"))
        if matches:
            return matches[0]
        mpp_files = list(r01904_dir.glob("*.mpp"))
        if mpp_files:
            return mpp_files[0]
    raise FileNotFoundError("No se encontro el modelo R019-04.")


def _r01904_plan_filename_candidates(safe_bp: str):
    bp = str(safe_bp or "").strip()
    return [
        f"Planificación - {bp}.mpp",
        f"PlanificaciÃ³n - {bp}.mpp",
        f"Planificacion - {bp}.mpp",
        f"R019-04 - {bp}.mpp",
    ]


def _resolve_r01904_plan_path(r01904_dir: Path, safe_bp: str) -> Path:
    names = _r01904_plan_filename_candidates(safe_bp)
    canonical = r01904_dir / names[0]
    existing = []
    for name in names:
        candidate = r01904_dir / name
        try:
            if candidate.exists():
                existing.append(candidate)
        except Exception:
            pass

    if not existing:
        return canonical

    legacy = [p for p in existing if p != canonical]

    # Always keep canonical file name as write destination.
    if canonical.exists():
        source = None
        synced = True
        if legacy:
            try:
                latest_legacy = max(legacy, key=lambda p: p.stat().st_mtime)
                if latest_legacy.stat().st_mtime > canonical.stat().st_mtime:
                    source = latest_legacy
            except Exception:
                source = legacy[0]
        if source is not None:
            synced = False
            for attempt in range(1, 4):
                try:
                    canonical.write_bytes(source.read_bytes())
                    synced = True
                    break
                except PermissionError:
                    time.sleep(0.4 * attempt)
                except Exception:
                    break
        if synced:
            for old in legacy:
                try:
                    if old.exists():
                        old.unlink()
                except Exception:
                    pass
        return canonical

    # Canonical does not exist: migrate newest legacy variant into canonical.
    try:
        source = max(legacy or existing, key=lambda p: p.stat().st_mtime)
    except Exception:
        source = (legacy or existing)[0]

    migrated = False
    for attempt in range(1, 4):
        try:
            source.replace(canonical)
            migrated = True
            break
        except PermissionError:
            time.sleep(0.4 * attempt)
        except Exception:
            break

    if not migrated:
        for attempt in range(1, 4):
            try:
                canonical.write_bytes(source.read_bytes())
                migrated = True
                break
            except PermissionError:
                time.sleep(0.4 * attempt)
            except Exception:
                break

    if migrated:
        for old in legacy:
            try:
                if old.exists():
                    old.unlink()
            except Exception:
                pass

    return canonical


def _ensure_r01904_created(bp: str) -> Path:
    r01904_dir = ISO_R01904_DOCS_DIR
    r01904_dir.mkdir(parents=True, exist_ok=True)
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    dest = _resolve_r01904_plan_path(r01904_dir, safe_bp)
    if dest.exists():
        return dest
    model = _find_r01904_model(ISO_DOCS_ROOT)
    last_err = None
    for attempt in range(1, 4):
        try:
            dest.write_bytes(model.read_bytes())
            return dest
        except PermissionError as e:
            last_err = e
            time.sleep(0.4 * attempt)
    if last_err:
        raise last_err
    return dest


def _generate_r01904(bp: str, allow_empty: bool = False) -> Path:
    import sys

    events, _ = _read_r01902_events_for_bp(bp)
    if not events:
        if allow_empty:
            return _ensure_r01904_created(bp)
        raise ValueError("R019-02 no tiene eventos para este BP.")

    iso_root = ISO_CODE_ROOT
    r01904_dir = ISO_R01904_DOCS_DIR
    r01904_dir.mkdir(parents=True, exist_ok=True)
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    dest = _resolve_r01904_plan_path(r01904_dir, safe_bp)

    csv_path = Path(tempfile.gettempdir()) / f"r01904_{safe_bp}.csv"
    with csv_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(
            fh,
            fieldnames=["fecha", "etapa", "area", "empresa", "descripcion", "resultado", "accion", "usuario"]
        )
        writer.writeheader()
        for ev in events:
            writer.writerow({
                "fecha": ev.get("fecha", ""),
                "etapa": ev.get("etapa", ""),
                "area": ev.get("area", ""),
                "empresa": ev.get("empresa", ""),
                "descripcion": ev.get("descripcion", ""),
                "resultado": ev.get("resultados", "") or ev.get("resultado", ""),
                "accion": ev.get("accion", ""),
                "usuario": ev.get("usuario", "")
            })

    script_path = iso_root / "Codigos" / "R019-04" / "fill_r01904.py"
    model_path = iso_root / "Codigos" / "R019-04" / "modelo_duraciones.json"
    if not script_path.exists():
        raise FileNotFoundError("No se encontro fill_r01904.py")

    cmd = [
        sys.executable,
        str(script_path),
        str(csv_path),
        "--modelo",
        str(model_path),
        "--out",
        str(dest)
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        err = (result.stderr or result.stdout or "").strip()
        raise RuntimeError(err or "No se pudo generar R019-04.")
    return dest

@app.route('/api/iso-r01904-generate', methods=['POST'])
def iso_r01904_generate():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or "").strip()
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400

    try:
        dest = _generate_r01904(bp, allow_empty=False)
        return jsonify({"status": "success", "file": str(dest)})
    except FileNotFoundError as e:
        return jsonify({"status": "error", "message": str(e)}), 404
    except ValueError as e:
        return jsonify({"status": "error", "message": str(e)}), 400
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo generar R019-04: {e}"}), 500


@app.route('/api/iso-r01902-approve', methods=['POST'])
def iso_r01902_approve():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or "").strip()
    row = data.get("row")
    status = (data.get("status") or "Aprobado").strip()
    accion = (data.get("accion") or "").strip()
    if not bp or not row:
        return jsonify({"status": "error", "message": "BP y fila requeridos"}), 400

    iso_docs_root = ISO_DOCS_ROOT
    iso_code_root = ISO_CODE_ROOT
    r01902_dir = ISO_R01902_DOCS_DIR
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp
    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)
    if target is None:
        return jsonify({"status": "error", "message": f"No se encontr\u00f3 el archivo R019-02 para {bp}"}), 404

    def _is_permission_error(err: Exception) -> bool:
        if isinstance(err, PermissionError):
            return True
        if isinstance(err, OSError):
            winerr = getattr(err, "winerror", None)
            if winerr in (32, 33):
                return True
            if getattr(err, "errno", None) == 13:
                return True
        msg = str(err).lower()
        return "permission denied" in msg or "being used by another process" in msg

    last_err = None
    for attempt in range(1, 4):
        wb = None
        try:
            wb = _load_workbook_quietly(str(target), keep_vba=target.suffix.lower() == ".xlsm")
            ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
            ws = wb[ws_name] if ws_name else wb.active
            ws.cell(row=int(row), column=6).value = status
            if accion:
                ws.cell(row=int(row), column=7).value = accion
            wb.save(target)
            return jsonify({"status": "success"})
        except Exception as e:
            last_err = e
            if _is_permission_error(e):
                time.sleep(0.4 * attempt)
                continue
            break
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass

    if last_err:
        if _is_permission_error(last_err):
            return jsonify({"status": "error", "message": "No se pudo aprobar el evento: el archivo R019-02 estÃ¡ en uso. Cierre Excel y reintente."}), 500
        return jsonify({"status": "error", "message": f"No se pudo aprobar el evento: {last_err}"}), 500
    return jsonify({"status": "error", "message": "No se pudo aprobar el evento."}), 500


@app.route('/api/iso-r01902-append', methods=['POST'])
def iso_r01902_append():
    data = request.get_json(silent=True) or {}
    bp = (data.get("bp") or data.get("BP") or "").strip()
    event = data.get("event") or {}
    if not bp:
        return jsonify({"status": "error", "message": "BP requerido"}), 400

    iso_docs_root = ISO_DOCS_ROOT
    iso_code_root = ISO_CODE_ROOT
    r01902_dir = ISO_R01902_DOCS_DIR
    safe_bp = re.sub(r"[^A-Za-z0-9-]+", "", bp) or bp

    candidates = [
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsm",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xlsx",
        r01902_dir / f"R019-02 Rev03 - {safe_bp}.xls",
    ]
    target = next((p for p in candidates if p.exists()), None)

    if target is None and r01902_dir.exists():
        for p in r01902_dir.glob("R019-02 Rev03 -*.*"):
            if safe_bp.lower() in p.stem.lower():
                target = p
                break

    if target is None or not target.exists():
        return jsonify({"status": "error", "message": f"No se encontr\u00f3 el archivo R019-02 para {bp}"}), 404

    try:
        codes_r01902 = iso_code_root / "Codigos" / "R019-02"
        mod02 = _load_fill_r01902_module(codes_r01902)
        if hasattr(mod02, "append_r01902_event"):
            row = mod02.append_r01902_event(target, event)
        else:
            # Fallback: write with openpyxl if function not available
            wb = _load_workbook_quietly(str(target), keep_vba=target.suffix.lower() == ".xlsm")
            try:
                ws_name = next((n for n in wb.sheetnames if n.strip().lower() == "listado"), None)
                ws = wb[ws_name] if ws_name else wb.active
                row = (ws.max_row or 3) + 1
                ws.cell(row=row, column=1).value = event.get("fecha") or ""
                ws.cell(row=row, column=2).value = event.get("etapa") or ""
                ws.cell(row=row, column=3).value = event.get("area") or ""
                ws.cell(row=row, column=4).value = event.get("empresa") or ""
                ws.cell(row=row, column=5).value = event.get("descripcion") or ""
                ws.cell(row=row, column=6).value = event.get("resultados") or ""
                ws.cell(row=row, column=7).value = event.get("accion") or ""
                ws.cell(row=row, column=8).value = event.get("usuario") or ""
                wb.save(target)
            finally:
                try:
                    wb.close()
                except Exception:
                    pass
        # NotificaciÃƒÂ³n de aprobaciÃƒÂ³n si aplica
        if event.get("requiere_aprobacion"):
            try:
                for approver in ("Luciano Cochis", "Lorenzo Contigiani"):
                    _add_notification(
                        approver,
                        "approval",
                        f"Aprobacion pendiente para {bp}",
                        {"bp": bp, "row": row, "etapa": event.get("etapa", "")}
                    )
            except Exception:
                pass

        r01903_updated = False
        r01903_error = None
        try:
            codes_r01903 = iso_code_root / "Codigos" / "R019-03"
            mod03 = _load_fill_r01903_module(codes_r01903)
            if hasattr(mod03, "update_r01903_status"):
                etapa = (event.get("etapa") or "").strip()
                situacion = (event.get("situacion") or "").strip()
                fecha_fin = None
                if etapa.lower() == "cierre":
                    fecha_fin = (event.get("fecha") or "").strip()
                template_path = None
                if hasattr(mod03, "resolve_template_path"):
                    template_path = mod03.resolve_template_path(iso_docs_root)
                r01903_updated = mod03.update_r01903_status(bp, etapa, situacion, fecha_fin, template_path=template_path)
            else:
                r01903_error = "fill_r01903.py no expone update_r01903_status"
        except Exception as e:
            r01903_error = f"No se pudo actualizar R019-03: {e}"
        r01904_error = None
        try:
            _generate_r01904(str(bp), allow_empty=True)
        except Exception as e:
            r01904_error = f"No se pudo generar R019-04: {e}"

        resp = {"status": "success", "row": row}
        if r01903_error:
            resp["r01903_error"] = r01903_error
        if r01904_error:
            resp["r01904_error"] = r01904_error
        resp["r01903_updated"] = bool(r01903_updated)
        return jsonify(resp)
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo guardar evento en R019-02: {e}"}), 500


def _iso_r01901_paths():
    iso_root = ISO_CODE_ROOT
    docs_dir = ISO_R01901_DOCS_DIR
    codes_dir = iso_root / "Codigos" / "R019-01"
    return iso_root, docs_dir, codes_dir


def _normalize_iso_bp(value) -> str:
    return re.sub(r"[^A-Za-z0-9-]+", "", str(value or "").strip()) or "BP-XXXXX"


def _r01901_filename_from_bp(bp: str, suffix: str = ".docx") -> str:
    safe_bp = _normalize_iso_bp(bp)
    file_bp = re.sub(r"^BP-", "BP", safe_bp)
    return f"R019-01-DATOS DE ENTRADA-{file_bp}{suffix}"


def _r01901_path_candidates(docs_dir: Path, bp: str) -> list[Path]:
    safe_bp = _normalize_iso_bp(bp)
    file_bp = re.sub(r"^BP-", "BP", safe_bp)
    candidates = [
        docs_dir / _r01901_filename_from_bp(safe_bp, ".docx"),
        docs_dir / _r01901_filename_from_bp(safe_bp, ".doc"),
    ]
    try:
        for entry in docs_dir.iterdir():
            if not entry.is_file():
                continue
            stem_up = entry.stem.upper()
            if "R019-01" in stem_up and file_bp.upper() in stem_up:
                candidates.append(entry)
    except Exception:
        pass

    unique = []
    seen = set()
    for candidate in candidates:
        key = str(candidate).lower()
        if key in seen:
            continue
        seen.add(key)
        unique.append(candidate)
    return unique


def _resolve_r01901_output_path(docs_dir: Path, bp: str, must_exist: bool = False) -> Path:
    candidates = _r01901_path_candidates(docs_dir, bp)
    if must_exist:
        for candidate in candidates:
            if candidate.exists():
                return candidate
    return candidates[0]


def _enrich_iso_payload_with_r01901_metadata(payload: dict, docs_dir: Path, bp: str) -> dict:
    enriched = dict(payload or {})
    safe_bp = _normalize_iso_bp(bp)
    output_path = _resolve_r01901_output_path(docs_dir, safe_bp, must_exist=False)
    enriched["BP"] = safe_bp
    enriched["R01901_File"] = output_path.name
    enriched["R01901_Path"] = str(output_path)
    return enriched


def _load_iso_payload_template(codes_dir: Path) -> dict:
    template_path = codes_dir / "payload_template.json"
    if not template_path.exists():
        return {}
    try:
        with template_path.open("r", encoding="utf-8") as fh:
            return json.load(fh)
    except Exception:
        return {}

def _iso_payloads_path() -> Path:
    ISO_PAYLOADS_FILE.parent.mkdir(parents=True, exist_ok=True)
    return ISO_PAYLOADS_FILE


def _load_iso_payloads() -> dict:
    path = _iso_payloads_path()
    if not path.exists():
        return {}
    try:
        with path.open("r", encoding="utf-8") as fh:
            data = json.load(fh) or {}
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _save_iso_payloads(data: dict) -> None:
    path = _iso_payloads_path()
    _write_json_file_atomic(path, data or {}, indent=2, ensure_ascii=False)


def _iso_payload_keys(payload: dict) -> list:
    keys = set()
    if not isinstance(payload, dict):
        return []
    num = payload.get("Numero_de_Registro") or payload.get("numero")
    if num is not None:
        s = str(num).strip()
        if s:
            keys.add(s)
        m = re.search(r"(\d+)", s)
        if m:
            keys.add(m.group(1))
    bp = payload.get("BP") or payload.get("bp")
    if bp:
        s = str(bp).strip()
        if s:
            keys.add(s)
    return list(keys)


def _upsert_iso_payload(payload: dict, merge: bool = False) -> None:
    if not isinstance(payload, dict):
        return
    data = _load_iso_payloads()
    if merge:
        for k in _iso_payload_keys(payload):
            existing = data.get(str(k))
            if isinstance(existing, dict):
                merged = dict(existing)
                merged.update(payload)
                payload = merged
                break
    keys = _iso_payload_keys(payload)
    if not keys:
        return
    for k in keys:
        data[str(k)] = payload
    _save_iso_payloads(data)


def _find_iso_payload(payloads: dict, key: str):
    if not key:
        return None
    key_str = str(key).strip()
    if not key_str:
        return None
    if key_str in payloads:
        return payloads.get(key_str)
    key_up = key_str.upper()
    for k, v in payloads.items():
        if str(k).strip().upper() == key_up:
            return v
    for v in payloads.values():
        if not isinstance(v, dict):
            continue
        num = v.get("Numero_de_Registro") or v.get("numero")
        if num is not None and str(num).strip() == key_str:
            return v
        bp = v.get("BP") or v.get("bp")
        if bp and str(bp).strip().upper() == key_up:
            return v
    m = re.search(r"(\d+)", key_str)
    if m:
        num_key = m.group(1)
        if num_key in payloads:
            return payloads.get(num_key)
    return None



def _load_fill_r01901_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01901.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontrÃ³ fill_r01901.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01901", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el mÃ³dulo fill_r01901.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _load_fill_r01903_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01903.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontrÃ³ fill_r01903.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01903", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el mÃ³dulo fill_r01903.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _load_fill_r01902_module(codes_dir: Path):
    import importlib.util
    module_path = codes_dir / "fill_r01902.py"
    if not module_path.exists():
        raise FileNotFoundError(f"No se encontrÃ³ fill_r01902.py en {codes_dir}")
    spec = importlib.util.spec_from_file_location("fill_r01902", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("No se pudo cargar el mÃ³dulo fill_r01902.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod


def _r01903_pending_path(iso_root: Path) -> Path:
    pending_dir = ISO_DOCS_ROOT
    return pending_dir / "R019-03-pendientes.json"


def _append_r01903_pending(iso_root: Path, payload: dict, reason=None) -> int:
    pending_path = _r01903_pending_path(iso_root)
    pending_path.parent.mkdir(parents=True, exist_ok=True)
    try:
        items = _load_json_file(pending_path, [], list)
    except Exception:
        items = []
    items.append({
        "created_at": datetime.now().isoformat(),
        "reason": reason or "",
        "payload": payload
    })
    _write_json_file_atomic(pending_path, items, indent=2, ensure_ascii=False)
    return len(items)


def _notifications_path() -> Path:
    NOTIFICATIONS_FILE.parent.mkdir(parents=True, exist_ok=True)
    return NOTIFICATIONS_FILE


def _normalize_user_key(value: str) -> str:
    return str(value or "").strip().lower()


def _destinatarios_path() -> Path:
    return ACTIVITY_CODE_DIR / "config" / "destinatarios.csv"


def _load_destinatarios() -> list:
    path = _destinatarios_path()
    if not path.exists():
        return []
    rows = []
    try:
        with open(path, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if not row or len(row) < 3:
                    continue
                nombre = str(row[0]).strip()
                apellido = str(row[1]).strip()
                email = str(row[2]).strip().lower()
                full = f"{nombre} {apellido}".strip()
                if full or email:
                    rows.append({
                        "nombre": nombre,
                        "apellido": apellido,
                        "email": email,
                        "full": full,
                        "key": _normalize_user_key(full)
                    })
    except Exception:
        return []
    return rows


def _resolve_email_from_name(full_name: str) -> str:
    target = _normalize_user_key(full_name)
    if not target:
        return ""
    for row in _load_destinatarios():
        if row.get("key") == target:
            return row.get("email", "")
    return ""


def _resolve_name_from_email(email: str) -> str:
    if not email:
        return ""
    email_key = email.strip().lower()
    for row in _load_destinatarios():
        if row.get("email", "").lower() == email_key:
            return row.get("full", "")
    return ""


def _resolve_user_email(user: str) -> str:
    if not user:
        return ""
    if "@" in user:
        return user.strip().lower()
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                users = json.load(f) or {}
            if user in users:
                email_val = users[user].get("email") or users[user].get("correo")
                if email_val:
                    return str(email_val).strip().lower()
    except Exception:
        return ""
    return ""


def _load_notifications() -> list:
    path = _notifications_path()
    if not path.exists():
        return []
    try:
        data = _load_json_file(path, [], list) or []
        return data if isinstance(data, list) else []
    except Exception:
        return []


def _save_notifications(items: list) -> None:
    path = _notifications_path()
    _write_json_file_atomic(path, items, indent=2, ensure_ascii=False)


def _resolve_display_name(user: str) -> str:
    if not user:
        return ""
    display_name = user
    email = ""
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, "r", encoding="utf-8") as f:
                users = json.load(f) or {}
            if user in users:
                display_name = users[user].get("display_name", user)
                email = users[user].get("email") or users[user].get("correo") or ""
            else:
                for uname, data in users.items():
                    email = (data.get("email") or data.get("correo") or "").strip().lower()
                    if email and email == user.strip().lower():
                        display_name = data.get("display_name", uname)
                        break
    except Exception:
        pass
    if "@" in user:
        email = user.strip().lower()
    if email:
        name_from_email = _resolve_name_from_email(email)
        if name_from_email:
            return name_from_email
    return display_name or user


def _add_notification(recipient_name: str, ntype: str, message: str, payload=None) -> dict:
    items = _load_notifications()
    recipient_email = ""
    if recipient_name and "@" in recipient_name:
        recipient_email = recipient_name.strip().lower()
    else:
        recipient_email = _resolve_email_from_name(recipient_name)
    entry = {
        "id": str(uuid.uuid4()),
        "recipient": recipient_name,
        "recipient_key": _normalize_user_key(recipient_name),
        "recipient_email": recipient_email,
        "type": ntype,
        "message": message,
        "payload": payload or {},
        "created_at": datetime.now().isoformat(),
        "read": False,
    }
    items.append(entry)
    _save_notifications(items)
    return entry

# --- ISO RESPONSABLES (Destinatarios) ---
@app.route('/api/iso-responsables', methods=['GET'])
def iso_responsables():
    try:
        config_dir = ACTIVITY_CODE_DIR / "config"
        destinatarios_file = config_dir / "destinatarios.csv"
        if not destinatarios_file.exists():
            return jsonify({
                "status": "error",
                "message": f"No se encontrÃ³ destinatarios.csv en {config_dir}"
            }), 404

        names = []
        seen = set()
        with open(destinatarios_file, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            header = next(reader, None)
            for row in reader:
                if not row or len(row) < 2:
                    continue
                nombre = str(row[0]).strip()
                apellido = str(row[1]).strip()
                full = f"{nombre} {apellido}".strip()
                key = full.lower()
                if not full or key in seen:
                    continue
                seen.add(key)
                names.append(full)

        return jsonify({"status": "success", "data": names})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-destinatarios', methods=['GET'])
def iso_destinatarios():
    try:
        rows = _load_destinatarios()
        data = [{"name": r.get("full", ""), "email": r.get("email", "")} for r in rows if r.get("full") or r.get("email")]
        return jsonify({"status": "success", "data": data})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


# --- ISO R019-03 PENDINGS SYNC ---
@app.route('/api/iso-r01903-pending-sync', methods=['POST'])
def iso_r01903_pending_sync():
    pending_path = _r01903_pending_path(ISO_DOCS_ROOT)
    if not pending_path.exists():
        return jsonify({"status": "success", "processed": 0, "remaining": 0})

    try:
        with open(pending_path, "r", encoding="utf-8") as f:
            items = json.load(f) or []
    except Exception as e:
        return jsonify({"status": "error", "message": f"No se pudo leer pendientes: {e}"}), 500

    if not items:
        return jsonify({"status": "success", "processed": 0, "remaining": 0})

    codes_r01903 = iso_root / "Codigos" / "R019-03"
    mod03 = _load_fill_r01903_module(codes_r01903)
    locked_cls = getattr(mod03, "FileLockedError", None)

    processed = 0
    remaining = []
    for idx, item in enumerate(items):
        payload = (item or {}).get("payload") or {}
        try:
            template_path = None
            if hasattr(mod03, "resolve_template_path"):
                template_path = mod03.resolve_template_path(ISO_DOCS_ROOT)
            mod03.generate_r01903(payload, inplace=True, template_path=template_path)
            processed += 1
        except Exception as e:
            if locked_cls and isinstance(e, locked_cls):
                remaining = items[idx:]
                break
            return jsonify({"status": "error", "message": f"Error al sincronizar pendientes: {e}"}), 500

    if remaining:
        _write_json_file_atomic(pending_path, remaining, indent=2, ensure_ascii=False)
    else:
        try:
            pending_path.unlink()
        except Exception:
            pass

    return jsonify({
        "status": "success",
        "processed": processed,
        "remaining": len(remaining)
    })


# --- ISO R019-01 GENERATION (WORD) ---
@app.route('/api/iso-generate-r01901', methods=['POST'])
def iso_generate_r01901():
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or data
    bp = data.get("bp") or payload.get("BP") or payload.get("bp") or "BP-XXXXX"

    iso_root, docs_dir, codes_dir = _iso_r01901_paths()
    if not docs_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontrÃ³ la carpeta R019-01 en {docs_dir}"
        }), 404
    if not codes_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontrÃ³ la carpeta de cÃ³digos R019-01 en {codes_dir}"
        }), 404

    template_payload = _load_iso_payload_template(codes_dir)
    merged = dict(template_payload)
    merged.update(payload or {})
    safe_bp = _normalize_iso_bp(bp)
    merged = _enrich_iso_payload_with_r01901_metadata(merged, docs_dir, safe_bp)
    filename = str(merged.get("R01901_File") or _r01901_filename_from_bp(safe_bp))
    output_path = Path(str(merged.get("R01901_Path") or _resolve_r01901_output_path(docs_dir, safe_bp)))

    if output_path.exists():
        return jsonify({
            "status": "error",
            "message": f"El archivo ya existe: {output_path}",
            "path": str(output_path)
        }), 409

    try:
        mod = _load_fill_r01901_module(codes_dir)
        if not hasattr(mod, "generate_r01901"):
            return jsonify({
                "status": "error",
                "message": "fill_r01901.py no expone generate_r01901"
            }), 500
        template_path = None
        try:
            if hasattr(mod, "resolve_template_path"):
                template_path = mod.resolve_template_path(iso_root / "R019-01")
        except Exception:
            template_path = None
        mod.generate_r01901(merged, output_path, template_path=template_path)
        # Notificar firma pendiente si Responsable de DiseÃƒÂ±o es distinto al usuario actual
        try:
            responsable = (merged.get("Responsable_DiseÃƒÂ±o") or merged.get("Responsable_DiseÃ±o") or "").strip()
            if responsable:
                actor = session.get("user") if session.get("user") else ""
                actor_name = _resolve_display_name(actor)
                if _normalize_user_key(responsable) != _normalize_user_key(actor_name):
                    _add_notification(
                        responsable,
                        "signature",
                        f"Firma pendiente para {safe_bp}",
                        {"bp": safe_bp, "numero": merged.get("Numero_de_Registro") or ""}
                    )
        except Exception:
            pass
        # Actualizar R019-03 (in-place) o dejar pendiente si el archivo estÃ¡ en uso
        pending_info = None
        try:
            codes_r01903 = iso_root / "Codigos" / "R019-03"
            mod03 = _load_fill_r01903_module(codes_r01903)
            if hasattr(mod03, "generate_r01903"):
                template_path = None
                if hasattr(mod03, "resolve_template_path"):
                    template_path = mod03.resolve_template_path(ISO_DOCS_ROOT)
                mod03.generate_r01903(merged, inplace=True, template_path=template_path)
            else:
                raise RuntimeError("fill_r01903.py no expone generate_r01903")
        except Exception as e:
            locked_cls = getattr(mod03, "FileLockedError", None)
            if locked_cls and isinstance(e, locked_cls):
                pending_count = _append_r01903_pending(iso_root, merged, str(e))
                pending_info = {
                    "pending": True,
                    "pending_count": pending_count,
                    "pending_message": "R019-03 en cola (archivo en uso)."
                }
            else:
                return jsonify({"status": "error", "message": f"R019-01 generado, pero fallÃ³ R019-03: {e}"}), 500

        # Crear R019-02 (in-place copia por BP)
        try:
            codes_r01902 = iso_root / "Codigos" / "R019-02"
            mod02 = _load_fill_r01902_module(codes_r01902)
            if hasattr(mod02, "generate_r01902"):
                r01902_dir = ISO_R01902_DOCS_DIR
                r01902_dir.mkdir(parents=True, exist_ok=True)
                r01902_name = f"R019-02 Rev03 - {safe_bp}.xlsm"
                r01902_path = r01902_dir / r01902_name
                if not r01902_path.exists():
                    mod02.generate_r01902(merged, r01902_path)
            else:
                raise RuntimeError("fill_r01902.py no expone generate_r01902")
        except Exception as e:
            return jsonify({"status": "error", "message": f"R019-01 generado, pero fallÃ³ R019-02: {e}"}), 500
        r01904_error = None
        try:
            _generate_r01904(str(safe_bp), allow_empty=True)
        except Exception as e:
            r01904_error = f"No se pudo generar R019-04: {e}"

        try:
            _upsert_iso_payload(merged)
        except Exception:
            pass

        resp = {
            "status": "success",
            "file": filename,
            "path": str(output_path),
            "root": str(docs_dir),
            "bp": safe_bp,
            "payload": merged,
        }
        if pending_info:
            resp.update(pending_info)
        if r01904_error:
            resp["r01904_error"] = r01904_error
        return jsonify(resp)
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500


@app.route('/api/iso-r01901-update', methods=['POST'])
def iso_update_r01901():
    data = request.get_json(silent=True) or {}
    payload = data.get("payload") or {}
    bp = data.get("bp") or payload.get("BP") or payload.get("bp") or "BP-XXXXX"

    iso_root, docs_dir, codes_dir = _iso_r01901_paths()
    if not docs_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontrÃ³ la carpeta R019-01 en {docs_dir}"
        }), 404
    if not codes_dir.exists():
        return jsonify({
            "status": "error",
            "message": f"No se encontrÃ³ la carpeta de cÃ³digos R019-01 en {codes_dir}"
        }), 404

    safe_bp = _normalize_iso_bp(bp)
    payload = _enrich_iso_payload_with_r01901_metadata(payload, docs_dir, safe_bp)
    output_path = _resolve_r01901_output_path(docs_dir, safe_bp, must_exist=True)
    filename = output_path.name
    if not output_path.exists():
        try:
            pending_payload = dict(payload or {})
            _upsert_iso_payload(pending_payload, merge=True)
        except Exception:
            pass
        return jsonify({
            "status": "success",
            "pending": True,
            "message": "Documento R019-01 aún no generado. Firma guardada y se aplicará al generar el archivo.",
            "file": filename,
            "path": str(output_path)
        }), 200

    try:
        mod = _load_fill_r01901_module(codes_dir)
        if not hasattr(mod, "update_r01901"):
            return jsonify({
                "status": "error",
                "message": "fill_r01901.py no expone update_r01901"
            }), 500
        mod.update_r01901(payload, output_path)
        try:
            _upsert_iso_payload(payload, merge=True)
        except Exception:
            pass
        return jsonify({"status": "success", "file": filename, "path": str(output_path), "bp": safe_bp, "payload": payload})
    except Exception as e:
        return jsonify({"status": "error", "message": str(e)}), 500






@app.route('/api/iso-r01901-payload', methods=['GET'])
def iso_r01901_payload():
    key = request.args.get('key') or request.args.get('id') or request.args.get('bp') or ''
    payloads = _load_iso_payloads()
    payload = _find_iso_payload(payloads, key)
    if not payload:
        return jsonify({"status": "error", "message": "No se encontro el registro."}), 404
    return jsonify({"status": "success", "payload": payload})


# --- MANUAL UPLOAD: LIST POs FROM resumen_all ---

@app.route('/api/manual-register-po', methods=['POST'])

def manual_register_po():

    data = request.json

    po_id = data.get('po_id')

    if not po_id:

        return jsonify({"status": "error", "message": "Missing PO ID"}), 400



    if not MANUAL_INGRESOS_FILE.exists():

        return jsonify({"status": "error", "message": "File resumen_all.csv not found"}), 404



    # 1. Read & Filter CSV

    encodings = ['utf-8-sig', 'cp1252', 'latin-1']

    content = None

    for enc in encodings:

        try:

            content = MANUAL_INGRESOS_FILE.read_text(encoding=enc)

            break

        except Exception:

            continue

    

    if content is None:

        return jsonify({"status": "error", "message": "Could not read resumen_all.csv"}), 500



    lines = [ln for ln in content.splitlines() if ln.strip()]

    if not lines:

         return jsonify({"status": "error", "message": "resumen_all is empty"}), 400



    # Validate delimiter

    first_line = lines[0]

    delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','

    

    reader = csv.reader(lines, delimiter=delimiter)

    all_rows = list(reader)

    if not all_rows:

        return jsonify({"status": "error", "message": "No rows in CSV"}), 400



    header = all_rows[0]

    data_rows = all_rows[1:]

    

    target_po_clean = str(po_id).strip().lower()

    

    filtered_rows = []

    

    for i, row in enumerate(data_rows):

        match = False

        val1 = ""

        val2 = ""

        

        # Check col 2 (index 1)

        if len(row) > 1:

            val1 = row[1].strip().lower()

            if target_po_clean in val1: match = True

        

        # Check col 3 (index 2)

        if len(row) > 2:

            val2 = row[2].strip().lower()

            if target_po_clean in val2: match = True

            

        if match:

            filtered_rows.append(row)



    if not filtered_rows:

        return jsonify({"status": "error", "message": "No matching rows found for this PO"}), 404



    # PADRON DE DUPLICADOS (DEBUG MODE)

    print(f"--- Verificando Duplicados para PO {po_id} ---")

    roots_to_scan = [

        BASE_DIR / "P2 - Purchase Order/En Progreso",

        BASE_DIR / "P2 - Purchase Order/Procesado"

    ]

    po_target_str = str(po_id).strip()

    

    dupe_found = None

    for root in roots_to_scan:

        if not root.exists(): 

            print(f"Ruta no existe: {root}")

            continue

            

        print(f"Escaneando raiz: {root.name}")

        

        # Mapeo de nombres amigables

        friendly_name = "Registros Pendientes" if "En Progreso" in root.name else "Historial de PO"

        

        for folder in root.iterdir():

            if not folder.is_dir(): continue

            

            # 1. Nombre Carpeta

            if re.search(rf"\bPO\s*[-_]?\s*{re.escape(po_target_str)}\b", folder.name, re.IGNORECASE):

                dupe_found = friendly_name

                print(f"!!! DUPLICADO ENCONTRADO POR NOMBRE: {dupe_found} ({folder.name})")

                break



            # 2. Contenido CSV

            res_csv = folder / "resumen.csv"

            if res_csv.exists():

                try:

                    c_txt = res_csv.read_text(encoding='latin-1', errors='ignore') 

                    if re.search(rf"\b{re.escape(po_target_str)}\b", c_txt):

                         dupe_found = friendly_name

                         print(f"!!! DUPLICADO ENCONTRADO POR CSV: {dupe_found} ({folder.name})")

                         break

                except Exception as e:

                    print(f"Error leyendo {folder.name}: {e}")

        if dupe_found: break

        

    if dupe_found:

         print(f"RECHAZADO: Duplicado en {dupe_found}")

         return jsonify({"status": "error", "message": f"La PO {po_id} ya existe en {dupe_found}"}), 400

    

    print("--- PO Valida, procediendo ---")



    # 2. Determine Next R Number

    if not IN_PROCESS_DIR.exists():

        IN_PROCESS_DIR.mkdir(parents=True, exist_ok=True)



    max_r = 0

    for item in IN_PROCESS_DIR.iterdir():

        if item.is_dir():

             # Match R number. Pattern: "Registro - R0060 - ..."

            match = re.search(r'Registro\s*-\s*R(\d+)', item.name, re.IGNORECASE)

            if match:

                try:

                    r_num = int(match.group(1))

                    if r_num > max_r: max_r = r_num

                except: pass

    

    next_r = max_r + 1

    # PREVIOUS: folder_name = f"Registro - R{next_r:04d} - Manual PO {po_id.upper()}"

    # REQUESTED FORMAT: "Registro - R0070 - 2026-01-14"

    today_str = datetime.now().strftime("%Y-%m-%d")

    folder_name = f"Registro - R{next_r:04d} - {today_str}"

    new_folder_path = IN_PROCESS_DIR / folder_name

    

    try:

        new_folder_path.mkdir(parents=True, exist_ok=True)

        

        # 3. Write resumen.csv

        out_csv = new_folder_path / "resumen.csv"

        

        # We write simply with utf-8-sig and delimiter ';'

        with open(out_csv, 'w', encoding='utf-8-sig', newline='') as f:

            writer = csv.writer(f, delimiter=';')

            writer.writerow(header)

            writer.writerows(filtered_rows)

            

        # Log it

        actor = session.get('user', 'system')

        log_action(actor, "Manual Register", f"Created {folder_name} with {len(filtered_rows)} rows")



        return jsonify({"status": "success", "message": f"Registro creado: {folder_name}", "folder_name": folder_name})

        

    except Exception as e:

        return jsonify({"status": "error", "message": f"Error creating folder/file: {e}"}), 500



@app.route('/api/run-step2-manual', methods=['POST'])

def run_step2_manual():

    try:

        script_dir = BASE_DIR / "Codigos"

        import sys

        data = request.get_json(silent=True) or {}
        folder_name = data.get("folder_name")

        # Reutilizamos la lÃ³gica de sincronizaciÃ³n en segundo plano: Step 2 (Entrantes) + Step 3 (En Progreso)

        start_sync_thread(source="manual", target_registro=folder_name)

        return jsonify({"status": "success", "message": "SincronizaciÃ³n manual iniciada en segundo plano"})





        

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/manual-po-list')

def manual_po_list():

    import sys

    # Actualizar CSVs (Run Step 1 dry-run)

    try:

        script_dir = BASE_DIR / "Codigos"

        updated = subprocess.run(

            [sys.executable, 'procesar_ingresos.py', '--just-update-csv', '--refresh-excel'],

            cwd=str(script_dir),

            capture_output=True,

            text=True

        )

        if updated.returncode != 0:
            return jsonify({"status": "error", "message": "No se pudo actualizar resumen_all (refresh Excel fall?)", "details": updated.stderr}), 500

    except Exception as e:
         return jsonify({"status": "error", "message": f"Failed to run procesar_ingresos update: {e}"}), 500



    if not MANUAL_INGRESOS_FILE.exists():

        return jsonify([])



    encodings = ['utf-8-sig', 'cp1252', 'latin-1']

    content = None

    for enc in encodings:

        try:

            content = MANUAL_INGRESOS_FILE.read_text(encoding=enc)

            break

        except Exception:

            continue



    if content is None:

        return jsonify({"status": "error", "message": "No se pudo leer resumen_all.csv"}), 500



    # Detect delimiter using first non-empty line

    lines = [ln for ln in content.splitlines() if ln.strip()]

    if not lines:

         return jsonify([])



    first_line = lines[0]

    delimiter = ';' if first_line.count(';') >= first_line.count(',') else ','



    reader = csv.DictReader(StringIO(content), delimiter=delimiter)



    # Aggregation Dict: PO -> { date: '...', count: 0 }

    po_data = {}



    for row in reader:

        if not isinstance(row, dict):

            continue

            

        # 1. Extract PO Number

        obs = (row.get('Observac_PO') or '').strip()

        po_val = row.get('PO') or ''

        

        po_str = None

        # Try to find digits in Observac_PO first, then PO column

        m = re.search(r'(\d+)', obs)

        if m:

            po_str = m.group(1)

        else:

            m2 = re.search(r'(\d+)', po_val)

            if m2:

                po_str = m2.group(1)

                

        if not po_str:

            continue



        # Clean leading zeros

        cleaned_po = po_str.lstrip('0') or po_str

        

        # 2. Extract Date (Try common keys or index based fallback if needed, but DictReader relies on keys)

        # Based on file check: Header is 'Columna_D' for the 4th column

        date_val = (row.get('Columna_D') or '').strip()

        

        # Fallback if 'Columna_D' key missing but 'Fecha' present or generic

        if not date_val:

            date_val = (row.get('Fecha') or '').strip()



        # 3. Aggregate

        if cleaned_po not in po_data:

            po_data[cleaned_po] = {

                "po": cleaned_po,

                "fecha_ingreso": date_val,

                "count": 0

            }

        

        po_data[cleaned_po]["count"] += 1

        # Update date if missing (or overwrite? usually dates are same for a PO block)

        if not po_data[cleaned_po]["fecha_ingreso"] and date_val:

            po_data[cleaned_po]["fecha_ingreso"] = date_val



    # Convert to list and sort

    final_list = list(po_data.values())



    def sort_key(item):

        try:

            return int(item['po'])

        except:

            return item['po']



    final_list.sort(key=sort_key, reverse=True) # Newest POs first usually



    return jsonify(final_list)



# --- API FOR HISTORIAL DETAILS ---

@app.route('/api/historial-po-details/<po_id>')

def get_historial_po_details(po_id):

    if not PROCESSED_PATH.exists():

        return jsonify([])

        

    target_folder = PROCESSED_PATH / po_id

    if not target_folder.exists():

        return jsonify({"status": "error", "message": "PO not found"}), 404

        

    products = []

    

    # Read approval metadata for global info

    app_meta = {}

    meta_path = target_folder / "approval_info.json"

    if meta_path.exists():

        try:

             with open(meta_path, 'r', encoding='utf-8') as f:

                 app_meta = json.load(f)

        except: pass



    # Scan for PDFs

    approved_items = []

    if isinstance(app_meta, dict):

        tmp = app_meta.get('approved_items', [])

        if isinstance(tmp, list):

            approved_items = [str(x) for x in tmp]

            

    for pdf in target_folder.glob('*.pdf'):

        stats = pdf.stat()

        mtime = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y %H:%M')

        

        # Derive status per product

        status_val = app_meta.get('status', 'Pendiente')

        if approved_items:

            status_val = "Aprobado" if pdf.name in approved_items else "Pendiente"

        

        products.append({

            "name": pdf.name,

            "status": status_val,

            "fecha_aprobado": app_meta.get('approved_at', '-'),

            "aprobado_por": app_meta.get('approved_by_name', app_meta.get('approved_by', '-')),

            "path": f"{po_id}/{pdf.name}",

            "size": stats.st_size

        })

        

    return jsonify(products)



@app.route('/api/historial-product-csv-match', methods=['POST'])

def get_historial_product_csv_match():

    data = request.json

    po_id = data.get('po_id')

    product_code = data.get('product_code')

    

    if not po_id or not product_code:

        return jsonify({"status": "error", "message": "Missing params"}), 400

        

    target_folder = PROCESSED_PATH / po_id

    if not target_folder.exists():

         return jsonify({"status": "error", "message": "PO not found"}), 404

         

    # Scan all CSVs in folder recursively

    match_data = None

    all_csvs = list(target_folder.rglob('*.csv'))

    

    # 1. Try Filename Match (Prioritize this to match Client-Side behavior)

    p_name_norm = re.sub(r'\.[a-z0-9]+$', '', product_code, flags=re.IGNORECASE)

    p_name_norm = re.sub(r'[^a-z0-9]', '', p_name_norm).lower()

    

    for csv_file in all_csvs:

        if "aux_matches" in csv_file.name.lower(): continue

        if "resumen" in csv_file.name.lower(): continue

        

        c_name_norm = re.sub(r'\.[a-z0-9]+$', '', csv_file.name, flags=re.IGNORECASE)

        c_name_norm = re.sub(r'[^a-z0-9]', '', c_name_norm).lower()

        

        # Loose match logic

        if p_name_norm and (p_name_norm == c_name_norm or p_name_norm in c_name_norm or c_name_norm in p_name_norm):

            try:

                content = None

                for enc in ['utf-8-sig', 'cp1252', 'latin-1']:

                    try:

                        with open(csv_file, 'r', encoding=enc) as f:

                            content = f.read()

                            break

                    except: continue

                

                if content:

                    from io import StringIO

                    f = StringIO(content)

                    delimiter = ';' if content.count(';') > content.count(',') else ','

                    reader = csv.reader(f, delimiter=delimiter)

                    rows = list(reader)

                    if len(rows) > 1:

                         headers = rows[0]

                         found_row = None

                         

                         # Try to find row matching product_code

                         p_code_search = product_code.strip().lower() if product_code else ""

                         

                         for r in rows[1:]:

                             if not r or len(r) == 0: continue

                             cell_val = r[0].strip().lower()

                             if p_code_search and p_code_search in cell_val:

                                 found_row = r

                                 break

                         

                         # Fallback to first row if not found

                         final_row = found_row if found_row else rows[1]

                         

                         match_data = {

                            "filename": csv_file.name,

                            "headers": headers,

                            "row": final_row

                        }

                         break

            except: continue



    if not match_data:

        # 2. Fallback: Content Scan

        for csv_file in all_csvs:

            if "aux_matches" in csv_file.name.lower(): continue

            if "resumen" in csv_file.name.lower(): continue

            

            try:

                content = None

                for enc in ['utf-8-sig', 'cp1252', 'latin-1']:

                    try:

                        with open(csv_file, 'r', encoding=enc) as f:

                            content = f.read()

                            break

                    except: continue

                    

                if not content: continue

                

                from io import StringIO

                f = StringIO(content)

                delimiter = ';' if content.count(';') > content.count(',') else ','

                reader = csv.reader(f, delimiter=delimiter)

                rows = list(reader)

                

                if len(rows) < 2: continue

                

                headers = rows[0]

                for row in rows[1:]:

                    if not row: continue

                    cell_val = row[0].strip().lower()

                    p_code = product_code.strip().lower()

                    if p_code in cell_val:

                        match_data = {

                            "filename": csv_file.name,

                            "headers": headers,

                            "row": row

                        }

                        break

                if match_data: break

            except: continue

        

    if match_data:

        return jsonify({"status": "success", "data": match_data})

    else:

        return jsonify({"status": "not_found", "message": "No matched CSV data"}), 200





# --- AUXILIAR ROUTES ---



@app.route('/api/auxiliar-indices')

def get_auxiliar_indices():

    files_data = []

    if not AUXILIAR_DIR.exists():

        return jsonify([])

        

    for file in AUXILIAR_DIR.glob('*.csv'):

        if file.is_file():

            stats = file.stat()

            mtime = datetime.fromtimestamp(stats.st_mtime).strftime('%d/%m/%Y %H:%M')

            author = "Oficina Tecnica"

            files_data.append({

                "name": file.name,

                "date": mtime,

                "author": author

            })

    return jsonify(files_data)



@app.route('/api/historial-po')
def get_historial_po():
    try:
        print("[HISTORY DEBUG] Starting get_historial_po...")
        po_list = []
        
        if not PROCESSED_PATH.exists():
            print(f"[HISTORY DEBUG] PROCESSED_PATH does not exist: {PROCESSED_PATH}")
            return jsonify([])

        # Load users for fallback name resolution
        users_map = {}
        if USERS_FILE.exists():
            try:
                with open(USERS_FILE, 'r') as f:
                    users_map = json.load(f)
            except Exception as e:
                print(f"[HISTORY DEBUG] Failed to load users map: {e}")

        # Scan PROCESSED_PATH for directories (POs)
        print(f"[HISTORY DEBUG] Scanning {PROCESSED_PATH}")
        
        for folder in PROCESSED_PATH.iterdir():
            if folder.is_dir():
                try:
                    # Basic Info
                    po_id = folder.name
                    # print(f"[HISTORY DEBUG] Found PO folder: {po_id}")
                    
                    po_data = {
                        "id": po_id,
                        "proveedor": "-", # Will try to fetch
                        "fecha_po": "-",
                        "fecha_ingreso": "-",
                        "fecha_aprobado": "-",
                        "aprobado_por": "-", # Display Name
                        "estado": "Procesado",
                        "counts": {"approved": 0, "total": 0}
                    }
                    
                    # 1. Folder Date (Ingreso approximation)
                    try:
                        stats = folder.stat()
                        # Use st_mtime as it's more reliable on some Windows configs for 'last write'
                        po_data["fecha_ingreso"] = datetime.fromtimestamp(stats.st_ctime).strftime('%d/%m/%Y')
                    except: pass
                    
                    # 2. Read Approval Metadata
                    app_meta = folder / "approval_info.json"
                    if app_meta.exists():
                        try:
                            with open(app_meta, 'r', encoding='utf-8') as f:
                                meta = json.load(f)
                                po_data["fecha_aprobado"] = meta.get("approved_at", "-")
                                po_data["estado"] = meta.get("status", "Procesado")
                                
                                # Author Name Logic
                                saved_name = meta.get("approved_by_name")
                                if saved_name:
                                    po_data["aprobado_por"] = saved_name
                                else:
                                    # Fallback to ID and lookup
                                    uid = meta.get("approved_by", "-")
                                    if uid in users_map:
                                        po_data["aprobado_por"] = users_map[uid].get('display_name', uid)
                                    else:
                                        po_data["aprobado_por"] = uid
                                        
                                # Counts
                                if "counts" in meta:
                                    po_data["counts"] = meta["counts"]
                        except Exception as e_meta:
                            print(f"[HISTORY DEBUG] Error reading metadata for {po_id}: {e_meta}")
                    
                    # 3. Read Provider/Date from resumen.csv (if available)
                    # (Placeholder logic preserved)

                    # Fallback Counts if not in metadata
                    if po_data["counts"]["total"] == 0:
                        pdfs = list(folder.glob('*.pdf'))
                        count_pdfs = len(pdfs)
                        if count_pdfs > 0:
                            po_data["counts"]["total"] = count_pdfs
                            po_data["counts"]["approved"] = count_pdfs # Assume 100% if no metadata
                    
                    po_list.append(po_data)
                except Exception as e_inner:
                    print(f"[HISTORY DEBUG] Error processing folder {folder.name}: {e_inner}")
                    continue

        print(f"[HISTORY DEBUG] Returning {len(po_list)} records.")
        return jsonify(po_list)

    except Exception as e:
        print(f"[HISTORY CRITICAL ERROR]: {str(e)}")
        # Return empty list or error object to prevent frontend crash if possible, or 500
        return jsonify({"error": str(e), "status": "error"}), 500



@app.route('/api/auxiliar-csv/<filename>')

def get_auxiliar_csv(filename):

    filename = os.path.basename(filename)

    target = AUXILIAR_DIR / filename

    

    if not target.exists():

        return jsonify({"status": "error", "message": f"File not found: {filename}"}), 404

        

    try:

        rows = []

        headers = []

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        

        content = None

        used_encoding = 'utf-8'

        

        for enc in encodings:

            try:

                with open(target, 'r', encoding=enc) as f:

                    content = f.read()

                    used_encoding = enc

                    break

            except UnicodeDecodeError:

                continue

                

        if content is None:

             return jsonify({"status": "error", "message": "Could not determine file encoding"}), 500

             

        from io import StringIO

        f = StringIO(content)

        

        try:

            sample = content[:1024]

            sniffer = csv.Sniffer()

            dialect = sniffer.sniff(sample)

        except csv.Error:

            dialect = 'excel' 

            

        if ';' in content.split('\n')[0] or ';' in content.split('\n')[1] if len(content.split('\n')) > 1 else False:

             if content.count(';') > content.count(','):

                 reader = csv.reader(f, delimiter=';')

             else:

                 reader = csv.reader(f, dialect)

        else:

             reader = csv.reader(f, dialect)



        rows_list = list(reader)

        

        if rows_list:

            if len(rows_list) > 1:

                raw_headers = rows_list[1]

                data_rows = rows_list[2:]

                

                valid_indices = [i for i, h in enumerate(raw_headers) if h and h.strip()]

                

            if valid_indices:

                headers = [raw_headers[i].strip() for i in valid_indices]

                rows = []

                for r in data_rows:

                    if not any(cell.strip() for cell in r):

                        continue

                    filtered_row = [r[i] if i < len(r) else "" for i in valid_indices]

                    rows.append(filtered_row)

            else:

                headers = rows_list[0]

                rows = rows_list[1:]

        else:

            headers = rows_list[0]

            rows = [row for row in rows if any(cell.strip() for cell in row)]



        return jsonify({"headers": headers, "rows": rows})

    except Exception as e:

        print(f"Error parsing CSV: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



# --- PLANILLA ROUTES ---



PLA_ROOT = BASE_DIR / "Auxiliares"



def get_planilla_folder(context_filename):

    stem = Path(context_filename).stem

    code = stem.split('(')[0].strip() if '(' in stem else stem.split(' ')[0]

    

    family_folder = None

    for item in PLA_ROOT.iterdir():

        if item.is_dir() and code in item.name and "Registro Planillas" in item.name:

            family_folder = item

            break

            

    if not family_folder:

        for item in PLA_ROOT.iterdir():

            if item.is_dir() and code in item.name:

                family_folder = item

                break

    

    if not family_folder:

        return PLA_ROOT / stem



    target_specific = family_folder / stem

    if not target_specific.exists():

        try:

            target_specific.mkdir(parents=True, exist_ok=True)

        except: pass

        

    return target_specific



@app.route('/api/planilla/list/<filename>')

def list_planillas(filename):

    filename = secure_filename(filename) 

    filename = os.path.basename(filename)

    

    folder = get_planilla_folder(filename)

    

    files = []

    if folder.exists():

        for f in folder.iterdir():

            if f.is_file():

                files.append(f.name)

                

    return jsonify({"folder": folder.name, "files": files})



@app.route('/api/planilla/upload', methods=['POST'])

def upload_planilla():

    if 'file' not in request.files or 'context' not in request.form:

        return jsonify({"status": "error", "message": "Missing file or context"}), 400

        

    file = request.files['file']

    context = request.form['context'] 

    

    if file.filename == '':

        return jsonify({"status": "error", "message": "No selected file"}), 400

        

    folder = get_planilla_folder(os.path.basename(context))

    if not folder.exists():

        try:

            folder.mkdir(parents=True, exist_ok=True)

        except Exception as e:

             return jsonify({"status": "error", "message": f"Could not create directory: {str(e)}"}), 500

             

    filename = secure_filename(file.filename)

    try:

        file.save(folder / filename)

        return jsonify({"status": "success", "message": "Planilla uploaded successfully"})

    except Exception as e:

        return jsonify({"status": "error", "message": str(e)}), 500



@app.route('/api/planilla/search-product', methods=['POST'])

def search_product_planilla():

    data = request.json

    context = data.get('context')

    query = data.get('query', '').lower()

    

    if not context or not query:

        return jsonify([])

        

    target = AUXILIAR_DIR / os.path.basename(context)

    if not target.exists():

        return jsonify([])

        

    results = []

    try:

        content = None

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        for enc in encodings:

            try:

                with open(target, 'r', encoding=enc) as f:

                    content = f.read()

                    break

            except: continue

            

        if content:

            from io import StringIO

            f = StringIO(content)

            delimiter = ';' if content.count(';') > content.count(',') else ','

            reader = csv.reader(f, delimiter=delimiter)

            rows = list(reader)

            

            if len(rows) > 1:

                headers = rows[1]

                data_rows = rows[2:]

                

                prod_idx = -1

                for i, h in enumerate(headers):

                    if 'PRODUCTO' in h.upper() or 'CODIGO' in h.upper() or 'ITEM' in h.upper():

                        prod_idx = i

                        break

                

                if prod_idx == -1: prod_idx = 1

                

                for row in data_rows:

                    if len(row) > prod_idx:

                        val = row[prod_idx].strip()

                        if query in val.lower():

                            if val not in results:

                                results.append(val)

                                if len(results) > 10: break

    except: pass

    

    return jsonify(results)



@app.route('/api/planilla/generate-pdf', methods=['POST'])

def generate_planilla_pdf():

    data = request.json

    product = data.get('product')

    # Validated, return URL to download

    # STUB RECOVERY

    return jsonify({"status": "error", "message": "Funcionalidad de PDF en mantenimiento"})



# --- ADVANCED AUXILIAR LOOKUP (PO LINKED) ---



@app.route('/api/product-aux-details', methods=['POST'])

def get_product_aux_details():

    try:

        data = request.json

        po_id = data.get('po_id')

        product_code_a = data.get('product_code_a')

        

        try:

            with open(os.path.join(SCRIPT_DIR, 'debug_backend.txt'), 'a') as df:

                df.write(f"\\n[{datetime.now()}] REQ: po={po_id}, code={product_code_a}\\n")

        except: pass



        print(f"DEBUG_AUX: START get_product_aux_details. po_id={po_id}, product={product_code_a}")



        if not po_id or not product_code_a:

            return jsonify({"status": "error", "message": "Missing params"}), 400



        po_path = DATA_DIR / po_id

        if not po_path.exists():

             # Check PROCESSED

             po_path = PROCESSED_PATH / po_id

             if not po_path.exists():

                return jsonify({"status": "error", "message": "PO not found"}), 404



        # 1. Find aux_matches.csv

        # 1. Find aux_matches.csv (More robust search)

        aux_matches_path = None

        for f in po_path.rglob('aux_matches.csv'):

            aux_matches_path = f

            break

            

        if not aux_matches_path or not aux_matches_path.exists():

             return jsonify({"status": "error", "message": "aux_matches.csv not found in PO"}), 404



        target_c = None

        target_d = None

        

        encodings = ['utf-8-sig', 'cp1252', 'latin-1']

        content = None

        for enc in encodings:

            try:

                with open(aux_matches_path, 'r', encoding=enc) as f:

                    content = f.read()

                    break

            except: continue

            

        if not content: return jsonify({"status": "error", "message": "Could not read aux_matches"}), 500

        

        from io import StringIO

        f = StringIO(content)

        delimiter = ';' if content.count(';') > content.count(',') else ','

        reader = csv.reader(f, delimiter=delimiter)

        

        product_code_clean = product_code_a.strip()
        norm_code = re.sub(r"[^A-Z0-9]", "", product_code_clean.upper())

        def load_precomputed_matches():
            aux_top_path = None
            for f in po_path.rglob('aux_top_matches.json'):
                aux_top_path = f
                break
            if not aux_top_path or not aux_top_path.exists():
                return None
            try:
                with open(aux_top_path, 'r', encoding='utf-8') as f:
                    data_top = json.load(f)
            except Exception:
                return None
            items = None
            if isinstance(data_top, dict):
                items = data_top.get('items') or data_top.get('matches') or data_top.get('data')
            if isinstance(items, dict):
                entry = items.get(norm_code) or items.get(product_code_clean) or items.get(product_code_clean.upper())
                if isinstance(entry, dict):
                    return entry.get('matches') or []
                if isinstance(entry, list):
                    return entry
            return None

        def read_aux_row(filename, match_val):
            if not filename or not match_val:
                return [], []
            target = AUXILIAR_DIR / filename
            if not target.exists():
                return [], []
            content_aux = None
            for enc in encodings:
                try:
                    with open(target, 'r', encoding=enc) as f:
                        content_aux = f.read()
                        break
                except:
                    continue
            if not content_aux:
                return [], []
            f_aux = StringIO(content_aux)
            delimiter_aux = ';' if content_aux.count(';') > content_aux.count(',') else ','
            reader_aux = csv.reader(f_aux, delimiter=delimiter_aux)
            rows_aux = list(reader_aux)
            headers = []
            row_data = []
            if len(rows_aux) > 1:
                headers = rows_aux[1][1:]
                for row in rows_aux:
                    if len(row) < 2:
                        continue
                    if row[1].strip() == match_val:
                        row_data = row[1:]
                        break
            return headers, row_data

        def enrich_match_rows(matches):
            if not matches:
                return matches
            for m in matches:
                if m.get('headers') and m.get('row'):
                    continue
                filename = (m.get('filename') or '').strip()
                match_val = (m.get('match_val') or '').strip()
                headers, row = read_aux_row(filename, match_val)
                m['headers'] = headers or []
                m['row'] = row or []
            return matches

        def collect_candidates(product_code_clean_local):
            from difflib import SequenceMatcher
            from io import StringIO
            candidates = []
            seen_rows = set()
            if AUXILIAR_DIR.exists():
                for item in AUXILIAR_DIR.iterdir():
                    if item.is_file() and item.suffix.lower() == '.csv':
                        try:
                            content_bf = None
                            for enc in ['utf-8-sig', 'cp1252', 'latin-1']:
                                try:
                                    with open(item, 'r', encoding=enc) as f:
                                        content_bf = f.read()
                                        break
                                except:
                                    continue
                            if not content_bf:
                                continue

                            f_bf = StringIO(content_bf)
                            delim_bf = ';' if content_bf.count(';') > content_bf.count(',') else ','
                            reader_bf = csv.reader(f_bf, delimiter=delim_bf)
                            rows_bf = list(reader_bf)

                            if len(rows_bf) > 1:
                                for r_idx, row in enumerate(rows_bf):
                                    if len(row) > 1:
                                        cell_val = row[1].strip()
                                        if not cell_val:
                                            continue

                                        score = 0
                                        if cell_val == product_code_clean_local:
                                            score = 100
                                        else:
                                            if abs(len(cell_val) - len(product_code_clean_local)) < 20:
                                                sm_ratio = SequenceMatcher(None, product_code_clean_local, cell_val).ratio()
                                                score = sm_ratio * 100
                                                import re
                                                def get_main_number(s):
                                                    nums = re.findall(r'\d+', s)
                                                    if not nums:
                                                        return None
                                                    return max(nums, key=len)
                                                n1 = get_main_number(product_code_clean_local)
                                                n2 = get_main_number(cell_val)
                                                penalty = 0
                                                if n1 and n2 and n1 != n2:
                                                    penalty += 50
                                                s_clean = product_code_clean_local.strip().lower()
                                                c_clean = cell_val.strip().lower()
                                                if len(product_code_clean_local) >= 3 and len(cell_val) >= 3:
                                                    if product_code_clean_local[:3] != cell_val[:3]:
                                                        penalty += 5
                                                prefix_same = False
                                                if '-' in product_code_clean_local and '-' in cell_val:
                                                    prefix_a = product_code_clean_local.rsplit('-', 1)[0]
                                                    prefix_b = cell_val.rsplit('-', 1)[0]
                                                    if prefix_a != prefix_b:
                                                        penalty += 20
                                                    else:
                                                        prefix_same = True
                                                        if product_code_clean_local != cell_val:
                                                            penalty += 2
                                                else:
                                                    if product_code_clean_local != cell_val:
                                                        penalty += 20
                                                if c_clean not in s_clean and s_clean not in c_clean:
                                                    penalty += 5 if prefix_same else 15
                                                score -= penalty
                                                if score < 0:
                                                    score = 0
                                        if score >= 65:
                                            unique_key = f"{item.name}_{r_idx}"
                                            if unique_key in seen_rows:
                                                continue
                                            seen_rows.add(unique_key)

                                            raw_h = rows_bf[1] if len(rows_bf) > 1 else []
                                            final_h = raw_h[1:] if len(raw_h) > 1 else raw_h
                                            row_d = row[1:]

                                            candidates.append({
                                                "score": score,
                                                "headers": final_h,
                                                "row": row_d,
                                                "filename": item.name,
                                                "match_val": cell_val
                                            })
                        except Exception as e:
                            print(f"Error scanning {item.name}: {e}")
                            continue

            candidates.sort(key=lambda x: x['score'], reverse=True)
            return candidates[:3]

        

        for i, row in enumerate(reader):

            if len(row) < 4: continue

            row0 = row[0].strip()
            row3 = row[3].strip()
            if row0 == product_code_clean or row3 == product_code_clean:

                target_c = row[2].strip() # Filename base

                target_d = row[3].strip() # Search Key

                break

                


        if not target_c:

             # Fallback: Brute force search in AUXILIAR_DIR

             print(f"Fallback Search for {product_code_clean} in {AUXILIAR_DIR}")

             try:
                with open(os.path.join(SCRIPT_DIR, 'debug_backend.txt'), 'a') as df:
                    df.write(f"Fallback Search Start for {product_code_clean} in {AUXILIAR_DIR}\n")

             except: pass

             precomputed_matches = load_precomputed_matches()
             top_matches = precomputed_matches if precomputed_matches is not None else collect_candidates(product_code_clean)

             if top_matches:

                 top_matches = enrich_match_rows(top_matches)

                 print(f"Returning top {len(top_matches)} matches")

                 return jsonify({

                     "status": "success",

                     "matches": top_matches

                 })


             return jsonify({"status": "not_found", "message": "No match found"}), 200


        # 2. Find Aux File (Smart Match)
        # 2. Find Aux File (Smart Match)

        target_c_base = os.path.splitext(target_c)[0]

        

        found_path = None

        

        # A. Try exact construction first

        try_name_1 = f"{target_c_base}_Auxiliar.csv"

        path_1 = AUXILIAR_DIR / try_name_1

        if path_1.exists():

            found_path = path_1

            aux_filename = try_name_1

            

        # B. Smart Match: Use text inside parentheses

        if not found_path and AUXILIAR_DIR.exists():

            paren_match = re.search(r'\((.*?)\)', target_c_base)

            if paren_match:

                key_text = paren_match.group(1).lower().strip() # e.g. "inspeccion de especiales"

                if len(key_text) > 3: # Safety check

                    for item in AUXILIAR_DIR.iterdir():

                        if item.is_file() and item.suffix.lower() == '.csv':

                            if key_text in item.name.lower():

                                found_path = item

                                aux_filename = item.name

                                break

        

        # C. Fallback: Loose partial match of the base name

        if not found_path and AUXILIAR_DIR.exists():

             search_base = target_c_base.lower().replace('_auxiliar','').strip()

             for item in AUXILIAR_DIR.iterdir():

                 if item.is_file() and item.suffix.lower() == '.csv':

                     if search_base in item.name.lower():

                         found_path = item

                         aux_filename = item.name

                         break

                         

        if found_path:

             aux_file_path = found_path

        else:

             return jsonify({"status": "error", "message": f"Aux file match not found for: {target_c}"}), 404

             

        # 3. Read Aux File and find Target D in Col B

        row_data = None

        final_headers = None

        

        content_aux = None

        for enc in encodings:

            try:

                with open(aux_file_path, 'r', encoding=enc) as f:

                    content_aux = f.read()

                    break

            except: continue

            

        if not content_aux: return jsonify({"status": "error", "message": "Could not read aux file"}), 500

        

        f_aux = StringIO(content_aux)

        delimiter_aux = ';' if content_aux.count(';') > content_aux.count(',') else ','

        reader_aux = csv.reader(f_aux, delimiter=delimiter_aux)

        rows_aux = list(reader_aux)

        

        if len(rows_aux) > 1:

            # User Request: Use Row 2 (index 1) as header

            # User Request: Skip Col 1 (index 0)

            

            raw_headers = rows_aux[1]

            # Slicing: Skip first column

            final_headers = raw_headers[1:]

            

            for row in rows_aux:

                if len(row) < 2: continue

                # Match Target D in Col B (index 1) - Search remains in original structure

                if row[1].strip() == target_d:

                    # Return data skipping first column

                    row_data = row[1:]

                    break

                    

        precomputed_matches = load_precomputed_matches()
        matches = precomputed_matches if precomputed_matches is not None else collect_candidates(product_code_clean)
        if matches:
            matches = enrich_match_rows(matches)
        else:
            matches = []

        if row_data:
            direct_entry = {
                "score": 100,
                "headers": final_headers,
                "row": row_data,
                "filename": aux_filename,
                "match_val": target_d
            }
            merged = [direct_entry]
            seen_keys = {(direct_entry["filename"], direct_entry.get("match_val", ""))}
            for cand in matches:
                key = (cand.get("filename"), cand.get("match_val", ""))
                if key in seen_keys:
                    continue
                merged.append(cand)
                seen_keys.add(key)
            merged.sort(key=lambda x: x.get("score", 0), reverse=True)
            merged = merged[:3]

            return jsonify({
                "status": "success",
                "headers": final_headers,
                "row": row_data,
                "filename": aux_filename,
                "matches": merged
            })

        else:
            if matches:
                return jsonify({"status": "success", "matches": matches})
            return jsonify({"status": "not_found", "message": "No match in aux file"}), 200



    except Exception as e:

        print(f"Aux Lookup Error: {e}")

        return jsonify({"status": "error", "message": str(e)}), 500



# --- API FOR SAVING AUX TARGET ---

@app.route('/api/save-aux-target', methods=['POST'])

def save_aux_target():

    try:

        data = request.json

        po_id = data.get('po_id')

        product_code = data.get('product_code')

        target = data.get('target')



        if not po_id or not product_code:

            return jsonify({'status': 'error', 'message': 'Missing fields'}), 400



        # Create/Update Metadata in PO Folder

        po_folder = DATA_DIR / po_id

        if not po_folder.exists():

            # Try PROCESSED? No, usually in process

            po_folder = PROCESSED_PATH / po_id

            if not po_folder.exists():

                return jsonify({'status': 'error', 'message': 'PO folder not found'}), 404



        target_file = po_folder / "aux_targets.json"

        

        current_data = {}

        if target_file.exists():

            try:

                with open(target_file, 'r', encoding='utf-8') as f:

                    current_data = json.load(f)

            except: pass



        current_data[product_code] = target

        

        _write_json_file_atomic(target_file, current_data, indent=4, ensure_ascii=False)

            

        return jsonify({'status': 'success'})

    except Exception as e:

        print(f"Save Target Error: {e}")

        return jsonify({'status': 'error', 'message': str(e)}), 500





# --- EXCEL UPDATE LOGIC ---

def process_excel_updates(po_folder, approved_items, po_id):

    """

    Updates the auxiliary Excel files (COPIES) based on approved items and saved targets.

    """

    from openpyxl.styles import Font, Alignment



    targets_file = po_folder / "aux_targets.json"

    matches_file = po_folder / "csv_Auxiliar/aux_matches.csv"

    

    targets = {}

    if not targets_file.exists():

        print(f"[EXCEL] No aux_targets.json for {po_id}; using automatic match fallback.")

    else:

        try:

            with open(targets_file, 'r', encoding='utf-8') as f:

                targets = json.load(f)

        except Exception as e:

            print(f"[EXCEL] Error reading targets: {e}. Using automatic match fallback.")



    # Load Matches (to resolve Type/Filename for New Lines)

    matches_map = {} # Code -> {Type, File...}

    if matches_file.exists():

        try:

            with open(matches_file, 'r', encoding='utf-8') as f:

                reader = csv.DictReader(f, delimiter=';')

                for row in reader:

                    matches_map[row['Codigo']] = row

        except Exception as e:

             print(f"[EXCEL] Error reading matches: {e}")



    # AUX_COPIES_DIR = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\Registro Planillas R016-01")
    AUX_COPIES_DIR = BASE_DIR / "Auxiliares/Registro Planillas R016-01"

    

    for prod_code in approved_items:

        # Resolve CSV Path & Real Code first

        clean_name = prod_code

        if clean_name.lower().endswith('.pdf'):

            clean_name = clean_name[:-4]

            

        csv_path = po_folder / f"csv/{clean_name}.csv"

        if not csv_path.exists():

             csv_path = po_folder / f"csv/{prod_code}.csv"

             

        if not csv_path.exists():

            print(f"[EXCEL] CSV not found for {prod_code}")

            continue



        row_data = None

        real_code = clean_name

        try:

            with open(csv_path, 'r', encoding='utf-8') as f:

                reader = csv.reader(f, delimiter=';')

                rows = list(reader)

                if len(rows) > 1:

                    row_data = rows[1]

                    if row_data: real_code = row_data[0]

        except Exception as e:

            print(f"[EXCEL] Error reading CSV {csv_path}: {e}")

            continue

            

        if not row_data: continue



        # Resolve Clean Code for Mapping (Fuzzy Match against Matches Map)

        # CSV might have "B&P-040-51310..." but Map has "51310"

        if real_code not in matches_map:

            sorted_keys = sorted(matches_map.keys(), key=len, reverse=True)

            for k in sorted_keys:

                if k in real_code:

                    print(f"[EXCEL] Fuzzy matched CSV code '{real_code}' to Map key '{k}'")

                    real_code = k

                    break



        # Resolve Target Action

        action_val = targets.get(prod_code)

        

        # --- FALLBACK ---

        if not action_val:

            match_info = matches_map.get(real_code) # Key by Code

            score = 0
            if cell_val == product_code_clean_local:
                score = 100
            else:
                if abs(len(cell_val) - len(product_code_clean_local)) < 20:
                    sm_ratio = SequenceMatcher(None, product_code_clean_local, cell_val).ratio()
                    score = sm_ratio * 100
                    import re
                    def get_main_number(s):
                        nums = re.findall(r'\d+', s)
                        if not nums:
                            return None
                        return max(nums, key=len)
                    n1 = get_main_number(product_code_clean_local)
                    n2 = get_main_number(cell_val)
                    penalty = 0
                    if n1 and n2 and n1 != n2:
                        penalty += 50
                    s_clean = product_code_clean_local.strip().lower()
                    c_clean = cell_val.strip().lower()
                    if len(product_code_clean_local) >= 3 and len(cell_val) >= 3:
                        if product_code_clean_local[:3] != cell_val[:3]:
                            penalty += 5
                    if '-' in product_code_clean_local and '-' in cell_val:
                        prefix_a = product_code_clean_local.rsplit('-', 1)[0]
                        prefix_b = cell_val.rsplit('-', 1)[0]
                        if prefix_a != prefix_b:
                            penalty += 20
                        else:
                            if product_code_clean_local != cell_val:
                                penalty += 5
                    else:
                        if product_code_clean_local != cell_val:
                            penalty += 20
                    if c_clean not in s_clean and s_clean not in c_clean:
                        penalty += 15
                    score -= penalty
                    if score < 0:
                        score = 0

            try: 

                score = float(match_info.get('Score', 0))

            except: pass

            

            if match_info and score >= 99.0:

                fname = match_info.get('Archivo_Auxiliar', '')

                if fname:

                    action_val = f"{fname}_0"

                    print(f"[EXCEL] Auto-default to Existing for {real_code} (Score {score})")

            

            if not action_val:

                action_val = "new"

                print(f"[EXCEL] Auto-default to New Line for {real_code}")

        

        # Determine Target File

        target_filename = None

        

        match_info = matches_map.get(real_code)

        if hasattr(match_info, 'get'):

             target_filename = match_info.get('Archivo_Auxiliar')



        if action_val == "new":

            # Direct use of identified file

            pass 

        else:

            # action_val can be either:
            # - a filename_with_optional_row (old format)
            # - a selected match code (new dropdown format)
            # Only override target file when action resembles filename metadata.
            try:

                parts = action_val.rsplit('_', 1)
                filename_from_val = parts[0]
                has_row_suffix = len(parts) > 1 and parts[1].isdigit()
                looks_like_file = any(tok in filename_from_val for tok in ["R016-01", ".csv", ".xls", ".xlsx", ".xlsm"])

                if has_row_suffix and filename_from_val and looks_like_file:

                    target_filename = filename_from_val

            except:

                pass



        if not target_filename:

            print(f"[EXCEL] No target file (Archivo_Auxiliar) found for '{real_code}'. Skipping.")

            continue



        # Clean filename (sometimes it might have paths? verify step3 output)

        if "\\" in target_filename or "/" in target_filename:

             target_filename = Path(target_filename).name



        # Resolve Real Excel File from CSV name (e.g. remove _Auxiliar.csv)

        if "_Auxiliar.csv" in target_filename:

            target_filename = target_filename.replace("_Auxiliar.csv", "")

        elif target_filename.lower().endswith(".csv"):

             target_filename = target_filename[:-4]

             

        # Search for .xlsx or .xlsm match

        found_excel = None

        for ext in ["", ".xlsx", ".xlsm", ".xls"]:

            p = AUX_COPIES_DIR / (target_filename + ext)

            if p.exists():

                found_excel = p

                break

        

        if not found_excel:

             print(f"[EXCEL] Excel file not found for stem '{target_filename}'. Skipping.")

             continue

             

        excel_path = found_excel

            

        try:

            is_xlsm = excel_path.suffix.lower() == '.xlsm'
            workbook_buffer = None

            

            # Load via binary stream to handle UNC/Network paths better

            wb = None

            try:

                with open(excel_path, 'rb') as f_stream:
                    workbook_buffer = io.BytesIO(f_stream.read())

                wb = _load_workbook_quietly(workbook_buffer, keep_vba=is_xlsm)

            except Exception as e_load:

                # If failed (likely password protected), try with 'bpb'

                print(f"[EXCEL] Standard load failed ({e_load}). Trying with password 'bpb'...")

                try:

                    import msoffcrypto

                    import io

                    

                    decrypted = io.BytesIO()

                    with open(excel_path, 'rb') as f_enc:

                        office_file = msoffcrypto.OfficeFile(f_enc)

                        office_file.load_key(password='bpb')

                        office_file.decrypt(decrypted)

                    

                    decrypted.seek(0)
                    workbook_buffer = decrypted
                    wb = _load_workbook_quietly(workbook_buffer, keep_vba=is_xlsm)

                    print(f"[EXCEL] Decryption successful.")

                except ImportError:

                    print("[EXCEL] CRITICAL: File is password protected. Please run in cmd: pip install msoffcrypto-tool")

                    raise e_load

                except Exception as e_dec:

                    print(f"[EXCEL] Decryption failed: {e_dec}")

                    raise e_load

            

            # Select "Auxiliar" Sheet

            ws = None

            if "Auxiliar" in wb.sheetnames:

                ws = wb["Auxiliar"]

            elif "Aux" in wb.sheetnames:

                ws = wb["Aux"]

            else:

                # Fallback to active but warn

                print(f"[EXCEL] Sheet 'Auxiliar' not found in {target_filename}, using active.")

                ws = wb.active



            START_ROW = 3 # User requested to skip rows 1 & 2

            

            found_row = None

            max_num = 0

            

            # Scan for Existing & Max Number

            # Iterate from START_ROW

            for r in range(START_ROW, ws.max_row + 1):

                # Col 1: Numero

                num_cell = ws.cell(row=r, column=1).value

                if isinstance(num_cell, (int, float)):

                    if num_cell > max_num: max_num = int(num_cell)

                

                if action_val != "new":

                    # Check Code in Col 2 (B)

                    code_cell = ws.cell(row=r, column=2).value

                    if code_cell and str(code_cell).strip() == prod_code:

                        found_row = r

                        # Don't break immediately if we need max_num? 

                        # Only need max_num for NEW lines. For existing, we replace.

                        break

            

            # Logic Application

            if action_val == "new":

                # Determine Next Num

                next_num = max_num + 1

                

                # Append Row

                # [Number] + [Code, Data...]

                final_row_values = [next_num] + row_data

                ws.append(final_row_values)

                

                # Apply Styles to the appended row (last row) (and fix data type if needed)

                last_row = ws.max_row

                for c_idx, val in enumerate(final_row_values, start=1):

                    cell = ws.cell(row=last_row, column=c_idx)

                    # Col 1 is Num, Col 2 is Code

                    if c_idx == 2:

                        cell.font = Font(bold=True)

                    elif c_idx > 2:

                        cell.alignment = Alignment(horizontal='right')

                        

                print(f"[EXCEL] Appended {real_code} to {target_filename} (No. {next_num})")

                

            elif found_row:

                # Update Existing Row

                # Skip Col 1 (Numero), Update from Col 2 match row_data

                for c_idx, val in enumerate(row_data, start=2): # Start at Col 2

                     cell = ws.cell(row=found_row, column=c_idx)

                     cell.value = val

                     # Apply Styles

                     if c_idx == 2:

                         cell.font = Font(bold=True)

                     else:

                         cell.alignment = Alignment(horizontal='right')

                         

                print(f"[EXCEL] Updated {real_code} in {target_filename} at row {found_row}")

                

            else:

                 # Existing requested but not found -> Append?

                 print(f"[EXCEL] Code {real_code} not found for update, appending new.")

                 next_num = max_num + 1

                 final_row_values = [next_num] + row_data

                 ws.append(final_row_values)

                 # Apply Styles fallback

                 last_row = ws.max_row

                 for c_idx, val in enumerate(final_row_values, start=1):

                    cell = ws.cell(row=last_row, column=c_idx)

                    if c_idx == 2:

                        cell.font = Font(bold=True)

                    elif c_idx > 2:

                        cell.alignment = Alignment(horizontal='right')

                 

                 print(f"[EXCEL] Appended (fallback) {real_code} to {target_filename}")



            save_path_str = str(excel_path)

            print(f"[EXCEL] Saving workbook to: {save_path_str}")

            wb.save(save_path_str)

            print(f"[EXCEL] Save operation completed.")

            

        except Exception as e:

            print(f"[EXCEL] Failed to process {target_filename}: {e}")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            if workbook_buffer is not None:
                try:
                    workbook_buffer.close()
                except Exception:
                    pass





# --- EXCEL UPDATE LOGIC V2 (SEARCH FIX + STYLES + CRYPTO) ---

def process_excel_updates_v2(po_folder, approved_items, po_id):
    """
    Updates the auxiliary Excel files (COPIES) based on approved items and saved targets.
    V2 Optimized: Batches updates per file to avoid opening/saving multiple times.
    """
    from openpyxl.styles import Font, Alignment
    import json
    import csv
    import openpyxl
    from pathlib import Path
    import msoffcrypto
    import io

    targets_file = po_folder / "aux_targets.json"
    matches_file = po_folder / "csv_Auxiliar/aux_matches.csv"
    
    targets = {}

    if not targets_file.exists():
        print(f"[EXCEL] No aux_targets.json for {po_id}; using automatic match fallback.", flush=True)
    else:
        try:
            with open(targets_file, 'r', encoding='utf-8') as f:
                targets = json.load(f)
        except Exception as e:
            print(f"[EXCEL] Error reading targets: {e}. Using automatic match fallback.")

    # Load Matches
    matches_map = {} 
    if matches_file.exists():
        try:
            with open(matches_file, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f, delimiter=';')
                for row in reader:
                    matches_map[row['Codigo']] = row
        except Exception as e:
             print(f"[EXCEL] Error reading matches: {e}")

    # AUX_COPIES_DIR = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\Registro Planillas R016-01")
    AUX_COPIES_DIR = BASE_DIR / "Auxiliares/Registro Planillas R016-01"
    
    # --- Phase 1: Group Updates by Target File ---
    updates_by_file = {} # { "filename": [ {item_data}, ... ] }

    for prod_code in approved_items:
        # Resolve CSV Path & Real Code
        clean_name = prod_code
        if clean_name.lower().endswith('.pdf'):
            clean_name = clean_name[:-4]
            
        csv_path = po_folder / f"csv/{clean_name}.csv"
        if not csv_path.exists():
             csv_path = po_folder / f"csv/{prod_code}.csv"
             
        if not csv_path.exists():
            print(f"[EXCEL] CSV not found for {prod_code}")
            continue

        row_data = None
        real_code = clean_name
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f, delimiter=';')
                rows = list(reader)
                if len(rows) > 1:
                    row_data = rows[1]
                    if row_data: real_code = row_data[0]
        except Exception as e:
            print(f"[EXCEL] Error reading CSV {csv_path}: {e}")
            continue
            
        if not row_data: continue

        # Capture raw code
        csv_raw_code = real_code

        # Safety: avoid cross-product writes when an item points to the wrong CSV row.
        expected_token = str(prod_code or "")
        if expected_token.lower().endswith('.pdf'):
            expected_token = expected_token[:-4]
        expected_token = re.sub(r"^\(?PO\d+\)?[-_\s]*", "", expected_token, flags=re.IGNORECASE)
        expected_token = re.sub(r"[\s_-]*REV[.\s_-]*[A-Z0-9]+$", "", expected_token, flags=re.IGNORECASE)
        expected_norm = re.sub(r"[^A-Z0-9]", "", expected_token.upper())
        raw_norm = re.sub(r"[^A-Z0-9]", "", str(csv_raw_code or "").upper())
        if expected_norm and raw_norm and expected_norm not in raw_norm:
            print(f"[EXCEL] WARNING: CSV mismatch for {prod_code}. CSV code={csv_raw_code}. Se omite para evitar cruce.")
            continue

        # Fuzzy Match Resolution
        if real_code not in matches_map:
            sorted_keys = sorted(matches_map.keys(), key=len, reverse=True)
            for k in sorted_keys:
                if k in real_code:
                    # print(f"[EXCEL] Fuzzy matched CSV code '{real_code}' to Map key '{k}'")
                    real_code = k
                    break

        # Resolve Target Action
        action_val = targets.get(prod_code)
        if not action_val and csv_raw_code:
             action_val = targets.get(csv_raw_code)
        if not action_val:
             action_val = targets.get(clean_name)
        
        # Fallback Logic
        if not action_val:
            match_info = matches_map.get(real_code)
            score = 0
            try: score = float(match_info.get('Score', 0) if match_info else 0)
            except: pass
            
            if match_info and score >= 99.0:
                fname = match_info.get('Archivo_Auxiliar', '')
                if fname:
                    action_val = f"{fname}_0"
            
            if not action_val:
                action_val = "new"

        # Determine Target Filename & Forced Row
        target_filename = None
        forced_row = None
        
        match_info = matches_map.get(real_code)
        if hasattr(match_info, 'get'):
             target_filename = match_info.get('Archivo_Auxiliar')

        if action_val != "new":
            try:
                parts = action_val.rsplit('_', 1)
                filename_from_val = parts[0]
                has_row_suffix = len(parts) > 1 and parts[1].isdigit()
                looks_like_file = any(tok in filename_from_val for tok in ["R016-01", ".csv", ".xls", ".xlsx", ".xlsm"])

                if has_row_suffix and filename_from_val and looks_like_file:
                    target_filename = filename_from_val
                    parsed_row = int(parts[1])
                    if parsed_row > 0:
                        forced_row = parsed_row
            except:
                pass

        if not target_filename:
            print(f"[EXCEL] No target file found for '{real_code}'. Skipping.")
            continue

        # Clean filename
        if "\\" in target_filename or "/" in target_filename:
             target_filename = Path(target_filename).name
        if "_Auxiliar.csv" in target_filename:
            target_filename = target_filename.replace("_Auxiliar.csv", "")
        elif target_filename.lower().endswith(".csv"):
             target_filename = target_filename[:-4]

        # Add to Group
        if target_filename not in updates_by_file:
            updates_by_file[target_filename] = []
        
        updates_by_file[target_filename].append({
            "real_code": real_code,
            "row_data": row_data,
            "action_val": action_val,
            "forced_row": forced_row,
            "csv_raw_code": csv_raw_code
        })

    # --- Phase 2: Execute Updates per File ---
    print(f"[EXCEL] Starting Batch Update for {len(updates_by_file)} Unique Files.", flush=True)
    
    for target_filename, items in updates_by_file.items():
        # Search for .xlsx or .xlsm
        found_excel = None
        for ext in ["", ".xlsx", ".xlsm", ".xls"]:
            p = AUX_COPIES_DIR / (target_filename + ext)
            if p.exists():
                found_excel = p
                break
        
        if not found_excel:
             print(f"[EXCEL] File {target_filename} not found. Skipping {len(items)} items.")
             continue
             
        excel_path = found_excel
        print(f"[EXCEL] Processing {target_filename} ({len(items)} updates)...")

        try:
            is_xlsm = excel_path.suffix.lower() == '.xlsm'
            workbook_buffer = None
            wb = None
            
            # Load Workbook (Once)
            try:
                with open(excel_path, 'rb') as f_stream:
                    workbook_buffer = io.BytesIO(f_stream.read())
                wb = _load_workbook_quietly(workbook_buffer, keep_vba=is_xlsm)
            except Exception as e_load:
                # Decryption Retry
                print(f"[EXCEL] Load failed ({e_load}). Trying decryption...")
                try:
                    decrypted = io.BytesIO()
                    with open(excel_path, 'rb') as f_enc:
                        office_file = msoffcrypto.OfficeFile(f_enc)
                        office_file.load_key(password='bpb')
                        office_file.decrypt(decrypted)
                    decrypted.seek(0)
                    workbook_buffer = decrypted
                    wb = _load_workbook_quietly(workbook_buffer, keep_vba=is_xlsm)
                except Exception as e_dec:
                    print(f"[EXCEL] Decryption failed: {e_dec}")
                    raise e_load

            # Identify Sheet
            ws = None
            if "Auxiliar" in wb.sheetnames: ws = wb["Auxiliar"]
            elif "Aux" in wb.sheetnames: ws = wb["Aux"]
            else: ws = wb.active

            START_ROW = 3
            
            # Iterate Items for THIS file
            for item in items:
                real_code = item["real_code"]
                row_data = item["row_data"]
                action = item["action_val"]
                forced_row = item["forced_row"]
                csv_raw_code = item["csv_raw_code"]

                # Determine Last Row (Dynamic, changes as we append)
                last_real_row = START_ROW
                for r in range(ws.max_row, START_ROW - 1, -1):
                    val = ws.cell(row=r, column=2).value
                    if val and str(val).strip():
                        last_real_row = r
                        break
                
                # Determine Next Num
                last_num = 0
                try:
                    val_num = ws.cell(row=last_real_row, column=1).value
                    if isinstance(val_num, (int, float)):
                        last_num = int(val_num)
                except: pass
                next_num = last_num + 1

                # Find Row for Update
                found_row_idx = forced_row
                if not found_row_idx:
                    search_term = str(csv_raw_code or real_code).strip()
                    for r in range(START_ROW, last_real_row + 1):
                        cell_val = str(ws.cell(row=r, column=2).value or "").strip()
                        if cell_val == search_term:
                            found_row_idx = r
                            break
                
                # Apply Write
                target_row = None
                if found_row_idx:
                    target_row = found_row_idx
                    ws.cell(row=target_row, column=1).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Updating {real_code} at Row {target_row}")
                elif action == "new":
                    target_row = last_real_row + 1
                    ws.cell(row=target_row, column=1, value=next_num).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Appending {real_code} at Row {target_row}")
                else:
                    # Fallback Append
                    target_row = last_real_row + 1
                    ws.cell(row=target_row, column=1, value=next_num).alignment = Alignment(horizontal='center')
                    print(f"[EXCEL] Appending (Fallback) {real_code} at Row {target_row}")

                # Write Data Cols
                for c_idx, val in enumerate(row_data, start=2):
                    cell = ws.cell(row=target_row, column=c_idx)
                    cell.value = val
                    if c_idx == 2: cell.font = Font(bold=True)
                    else: cell.alignment = Alignment(horizontal='right')

            # Save Workbook (Once)
            wb.save(str(excel_path))
            print(f"[EXCEL] Saved {target_filename} successfully.")

            # Update Index CSV (Once)
            try:
                # INDICES_DIR = Path(r"\\192.168.0.13\lcontigiani\Oficina Tecnica\Registro de Control de Producto\Auxiliares\indices_auxiliar")
                INDICES_DIR = BASE_DIR / "Auxiliares/indices_auxiliar"
                if not INDICES_DIR.exists(): INDICES_DIR.mkdir(parents=True, exist_ok=True)
                
                csv_out = INDICES_DIR / f"{excel_path.stem}_Auxiliar.csv"
                with open(csv_out, 'w', newline='', encoding='utf-8') as f_csv:
                    writer = csv.writer(f_csv, delimiter=';')
                    for row in ws.iter_rows(values_only=True):
                        clean_row = [str(c) if c is not None else "" for c in row]
                        writer.writerow(clean_row)
            except Exception as e_csv:
                print(f"[EXCEL] Warning: CSV Index update failed: {e_csv}")

            if wb is not None:
                wb.close()
            if workbook_buffer is not None:
                workbook_buffer.close()

        except Exception as e:
            print(f"[EXCEL] Failed to process {target_filename}: {e}")
            if "Permission denied" in str(e):
                 raise Exception(f"El Registro {target_filename} se encuentra abierto.")
        finally:
            if wb is not None:
                try:
                    wb.close()
                except Exception:
                    pass
            if workbook_buffer is not None:
                try:
                    workbook_buffer.close()
                except Exception:
                    pass
            raise Exception(f"Error en archivo '{target_filename}': {e}")



@app.route('/api/submit-activity', methods=['POST'])

def submit_activity():

    try:

        data = request.json

        date_str = data.get('date') # Format DD/MM/YYYY

        activities = data.get('activities')

        

        if not date_str or not activities:

            return jsonify({'status': 'error', 'content': 'Datos incompletos'})

            

        # Parse date to safe filename

        safe_date = date_str.replace('/', '-')

        

        # Base Path for Activities

        base_dir = os.path.dirname(os.path.abspath(__file__))

        activities_dir = os.path.join(base_dir, 'Datos', 'Actividades')

        if not os.path.exists(activities_dir):

            os.makedirs(activities_dir)

            

        file_path = os.path.join(activities_dir, f'registro_{safe_date}.json')

        

        record = {

            'date': date_str,

            'timestamp': datetime.now().isoformat(),

            'activities': activities,

            'user': session.get('user'),

            'status': 'completed'

        }

        

        _write_json_file_atomic(file_path, record, indent=4, ensure_ascii=False)

            

        return jsonify({'status': 'success'})

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/approve-activity', methods=['POST'])
def approve_activity():
    try:
        data = request.json
        date_str = data.get('date')
        if not date_str:
            return jsonify({'status': 'error', 'content': 'Fecha requerida'})

        safe_date = date_str.replace('/', '-')
        base_dir = os.path.dirname(os.path.abspath(__file__))
        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        if not os.path.exists(file_path):
            return jsonify({'status': 'error', 'content': 'Registro no encontrado'})

        with open(file_path, 'r', encoding='utf-8') as f:
            record = json.load(f)

        if record.get('status') == 'approved':
            return jsonify({'status': 'error', 'content': 'El registro ya fue aprobado.'})

        # --- Guardado en CSV + estado mailer ---
        try:
            mailer_state_path = ACTIVITY_CODE_DIR / "data" / "activity_mailer_state.json"
            if not mailer_state_path.exists():
                return jsonify({'status': 'error', 'content': 'Estado de correos no encontrado.'})

            with open(mailer_state_path, 'r', encoding='utf-8') as f:
                mailer_data = json.load(f)

            try:
                dt_obj = datetime.strptime(date_str, '%d/%m/%Y')
                iso_date = dt_obj.strftime('%Y-%m-%d')
            except ValueError:
                return jsonify({'status': 'error', 'content': 'Formato de fecha invalido.'})

            token = None
            target_campaign = None
            for camp in mailer_data.get('campaigns', []):
                if camp.get('date') == iso_date:
                    token = camp.get('token')
                    target_campaign = camp
                    break

            if not token or not target_campaign:
                return jsonify({'status': 'error', 'content': 'No se encontro la solicitud (token) para esa fecha.'})

            user_email = record.get('user') or session.get('user')

            if user_email:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    users_db = json.load(f)

                user_data = users_db.get(user_email)
                if user_data:
                    full_name = user_data.get('display_name', user_email)

                    def clean_csv_val(val):
                        return str(val).replace(';', ',').replace('\n', ' ').replace('\r', ' ').strip()

                    legacy_root = ACTIVITY_CODE_DIR / "data"
                    global_csv = legacy_root / "base_datos_respuestas.csv"
                    individual_csv = legacy_root / "respuestas_csv" / f"{full_name}.csv"

                    now = datetime.now()
                    time_str = now.strftime("%H:%M:%S")

                    activities = record.get('activities', [])

                    body_blocks = []
                    for act in activities:
                        proj = clean_csv_val(act.get('project', ''))
                        time_val = str(act.get('time', '0')).replace('.', ',')
                        desc = clean_csv_val(act.get('description', ''))
                        body_blocks.append(f"Proyecto: {proj}\\nTiempo destinado: {time_val} horas\\nDescripcion: {desc}")
                    body_text = "\\n\\n".join(body_blocks) if body_blocks else ""

                    responses = target_campaign.setdefault('responses', {})
                    responses[user_email] = {
                        "from": user_email,
                        "from_name": full_name,
                        "subject": f"Re: {target_campaign.get('subject')}",
                        "date": now.strftime("%a, %d %b %Y %H:%M:%S"),
                        "received_at": now.isoformat(),
                        "snippet": body_text,
                        "body": body_text,
                        "uid": 0,
                        "saved_path": "WEB_SUBMISSION"
                    }

                    with open(global_csv, 'a', newline='', encoding='utf-8') as lg:
                        writer = csv.writer(lg, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                        for act in activities:
                            proj = clean_csv_val(act.get('project', ''))
                            time_val = str(act.get('time', '0')).replace('.', ',')
                            desc = clean_csv_val(act.get('description', ''))
                            writer.writerow([date_str, time_str, token, full_name, user_email, proj, proj, time_val, desc])

                    individual_csv.parent.mkdir(parents=True, exist_ok=True)
                    with open(individual_csv, 'a', newline='', encoding='utf-8') as li:
                        writer = csv.writer(li, delimiter=';', quoting=csv.QUOTE_MINIMAL)
                        for act in activities:
                            proj = clean_csv_val(act.get('project', ''))
                            time_val = str(act.get('time', '0')).replace('.', ',')
                            desc = clean_csv_val(act.get('description', ''))
                            writer.writerow([date_str, time_str, token, proj, time_val, proj, desc])

                    _write_json_file_atomic(mailer_state_path, mailer_data, indent=2, ensure_ascii=False)

        except Exception as e:
            print(f"Error saving legacy CSVs: {e}")
            raise e

        record['status'] = 'approved'
        record['approved_at'] = datetime.now().isoformat()

        _write_json_file_atomic(file_path, record, indent=4, ensure_ascii=False)

        return jsonify({'status': 'success'})

    except Exception as e:
        return jsonify({'status': 'error', 'content': str(e)})

@app.route('/api/get-activity-status', methods=['GET'])

def get_activity_status():

    try:

        date_str = request.args.get('date')

        if not date_str:

            return jsonify({'status': 'error'})

            

        safe_date = date_str.replace('/', '-')

        base_dir = os.path.dirname(os.path.abspath(__file__))

        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        

        if os.path.exists(file_path):

             with open(file_path, 'r', encoding='utf-8') as f:

                 data = json.load(f)

             is_approved = data.get('status') == 'approved'

             return jsonify({'status': 'success', 'completed': True, 'approved': is_approved})

        else:

             return jsonify({'status': 'success', 'completed': False})

             

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/get-activity', methods=['GET'])

def get_activity():

    try:

        date_str = request.args.get('date')

        if not date_str:

            return jsonify({'status': 'error', 'content': 'Fecha requerida'})

            

        safe_date = date_str.replace('/', '-')

        base_dir = os.path.dirname(os.path.abspath(__file__))

        file_path = os.path.join(base_dir, 'Datos', 'Actividades', f'registro_{safe_date}.json')

        

        if os.path.exists(file_path):

            with open(file_path, 'r', encoding='utf-8') as f:

                data = json.load(f)

                return jsonify({'status': 'success', 'activities': data.get('activities', [])})

        else:

             return jsonify({'status': 'not_found'})

             

    except Exception as e:

        return jsonify({'status': 'error', 'content': str(e)})



@app.route('/api/glossary', methods=['GET'])

def get_glossary():

    # Dynamic path resolution for portability
    # BASE_DIR points to '.../Registro de Control de Producto'
    glossary_path = ACTIVITY_CODE_DIR / "config" / "glosario_proyectos.log"

    try:

        if os.path.exists(glossary_path):

            with open(glossary_path, 'r', encoding='utf-8', errors='ignore') as f:

                content = f.read()

            return jsonify({'status': 'success', 'content': content})

        else:

            return jsonify({'status': 'error', 'content': 'El archivo de glosario (glosario_proyectos.log) no fue encontrado.'})

    except Exception as e:

        return jsonify({'status': 'error', 'content': f'Error al leer el glosario: {str(e)}'})





@app.route('/api/glossary/add', methods=['POST'])
def add_glossary_project():
    data = request.get_json(silent=True) or {}
    project = str(data.get('project') or '').strip()

    if not project:
        return jsonify({'status': 'error', 'message': 'Nombre de proyecto requerido.'}), 400

    glossary_path = ACTIVITY_CODE_DIR / "config" / "glosario_proyectos.log"

    try:
        glossary_path.parent.mkdir(parents=True, exist_ok=True)

        entries = []
        if glossary_path.exists():
            entries = [line.strip() for line in glossary_path.read_text(encoding='utf-8', errors='ignore').splitlines() if line.strip()]

        # Evita duplicados case-insensitive
        if any(project.lower() == item.lower() for item in entries):
            return jsonify({'status': 'success', 'message': 'Proyecto ya existente.', 'project': project, 'already_exists': True})

        entries.append(project)
        glossary_path.write_text('\n'.join(entries) + '\n', encoding='utf-8')

        return jsonify({'status': 'success', 'message': 'Proyecto agregado.', 'project': project})
    except Exception as e:
        return jsonify({'status': 'error', 'message': f'Error al guardar glosario: {str(e)}'}), 500

@app.route('/api/activity-history', methods=['GET'])

def get_activity_history():

    if not session.get('user'):

         return jsonify({"status": "error", "message": "No autenticado"}), 401

    

    # Paths (Dynamic based on BASE_DIR)
    # BASE_DIR points to '.../Registro de Control de Producto'
    # We need to go up to 'Oficina Tecnica' then down to 'Registro de Actividad'
    ACTIVITY_BASE = ACTIVITY_CODE_DIR
    
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"



    user_email = session.get('user')

    print(f"[HISTORY] User from session: {user_email}")



    target_filename = None

    

    try:

        # 1. Lookup User

        if not DESTINATARIOS_FILE.exists():

             print(f"[HISTORY] Destinatarios file not found: {DESTINATARIOS_FILE}")

             return jsonify({'status': 'error', 'message': 'Archivo de destinatarios no encontrado.'}), 500

             

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:

            reader = csv.DictReader(f) # Headers: NOMBRE,APELLIDO,EMAIL

            for row in reader:

                # print(f"Checking row: {row}")

                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():

                    # Found

                    target_filename = f"{row['NOMBRE'].strip()} {row['APELLIDO'].strip()}.csv"

                    print(f"[HISTORY] Found target filename: {target_filename}")

                    break

        

        if not target_filename:

             print(f"[HISTORY] User {user_email} not found in destinatarios.")

             return jsonify({'status': 'error', 'message': 'Usuario no encontrado en lista de destinatarios.'}), 404



        # 2. Read CSV

        csv_path = DATA_DIR / target_filename

        print(f"[HISTORY] Looking for CSV at: {csv_path}")

        

        if not csv_path.exists():

             print(f"[HISTORY] CSV file does not exist.")

             return jsonify({'status': 'success', 'data': [], 'columns': []})

        

        data_rows = []

        columns = []

        

        def read_csv_data(enc):

            local_rows = []

            with open(csv_path, 'r', encoding=enc) as f:

                reader = csv.reader(f, delimiter=';')

                local_rows = list(reader)

            return local_rows



        try:

            try:

                rows = read_csv_data('utf-8')

            except UnicodeDecodeError:

                print(f"[HISTORY] UTF-8 failed, trying latin-1")

                rows = read_csv_data('latin-1')

            

            if len(rows) > 0:

                columns = rows[0]

                data_rows = rows[1:] # Skip header

        except Exception as e:

            print(f"[HISTORY] Error reading CSV: {e}")

            return jsonify({'status': 'error', 'message': f'Error leyendo archivo de historial: {str(e)}'}), 500

        

        print(f"[HISTORY] Found {len(data_rows)} rows.")

    except Exception as e:
        print(f"[HISTORY] Error general: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', 'data': data_rows, 'columns': columns})


@app.route('/api/activity-stats', methods=['GET'])
def get_activity_stats():
    """
    Devuelve horas acumuladas por proyecto para el usuario logueado.
    Filtros opcionales ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD.
    """
    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401

    ACTIVITY_BASE = ACTIVITY_CODE_DIR
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"

    user_email = session.get('user')
    target_filename = None

    # Rango de fechas
    start_param = request.args.get('start_date')
    end_param = request.args.get('end_date')
    start_dt = None
    end_dt = None
    try:
        if start_param:
            start_dt = datetime.strptime(start_param, '%Y-%m-%d').date()
        if end_param:
            end_dt = datetime.strptime(end_param, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'status': 'error', 'message': 'Fechas invÃ¡lidas'}), 400

    # Buscar archivo del usuario
    try:
        if not DESTINATARIOS_FILE.exists():
             return jsonify({'status': 'error', 'message': 'Archivo de destinatarios no encontrado.'}), 500

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():
                    target_filename = f"{row['NOMBRE'].strip()} {row['APELLIDO'].strip()}.csv"
                    break

        if not target_filename:
             return jsonify({'status': 'error', 'message': 'Usuario no encontrado en lista de destinatarios.'}), 404

        csv_path = DATA_DIR / target_filename
        if not csv_path.exists():
             return jsonify({'status': 'success', 'data': [], 'total': 0})

        def read_csv(enc):
            with open(csv_path, 'r', encoding=enc) as f:
                return list(csv.reader(f, delimiter=';'))

        try:
            rows = read_csv('utf-8')
        except Exception:
            rows = read_csv('latin-1')

        totals = {}
        records = []
        for idx, row in enumerate(rows):
            if not row or len(row) < 5:
                continue
            # Skip header row explicitly (first line) even if text is not exactly "Fecha"
            if idx == 0 or str(row[0]).lower().strip() == 'fecha':
                continue

            # Fecha en CSV es dd/mm/yyyy
            try:
                d = datetime.strptime(row[0], '%d/%m/%Y').date()
            except Exception:
                d = None

            if start_dt and d and d < start_dt:
                continue
            if end_dt and d and d > end_dt:
                continue

            proj = str(row[3]).strip() if len(row) > 3 else ''
            # horas viene en la columna 7 (index 7) en base_datos_respuestas.csv; fallback a col 4 si no existe
            time_str = str(row[7]).strip() if len(row) > 7 else (str(row[4]).strip() if len(row) > 4 else '0')
            # Extraer nÃºmero (soporta "1", "1,5", "1.5", "1 Horas")
            num_match = re.search(r'[-+]?[0-9]*[\\.,]?[0-9]+', time_str)
            time_clean = num_match.group(0) if num_match else '0'
            try:
                hours = float(time_clean.replace(',', '.'))
            except Exception:
                hours = 0.0

            if not proj:
                proj = 'Sin Proyecto Asignado'
            totals[proj] = totals.get(proj, 0.0) + hours
            record_desc = ''
            if len(row) > 5:
                record_desc = str(row[5]).strip()
            records.append({
                "project": proj,
                "hours": round(hours, 2),
                "user": user_email,
                "date": row[0] if len(row) > 0 else '',
                "time": row[1] if len(row) > 1 else '',
                "token": row[2] if len(row) > 2 else '',
                "description": record_desc
            })

        data_rows = [{'project': k, 'hours': round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[1], reverse=True)]
        total_hours = round(sum(totals.values()), 2)

        return jsonify({'status': 'success', 'data': data_rows, 'total': total_hours, 'records': records})

    except Exception as e:
        print(f"[STATS] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



@app.route('/api/activity-stats-global', methods=['GET'])
def get_activity_stats_global():
    """
    Devuelve horas acumuladas por proyecto considerando todos los CSV de respuestas.
    Filtros opcionales ?start_date=YYYY-MM-DD&end_date=YYYY-MM-DD.
    """
    if not session.get('user'):
         return jsonify({"status": "error", "message": "No autenticado"}), 401

    ACTIVITY_BASE = ACTIVITY_CODE_DIR
    MASTER_CSV = ACTIVITY_BASE / "data/base_datos_respuestas.csv"

    start_param = request.args.get('start_date')
    end_param = request.args.get('end_date')
    start_dt = None
    end_dt = None
    try:
        if start_param:
            start_dt = datetime.strptime(start_param, '%Y-%m-%d').date()
        if end_param:
            end_dt = datetime.strptime(end_param, '%Y-%m-%d').date()
    except Exception:
        return jsonify({'status': 'error', 'message': 'Fechas invÃ¡lidas'}), 400

    def read_rows(csv_path):
        try:
            with open(csv_path, 'r', encoding='utf-8') as f:
                return list(csv.reader(f, delimiter=';'))
        except Exception:
            with open(csv_path, 'r', encoding='latin-1') as f:
                return list(csv.reader(f, delimiter=';'))

    totals = {}
    records = []
    try:
        if not MASTER_CSV.exists():
             return jsonify({'status': 'success', 'data': [], 'total': 0, 'records': []})

        try:
            rows = read_rows(MASTER_CSV)
        except Exception as e:
            print(f"[STATS][GLOBAL] Error leyendo maestro: {e}")
            return jsonify({'status': 'error', 'message': 'No se pudo leer base_datos_respuestas.csv'}), 500

        users_map = {}
        users_by_display = {}
        try:
            if USERS_FILE.exists():
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    users_map = json.load(f) or {}
                for uname, udata in users_map.items():
                    display = str((udata or {}).get('display_name') or '').strip().lower()
                    if display:
                        users_by_display[display] = udata or {}
        except Exception as e:
            print(f"[STATS][GLOBAL] Warning loading users.json: {e}")

        def resolve_area(user_email, user_name):
            role = ''
            email_key = str(user_email or '').strip().lower()
            if email_key and email_key in users_map:
                role = str((users_map[email_key] or {}).get('role') or '').strip().lower()
            if not role:
                display_key = str(user_name or '').strip().lower()
                if display_key and display_key in users_by_display:
                    role = str((users_by_display[display_key] or {}).get('role') or '').strip().lower()
            return 'Calidad' if role == 'calidad' else 'Oficina Tecnica'

        for idx_row, row in enumerate(rows):
            if not row or len(row) < 5:
                continue
            if idx_row == 0 or str(row[0]).lower().strip() == 'fecha':
                continue

            try:
                d = datetime.strptime(row[0], '%d/%m/%Y').date()
            except Exception:
                d = None

            if start_dt and d and d < start_dt:
                continue
            if end_dt and d and d > end_dt:
                continue

            # Columnas: 0 Fecha, 1 Hora, 2 Token, 3 Nombre, 4 Email, 5 ProyectoInicial, 6 ProyectoFinal, 7 Tiempo, 8 Descripcion
            user_val = str(row[3]).strip() if len(row) > 3 else 'Desconocido'
            user_email = str(row[4]).strip() if len(row) > 4 else ''
            proj = str(row[6]).strip() if len(row) > 6 else (str(row[5]).strip() if len(row) > 5 else '')
            time_str = str(row[7]).strip() if len(row) > 7 else '0'
            num_match = re.search(r'[-+]?[0-9]*[\\.,]?[0-9]+', time_str)
            time_clean = num_match.group(0) if num_match else '0'
            try:
                hours = float(time_clean.replace(',', '.'))
            except Exception:
                hours = 0.0
            if not proj:
                proj = 'Sin Proyecto Asignado'
            totals[proj] = totals.get(proj, 0.0) + hours
            record_desc = ''
            if len(row) > 8:
                record_desc = str(row[8]).strip()
            elif len(row) > 5:
                record_desc = str(row[5]).strip()
            records.append({
                "project": proj,
                "hours": round(hours, 2),
                "user": user_val,
                "area": resolve_area(user_email, user_val),
                "date": row[0] if len(row) > 0 else '',
                "time": row[1] if len(row) > 1 else '',
                "token": row[2] if len(row) > 2 else '',
                "description": record_desc
            })

        data_rows = [{'project': k, 'hours': round(v, 2)} for k, v in sorted(totals.items(), key=lambda x: x[1], reverse=True)]
        total_hours = round(sum(totals.values()), 2)

        return jsonify({'status': 'success', 'data': data_rows, 'total': total_hours, 'records': records})
    except Exception as e:
        print(f"[STATS][GLOBAL] Error general: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500







# -------------------------------------------------------------------------

# ACTIVITY TRACKING EXTENSION

# -------------------------------------------------------------------------



ACTIVITY_BASE_PATH = ACTIVITY_CODE_DIR

ACTIVITY_STATE_FILE = ACTIVITY_BASE_PATH / "data/activity_mailer_state.json"

ACTIVITY_CSV_BASE = ACTIVITY_BASE_PATH / "data/base_datos_respuestas.csv"

ACTIVITY_CSV_USER_DIR = ACTIVITY_BASE_PATH / "data/respuestas_csv"

DESTINATARIOS_FILE = ACTIVITY_BASE_PATH / "config/destinatarios.csv"

QUALITY_BASE_PATH = QUALITY_ROOT
QUALITY_WORKBOOK_PATH = QUALITY_BASE_PATH / "En control de Calidad-Lorenzo.xlsx"
QUALITY_CSV_CACHE_DIR = QUALITY_BASE_PATH / "_csv_cache"
QUALITY_OBSERVATIONS_FILE = QUALITY_BASE_PATH / "quality_observations.json"
QUALITY_PENDING_CATEGORIES_FILE = QUALITY_BASE_PATH / "quality_pending_categories.json"
QUALITY_PENDING_HANDLERS_FILE = QUALITY_BASE_PATH / "quality_pending_handlers.json"
QUALITY_SHEET_CACHE_MAP = {
    'Control de Calidad': QUALITY_CSV_CACHE_DIR / "control_de_calidad.csv",
    'Aprobaciones': QUALITY_CSV_CACHE_DIR / "aprobaciones.csv"
}
QUALITY_CSV_LOCK = threading.Lock()
QUALITY_OBSERVATIONS_LOCK = threading.Lock()
QUALITY_PENDING_CATEGORIES_LOCK = threading.Lock()
QUALITY_PENDING_HANDLERS_LOCK = threading.Lock()
QUALITY_HISTORY_CACHE_LOCK = threading.Lock()
QUALITY_PENDING_HIDDEN_CATEGORY_ID = 'hidden'
QUALITY_PENDING_HIDDEN_CATEGORY_NAME = 'Ocultos'
QUALITY_PENDING_HIDDEN_CATEGORY_COLOR = '#5c6370'
QUALITY_HISTORY_CACHE = {
    'csv_mtime': None,
    'obs_mtime': None,
    'data': None
}


def _can_access_quality_module():
    role = str(session.get('role') or '').strip().lower()
    return role in {'admin', 'calidad'}


def _normalize_sheet_header(value):
    text = unicodedata.normalize('NFD', str(value or '').strip())
    text = ''.join(ch for ch in text if unicodedata.category(ch) != 'Mn')
    text = re.sub(r'\s+', ' ', text).strip().lower()
    return text


def _quality_tracking_key(item, produc):
    return f"{str(item or '').strip().lower()}||{str(produc or '').strip().lower()}"


def _quality_pending_row_key(item, produc, oc_numero='', fecha_ing=''):
    return "||".join([
        str(item or '').strip().lower(),
        str(produc or '').strip().lower(),
        str(oc_numero or '').strip().lower(),
        str(fecha_ing or '').strip().lower()
    ])


def _quality_load_observation_registry():
    if not QUALITY_OBSERVATIONS_FILE.exists():
        return {'items': {}}

    try:
        with QUALITY_OBSERVATIONS_FILE.open('r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'items': {}}
        items = data.get('items')
        if not isinstance(items, dict):
            data['items'] = {}
        return data
    except Exception as e:
        print(f"[QUALITY][OBS] Warning loading observations registry: {e}")
        return {'items': {}}


def _quality_save_observation_registry(registry):
    QUALITY_BASE_PATH.mkdir(parents=True, exist_ok=True)
    _write_json_file_atomic(QUALITY_OBSERVATIONS_FILE, registry, indent=2, ensure_ascii=False)


def _quality_load_pending_categories_state():
    if not QUALITY_PENDING_CATEGORIES_FILE.exists():
        return _quality_ensure_hidden_category({'categories': [], 'assignments': {}})

    try:
        with QUALITY_PENDING_CATEGORIES_FILE.open('r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'categories': [], 'assignments': {}}
        categories = data.get('categories')
        assignments = data.get('assignments')
        data['categories'] = categories if isinstance(categories, list) else []
        data['assignments'] = assignments if isinstance(assignments, dict) else {}
        return _quality_ensure_hidden_category(data)
    except Exception as e:
        print(f"[QUALITY][CAT] Warning loading categories state: {e}")
        return _quality_ensure_hidden_category({'categories': [], 'assignments': {}})


def _quality_save_pending_categories_state(state):
    QUALITY_BASE_PATH.mkdir(parents=True, exist_ok=True)
    _write_json_file_atomic(QUALITY_PENDING_CATEGORIES_FILE, state, indent=2, ensure_ascii=False)


def _quality_load_pending_handlers_state():
    if not QUALITY_PENDING_HANDLERS_FILE.exists():
        return {'assignments': {}}
    try:
        with QUALITY_PENDING_HANDLERS_FILE.open('r', encoding='utf-8') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            return {'assignments': {}}
        assignments = data.get('assignments')
        data['assignments'] = assignments if isinstance(assignments, dict) else {}
        return data
    except Exception as e:
        print(f"[QUALITY][HANDLER] Warning loading handlers state: {e}")
        return {'assignments': {}}


def _quality_save_pending_handlers_state(state):
    QUALITY_BASE_PATH.mkdir(parents=True, exist_ok=True)
    _write_json_file_atomic(QUALITY_PENDING_HANDLERS_FILE, state, indent=2, ensure_ascii=False)


def _quality_normalize_category_color(value):
    color = str(value or '').strip()
    if re.fullmatch(r'#[0-9a-fA-F]{6}', color):
        return color.lower()
    return '#4aa3ff'


def _quality_ensure_hidden_category(state):
    if not isinstance(state, dict):
        state = {}

    categories = state.get('categories')
    assignments = state.get('assignments')
    categories = categories if isinstance(categories, list) else []
    assignments = assignments if isinstance(assignments, dict) else {}

    normalized_categories = []
    hidden_category = None
    migrated_ids = set()

    for category in categories:
        if not isinstance(category, dict):
            continue
        category_id = str(category.get('id') or '').strip()
        category_name = str(category.get('name') or '').strip()
        if not category_id and not category_name:
            continue

        is_hidden = (
            category_id == QUALITY_PENDING_HIDDEN_CATEGORY_ID or
            category_name.lower() == QUALITY_PENDING_HIDDEN_CATEGORY_NAME.lower()
        )
        if is_hidden:
            migrated_ids.add(category_id)
            if hidden_category is None:
                hidden_category = {
                    'id': QUALITY_PENDING_HIDDEN_CATEGORY_ID,
                    'name': QUALITY_PENDING_HIDDEN_CATEGORY_NAME,
                    'color': _quality_normalize_category_color(category.get('color') or QUALITY_PENDING_HIDDEN_CATEGORY_COLOR),
                    'updated_at': category.get('updated_at', ''),
                    'updated_by': category.get('updated_by', '')
                }
            continue

        normalized_categories.append(category)

    if hidden_category is None:
        hidden_category = {
            'id': QUALITY_PENDING_HIDDEN_CATEGORY_ID,
            'name': QUALITY_PENDING_HIDDEN_CATEGORY_NAME,
            'color': QUALITY_PENDING_HIDDEN_CATEGORY_COLOR,
            'updated_at': '',
            'updated_by': ''
        }

    if migrated_ids:
        assignments = {
            row_key: (QUALITY_PENDING_HIDDEN_CATEGORY_ID if str(assigned_id or '').strip() in migrated_ids else assigned_id)
            for row_key, assigned_id in assignments.items()
        }

    state['categories'] = [hidden_category, *normalized_categories]
    state['assignments'] = assignments
    return state


def _quality_sort_categories(categories):
    valid_categories = [category for category in categories if isinstance(category, dict) and category.get('id')]
    hidden_category = None
    regular_categories = []
    for category in valid_categories:
        if str(category.get('id') or '').strip() == QUALITY_PENDING_HIDDEN_CATEGORY_ID:
            hidden_category = category
        else:
            regular_categories.append(category)

    sorted_categories = sorted(
        regular_categories,
        key=lambda category: str(category.get('name') or '').strip().lower()
    )
    return ([hidden_category] if hidden_category else []) + sorted_categories


def _quality_register_snapshot_pairs(snapshot_records):
    with QUALITY_OBSERVATIONS_LOCK:
        registry = _quality_load_observation_registry()
        items = registry.setdefault('items', {})
        changed = False

        for record in snapshot_records or []:
            item = _format_quality_value(record.get('item'))
            produc = _format_quality_value(record.get('produc'))
            if not item and not produc:
                continue

            key = _quality_tracking_key(item, produc)
            existing = items.get(key) if isinstance(items.get(key), dict) else {}
            subgroup = _format_quality_value(record.get('subgrupo'))

            payload = {
                'item': item,
                'produc': produc,
                'subgrupo': subgroup or existing.get('subgrupo', ''),
                'observacion': _format_quality_value(existing.get('observacion')),
                'updated_at': existing.get('updated_at', ''),
                'updated_by': existing.get('updated_by', '')
            }
            if payload != existing:
                items[key] = payload
                changed = True

        if changed:
            registry['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_observation_registry(registry)

        return items


def _quality_update_observation(item, produc, observacion, updated_by=''):
    item = _format_quality_value(item)
    produc = _format_quality_value(produc)
    if not item and not produc:
        raise ValueError('item/produc requeridos')

    with QUALITY_OBSERVATIONS_LOCK:
        registry = _quality_load_observation_registry()
        items = registry.setdefault('items', {})
        key = _quality_tracking_key(item, produc)
        existing = items.get(key) if isinstance(items.get(key), dict) else {}

        items[key] = {
            'item': item or existing.get('item', ''),
            'produc': produc or existing.get('produc', ''),
            'subgrupo': existing.get('subgrupo', ''),
            'observacion': str(observacion or '').strip(),
            'updated_at': datetime.now().isoformat(timespec='seconds'),
            'updated_by': _format_quality_value(updated_by)
        }
        registry['updated_at'] = datetime.now().isoformat(timespec='seconds')
        _quality_save_observation_registry(registry)
        return items[key]


def _quality_load_observation_items():
    registry = _quality_load_observation_registry()
    items = registry.get('items')
    return items if isinstance(items, dict) else {}


def _quality_get_mtime(path_obj):
    try:
        return path_obj.stat().st_mtime
    except Exception:
        return None


def _format_quality_value(value):
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, date):
        return value.strftime('%Y-%m-%d')
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _quality_force_save_workbook():
    if os.name != 'nt':
        return
    if not QUALITY_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {QUALITY_WORKBOOK_PATH}')

    workbook_path = str(QUALITY_WORKBOOK_PATH)
    ps_script = f"""
$ErrorActionPreference = 'Stop'
$path = @'
{workbook_path}
'@
$excel = $null
$workbook = $null
try {{
    $excel = New-Object -ComObject Excel.Application
    $excel.Visible = $false
    $excel.DisplayAlerts = $false
    $excel.AskToUpdateLinks = $false
    $excel.EnableEvents = $false
    $workbook = $excel.Workbooks.Open($path, 0, $false)
    try {{ $excel.CalculateFullRebuild() }} catch {{}}
    $workbook.Save()
}}
finally {{
    if ($workbook -ne $null) {{ $workbook.Close($true) }}
    if ($excel -ne $null) {{ $excel.Quit() }}
    [System.GC]::Collect()
    [System.GC]::WaitForPendingFinalizers()
}}
"""
    subprocess.run(
        [
            "powershell",
            "-NoProfile",
            "-NonInteractive",
            "-ExecutionPolicy",
            "Bypass",
            "-Command",
            ps_script,
        ],
        check=True,
        timeout=120,
        capture_output=True,
        text=True,
    )


def _quality_sync_csv_exports(force=False):
    if not QUALITY_WORKBOOK_PATH.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {QUALITY_WORKBOOK_PATH}')

    workbook_mtime = QUALITY_WORKBOOK_PATH.stat().st_mtime
    csv_paths = list(QUALITY_SHEET_CACHE_MAP.values())
    needs_refresh = force or any(
        (not path.exists()) or path.stat().st_mtime < workbook_mtime
        for path in csv_paths
    )
    if not needs_refresh:
        return

    with QUALITY_CSV_LOCK:
        workbook_mtime = QUALITY_WORKBOOK_PATH.stat().st_mtime
        needs_refresh = force or any(
            (not path.exists()) or path.stat().st_mtime < workbook_mtime
            for path in csv_paths
        )
        if not needs_refresh:
            return

        try:
            _quality_force_save_workbook()
        except Exception as e:
            print(f"[QUALITY][CACHE] Warning forcing workbook save: {e}")
        workbook_mtime = QUALITY_WORKBOOK_PATH.stat().st_mtime

        QUALITY_CSV_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        wb = _load_workbook_quietly(str(QUALITY_WORKBOOK_PATH), read_only=True, data_only=True)
        try:
            for sheet_name, csv_path in QUALITY_SHEET_CACHE_MAP.items():
                if sheet_name not in wb.sheetnames:
                    raise ValueError(f'Hoja no encontrada: {sheet_name}')

                ws = wb[sheet_name]
                tmp_path = csv_path.with_suffix(csv_path.suffix + ".tmp")
                with tmp_path.open('w', newline='', encoding='utf-8-sig') as f:
                    writer = csv.writer(f)
                    for row in ws.iter_rows(values_only=True):
                        writer.writerow([_format_quality_value(cell) for cell in row])
                tmp_path.replace(csv_path)
        finally:
            wb.close()


def _quality_sheet_rows(sheet_name):
    _quality_sync_csv_exports()

    csv_path = QUALITY_SHEET_CACHE_MAP.get(sheet_name)
    if not csv_path:
        raise ValueError(f'Hoja no configurada: {sheet_name}')
    if not csv_path.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {csv_path}')

    with csv_path.open('r', newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header_row = next(reader, None)
        if not header_row:
            return []

        header_map = {}
        for idx, raw_header in enumerate(header_row):
            normalized = _normalize_sheet_header(raw_header)
            if normalized and normalized not in header_map:
                header_map[normalized] = idx

        data_rows = []
        for row in reader:
            if not row:
                continue
            has_values = any(cell is not None and str(cell).strip() != '' for cell in row)
            if not has_values:
                continue
            data_rows.append((tuple(row), header_map))

        return data_rows


@app.route('/api/quality/cache/refresh', methods=['POST'])
def refresh_quality_csv_cache():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module() and session.get('role') != 'admin':
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        _quality_sync_csv_exports(force=True)
        return jsonify({
            'status': 'success',
            'data': {
                'generated_at': datetime.now().isoformat(timespec='seconds')
            }
        })
    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 404
    except Exception as e:
        print(f"[QUALITY][CACHE] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


def _quality_get_value(row, header_map, *header_names):
    for header_name in header_names:
        idx = header_map.get(_normalize_sheet_header(header_name))
        if idx is not None and idx < len(row):
            return row[idx]
    return None


def _quality_parse_number(value):
    if value in (None, ''):
        return None
    try:
        return float(str(value).replace(',', '.').strip())
    except Exception:
        return None


def _quality_parse_date(value):
    if value in (None, ''):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value

    raw = str(value).strip()
    if not raw:
        return None

    for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%Y/%m/%d', '%Y-%m-%d %H:%M:%S'):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue
    return None


def _quality_build_control_snapshot():
    records = []
    ingress_pair = {}
    ingress_full = {}
    ingress_product = {}
    ingress_item = {}
    subgroup_pair = {}
    total_pair = {}
    total_full = {}
    total_product = {}
    total_item = {}

    for row, header_map in _quality_sheet_rows('Control de Calidad'):
        item = _format_quality_value(_quality_get_value(row, header_map, 'item'))
        produc = _format_quality_value(_quality_get_value(row, header_map, 'produc'))
        subgrupo = _format_quality_value(_quality_get_value(row, header_map, 'SubGrupo', 'Sub Grupo', 'Subgrupo'))
        oc_numero = _format_quality_value(_quality_get_value(row, header_map, 'OC numero', 'OC Numero'))
        canti_real = _format_quality_value(_quality_get_value(row, header_map, 'canti_real'))
        canti_real_item = _format_quality_value(_quality_get_value(row, header_map, 'canti_real_item'))
        fecha_ing_raw = _quality_get_value(row, header_map, 'Fecha_ing')
        fecha_partida_raw = _quality_get_value(row, header_map, 'Fecha Partida')

        total_qty = _quality_parse_number(canti_real) or 0.0
        remaining_qty = _quality_parse_number(canti_real_item) or 0.0
        approved_qty = max(total_qty - remaining_qty, 0.0)
        if total_qty > 0:
            approved_qty = min(approved_qty, total_qty)
            raw_progress_pct = (approved_qty / total_qty) * 100
            if approved_qty >= total_qty:
                progress_pct = 100
            else:
                progress_pct = min(99, int(raw_progress_pct))
        else:
            progress_pct = 0

        fecha_ing_date = _quality_parse_date(fecha_ing_raw)
        fecha_partida_date = _quality_parse_date(fecha_partida_raw)
        formatted_ing = _format_quality_value(fecha_ing_raw)
        formatted_partida = _format_quality_value(fecha_partida_raw)

        record = {
            'item': item,
            'produc': produc,
            'subgrupo': subgrupo,
            'oc_numero': oc_numero,
            'fecha_ing': formatted_ing,
            'fecha_ing_iso': fecha_ing_date.isoformat() if fecha_ing_date else '',
            'fecha_partida': formatted_partida,
            'fecha_partida_iso': fecha_partida_date.isoformat() if fecha_partida_date else '',
            'approved_qty': approved_qty,
            'total_qty': total_qty,
            'progress_pct': progress_pct,
            'is_approved': total_qty > 0 and approved_qty >= total_qty
        }
        records.append(record)

        if fecha_ing_date:
            full_key = (item, produc, oc_numero)
            product_key = (item, produc)
            pair_key = (item, produc)
            if pair_key not in ingress_pair or fecha_ing_date < ingress_pair[pair_key]:
                ingress_pair[pair_key] = fecha_ing_date
            if full_key not in ingress_full or fecha_ing_date < ingress_full[full_key]:
                ingress_full[full_key] = fecha_ing_date
            if product_key not in ingress_product or fecha_ing_date < ingress_product[product_key]:
                ingress_product[product_key] = fecha_ing_date
            if item and (item not in ingress_item or fecha_ing_date < ingress_item[item]):
                ingress_item[item] = fecha_ing_date

        full_key = (item, produc, oc_numero)
        product_key = (item, produc)
        pair_key = (item, produc)
        if subgrupo and pair_key not in subgroup_pair:
            subgroup_pair[pair_key] = subgrupo
        total_pair[pair_key] = max(total_qty, total_pair.get(pair_key, 0.0))
        total_full[full_key] = max(total_qty, total_full.get(full_key, 0.0))
        total_product[product_key] = max(total_qty, total_product.get(product_key, 0.0))
        if item:
            total_item[item] = max(total_qty, total_item.get(item, 0.0))

    return {
        'records': records,
        'ingress_pair': ingress_pair,
        'ingress_full': ingress_full,
        'ingress_product': ingress_product,
        'ingress_item': ingress_item,
        'subgroup_pair': subgroup_pair,
        'total_pair': total_pair,
        'total_full': total_full,
        'total_product': total_product,
        'total_item': total_item
    }


def _read_quality_pending_records():
    snapshot = _quality_build_control_snapshot()
    registry_items = _quality_register_snapshot_pairs(snapshot['records'])
    handler_state = _quality_load_pending_handlers_state()
    handler_assignments = handler_state.get('assignments') if isinstance(handler_state, dict) else {}
    handler_assignments = handler_assignments if isinstance(handler_assignments, dict) else {}
    records = []
    for record in snapshot['records']:
        key = _quality_tracking_key(record['item'], record['produc'])
        tracked = registry_items.get(key, {}) if isinstance(registry_items, dict) else {}
        row_key = _quality_pending_row_key(record['item'], record['produc'], record['oc_numero'], record['fecha_ing'])
        assigned_handler = handler_assignments.get(row_key) if isinstance(handler_assignments.get(row_key), dict) else {}
        records.append({
            'item': record['item'],
            'produc': record['produc'],
            'subgrupo': record.get('subgrupo', ''),
            'row_key': row_key,
            'oc_numero': record['oc_numero'],
            'observaciones': _format_quality_value(tracked.get('observacion')),
            'canti_real': _format_quality_value(record['total_qty']),
            'fecha_ing': record['fecha_ing'],
            'approved_qty': _format_quality_value(record['approved_qty']),
            'total_qty': _format_quality_value(record['total_qty']),
            'progress_pct': record['progress_pct'],
            'is_approved': record['is_approved'],
            'encargado': _format_quality_value(assigned_handler.get('name')),
            'encargado_updated_at': _format_quality_value(assigned_handler.get('updated_at')),
            'encargado_updated_by': _format_quality_value(assigned_handler.get('updated_by'))
        })
    return records


def _read_quality_history_records():
    _quality_sync_csv_exports()
    history_csv_mtime = _quality_get_mtime(QUALITY_SHEET_CACHE_MAP['Aprobaciones'])
    observations_mtime = _quality_get_mtime(QUALITY_OBSERVATIONS_FILE)

    with QUALITY_HISTORY_CACHE_LOCK:
        if (
            QUALITY_HISTORY_CACHE.get('data') is not None and
            QUALITY_HISTORY_CACHE.get('csv_mtime') == history_csv_mtime and
            QUALITY_HISTORY_CACHE.get('obs_mtime') == observations_mtime
        ):
            return QUALITY_HISTORY_CACHE['data']

    registry_items = _quality_load_observation_items()
    records = []
    csv_path = QUALITY_SHEET_CACHE_MAP['Aprobaciones']
    if not csv_path.exists():
        raise FileNotFoundError(f'Archivo no encontrado: {csv_path}')

    with csv_path.open('r', newline='', encoding='utf-8-sig') as f:
        reader = csv.reader(f)
        header_row = next(reader, None)
        if not header_row:
            return []

        header_map = {}
        for idx, raw_header in enumerate(header_row):
            normalized = _normalize_sheet_header(raw_header)
            if normalized and normalized not in header_map:
                header_map[normalized] = idx

        for row in reader:
            if not row or not any(cell is not None and str(cell).strip() != '' for cell in row):
                continue
            item = _format_quality_value(_quality_get_value(row, header_map, 'item'))
            produc = _format_quality_value(_quality_get_value(row, header_map, 'produc'))
            key = _quality_tracking_key(item, produc)
            tracked = registry_items.get(key, {}) if isinstance(registry_items, dict) else {}
            records.append({
                'item': item,
                'produc': produc,
                'ubi': _format_quality_value(_quality_get_value(row, header_map, 'ubi', 'Ubi', 'ubicacion', 'Ubicacion')),
                'observaciones': _format_quality_value(tracked.get('observacion')),
                'oc_numero': _format_quality_value(_quality_get_value(row, header_map, 'OC numero', 'OC Numero')),
                'canti_real': _format_quality_value(_quality_get_value(row, header_map, 'canti_real', 'CANTI_EXIS')),
                'fecha': _format_quality_value(_quality_get_value(row, header_map, 'fecha')),
                'name': _format_quality_value(_quality_get_value(row, header_map, 'name'))
            })

    with QUALITY_HISTORY_CACHE_LOCK:
        QUALITY_HISTORY_CACHE['csv_mtime'] = history_csv_mtime
        QUALITY_HISTORY_CACHE['obs_mtime'] = observations_mtime
        QUALITY_HISTORY_CACHE['data'] = records

    return records


def _read_quality_admin_stats():
    snapshot = _quality_build_control_snapshot()
    registry_items = _quality_register_snapshot_pairs(snapshot['records'])
    ingress_pair = snapshot['ingress_pair']
    subgroup_pair = snapshot.get('subgroup_pair', {})
    total_pair = snapshot['total_pair']

    history = []
    timed_records = 0

    for row, header_map in _quality_sheet_rows('Aprobaciones'):
        item = _format_quality_value(_quality_get_value(row, header_map, 'item'))
        produc = _format_quality_value(_quality_get_value(row, header_map, 'produc'))
        oc_numero = _format_quality_value(_quality_get_value(row, header_map, 'OC numero', 'OC Numero'))
        approval_raw = _quality_get_value(row, header_map, 'fecha')
        ingress_raw = _quality_get_value(row, header_map, 'Fecha_ing', 'fecha_ing', 'Fecha Ing')
        approval_date = _quality_parse_date(approval_raw)
        approval_iso = approval_date.isoformat() if approval_date else ''
        ingress_date = _quality_parse_date(ingress_raw)
        qty_num = _quality_parse_number(_quality_get_value(row, header_map, 'canti_real', 'CANTI_EXIS')) or 0.0
        author = _format_quality_value(_quality_get_value(row, header_map, 'name'))
        tracked = registry_items.get(_quality_tracking_key(item, produc), {}) if isinstance(registry_items, dict) else {}

        pair_key = (item, produc)
        match_level = 'approval' if ingress_date else ''
        if not ingress_date:
            ingress_date = ingress_pair.get(pair_key)
            match_level = 'pair' if ingress_date else ''

        total_qty = total_pair.get(pair_key)
        if total_qty is None:
            total_qty = 0.0

        control_days = None
        if ingress_date and approval_date:
            delta_days = (approval_date - ingress_date).days
            if delta_days >= 0:
                control_days = delta_days
                timed_records += 1

        history.append({
            'item': item,
            'produc': produc,
            'subgrupo': _format_quality_value(subgroup_pair.get(pair_key) or tracked.get('subgrupo')),
            'observaciones': _format_quality_value(tracked.get('observacion')),
            'oc_numero': oc_numero,
            'ubi': _format_quality_value(_quality_get_value(row, header_map, 'ubi', 'Ubi', 'ubicacion', 'Ubicacion')),
            'approval_date': approval_iso,
            'approval_date_label': _format_quality_value(approval_raw),
            'author': author,
            'controlled_qty': qty_num,
            'total_qty': total_qty,
            'control_days': control_days,
            'fecha_ing_iso': ingress_date.isoformat() if ingress_date else '',
            'match_level': match_level
        })

    pending_records = snapshot['records']
    pending_count = sum(1 for record in pending_records if not record['is_approved'])
    approved_snapshot_count = sum(1 for record in pending_records if record['is_approved'])

    return {
        'generated_at': datetime.now().isoformat(timespec='seconds'),
        'summary': {
            'pending_count': pending_count,
            'approved_snapshot_count': approved_snapshot_count,
            'pending_total': len(pending_records),
            'history_total': len(history),
            'timed_history_total': timed_records
        },
        'pending_records': pending_records,
        'history_records': history
    }


@app.route('/api/quality/pending', methods=['GET'])
def get_quality_pending_records():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        data = _read_quality_pending_records()
        return jsonify({'status': 'success', 'data': data})
    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 404
    except Exception as e:
        print(f"[QUALITY][PENDING] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/history', methods=['GET'])
def get_quality_history_records():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        data = _read_quality_history_records()
        return jsonify({'status': 'success', 'data': data})
    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 404
    except Exception as e:
        print(f"[QUALITY][HISTORY] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/observation', methods=['POST'])
def save_quality_observation():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    data = request.get_json(silent=True) or {}
    item = _format_quality_value(data.get('item'))
    produc = _format_quality_value(data.get('produc'))
    observacion = str(data.get('observacion') or '').strip()

    if not item and not produc:
        return jsonify({'status': 'error', 'message': 'item/produc requeridos'}), 400

    try:
        saved = _quality_update_observation(
            item=item,
            produc=produc,
            observacion=observacion,
            updated_by=session.get('user') or ''
        )
        return jsonify({'status': 'success', 'data': saved})
    except Exception as e:
        print(f"[QUALITY][OBS] Error saving observation: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-categories', methods=['GET'])
def get_quality_pending_categories():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        state = _quality_load_pending_categories_state()
        state['categories'] = _quality_sort_categories(state.get('categories', []))
        return jsonify({'status': 'success', 'data': state})
    except Exception as e:
        print(f"[QUALITY][CAT] Error loading categories: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-categories', methods=['POST'])
def create_quality_pending_category():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    data = request.get_json(silent=True) or {}
    name = str(data.get('name') or '').strip()
    color = _quality_normalize_category_color(data.get('color'))
    if not name:
        return jsonify({'status': 'error', 'message': 'Ingrese un nombre para la categoria.'}), 400

    try:
        with QUALITY_PENDING_CATEGORIES_LOCK:
            state = _quality_load_pending_categories_state()
            categories = state.setdefault('categories', [])
            if any(str(category.get('name') or '').strip().lower() == name.lower() for category in categories):
                return jsonify({'status': 'error', 'message': 'Ya existe una categoria con ese nombre.'}), 400

            category = {
                'id': uuid.uuid4().hex,
                'name': name,
                'color': color,
                'updated_at': datetime.now().isoformat(timespec='seconds'),
                'updated_by': _format_quality_value(session.get('user'))
            }
            categories.append(category)
            state['categories'] = _quality_sort_categories(categories)
            state = _quality_ensure_hidden_category(state)
            state['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_pending_categories_state(state)

        return jsonify({'status': 'success', 'data': category})
    except Exception as e:
        print(f"[QUALITY][CAT] Error creating category: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-categories/<category_id>', methods=['PUT'])
def update_quality_pending_category(category_id):
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    data = request.get_json(silent=True) or {}
    name = str(data.get('name') or '').strip()
    color = _quality_normalize_category_color(data.get('color'))
    if not category_id:
        return jsonify({'status': 'error', 'message': 'Categoria invalida.'}), 400
    if not name:
        return jsonify({'status': 'error', 'message': 'Ingrese un nombre para la categoria.'}), 400

    try:
        with QUALITY_PENDING_CATEGORIES_LOCK:
            state = _quality_load_pending_categories_state()
            categories = state.setdefault('categories', [])
            target = None
            for category in categories:
                if str(category.get('id')) == str(category_id):
                    target = category
                    break
            if not target:
                return jsonify({'status': 'error', 'message': 'Categoria no encontrada.'}), 404
            if str(category_id) == QUALITY_PENDING_HIDDEN_CATEGORY_ID:
                name = QUALITY_PENDING_HIDDEN_CATEGORY_NAME
            if any(
                str(category.get('id')) != str(category_id) and
                str(category.get('name') or '').strip().lower() == name.lower()
                for category in categories
            ):
                return jsonify({'status': 'error', 'message': 'Ya existe una categoria con ese nombre.'}), 400

            target['name'] = name
            target['color'] = color
            target['updated_at'] = datetime.now().isoformat(timespec='seconds')
            target['updated_by'] = _format_quality_value(session.get('user'))
            state['categories'] = _quality_sort_categories(categories)
            state = _quality_ensure_hidden_category(state)
            state['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_pending_categories_state(state)

        return jsonify({'status': 'success', 'data': target})
    except Exception as e:
        print(f"[QUALITY][CAT] Error updating category: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-categories/<category_id>', methods=['DELETE'])
def delete_quality_pending_category(category_id):
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        with QUALITY_PENDING_CATEGORIES_LOCK:
            state = _quality_load_pending_categories_state()
            categories = state.setdefault('categories', [])
            assignments = state.setdefault('assignments', {})
            if str(category_id) == QUALITY_PENDING_HIDDEN_CATEGORY_ID:
                return jsonify({'status': 'error', 'message': 'La categoria Ocultos no se puede borrar.'}), 400
            remaining = [category for category in categories if str(category.get('id')) != str(category_id)]
            if len(remaining) == len(categories):
                return jsonify({'status': 'error', 'message': 'Categoria no encontrada.'}), 404

            state['categories'] = _quality_sort_categories(remaining)
            state['assignments'] = {
                row_key: assigned_id
                for row_key, assigned_id in assignments.items()
                if str(assigned_id) != str(category_id)
            }
            state = _quality_ensure_hidden_category(state)
            state['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_pending_categories_state(state)

        return jsonify({'status': 'success'})
    except Exception as e:
        print(f"[QUALITY][CAT] Error deleting category: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-categories/assign', methods=['POST'])
def assign_quality_pending_category():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    data = request.get_json(silent=True) or {}
    row_key = str(data.get('row_key') or '').strip()
    category_id = str(data.get('category_id') or '').strip()
    if not row_key:
        return jsonify({'status': 'error', 'message': 'Fila invalida.'}), 400

    try:
        with QUALITY_PENDING_CATEGORIES_LOCK:
            state = _quality_load_pending_categories_state()
            categories = state.setdefault('categories', [])
            assignments = state.setdefault('assignments', {})

            if category_id:
                if not any(str(category.get('id')) == category_id for category in categories):
                    return jsonify({'status': 'error', 'message': 'Categoria no encontrada.'}), 404
                assignments[row_key] = category_id
            else:
                assignments.pop(row_key, None)

            state = _quality_ensure_hidden_category(state)
            state['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_pending_categories_state(state)

        return jsonify({'status': 'success', 'data': {'row_key': row_key, 'category_id': category_id}})
    except Exception as e:
        print(f"[QUALITY][CAT] Error assigning category: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/quality/pending-handler', methods=['POST'])
def assign_quality_pending_handler():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if not _can_access_quality_module():
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    data = request.get_json(silent=True) or {}
    row_key = str(data.get('row_key') or '').strip()
    if not row_key:
        return jsonify({'status': 'error', 'message': 'Fila invalida.'}), 400

    handler_name = _format_quality_value(data.get('name') or session.get('user') or '')
    if not handler_name:
        return jsonify({'status': 'error', 'message': 'Usuario invalido.'}), 400

    try:
        with QUALITY_PENDING_HANDLERS_LOCK:
            state = _quality_load_pending_handlers_state()
            assignments = state.setdefault('assignments', {})
            assignments[row_key] = {
                'name': handler_name,
                'updated_at': datetime.now().isoformat(timespec='seconds'),
                'updated_by': _format_quality_value(session.get('user'))
            }
            state['updated_at'] = datetime.now().isoformat(timespec='seconds')
            _quality_save_pending_handlers_state(state)

        return jsonify({'status': 'success', 'data': {'row_key': row_key, 'name': handler_name}})
    except Exception as e:
        print(f"[QUALITY][HANDLER] Error assigning handler: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/admin/quality-stats', methods=['GET'])
def get_admin_quality_stats():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401
    if session.get('role') != 'admin':
        return jsonify({"status": "error", "message": "No autorizado"}), 403

    try:
        data = _read_quality_admin_stats()
        return jsonify({'status': 'success', 'data': data})
    except FileNotFoundError as e:
        return jsonify({'status': 'error', 'message': str(e)}), 404
    except Exception as e:
        print(f"[QUALITY][ADMIN_STATS] Error: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500



@app.route('/api/activity-pending', methods=['GET'])
def get_pending_activity():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    # Resolver email real del usuario (en users.json el username puede no ser el email completo)
    raw_user = session.get('user')
    user_email = raw_user.strip().lower() if raw_user else ''
    try:
        if USERS_FILE.exists() and raw_user and '@' not in raw_user:
            with open(USERS_FILE, 'r', encoding='utf-8') as f:
                users_db = json.load(f)
            if raw_user in users_db:
                email_val = users_db[raw_user].get('email') or users_db[raw_user].get('correo')
                if email_val:
                    user_email = email_val.strip().lower()
    except Exception as e:
        print(f"[ACTIVITY] Warning resolving email: {e}")

    pending = []

    try:
        if not ACTIVITY_STATE_FILE.exists():
            return jsonify({'status': 'success', 'data': [], 'debug_user': user_email})

        with open(ACTIVITY_STATE_FILE, 'r', encoding='utf-8') as f:
            state = json.load(f)

        campaigns = state.get('campaigns', [])
        today = datetime.now().date()

        # Tomar todas las campaÃ±as con fecha <= hoy donde este usuario no tenga respuesta (ignorando recipients)
        for camp in campaigns:
            camp_date = camp.get('date')
            show_camp = False
            try:
                if camp_date:
                    d_obj = datetime.strptime(camp_date, '%Y-%m-%d').date()
                    
                    # 3 Working Days Filter
                    wdays = 0
                    curr = d_obj
                    while curr < today:
                         curr += timedelta(days=1)
                         if curr.weekday() < 5:
                             wdays += 1
                    
                    if wdays < 3:
                        show_camp = True
            except Exception:
                pass
            
            if not show_camp:
                continue

            responses = camp.get('responses', {})
            
            # Check if user is in recipients list
            recipients = [r.strip().lower() for r in camp.get('recipients', [])]
            if user_email not in recipients:
                continue

            if any(email_key.strip().lower() == user_email for email_key in responses):
                continue

            pending.append({
                'token': camp.get('token'),
                'date': camp.get('date'),
                'subject': camp.get('subject'),
                'type': 'activity_record',
                'debug_recipients': camp.get('recipients', [])
            })

    except Exception as e:
        print(f"[ACTIVITY] Error checking pending: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500

    return jsonify({'status': 'success', 'data': pending, 'debug_user': user_email, 'debug_pending_count': len(pending)})

@app.route('/api/activity-submit', methods=['POST'])
def submit_activity_record():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    data = request.json
    token = data.get('token') if data else None
    project = data.get('project') if data else None
    time_val = data.get('time') if data else None
    description = data.get('description') if data else None

    # Resolver email y nombre
    raw_user = session.get('user')
    user_email = raw_user.strip().lower() if raw_user else ''
    
    # 1. First, ensure we have the correct email from users.json if possible
    try:
        if USERS_FILE.exists():
            with open(USERS_FILE, 'r') as f:
                users_db = json.load(f)
            key = raw_user if raw_user in users_db else user_email
            if key in users_db:
                email_val = users_db[key].get('email') or users_db[key].get('correo')
                if email_val:
                    user_email = email_val.strip().lower()
    except Exception as e:
        print(f"[ACTIVITY] Warning resolving user email: {e}")

    user_display_name = None

    # 2. Try resolving name from DESTINATARIOS_FILE (Highest Priority)
    if DESTINATARIOS_FILE.exists():
         try:
             with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
                 reader = csv.reader(f)
                 # Expecting header: NOMBRE,APELLIDO,EMAIL
                 header = next(reader, None) 
                 for row in reader:
                     if len(row) >= 3:
                         csv_email = row[2].strip().lower()
                         if csv_email == user_email:
                             user_display_name = f"{row[0].strip()} {row[1].strip()}"
                             break
         except Exception as e:
             print(f"[ACTIVITY] Warning reading destinatarios.csv: {e}")

    # 3. Fallback to users.json (if present)
    if not user_display_name and USERS_FILE.exists():
        try:
             with open(USERS_FILE, 'r') as f:
                users_db = json.load(f)
             key = raw_user if raw_user in users_db else user_email
             if key in users_db:
                 user_display_name = users_db[key].get('display_name')
        except: pass

    # 4. Final Fallback: Format from email
    if not user_display_name:
        user_display_name = user_email.split('@')[0].replace('.', ' ').title() if '@' in user_email else user_email

    if not token or not project:
        return jsonify({'status': 'error', 'message': 'Datos incompletos'}), 400

    try:
        if not ACTIVITY_STATE_FILE.exists():
            return jsonify({'status': 'error', 'message': 'Archivo de estado no encontrado'}), 500

        with open(ACTIVITY_STATE_FILE, 'r', encoding='utf-8') as f:
            state = json.load(f)

        campaign_idx = -1
        target_camp = None
        for idx, camp in enumerate(state.get('campaigns', [])):
            if camp.get('token') == token:
                campaign_idx = idx
                target_camp = camp
                break

        if not target_camp:
            return jsonify({'status': 'error', 'message': 'Token invalido o expirado'}), 404

        now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        body_mock = f'Proyecto: {project}\nTiempo destinado: {time_val}\nDescripci?n: {description}'

        response_obj = {
            "from": user_email,
            "from_name": user_display_name,
            "subject": f"Re: {target_camp.get('subject')}",
            "date": now_str,
            "received_at": datetime.now().isoformat(),
            "snippet": body_mock,
            "body": body_mock,
            "uid": 0,
            "saved_path": "WEB_SUBMISSION"
        }

        state['campaigns'][campaign_idx].setdefault('responses', {})
        
        # Append logic if user already responded (to support multi-project submissions)
        if user_email in state['campaigns'][campaign_idx]['responses']:
            existing = state['campaigns'][campaign_idx]['responses'][user_email]
            existing['body'] += "\n----------------\n" + body_mock
            existing['snippet'] = existing['body']
            existing['received_at'] = response_obj['received_at']
        else:
            state['campaigns'][campaign_idx]['responses'][user_email] = response_obj

        _write_json_file_atomic(ACTIVITY_STATE_FILE, state, indent=2, ensure_ascii=False)

        def clean_csv(val):
            return str(val).replace(';', ',').replace('\\n', ' ').strip()

        date_str = target_camp.get('date')
        try:
            dobj = datetime.strptime(date_str, '%Y-%m-%d')
            date_str_csv = dobj.strftime('%d/%m/%Y')
        except:
            date_str_csv = date_str

        time_now_str = datetime.now().strftime('%H:%M:%S')

        # Separar multiples proyectos (coma, punto y coma o salto de linea)
        projects_list = [p.strip() for p in re.split(r"[;,\n]+", str(project or '')) if p.strip()]
        if not projects_list:
            projects_list = [str(project or '')]

        # Master CSV (una fila por proyecto)
        try:
            with open(ACTIVITY_CSV_BASE, 'a', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                for proj in projects_list:
                    master_row = [
                        date_str_csv,
                        time_now_str,
                        token,
                        user_display_name,
                        user_email,
                        clean_csv(proj),
                        clean_csv(proj),
                        clean_csv(time_val),
                        clean_csv(description)
                    ]
                    writer.writerow(master_row)
        except Exception as e:
            print(f"[ACTIVITY] Error writing Master CSV: {e}")

        # User CSV (una fila por proyecto)
        safe_name = re.sub(r'[<>:"/\|?*]', '_', user_display_name).strip()
        user_csv_path = ACTIVITY_CSV_USER_DIR / f"{safe_name}.csv"
        try:
            write_header = not user_csv_path.exists()
            with open(user_csv_path, 'a', encoding='utf-8-sig', newline='') as f:
                writer = csv.writer(f, delimiter=';')
                if write_header:
                    writer.writerow(["Fecha", "Hora", "Token", "ProyectoFinal", "Tiempo", "Registro", "Observaciones"])
                for proj in projects_list:
                    user_row = [
                        date_str_csv,
                        time_now_str,
                        token,
                        clean_csv(proj),
                        clean_csv(time_val),
                        clean_csv(description),
                        ""
                    ]
                    writer.writerow(user_row)
        except Exception as e:
            print(f"[ACTIVITY] Error writing User CSV: {e}")

        return jsonify({'status': 'success', 'message': 'Registro guardado correctamente'})

    except Exception as e:
        print(f"[ACTIVITY] Error submitting: {e}")
        return jsonify({'status': 'error', 'message': str(e)}), 500


@app.route('/api/activity-update', methods=['POST'])
def update_activity_entry():
    if not session.get('user'):
        return jsonify({"status": "error", "message": "No autenticado"}), 401

    data = request.get_json(silent=True) or {}
    token = str(data.get('token') or '').strip()
    new_desc = str(data.get('description') or '').strip()
    new_time = str(data.get('time') or '').strip()
    original_project = str(data.get('original_project') or '').strip()
    new_project = str(data.get('project') or '').strip()

    if not token or not new_desc or not new_time:
        return jsonify({"status": "error", "message": "Faltan datos requeridos"}), 400

    ACTIVITY_BASE = ACTIVITY_CODE_DIR
    CONFIG_DIR = ACTIVITY_BASE / "config"
    DATA_DIR = ACTIVITY_BASE / "data/respuestas_csv"
    DESTINATARIOS_FILE = CONFIG_DIR / "destinatarios.csv"

    def sanitize_display_name(name):
        return re.sub(r'[<>:"/\|?*]', '_', str(name or '')).strip()

    def unique_values(values):
        result = []
        seen = set()
        for v in values:
            key = str(v or '').strip()
            if not key:
                continue
            lower_key = key.lower()
            if lower_key in seen:
                continue
            seen.add(lower_key)
            result.append(key)
        return result

    def read_rows_with_fallback(path_obj):
        last_error = None
        for enc in ('utf-8-sig', 'utf-8', 'latin-1'):
            try:
                with open(path_obj, 'r', encoding=enc, newline='') as f:
                    reader = csv.reader(f, delimiter=';')
                    return list(reader), enc
            except UnicodeDecodeError as e:
                last_error = e
                continue
        if last_error:
            raise last_error
        return [], 'utf-8-sig'

    raw_user = session.get('user')
    user_email = raw_user.strip().lower() if raw_user else ''
    user_display_names = []

    try:
        if USERS_FILE.exists():
            try:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    users_db = json.load(f)
                key = raw_user if raw_user in users_db else user_email
                if key in users_db:
                    email_val = users_db[key].get('email') or users_db[key].get('correo')
                    if email_val:
                        user_email = str(email_val).strip().lower()
            except Exception as e:
                print(f"[ACTIVITY-UPDATE] Warning resolving user email: {e}")

        if not DESTINATARIOS_FILE.exists():
            return jsonify({'status': 'error', 'message': 'Configuracion no encontrada.'}), 500

        with open(DESTINATARIOS_FILE, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            for row in reader:
                if row.get('EMAIL', '').strip().lower() == str(user_email).strip().lower():
                    first_name = row.get('NOMBRE', '').strip()
                    last_name = row.get('APELLIDO', '').strip()
                    full_name = f"{first_name} {last_name}".strip()
                    if full_name:
                        user_display_names.append(full_name)

        if USERS_FILE.exists():
            try:
                with open(USERS_FILE, 'r', encoding='utf-8') as f:
                    users_db = json.load(f)
                key = raw_user if raw_user in users_db else user_email
                if key in users_db:
                    display_name = users_db[key].get('display_name')
                    if display_name:
                        user_display_names.append(str(display_name).strip())
            except Exception as e:
                print(f"[ACTIVITY-UPDATE] Warning resolving display name: {e}")

        if '@' in user_email:
            user_display_names.append(user_email.split('@')[0].replace('.', ' ').title())
        elif user_email:
            user_display_names.append(user_email)

        candidate_names = unique_values(user_display_names)
        candidate_filenames = unique_values(
            [f"{name}.csv" for name in candidate_names] +
            [f"{sanitize_display_name(name)}.csv" for name in candidate_names]
        )

        csv_path = None
        for filename in candidate_filenames:
            candidate_path = DATA_DIR / filename
            if candidate_path.exists():
                csv_path = candidate_path
                break

        if not csv_path:
            return jsonify({'status': 'error', 'message': 'Archivo de registros no encontrado.'}), 404

        rows, encoding_used = read_rows_with_fallback(csv_path)
        if not rows:
            return jsonify({'status': 'error', 'message': 'Archivo de registros vacio.'}), 404

        header = rows[0]
        body = rows[1:]
        updated_rows = [header]
        found = False

        def ensure_len(row_obj, idx):
            while len(row_obj) <= idx:
                row_obj.append("")
            return row_obj

        for row in body:
            row_token = row[2].strip() if len(row) > 2 else ''
            row_project = row[3].strip() if len(row) > 3 else ''

            row_matches = row_token == token
            if row_matches and original_project:
                row_matches = row_project == original_project
            elif row_matches and found:
                row_matches = False

            if row_matches:
                row = ensure_len(row, 6)
                if new_project:
                    row[3] = new_project
                row[4] = new_time
                row[5] = new_desc

                timestamp = datetime.now().strftime("%d/%m %H:%M")
                mod_note = f"[Modificado el {timestamp}]"
                if row[6].strip():
                    row[6] = f"{row[6]} | {mod_note}"
                else:
                    row[6] = mod_note

                found = True

            updated_rows.append(row)

        if not found:
            return jsonify({'status': 'error', 'message': 'Registro no encontrado.'}), 404

        write_encoding = 'latin-1' if encoding_used == 'latin-1' else 'utf-8-sig'
        with open(csv_path, 'w', encoding=write_encoding, newline='') as f:
            writer = csv.writer(f, delimiter=';')
            writer.writerows(updated_rows)

        return jsonify({'status': 'success', 'message': 'Registro actualizado correctamente.'})

    except Exception as e:
        print(f"[ACTIVITY-UPDATE] Error: {e}")
        return jsonify({'status': 'error', 'message': f'Error interno: {str(e)}'}), 500

if __name__ == '__main__':
    import webbrowser
    from threading import Timer
    import os

    server_host = os.environ.get("BPB_SERVER_HOST", "0.0.0.0")
    browser_host = os.environ.get("BPB_BROWSER_HOST", "127.0.0.1")
    server_port = int(os.environ.get("BPB_SERVER_PORT", "5000"))
    browser_url = f"http://{browser_host}:{server_port}"

    # Only open browser in the main process
    if os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        pass

    def open_browser():
        try:
            # Try to find Chrome
            chrome_path = "C:/Program Files/Google/Chrome/Application/chrome.exe %s"
            if not os.path.exists("C:/Program Files/Google/Chrome/Application/chrome.exe"):
                chrome_path = "C:/Program Files (x86)/Google/Chrome/Application/chrome.exe %s"
            
            webbrowser.get(chrome_path).open(browser_url)
        except Exception:
            webbrowser.open(browser_url)

    Timer(1.5, open_browser).start()

    try:
        from waitress import serve
        print("----------------------------------------------------------------")
        print(" INICIANDO SERVIDOR DE PRODUCCION (WAITRESS)")
        print(" Workers (Threads): 5")
        print(f" URL: http://{server_host}:{server_port}")
        print("----------------------------------------------------------------")
        serve(app, host=server_host, port=server_port, threads=5)
    except ImportError:
        print("Waitress no encontrado, instalando fallback...")
        app.run(host=server_host, port=server_port, debug=False)
