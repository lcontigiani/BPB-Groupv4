from __future__ import annotations

import json
import re
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import openpyxl

try:
    import psycopg  # type: ignore
except Exception:
    psycopg = None

try:
    import psycopg2  # type: ignore
except Exception:
    psycopg2 = None


def _utc_now() -> datetime:
    return datetime.now(timezone.utc)


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    return re.sub(r"\s+", " ", text)


def _format_value(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, date):
        return value.strftime("%Y-%m-%d")
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _parse_date(value: Any) -> Optional[date]:
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    raw = str(value).strip()
    if not raw:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(raw, fmt).date()
        except Exception:
            continue
    return None


def _parse_number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(str(value).replace(",", ".").strip())
    except Exception:
        return 0.0


def _tracking_key(item: str, produc: str) -> str:
    return f"{str(item or '').strip().lower()}||{str(produc or '').strip().lower()}"


def _row_key(item: str, produc: str, oc_numero: str = "", fecha_ing: str = "") -> str:
    return "||".join(
        [
            str(item or "").strip().lower(),
            str(produc or "").strip().lower(),
            str(oc_numero or "").strip().lower(),
            str(fecha_ing or "").strip().lower(),
        ]
    )


def _is_ar7_item(item: str) -> bool:
    return bool(re.fullmatch(r"AR7\d+", _format_value(item).upper()))


def _load_json(path: Path, default: Any) -> Any:
    if not path.exists():
        return default
    try:
        with path.open("r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return default


def _ensure_hidden_category(state: Dict[str, Any]) -> Dict[str, Any]:
    categories = state.get("categories") if isinstance(state.get("categories"), list) else []
    assignments = state.get("assignments") if isinstance(state.get("assignments"), dict) else {}
    hidden = None
    regular: List[Dict[str, Any]] = []
    migrated_ids = set()
    for category in categories:
        if not isinstance(category, dict):
            continue
        category_id = str(category.get("id") or "").strip()
        category_name = str(category.get("name") or "").strip()
        is_hidden = category_id == "hidden" or category_name.lower() == "ocultos"
        if is_hidden:
            migrated_ids.add(category_id)
            if hidden is None:
                hidden = {
                    "id": "hidden",
                    "name": "Ocultos",
                    "color": str(category.get("color") or "#5c6370").strip() or "#5c6370",
                    "updated_at": str(category.get("updated_at") or ""),
                    "updated_by": str(category.get("updated_by") or ""),
                }
            continue
        regular.append(category)
    if hidden is None:
        hidden = {"id": "hidden", "name": "Ocultos", "color": "#5c6370", "updated_at": "", "updated_by": ""}
    if migrated_ids:
        assignments = {
            row_key: ("hidden" if str(assigned_id or "").strip() in migrated_ids else assigned_id)
            for row_key, assigned_id in assignments.items()
        }
    regular.sort(key=lambda c: str(c.get("name") or "").strip().lower())
    return {"categories": [hidden, *regular], "assignments": assignments}


@dataclass
class QualitySyncResult:
    pending_locations: int
    pending_snapshot: int
    approvals: int


class QualityPostgresStore:
    def __init__(
        self,
        dsn: str,
        workbook_path: Path,
        observations_file: Path,
        categories_file: Path,
        handlers_file: Path,
    ) -> None:
        self.dsn = str(dsn or "").strip()
        self.workbook_path = Path(workbook_path)
        self.observations_file = Path(observations_file)
        self.categories_file = Path(categories_file)
        self.handlers_file = Path(handlers_file)
        self.driver = self._resolve_driver()

    @staticmethod
    def is_driver_available() -> bool:
        return psycopg is not None or psycopg2 is not None

    @staticmethod
    def driver_name() -> str:
        if psycopg is not None:
            return "psycopg"
        if psycopg2 is not None:
            return "psycopg2"
        return ""

    def _resolve_driver(self) -> str:
        if psycopg is not None:
            return "psycopg"
        if psycopg2 is not None:
            return "psycopg2"
        raise RuntimeError("No PostgreSQL driver available. Install psycopg or psycopg2.")

    def connect(self):
        if self.driver == "psycopg":
            conn = psycopg.connect(self.dsn)
            conn.autocommit = False
            return conn
        conn = psycopg2.connect(self.dsn)
        conn.autocommit = False
        return conn

    def _fetch_all(self, conn, sql: str, params: Optional[Iterable[Any]] = None) -> List[Dict[str, Any]]:
        with conn.cursor() as cur:
            cur.execute(sql, tuple(params or ()))
            rows = cur.fetchall()
            columns = [desc[0] for desc in cur.description] if cur.description else []
        return [dict(zip(columns, row)) for row in rows]

    def _fetch_one(self, conn, sql: str, params: Optional[Iterable[Any]] = None) -> Optional[Dict[str, Any]]:
        rows = self._fetch_all(conn, sql, params)
        return rows[0] if rows else None

    def ensure_schema(self) -> None:
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_meta (
                        key TEXT PRIMARY KEY,
                        value TEXT NOT NULL
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_sync_runs (
                        run_id TEXT PRIMARY KEY,
                        status TEXT NOT NULL,
                        trigger_source TEXT NOT NULL,
                        triggered_by TEXT NOT NULL,
                        started_at TIMESTAMPTZ NOT NULL,
                        finished_at TIMESTAMPTZ NULL,
                        rows_pending_locations INTEGER NOT NULL DEFAULT 0,
                        rows_pending_snapshot INTEGER NOT NULL DEFAULT 0,
                        rows_approvals INTEGER NOT NULL DEFAULT 0,
                        workbook_mtime DOUBLE PRECISION NULL,
                        error_message TEXT NOT NULL DEFAULT ''
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_observations (
                        item TEXT NOT NULL,
                        produc TEXT NOT NULL,
                        subgrupo TEXT NOT NULL DEFAULT '',
                        observacion TEXT NOT NULL DEFAULT '',
                        updated_at TIMESTAMPTZ NULL,
                        updated_by TEXT NOT NULL DEFAULT '',
                        PRIMARY KEY (item, produc)
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_pending_categories (
                        category_id TEXT PRIMARY KEY,
                        name TEXT NOT NULL UNIQUE,
                        color TEXT NOT NULL,
                        updated_at TIMESTAMPTZ NULL,
                        updated_by TEXT NOT NULL DEFAULT ''
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_pending_category_assignments (
                        row_key TEXT PRIMARY KEY,
                        category_id TEXT NOT NULL REFERENCES quality_pending_categories(category_id) ON DELETE CASCADE,
                        updated_at TIMESTAMPTZ NULL,
                        updated_by TEXT NOT NULL DEFAULT ''
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_pending_handlers (
                        row_key TEXT PRIMARY KEY,
                        handler_name TEXT NOT NULL,
                        updated_at TIMESTAMPTZ NULL,
                        updated_by TEXT NOT NULL DEFAULT ''
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_pending_locations (
                        location_key TEXT PRIMARY KEY,
                        row_key TEXT NOT NULL,
                        item TEXT NOT NULL,
                        produc TEXT NOT NULL,
                        subgrupo TEXT NOT NULL DEFAULT '',
                        oc_numero TEXT NOT NULL DEFAULT '',
                        fecha_ing TEXT NOT NULL DEFAULT '',
                        fecha_ing_iso TEXT NOT NULL DEFAULT '',
                        fecha_partida TEXT NOT NULL DEFAULT '',
                        fecha_partida_iso TEXT NOT NULL DEFAULT '',
                        ubi TEXT NOT NULL DEFAULT '',
                        total_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        remaining_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        approved_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        progress_pct INTEGER NOT NULL DEFAULT 0,
                        is_approved BOOLEAN NOT NULL DEFAULT FALSE,
                        is_ar7_item BOOLEAN NOT NULL DEFAULT FALSE,
                        imported_at TIMESTAMPTZ NOT NULL
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_pending_snapshot (
                        row_key TEXT PRIMARY KEY,
                        item TEXT NOT NULL,
                        produc TEXT NOT NULL,
                        subgrupo TEXT NOT NULL DEFAULT '',
                        oc_numero TEXT NOT NULL DEFAULT '',
                        fecha_ing TEXT NOT NULL DEFAULT '',
                        fecha_ing_iso TEXT NOT NULL DEFAULT '',
                        total_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        remaining_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        approved_qty DOUBLE PRECISION NOT NULL DEFAULT 0,
                        progress_pct INTEGER NOT NULL DEFAULT 0,
                        is_approved BOOLEAN NOT NULL DEFAULT FALSE,
                        is_ar7_item BOOLEAN NOT NULL DEFAULT FALSE,
                        imported_at TIMESTAMPTZ NOT NULL
                    )
                    """
                )
                cur.execute(
                    """
                    CREATE TABLE IF NOT EXISTS quality_approvals (
                        approval_key TEXT PRIMARY KEY,
                        item TEXT NOT NULL,
                        produc TEXT NOT NULL,
                        ubi TEXT NOT NULL DEFAULT '',
                        oc_numero TEXT NOT NULL DEFAULT '',
                        canti_real DOUBLE PRECISION NOT NULL DEFAULT 0,
                        fecha TEXT NOT NULL DEFAULT '',
                        fecha_iso TEXT NOT NULL DEFAULT '',
                        fecha_ing TEXT NOT NULL DEFAULT '',
                        fecha_ing_iso TEXT NOT NULL DEFAULT '',
                        author_name TEXT NOT NULL DEFAULT '',
                        imported_at TIMESTAMPTZ NOT NULL
                    )
                    """
                )
                cur.execute(
                    "INSERT INTO quality_meta(key, value) VALUES (%s, %s) ON CONFLICT (key) DO NOTHING",
                    ("schema_version", "1"),
                )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def validate_connection(self) -> None:
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute("SELECT 1")
                cur.fetchone()
            conn.commit()
        finally:
            conn.close()

    def bootstrap_json_state(self) -> None:
        conn = self.connect()
        try:
            self.ensure_schema()
            with conn.cursor() as cur:
                cur.execute("SELECT COUNT(*) FROM quality_observations")
                obs_count = int(cur.fetchone()[0] or 0)
                cur.execute("SELECT COUNT(*) FROM quality_pending_categories")
                cat_count = int(cur.fetchone()[0] or 0)
                cur.execute("SELECT COUNT(*) FROM quality_pending_handlers")
                handler_count = int(cur.fetchone()[0] or 0)

                if obs_count == 0:
                    data = _load_json(self.observations_file, {"items": {}})
                    for _, entry in (data.get("items") or {}).items():
                        if not isinstance(entry, dict):
                            continue
                        cur.execute(
                            """
                            INSERT INTO quality_observations(item, produc, subgrupo, observacion, updated_at, updated_by)
                            VALUES (%s, %s, %s, %s, %s, %s)
                            ON CONFLICT (item, produc) DO NOTHING
                            """,
                            (
                                _format_value(entry.get("item")),
                                _format_value(entry.get("produc")),
                                _format_value(entry.get("subgrupo")),
                                _format_value(entry.get("observacion")),
                                _format_value(entry.get("updated_at")) or None,
                                _format_value(entry.get("updated_by")),
                            ),
                        )

                if cat_count == 0:
                    state = _ensure_hidden_category(_load_json(self.categories_file, {"categories": [], "assignments": {}}))
                    for category in state.get("categories") or []:
                        if not isinstance(category, dict):
                            continue
                        cur.execute(
                            """
                            INSERT INTO quality_pending_categories(category_id, name, color, updated_at, updated_by)
                            VALUES (%s, %s, %s, %s, %s)
                            ON CONFLICT (category_id) DO NOTHING
                            """,
                            (
                                _format_value(category.get("id")),
                                _format_value(category.get("name")),
                                _format_value(category.get("color")) or "#4aa3ff",
                                _format_value(category.get("updated_at")) or None,
                                _format_value(category.get("updated_by")),
                            ),
                        )
                    for rk, category_id in (state.get("assignments") or {}).items():
                        cur.execute(
                            """
                            INSERT INTO quality_pending_category_assignments(row_key, category_id, updated_at, updated_by)
                            VALUES (%s, %s, NULL, '')
                            ON CONFLICT (row_key) DO NOTHING
                            """,
                            (_format_value(rk), _format_value(category_id)),
                        )

                if handler_count == 0:
                    state = _load_json(self.handlers_file, {"assignments": {}})
                    for rk, handler in (state.get("assignments") or {}).items():
                        if not isinstance(handler, dict):
                            continue
                        cur.execute(
                            """
                            INSERT INTO quality_pending_handlers(row_key, handler_name, updated_at, updated_by)
                            VALUES (%s, %s, %s, %s)
                            ON CONFLICT (row_key) DO NOTHING
                            """,
                            (
                                _format_value(rk),
                                _format_value(handler.get("name")),
                                _format_value(handler.get("updated_at")) or None,
                                _format_value(handler.get("updated_by")),
                            ),
                        )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def create_sync_run(self, trigger_source: str, triggered_by: str) -> str:
        run_id = uuid.uuid4().hex
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO quality_sync_runs(
                        run_id, status, trigger_source, triggered_by, started_at, finished_at,
                        rows_pending_locations, rows_pending_snapshot, rows_approvals, workbook_mtime, error_message
                    ) VALUES (%s, %s, %s, %s, %s, NULL, 0, 0, 0, NULL, '')
                    """,
                    (run_id, "running", trigger_source, triggered_by, _utc_now()),
                )
            conn.commit()
            return run_id
        finally:
            conn.close()

    def update_sync_run_success(self, run_id: str, result: QualitySyncResult, workbook_mtime: Optional[float]) -> None:
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE quality_sync_runs
                    SET status = %s,
                        finished_at = %s,
                        rows_pending_locations = %s,
                        rows_pending_snapshot = %s,
                        rows_approvals = %s,
                        workbook_mtime = %s,
                        error_message = ''
                    WHERE run_id = %s
                    """,
                    (
                        "success",
                        _utc_now(),
                        int(result.pending_locations),
                        int(result.pending_snapshot),
                        int(result.approvals),
                        workbook_mtime,
                        run_id,
                    ),
                )
            conn.commit()
        finally:
            conn.close()

    def update_sync_run_failure(self, run_id: str, error_message: str, workbook_mtime: Optional[float]) -> None:
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE quality_sync_runs
                    SET status = %s,
                        finished_at = %s,
                        workbook_mtime = %s,
                        error_message = %s
                    WHERE run_id = %s
                    """,
                    ("failed", _utc_now(), workbook_mtime, str(error_message or "")[:4000], run_id),
                )
            conn.commit()
        finally:
            conn.close()

    def get_latest_sync_status(self) -> Dict[str, Any]:
        conn = self.connect()
        try:
            row = self._fetch_one(
                conn,
                """
                SELECT run_id, status, trigger_source, triggered_by, started_at, finished_at,
                       rows_pending_locations, rows_pending_snapshot, rows_approvals, workbook_mtime, error_message
                FROM quality_sync_runs
                ORDER BY started_at DESC
                LIMIT 1
                """,
            )
            snapshot_row = self._fetch_one(conn, "SELECT COUNT(*) AS total FROM quality_pending_snapshot")
            return {
                "run_id": row.get("run_id") if row else "",
                "status": row.get("status") if row else "idle",
                "trigger_source": row.get("trigger_source") if row else "",
                "triggered_by": row.get("triggered_by") if row else "",
                "started_at": row.get("started_at").isoformat() if row and row.get("started_at") else "",
                "finished_at": row.get("finished_at").isoformat() if row and row.get("finished_at") else "",
                "rows_pending_locations": int((row or {}).get("rows_pending_locations") or 0),
                "rows_pending_snapshot": int((row or {}).get("rows_pending_snapshot") or 0),
                "rows_approvals": int((row or {}).get("rows_approvals") or 0),
                "workbook_mtime": (row or {}).get("workbook_mtime"),
                "error": (row or {}).get("error_message") or "",
                "snapshot_count": int((snapshot_row or {}).get("total") or 0),
            }
        finally:
            conn.close()

    def should_sync(self, max_age_seconds: int) -> bool:
        status = self.get_latest_sync_status()
        if status.get("status") == "running":
            return False
        if int(status.get("snapshot_count") or 0) <= 0:
            return True
        finished_at_raw = status.get("finished_at") or ""
        if not finished_at_raw:
            return True
        try:
            finished_at = datetime.fromisoformat(str(finished_at_raw).replace("Z", "+00:00"))
        except Exception:
            return True
        age = _utc_now() - finished_at.astimezone(timezone.utc)
        return age.total_seconds() > max_age_seconds

    def _sheet_rows(self, workbook, sheet_name: str) -> List[Tuple[Tuple[Any, ...], Dict[str, int], int]]:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(f"Hoja no encontrada: {sheet_name}")
        ws = workbook[sheet_name]
        rows = ws.iter_rows(values_only=True)
        header_row = next(rows, None)
        if not header_row:
            return []
        header_map: Dict[str, int] = {}
        for idx, raw_header in enumerate(header_row):
            normalized = _normalize_header(raw_header)
            if normalized and normalized not in header_map:
                header_map[normalized] = idx
        data_rows: List[Tuple[Tuple[Any, ...], Dict[str, int], int]] = []
        for sheet_row_index, row in enumerate(rows, start=2):
            if not row or not any(cell is not None and str(cell).strip() != "" for cell in row):
                continue
            data_rows.append((tuple(row), header_map, sheet_row_index))
        return data_rows

    def _get_value(self, row: Tuple[Any, ...], header_map: Dict[str, int], *header_names: str) -> Any:
        for header_name in header_names:
            idx = header_map.get(_normalize_header(header_name))
            if idx is not None and idx < len(row):
                return row[idx]
        return None

    def _build_pending_payload(self, workbook) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
        pending_locations: List[Dict[str, Any]] = []
        grouped: Dict[str, Dict[str, Any]] = {}
        for row, header_map, sheet_row_index in self._sheet_rows(workbook, "Control de Calidad"):
            item = _format_value(self._get_value(row, header_map, "item"))
            produc = _format_value(self._get_value(row, header_map, "produc"))
            subgrupo = _format_value(self._get_value(row, header_map, "SubGrupo", "Sub Grupo", "Subgrupo"))
            oc_numero = _format_value(self._get_value(row, header_map, "OC numero", "OC Numero"))
            canti_real = _parse_number(self._get_value(row, header_map, "canti_real"))
            canti_real_item = _parse_number(self._get_value(row, header_map, "canti_real_item"))
            fecha_ing_raw = self._get_value(row, header_map, "Fecha_ing", "fecha_ing", "Fecha Ing")
            fecha_partida_raw = self._get_value(row, header_map, "Fecha Partida")
            ubi = _format_value(self._get_value(row, header_map, "ubi", "Ubi", "ubicacion", "Ubicacion"))

            fecha_ing_date = _parse_date(fecha_ing_raw)
            fecha_partida_date = _parse_date(fecha_partida_raw)
            fecha_ing = _format_value(fecha_ing_raw)
            fecha_partida = _format_value(fecha_partida_raw)
            total_qty = max(canti_real, 0.0)
            remaining_qty = max(canti_real_item, 0.0)
            if total_qty > 0:
                remaining_qty = min(remaining_qty, total_qty)
            approved_qty = max(total_qty - remaining_qty, 0.0)
            progress_pct = 0
            if total_qty > 0:
                raw_progress = (approved_qty / total_qty) * 100
                progress_pct = 100 if approved_qty >= total_qty else min(99, int(raw_progress))
            row_key = _row_key(item, produc, oc_numero, fecha_ing)
            location = {
                "location_key": f"{row_key}||{sheet_row_index}",
                "row_key": row_key,
                "item": item,
                "produc": produc,
                "subgrupo": subgrupo,
                "oc_numero": oc_numero,
                "fecha_ing": fecha_ing,
                "fecha_ing_iso": fecha_ing_date.isoformat() if fecha_ing_date else "",
                "fecha_partida": fecha_partida,
                "fecha_partida_iso": fecha_partida_date.isoformat() if fecha_partida_date else "",
                "ubi": ubi,
                "total_qty": total_qty,
                "remaining_qty": remaining_qty,
                "approved_qty": approved_qty,
                "progress_pct": progress_pct,
                "is_approved": total_qty > 0 and approved_qty >= total_qty,
                "is_ar7_item": _is_ar7_item(item),
            }
            pending_locations.append(location)

            bucket = grouped.get(row_key)
            if bucket is None:
                bucket = {
                    "row_key": row_key,
                    "item": item,
                    "produc": produc,
                    "subgrupo": subgrupo,
                    "oc_numero": oc_numero,
                    "fecha_ing": fecha_ing,
                    "fecha_ing_iso": fecha_ing_date.isoformat() if fecha_ing_date else "",
                    "total_qty": total_qty,
                    "remaining_qty": 0.0,
                    "signatures": set(),
                    "is_ar7_item": _is_ar7_item(item),
                }
                grouped[row_key] = bucket
            bucket["total_qty"] = max(float(bucket.get("total_qty") or 0.0), total_qty)
            if not bucket.get("subgrupo") and subgrupo:
                bucket["subgrupo"] = subgrupo
            signature = (ubi, remaining_qty, total_qty)
            if signature not in bucket["signatures"]:
                bucket["signatures"].add(signature)
                bucket["remaining_qty"] += remaining_qty

        pending_snapshot: List[Dict[str, Any]] = []
        for bucket in grouped.values():
            total_qty = max(float(bucket.get("total_qty") or 0.0), 0.0)
            remaining_qty = max(float(bucket.get("remaining_qty") or 0.0), 0.0)
            if total_qty > 0:
                remaining_qty = min(remaining_qty, total_qty)
            approved_qty = max(total_qty - remaining_qty, 0.0)
            progress_pct = 0
            if total_qty > 0:
                raw_progress = (approved_qty / total_qty) * 100
                progress_pct = 100 if approved_qty >= total_qty else min(99, int(raw_progress))
            pending_snapshot.append(
                {
                    "row_key": bucket["row_key"],
                    "item": bucket["item"],
                    "produc": bucket["produc"],
                    "subgrupo": bucket.get("subgrupo", ""),
                    "oc_numero": bucket["oc_numero"],
                    "fecha_ing": bucket["fecha_ing"],
                    "fecha_ing_iso": bucket.get("fecha_ing_iso", ""),
                    "total_qty": total_qty,
                    "remaining_qty": remaining_qty,
                    "approved_qty": approved_qty,
                    "progress_pct": progress_pct,
                    "is_approved": total_qty > 0 and approved_qty >= total_qty,
                    "is_ar7_item": bool(bucket.get("is_ar7_item")),
                }
            )
        return pending_locations, pending_snapshot

    def _build_approvals_payload(self, workbook) -> List[Dict[str, Any]]:
        approvals: List[Dict[str, Any]] = []
        for row, header_map, sheet_row_index in self._sheet_rows(workbook, "Aprobaciones"):
            item = _format_value(self._get_value(row, header_map, "item"))
            produc = _format_value(self._get_value(row, header_map, "produc"))
            ubi = _format_value(self._get_value(row, header_map, "ubi", "Ubi", "ubicacion", "Ubicacion"))
            oc_numero = _format_value(self._get_value(row, header_map, "OC numero", "OC Numero"))
            canti_real = _parse_number(self._get_value(row, header_map, "canti_real", "CANTI_EXIS"))
            fecha_raw = self._get_value(row, header_map, "fecha")
            fecha_ing_raw = self._get_value(row, header_map, "Fecha_ing", "fecha_ing", "Fecha Ing")
            author_name = _format_value(self._get_value(row, header_map, "name"))
            fecha_date = _parse_date(fecha_raw)
            fecha_ing_date = _parse_date(fecha_ing_raw)
            fecha = _format_value(fecha_raw)
            fecha_ing = _format_value(fecha_ing_raw)
            approvals.append(
                {
                    "approval_key": f"{_tracking_key(item, produc)}||{oc_numero.lower()}||{fecha.lower()}||{ubi.lower()}||{sheet_row_index}",
                    "item": item,
                    "produc": produc,
                    "ubi": ubi,
                    "oc_numero": oc_numero,
                    "canti_real": max(canti_real, 0.0),
                    "fecha": fecha,
                    "fecha_iso": fecha_date.isoformat() if fecha_date else "",
                    "fecha_ing": fecha_ing,
                    "fecha_ing_iso": fecha_ing_date.isoformat() if fecha_ing_date else "",
                    "author_name": author_name,
                }
            )
        return approvals

    def sync_from_workbook(self) -> QualitySyncResult:
        self.ensure_schema()
        imported_at = _utc_now()
        workbook = openpyxl.load_workbook(str(self.workbook_path), read_only=True, data_only=True)
        try:
            pending_locations, pending_snapshot = self._build_pending_payload(workbook)
            approvals = self._build_approvals_payload(workbook)
        finally:
            workbook.close()

        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM quality_pending_locations")
                cur.execute("DELETE FROM quality_pending_snapshot")
                cur.execute("DELETE FROM quality_approvals")

                if pending_locations:
                    cur.executemany(
                        """
                        INSERT INTO quality_pending_locations(
                            location_key, row_key, item, produc, subgrupo, oc_numero, fecha_ing, fecha_ing_iso,
                            fecha_partida, fecha_partida_iso, ubi, total_qty, remaining_qty, approved_qty,
                            progress_pct, is_approved, is_ar7_item, imported_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            (
                                row["location_key"], row["row_key"], row["item"], row["produc"], row["subgrupo"],
                                row["oc_numero"], row["fecha_ing"], row["fecha_ing_iso"], row["fecha_partida"],
                                row["fecha_partida_iso"], row["ubi"], row["total_qty"], row["remaining_qty"],
                                row["approved_qty"], row["progress_pct"], row["is_approved"], row["is_ar7_item"], imported_at,
                            )
                            for row in pending_locations
                        ],
                    )

                if pending_snapshot:
                    cur.executemany(
                        """
                        INSERT INTO quality_pending_snapshot(
                            row_key, item, produc, subgrupo, oc_numero, fecha_ing, fecha_ing_iso,
                            total_qty, remaining_qty, approved_qty, progress_pct, is_approved, is_ar7_item, imported_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            (
                                row["row_key"], row["item"], row["produc"], row["subgrupo"], row["oc_numero"],
                                row["fecha_ing"], row["fecha_ing_iso"], row["total_qty"], row["remaining_qty"],
                                row["approved_qty"], row["progress_pct"], row["is_approved"], row["is_ar7_item"], imported_at,
                            )
                            for row in pending_snapshot
                        ],
                    )

                if approvals:
                    cur.executemany(
                        """
                        INSERT INTO quality_approvals(
                            approval_key, item, produc, ubi, oc_numero, canti_real, fecha, fecha_iso,
                            fecha_ing, fecha_ing_iso, author_name, imported_at
                        ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                        """,
                        [
                            (
                                row["approval_key"], row["item"], row["produc"], row["ubi"], row["oc_numero"],
                                row["canti_real"], row["fecha"], row["fecha_iso"], row["fecha_ing"],
                                row["fecha_ing_iso"], row["author_name"], imported_at,
                            )
                            for row in approvals
                        ],
                    )
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

        return QualitySyncResult(
            pending_locations=len(pending_locations),
            pending_snapshot=len(pending_snapshot),
            approvals=len(approvals),
        )

    def fetch_pending_records(self) -> List[Dict[str, Any]]:
        conn = self.connect()
        try:
            rows = self._fetch_all(
                conn,
                """
                SELECT s.row_key, s.item, s.produc, s.subgrupo, s.oc_numero, s.fecha_ing, s.fecha_ing_iso,
                       s.total_qty, s.remaining_qty, s.approved_qty, s.progress_pct, s.is_approved, s.is_ar7_item,
                       COALESCE(o.observacion, '') AS observaciones,
                       COALESCE(h.handler_name, '') AS encargado,
                       COALESCE(h.updated_by, '') AS encargado_updated_by,
                       COALESCE(TO_CHAR(h.updated_at AT TIME ZONE 'UTC', 'YYYY-MM-DD"T"HH24:MI:SS'), '') AS encargado_updated_at
                FROM quality_pending_snapshot s
                LEFT JOIN quality_observations o ON o.item = s.item AND o.produc = s.produc
                LEFT JOIN quality_pending_handlers h ON h.row_key = s.row_key
                WHERE NOT s.is_approved
                ORDER BY s.fecha_ing_iso DESC, s.produc ASC, s.oc_numero ASC
                """,
            )
            return [
                {
                    "row_key": row.get("row_key") or "",
                    "item": row.get("item") or "",
                    "produc": row.get("produc") or "",
                    "subgrupo": row.get("subgrupo") or "",
                    "oc_numero": row.get("oc_numero") or "",
                    "fecha_ing": row.get("fecha_ing") or "",
                    "fecha_ing_iso": row.get("fecha_ing_iso") or "",
                    "observaciones": row.get("observaciones") or "",
                    "canti_real": _format_value(row.get("total_qty")),
                    "approved_qty": _format_value(row.get("approved_qty")),
                    "total_qty": _format_value(row.get("total_qty")),
                    "progress_pct": int(row.get("progress_pct") or 0),
                    "is_approved": bool(row.get("is_approved")),
                    "is_ar7_item": bool(row.get("is_ar7_item")),
                    "encargado": row.get("encargado") or "",
                    "encargado_updated_at": row.get("encargado_updated_at") or "",
                    "encargado_updated_by": row.get("encargado_updated_by") or "",
                }
                for row in rows
            ]
        finally:
            conn.close()

    def fetch_history_records(self) -> List[Dict[str, Any]]:
        conn = self.connect()
        try:
            rows = self._fetch_all(
                conn,
                """
                SELECT a.item, a.produc, a.ubi, a.oc_numero, a.canti_real, a.fecha, a.fecha_iso,
                       a.fecha_ing, a.fecha_ing_iso, a.author_name,
                       COALESCE(o.observacion, '') AS observaciones
                FROM quality_approvals a
                LEFT JOIN quality_observations o ON o.item = a.item AND o.produc = a.produc
                ORDER BY a.fecha_iso DESC, a.produc ASC, a.oc_numero ASC, a.ubi ASC
                """,
            )
            return [
                {
                    "item": row.get("item") or "",
                    "produc": row.get("produc") or "",
                    "ubi": row.get("ubi") or "",
                    "oc_numero": row.get("oc_numero") or "",
                    "canti_real": _format_value(row.get("canti_real")),
                    "fecha": row.get("fecha") or "",
                    "fecha_iso": row.get("fecha_iso") or "",
                    "fecha_ing": row.get("fecha_ing") or "",
                    "fecha_ing_iso": row.get("fecha_ing_iso") or "",
                    "name": row.get("author_name") or "",
                    "observaciones": row.get("observaciones") or "",
                }
                for row in rows
            ]
        finally:
            conn.close()

    def fetch_pending_categories_state(self) -> Dict[str, Any]:
        conn = self.connect()
        try:
            categories = self._fetch_all(
                conn,
                """
                SELECT category_id, name, color, updated_at, updated_by
                FROM quality_pending_categories
                ORDER BY CASE WHEN category_id = 'hidden' THEN 0 ELSE 1 END, LOWER(name)
                """,
            )
            assignments_rows = self._fetch_all(conn, "SELECT row_key, category_id FROM quality_pending_category_assignments")
            state = {
                "categories": [
                    {
                        "id": row.get("category_id") or "",
                        "name": row.get("name") or "",
                        "color": row.get("color") or "#4aa3ff",
                        "updated_at": row.get("updated_at").isoformat() if row.get("updated_at") else "",
                        "updated_by": row.get("updated_by") or "",
                    }
                    for row in categories
                ],
                "assignments": {row.get("row_key") or "": row.get("category_id") or "" for row in assignments_rows},
            }
            return _ensure_hidden_category(state)
        finally:
            conn.close()

    def save_observation(self, item: str, produc: str, observacion: str, updated_by: str = "") -> Dict[str, Any]:
        item = _format_value(item)
        produc = _format_value(produc)
        if not item and not produc:
            raise ValueError("item/produc requeridos")
        conn = self.connect()
        try:
            existing = self._fetch_one(conn, "SELECT subgrupo FROM quality_observations WHERE item = %s AND produc = %s", (item, produc))
            if not existing:
                existing = self._fetch_one(
                    conn,
                    "SELECT subgrupo FROM quality_pending_snapshot WHERE item = %s AND produc = %s LIMIT 1",
                    (item, produc),
                ) or {"subgrupo": ""}
            now = _utc_now()
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO quality_observations(item, produc, subgrupo, observacion, updated_at, updated_by)
                    VALUES (%s, %s, %s, %s, %s, %s)
                    ON CONFLICT (item, produc) DO UPDATE
                    SET observacion = EXCLUDED.observacion,
                        updated_at = EXCLUDED.updated_at,
                        updated_by = EXCLUDED.updated_by,
                        subgrupo = CASE WHEN quality_observations.subgrupo = '' THEN EXCLUDED.subgrupo ELSE quality_observations.subgrupo END
                    """,
                    (item, produc, _format_value(existing.get("subgrupo")), str(observacion or "").strip(), now, _format_value(updated_by)),
                )
            conn.commit()
            return {
                "item": item,
                "produc": produc,
                "subgrupo": _format_value(existing.get("subgrupo")),
                "observacion": str(observacion or "").strip(),
                "updated_at": now.isoformat(),
                "updated_by": _format_value(updated_by),
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def create_category(self, name: str, color: str, updated_by: str = "") -> Dict[str, Any]:
        category = {
            "id": uuid.uuid4().hex,
            "name": str(name or "").strip(),
            "color": str(color or "#4aa3ff").strip() or "#4aa3ff",
            "updated_at": _utc_now(),
            "updated_by": _format_value(updated_by),
        }
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO quality_pending_categories(category_id, name, color, updated_at, updated_by)
                    VALUES (%s, %s, %s, %s, %s)
                    """,
                    (category["id"], category["name"], category["color"], category["updated_at"], category["updated_by"]),
                )
            conn.commit()
            return {
                "id": category["id"],
                "name": category["name"],
                "color": category["color"],
                "updated_at": category["updated_at"].isoformat(),
                "updated_by": category["updated_by"],
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def update_category(self, category_id: str, name: str, color: str, updated_by: str = "") -> Dict[str, Any]:
        now = _utc_now()
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    UPDATE quality_pending_categories
                    SET name = %s, color = %s, updated_at = %s, updated_by = %s
                    WHERE category_id = %s
                    """,
                    (str(name or "").strip(), str(color or "#4aa3ff").strip() or "#4aa3ff", now, _format_value(updated_by), str(category_id or "").strip()),
                )
                if cur.rowcount <= 0:
                    raise KeyError("Categoria no encontrada.")
            conn.commit()
            return {
                "id": str(category_id or "").strip(),
                "name": str(name or "").strip(),
                "color": str(color or "#4aa3ff").strip() or "#4aa3ff",
                "updated_at": now.isoformat(),
                "updated_by": _format_value(updated_by),
            }
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def delete_category(self, category_id: str) -> None:
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute("DELETE FROM quality_pending_categories WHERE category_id = %s", (str(category_id or "").strip(),))
                if cur.rowcount <= 0:
                    raise KeyError("Categoria no encontrada.")
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def assign_category(self, row_key: str, category_id: str, updated_by: str = "") -> Dict[str, Any]:
        now = _utc_now()
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                if str(category_id or "").strip():
                    cur.execute(
                        """
                        INSERT INTO quality_pending_category_assignments(row_key, category_id, updated_at, updated_by)
                        VALUES (%s, %s, %s, %s)
                        ON CONFLICT (row_key) DO UPDATE
                        SET category_id = EXCLUDED.category_id,
                            updated_at = EXCLUDED.updated_at,
                            updated_by = EXCLUDED.updated_by
                        """,
                        (str(row_key or "").strip(), str(category_id or "").strip(), now, _format_value(updated_by)),
                    )
                else:
                    cur.execute("DELETE FROM quality_pending_category_assignments WHERE row_key = %s", (str(row_key or "").strip(),))
            conn.commit()
            return {"row_key": str(row_key or "").strip(), "category_id": str(category_id or "").strip()}
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def assign_handler(self, row_key: str, handler_name: str, updated_by: str = "") -> Dict[str, Any]:
        now = _utc_now()
        conn = self.connect()
        try:
            with conn.cursor() as cur:
                cur.execute(
                    """
                    INSERT INTO quality_pending_handlers(row_key, handler_name, updated_at, updated_by)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (row_key) DO UPDATE
                    SET handler_name = EXCLUDED.handler_name,
                        updated_at = EXCLUDED.updated_at,
                        updated_by = EXCLUDED.updated_by
                    """,
                    (str(row_key or "").strip(), _format_value(handler_name), now, _format_value(updated_by)),
                )
            conn.commit()
            return {"row_key": str(row_key or "").strip(), "name": _format_value(handler_name)}
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def fetch_admin_stats(self) -> Dict[str, Any]:
        pending_records = self.fetch_pending_records()
        conn = self.connect()
        try:
            summary_row = self._fetch_one(
                conn,
                """
                SELECT
                    COUNT(*) FILTER (WHERE NOT is_approved) AS pending_count,
                    COUNT(*) FILTER (WHERE is_approved) AS approved_count,
                    COUNT(*) AS total_count
                FROM quality_pending_snapshot
                """,
            ) or {}
            history_rows = self._fetch_all(
                conn,
                """
                SELECT a.item, a.produc, a.ubi, a.oc_numero, a.canti_real, a.fecha, a.fecha_iso,
                       a.fecha_ing, a.fecha_ing_iso, a.author_name,
                       COALESCE(o.observacion, '') AS observaciones,
                       COALESCE(s.subgrupo, '') AS subgrupo,
                       COALESCE(s.total_qty, 0) AS total_qty
                FROM quality_approvals a
                LEFT JOIN quality_observations o ON o.item = a.item AND o.produc = a.produc
                LEFT JOIN quality_pending_snapshot s
                  ON s.item = a.item AND s.produc = a.produc AND s.oc_numero = a.oc_numero
                ORDER BY a.fecha_iso DESC, a.produc ASC, a.oc_numero ASC
                """,
            )
        finally:
            conn.close()

        history_records: List[Dict[str, Any]] = []
        timed_records = 0
        for row in history_rows:
            approval_date = _parse_date(row.get("fecha_iso") or row.get("fecha"))
            ingress_date = _parse_date(row.get("fecha_ing_iso") or row.get("fecha_ing"))
            control_days = None
            if approval_date and ingress_date:
                delta_days = (approval_date - ingress_date).days
                if delta_days >= 0:
                    control_days = delta_days
                    timed_records += 1
            history_records.append(
                {
                    "item": row.get("item") or "",
                    "produc": row.get("produc") or "",
                    "subgrupo": row.get("subgrupo") or "",
                    "observaciones": row.get("observaciones") or "",
                    "oc_numero": row.get("oc_numero") or "",
                    "ubi": row.get("ubi") or "",
                    "approval_date": row.get("fecha_iso") or row.get("fecha") or "",
                    "approval_date_label": row.get("fecha") or "",
                    "author": row.get("author_name") or "",
                    "controlled_qty": float(row.get("canti_real") or 0.0),
                    "total_qty": float(row.get("total_qty") or 0.0),
                    "control_days": control_days,
                    "fecha_ing_iso": row.get("fecha_ing_iso") or row.get("fecha_ing") or "",
                    "match_level": "approval" if row.get("fecha_ing_iso") else "",
                }
            )

        return {
            "generated_at": _utc_now().isoformat(timespec="seconds"),
            "summary": {
                "pending_count": int(summary_row.get("pending_count") or 0),
                "approved_snapshot_count": int(summary_row.get("approved_count") or 0),
                "pending_total": int(summary_row.get("total_count") or 0),
                "history_total": len(history_records),
                "timed_history_total": timed_records,
            },
            "pending_records": pending_records,
            "history_records": history_records,
        }
