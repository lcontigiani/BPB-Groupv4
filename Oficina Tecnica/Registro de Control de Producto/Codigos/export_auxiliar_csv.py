import argparse
import io
import logging
import re
import csv
from pathlib import Path
from typing import Optional

try:
    import openpyxl
except ImportError:
    openpyxl = None

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None

try:
    import xlrd
except ImportError:
    xlrd = None


def setup_logging():
    logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")


def load_wb(path: Path, password: str = "bpb"):
    if openpyxl is None:
        return None
    try:
        return openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        if msoffcrypto is None:
            logging.warning("Fallo al abrir %s (posiblemente encriptado y falta msoffcrypto): %s", path, e)
            return None
        try:
            bio = io.BytesIO()
            with open(path, "rb") as f:
                office_file = msoffcrypto.OfficeFile(f)
                office_file.load_key(password=password)
                office_file.decrypt(bio)
            bio.seek(0)
            return openpyxl.load_workbook(bio, read_only=True, data_only=True)
        except Exception as e2:
            logging.warning("Fallo al desencriptar/abrir %s: %s", path, e2)
            return None


def load_xls(path: Path):
    if xlrd is None:
        return None
    try:
        # xlrd opens .xls files
        return xlrd.open_workbook(filename=str(path), formatting_info=False)
    except Exception as e:
        logging.warning("Fallo al abrir .xls %s con xlrd: %s", path, e)
        return None



def _export_aux_sheet_xlsx(xls_path: Path, dest_dir: Path, password: str = "bpb"):
    wb = load_wb(xls_path, password)
    if wb is None:
        logging.warning("No se pudo abrir %s", xls_path)
        return
    sheet_name = None
    for name in wb.sheetnames:
        if name and name.strip().lower() == "auxiliar":
            sheet_name = name
            break
    if not sheet_name:
        wb.close()
        logging.warning("Hoja Auxiliar no encontrada en %s", xls_path)
        return
    sh = wb[sheet_name]
    rows = []
    for row in sh.iter_rows(values_only=True):
        rows.append([cell if cell is not None else "" for cell in row])
    wb.close()
    _write_csv(rows, dest_dir, xls_path)

def export_aux_sheet_xls(xls_path: Path, dest_dir: Path):
    wb = load_xls(xls_path)
    if wb is None:
        logging.warning("No se pudo abrir .xls %s", xls_path)
        return
    
    sheet_name = None
    # xlrd sheet names
    for name in wb.sheet_names():
        if name and name.strip().lower() == "auxiliar":
            sheet_name = name
            break
            
    if not sheet_name:
        logging.warning("Hoja Auxiliar no encontrada en %s", xls_path)
        return
        
    sh = wb.sheet_by_name(sheet_name)
    rows = []
    # xlrd rows
    for rx in range(sh.nrows):
        row_vals = sh.row_values(rx)
        rows.append([c if c is not None else "" for c in row_vals])
        
    _write_csv(rows, dest_dir, xls_path)

def _write_csv(rows, dest_dir, xls_path):
    dest_dir.mkdir(parents=True, exist_ok=True)
    out_path = dest_dir / f"{xls_path.stem}_Auxiliar.csv"
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f, delimiter=";")
        writer.writerows(rows)
    logging.info("Exportado Auxiliar -> %s", out_path)

def export_aux_sheet(xls_path: Path, dest_dir: Path, password: str = "bpb"):
    # Dispatcher based on extension
    if xls_path.suffix.lower() == '.xls':
        export_aux_sheet_xls(xls_path, dest_dir)
    else:
        _export_aux_sheet_xlsx(xls_path, dest_dir, password)


def main():
    parser = argparse.ArgumentParser(description="Exporta hoja Auxiliar de R016-01 a CSV.")
    parser.add_argument("--root", required=True, help="Carpeta con R016-01*.xls*")
    parser.add_argument("--out", required=True, help="Carpeta destino para CSVs")
    parser.add_argument("--password", default="bpb")
    args = parser.parse_args()

    setup_logging()
    root = Path(args.root)
    out = Path(args.out)
    for xls in root.glob("R016-01*.xls*"):
        if not xls.is_file():
            continue
        if any("obsole" in part.lower() for part in xls.parts):
            continue
        # Check valid extensions
        # Check valid extensions
        if xls.suffix.lower() not in ['.xlsx', '.xlsm', '.xls']:
            continue
        export_aux_sheet(xls, out, args.password)


if __name__ == "__main__":
    main()
