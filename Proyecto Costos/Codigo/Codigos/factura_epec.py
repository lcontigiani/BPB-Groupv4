import json
import os
import re
import unicodedata
import urllib.request
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple

from PyPDF2 import PdfReader
import openpyxl
from openpyxl.cell.cell import MergedCell

from config import EPEC_FACTURAS_DIR, EPEC_EXCEL_PATH, EPEC_SHEET_NAME
from utils import log


def _normalize(text: str) -> str:
    if not isinstance(text, str):
        text = str(text)
    return "".join(c for c in unicodedata.normalize("NFD", text) if unicodedata.category(c) != "Mn").lower()


def _to_float(num_str: str) -> Optional[float]:
    try:
        clean = num_str.replace(".", "").replace(",", ".")
        return float(clean)
    except Exception:
        return None


def _extract_currency_numbers(text: str) -> List[str]:
    if not isinstance(text, str):
        text = str(text)
    return re.findall(r"\d{1,3}(?:\.\d{3})*,\d{2}", text)


def _next_month_first(dt: datetime) -> datetime:
    if dt.month == 12:
        return datetime(dt.year + 1, 1, 1)
    return datetime(dt.year, dt.month + 1, 1)


def _fetch_json(url: str) -> Optional[dict]:
    try:
        with urllib.request.urlopen(url, timeout=10) as resp:
            return json.loads(resp.read().decode("utf-8"))
    except Exception as exc:
        log(f"No se pudo obtener JSON desde {url}: {exc}", "WARN")
        return None


def _dolar_blue_venta(fecha: datetime) -> Optional[float]:
    url = f"https://api.bluelytics.com.ar/v2/historical?day={fecha.strftime('%Y-%m-%d')}"
    data = _fetch_json(url)
    if not data:
        return None
    try:
        val = data["blue"]["value_sell"]
        return float(val)
    except Exception:
        return None


def _sum_charges(amounts: List[float]) -> Optional[Tuple[float, List[float]]]:
    if not amounts:
        return None
    try:
        # Heuristic: first block of three amounts after small fees (~74.30) is the net EPEC trio.
        trio = None
        for idx, val in enumerate(amounts):
            if abs(val - 74.30) < 0.01 and idx + 3 <= len(amounts):
                candidate = amounts[idx + 1 : idx + 4]
                if len(candidate) == 3:
                    trio = candidate
                    break
        if trio is None:
            for i in range(1, len(amounts) - 2):
                a, b, c = amounts[i : i + 3]
                if a > 100000 and b > 100000 and c > 100000 and amounts[i - 1] < 100000:
                    trio = [a, b, c]
                    break
        if trio:
            return sum(trio), trio
    except Exception:
        pass
        return None


def _extract_power_energy(text: str) -> Dict[str, Optional[float]]:
    lines = [ln.strip() for ln in text.splitlines() if ln.strip()]
    values: Dict[str, Optional[float]] = {
        "pot_adq": None,
        "pot_fuera": None,
        "pot_pico": None,
        "ene_pico": None,
        "ene_resto": None,
        "ene_valle": None,
    }

    start = None
    for idx, ln in enumerate(lines):
        if "(0,10%)" in ln:
            start = idx + 1
            break
    if start is None:
        return values

    numeric_lines = []
    for ln in lines[start:]:
        if re.fullmatch(r"[0-9,\.]+", ln):
            numeric_lines.append(ln)
        if len(numeric_lines) >= 6:
            break

    if len(numeric_lines) >= 5:
        try:
            values["pot_adq"] = float(numeric_lines[0].replace(",", "."))
            values["pot_fuera"] = float(numeric_lines[1].replace(",", "."))
            values["pot_pico"] = float(numeric_lines[2].replace(",", "."))
            values["ene_pico"] = float(numeric_lines[3].replace(",", "."))
            values["ene_resto"] = float(numeric_lines[4].replace(",", "."))
            # The sixth line packs valle + monetary pieces; take leading digits as energy valle.
            m = re.match(r"^(\d{3,5})", numeric_lines[5])
            if m:
                leading = m.group(1)
                if len(leading) > 4:
                    leading = leading[:4]
                values["ene_valle"] = float(leading)
        except Exception:
            pass

    return values


def _extract_amount_groups(text: str) -> Tuple[List[float], List[float], Optional[float]]:
    """Devuelve (grupo_potencia, grupo_energia, cargo_cosfi) en pesos."""
    currs = _extract_currency_numbers(text)
    vals = [_to_float(c) for c in currs if _to_float(c) is not None]
    # Filtramos montos entre 50k y 2M para aislar los 3+3 principales.
    main = [v for v in vals if 50000 <= v <= 2000000]
    grupo_pot = main[0:3] if len(main) >= 3 else []
    grupo_ene = main[3:6] if len(main) >= 6 else []
    # Cargo coseno fi: primer valor entre 10k y 100k que no usemos ya.
    used = set(grupo_pot + grupo_ene)
    cargo_cosfi = None
    for v in vals:
        if v in used:
            continue
        if 10000 <= v <= 100000:
            cargo_cosfi = v
            break
    return grupo_pot, grupo_ene, cargo_cosfi


def _parse_period(text: str) -> Tuple[Optional[datetime], Optional[str]]:
    # Example: "30/09/2025 al 31/10/2025"
    m = re.search(r"(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})", text)
    if not m:
        return None, None
    fin_raw = m.group(2)
    try:
        fin_dt = datetime.strptime(fin_raw, "%d/%m/%Y")
        month_start = datetime(fin_dt.year, fin_dt.month, 1)
        return month_start, f"{m.group(1)} al {m.group(2)}"
    except Exception:
        return None, f"{m.group(1)} al {m.group(2)}"


def _extract_totals(text: str) -> Tuple[Optional[float], Optional[float]]:
    m = re.search(
        r"VENCIMIENTO\s+TOTAL\s+FACTURADO\s+TOTAL\s+A\s+PAGAR\s+(\d{2}/\d{2}/\d{4})\s+([\d\.,]+)\s+([\d\.,]+)",
        _normalize(text).replace("  ", " "),
    )
    if not m:
        # fallback: first two currency matches (often duplicated)
        amounts = _extract_currency_numbers(text)
        vals = [_to_float(a) for a in amounts[:2] if _to_float(a) is not None]
        if len(vals) >= 2:
            return vals[0], vals[1]
        return None, None
    total_facturado = _to_float(m.group(2))
    total_pagar = _to_float(m.group(3))
    return total_facturado, total_pagar


def _extract_percentages(text: str) -> Tuple[Optional[float], Optional[float], Optional[float], Optional[float]]:
    if not isinstance(text, str):
        text = str(text)
    raw = re.findall(r"(\d{1,2},\d{2})%", text)
    vals = []
    for r in raw:
        try:
            vals.append(float(r.replace(",", ".")) / 100.0)
        except Exception:
            continue
    while len(vals) < 5:
        vals.append(None)
    return vals[0], vals[2], vals[3], vals[4]  # 0:6.50, 1:27.00 (skip), 2:6.00, 3:0.40, 4:0.10


def _extract_cosfi(text: str) -> Optional[float]:
    m = re.search(r"0[,\.]9\d{2}", text)
    if m:
        return float(m.group(0).replace(",", "."))
    return None


def parse_factura(pdf_path: str) -> Optional[Dict[str, object]]:
    try:
        reader = PdfReader(pdf_path)
        full_text = "\n".join(str(page.extract_text() or "") for page in reader.pages)
    except Exception as exc:
        log(f"No se pudo leer PDF {pdf_path}: {exc}", "ERROR")
        return None

    power_energy = _extract_power_energy(full_text)
    grupo_pot, grupo_ene, cargo_cosfi = _extract_amount_groups(full_text)
    currency_strings = _extract_currency_numbers(full_text)
    currency_values = [_to_float(a) for a in currency_strings if _to_float(a) is not None]
    charges = _sum_charges(currency_values)
    total_facturado, total_pagar = _extract_totals(full_text)
    periodo_date, periodo_label = _parse_period(full_text)
    fdes, ordimp, dto2298, ersepr = _extract_percentages(full_text)
    cosfi_val = _extract_cosfi(full_text)
    breakdown_from_sum = charges[1] if charges and len(charges) > 1 else None
    if isinstance(breakdown_from_sum, tuple):
        breakdown_from_sum = list(breakdown_from_sum)

    return {
        "texto": full_text,
        "power_energy": power_energy,
        "charges_sum": charges[0] if charges else None,
        "charges_breakdown": grupo_pot or breakdown_from_sum,
        "energy_breakdown": grupo_ene,
        "cosfi_charge": cargo_cosfi,
        "cosfi_value": cosfi_val,
        "total_facturado": total_facturado,
        "total_pagar": total_pagar,
        "periodo_date": periodo_date,
        "periodo_label": periodo_label,
        "impuestos": {
            "fdes": fdes,
            "ordimp": ordimp,
            "dto2298": dto2298,
            "ersep": ersepr,
        },
    }


def _excel_cached_dolar(ws) -> Optional[float]:
    candidates = [("H3",), ("L3",), ("K3", "L3")]
    for candidate in candidates:
        try:
            if len(candidate) == 1:
                val = ws[candidate[0]].value
            else:
                label_cell, value_cell = candidate
                label_val = ws[label_cell].value if label_cell else None
                if isinstance(label_val, str) and "DOLAR" not in label_val.upper():
                    continue
                val = ws[value_cell].value
            if isinstance(val, (int, float)):
                return float(val)
            if isinstance(val, str):
                return float(val.replace(",", "."))
        except Exception:
            continue
    return None


def _dolar_para_periodo(periodo_date: datetime, ws_data) -> Optional[float]:
    target = _next_month_first(periodo_date)
    rate = _dolar_blue_venta(target)
    if rate:
        return rate
    # Fallback al dolar en planilla si no hay internet/API.
    return _excel_cached_dolar(ws_data)


def _find_row(ws) -> int:
    for row in range(10, 150):
        if not any(ws.cell(row=row, column=col).value for col in range(2, 8)):
            return row
    return ws.max_row + 1


def _find_row_by_date(ws, target_date: datetime) -> Optional[int]:
    for row in range(10, 150):
        val = ws.cell(row=row, column=2).value
        if isinstance(val, datetime) and val.date() == target_date.date():
            return row
    return None


def _footer_start_row(ws) -> int:
    for r in range(10, ws.max_row + 1):
        val = ws.cell(row=r, column=3).value
        if isinstance(val, str) and "APROBADO" in val.upper():
            return r
    return ws.max_row + 1


def _find_or_make_space(ws, target_date: datetime) -> int:
    """Ubica la fila por fecha; si no existe, inserta fila manteniendo el footer intacto."""
    footer_row = _footer_start_row(ws)
    data_rows: List[int] = []
    for row in range(10, footer_row):
        val = ws.cell(row=row, column=2).value
        if isinstance(val, datetime):
            data_rows.append(row)

    # Si ya existe el periodo, se reutiliza esa fila.
    for row in data_rows:
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, datetime) and val.date() == target_date.date():
            return row

    # Insertar manteniendo orden descendente (mas nuevo arriba)
    insert_row = footer_row - 1
    for row in data_rows:
        val = ws.cell(row=row, column=2).value
        if val and isinstance(val, datetime) and target_date > val:
            insert_row = row
            break
        insert_row = max(insert_row, row + 1)

    if insert_row >= footer_row:
        insert_row = footer_row - 1

    ws.insert_rows(insert_row)
    return insert_row


def _snapshot_row_styles(ws, row: int, max_col: int = 30) -> List[object]:
    styles = []
    for c in range(1, max_col + 1):
        styles.append(ws.cell(row=row, column=c)._style)
    return styles


def _apply_row_styles(ws, row: int, styles: List[object]):
    for c, st in enumerate(styles, start=1):
        ws.cell(row=row, column=c)._style = st


def _first_empty_row(ws, footer_row: int) -> Optional[int]:
    for r in range(10, footer_row):
        if not any(ws.cell(row=r, column=c).value not in (None, "") for c in range(2, 22)):
            return r
    return None


def _set_value(ws, row: int, col: int, val):
    cell = ws.cell(row=row, column=col)
    if isinstance(cell, MergedCell):
        for rng in ws.merged_cells.ranges:
            if (row, col) in rng:
                cell = ws.cell(row=rng.min_row, column=rng.min_col)
                break
    cell.value = val


def _clear_row_values(ws, row: int, max_col: int = 30):
    for c in range(1, max_col + 1):
        _set_value(ws, row, c, None)


def _unmerge_row(ws, row: int):
    """Elimina merges que atraviesan la fila indicada."""
    to_remove = []
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row:
            to_remove.append(rng)
    for rng in to_remove:
        ws.unmerge_cells(str(rng))


def _apply_spacer_from_above(ws):
    """Toma la primera fila vacia antes del footer y le aplica el formato de la fila superior, limpiando valores."""
    footer_row = _footer_start_row(ws)
    spacer_row = _first_empty_row(ws, footer_row)
    if not spacer_row:
        return
    if spacer_row <= 10:
        return
    source_row = spacer_row - 1
    _unmerge_row(ws, spacer_row)
    styles = _snapshot_row_styles(ws, source_row)
    _apply_row_styles(ws, spacer_row, styles)
    _clear_row_values(ws, spacer_row, max_col=22)


def _copy_row_style(ws, source_row: int, target_row: int, max_col: int = 25):
    for col in range(1, max_col + 1):
        s = ws.cell(row=source_row, column=col)
        t = ws.cell(row=target_row, column=col)
        t._style = s._style


def _ensure_aprobado_merge(ws):
    # No tocar footer para no mover ni modificar el texto/boton.
    return


def _ensure_data_merges(ws, footer_row: int):
    return


def _ensure_spacer_merge(ws, footer_row: int):
    return


def _sync_row_formulas(ws):
    """Ajusta formulas dependientes de la fila (T/U) para evitar referencias corridas al insertar filas."""
    for row in range(10, ws.max_row + 1):
        if not isinstance(ws.cell(row=row, column=2).value, datetime):
            continue
        ws.cell(row=row, column=20, value=f"=(G{row}+K{row}+M{row})*SUM(P{row}:S{row})")
        ws.cell(row=row, column=21, value=f"=G{row}+K{row}+M{row}+T{row}")


def actualizar_excel(datos: Dict[str, object]) -> bool:
    if not os.path.isfile(EPEC_EXCEL_PATH):
        log(f"No se encontro Excel destino: {EPEC_EXCEL_PATH}", "ERROR")
        return False

    try:
        wb = openpyxl.load_workbook(EPEC_EXCEL_PATH, data_only=False, keep_vba=True)
        wb_data = openpyxl.load_workbook(EPEC_EXCEL_PATH, data_only=True, keep_vba=True)
    except Exception as exc:
        log(f"No se pudo abrir Excel {EPEC_EXCEL_PATH}: {exc}", "ERROR")
        return False

    if EPEC_SHEET_NAME not in wb.sheetnames:
        log(f"No se encontro hoja {EPEC_SHEET_NAME} en Excel EPEC.", "ERROR")
        return False
    ws = wb[EPEC_SHEET_NAME]
    ws_data = wb_data[EPEC_SHEET_NAME]

    periodo_date: Optional[datetime] = datos.get("periodo_date")  # type: ignore
    if not periodo_date:
        log("No se pudo determinar la fecha de periodo; se omite carga.", "WARN")
        return False

    # Insertar fila y aplicar formato de la fila inferior (sin tocar resto)
    target_row = _find_or_make_space(ws, periodo_date)
    footer_row_after = _footer_start_row(ws)
    row_below = min(target_row + 1, footer_row_after - 1)
    below_styles = _snapshot_row_styles(ws, row_below)
    _apply_row_styles(ws, target_row, below_styles)

    # Asegurar que la fila vacia (inmediatamente despues de la ultima con datos) mantenga formato de la fila superior y quede sin contenido.
    _apply_spacer_from_above(ws)
    next_month_first = _next_month_first(periodo_date)
    dolar_ref = _dolar_para_periodo(periodo_date, ws_data) or 1.0
    charges_sum = datos.get("charges_sum")  # type: ignore
    charges_breakdown: List[float] = datos.get("charges_breakdown") or []  # type: ignore
    energy_breakdown: List[float] = datos.get("energy_breakdown") or []  # type: ignore
    cosfi_charge = datos.get("cosfi_charge")  # type: ignore
    cosfi_val = datos.get("cosfi_value")  # type: ignore
    impuestos = datos.get("impuestos") or {}  # type: ignore

    _set_value(ws, target_row, 2, periodo_date)
    if datos["power_energy"]["pot_adq"] is not None:  # type: ignore
        _set_value(ws, target_row, 4, datos["power_energy"]["pot_adq"])  # type: ignore
    if datos["power_energy"]["pot_fuera"] is not None:  # type: ignore
        _set_value(ws, target_row, 5, datos["power_energy"]["pot_fuera"])  # type: ignore
    if datos["power_energy"]["pot_pico"] is not None:  # type: ignore
        _set_value(ws, target_row, 6, datos["power_energy"]["pot_pico"])  # type: ignore
    if datos["power_energy"]["ene_pico"] is not None:  # type: ignore
        _set_value(ws, target_row, 8, datos["power_energy"]["ene_pico"])  # type: ignore
    if datos["power_energy"]["ene_resto"] is not None:  # type: ignore
        _set_value(ws, target_row, 9, datos["power_energy"]["ene_resto"])  # type: ignore
    if datos["power_energy"]["ene_valle"] is not None:  # type: ignore
        _set_value(ws, target_row, 10, datos["power_energy"]["ene_valle"])  # type: ignore

    if charges_sum:
        if charges_breakdown:
            parts = "+".join(f"{v:.2f}" for v in charges_breakdown)
            _set_value(ws, target_row, 7, f"=({parts})/{dolar_ref:.2f}")
        else:
            _set_value(ws, target_row, 7, f"={charges_sum:.2f}/{dolar_ref:.2f}")

    if energy_breakdown:
        parts = "+".join(f"{v:.2f}" for v in energy_breakdown)
        _set_value(ws, target_row, 11, f"=({parts})/{dolar_ref:.2f}")

    if cosfi_val:
        _set_value(ws, target_row, 12, round(cosfi_val, 3))
    if cosfi_charge:
        _set_value(ws, target_row, 13, f"={cosfi_charge:.2f}/{dolar_ref:.2f}")

    # Periodo leído (texto a dos celdas)
    if datos.get("periodo_label") and isinstance(datos.get("periodo_label"), str):
        m = re.search(r"(\d{2}/\d{2}/\d{4})\s+al\s+(\d{2}/\d{2}/\d{4})", datos["periodo_label"])  # type: ignore
        if m:
            try:
                desde = datetime.strptime(m.group(1), "%d/%m/%Y")
                hasta = datetime.strptime(m.group(2), "%d/%m/%Y")
                _set_value(ws, target_row, 14, desde)
                _set_value(ws, target_row, 15, hasta)
            except Exception:
                pass

    # Impuestos no deducibles
    _set_value(ws, target_row, 16, impuestos.get("fdes"))
    _set_value(ws, target_row, 17, impuestos.get("ordimp"))
    _set_value(ws, target_row, 18, impuestos.get("dto2298"))
    _set_value(ws, target_row, 19, impuestos.get("ersep"))

    # Observacion con dolar de referencia del 1er dia del mes siguiente
    obs_date = next_month_first.strftime("%d/%m/%y")
    obs_rate_str = f"{round(dolar_ref):,.0f}".replace(",", ".")
    _set_value(ws, target_row, 22, f"Ajustado a Valor Dólar ${obs_rate_str} del dia {obs_date}")

    # Reajustar formulas dependientes de fila
    _sync_row_formulas(ws)

    # Ajustes de formatos clave
    ws.cell(row=target_row, column=2).number_format = "dd/mm/yyyy"
    for col in (7, 11, 13):
        ws.cell(row=target_row, column=col).number_format = '"$"#,##0.00'

    # No tocar merges ni footer

    try:
        wb.save(EPEC_EXCEL_PATH)
        log(f"Factura EPEC cargada en fila {target_row} ({EPEC_SHEET_NAME}) usando {EPEC_EXCEL_PATH}", "INFO")
        return True
    except Exception as exc:
        log(f"No se pudo guardar Excel {EPEC_EXCEL_PATH}: {exc}", "ERROR")
        return False


def procesar_factura(pdf_path: str) -> bool:
    if not pdf_path.lower().endswith(".pdf"):
        return False
    if not os.path.abspath(pdf_path).lower().startswith(os.path.abspath(EPEC_FACTURAS_DIR).lower()):
        return False

    if not os.path.isfile(pdf_path):
        log(f"PDF no encontrado para procesar: {pdf_path}", "WARN")
        return False

    datos = parse_factura(pdf_path)
    if not datos:
        return False

    ok = actualizar_excel(datos)
    return ok
